# =============================================================================
# 🔒 ARCHIVO BLOQUEADO — VER CLAUDE.md
# No modificar sin orden explícita de KIRVY.
# Versión: v5 | Fecha: 30-Jun-2026
# Columnas nómina/préstamos: Nombre | Nota | Cta. Contable | Importe | Fecha
# =============================================================================
"""
conciliacion_nomina.py — Automatiza pagos bancarios BBVA Net Cash.
Requiere: pandas, openpyxl, pdfplumber
"""
import os, re, unicodedata
import pandas as pd
import pdfplumber
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL     = "1F4E78"
TOTAL_FILL      = "D9D9D9"
FILL_OK         = "E2EFDA"
FILL_DIFF       = "FCE4D6"
FILL_ONLY       = "FFF2CC"
FILL_COMPLEMENTO= "D9E1F2"
_THIN  = Side(style="thin", color="CCCCCC")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

def norm(s):
    s = unicodedata.normalize("NFKD", str(s).strip())
    s = s.encode("ascii", "ignore").decode("utf-8")
    return re.sub(r"\s+", " ", s).upper().strip()

def style_header(ws, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial")
        c.fill      = PatternFill("solid", start_color=HEADER_FILL)
        c.border    = BORDER
        c.alignment = Alignment(horizontal="center")

def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

def _norm_col(s):
    s = unicodedata.normalize("NFKD", str(s).strip())
    s = s.encode("ascii", "ignore").decode("utf-8")
    return re.sub(r"\s+", " ", s).lower().strip()

def load_poliza(excel_path):
    """Lee el catálogo de cuentas. Devuelve dict con 'empleados' y 'prestamos'
    (cada uno: DataFrame indexado por nombre_norm con columnas Cuenta, nombre).
    Compatible con el formato: dos tablas en una sola hoja —
      col 0-1 = PRESTAMOS, col 3-4 = EMPLEADOS."""
    df_raw = pd.read_excel(excel_path, sheet_name=0, header=None, dtype=str)

    def _parse_tabla(col_cta, col_nom):
        rows = []
        for _, row in df_raw.iterrows():
            cta  = row.get(col_cta, None)
            nom  = row.get(col_nom, None)
            if pd.isna(cta) or pd.isna(nom):
                continue
            cta_s = str(cta).strip()
            nom_s = " ".join(str(nom).split())
            if not cta_s or nom_s.lower() in ("nombre empleado", "nombre del empleado"):
                continue
            rows.append({"Cuenta": cta_s, "nombre": nom_s, "nombre_norm": norm(nom_s)})
        if not rows:
            return pd.DataFrame(columns=["Cuenta","nombre","nombre_norm"]).set_index("nombre_norm")
        return pd.DataFrame(rows).set_index("nombre_norm")

    n_cols = len(df_raw.columns)
    if n_cols >= 5:
        prestamos = _parse_tabla(0, 1)
        empleados = _parse_tabla(3, 4)
    else:
        empleados = _parse_tabla(0, 1)
        prestamos = pd.DataFrame(columns=["Cuenta","nombre","nombre_norm"]).set_index("nombre_norm")

    return {"empleados": empleados, "prestamos": prestamos}

def _get_words(pdf_path):
    result = []
    with pdfplumber.open(pdf_path) as pdf:
        for pi, page in enumerate(pdf.pages):
            for w in page.extract_words(x_tolerance=3, y_tolerance=3):
                result.append((pi, w["x0"], w["top"], w["text"]))
    return result

def _full_text(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        return "\n".join(p.extract_text() or "" for p in pdf.pages)

def detect_template(pdf_path):
    text = _full_text(pdf_path)
    if "Consulta Dispersi" in text: return "dispersion"
    if "Pagos y Transferencias" in text: return "pagos_transferencias"
    raise ValueError("No reconozco el tipo de PDF: " + pdf_path)

def _meta_dispersion(text):
    def grab(pat, d=""):
        m = re.search(pat, text); return m.group(1).strip() if m else d
    return {"descripcion":     grab(r"Descripci[o\xf3]n:\s*(.+?)(?:\n|\s+N[u\xfa]mero\b|\s+Fecha\b|$)"),
            "fecha_operacion": grab(r"Fecha de operaci[o\xf3]n:\s*([\d/]+)"),
            "folio":           grab(r"Folio del lote:\s*(\S+)"),
            "importe_total":   grab(r"Importe total:\s*([\d,\.]+)"),
            "num_registros":   grab(r"N[u\xfa]mero de registros:\s*(\d+)")}

def _meta_pagos(text):
    def grab(pat, d=""):
        m = re.search(pat, text); return m.group(1).strip() if m else d
    return {"descripcion":       grab(r"Descripci[o\xf3]n:\s*(.+?)(?:\n|\s+N[u\xfa]mero\b|\s+Fecha\b|$)"),
            "fecha_creacion":    grab(r"Fecha de Creaci[o\xf3]n:\s*([\d/]+)"),
            "fecha_aplicacion":  grab(r"Fecha de aplicaci[o\xf3]n:\s*([\d/]+)"),
            "folio":             grab(r"Folio:\s*(\S+)"),
            "importe_cargo_mxp": grab(r"MXP:\s*([\d,\.]+)")}

def _assign_to_nearest_anchor(words, anchor_list):
    result = {}
    for top, t in words:
        if not anchor_list: continue
        idx = min(anchor_list, key=lambda a: abs(a[1] - top))[0]
        result[idx] = (result.get(idx, "") + " " + t).strip()
    return result

def _table_start_top(ws):
    tops = [top for x0, top, t in ws if t == "Detalle"]
    return tops[0] if tops else -1

def _footer_top(ws):
    all_tops = [top for _, top, _ in ws]
    if not all_tops: return float("inf")
    max_top = max(all_tops)
    tops = [top for x0, top, t in ws if t == "BBVA" and top > max_top * 0.7]
    return min(tops) if tops else float("inf")

def _is_header_label(t):
    return any(c.islower() for c in t)

def _find_col_positions(ws_page, start_top):
    positions = {}
    for x0, top, t in ws_page:
        if start_top - 60 < top < start_top + 90:
            t_norm = t.replace("\xf3", "o").replace("\xf2", "o")
            if t in ("Importe",): positions["importe"] = x0
            elif t in ("Nombre",): positions["nombre"] = x0
            elif t in ("C\xf3digo", "Codigo", "Código"): positions["codigo"] = x0
    return positions

def parse_dispersion(pdf_path):
    text = _full_text(pdf_path)
    meta = _meta_dispersion(text)
    words = _get_words(pdf_path)
    rows = []
    by_page = {}
    for pi, x0, top, t in words:
        by_page.setdefault(pi, []).append((x0, top, t))
    cached_cols = {}
    for pi, ws in sorted(by_page.items()):
        start_top = _table_start_top(ws) if pi == 0 else -1
        all_tops = [top for _, top, _ in ws]
        if all_tops:
            _max_t = max(all_tops)
            _bbva_footer = [top for x0, top, t in ws if t == "BBVA" and x0 < 50 and top > _max_t * 0.7]
            end_top = min(_bbva_footer) if _bbva_footer else float("inf")
        else:
            end_top = float("inf")
        if pi == 0:
            cached_cols = _find_col_positions(ws, start_top)
        cols = cached_cols
        imp_x = cols.get("importe", 150)
        nom_x = cols.get("nombre",  200)
        cod_x = cols.get("codigo",  nom_x + 70)
        ws_f = [(x0, round(top, 1), t) for x0, top, t in ws
                if start_top < top < end_top]
        anchors = sorted(set((top, t) for x0, top, t in ws_f
                              if x0 < 50 and re.fullmatch(r"\d{1,2}", t)),
                         key=lambda a: a[0])
        if not anchors: continue
        anchor_list = [(i, top) for i, (top, _) in enumerate(anchors)]
        anchor_tops = {top for top, _ in anchors}
        cuenta_by_top, importe_words, nombre_words = {}, [], []
        for x0, top, t in ws_f:
            if 50 <= x0 < imp_x - 20 and top in anchor_tops and re.fullmatch(r"\d+", t):
                cuenta_by_top[top] = t
            elif imp_x - 25 <= x0 <= imp_x + 25 and re.search(r"[\d,]+\.\d{2}", t):
                importe_words.append((top, t))
            elif nom_x - 25 <= x0 < cod_x - 12 and not _is_header_label(t):
                nombre_words.append((top, t))
        nombre_by_idx  = _assign_to_nearest_anchor(nombre_words,  anchor_list)
        importe_by_idx = _assign_to_nearest_anchor(importe_words, anchor_list)
        for idx, (top, rn) in enumerate(anchors):
            imp = importe_by_idx.get(idx, "0")
            rows.append({"row": int(rn),
                         "cuenta_abono": cuenta_by_top.get(top, "").strip(),
                         "importe": float(str(imp).replace(",","") or 0),
                         "nombre":  nombre_by_idx.get(idx, "").strip()})
    rows.sort(key=lambda r: r["row"])
    return meta, rows

def parse_pagos_transferencias(pdf_path):
    _DIVISAS = {"MXP","USD","EUR","CAD","GBP","JPY","SEK","CHF"}
    text = _full_text(pdf_path)
    meta = _meta_pagos(text)
    words = _get_words(pdf_path)
    rows = []
    by_page = {}
    for pi, x0, top, t in words:
        by_page.setdefault(pi, []).append((x0, top, t))
    for pi, ws in by_page.items():
        start_top = _table_start_top(ws) if pi == 0 else -1
        end_top   = _footer_top(ws)
        ws = [(x0, round(top, 1), t) for x0, top, t in ws
              if start_top < top < end_top]
        anchors = sorted(set((top, t) for x0, top, t in ws
                              if x0 < 50 and re.fullmatch(r"\d{1,2}", t)),
                         key=lambda a: a[0])
        if not anchors: continue
        anchor_list = [(i, top) for i, (top, _) in enumerate(anchors)]
        banco_words, monto_words, motivo_words, titular_words = [], [], [], []
        for x0, top, t in ws:
            if 200 <= x0 < 263 and not _is_header_label(t):
                banco_words.append((top, t))
            elif 258 <= x0 < 295 and re.search(r"[\d,]+\.\d{2}", t):
                monto_words.append((top, t))
            elif 295 <= x0 < 368 and not _is_header_label(t) and t not in _DIVISAS:
                motivo_words.append((top, t))
            elif 375 <= x0 < 430 and not _is_header_label(t) and t not in _DIVISAS:
                if t not in {"MISMO","DIA","SIGUIENTE","Disponibilidad"}:
                    titular_words.append((top, t))
        banco_by_idx   = _assign_to_nearest_anchor(banco_words,   anchor_list)
        monto_by_idx   = _assign_to_nearest_anchor(monto_words,   anchor_list)
        motivo_by_idx  = _assign_to_nearest_anchor(motivo_words,  anchor_list)
        titular_by_idx = _assign_to_nearest_anchor(titular_words, anchor_list)
        for idx, (top, rn) in enumerate(anchors):
            ms = monto_by_idx.get(idx, "0").replace(",","") or "0"
            rows.append({"no": int(rn),
                         "banco_destino": banco_by_idx.get(idx,"").strip(),
                         "monto":  float(ms),
                         "motivo": motivo_by_idx.get(idx,"").strip(),
                         "titular":titular_by_idx.get(idx,"").strip()})
    rows.sort(key=lambda r: r["no"])
    return meta, rows

def _resolver_excel_df(excel_map):
    if isinstance(excel_map, dict):
        return excel_map.get("empleados", pd.DataFrame(columns=["Cuenta","nombre","nombre_norm"]).set_index("nombre_norm"))
    return excel_map

def _construir_merge_conciliacion(excel_map, pdf_rows):
    excel_df_src = _resolver_excel_df(excel_map)
    pdf_df = pd.DataFrame(pdf_rows)
    pdf_df["nombre"] = pdf_df["nombre"].apply(lambda s: " ".join(str(s).split()))
    pdf_df["nombre_norm"] = pdf_df["nombre"].apply(norm)
    agg_spec = {"nombre": "first", "importe": "sum"}
    if "cuenta_abono" in pdf_df.columns:
        agg_spec["cuenta_abono"] = "first"
    pdf_g = pdf_df.groupby("nombre_norm", as_index=False).agg(agg_spec)
    excel_df = excel_df_src.reset_index()
    excel_dedup = excel_df.drop_duplicates(subset=["nombre_norm"], keep="first")
    merged = pd.merge(excel_dedup[["nombre_norm", "Cuenta"]],
                      pdf_g, on="nombre_norm", how="right", indicator=True)
    return merged

def _escribir_hoja_conciliacion(ws, merged, fecha=""):
    # Sin columna Cta. Bancaria — siempre: Nombre | Nota | Cta. Contable | Importe | Fecha
    headers = ["Nombre", "Nota", "Cta. Contable", "Importe", "Fecha"]
    imp_col = 4
    style_header(ws, headers)
    row_idx = 2
    for _, r in merged.iterrows():
        status = str(r.get("_merge", ""))
        importe = r.get("importe", 0) or 0
        if status == "both":
            fill, label = FILL_OK, "Conciliado"
        else:
            fill, label = FILL_ONLY, "Sin poliza"
        nombre   = r.get("nombre", "") or ""
        cta_cont = r.get("Cuenta", "") or ""
        vals = [nombre, label, cta_cont, importe, fecha]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=c, value=v)
            cell.font   = Font(name="Arial")
            cell.border = BORDER
            cell.fill   = PatternFill("solid", start_color=fill)
            if c == imp_col:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            if c == imp_col + 1:   # Fecha
                cell.alignment = Alignment(horizontal="center")
        row_idx += 1
    n_cols = len(headers)
    imp_col_letter = chr(ord("A") + imp_col - 1)
    ws.cell(row=row_idx, column=1, value="TOTAL").font = Font(bold=True, name="Arial")
    ws.cell(row=row_idx, column=imp_col,
            value=f"=SUM({imp_col_letter}2:{imp_col_letter}{row_idx-1})").font = Font(bold=True, name="Arial")
    ws.cell(row=row_idx, column=imp_col).number_format = "#,##0.00"
    for c in range(1, n_cols + 1):
        ws.cell(row=row_idx, column=c).border = BORDER
        ws.cell(row=row_idx, column=c).fill = PatternFill("solid", start_color=TOTAL_FILL)
    autosize(ws, [34, 16, 16, 14, 14])

def escribir_conciliacion(excel_map, pdf_rows, meta, out_path):
    merged = _construir_merge_conciliacion(excel_map, pdf_rows)
    wb = Workbook(); ws = wb.active; ws.title = "Conciliacion"
    _escribir_hoja_conciliacion(ws, merged)
    wb.save(out_path)

def _safe_sheet_name(name, used):
    name = re.sub(r"[\\/*?:\[\]]","_", str(name))[:31]
    base, n = name, 1
    while name in used:
        suffix = f"_{n}"; name = base[:31-len(suffix)] + suffix; n += 1
    used.add(name); return name

def _escribir_hoja_complemento(ws, rows, excel_map=None, meta=None):
    _em = _resolver_excel_df(excel_map) if excel_map is not None else None
    fecha = ""
    if meta:
        fecha = meta.get("fecha_aplicacion") or meta.get("fecha_operacion") or ""
    style_header(ws, ["Titular","Cta. Contable","Monto","Fecha","Motivo"])
    row_idx = 2
    for r in rows:
        titular = r.get("titular","") or ""
        cta = ""
        if _em is not None:
            nn = norm(titular)
            cta = _em.at[nn, "Cuenta"] if nn in _em.index else ""
        fill = FILL_OK if cta else FILL_COMPLEMENTO
        vals = [titular, cta, r.get("monto",0), fecha, r.get("motivo","")]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=c, value=v)
            cell.font = Font(name="Arial"); cell.border = BORDER
            cell.fill = PatternFill("solid", start_color=fill)
            if c==3: cell.number_format="#,##0.00"; cell.alignment=Alignment(horizontal="right")
            if c==4: cell.alignment=Alignment(horizontal="center")
        row_idx += 1
    ws.cell(row=row_idx,column=1,value="TOTAL").font=Font(bold=True,name="Arial")
    ws.cell(row=row_idx,column=3,value=f"=SUM(C2:C{row_idx-1})").font=Font(bold=True,name="Arial")
    ws.cell(row=row_idx,column=3).number_format="#,##0.00"
    for c in range(1,6):
        ws.cell(row=row_idx,column=c).border=BORDER
        ws.cell(row=row_idx,column=c).fill=PatternFill("solid",start_color=TOTAL_FILL)
    autosize(ws,[32,16,14,14,28])

def escribir_complementos(pdf_paths, out_path):
    wb = Workbook(); used = set(); first = True
    for path in pdf_paths:
        meta, rows = parse_pagos_transferencias(path)
        nm = _safe_sheet_name((meta.get("descripcion") or "Compl")[:28], used)
        ws = wb.active if first else wb.create_sheet(nm)
        if first: ws.title = nm; first = False
        _escribir_hoja_complemento(ws, rows)
    wb.save(out_path)

def _recolectar_vacaciones(pdf_paths):
    semanas = []
    for path in pdf_paths:
        tipo = detect_template(path)
        if tipo == "dispersion":
            meta, rows = parse_dispersion(path)
            registros = [(r["nombre"], r["cuenta_abono"], r["importe"]) for r in rows]
            fecha = meta.get("fecha_operacion",""); folio = meta.get("folio","")
        else:
            meta, rows = parse_pagos_transferencias(path)
            registros = [(r["titular"], r["banco_destino"], r["monto"]) for r in rows]
            fecha = meta.get("fecha_aplicacion",""); folio = meta.get("folio","")
        semanas.append((fecha, folio, registros))
    return semanas

def _escribir_hoja_vacaciones(ws, registros, fecha, folio, excel_map):
    _em = _resolver_excel_df(excel_map)
    style_header(ws, ["Nombre","Cta. Contable","Importe","Fecha"])
    row_idx = 2
    for nombre, _cuenta_pdf, importe in registros:
        nn = norm(nombre)
        cuenta = _em.at[nn,"Cuenta"] if nn in _em.index else "Sin coincidencia"
        fill = FILL_OK if cuenta != "Sin coincidencia" else FILL_ONLY
        ws.cell(row=row_idx,column=1,value=nombre).font=Font(name="Arial")
        ws.cell(row=row_idx,column=2,value=cuenta).font=Font(name="Arial")
        ws.cell(row=row_idx,column=3,value=round(float(importe or 0),2)).number_format="#,##0.00"
        ws.cell(row=row_idx,column=4,value=fecha).font=Font(name="Arial")
        for c in range(1,5):
            cell=ws.cell(row=row_idx,column=c); cell.border=BORDER
            cell.fill=PatternFill("solid",start_color=fill)
            if c==3: cell.alignment=Alignment(horizontal="right")
            if c==4: cell.alignment=Alignment(horizontal="center")
        row_idx += 1
    ws.cell(row=row_idx,column=1,value="VACACIONES").font=Font(bold=True,name="Arial")
    ws.cell(row=row_idx,column=3,value=f"=SUM(C2:C{row_idx-1})").font=Font(bold=True,name="Arial")
    ws.cell(row=row_idx,column=3).number_format="#,##0.00"
    for c in range(1,5):
        ws.cell(row=row_idx,column=c).border=BORDER
        ws.cell(row=row_idx,column=c).fill=PatternFill("solid",start_color=TOTAL_FILL)
    autosize(ws,[32,20,14,14])

def escribir_vacaciones(excel_map, pdf_paths, out_path):
    semanas = _recolectar_vacaciones(pdf_paths)
    wb = Workbook(); ws_res = wb.active; ws_res.title = "Resumen"
    used = {"Resumen"}
    for fecha, folio, registros in semanas:
        sname = _safe_sheet_name(f"Vac {fecha}"[:28], used)
        ws = wb.create_sheet(sname)
        _escribir_hoja_vacaciones(ws, registros, fecha, folio, excel_map)
    wb.save(out_path)

def escribir_todo_en_uno(excel_map, nomina_paths, complementos_paths, vacaciones_paths, out_path, prestamos_paths=None):
    wb = Workbook(); ws_res = wb.active; ws_res.title = "Resumen general"
    used = {"Resumen general"}; resumen_rows = []
    _cat_empleados = excel_map.get("empleados", excel_map) if isinstance(excel_map, dict) else excel_map
    _cat_prestamos = excel_map.get("prestamos", _cat_empleados) if isinstance(excel_map, dict) else excel_map

    for path in nomina_paths:
        meta, rows = parse_dispersion(path)
        sname = _safe_sheet_name(f"Concil {meta.get('descripcion','')}"[:28], used)
        ws = wb.create_sheet(sname)
        _escribir_hoja_conciliacion(ws, _construir_merge_conciliacion(_cat_empleados, rows), fecha=meta.get('fecha_operacion',''))
        resumen_rows.append(("Nomina", meta.get("descripcion",""), meta.get("fecha_operacion",""), len(rows), sum(r["importe"] for r in rows)))

    for path in complementos_paths:
        meta, rows = parse_pagos_transferencias(path)
        sname = _safe_sheet_name(f"Compl {meta.get('descripcion','')}"[:28], used)
        ws = wb.create_sheet(sname)
        _escribir_hoja_complemento(ws, rows, excel_map=_cat_empleados, meta=meta)
        resumen_rows.append(("Complemento", meta.get("descripcion",""), meta.get("fecha_aplicacion",""), len(rows), sum(r["monto"] for r in rows)))

    for fecha, folio, registros in _recolectar_vacaciones(vacaciones_paths):
        sname = _safe_sheet_name(f"Vac {fecha}"[:28], used)
        ws = wb.create_sheet(sname)
        _escribir_hoja_vacaciones(ws, registros, fecha, folio, _cat_empleados)
        resumen_rows.append(("Vacaciones", f"Semana al {fecha}", fecha, len(registros), sum(r[2] for r in registros)))

    for path in (prestamos_paths or []):
        try:
            tipo = detect_template(path)
            meta, rows = parse_dispersion(path) if tipo == "dispersion" else parse_pagos_transferencias(path)
            sname = _safe_sheet_name(f"Prest {meta.get('descripcion','')}"[:28], used)
            ws = wb.create_sheet(sname)
            _escribir_hoja_conciliacion(ws, _construir_merge_conciliacion(_cat_prestamos, rows), fecha=meta.get('fecha_operacion', meta.get('fecha_aplicacion','')))
            total = sum(r.get("importe", r.get("monto", 0)) for r in rows)
            resumen_rows.append(("Prestamo", meta.get("descripcion",""), meta.get("fecha_operacion", meta.get("fecha_aplicacion","")), len(rows), total))
        except Exception:
            pass

    style_header(ws_res, ["Tipo","Descripcion","Fecha","Registros","Total"])
    row = 2
    for tipo, desc, fecha, n_reg, total in resumen_rows:
        ws_res.cell(row=row,column=1,value=tipo).font=Font(name="Arial")
        ws_res.cell(row=row,column=2,value=desc).font=Font(name="Arial")
        ws_res.cell(row=row,column=3,value=fecha).font=Font(name="Arial")
        ws_res.cell(row=row,column=4,value=n_reg).alignment=Alignment(horizontal="center")
        ws_res.cell(row=row,column=5,value=round(total,2)).number_format="#,##0.00"
        for c in range(1,6):
            ws_res.cell(row=row,column=c).border=BORDER
            ws_res.cell(row=row,column=c).fill=PatternFill("solid",start_color="D9E1F2")
            ws_res.cell(row=row,column=c).font=Font(name="Arial")
        row += 1
    ws_res.cell(row=row,column=1,value="TOTAL GENERAL").font=Font(bold=True,name="Arial")
    ws_res.cell(row=row,column=5,value=f"=SUM(E2:E{row-1})").font=Font(bold=True,name="Arial")
    ws_res.cell(row=row,column=5).number_format="#,##0.00"
    for c in range(1,6):
        ws_res.cell(row=row,column=c).border=BORDER
        ws_res.cell(row=row,column=c).fill=PatternFill("solid",start_color=TOTAL_FILL)
    autosize(ws_res,[18,32,14,12,16])
    wb.move_sheet(ws_res.title, offset=-len(wb.sheetnames))
    wb.save(out_path)

# ------------------------------------------------------------------ #
# Pagos Bancarios: nombre + N de cuenta + importe (3 columnas)
# ------------------------------------------------------------------ #

def _es_prestamo_texto(texto):
    """Devuelve True si el texto (normalizado) contiene 'prestamo'."""
    t = unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode().lower()
    return "prestamo" in t

def _extraer_filas_pdf(path):
    tipo = detect_template(path)
    if tipo == "dispersion":
        meta, rows = parse_dispersion(path)
        # Toda la dispersión es préstamo si la descripción lo indica
        desc_global = meta.get("descripcion", "") or os.path.basename(path)
        es_prest_global = _es_prestamo_texto(desc_global)
        filas = [{"nombre": r["nombre"], "nombre_alt": "", "importe": r["importe"],
                  "es_prestamo": es_prest_global} for r in rows]
    else:
        meta, rows = parse_pagos_transferencias(path)
        filas = []
        for r in rows:
            motivo = r["motivo"].strip()
            es_prest = _es_prestamo_texto(motivo)
            filas.append({"nombre": r["titular"].strip() or motivo,
                          "nombre_alt": motivo,
                          "importe": r["monto"],
                          "es_prestamo": es_prest})
    return meta, filas, tipo

def _resolver_nombre_cuenta(nombre, nombre_alt, catalogo, es_prestamo=False):
    """Busca la cuenta del empleado en el catálogo correcto.
    catalogo puede ser dict plano {nombre_norm: cuenta}
    o dict estructurado {"empleados": {...}, "prestamos": {...}}."""
    if isinstance(catalogo, dict) and ("empleados" in catalogo or "prestamos" in catalogo):
        sub = catalogo.get("prestamos" if es_prestamo else "empleados", {})
    else:
        sub = catalogo  # retrocompatibilidad
    n_cuenta = sub.get(norm(nombre), "")
    if n_cuenta: return nombre, n_cuenta
    if nombre_alt:
        n2 = sub.get(norm(nombre_alt), "")
        if n2: return nombre_alt, n2
    return nombre, ""

def _escribir_hoja_pagos(ws, filas, catalogo, meta, desc_extra=""):
    fecha_op = meta.get("fecha_operacion") or meta.get("fecha_aplicacion") or ""
    style_header(ws, ["Nombre del empleado", "N° de cuenta", "Importe", "Fecha de operación"])
    row_idx = 2
    for r in filas:
        importe = r["importe"]
        es_prestamo = r.get("es_prestamo", False)
        nombre, n_cuenta = _resolver_nombre_cuenta(r["nombre"], r.get("nombre_alt",""), catalogo, es_prestamo)
        fill = FILL_OK if n_cuenta else FILL_ONLY
        ws.cell(row=row_idx,column=1,value=nombre).font=Font(name="Arial")
        ws.cell(row=row_idx,column=2,value=n_cuenta).font=Font(name="Arial")
        ws.cell(row=row_idx,column=3,value=round(importe,2)).number_format="#,##0.00"
        ws.cell(row=row_idx,column=4,value=fecha_op).font=Font(name="Arial")
        for c in range(1,5):
            cell=ws.cell(row=row_idx,column=c); cell.border=BORDER
            cell.fill=PatternFill("solid",start_color=fill)
            if c==3: cell.alignment=Alignment(horizontal="right")
            if c==4: cell.alignment=Alignment(horizontal="center")
        row_idx += 1
    ws.cell(row=row_idx,column=1,value="TOTAL").font=Font(bold=True,name="Arial")
    ws.cell(row=row_idx,column=3,value=f"=SUM(C2:C{row_idx-1})").font=Font(bold=True,name="Arial")
    ws.cell(row=row_idx,column=3).number_format="#,##0.00"
    for c in range(1,5):
        ws.cell(row=row_idx,column=c).border=BORDER
        ws.cell(row=row_idx,column=c).fill=PatternFill("solid",start_color=TOTAL_FILL)
    autosize(ws,[36,18,16,16])
    nota_row = row_idx + 2
    desc  = meta.get("descripcion") or ""
    folio = meta.get("folio") or ""
    info  = f"{desc}  |  Fecha: {fecha_op}  |  Folio: {folio}"
    if desc_extra: info = f"{desc_extra}  --  {info}"
    ws.cell(row=nota_row,column=1,value=info).font=Font(italic=True,color="666666",name="Arial",size=8)

def escribir_pagos_bancarios(filas, meta, catalogo, out_path, desc_extra=""):
    wb = Workbook(); ws = wb.active; ws.title = "Pagos Bancarios"
    _escribir_hoja_pagos(ws, filas, catalogo, meta, desc_extra)
    wb.save(out_path)

def escribir_pagos_bancarios_todo(pdf_paths, catalogo, out_path):
    wb = Workbook(); used_names = set(); resumen_rows = []
    ws_resumen = wb.active; ws_resumen.title = "Resumen"; used_names.add("Resumen")
    for path in pdf_paths:
        meta, filas, _ = _extraer_filas_pdf(path)
        desc = meta.get("descripcion") or os.path.splitext(os.path.basename(path))[0]
        sname = _safe_sheet_name(desc[:28], used_names)
        ws = wb.create_sheet(sname)
        _escribir_hoja_pagos(ws, filas, catalogo, meta)
        total = sum(r["importe"] for r in filas)
        resumen_rows.append((desc, sname,
                             meta.get("fecha_operacion") or meta.get("fecha_aplicacion") or "",
                             len(filas), total))
    style_header(ws_resumen, ["Descripcion","Hoja","Fecha","Registros","Importe total"])
    row = 2
    for desc, hoja, fecha, n_reg, total in resumen_rows:
        ws_resumen.cell(row=row,column=1,value=desc).font=Font(name="Arial")
        ws_resumen.cell(row=row,column=2,value=hoja).font=Font(name="Arial")
        ws_resumen.cell(row=row,column=3,value=fecha).font=Font(name="Arial")
        ws_resumen.cell(row=row,column=4,value=n_reg).alignment=Alignment(horizontal="center")
        ws_resumen.cell(row=row,column=5,value=round(total,2)).number_format="#,##0.00"
        for c in range(1,6):
            cell=ws_resumen.cell(row=row,column=c); cell.border=BORDER
            cell.fill=PatternFill("solid",start_color="D9E1F2"); cell.font=Font(name="Arial")
        row += 1
    ws_resumen.cell(row=row,column=1,value="TOTAL GENERAL").font=Font(bold=True,name="Arial")
    ws_resumen.cell(row=row,column=5,value=f"=SUM(E2:E{row-1})").font=Font(bold=True,name="Arial")
    ws_resumen.cell(row=row,column=5).number_format="#,##0.00"
    for c in range(1,6):
        ws_resumen.cell(row=row,column=c).border=BORDER
        ws_resumen.cell(row=row,column=c).fill=PatternFill("solid",start_color=TOTAL_FILL)
    autosize(ws_resumen,[36,26,14,10,16])
    wb.move_sheet(ws_resumen.title, offset=-len(wb.sheetnames))
    wb.save(out_path)
