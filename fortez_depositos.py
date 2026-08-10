"""
DEPOSITOS BANCARIOS FORTEZ
Genera póliza Excel a partir de plantilla (hoja CUENTAS) + estado de cuenta.
"""
import openpyxl
from openpyxl.styles import (Font, Alignment, PatternFill, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from collections import defaultdict
import re, sys, os

# ── Aliases adicionales (palabras clave que no están en el nombre de cuenta) ──
# key = palabra clave del nombre de cuenta (UPPERCASE), value = lista de alias
_EXTRA_ALIASES = {
    "BANORTE":    ["SUPER SERV VALLEJO", "SUPER SERV PARIFERICO", "SUPER SERV PERIFÉRICO",
                   "SUPER SERV PERIFERICO", "SERV MENA", "BANORTE"],
    "EDENRED":    ["EDENRED", "EDENR"],
    "PLUXE":      ["PLUXEE", "PLUXE"],
    "EFECTIVALE": ["EFECTIVALE", "DEL CLIENTE EFECTI"],
    "EFECTIVO":   ["DEP.EFECTIVO", "DEPOSITO EN EFECTIVO", "COMETRO"],
    "AMERICAN EXPRESS": ["AMERICAN EXPRESS", "AMERICAN"],
}

def _extraer_keyword(nombre: str) -> str:
    """Saca el identificador principal del nombre de cuenta."""
    n = nombre.strip()
    for pref in ["Depositos en transito por ", "Deposito en transito TARJETA ",
                 "Deposito en transito por ", "Deposito en transito "]:
        if n.lower().startswith(pref.lower()):
            return n[len(pref):].strip().upper()
    return n.strip().upper()

def _leer_cuentas(ws_cuentas):
    """Retorna (cargos_list, abonos_list)."""
    cargos, abonos = [], []
    r = 3
    while ws_cuentas.cell(r, 1).value:
        cargos.append({
            "num":   str(ws_cuentas.cell(r, 1).value).strip(),
            "banco": str(ws_cuentas.cell(r, 2).value or "").strip(),
        })
        r += 1
    r = 3
    while ws_cuentas.cell(r, 4).value:
        nombre = str(ws_cuentas.cell(r, 5).value or "").strip()
        kw = _extraer_keyword(nombre)
        aliases = _EXTRA_ALIASES.get(kw, [kw])
        abonos.append({
            "num":     str(ws_cuentas.cell(r, 4).value).strip(),
            "nombre":  nombre,
            "kw":      kw,
            "aliases": aliases,
        })
        r += 1
    return cargos, abonos

def _clasificar(desc: str, abonos: list):
    """Retorna índice del abono que coincide, o None."""
    du = str(desc).upper()
    # Ordenar por longitud del alias más largo (mayor especificidad primero)
    ranked = sorted(enumerate(abonos),
                    key=lambda x: max(len(a) for a in x[1]["aliases"]),
                    reverse=True)
    for idx, ab in ranked:
        for alias in sorted(ab["aliases"], key=len, reverse=True):
            if alias.upper() in du:
                return idx
    return None

# ── Estilos ───────────────────────────────────────────────────────────────────
_AZUL_OSC  = "1E3A8A"
_AZUL_MED  = "2563EB"
_AZUL_CLAR = "DBEAFE"
_VERDE_CLAR = "D1FAE5"
_NARANJA_CLR = "FEF3C7"
_GRIS_CLR   = "F3F4F6"
_BLANCO     = "FFFFFF"
_ROJO       = "DC2626"

def _fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def _font(bold=False, color="000000", sz=10):
    return Font(bold=bold, color=color, size=sz)
def _align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _fmt_num(ws, r, c, val):
    cell = ws.cell(r, c)
    cell.value = val
    cell.number_format = '#,##0.00'
    cell.alignment = _align("right", "center", False)
    cell.border = _border()

def generar_poliza_fortez(
        ruta_plantilla: str,
        ruta_estado_cuenta: str,
        ruta_salida: str,
        banco_override: str = None,  # "BANORTE" | "BBVA" | None=auto
):
    # 1. Leer CUENTAS ─────────────────────────────────────────────────────────
    wb_tmpl = openpyxl.load_workbook(ruta_plantilla)
    if "CUENTAS" not in wb_tmpl.sheetnames:
        hojas = ", ".join(wb_tmpl.sheetnames)
        raise ValueError(
            f"La plantilla no contiene la hoja 'CUENTAS'. "
            f"Hojas encontradas: {hojas}"
        )
    cargos, abonos = _leer_cuentas(wb_tmpl["CUENTAS"])
    # Banco desde la primera cuenta cargo
    banco_carta = cargos[0]["banco"].upper() if cargos else "BANORTE"
    banco = (banco_override or banco_carta).upper()
    banco_str = f"DEPOSITOS DE {banco}"

    # 2. Leer estado de cuenta ────────────────────────────────────────────────
    wb_ec = openpyxl.load_workbook(ruta_estado_cuenta)
    ws_mv = wb_ec["Movimientos"]
    movs = []
    for r in range(3, ws_mv.max_row + 1):
        fecha = ws_mv.cell(r, 1).value
        desc  = ws_mv.cell(r, 2).value
        dep   = ws_mv.cell(r, 3).value
        if not isinstance(fecha, datetime) or not dep:
            continue
        movs.append((fecha.date(), str(desc or ""), float(dep)))

    # 3. Clasificar ───────────────────────────────────────────────────────────
    rows = [(f, d, m, _clasificar(d, abonos)) for f, d, m in movs]

    # 4. Construir Excel ──────────────────────────────────────────────────────
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "POLIZA"

    nc = len(cargos)
    na = len(abonos)
    # Mapa de columnas (1-based)
    C_TIPO   = 1; C_FECHA  = 2; C_REF   = 3; C_CONC   = 4
    C_ERR    = 5; C_UIDD   = 6; C_NUMPOL = 7; C_PROC  = 8
    C_CARGO_0 = 9
    C_TOTAL_C = C_CARGO_0 + nc          # columna TOTAL CARGOS
    C_ABONO_0 = C_TOTAL_C + 1           # primera columna de abono
    C_TOTAL_A = C_ABONO_0 + na          # columna TOTAL ABONOS
    C_CONC_F  = C_TOTAL_A + 1           # CONCILIACION
    TOTAL_COLS = C_CONC_F

    # ── ROW 1: numeración 0-based ─────────────────────────────────────────────
    for c in range(1, TOTAL_COLS + 1):
        cell = ws.cell(1, c)
        cell.value = c - 1
        cell.font = _font(bold=True, color=_BLANCO, sz=9)
        cell.fill = _fill(_AZUL_OSC)
        cell.alignment = _align()
        cell.border = _border()

    # ── ROW 2: números de cuenta (solo en cols de cuentas) ───────────────────
    for c in range(1, TOTAL_COLS + 1):
        cell = ws.cell(2, c)
        cell.border = _border()
        cell.font = _font(bold=True, sz=9)
        cell.alignment = _align()
        if C_CARGO_0 <= c <= C_CARGO_0 + nc - 1:
            cell.value = cargos[c - C_CARGO_0]["num"]
            cell.fill = _fill(_AZUL_CLAR)
        elif C_TOTAL_C == c:
            cell.fill = _fill(_NARANJA_CLR)
        elif C_ABONO_0 <= c <= C_ABONO_0 + na - 1:
            cell.value = abonos[c - C_ABONO_0]["num"]
            cell.fill = _fill(_VERDE_CLAR)
        elif C_TOTAL_A == c:
            cell.fill = _fill(_NARANJA_CLR)
        else:
            cell.fill = _fill(_GRIS_CLR)

    # ── ROW 3: encabezados ────────────────────────────────────────────────────
    hdrs_fijo = ["TIPO DE POLIZA","Fecha","REFERENCIA","CONCEPTO",
                 "ERROR","UIDD","NUM POLIZA","PROCESADO"]
    for c, h in enumerate(hdrs_fijo, start=1):
        cell = ws.cell(3, c)
        cell.value = h
        cell.font = _font(bold=True, color=_BLANCO, sz=9)
        cell.fill = _fill(_AZUL_MED)
        cell.alignment = _align(wrap=True)
        cell.border = _border()
    for c in range(C_CARGO_0, C_CARGO_0 + nc):
        cell = ws.cell(3, c)
        cell.value = cargos[c - C_CARGO_0]["banco"]
        cell.font = _font(bold=True, sz=9)
        cell.fill = _fill(_AZUL_CLAR)
        cell.alignment = _align(wrap=True)
        cell.border = _border()
    cell = ws.cell(3, C_TOTAL_C)
    cell.value = "TOTAL CARGOS"; cell.font = _font(bold=True, sz=9)
    cell.fill = _fill(_NARANJA_CLR); cell.alignment = _align(wrap=True); cell.border = _border()
    for c in range(C_ABONO_0, C_ABONO_0 + na):
        cell = ws.cell(3, c)
        cell.value = abonos[c - C_ABONO_0]["nombre"]
        cell.font = _font(bold=True, sz=9)
        cell.fill = _fill(_VERDE_CLAR)
        cell.alignment = _align(wrap=True)
        cell.border = _border()
    cell = ws.cell(3, C_TOTAL_A)
    cell.value = "TOTAL ABONOS"; cell.font = _font(bold=True, sz=9)
    cell.fill = _fill(_NARANJA_CLR); cell.alignment = _align(wrap=True); cell.border = _border()
    cell = ws.cell(3, C_CONC_F)
    cell.value = "CONCILIACION"; cell.font = _font(bold=True, color=_BLANCO, sz=9)
    cell.fill = _fill(_AZUL_OSC); cell.alignment = _align(wrap=True); cell.border = _border()

    # ── ROWS 4+: datos (solo filas con match; sin match se omiten) ──────────
    fila_datos = 4   # contador de fila real en la hoja
    row_parity = 0   # para alternar color sin huecos

    def cel(ws_ref, r, col, val=None, bold=False, align_h="left", num_fmt=None, bg_col=None):
        cell = ws_ref.cell(r, col)
        cell.value = val
        cell.font = _font(bold=bold, sz=9)
        cell.fill = _fill(bg_col or _BLANCO)
        cell.alignment = _align(align_h, "center", False)
        cell.border = _border()
        if num_fmt:
            cell.number_format = num_fmt

    for fecha, desc, monto, ab_idx in rows:
        if ab_idx is None:
            continue  # sin match → no se escribe fila

        r   = fila_datos
        bg  = _BLANCO if row_parity % 2 == 0 else _GRIS_CLR
        row_parity += 1
        fila_datos += 1

        cel(ws, r, C_TIPO,  "I", bold=True, align_h="center", bg_col=bg)
        cel(ws, r, C_FECHA, fecha, align_h="center", bg_col=bg)
        ws.cell(r, C_FECHA).number_format = "DD/MM/YYYY"
        cel(ws, r, C_REF,   banco_str, bg_col=bg)
        cel(ws, r, C_CONC,  banco_str, bg_col=bg)
        cel(ws, r, C_ERR,   None, bg_col=bg)
        cel(ws, r, C_UIDD,  None, bg_col=bg)
        cel(ws, r, C_NUMPOL,None, bg_col=bg)
        cel(ws, r, C_PROC,  None, bg_col=bg)

        # Cargo
        for c in range(C_CARGO_0, C_CARGO_0 + nc):
            _fmt_num(ws, r, c, monto)
            ws.cell(r, c).fill = _fill(_AZUL_CLAR if bg == _BLANCO else "C7D7F5")

        # TOTAL CARGOS
        _fmt_num(ws, r, C_TOTAL_C, monto)
        ws.cell(r, C_TOTAL_C).fill = _fill(_NARANJA_CLR if bg == _BLANCO else "FDE68A")

        # Abonos
        for c in range(C_ABONO_0, C_ABONO_0 + na):
            val = monto if (c - C_ABONO_0) == ab_idx else None
            _fmt_num(ws, r, c, val)
            ws.cell(r, c).fill = _fill(_VERDE_CLAR if bg == _BLANCO else "A7F3D0")

        # TOTAL ABONOS
        _fmt_num(ws, r, C_TOTAL_A, monto)
        ws.cell(r, C_TOTAL_A).fill = _fill(_NARANJA_CLR if bg == _BLANCO else "FDE68A")

        # CONCILIACION = 0
        _fmt_num(ws, r, C_CONC_F, None)
        ws.cell(r, C_CONC_F).fill = _fill(bg)

    # ── Anchos de columna ─────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(C_TIPO)].width   = 6
    ws.column_dimensions[get_column_letter(C_FECHA)].width  = 12
    ws.column_dimensions[get_column_letter(C_REF)].width    = 22
    ws.column_dimensions[get_column_letter(C_CONC)].width   = 22
    for c in [C_ERR, C_UIDD, C_NUMPOL, C_PROC]:
        ws.column_dimensions[get_column_letter(c)].width = 8
    for c in range(C_CARGO_0, C_CARGO_0 + nc):
        ws.column_dimensions[get_column_letter(c)].width = 15
    ws.column_dimensions[get_column_letter(C_TOTAL_C)].width = 15
    for c in range(C_ABONO_0, C_ABONO_0 + na):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.column_dimensions[get_column_letter(C_TOTAL_A)].width = 15
    ws.column_dimensions[get_column_letter(C_CONC_F)].width  = 14

    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 35

    # Fijar primeras 3 filas
    ws.freeze_panes = "A4"

    # ── Hoja RESUMEN ──────────────────────────────────────────────────────────
    wr = wb_out.create_sheet("RESUMEN")

    # Totales por cuenta abono
    from collections import defaultdict as _dd
    tot_por_cuenta = _dd(float)     # idx → monto
    cnt_por_cuenta = _dd(int)
    total_sin_cls_monto = 0.0
    total_sin_cls_cnt   = 0
    for _, _, monto, ab_idx in rows:
        if ab_idx is None:
            total_sin_cls_monto += monto
            total_sin_cls_cnt   += 1
        else:
            tot_por_cuenta[ab_idx] += monto
            cnt_por_cuenta[ab_idx] += 1

    # Encabezado RESUMEN
    hdrs_r = ["N° Cuenta", "Nombre de la cuenta", "Movimientos", "Importe"]
    for c, h in enumerate(hdrs_r, start=1):
        cell = wr.cell(1, c)
        cell.value = h
        cell.font = _font(bold=True, color=_BLANCO, sz=10)
        cell.fill = _fill(_AZUL_OSC)
        cell.alignment = _align()
        cell.border = _border()

    # Cargos (resumen de la cuenta cargo)
    grand_total_cargo = sum(tot_por_cuenta.values())
    grand_cnt_cargo   = sum(cnt_por_cuenta.values())

    fila_r = 2
    # Sección CARGOS
    for ci, cargo in enumerate(cargos):
        bg_c = _AZUL_CLAR
        wr.cell(fila_r, 1).value = cargo["num"]
        wr.cell(fila_r, 2).value = cargo["banco"]
        wr.cell(fila_r, 3).value = grand_cnt_cargo
        wr.cell(fila_r, 4).value = grand_total_cargo
        for c in range(1, 5):
            cell = wr.cell(fila_r, c)
            cell.font = _font(bold=True, sz=10)
            cell.fill = _fill(bg_c)
            cell.alignment = _align("left" if c == 2 else "center", "center", False)
            cell.border = _border()
        wr.cell(fila_r, 4).number_format = '#,##0.00'
        fila_r += 1

    # Separador
    fila_r += 1

    # Sección ABONOS
    for ai, ab in enumerate(abonos):
        monto_ab = tot_por_cuenta.get(ai, 0.0)
        cnt_ab   = cnt_por_cuenta.get(ai, 0)
        bg_a = _VERDE_CLAR
        wr.cell(fila_r, 1).value = ab["num"]
        wr.cell(fila_r, 2).value = ab["nombre"]
        wr.cell(fila_r, 3).value = cnt_ab
        wr.cell(fila_r, 4).value = monto_ab if monto_ab else None
        for c in range(1, 5):
            cell = wr.cell(fila_r, c)
            cell.font = _font(sz=10)
            cell.fill = _fill(bg_a)
            cell.alignment = _align("left" if c == 2 else "center", "center", False)
            cell.border = _border()
        wr.cell(fila_r, 4).number_format = '#,##0.00'
        fila_r += 1

    # Fila TOTAL ABONOS
    fila_r += 1
    for c, val in enumerate([None, "TOTAL ABONOS", grand_cnt_cargo, grand_total_cargo], start=1):
        cell = wr.cell(fila_r, c)
        cell.value = val
        cell.font = _font(bold=True, color=_BLANCO, sz=10)
        cell.fill = _fill(_AZUL_MED)
        cell.alignment = _align("left" if c == 2 else "center", "center", False)
        cell.border = _border()
    wr.cell(fila_r, 4).number_format = '#,##0.00'

    # Fila SIN CLASIFICAR (si los hay)
    if total_sin_cls_cnt:
        fila_r += 2
        for c, val in enumerate([None, "SIN CLASIFICAR", total_sin_cls_cnt, total_sin_cls_monto], start=1):
            cell = wr.cell(fila_r, c)
            cell.value = val
            cell.font = _font(bold=True, sz=10)
            cell.fill = _fill("FEE2E2")
            cell.alignment = _align("left" if c == 2 else "center", "center", False)
            cell.border = _border()
        wr.cell(fila_r, 4).number_format = '#,##0.00'

    wr.column_dimensions["A"].width = 20
    wr.column_dimensions["B"].width = 45
    wr.column_dimensions["C"].width = 14
    wr.column_dimensions["D"].width = 18
    wr.row_dimensions[1].height = 22

    wb_out.save(ruta_salida)

    # ── Resumen ───────────────────────────────────────────────────────────────
    total_monto = sum(m for _, _, m, _ in rows)
    clasificados = [(f, d, m, i) for f, d, m, i in rows if i is not None]
    sin_cls = [(f, d, m, i) for f, d, m, i in rows if i is None]
    total_cls = sum(m for _, _, m, _ in clasificados)

    print(f"\n{'='*60}")
    print(f"  DEPOSITOS BANCARIOS FORTEZ — Resumen")
    print(f"{'='*60}")
    print(f"  Total depósitos procesados : {len(rows):>6,}")
    print(f"  Clasificados               : {len(clasificados):>6,}  ${total_cls:>14,.2f}")
    print(f"  Sin clasificar             : {len(sin_cls):>6,}  ${total_monto - total_cls:>14,.2f}")
    print(f"  Importe total              :          ${total_monto:>14,.2f}")
    print()
    print("  Distribución por cuenta abono:")
    from collections import Counter
    dist = Counter(abonos[i]["nombre"][:40] if i is not None else "SIN CLASIFICAR"
                   for _, _, _, i in rows)
    for k, v in dist.most_common():
        total_k = sum(m for _, _, m, i in rows
                      if (abonos[i]["nombre"][:40] if i is not None else "SIN CLASIFICAR") == k)
        print(f"    {v:5,}  {k:<42}  ${total_k:>14,.2f}")
    if sin_cls:
        print(f"\n  Ejemplos sin clasificar:")
        for f, d, m, _ in sin_cls[:5]:
            print(f"    {f}  ${m:>10,.2f}  {d[:60]}")

    return len(rows), len(sin_cls)
