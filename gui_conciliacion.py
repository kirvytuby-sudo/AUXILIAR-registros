#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gui_conciliacion.py
====================
Interfaz grÃ¡fica con pestaÃ±as para ejecutar conciliacion_nomina.py sin
tener que escribir comandos en la terminal.

PestaÃ±as disponibles:
  1. ConciliaciÃ³n  â nÃ³mina principal, complementos, vacaciones
  2. PrÃ©stamos     â PDFs de prÃ©stamos detectados
  3. Reportes      â resumen consolidado y filtros por fecha
  4. ConfiguraciÃ³n â rutas por defecto, carpeta de salida
h
Requiere: pandas, openpyxl, pdfplumber
"""

import os
import json
import re
import sys
import threading
import traceback
from datetime import datetime

# Redirigir errores a log cuando corre sin consola (pythonw)
if sys.stderr is None or not hasattr(sys.stderr, 'write'):
    _log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "error_log.txt")
    sys.stderr = open(_log_path, "w", encoding="utf-8")

import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    _DND_OK = True
except ImportError:
    _DND_OK = False

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    import conciliacion_nomina as cn
except ImportError:
    cn = None


# ------------------------------------------------------------------ #
# Paleta de colores
# ------------------------------------------------------------------ #
COLOR_FONDO          = "#FFF0F5"   # canvas â azul cielo claro
COLOR_TARJETA        = "#FFFFFF"   # tarjeta â azul muy suave
COLOR_FUCSIA         = "#1E3A8A"   # denim â acento principal
COLOR_FUCSIA_OSCURO  = "#0F172A"   # ink â azul marino oscuro
COLOR_FUCSIA_SUAVE   = "#DBEAFE"   # line â borde azul suave
COLOR_FUCSIA_MID     = "#1D4ED8"   # rope â tono medio marino
COLOR_TEXTO          = "#0F172A"   # ink â texto principal
COLOR_BLANCO         = "#FFFFFF"
COLOR_OK             = "#1D4ED8"   # denim â Ã©xito en log
COLOR_ERROR          = "#E14B3D"   # coral â error en log
COLOR_AZUL           = "#1E3A8A"   # denim
COLOR_AZUL_SUAVE     = "#EFF6FF"   # canvas suave
COLOR_VERDE          = "#1D4ED8"   # rope
COLOR_VERDE_SUAVE    = "#DBEAFE"   # canvas deep
COLOR_NARANJA        = "#F43F8A"   # rust â bancos
COLOR_NARANJA_SUAVE  = "#FDF2F8"   # gold light


CATEGORIA_LABELS = {
    "nomina":        "NÃ³mina principal",
    "complementos":  "Complementos",
    "vacaciones":    "Vacaciones",
    "prestamos":     "PrÃ©stamos",
    "no_reconocido": "No reconocido",
    "error":         "No se pudo leer",
}


# ------------------------------------------------------------------ #
# Helpers
# ------------------------------------------------------------------ #

def clasificar_pdf(path):
    """Devuelve (categoria, descripcion_legible) para un PDF."""
    try:
        tipo = cn.detect_template(path)
    except Exception:
        return "no_reconocido", os.path.basename(path)

    try:
        if tipo == "dispersion":
            meta, _ = cn.parse_dispersion(path)
        else:
            meta, _ = cn.parse_pagos_transferencias(path)
    except Exception:
        return "error", os.path.basename(path)

    desc = (meta.get("descripcion") or "").upper()
    if "VAC" in desc:
        categoria = "vacaciones"
    elif "PRESTAMO" in desc or "PRÃSTAMO" in desc:
        categoria = "prestamos"
    elif tipo == "dispersion":
        categoria = "nomina"
    else:
        categoria = "complementos"

    descripcion = meta.get("descripcion") or os.path.basename(path)
    return categoria, descripcion


def sanitizar_nombre(texto, default="archivo"):
    texto = re.sub(r'[\\/*?:"<>|]', "", str(texto or default)).strip()
    texto = re.sub(r"\s+", "_", texto)
    return texto[:60] or default


def _make_listbox(parent, bg_card):
    """Crea un Listbox con scrollbar y devuelve (frame, listbox)."""
    frame = tk.Frame(parent, bg=bg_card)
    sb = ttk.Scrollbar(frame, orient="vertical")
    lb = tk.Listbox(
        frame, selectmode="extended", yscrollcommand=sb.set, exportselection=False,
        bg=COLOR_BLANCO, fg=COLOR_TEXTO,
        selectbackground=COLOR_FUCSIA, selectforeground=COLOR_BLANCO,
        relief="flat", highlightthickness=1,
        highlightbackground=COLOR_FUCSIA_SUAVE, highlightcolor=COLOR_FUCSIA,
        font=("Segoe UI", 9),
    )
    sb.config(command=lb.yview)
    lb.pack(side="left", fill="both", expand=True)
    sb.pack(side="right", fill="y")
    return frame, lb


# ================================================================== #
# VENTANA DE TRABAJO
# ================================================================== #

def _dibujar_brujula(cv, size=38):
    """
    Intenta cargar logo.png desde la carpeta de la app (sin fondo blanco).
    Si no existe o falla, dibuja la brÃºjula vectorial de respaldo.
    """
    import os
    _DIR = os.path.dirname(os.path.abspath(__file__))
    # Intentar variantes de nombre (logo_ready.png > logo.png > logo.png.png)
    for _cand in ("logo_ready.png", "logo.png", "logo.png.png"):
        logo_path = os.path.join(_DIR, _cand)
        if os.path.isfile(logo_path):
            break
    if os.path.isfile(logo_path):
        try:
            from PIL import Image, ImageTk
            img = Image.open(logo_path).convert("RGBA")
            # Quitar fondo blanco/casi-blanco
            data = img.getdata()
            new_data = []
            for r2, g2, b2, a2 in data:
                if r2 > 238 and g2 > 238 and b2 > 238:
                    new_data.append((r2, g2, b2, 0))
                else:
                    new_data.append((r2, g2, b2, a2))
            img.putdata(new_data)
            img = img.resize((size, size), Image.LANCZOS)
            # Componer sobre fondo harbor
            bg = Image.new('RGBA', (size, size), (174, 214, 242, 255))
            bg.paste(img, mask=img.split()[3])
            photo = ImageTk.PhotoImage(bg)
            cv._logo_photo = photo  # evitar GC
            cv.create_image(size // 2, size // 2, image=photo, anchor='center')
            return
        except Exception:
            pass  # fallback a brÃºjula
    # ââ BrÃºjula vectorial de respaldo ââââââââââââââââââââââââââââââââ
    cx = cy = size // 2
    r = cx - 2
    cv.create_oval(2, 2, size - 2, size - 2, outline="#F2C572", width=1.5)
    cv.create_polygon(cx, cy - r + 2, cx + 4, cy - 3, cx - 4, cy - 3,
                      fill="#E14B3D", outline="")
    cv.create_polygon(cx, cy + r - 2, cx - 4, cy + 3, cx + 4, cy + 3,
                      fill="#4F46C7", outline="")
    cv.create_polygon(cx - r + 2, cy, cx - 3, cy - 4, cx - 3, cy + 4,
                      fill="#1E6FBF", outline="")
    cv.create_polygon(cx + r - 2, cy, cx + 3, cy + 4, cx + 3, cy - 4,
                      fill="#E8A23A", outline="")
    cv.create_oval(cx - 2, cy - 2, cx + 2, cy + 2,
                   fill="#0B2A3D", outline="#F2C572", width=1)


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# BARRA DE PROGRESO PERSONALIZADA  (estilo aventura marina)
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
class FunkyProgressBar(tk.Canvas):
    """Barra de progreso animada â bloques segmentados estilo aventura."""

    # Paleta: rojo oscuro â naranja â Ã¡mbar â dorado (cÃ­clico)
    _SEG_COLORS = [
        (204, 34,   0),   # rojo oscuro
        (255, 107,  0),   # naranja
        (255, 140,  0),   # naranja Ã¡mbar
        (255, 184,  0),   # amarillo Ã¡mbar
        (255, 215,  0),   # dorado
        (255, 140,  0),   # naranja Ã¡mbar
        (255, 107,  0),   # naranja
    ]
    _SEG_W   = 18
    _SEG_GAP = 3
    _STRIPE_W = 21   # alias para compatibilidad

    def __init__(self, master, maximum=100, height=44, **kw):
        kw.setdefault("highlightthickness", 0)
        kw.setdefault("bd", 0)
        # Obtener color de fondo del parent â compatible con tk Y ttk
        # (ttk.Frame no soporta cget("bg"), asÃ­ que se captura el error)
        _fallback_bg = "#CFE7F8"
        if hasattr(master, "cget"):
            try:
                _fallback_bg = master.cget("bg")
            except Exception:
                try:
                    _fallback_bg = master.cget("background")
                except Exception:
                    pass
        kw.setdefault("bg", _fallback_bg)
        super().__init__(master, height=height, **kw)
        self._maximum = float(maximum)
        self._value   = 0.0
        self._label   = "Loading..."
        self._after_id = None
        self._h = height
        self._stripe_offset = 0.0
        self.bind("<Configure>", lambda e: self._draw())

    # ââ ttk.Progressbar-compatible API ââââââââââââââââââââââââââââââââââââââââ
    def __getitem__(self, key):
        if key == "value":  return self._value
        if key == "maximum": return self._maximum
        raise KeyError(key)

    def __setitem__(self, key, value):
        if key == "value":
            self._value = float(value)
            self._draw()
        elif key == "maximum":
            self._maximum = float(value)
        else:
            raise KeyError(key)

    # ââ Dibujo ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    def _load_side_imgs(self, H):
        """Carga y escala pb_izquierda/derecha, cacheadas por altura."""
        from PIL import Image as _Img
        import os
        cache = FunkyProgressBar.__dict__.get("_side_cache", {})
        if H in cache:
            return cache[H]
        try:
            _d = os.path.dirname(os.path.abspath(__file__))
            def _load(name):
                p = os.path.join(_d, name)
                if not os.path.exists(p):
                    return None
                raw = _Img.open(p).convert("RGBA")
                iw = int(raw.width * H / raw.height)
                return raw.resize((iw, H), _Img.LANCZOS)
            result = (_load("pb_izquierda.png"), _load("pb_derecha.png"))
        except Exception:
            result = (None, None)
        FunkyProgressBar._side_cache = {**cache, H: result}
        return result

    def _draw(self):
        try:
            from PIL import Image, ImageDraw, ImageFont, ImageTk
        except ImportError:
            self._draw_fallback()
            return

        W = self.winfo_width() or 400
        H = self._h
        pct = min(100.0, max(0.0, self._value / self._maximum * 100))

        # ââ ImÃ¡genes laterales ââââââââââââââââââââââââââââââââââââââââââââ
        il, ir = self._load_side_imgs(H)
        lw = (il.width - 12) if il else 0   # ancho reservado izquierda (con solapado)
        rw = (ir.width - 12) if ir else 0   # ancho reservado derecha

        # ââ Track horizontal: ocupa la zona entre los personajes ââââââââââ
        # La barra ocupa el tercio central verticalmente
        bvt  = H // 3        # bar vertical top
        bvb  = H - H // 3   # bar vertical bottom
        bh   = bvb - bvt    # bar height

        img = Image.new("RGBA", (W, H), (0, 0, 0, 0))
        d   = ImageDraw.Draw(img)

        pad  = 2
        BL   = lw + pad          # borde izquierdo de la barra
        BR   = W - rw - pad      # borde derecho de la barra
        r_out = bh // 2 - pad
        r_in  = max(1, r_out - 3)

        # CÃ¡psula exterior oscura
        self._draw_rounded(d, BL, bvt+pad, BR, bvb-pad, r_out,
                           fill=(30, 34, 45, 255), outline=(60, 65, 80, 255))
        # Track interior casi negro
        tp = pad + 3
        self._draw_rounded(d, BL+tp, bvt+tp, BR-tp, bvb-tp, r_in,
                           fill=(15, 17, 23, 255), outline=None)

        # ââ Bloques segmentados âââââââââââââââââââââââââââââââââââââââââââ
        track_x0 = BL + tp
        track_x1 = BR - tp
        track_y0 = bvt + tp
        track_y1 = bvb - tp
        fill_x1  = track_x0 + int((track_x1 - track_x0) * pct / 100)
        trk_h    = track_y1 - track_y0
        sw = self._SEG_W
        sg = self._SEG_GAP
        nc = len(self._SEG_COLORS)

        if pct > 0 and fill_x1 > track_x0 + 2:
            seg_img = Image.new("RGBA", (W, H), (0, 0, 0, 0))
            sd      = ImageDraw.Draw(seg_img)
            off     = int(self._stripe_offset) % (sw + sg)
            x   = track_x0 - off
            idx = 0
            while x < fill_x1:
                x_end = min(x + sw, fill_x1)
                if x_end > track_x0 and x_end > x:
                    x0 = max(x, track_x0)
                    col = self._SEG_COLORS[idx % nc]
                    sd.rectangle([x0, track_y0+1, x_end-1, track_y1-1], fill=col+(255,))
                    br = tuple(min(255, int(c*1.4)) for c in col) + (180,)
                    sd.rectangle([x0, track_y0+1, x_end-1, track_y0+3], fill=br)
                    dk = tuple(int(c*0.55) for c in col) + (255,)
                    sd.rectangle([x0, track_y1-3, x_end-1, track_y1-1], fill=dk)
                idx += 1
                x += sw + sg

            mask = Image.new("L", (W, H), 0)
            md   = ImageDraw.Draw(mask)
            self._draw_rounded(md, track_x0+1, track_y0+1, fill_x1, track_y1-1,
                               r_in, fill=255, outline=None)
            img.paste(seg_img, mask=mask)

            shine = Image.new("RGBA", (W, H), (0, 0, 0, 0))
            shd   = ImageDraw.Draw(shine)
            self._draw_rounded(shd, track_x0+2, track_y0+2, fill_x1-2,
                               track_y0 + trk_h//3, max(1, r_in//2),
                               fill=(255, 255, 255, 45), outline=None)
            img = Image.alpha_composite(img, shine)

        # ââ Texto en pill navy oscuro âââââââââââââââââââââââââââââââââââââ
        d2 = ImageDraw.Draw(img)
        if pct >= 100:
            txt = "â  Listo"
            tc  = (120, 255, 120, 255)
        elif pct > 0:
            txt = f"Loading...  {int(pct)} %"
            tc  = (240, 240, 240, 255)
        else:
            txt = self._label
            tc  = (170, 170, 200, 255)

        try:
            font = ImageFont.truetype("segoeuib.ttf", max(9, bh - 6))
        except Exception:
            font = ImageFont.load_default()
        bb = d2.textbbox((0, 0), txt, font=font)
        tw, th = bb[2]-bb[0], bb[3]-bb[1]
        tx = (BL + BR - tw) // 2
        ty = (bvt + bvb - th) // 2 - 1
        pp = 10
        self._draw_rounded(d2, tx-pp, ty-4, tx+tw+pp, ty+th+4, 8,
                           fill=(20, 30, 55, 210), outline=None)
        d2.text((tx, ty), txt, fill=tc, font=font)

        # ââ Pegar personajes sobre la barra âââââââââââââââââââââââââââââââ
        if il:
            img.paste(il, (0, 0), il)
        if ir:
            img.paste(ir, (W - ir.width, 0), ir)

        # ââ Convertir a PhotoImage âââââââââââââââââââââââââââââââââââââââââ
        bg_hex = self.cget("bg")
        try:
            rbg = tuple(int(bg_hex.lstrip("#")[i:i+2], 16) for i in (0, 2, 4))
        except Exception:
            rbg = (207, 231, 248)
        bg_img = Image.new("RGB", (W, H), rbg)
        bg_img.paste(img, mask=img.split()[3])
        photo = ImageTk.PhotoImage(bg_img)
        self._photo = photo
        self.delete("all")
        self.create_image(0, 0, anchor="nw", image=photo)

    @staticmethod
    def _draw_rounded(draw, x0, y0, x1, y1, r, fill, outline=None):
        if x1 <= x0 or y1 <= y0:
            return
        r = max(0, min(r, (x1-x0)//2, (y1-y0)//2))
        lw = 2 if outline else 0
        draw.rectangle([x0+r, y0, x1-r, y1], fill=fill)
        draw.rectangle([x0, y0+r, x1, y1-r], fill=fill)
        for cx, cy in [(x0+r,y0+r),(x1-r,y0+r),(x0+r,y1-r),(x1-r,y1-r)]:
            draw.ellipse([cx-r, cy-r, cx+r, cy+r], fill=fill)
        if outline and lw:
            draw.rounded_rectangle([x0, y0, x1, y1], radius=r, outline=outline, width=lw)

    def _draw_fallback(self):
        """Fallback si PIL no estÃ¡ disponible."""
        W = self.winfo_width() or 400
        H = self._h
        pct = min(100.0, max(0.0, self._value / self._maximum * 100))
        self.delete("all")
        self.create_rectangle(0, 0, W, H, fill="#1A1A2E", outline="")
        fw = int((W-8) * pct / 100)
        if fw > 0:
            self.create_rectangle(4, 4, 4+fw, H-4, fill="#FF6B00", outline="")
        txt = f"Loading... {int(pct)} %" if pct > 0 else ""
        if txt:
            self.create_text(W//2, H//2, text=txt, fill="white",
                            font=("Segoe UI", 9, "bold"))

    # ââ AnimaciÃ³n âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    def animate_step(self):
        """Desliza bloques en cada tick de animaciÃ³n."""
        self._stripe_offset = (self._stripe_offset + 2) % (self._SEG_W + self._SEG_GAP)
        self._draw()


def _hex_adjust(hex_col, factor):
    """Aclara (factor>1) u oscurece (factor<1) un color hexadecimal."""
    r = max(0, min(255, int(int(hex_col[1:3], 16) * factor)))
    g = max(0, min(255, int(int(hex_col[3:5], 16) * factor)))
    b = max(0, min(255, int(int(hex_col[5:7], 16) * factor)))
    return f"#{r:02X}{g:02X}{b:02X}"


class WorkspaceWindow(tk.Toplevel):
    """Una ventana de trabajo independiente (empresa / proyecto)."""

    # CatÃ¡logo de mÃ³dulos disponibles: (clave, etiqueta_pestaÃ±a, mÃ©todo_constructor)
    MODULOS = [
        ("pagos_bancarios",  "  ð¼ Pagos Bancarios  ",    "_tab_conciliacion"),
        ("prestamos",        "  ð³ PrÃ©stamos  ",           "_tab_prestamos"),
        ("provision_nomina", "  ð ProvisiÃ³n de NÃ³mina  ", "_tab_provision_nomina"),
        ("ventas_dia",       "  â½ Ventas del DÃ­a  ",      "_tab_ventas_grupo"),
        ("estado_cuenta",    "  ð¦ Estado de Cuenta  ",   "_tab_estado_cuenta"),
        ("reconciliacion",   "  ð ReconciliaciÃ³n  ",      "_tab_reconciliacion"),
        ("visor",            "  ð Visor de Resultados  ", "_tab_visor"),
        ("configuracion",    "  â  ConfiguraciÃ³n  ",       "_tab_configuracion"),
        ("concilia_sat",     "  ð ConciliaciÃ³n SAT  ",    "_tab_concilia_sat"),
        ("conc_banco_aux",   "  ð Banco vs Auxiliar  ",   "_tab_conc_banco_aux"),
        ("depositos_bancarios", "  ð¦ DepÃ³sitos Bancarios  ", "_tab_depositos_bancarios"),
    ]

    # Color de acento por mÃ³dulo (orden visual del mockup RumboERP)
    _MODULE_COLORS = {
        "pagos_bancarios":  "#E8A23A",
        "prestamos":        "#D9527A",
        "provision_nomina": "#1E6FBF",
        "ventas_dia":       "#E14B3D",
        "vd_concilia":      "#B45309",
        "estado_cuenta":    "#C2611A",
        "reconciliacion":   "#2E7D32",
        "visor":            "#4F46C7",
        "configuracion":    "#8C3B73",
        "concilia_sat":     "#0A7A4A",
        "conc_banco_aux":   "#6B3FA0",
        "depositos_bancarios": "#1565C0",
    }

    # (emoji, etiqueta ribbon, color acento)
    _MODULE_INFO = {
        "pagos_bancarios":  ("ð¼", "PAGOS",      "#E8A23A"),
        "prestamos":        ("ð³", "PRÃSTAMOS",  "#D9527A"),
        "provision_nomina": ("ð", "PROVISIÃN",  "#1E6FBF"),
        "ventas_dia":       ("â½", "VENTAS",      "#E14B3D"),
        "vd_concilia":      ("ð", "DESP/VTA",   "#B45309"),
        "estado_cuenta":    ("ð¦", "BANCOS",      "#C2611A"),
        "reconciliacion":   ("ð", "RECONCILIA",  "#2E7D32"),
        "visor":            ("ð", "VISOR",       "#4F46C7"),
        "configuracion":    ("â",  "CONFIG",      "#8C3B73"),
        "concilia_sat":     ("ð", "CONCILIA",     "#0A7A4A"),
        "conc_banco_aux":   ("ð", "BCO/AUX",     "#6B3FA0"),
        "depositos_bancarios": ("ð¦", "DEPÃSITOS",   "#1565C0"),
    }

    def __init__(self, master, nombre="LA SANITARIA", modulos=None,
                 modulos_disponibles=None, on_modulos_change=None):
        super().__init__(master)
        self.nombre = nombre
        self.title("AUXILIAR DE REGISTROS")
        self.geometry("1100x780")
        self.minsize(900, 640)
        self.state("zoomed")
        self.configure(bg=COLOR_FONDO)

        # Todos los mÃ³dulos disponibles para todas las empresas
        all_keys = [k for k, _, _ in WorkspaceWindow.MODULOS]
        self._modulos_disponibles = all_keys

        # MÃ³dulos activos y callback para guardar cambios
        self._modulos_activos = [m for m in (modulos or []) if m in self._modulos_disponibles]
        self._tab_frames = {}   # key -> tab_id para poder quitar sin buscar por texto
        self._on_modulos_change = on_modulos_change

        # Variables de instancia (globales â backward compat con otros mÃ³dulos)
        self.carpeta    = tk.StringVar(value=os.getcwd())
        self.excel_path = tk.StringVar(value="")

        # Variables INDEPENDIENTES por mÃ³dulo
        self.carpeta_pagos      = tk.StringVar(value=os.getcwd())
        self.excel_path_pagos   = tk.StringVar(value="")
        self.carpeta_prestamos  = tk.StringVar(value=os.getcwd())
        self.prov_xmls  = []
        _plantilla_default = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            "PLANTILLA PARA NOMINA DA LA SANITARIA.xlsx")
        self.prov_plantilla = tk.StringVar(
            value=_plantilla_default if os.path.isfile(_plantilla_default) else ""
        )
        self.salida_dir     = tk.StringVar(value="")
        self.prefijo_salida = tk.StringVar(value="")

        self.archivos_por_categoria = {
            "nomina": {}, "complementos": {}, "vacaciones": {}, "prestamos": {}
        }
        # Inicializar widgets opcionales (solo existen si su tab estÃ¡ cargada)
        self.listboxes    = {}
        self.lb_prestamos = None
        self.otros_label  = None

        self._construir_ui()
        self.protocol("WM_DELETE_WINDOW", self._on_cerrar_workspace)

        if cn is None:
            self._log(
                "â   No encontrÃ© 'conciliacion_nomina.py' en esta carpeta.\n"
                "   Copia este archivo a la misma carpeta y vuelve a abrir la interfaz.",
                error=True,
            )

    # ---------------------------------------------------------------- #
    # HELPERS VISUALES (logo + indicadores de mÃ³dulo)
    # ---------------------------------------------------------------- #
    def _dibujar_brujula(self, cv, size=38):
        """Delega a la funciÃ³n de mÃ³dulo."""
        _dibujar_brujula(cv, size)

    def _make_tab_dot(self, color):
        """PhotoImage cuadrado 10Ã10 del color de acento del mÃ³dulo."""
        img = tk.PhotoImage(width=10, height=10)
        row = " ".join([color] * 10)
        for y in range(10):
            img.put("{" + row + "}", to=(0, y))
        return img

    # ---------------------------------------------------------------- #
    # RIBBON DE MÃDULOS
    # ---------------------------------------------------------------- #
    _RIBBON_BG     = "#111827"   # sidebar â dark navy
    _RIBBON_ACTIVE = "#1E3A8A"   # sidebar item activo â royal blue
    _RIBBON_HOVER  = "#1F2937"   # sidebar item hover

    def _rebuild_ribbon(self):
        """Reconstruye los botones de la sidebar (muestra todos los mÃ³dulos disponibles)."""
        if not hasattr(self, "_ribbon_frame") or not self._ribbon_frame.winfo_exists():
            return
        prev = getattr(self, "_ribbon_active", None)
        self._ribbon_active = None
        self._ribbon_btns = {}
        self._ribbon_photos = {}
        self._ribbon_set_fns = {}
        for w in self._ribbon_frame.winfo_children():
            w.destroy()
        for key in self._modulos_disponibles:
            info = self._MODULE_INFO.get(key)
            if info:
                self._ribbon_add_btn(key, *info)
        target = prev if prev in self._modulos_disponibles else (
            self._modulos_disponibles[0] if self._modulos_disponibles else None)
        if target:
            self._ribbon_select(target)

    def _ribbon_add_btn(self, key, emoji, label, color):
        """Ãtem de mÃ³dulo en la barra lateral izquierda (sidebar vertical)."""
        # Nombre completo desde MODULOS (label limpio, sin espacios ni emoji)
        full_name = next(
            (lbl.strip() for k, lbl, _ in self.MODULOS if k == key), label
        )
        BG_IDLE   = self._RIBBON_BG
        BG_ACTIVE = self._RIBBON_ACTIVE
        BG_HOVER  = self._RIBBON_HOVER
        FG_IDLE   = "#9CA3AF"
        FG_ACTIVE = "#FFFFFF"

        outer = tk.Frame(self._ribbon_frame, bg=BG_IDLE, cursor="hand2")
        outer.pack(fill="x", side="top")

        # Indicador lateral izquierdo (rosa cuando activo)
        accent = tk.Frame(outer, bg=BG_IDLE, width=3)
        accent.pack(side="left", fill="y")

        # Emoji del mÃ³dulo
        ico = tk.Label(outer, text=emoji, bg=BG_IDLE, fg=FG_IDLE,
                       font=("Segoe UI Emoji", 11), padx=8, pady=8)
        ico.pack(side="left")

        # Nombre del mÃ³dulo
        lbl_tx = tk.Label(outer, text=full_name, bg=BG_IDLE, fg=FG_IDLE,
                          font=("Segoe UI", 8, "bold"), anchor="w")
        lbl_tx.pack(side="left", fill="x", expand=True, pady=8)

        # Bindings de clic
        for w in (outer, ico, lbl_tx, accent):
            w.bind("<Button-1>", lambda e, k=key: self._ribbon_select(k))

        def _ent(e):
            if self._ribbon_active != key:
                for w in (outer, ico, lbl_tx, accent):
                    w.config(bg=BG_HOVER)
        def _lve(e):
            if self._ribbon_active != key:
                for w in (outer, ico, lbl_tx, accent):
                    w.config(bg=BG_IDLE)
        for w in (outer, ico, lbl_tx, accent):
            w.bind("<Enter>", _ent)
            w.bind("<Leave>", _lve)

        def _set_state(active):
            if active:
                for w in (outer, ico, lbl_tx):
                    w.config(bg=BG_ACTIVE)
                ico.config(fg=FG_ACTIVE)
                lbl_tx.config(fg=FG_ACTIVE)
                accent.config(bg="#F43F8A")     # indicador rosa
            else:
                for w in (outer, ico, lbl_tx, accent):
                    w.config(bg=BG_IDLE)
                ico.config(fg=FG_IDLE)
                lbl_tx.config(fg=FG_IDLE)

        self._ribbon_btns[key]    = (outer, ico, lbl_tx)
        self._ribbon_set_fns[key] = _set_state

    def _ribbon_select(self, key):
        """Clic en Ã­cono: navega al tab existente o lo crea si es la primera vez."""
        # Desactivar anterior en ribbon
        if self._ribbon_active and self._ribbon_active in self._ribbon_set_fns:
            try:
                self._ribbon_set_fns[self._ribbon_active](False)
            except Exception:
                pass
        self._ribbon_active = key
        # Activar Ã­cono
        if key in self._ribbon_set_fns:
            try:
                self._ribbon_set_fns[key](True)
            except Exception:
                pass
        # Verificar que la referencia guardada siga viva
        if key in self._tab_frames:
            if self._tab_frames[key] not in self.nb.tabs():
                del self._tab_frames[key]   # referencia muerta â permitir recreaciÃ³n
        # Crear tab solo si no existe todavÃ­a
        if key not in self._tab_frames:
            self._cargar_tab(key)
            # Registrar en mÃ³dulos activos para persistencia
            if key not in self._modulos_activos:
                self._modulos_activos.append(key)
                if self._on_modulos_change:
                    try:
                        self._on_modulos_change(self.nombre, self._modulos_activos)
                    except Exception:
                        pass
        # Navegar al tab
        self._asegurar_x_tabs()   # garantizar â en todos los tabs
        if key in self._tab_frames:
            try:
                self.nb.select(self._tab_frames[key])
            except Exception:
                pass

    # ---------------------------------------------------------------- #
    # CONSTRUCCIÃN PRINCIPAL
    # ---------------------------------------------------------------- #
    def _construir_ui(self):
        self._tab_dot_imgs = {}
        self._btn_cinta    = {}

        # ââ LAYOUT: sidebar izquierda + contenido derecho ââââââââââââ
        main = tk.Frame(self, bg=COLOR_FONDO)
        main.pack(fill="both", expand=True)

        # ââ SIDEBAR ââââââââââââââââââââââââââââââââââââââââââââââââââ
        sidebar = tk.Frame(main, bg=self._RIBBON_BG, width=178)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        # Header: logo brÃºjula + nombre empresa
        hdr_side = tk.Frame(sidebar, bg=self._RIBBON_BG)
        hdr_side.pack(fill="x")
        logo_cv = tk.Canvas(hdr_side, width=34, height=34,
                            bg=self._RIBBON_BG, highlightthickness=0)
        logo_cv.pack(side="left", padx=(12, 6), pady=12)
        self._dibujar_brujula(logo_cv, 34)
        emp_fr = tk.Frame(hdr_side, bg=self._RIBBON_BG)
        emp_fr.pack(side="left", fill="x", expand=True)
        tk.Label(emp_fr, text=self.nombre.upper(),
                 bg=self._RIBBON_BG, fg="#FFFFFF",
                 font=("Segoe UI", 9, "bold"), anchor="w").pack(anchor="w")
        tk.Label(emp_fr, text="Sistema auxiliar",
                 bg=self._RIBBON_BG, fg="#6B7280",
                 font=("Segoe UI", 7), anchor="w").pack(anchor="w")

        # Separador
        tk.Frame(sidebar, bg="#1F2937", height=1).pack(fill="x", padx=10)
        tk.Label(sidebar, text="MÃDULOS", bg=self._RIBBON_BG, fg="#6B7280",
                 font=("Segoe UI", 7, "bold"), anchor="w",
                 padx=16).pack(fill="x", pady=(10, 2))

        # Frame de mÃ³dulos (vertical)
        self._ribbon_active = None
        self._ribbon_btns   = {}
        self._ribbon_photos = {}
        self._ribbon_frame  = tk.Frame(sidebar, bg=self._RIBBON_BG)
        self._ribbon_frame.pack(fill="x")

        # Botones de acciÃ³n fijos al fondo de la sidebar
        def _side_action(txt, cmd):
            b = tk.Label(sidebar, text=txt, bg=self._RIBBON_BG, fg="#9CA3AF",
                         font=("Segoe UI", 8), anchor="w", padx=16, pady=6,
                         cursor="hand2")
            b.pack(fill="x", side="bottom")
            b.bind("<Button-1>", lambda e: cmd())
            b.bind("<Enter>",    lambda e, _b=b: _b.config(fg="#FFFFFF"))
            b.bind("<Leave>",    lambda e, _b=b: _b.config(fg="#9CA3AF"))
            return b
        tk.Frame(sidebar, bg="#1F2937", height=1).pack(
            fill="x", padx=10, side="bottom")
        self._btn_modulos = None   # botÃ³n removido; referencia conservada para compatibilidad
        _side_action("â  Empresas", self._mostrar_lanzador)

        # ââ CONTENIDO DERECHO âââââââââââââââââââââââââââââââââââââââââ
        self._content_frame = tk.Frame(main, bg=COLOR_FONDO)
        self._content_frame.pack(side="right", fill="both", expand=True)

        # Mini-barra superior (tÃ­tulo)
        top_bar = tk.Frame(self._content_frame, bg=self._RIBBON_ACTIVE, height=40)
        top_bar.pack(fill="x")
        top_bar.pack_propagate(False)
        tk.Label(top_bar, text="AUXILIAR DE REGISTROS",
                 bg=self._RIBBON_ACTIVE, fg="#FFFFFF",
                 font=("Georgia", 11, "bold")).pack(side="left", padx=14)
        tk.Frame(self._content_frame, bg=COLOR_FUCSIA, height=1).pack(fill="x")

        self._panel_superior()

        # Log y barra de progreso se reservan al fondo ANTES del notebook
        # para que el notebook no consuma todo el espacio vertical
        self._panel_log()   # packs side="bottom" internamente

        _pb_global_frame = tk.Frame(self._content_frame, bg=COLOR_FONDO)
        _pb_global_frame.pack(side="bottom", fill="x", padx=8, pady=(2, 2))
        self._global_pb = FunkyProgressBar(_pb_global_frame, maximum=100, height=70)
        self._global_pb.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self._global_pb_lbl = tk.Label(_pb_global_frame, text="Listo",
                                        bg=COLOR_FONDO, fg=COLOR_FUCSIA_OSCURO,
                                        font=("Segoe UI", 8, "bold"), width=12, anchor="w")
        self._global_pb_lbl.pack(side="left")

        # Notebook en el medio â ahora expande solo el espacio que queda
        self.nb = ttk.Notebook(self._content_frame)
        self.nb.pack(fill="both", expand=True, padx=8, pady=(0, 2))
        self._habilitar_drag_tabs(self.nb)
        self.nb.bind("<ButtonRelease-1>", self._on_tab_close_click)

        self._rebuild_ribbon()
        if not self._modulos_activos:
            self._mostrar_hint_vacio()

    def _asegurar_x_tabs(self):
        """Agrega el sÃ­mbolo â a cualquier pestaÃ±a que no lo tenga todavÃ­a."""
        try:
            for tid in self.nb.tabs():
                txt = self.nb.tab(tid, "text")
                if "â" not in txt:          # â U+2715
                    self.nb.tab(tid, text=txt.rstrip() + "  â")
        except Exception:
            pass

    def _on_tab_close_click(self, event):
        """Cierra la pestaÃ±a si el clic cae sobre el â al final del label."""
        # Identificar quÃ© tab estÃ¡ bajo el cursor (sin filtrar por element,
        # porque nb.identify() devuelve valores distintos segÃºn el tema/SO)
        try:
            idx = self.nb.index("@%d,%d" % (event.x, event.y))
        except Exception:
            return
        try:
            tab_text = self.nb.tab(idx, "text")
            if "â" not in tab_text:
                return
            # Escanear a la derecha para hallar el borde derecho del tab
            scan_x = event.x + 1
            nb_w = self.nb.winfo_width()
            while scan_x < nb_w:
                try:
                    if self.nb.index("@%d,%d" % (scan_x, event.y)) != idx:
                        break
                except Exception:
                    break
                scan_x += 1
            # El â mÃ¡s padding del tema ocupa los Ãºltimos ~40 px del tab
            if scan_x - event.x > 40:
                return
            # Identificar el key y cerrar vÃ­a _toggle_modulo
            tab_id = self.nb.tabs()[idx]
            key = next((k for k, tid in self._tab_frames.items() if tid == tab_id), None)
            if key:
                self._toggle_modulo(key)
        except Exception:
            pass

    def _refrescar_cinta(self):
        """Actualiza colores y visibilidad de la cinta segÃºn mÃ³dulos disponibles/activos."""
        for key, b in self._btn_cinta.items():
            # Mostrar solo mÃ³dulos disponibles para esta empresa
            if key in self._modulos_disponibles:
                b.pack(side="left", padx=2, pady=4)
                if key in self._modulos_activos:
                    b.config(bg=COLOR_FUCSIA, fg=COLOR_BLANCO,
                             activebackground="#1558A0", activeforeground=COLOR_BLANCO)
                else:
                    b.config(bg=COLOR_FUCSIA_OSCURO, fg="#AED6F2",
                             activebackground="#0B2A3D", activeforeground=COLOR_BLANCO)
            else:
                b.pack_forget()

    def _toggle_modulo(self, key):
        """Activa o desactiva un mÃ³dulo desde la cinta."""
        if key not in self._modulos_disponibles:
            return  # no disponible para esta empresa
        if key in self._modulos_activos:
            # Quitar: usar referencia directa al widget
            tab_id = self._tab_frames.pop(key, None)
            if tab_id:
                try:
                    self.nb.forget(tab_id)
                except Exception:
                    pass
            self._modulos_activos.remove(key)
        else:
            # Quitar hint de inicio si existe
            try:
                for i in range(self.nb.index("end")):
                    if self.nb.tab(i, "text").strip() == "Inicio":
                        self.nb.forget(i)
                        break
            except Exception:
                pass
            self._modulos_activos.append(key)
            self._cargar_tab(key)
        if not self._modulos_activos:
            self._mostrar_hint_vacio()
        if self._on_modulos_change:
            try:
                self._on_modulos_change(self.nombre, self._modulos_activos)
            except Exception:
                pass
        self._refrescar_cinta()

    def _mostrar_hint_vacio(self):
        """Sin mÃ³dulos activos: vaciar el notebook completamente."""
        for tab_id in list(self.nb.tabs()):
            try:
                self.nb.forget(tab_id)
            except Exception:
                pass
        self._tab_frames.clear()

    def _cargar_tab(self, key):
        """Construye la pestaÃ±a de un mÃ³dulo por su clave (no agrega si ya existe)."""
        # Verificar referencia guardada
        if key in self._tab_frames:
            if self._tab_frames[key] in self.nb.tabs():
                return  # ya existe y estÃ¡ activa â no duplicar
            else:
                del self._tab_frames[key]  # referencia muerta, limpiar
        # Verificar que ningÃºn tab existente en el notebook es ya este mÃ³dulo
        # (por si _tab_frames se desincronizÃ³)
        label_esperado = next((lbl.strip() for k, lbl, _ in self.MODULOS if k == key), None)
        if label_esperado:
            for tid in self.nb.tabs():
                try:
                    if self.nb.tab(tid, "text").strip() == label_esperado:
                        self._tab_frames[key] = tid  # re-sincronizar
                        return
                except Exception:
                    pass
        for k, _, metodo in self.MODULOS:
            if k == key:
                n_antes = self.nb.index("end")
                getattr(self, metodo)(self.nb)
                n_despues = self.nb.index("end")
                if n_despues > n_antes:
                    tab_id = self.nb.tabs()[-1]
                    self._tab_frames[key] = tab_id
                    _color = self._MODULE_COLORS.get(key, "#1E6FBF")
                    _dot = self._make_tab_dot(_color)
                    self._tab_dot_imgs[key] = _dot
                    # AÃ±adir botÃ³n â al texto de la pestaÃ±a
                    try:
                        _txt = self.nb.tab(tab_id, "text").rstrip()
                        self.nb.tab(tab_id, text=_txt + "  â")
                    except Exception:
                        pass
                return

    def _agregar_modulo_dialog(self):
        """DiÃ¡logo para seleccionar mÃ³dulos a agregar."""
        disponibles = [(k, lbl) for k, lbl, _ in self.MODULOS
                       if k not in self._modulos_activos]
        if not disponibles:
            messagebox.showinfo("MÃ³dulos", "Ya tienes todos los mÃ³dulos disponibles en esta ventana.")
            return

        win = tk.Toplevel(self)
        win.title("Agregar mÃ³dulo")
        win.geometry("360x280")
        win.resizable(False, False)
        win.configure(bg=COLOR_FONDO)
        win.grab_set()

        ttk.Label(win, text="Selecciona los mÃ³dulos a agregar:",
                  style="Seccion.TLabel").pack(padx=20, pady=(18, 8), anchor="w")

        vars_check = {}
        for key, lbl in disponibles:
            v = tk.BooleanVar()
            vars_check[key] = v
            tk.Checkbutton(win, text=lbl.strip(), variable=v,
                           bg=COLOR_FONDO, fg=COLOR_TEXTO,
                           selectcolor=COLOR_FUCSIA_SUAVE,
                           font=("Segoe UI", 10),
                           activebackground=COLOR_FONDO).pack(anchor="w", padx=30, pady=4)

        def _confirmar():
            seleccionados = [k for k, v in vars_check.items() if v.get()]
            for key in seleccionados:
                # Quitar pestaÃ±a de inicio si es la primera vez
                if not self._modulos_activos:
                    try:
                        self.nb.forget(0)
                    except Exception:
                        pass
                self._cargar_tab(key)
                self._modulos_activos.append(key)
            if seleccionados and self._on_modulos_change:
                self._on_modulos_change(self.nombre, self._modulos_activos)
            win.destroy()

        ttk.Button(win, text="Agregar", style="Grande.TButton",
                   command=_confirmar).pack(pady=16)

    def _habilitar_drag_tabs(self, nb):
        """Permite reordenar pestaÃ±as del Notebook arrastrÃ¡ndolas."""
        _d = {}

        def _tab_en(x, y):
            try:
                return nb.index(f"@{x},{y}")
            except tk.TclError:
                return None

        def _press(event):
            idx = _tab_en(event.x, event.y)
            if idx is not None:
                _d['src']     = idx
                _d['x0']      = event.x
                _d['y0']      = event.y
                _d['activo']  = False   # solo True si hay movimiento real

        def _motion(event):
            if 'src' not in _d:
                return
            if abs(event.x - _d['x0']) > 8 or abs(event.y - _d['y0']) > 8:
                _d['activo'] = True
                nb.configure(cursor="fleur")

        def _release(event):
            nb.configure(cursor="")
            activo = _d.pop('activo', False)
            src    = _d.pop('src',    None)
            _d.clear()
            if not activo or src is None:
                return
            dst = _tab_en(event.x, event.y)
            if dst is None or dst == src:
                return
            # Guardar widget y opciones antes de modificar
            try:
                tabs   = nb.tabs()
                widget = nb.nametowidget(tabs[src])
                texto  = nb.tab(src, 'text')
            except Exception:
                return
            nb.forget(src)
            try:
                nb.insert(dst, widget, text=texto)
                nb.select(dst)
            except Exception:
                return
            # Sincronizar orden de mÃ³dulos
            nuevo_orden = []
            for i in range(nb.index("end")):
                lbl = nb.tab(i, "text").strip()
                for k, l, _ in self.MODULOS:
                    if l.strip() == lbl:
                        nuevo_orden.append(k); break
            self._modulos_activos = nuevo_orden
            if self._on_modulos_change:
                try:
                    self._on_modulos_change(self.nombre, self._modulos_activos)
                except Exception:
                    pass
            self._refrescar_cinta()

        nb.bind("<ButtonPress-1>",   _press,   add="+")
        nb.bind("<B1-Motion>",       _motion,  add="+")
        nb.bind("<ButtonRelease-1>", _release, add="+")

    def _on_cerrar_workspace(self):
        """Cierra esta Ã¡rea de trabajo y re-muestra el launcher si es la Ãºltima."""
        self.destroy()
        try:
            app = self.master
            abiertas = [w for w in app.workspaces.values() if w.winfo_exists()]
            if not abiertas:
                # No quedan Ã¡reas abiertas â mostrar el launcher
                app.deiconify()
                app.lift()
                app.focus_force()
        except Exception:
            pass

    def _configurar_modulos_dropdown(self):
        """Despliega un menÃº hacia abajo debajo del botÃ³n â MÃ³dulos."""
        btn = self._btn_modulos

        # Si ya hay un dropdown abierto, cerrarlo
        if getattr(self, "_modulos_popup", None):
            try:
                self._modulos_popup.destroy()
            except Exception:
                pass
            self._modulos_popup = None
            return

        # PosiciÃ³n: justo debajo del botÃ³n
        x = btn.winfo_rootx()
        y = btn.winfo_rooty() + btn.winfo_height()

        pop = tk.Toplevel(self)
        pop.overrideredirect(True)
        pop.configure(bg=COLOR_FONDO)
        pop.geometry(f"+{x}+{y}")
        self._modulos_popup = pop

        # Borde sutil
        frame = tk.Frame(pop, bg="#3E6E96", padx=1, pady=1)
        frame.pack(fill="both", expand=True)
        inner = tk.Frame(frame, bg=COLOR_FONDO)
        inner.pack(fill="both", expand=True)

        checks = {}
        for key, lbl, _ in self.MODULOS:
            # Checkbox refleja si la pestaÃ±a estÃ¡ activa actualmente
            var = tk.BooleanVar(value=key in self._modulos_activos)
            checks[key] = var
            cb = tk.Checkbutton(inner, text=lbl.strip(), variable=var,
                                bg=COLOR_FONDO, fg=COLOR_TEXTO,
                                selectcolor=COLOR_FUCSIA_SUAVE,
                                activebackground="#B7D9EF",
                                font=("Segoe UI", 10), anchor="w",
                                padx=14, pady=3, cursor="hand2")
            cb.pack(fill="x")

        sep = tk.Frame(inner, bg="#A8CDE8", height=1)
        sep.pack(fill="x", padx=8, pady=(4, 0))

        def aplicar():
            seleccionados = [k for k, var in checks.items() if var.get()]
            # Activar los nuevos (NO crear tabs aquÃ­ â solo al hacer clic en ribbon)
            for k in seleccionados:
                if k not in self._modulos_activos:
                    self._modulos_activos.append(k)
            # Desactivar los desmarcados
            for k in list(self._modulos_activos):
                if k not in seleccionados:
                    self._modulos_activos.remove(k)
                    # Quitar la pestaÃ±a usando _tab_frames (fiable)
                    tab_id = self._tab_frames.pop(k, None)
                    if tab_id:
                        try:
                            self.nb.forget(tab_id)
                        except Exception:
                            pass
            self._modulos_disponibles = list(dict.fromkeys(
                [k for k, _, _ in self.MODULOS]))
            self._refrescar_cinta()
            if not self._modulos_activos:
                self._mostrar_hint_vacio()
            if self._on_modulos_change:
                try:
                    self._on_modulos_change(self.nombre, self._modulos_activos)
                except Exception:
                    pass
            self._rebuild_ribbon()
            cerrar()

        def cerrar(event=None):
            try:
                pop.destroy()
            except Exception:
                pass
            self._modulos_popup = None

        tk.Button(inner, text="Aplicar", command=aplicar,
                  bg=COLOR_FUCSIA, fg=COLOR_BLANCO,
                  font=("Segoe UI", 9, "bold"), relief="flat",
                  padx=12, pady=5, cursor="hand2",
                  activebackground=COLOR_FUCSIA_OSCURO).pack(pady=8)

        # Cerrar al hacer clic fuera
        pop.bind("<FocusOut>", lambda e: self.after(100, lambda: cerrar() if self._modulos_popup else None))
        pop.focus_set()

    def _mostrar_lanzador(self):
        """Muestra el lanzador AUXILIAR DE REGISTROS."""
        try:
            launcher = self.master._launcher_top
            launcher.wm_attributes("-alpha", 1.0)
            launcher.lift()
            launcher.focus_force()
        except Exception:
            pass

    def _quitar_modulo_activo(self):
        """DiÃ¡logo para seleccionar quÃ© mÃ³dulos quitar."""
        if not self._modulos_activos:
            messagebox.showinfo("Sin mÃ³dulos", "No hay mÃ³dulos activos en esta ventana.")
            return

        win = tk.Toplevel(self)
        win.title("Quitar mÃ³dulo")
        win.geometry("360x260")
        win.resizable(False, False)
        win.configure(bg=COLOR_FONDO)
        win.grab_set()

        ttk.Label(win, text="Selecciona los mÃ³dulos a quitar:",
                  style="Seccion.TLabel").pack(padx=20, pady=(18, 8), anchor="w")

        vars_check = {}
        for key, lbl, _ in self.MODULOS:
            if key in self._modulos_activos:
                v = tk.BooleanVar()
                vars_check[key] = v
                tk.Checkbutton(win, text=lbl.strip(), variable=v,
                               bg=COLOR_FONDO, fg=COLOR_TEXTO,
                               selectcolor=COLOR_FUCSIA_SUAVE,
                               font=("Segoe UI", 10),
                               activebackground=COLOR_FONDO).pack(anchor="w", padx=30, pady=4)

        def _confirmar():
            a_quitar = [k for k, v in vars_check.items() if v.get()]
            if not a_quitar:
                win.destroy(); return
            for key in a_quitar:
                tab_id = self._tab_frames.pop(key, None)
                if tab_id:
                    try:
                        self.nb.forget(tab_id)
                    except Exception:
                        pass
                if key in self._modulos_activos:
                    self._modulos_activos.remove(key)
            if self._on_modulos_change:
                try:
                    self._on_modulos_change(self.nombre, self._modulos_activos)
                except Exception:
                    pass
            if not self.nb.tabs():
                self._mostrar_hint_vacio()
            win.destroy()

        ttk.Button(win, text="Quitar seleccionados", style="Grande.TButton",
                   command=_confirmar).pack(pady=16)

    # ---------------------------------------------------------------- #
    # PANEL SUPERIOR
    # ---------------------------------------------------------------- #
    def _panel_superior(self):
        # El panel de Carpeta/CatÃ¡logo/Escanear ahora vive dentro de cada mÃ³dulo.
        # AquÃ­ solo queda el botÃ³n global de Instrucciones.
        fila_botones = ttk.Frame(self._content_frame)
        fila_botones.pack(pady=(6, 4))
        ttk.Button(fila_botones, text="â  Instrucciones",
                   command=self._mostrar_instrucciones).pack(side="left")

    def _mostrar_instrucciones(self):
        """Ventana con instrucciones organizadas por mÃ³dulo."""
        win = tk.Toplevel(self)
        win.title("â Instrucciones de uso")
        win.geometry("780x640")
        win.resizable(True, True)
        win.configure(bg=COLOR_FONDO)
        win.grab_set()

        tk.Label(win,
            text=f"ð  {self.nombre} â GuÃ­a por mÃ³dulo",
            bg=COLOR_FONDO, fg=COLOR_FUCSIA,
            font=("Segoe UI", 13, "bold"),
        ).pack(pady=(16, 2))
        tk.Label(win,
            text="Selecciona un mÃ³dulo para ver sus instrucciones.",
            bg=COLOR_FONDO, fg=COLOR_TEXTO,
            font=("Segoe UI", 9),
        ).pack(pady=(0, 8))

        body = tk.Frame(win, bg=COLOR_FONDO)
        body.pack(fill="both", expand=True, padx=14, pady=(0, 4))
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        # ââ MÃ³dulos disponibles con sus instrucciones ââ
        INSTRUCCIONES = {
            "conciliacion": {
                "icono": "ð¼", "titulo": "Pagos Bancarios",
                "pasos": [
                    ("Escanea tu carpeta", "Haz clic en ð Escanear carpeta y selecciona la carpeta donde estÃ¡n los PDFs de dispersiÃ³n o comprobantes de pago de nÃ³mina."),
                    ("Selecciona archivos", "Se muestran 3 listas separadas:\nâ¸ NÃ³mina principal\nâ¸ Complementos\nâ¸ Vacaciones\nUsa 'Todos' o 'Ninguno' para seleccionar masivamente, o marca uno a uno."),
                    ("Procesa", "â¸ BotÃ³n individual (NÃ³mina / Complementos / Vacaciones): genera un Excel solo de esa categorÃ­a.\nâ¸ ð¦ Generar TODO: procesa las 3 listas y las combina en un solo archivo Excel consolidado."),
                    ("Descarga el resultado", "Al terminar aparecen los botones:\nâ¸ ð Abrir archivo â abre el Excel directamente.\nâ¸ ð¾ Guardar como â guarda una copia en la ubicaciÃ³n que elijas."),
                ],
            },
            "prestamos": {
                "icono": "ð³", "titulo": "PrÃ©stamos",
                "pasos": [
                    ("Escanea la carpeta", "Haz clic en ð Escanear carpeta para detectar automÃ¡ticamente los PDFs de prÃ©stamos de nÃ³mina."),
                    ("Selecciona empleados", "Marca los archivos que quieres procesar. Puedes usar 'Todos' o buscar por nombre."),
                    ("Exporta a Excel", "Haz clic en ð¤ Exportar prÃ©stamos a Excel para generar el reporte con montos, descuentos y saldos pendientes."),
                    ("Descarga", "Al finalizar, abre el archivo directamente con ð Abrir o guÃ¡rdalo con ð¾ Guardar como."),
                ],
            },
            "provision": {
                "icono": "ð", "titulo": "ProvisiÃ³n de NÃ³mina",
                "pasos": [
                    ("Selecciona la plantilla SINUBE", "Haz clic en ð Seleccionar plantilla y elige el archivo Excel base de SINUBE que contiene las columnas de dispersiÃ³n."),
                    ("Carga los XMLs del SAT", "Haz clic en â Agregar XMLs y selecciona los CFDIs de nÃ³mina emitidos por el SAT (pueden ser varios a la vez)."),
                    ("Procesa", "Haz clic en â¶ Procesar XMLs. La app lee cada CFDI y llena automÃ¡ticamente las columnas de la plantilla SINUBE con los datos de cada empleado."),
                    ("Descarga el resultado", "El archivo Excel generado aparece abajo. Usa ð Abrir para revisarlo o ð¾ Guardar como para guardarlo donde prefieras."),
                ],
            },
            "ventas": {
                "icono": "â½", "titulo": "Ventas del DÃ­a",
                "pasos": [
                    ("Carga el control de despachos", "Haz clic en ð Seleccionar Excel y elige el archivo de Control de Despachos del dÃ­a (generado por el sistema de la gasolinera)."),
                    ("Selecciona la plantilla de cuentas", "Elige la plantilla de cuentas contables que define las cuentas de mayor para la pÃ³liza."),
                    ("Genera la pÃ³liza", "Haz clic en â½ Generar PÃ³liza Ventas del DÃ­a. La app calcula totales por turno, producto y forma de pago y crea la pÃ³liza contable en Excel."),
                    ("Descarga", "El archivo Excel con la pÃ³liza aparece al terminar. Ãbrelo con ð Abrir Excel o guÃ¡rdalo con ð¾ Guardar como."),
                ],
            },
            "estado_cuenta": {
                "icono": "ð¦", "titulo": "Estado de Cuenta",
                "pasos": [
                    ("Selecciona el banco", "Haz clic en el banco correspondiente al estado de cuenta que vas a procesar (BBVA, Banorte, HSBC, Inbursa, etc.)."),
                    ("Carga el PDF", "Arrastra el PDF directamente a la zona azul, o haz clic en ella para buscar el archivo en tu equipo."),
                    ("Extrae las tablas", "Haz clic en â¬ Extraer tablas y generar Excel. La app lee el PDF y detecta automÃ¡ticamente los movimientos bancarios (depÃ³sitos, retiros, saldo)."),
                    ("Revisa y herramientas", "VerÃ¡s la tabla de movimientos. Puedes:\nâ¸ Copiar â copiar al portapapeles.\nâ¸ Î£ Autosuma â sumar columnas seleccionadas.\nâ¸ Quitar duplicados â eliminar filas repetidas.\nâ¸ Ordenar AâZ / ZâA â ordenar por fecha o monto.\nâ¸ ð Buscar â filtrar por descripciÃ³n o monto.\nâ¸ ð Abrir Excel â abrir el archivo generado."),
                ],
            },
            "visor": {
                "icono": "ð", "titulo": "Visor de Resultados",
                "pasos": [
                    ("Abre un archivo Excel", "Haz clic en ð Abrir archivo y selecciona cualquier Excel generado por la app (pÃ³lizas, conciliaciones, nÃ³mina, etc.)."),
                    ("Navega entre hojas", "Usa las pestaÃ±as en la parte inferior para cambiar de hoja dentro del archivo."),
                    ("Filtra y busca", "Usa el campo de bÃºsqueda para filtrar filas. Haz clic en los encabezados de columna para ordenar los datos."),
                    ("Exporta o imprime", "Desde el visor puedes copiar datos al portapapeles o imprimir directamente el reporte."),
                ],
            },
            "configuracion": {
                "icono": "â", "titulo": "ConfiguraciÃ³n",
                "pasos": [
                    ("Carpeta de trabajo", "Define la carpeta raÃ­z donde la app buscarÃ¡ y guardarÃ¡ los archivos. Al cambiarla, todas las rutas se actualizan automÃ¡ticamente."),
                    ("Carpeta de salida", "Subcarpeta donde se guardan los Excel generados. Por defecto es 'resultados' dentro de la carpeta de trabajo."),
                    ("Prefijo de archivos", "Texto que se agrega al inicio del nombre de cada archivo generado, p. ej. 'MAYO2026_'. Ãtil para organizar por periodo."),
                    ("Empresa activa", "Nombre de la empresa o razÃ³n social que aparece en los encabezados de los reportes generados."),
                ],
            },
            "concilia_sat": {
                "icono": "ð", "titulo": "ConciliaciÃ³n SAT",
                "pasos": [
                    ("Carga el Excel de Control de Despachos", "Arrastra el archivo Excel de Control de Despachos a la zona verde izquierda, o haz clic en ð Seleccionar Excel. Debe contener la columna FolioFiscal (UUID)."),
                    ("Carga los XMLs del SAT", "Arrastra uno o varios archivos XML (CFDIs del SAT) a la zona verde derecha, o haz clic en ð Seleccionar XMLs. Puedes cargar mÃºltiples XMLs a la vez."),
                    ("Procesa la conciliaciÃ³n", "Haz clic en ð Procesar ConciliaciÃ³n. La app compara el UUID de cada CFDI contra la columna FolioFiscal del Excel, detectando coincidencias y diferencias de importe."),
                    ("Revisa el reporte", "Se genera un Excel con 13 columnas:\nâ¸ Folio Fiscal XML y Folio Fiscal Excel\nâ¸ Â¿Coincide? (â SÃ / â NO)\nâ¸ Fecha, Receptor, Concepto\nâ¸ SubTotal, IVA y Total (XML)\nâ¸ Importe Excel y Diferencia\nâ¸ Estado (CONCILIADO / SIN MATCH)"),
                    ("Abre los resultados", "Al terminar aparecen dos botones:\nâ¸ ð Abrir ConciliaciÃ³n â abre el reporte completo.\nâ¸ ð Abrir Excel con UUID â abre el Control de Despachos original."),
                ],
            },
        }

        # ââ Panel izquierdo: lista de mÃ³dulos (con scrollbar) ââ
        left_outer = tk.Frame(body, bg=COLOR_FUCSIA_OSCURO, width=190)
        left_outer.pack_propagate(False)
        left_outer.grid(row=0, column=0, sticky="ns", padx=(0, 10))
        left_outer.rowconfigure(0, weight=1)
        left_outer.columnconfigure(0, weight=1)
        left_cv = tk.Canvas(left_outer, bg=COLOR_FUCSIA_OSCURO, highlightthickness=0, width=190)
        left_cv.grid(row=0, column=0, sticky="nsew")
        left_vsb = ttk.Scrollbar(left_outer, orient="vertical", command=left_cv.yview)
        left_vsb.grid(row=0, column=1, sticky="ns")
        left_cv.configure(yscrollcommand=left_vsb.set)
        left = tk.Frame(left_cv, bg=COLOR_FUCSIA_OSCURO)
        _left_win = left_cv.create_window((0, 0), window=left, anchor="nw")
        def _left_resize(e):
            left_cv.itemconfig(_left_win, width=e.width)
            left_cv.configure(scrollregion=left_cv.bbox("all"))
        left_cv.bind("<Configure>", _left_resize)
        left.bind("<Configure>", lambda e: left_cv.configure(scrollregion=left_cv.bbox("all")))

        # ââ Panel derecho: contenido ââ
        right = tk.Frame(body, bg=COLOR_BLANCO, highlightthickness=1,
                         highlightbackground=COLOR_FUCSIA_SUAVE)
        right.grid(row=0, column=1, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)

        content_title = tk.Label(right, text="", bg=COLOR_FUCSIA, fg=COLOR_BLANCO,
            font=("Segoe UI", 11, "bold"), anchor="w", padx=14, pady=10)
        content_title.grid(row=0, column=0, sticky="ew")

        scroll_r = ttk.Scrollbar(right, orient="vertical")
        scroll_r.grid(row=1, column=1, sticky="ns")
        txt = tk.Text(right, yscrollcommand=scroll_r.set, state="normal", wrap="word",
            bg=COLOR_BLANCO, fg=COLOR_TEXTO, relief="flat",
            font=("Segoe UI", 10), padx=16, pady=12, spacing3=6)
        scroll_r.config(command=txt.yview)
        txt.grid(row=1, column=0, sticky="nsew")
        txt.tag_config("step",  font=("Segoe UI", 10, "bold"), foreground=COLOR_FUCSIA_OSCURO)
        txt.tag_config("desc",  font=("Segoe UI", 10))
        txt.tag_config("sep",   font=("Segoe UI", 6))

        _btns_izq = {}

        def mostrar(key):
            for k, b in _btns_izq.items():
                b.config(bg=COLOR_FUCSIA_OSCURO, fg=COLOR_BLANCO)
            if key in _btns_izq:
                _btns_izq[key].config(bg=COLOR_BLANCO, fg=COLOR_FUCSIA_OSCURO)
            info = INSTRUCCIONES[key]
            content_title.config(text=f"  {info['icono']}  {info['titulo']}")
            txt.config(state="normal")
            txt.delete("1.0", "end")
            for i, (titulo_paso, desc) in enumerate(info["pasos"], 1):
                txt.insert("end", f"\n  {i}.  {titulo_paso}\n", "step")
                txt.insert("end", f"     {desc}\n\n", "desc")
            txt.config(state="disabled")
            txt.see("1.0")

        for key, info in INSTRUCCIONES.items():
            b = tk.Button(left,
                text=f"  {info['icono']}  {info['titulo']}",
                bg=COLOR_FUCSIA_OSCURO, fg=COLOR_BLANCO,
                font=("Segoe UI", 9), relief="flat", anchor="w",
                padx=10, pady=7, cursor="hand2",
                activebackground=COLOR_BLANCO, activeforeground=COLOR_FUCSIA_OSCURO,
                command=lambda k=key: mostrar(k))
            b.pack(fill="x", pady=1)
            _btns_izq[key] = b

        # Mostrar el primero por defecto
        first = next(iter(INSTRUCCIONES))
        mostrar(first)

        ttk.Button(win, text="â  Entendido", command=win.destroy).pack(pady=10)

    # ---------------------------------------------------------------- #
    # PESTAÃA 1: CONCILIACIÃN
    # ---------------------------------------------------------------- #
    def _tab_conciliacion(self, nb):
        tab = ttk.Frame(nb)
        nb.add(tab, text="  \U0001f4bc Pagos Bancarios  ")

        # âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
        # BARRA INFERIOR â resultado (pack side="bottom" PRIMERO para
        # garantizar visibilidad independiente del espacio disponible)
        # âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
        self._conc_resultado_path = None

        _bot = tk.Frame(tab, bg=COLOR_FUCSIA)
        _bot.pack(side="bottom", fill="x")

        # Fila de encabezado: nombre + botones
        _brow0 = tk.Frame(_bot, bg=COLOR_FUCSIA, height=40)
        _brow0.pack(side="top", fill="x")
        _brow0.pack_propagate(False)

        tk.Label(_brow0, text="\U0001f4c4  Reporte:",
                 bg=COLOR_FUCSIA, fg=COLOR_BLANCO,
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=(12, 4))
        self._conc_lbl_archivo = tk.Label(
            _brow0, text="\u2014  (genera el reporte primero)",
            bg=COLOR_FUCSIA, fg="#BFDBFE",
            font=("Segoe UI", 9))
        self._conc_lbl_archivo.pack(side="left", padx=4)
        ttk.Button(_brow0, text="\U0001f4be  Guardar como...",
                   command=self._conc_guardar_como).pack(
                       side="right", padx=6, pady=4)
        ttk.Button(_brow0, text="\U0001f4c2  Abrir en Excel",
                   command=self._conc_abrir).pack(
                       side="right", padx=(0, 4), pady=4)

        # Fila de vista previa: Treeview
        _brow1 = tk.Frame(_bot, bg=COLOR_FONDO)
        _brow1.pack(side="top", fill="x")

        self._conc_tv = ttk.Treeview(_brow1, show="headings",
                                      height=4, selectmode="browse")
        _tv_vsb = ttk.Scrollbar(_brow1, orient="vertical",
                                 command=self._conc_tv.yview)
        _tv_hsb = ttk.Scrollbar(_brow1, orient="horizontal",
                                 command=self._conc_tv.xview)
        self._conc_tv.configure(yscrollcommand=_tv_vsb.set,
                                 xscrollcommand=_tv_hsb.set)
        _tv_hsb.pack(side="bottom", fill="x", padx=4)
        _tv_vsb.pack(side="right", fill="y", pady=4)
        self._conc_tv.pack(side="left", fill="both", expand=True, padx=4, pady=4)

        # Placeholder inicial
        self._conc_tv["columns"] = ["msg"]
        self._conc_tv.heading("msg",
            text="\U0001f4ca  Genera el reporte para ver la vista previa aqu\xed")
        self._conc_tv.column("msg", width=500, anchor="center")
        self._conc_tv.insert("", "end", values=("\u2014",))

        # âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
        # ÃREA SUPERIOR â configuraciÃ³n + listas + botÃ³n generar
        # (pack side="top", expand=True llena el espacio restante)
        # âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
        top = ttk.Frame(tab)
        top.pack(side="top", fill="both", expand=True)
        top.columnconfigure(0, weight=1)
        top.columnconfigure(1, weight=1)
        top.columnconfigure(2, weight=1)
        top.rowconfigure(1, weight=1)

        # ââ Panel de configuraciÃ³n âââââââââââââââââââââââââââââââââ
        cfg_pagos = ttk.Frame(top, style="Tarjeta.TFrame")
        cfg_pagos.grid(row=0, column=0, columnspan=3,
                       sticky="ew", padx=6, pady=(6, 2))
        inner_p = ttk.Frame(cfg_pagos, style="Tarjeta.TFrame")
        inner_p.pack(fill="x", padx=10, pady=4)

        ttk.Label(inner_p, text="\U0001f4c1 Carpeta de trabajo:", width=22,
                  style="Tarjeta.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Entry(inner_p, textvariable=self.carpeta_pagos).grid(
            row=0, column=1, sticky="ew", padx=6, pady=2)
        ttk.Button(inner_p, text="ð PDFs...",
                   command=self._elegir_carpeta_pagos).grid(
                       row=0, column=2, padx=4)

        ttk.Label(inner_p, text="\U0001f4ca Cat\xe1logo de cuentas:", width=22,
                  style="Tarjeta.TLabel").grid(row=1, column=0, sticky="w")
        ttk.Entry(inner_p, textvariable=self.excel_path_pagos).grid(
            row=1, column=1, sticky="ew", padx=6, pady=2)
        ttk.Button(inner_p, text="Elegir...",
                   command=self._elegir_excel_pagos).grid(
                       row=1, column=2, padx=4)
        inner_p.columnconfigure(1, weight=1)

        ttk.Button(cfg_pagos, text="\U0001f50d  Escanear carpeta",
                   style="Grande.TButton",
                   command=self._escanear_pagos).pack(pady=(2, 6))

        # ââ Tres tarjetas con listas âââââââââââââââââââââââââââââââ
        self.listboxes = {}
        columnas = [
            ("nomina",       "\U0001f4bc N\xf3mina principal",
             "\u25b6  Conciliar seleccionados"),
            ("complementos", "\U0001f497 Complementos",
             "\u25b6  Procesar complementos"),
            ("vacaciones",   "\U0001f334 Vacaciones",
             "\u25b6  Procesar vacaciones"),
        ]
        for i, (key, titulo, btn_txt) in enumerate(columnas):
            card = ttk.Frame(top, style="Tarjeta.TFrame")
            card.grid(row=1, column=i, sticky="nsew", padx=6, pady=4)
            card.columnconfigure(0, weight=1)
            card.rowconfigure(1, weight=1)

            ttk.Label(card, text=titulo,
                      style="Encabezado.TLabel").grid(
                row=0, column=0, sticky="w", padx=10, pady=(8, 2))

            lbf, lb = _make_listbox(card, COLOR_TARJETA)
            lbf.grid(row=1, column=0, sticky="nsew", padx=10, pady=4)
            lb.config(height=6)
            self.listboxes[key] = lb

            sel_frame = ttk.Frame(card, style="Tarjeta.TFrame")
            sel_frame.grid(row=2, column=0, sticky="ew", padx=10)
            ttk.Button(sel_frame, text="\u2714 Todos",
                       command=lambda k=key: self._sel_todos(k)).pack(
                           side="left", padx=2, pady=3)
            ttk.Button(sel_frame, text="\u2718 Ninguno",
                       command=lambda k=key: self._sel_ninguno(k)).pack(
                           side="left", padx=2, pady=3)

            ttk.Button(card, text=btn_txt,
                       command=lambda k=key: self._ejecutar(k)).grid(
                row=3, column=0, sticky="ew", padx=10, pady=(2, 8))

        # ââ BotÃ³n "Generar TODO" âââââââââââââââââââââââââââââââââââ
        ttk.Button(top,
            text="\U0001f4e6  Generar TODO en un solo Excel"
                 "  (usa lo seleccionado en las 3 listas)",
            style="Grande.TButton",
            command=self._ejecutar_todo,
        ).grid(row=2, column=0, columnspan=3,
               sticky="ew", padx=6, pady=(0, 4))

        self.otros_label = ttk.Label(
            top, text="", foreground=COLOR_FUCSIA_OSCURO)
        self.otros_label.grid(
            row=3, column=0, columnspan=3, sticky="w", padx=10)

        # ââ Barra de progreso ââââââââââââââââââââââââââââââââââââââ
        _conc_pb_frame = ttk.Frame(top)
        _conc_pb_frame.grid(row=4, column=0, columnspan=3,
                            sticky="ew", padx=8, pady=(2, 2))
        _conc_pb_frame.grid_remove()
        self._conc_pb_frame = _conc_pb_frame
        self._conc_pb = FunkyProgressBar(
            _conc_pb_frame, maximum=100, height=60)
        self._conc_pb.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self._conc_pb_lbl = ttk.Label(
            _conc_pb_frame, text="0 %", width=8,
            foreground=COLOR_FUCSIA_OSCURO,
            font=("Segoe UI", 8, "bold"))
        self._conc_pb_lbl.pack(side="left")

    def _conc_set_resultado(self, path):
        """Registra el archivo generado por ConciliaciÃ³n y actualiza la barra."""
        self._conc_resultado_path = path
        nombre = os.path.basename(path) if path else "â"
        self._conc_lbl_archivo.config(text=nombre)

    def _conc_abrir(self):
        """Abre el archivo de resultado en el programa predeterminado."""
        if self._conc_resultado_path and os.path.exists(self._conc_resultado_path):
            os.startfile(self._conc_resultado_path)
        else:
            # Si no hay resultado aÃºn, permitir elegir un archivo manualmente
            self.lift()
            self.focus_force()
            self.update()
            ruta = filedialog.askopenfilename(
                parent=self,
                title="Abrir archivo de ConciliaciÃ³n",
                filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")],
            )
            if ruta:
                self._conc_set_resultado(ruta)
                os.startfile(ruta)

    def _conc_guardar_como(self):
        """Guarda una copia del resultado en la ubicaciÃ³n elegida."""
        if not self._conc_resultado_path or \
                not os.path.exists(self._conc_resultado_path):
            messagebox.showwarning(
                "Sin archivo", "Primero ejecuta una conciliaciÃ³n o abre un archivo.")
            return
        self.lift()
        self.focus_force()
        self.update()
        destino = filedialog.asksaveasfilename(
            parent=self,
            title="Guardar conciliaciÃ³n comoâ¦",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialdir=os.path.dirname(self._conc_resultado_path),
            initialfile=os.path.basename(self._conc_resultado_path),
        )
        if destino:
            import shutil
            shutil.copy2(self._conc_resultado_path, destino)
            messagebox.showinfo("Guardado", f"Archivo guardado en:\n{destino}")

    # ---------------------------------------------------------------- #
    # PESTAÃA 2: PRÃSTAMOS
    # ---------------------------------------------------------------- #
    def _tab_prestamos(self, nb):
        tab = ttk.Frame(nb)
        nb.add(tab, text="  ð³ PrÃ©stamos  ")
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)

        # ââ Panel de configuraciÃ³n propio de PrÃ©stamos ââââââââââââââââ
        cfg_prest = ttk.Frame(tab, style="Tarjeta.TFrame")
        cfg_prest.grid(row=0, column=0, sticky="ew", padx=6, pady=(8, 2))
        inner_pr = ttk.Frame(cfg_prest, style="Tarjeta.TFrame")
        inner_pr.pack(fill="x", padx=10, pady=6)

        ttk.Label(inner_pr, text="ð Carpeta de trabajo:", width=22,
                  style="Tarjeta.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Entry(inner_pr, textvariable=self.carpeta_prestamos).grid(
            row=0, column=1, sticky="ew", padx=6, pady=2)
        ttk.Button(inner_pr, text="Elegir...",
                   command=self._elegir_carpeta_prestamos).grid(row=0, column=2, padx=4)

        inner_pr.columnconfigure(1, weight=1)

        ttk.Button(cfg_prest, text="ð  Escanear carpeta",
                   style="Grande.TButton",
                   command=self._escanear_prestamos).pack(pady=(2, 8))
        # âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

        info = ttk.Frame(tab, style="Tarjeta.TFrame")
        info.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 4))
        ttk.Label(info,
            text="PDFs de prÃ©stamos detectados en la carpeta. "
                 "Usa el botÃ³n para exportar el listado a Excel.",
            style="Tarjeta.TLabel", wraplength=800,
        ).pack(padx=12, pady=8)

        card = ttk.Frame(tab, style="Tarjeta.TFrame")
        card.grid(row=2, column=0, sticky="nsew", padx=8, pady=(0, 8))
        card.columnconfigure(0, weight=1)
        card.rowconfigure(0, weight=1)

        lbf, self.lb_prestamos = _make_listbox(card, COLOR_TARJETA)
        lbf.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        btn_row = ttk.Frame(card, style="Tarjeta.TFrame")
        btn_row.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))
        ttk.Button(btn_row, text="â Todos",
                   command=lambda: self.lb_prestamos.select_set(0, "end")).pack(side="left", padx=2)
        ttk.Button(btn_row, text="â Ninguno",
                   command=lambda: self.lb_prestamos.select_clear(0, "end")).pack(side="left", padx=2)
        ttk.Button(btn_row, text="ð  Exportar listado a Excel",
                   command=self._exportar_prestamos).pack(side="right", padx=2)

    # ---------------------------------------------------------------- #
    # PESTAÃA 3: PROVISIÃN DE NÃMINA (XML â SINUBE)
    # ---------------------------------------------------------------- #
    def _tab_provision_nomina(self, nb):
        tab = ttk.Frame(nb)
        nb.add(tab, text="  ð ProvisÃ³n de NÃ³mina  ")
        self._prov_resultado_path = None
        self.prov_xmls = getattr(self, "prov_xmls", [])
        tab.columnconfigure(0, weight=1)

        # ââ Panel de configuraciÃ³n: plantilla Excel âââââââââââââââââââââââââââ
        cfg = ttk.Frame(tab, style="Tarjeta.TFrame")
        cfg.grid(row=0, column=0, sticky="ew", padx=6, pady=(8, 2))
        inner = ttk.Frame(cfg, style="Tarjeta.TFrame")
        inner.pack(fill="x", padx=10, pady=8)
        ttk.Label(inner, text="ð  Plantilla Excel:", width=20,
                  style="Tarjeta.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Entry(inner, textvariable=self.prov_plantilla).grid(
            row=0, column=1, sticky="ew", padx=6, pady=2)
        ttk.Button(inner, text="Elegir...",
                   command=self._prov_elegir_plantilla).grid(row=0, column=2, padx=4)
        inner.columnconfigure(1, weight=1)

        # ââ Panel de XMLs âââââââââââââââââââââââââââââââââââââââââââââââââââââ
        xml_frame = ttk.Frame(tab, style="Tarjeta.TFrame")
        xml_frame.grid(row=1, column=0, sticky="nsew", padx=6, pady=(4, 2))
        tab.rowconfigure(1, weight=1)
        xml_frame.columnconfigure(0, weight=1)
        xml_frame.rowconfigure(1, weight=1)

        ttk.Label(xml_frame, text="ð  Archivos XML (CFDI de NÃ³mina):",
                  style="Tarjeta.TLabel").grid(
            row=0, column=0, sticky="w", padx=10, pady=(8, 2))

        lb_frame = ttk.Frame(xml_frame)
        lb_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 4))
        lb_frame.columnconfigure(0, weight=1)
        lb_frame.rowconfigure(0, weight=1)

        self.lb_prov_xml = tk.Listbox(
            lb_frame, height=10, selectmode="extended",
            font=("Segoe UI", 9), bg="white", fg=COLOR_FUCSIA_OSCURO,
            selectbackground=COLOR_FUCSIA, selectforeground="white",
            relief="flat", highlightthickness=1,
            highlightbackground=COLOR_FUCSIA_SUAVE)
        self.lb_prov_xml.grid(row=0, column=0, sticky="nsew")
        _sb_prov = ttk.Scrollbar(lb_frame, orient="vertical",
                                 command=self.lb_prov_xml.yview)
        _sb_prov.grid(row=0, column=1, sticky="ns")
        self.lb_prov_xml.configure(yscrollcommand=_sb_prov.set)

        btn_bar = ttk.Frame(xml_frame)
        btn_bar.grid(row=2, column=0, sticky="e", padx=10, pady=(0, 8))
        ttk.Button(btn_bar, text="â  Agregar XMLs",
                   command=self._prov_agregar_xmls).pack(side="left", padx=3)
        ttk.Button(btn_bar, text="â  Quitar",
                   command=self._prov_quitar_xml).pack(side="left", padx=3)
        ttk.Button(btn_bar, text="ð  Limpiar todo",
                   command=self._prov_limpiar_xmls).pack(side="left", padx=3)

        # ââ BotÃ³n Procesar ââââââââââââââââââââââââââââââââââââââââââââââââââââ
        ttk.Button(tab, text="âï¸  Procesar ProvisiÃ³n de NÃ³mina",
                   style="Grande.TButton",
                   command=self._prov_procesar).grid(
            row=2, column=0, pady=(4, 6))

        # ââ Barra de progreso âââââââââââââââââââââââââââââââââââââââââââââââââ
        _prov_pb_frame = tk.Frame(tab, bg=COLOR_FONDO)
        _prov_pb_frame.grid(row=3, column=0, sticky="ew", padx=8, pady=(2, 2))
        _prov_pb_frame.grid_remove()
        self._prov_pb_frame = _prov_pb_frame
        self._prov_pb = FunkyProgressBar(_prov_pb_frame, maximum=100, height=70)
        self._prov_pb.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self._prov_pb_lbl = ttk.Label(
            _prov_pb_frame, text="0 %", width=8,
            foreground=COLOR_FUCSIA_OSCURO, font=("Segoe UI", 8, "bold"))
        self._prov_pb_lbl.pack(side="left")

        # ââ Barra Abrir / Guardar como ââââââââââââââââââââââââââââââââââââââââ
        self._prov_barra_descarga = ttk.Frame(tab, style="Tarjeta.TFrame")
        self._prov_barra_descarga.grid(
            row=4, column=0, sticky="ew", padx=6, pady=(2, 8))
        self._prov_barra_descarga.grid_remove()
        ttk.Label(self._prov_barra_descarga,
                  text="ð  Resultado:", style="Tarjeta.TLabel").pack(
            side="left", padx=(10, 4), pady=8)
        self._prov_lbl_archivo = ttk.Label(
            self._prov_barra_descarga, text="â",
            foreground=COLOR_FUCSIA_OSCURO,
            font=("Segoe UI", 9, "bold"),
            background=COLOR_TARJETA)
        self._prov_lbl_archivo.pack(side="left", padx=4, pady=8)
        ttk.Button(self._prov_barra_descarga,
                   text="ð  Abrir archivo",
                   command=self._prov_abrir_resultado).pack(
            side="right", padx=4, pady=6)
        ttk.Button(self._prov_barra_descarga,
                   text="ð¾  Guardar como...",
                   command=self._prov_guardar_como).pack(
            side="right", padx=4, pady=6)

    def _prov_elegir_plantilla(self):
        self.lift()
        self.focus_force()
        self.update()
        _cur = self.prov_plantilla.get()
        _init = os.path.dirname(_cur) if _cur and os.path.exists(_cur) else os.getcwd()
        r = filedialog.askopenfilename(
            parent=self,
            initialdir=_init,
            title="Selecciona la plantilla Excel SINUBE",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos los archivos", "*.*")],
        )
        if r:
            self.prov_plantilla.set(r)

    def _prov_agregar_xmls(self):
        self.lift()
        self.focus_force()
        self.update()
        _cur = self.prov_plantilla.get()
        _init = os.path.dirname(_cur) if _cur and os.path.exists(_cur) else os.getcwd()
        rutas = filedialog.askopenfilenames(
            parent=self,
            initialdir=_init,
            title="Selecciona archivos XML (CFDI)",
            filetypes=[("XML", "*.xml"), ("Todos los archivos", "*.*")],
        )
        for r in rutas:
            if r not in self.prov_xmls:
                self.prov_xmls.append(r)
                self.lb_prov_xml.insert("end", os.path.basename(r))

    def _prov_quitar_xml(self):
        sel = list(self.lb_prov_xml.curselection())
        for i in reversed(sel):
            self.lb_prov_xml.delete(i)
            self.prov_xmls.pop(i)

    def _prov_limpiar_xmls(self):
        self.lb_prov_xml.delete(0, "end")
        self.prov_xmls.clear()

    def _prov_abrir_resultado(self):
        if self._prov_resultado_path and os.path.exists(self._prov_resultado_path):
            try:
                os.startfile(self._prov_resultado_path)
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{e}")
        else:
            messagebox.showwarning("Sin archivo", "No hay archivo generado todavÃ­a.")

    def _prov_guardar_como(self):
        if not self._prov_resultado_path or not os.path.exists(self._prov_resultado_path):
            messagebox.showwarning("Sin archivo", "No hay archivo generado todavÃ­a.")
            return
        self.lift()
        self.focus_force()
        self.update()
        destino = filedialog.asksaveasfilename(
            parent=self,
            initialdir=os.path.dirname(self._prov_resultado_path),
            initialfile=os.path.basename(self._prov_resultado_path),
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
            title="Guardar ProvisiÃ³n de NÃ³mina como...",
        )
        if destino:
            import shutil
            shutil.copy2(self._prov_resultado_path, destino)
            self._log(f"â Guardado en: {destino}", ok=True)
            messagebox.showinfo("Guardado", f"Archivo guardado en:\n{destino}")

    def _prov_procesar(self):
        if not self.prov_xmls:
            messagebox.showwarning("Sin XMLs", "Agrega al menos un archivo XML.")
            return
        if not self.prov_plantilla.get():
            messagebox.showwarning("Sin plantilla", "Elige primero la plantilla Excel SINUBE.")
            return
        # Ocultar barra de descarga anterior
        self._prov_barra_descarga.grid_remove()
        self._prov_resultado_path = None
        self._prov_pb_frame.grid()
        self._pb_iniciar(self._prov_pb, self._prov_pb_lbl)
        threading.Thread(target=self._prov_procesar_hilo, daemon=True).start()

    def _prov_procesar_hilo(self):
        try:
            import xml.etree.ElementTree as ET
            import openpyxl, re as _re, shutil as _shutil
            from openpyxl.utils import get_column_letter
            from datetime import datetime

            def norm(s):
                return _re.sub(r'\s+', ' ', str(s or '').strip().upper())

            plantilla = self.prov_plantilla.get()
            base_plantilla = os.path.dirname(os.path.abspath(plantilla))
            out_dir = self._resultados_dir(base_hint=base_plantilla)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            prefijo = self.prefijo_salida.get().strip()
            nombre_out = f"{prefijo}Provision_Nomina_{ts}.xlsx"
            out_path = os.path.join(out_dir, nombre_out)

            _shutil.copy2(plantilla, out_path)
            try:
                os.chmod(out_path, 0o644)
            except Exception:
                pass
            wb = openpyxl.load_workbook(out_path)

            # ââ Auto-detectar hoja POLIZA âââââââââââââââââââââââââââââââââââââ
            ws = None
            for sn in wb.sheetnames:
                if sn.strip().upper() in ('POLIZA', 'PÃLIZA', 'POLIZA IA'):
                    ws = wb[sn]
                    break
            if ws is None:
                ws = wb.active

            # Limpiar filas de datos (mantener encabezados filas 1-3)
            for r in range(ws.max_row, 3, -1):
                ws.delete_rows(r)

            # ââ Leer hoja CUENTAS âââââââââââââââââââââââââââââââââââââââââââââ
            # A/B  â Percepciones (SUELDO, AYUDA DESPENSA, â¦)
            # G/H  â Empleados regulares  (210-01-xxxx)
            # J/K  â Asimilados a salarios (210-05-xxxx)
            # M/N  â Deducciones (ISR, IMSS, â¦)
            ws_c = None
            for sn in wb.sheetnames:
                if sn.strip().upper() == 'CUENTAS':
                    ws_c = wb[sn]
                    break

            perc_cat    = []   # [(cuenta, nombre_upper)]
            ded_cat     = []   # [(cuenta, nombre_upper)]
            emp_reg     = {}   # nombre_upper -> cuenta 210-01-xxxx
            emp_asim    = {}   # nombre_upper -> cuenta 210-05-xxxx
            emp_prestamo = {}  # nombre_upper -> cuenta 107-01-xxxx  (cols D/E)

            if ws_c is not None:
                for _row in ws_c.iter_rows(min_row=2, values_only=True):
                    _a = _row[0]  if len(_row) >  0 else None
                    _b = _row[1]  if len(_row) >  1 else None
                    _d = _row[3]  if len(_row) >  3 else None
                    _e = _row[4]  if len(_row) >  4 else None
                    _g = _row[6]  if len(_row) >  6 else None
                    _h = _row[7]  if len(_row) >  7 else None
                    _j = _row[9]  if len(_row) >  9 else None
                    _k = _row[10] if len(_row) > 10 else None
                    _m = _row[12] if len(_row) > 12 else None
                    _n = _row[13] if len(_row) > 13 else None
                    if _a and _b: perc_cat.append((str(_a).strip(), norm(_b)))
                    if _d and _e: emp_prestamo[norm(_e)] = str(_d).strip()
                    if _g and _h: emp_reg[norm(_h)]  = str(_g).strip()
                    if _j and _k: emp_asim[norm(_k)] = str(_j).strip()
                    if _m and _n: ded_cat.append((str(_m).strip(), norm(_n)))

            self.after(0, self._log,
                f"  CUENTAS: {len(perc_cat)} percepciones, "
                f"{len(emp_reg)+len(emp_asim)} empleados, "
                f"{len(emp_prestamo)} prÃ©stamos, "
                f"{len(ded_cat)} deducciones.")

            # ââ Localizar TOTAL 1 y TOTAL 2 en fila 3 ââââââââââââââââââââââââ
            col_tot1 = col_tot2 = None
            for c in range(1, ws.max_column + 1):
                v = norm(ws.cell(3, c).value or '')
                if 'TOTAL 1' in v:   col_tot1 = c
                elif 'TOTAL 2' in v: col_tot2 = c

            # ââ Insertar percepciones ANTES de TOTAL 1 âââââââââââââââââââââââ
            col_perc_start = col_tot1   # posiciÃ³n inicial; TOTAL 1 = SUM(col_perc_start:col_tot1-1)
            perc_cols = {}   # nombre_upper -> columna
            for i, (cta, nombre) in enumerate(perc_cat):
                col_ins = col_tot1 + i
                ws.insert_cols(col_ins)
                ws.cell(2, col_ins, value=cta)
                ws.cell(3, col_ins, value=nombre)
                perc_cols[nombre] = col_ins
            col_tot1 += len(perc_cat)
            col_tot2 += len(perc_cat)

            # ââ Insertar deducciones DESPUÃS de TOTAL 1 ââââââââââââââââââââââ
            ded_cols = {}   # nombre_upper -> columna
            for i, (cta, nombre) in enumerate(ded_cat):
                col_ins = col_tot1 + 1 + i
                ws.insert_cols(col_ins)
                ws.cell(2, col_ins, value=cta)
                ws.cell(3, col_ins, value=nombre)
                ded_cols[nombre] = col_ins
            col_tot2 += len(ded_cat)

            self.after(0, self._log,
                f"  Estructura: TOTAL 1=col{col_tot1}  |  TOTAL 2=col{col_tot2}")

            # ââ Parsear XMLs ââââââââââââââââââââââââââââââââââââââââââââââââââ
            self.after(0, self._log,
                f"Leyendo {len(self.prov_xmls)} archivo(s) XML...")
            filas = []
            for xml_path in self.prov_xmls:
                try:
                    root = ET.parse(xml_path).getroot()
                    receptor = nomina = None
                    uuid = ''
                    for el in root.iter():
                        local = el.tag.split('}')[-1] if '}' in el.tag else el.tag
                        if local == 'Receptor' and receptor is None:
                            receptor = el
                        elif local == 'Nomina' and 'FechaPago' in el.attrib and nomina is None:
                            nomina = el
                        elif local == 'TimbreFiscalDigital' and not uuid:
                            uuid = el.get('UUID', '')

                    if nomina is None:
                        self.after(0, self._log,
                            f"  â  {os.path.basename(xml_path)} â sin nodo Nomina, se omite.", True)
                        continue

                    # CRÃTICO: usar 'is not None' â bool(Element) es False si no tiene hijos
                    nombre = (_re.sub(r'\s+', ' ', receptor.get('Nombre', '').strip())
                              if receptor is not None else '')
                    rfc    = (receptor.get('Rfc', '').strip()
                              if receptor is not None else '')
                    total  = float(root.get('Total', 0))
                    fecha  = datetime.strptime(nomina.get('FechaPago'), '%Y-%m-%d')

                    perc = {}
                    for el in root.iter():
                        if el.tag.split('}')[-1] == 'Percepcion':
                            c = norm(el.get('Concepto', ''))
                            v = (float(el.get('ImporteGravado', 0)) +
                                 float(el.get('ImporteExento',  0)))
                            if c and v:
                                perc[c] = perc.get(c, 0) + v
                    ded = {}
                    for el in root.iter():
                        if el.tag.split('}')[-1] == 'Deduccion':
                            c = norm(el.get('Concepto', ''))
                            v = float(el.get('Importe', 0))
                            if c and v:
                                ded[c] = ded.get(c, 0) + v

                    # OtroPago con PRESTAMO â percepciÃ³n especial (107-01-xxxx, lado percepciones)
                    otro_prest = 0.0
                    for el in root.iter():
                        if el.tag.split('}')[-1] == 'OtroPago':
                            c = norm(el.get('Concepto', ''))
                            v = float(el.get('Importe', 0))
                            if 'PRESTAMO' in c and v:
                                otro_prest += v

                    # Detectar si es asimilado: tiene percepciÃ³n ASIMILABLES A SALARIOS
                    es_asimilado = 'ASIMILABLES A SALARIOS' in perc

                    filas.append({
                        'fecha':       fecha,
                        'rfc':         rfc,
                        'nombre':      nombre,
                        'nombre_up':   norm(nombre),
                        'uuid':        uuid,
                        'total':       total,
                        'perc':        perc,
                        'ded':         ded,
                        'otro_prest':  otro_prest,
                        'asimilado':   es_asimilado,
                        'ref':         f"NOMINA DEL {fecha.strftime('%d/%m/%Y')}",
                    })
                    self.after(0, self._log,
                        f"  â {os.path.basename(xml_path)}  |  {nombre}  |  ${total:,.2f}",
                        False, True)

                except Exception as e:
                    self.after(0, self._log,
                        f"  â {os.path.basename(xml_path)}: {e}", True)

            if not filas:
                self.after(0, self._log, "No se procesÃ³ ningÃºn XML.", True)
                self.after(0, self._pb_error, self._prov_pb, self._prov_pb_lbl)
                self.after(0, self._prov_pb_frame.grid_remove)
                return

            filas.sort(key=lambda f: f['fecha'])

            # ââ Insertar columnas OtroPago-PRESTAMO en PERCEPCIONES (antes TOTAL 1) ââ
            # Un empleado con OtroPago PRESTAMO recibe un prÃ©stamo â va en percepciones
            # con cuenta 107-01-xxxx (CUENTAS D/E)
            otro_prest_cols = {}   # nombre_upper -> columna
            vistos_otro_prest = []
            for f in filas:
                if f['otro_prest'] and f['nombre_up'] not in vistos_otro_prest:
                    vistos_otro_prest.append(f['nombre_up'])

            for nu in vistos_otro_prest:
                cta = emp_prestamo.get(nu, '')
                col_ins = col_tot1   # justo antes de TOTAL 1
                ws.insert_cols(col_ins)
                ws.cell(2, col_ins, value=cta)
                ws.cell(3, col_ins, value=f"PREST.OTROP. {nu}")
                otro_prest_cols[nu] = col_ins
                col_tot1 += 1
                col_tot2 += 1
                if cta:
                    self.after(0, self._log,
                        f"  + OtroPago PrÃ©stamo {nu}  â  {cta}  (col {col_ins})")
                else:
                    self.after(0, self._log,
                        f"  â  Sin cuenta prÃ©stamo en CUENTAS D/E: '{nu}'", True)

            # CRÃTICO: insertar OtroPago antes de TOTAL 1 desplaza ded_cols â corregir Ã­ndices
            if otro_prest_cols:
                n_shift = len(otro_prest_cols)
                ded_cols = {k: v + n_shift for k, v in ded_cols.items()}

            # ââ Insertar columnas de PRÃSTAMO por empleado (entre ded y neto) ââ
            # Se crean sÃ³lo para empleados que tienen alguna deducciÃ³n con "PRESTAMO"
            prest_cols = {}   # nombre_upper -> columna  (107-01-xxxx)
            vistos_prest = []
            for f in filas:
                for c_name in f['ded']:
                    if 'PRESTAMO' in c_name and f['nombre_up'] not in vistos_prest:
                        vistos_prest.append(f['nombre_up'])
                        break

            for nu in vistos_prest:
                cta = emp_prestamo.get(nu, '')
                col_ins = col_tot2
                ws.insert_cols(col_ins)
                ws.cell(2, col_ins, value=cta)
                ws.cell(3, col_ins, value=f"PREST. {nu}")
                prest_cols[nu] = col_ins
                col_tot2 += 1
                if cta:
                    self.after(0, self._log,
                        f"  + PrÃ©stamo {nu}  â  {cta}  (col {col_ins})")
                else:
                    self.after(0, self._log,
                        f"  â  Sin cuenta prÃ©stamo en CUENTAS D/E: '{nu}'", True)

            # ââ Insertar columnas de empleados ANTES de TOTAL 2 ââââââââââââââ
            vistos = []
            for f in filas:
                if f['nombre_up'] and f['nombre_up'] not in vistos:
                    vistos.append(f['nombre_up'])

            emp_cols = {}   # nombre_upper -> columna
            for nu in vistos:
                cta = emp_reg.get(nu) or emp_asim.get(nu) or ''
                col_ins = col_tot2
                ws.insert_cols(col_ins)
                ws.cell(2, col_ins, value=cta)
                ws.cell(3, col_ins, value=nu)
                emp_cols[nu] = col_ins
                col_tot2 += 1
                if cta:
                    self.after(0, self._log,
                        f"  + {nu}  â  {cta}  (col {col_ins})")
                else:
                    self.after(0, self._log,
                        f"  â  Sin cuenta en CUENTAS: '{nu}'", True)

            # ââ Columna CONCILIACION al final (TOTAL 1 â TOTAL 2) ââââââââââââ
            col_conc = col_tot2 + 1
            ws.cell(3, col_conc, value='CONCILIACION')

            # ââ Fila 1: numeraciÃ³n 0-based + estilo profesional âââââââââââââââ
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            _fill_num  = PatternFill("solid", fgColor="1F3864")   # azul marino
            _font_num  = Font(name="Segoe UI", bold=True, color="FFFFFF", size=8)
            _align_ctr = Alignment(horizontal="center", vertical="center")
            _side_thin = Side(style="thin", color="3A6BAF")
            _border_num = Border(
                left=_side_thin, right=_side_thin,
                top=_side_thin,  bottom=_side_thin,
            )
            ws.row_dimensions[1].height = 16
            for _c in range(1, col_conc + 1):
                _cell = ws.cell(1, _c, value=_c - 1)
                _cell.fill      = _fill_num
                _cell.font      = _font_num
                _cell.alignment = _align_ctr
                _cell.border    = _border_num

            # ââ Reconstruir mapa de encabezados (fila 3) âââââââââââââââââââââ
            col_meta = {}
            for c in range(1, ws.max_column + 1):
                v = norm(ws.cell(3, c).value or '')
                if v:
                    col_meta[v] = c

            # ââ Rastrear conceptos sin columna (para diagnÃ³stico) âââââââââââââ
            perc_no_mapeadas = set()
            ded_no_mapeadas  = set()

            # ââ Escribir filas de datos âââââââââââââââââââââââââââââââââââââââ
            self.after(0, self._log,
                f"Escribiendo {len(filas)} fila(s)...")
            nf = 4
            for fila in filas:
                # Meta
                if 'TIPO DE POLIZA' in col_meta:
                    ws.cell(nf, col_meta['TIPO DE POLIZA'], value='D')
                if 'FECHA' in col_meta:
                    ws.cell(nf, col_meta['FECHA'],
                            value=fila['fecha']).number_format = 'DD/MM/YYYY'
                if 'REFERENCIA' in col_meta:
                    ws.cell(nf, col_meta['REFERENCIA'], value=fila['ref'])
                if 'CONCEPTO' in col_meta:
                    ws.cell(nf, col_meta['CONCEPTO'], value=fila['ref'])
                if 'UIDD' in col_meta:
                    ws.cell(nf, col_meta['UIDD'],
                            value=f"{fila['uuid']};{fila['rfc']};{fila['total']}")

                # Percepciones
                for c_name, c_val in fila['perc'].items():
                    _c = perc_cols.get(c_name)
                    if _c and c_val:
                        ws.cell(nf, _c, value=c_val).number_format = '#,##0.00'
                    elif c_val and not _c:
                        perc_no_mapeadas.add(c_name)

                # Deducciones
                # Reglas especiales:
                #   PRESTAMO* â columna individual del empleado (107-01-xxxx, prest_cols)
                #   ISR + asimilado â ISR ASIMILADOS (216-02)
                _prest_acum = 0.0   # acumula todos los prÃ©stamos del empleado en esta fila
                for c_name, c_val in fila['ded'].items():
                    if 'PRESTAMO' in c_name:
                        _prest_acum += c_val
                        continue   # se escribe abajo, acumulado
                    col_key = c_name
                    if fila.get('asimilado') and c_name == 'ISR':
                        col_key = 'ISR ASIMILADOS'
                    _c = ded_cols.get(col_key)
                    if _c and c_val:
                        ws.cell(nf, _c, value=c_val).number_format = '#,##0.00'
                    elif c_val and not _c:
                        ded_no_mapeadas.add(col_key)
                # Escribir total de prÃ©stamos en columna del empleado (107-01-xxxx)
                if _prest_acum:
                    _cp = prest_cols.get(fila['nombre_up'])
                    if _cp:
                        ws.cell(nf, _cp,
                                value=_prest_acum).number_format = '#,##0.00'
                    else:
                        ded_no_mapeadas.add('PRESTAMO (sin col empleado)')

                # OtroPago PRESTAMO â columna percepciÃ³n del empleado (107-01-xxxx)
                if fila['otro_prest']:
                    _cp_otro = otro_prest_cols.get(fila['nombre_up'])
                    if _cp_otro:
                        ws.cell(nf, _cp_otro,
                                value=fila['otro_prest']).number_format = '#,##0.00'
                    else:
                        perc_no_mapeadas.add('PRESTAMO OtroPago (sin col empleado)')

                # Neto del empleado (Total del CFDI)
                _ce = emp_cols.get(fila['nombre_up'])
                if _ce:
                    ws.cell(nf, _ce, value=fila['total']).number_format = '#,##0.00'

                # TOTAL 1 = SUM de TODA la secciÃ³n de percepciones (col_perc_start â col_tot1-1)
                # Incluye percepciones fijas + OtroPago PRESTAMO por empleado
                p_s = get_column_letter(col_perc_start)
                p_e = get_column_letter(col_tot1 - 1)
                ws.cell(nf, col_tot1,
                    value=f"=SUM({p_s}{nf}:{p_e}{nf})").number_format = '#,##0.00'

                # TOTAL 2 = SUM de deducciones + columnas de empleados
                d_s = get_column_letter(col_tot1 + 1)
                d_e = get_column_letter(col_tot2 - 1)
                ws.cell(nf, col_tot2,
                    value=f"=SUM({d_s}{nf}:{d_e}{nf})").number_format = '#,##0.00'

                # CONCILIACION = TOTAL 1 â TOTAL 2  (debe ser 0)
                t1l = get_column_letter(col_tot1)
                t2l = get_column_letter(col_tot2)
                ws.cell(nf, col_conc,
                    value=f"={t1l}{nf}-{t2l}{nf}").number_format = '#,##0.00'

                nf += 1

            # ââ Estilos profesionales âââââââââââââââââââââââââââââââââââââââââ
            # Paleta corporativa azul-Ã¡mbar-verde
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

            def _fill(hex6):
                return PatternFill("solid", fgColor=hex6)
            def _font(bold=False, color="000000", size=9, italic=False):
                return Font(name="Segoe UI", bold=bold, italic=italic,
                            color=color, size=size)
            def _border(color="C9D9EE"):
                s = Side(style="thin", color=color)
                return Border(left=s, right=s, top=s, bottom=s)

            C_NAVY   = "1F3864"   # azul marino oscuro  (fila 1 â ya hecho)
            C_MID    = "2F5496"   # azul medio           (fila 2 â cuentas)
            C_BLUE   = "4472C4"   # azul header          (fila 3)
            C_AMBER  = "FFC000"   # Ã¡mbar TOTAL          (encabezado totales)
            C_AMBER2 = "FFF2CC"   # Ã¡mbar claro          (datos totales)
            C_GREEN  = "375623"   # verde oscuro         (encabezado conciliacion)
            C_GREEN2 = "E2EFDA"   # verde claro          (datos conciliacion)
            C_ROW1   = "FFFFFF"   # blanco filas impares
            C_ROW2   = "D6E4F7"   # azul muy claro filas pares
            C_WHITE  = "FFFFFF"
            FMT_NUM  = '#,##0.00'

            brd = _border()

            # ââ Fila 2 (cuentas contables) ââââââââââââââââââââââââââââââââââââ
            ws.row_dimensions[2].height = 14
            for c in range(1, col_conc + 1):
                cell = ws.cell(2, c)
                if c in (col_tot1, col_tot2):
                    cell.fill = _fill(C_AMBER)
                    cell.font = _font(italic=True, color=C_NAVY, size=8)
                elif c == col_conc:
                    cell.fill = _fill(C_GREEN)
                    cell.font = _font(italic=True, color=C_WHITE, size=8)
                else:
                    cell.fill = _fill(C_MID)
                    cell.font = _font(color=C_WHITE, size=8, italic=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = brd

            # ââ Fila 3 (encabezados de columnas) âââââââââââââââââââââââââââââ
            ws.row_dimensions[3].height = 42
            for c in range(1, col_conc + 1):
                cell = ws.cell(3, c)
                if c in (col_tot1, col_tot2):
                    cell.fill = _fill(C_AMBER)
                    cell.font = _font(bold=True, color=C_NAVY, size=9)
                elif c == col_conc:
                    cell.fill = _fill(C_GREEN)
                    cell.font = _font(bold=True, color=C_WHITE, size=9)
                else:
                    cell.fill = _fill(C_BLUE)
                    cell.font = _font(bold=True, color=C_WHITE, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                cell.border = brd

            # ââ Filas de datos (4 â¦ nf-1) âââââââââââââââââââââââââââââââââââââ
            for r in range(4, nf):
                ws.row_dimensions[r].height = 14
                fondo = _fill(C_ROW1 if r % 2 != 0 else C_ROW2)
                for c in range(1, col_conc + 1):
                    cell = ws.cell(r, c)
                    # Formato numÃ©rico a TODAS las columnas desde percepciones
                    if c >= col_perc_start:
                        cell.number_format = FMT_NUM
                    # Fondo segÃºn tipo de columna
                    if c in (col_tot1, col_tot2):
                        cell.fill = _fill(C_AMBER2)
                        cell.font = _font(bold=True, color=C_NAVY, size=9)
                    elif c == col_conc:
                        cell.fill = _fill(C_GREEN2)
                        cell.font = _font(bold=True, color=C_NAVY, size=9)
                    else:
                        cell.fill = fondo
                        cell.font = _font(size=9)
                    cell.alignment = Alignment(horizontal="right" if c >= col_perc_start
                                               else "left", vertical="center")
                    cell.border = brd

            wb.save(out_path)
            wb.close()

            # ââ DiagnÃ³stico de conceptos no mapeados ââââââââââââââââââââââââââ
            if perc_no_mapeadas:
                self.after(0, self._log,
                    f"â  Percepciones en XML sin columna en CUENTAS A/B "
                    f"(pueden causar CONCILIACION â  0): "
                    + ", ".join(sorted(perc_no_mapeadas)), True)
            if ded_no_mapeadas:
                self.after(0, self._log,
                    f"â  Deducciones en XML sin columna en CUENTAS M/N "
                    f"(pueden causar CONCILIACION â  0): "
                    + ", ".join(sorted(ded_no_mapeadas)), True)
            emp_sin_cuenta = [nu for nu in vistos
                              if not (emp_reg.get(nu) or emp_asim.get(nu))]
            if emp_sin_cuenta:
                self.after(0, self._log,
                    f"â  Empleados sin cuenta en CUENTAS G/H o J/K "
                    f"(neto no escrito â CONCILIACION â  0): "
                    + ", ".join(emp_sin_cuenta), True)

            self.after(0, self._log,
                f"â {len(filas)} XML(s) â resultados/{nombre_out}", False, True)
            self.after(0, self._log, "Listo.", False, True)
            self.after(0, self._pb_detener, self._prov_pb, self._prov_pb_lbl)
            self.after(0, self._prov_pb_frame.grid_remove)
            self.after(0, self._prov_mostrar_descarga, out_path)

        except Exception as exc:
            import traceback as _tb
            self.after(0, self._pb_error, self._prov_pb, self._prov_pb_lbl)
            self.after(0, self._prov_pb_frame.grid_remove)
            self.after(0, self._log,
                f"â Error general: {exc}\n{_tb.format_exc()}", True)


    def _prov_mostrar_descarga(self, out_path):
        """Muestra la barra de descarga con el nombre del archivo generado."""
        self._prov_resultado_path = out_path
        self._prov_lbl_archivo.config(text=os.path.basename(out_path))
        self._prov_barra_descarga.grid()

    # ---------------------------------------------------------------- #
    # PESTAÃA: ESTADO DE CUENTA
    # ---------------------------------------------------------------- #
    def _tab_estado_cuenta(self, nb):
        outer = ttk.Frame(nb)
        nb.add(outer, text="  ð¦ Estado de Cuenta  ")
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(0, weight=1)

        self._ec_archivo      = tk.StringVar(value="")
        self._ec_resultado    = None
        self._ec_banco        = tk.StringVar(value="Banorte DÃ©bito")
        self._ec_filas_todas  = []
        self._ec_sort_col     = None
        self._ec_sort_asc     = True
        self._ec_busqueda     = tk.StringVar()

        # ââ Canvas scrollable: todo el mÃ³dulo en una sola Ã¡rea con scroll ââ
        _cv = tk.Canvas(outer, bg=COLOR_FONDO, highlightthickness=0)
        _vsb = ttk.Scrollbar(outer, orient="vertical", command=_cv.yview)
        _cv.configure(yscrollcommand=_vsb.set)
        _cv.grid(row=0, column=0, sticky="nsew")
        _vsb.grid(row=0, column=1, sticky="ns")

        inn = tk.Frame(_cv, bg=COLOR_FONDO)
        inn.columnconfigure(0, weight=1)
        _win_id = _cv.create_window((0, 0), window=inn, anchor="nw")

        def _sync_width(e):
            _cv.itemconfig(_win_id, width=e.width)
        _cv.bind("<Configure>", _sync_width)

        def _sync_scroll(e=None):
            _cv.configure(scrollregion=_cv.bbox("all"))
        inn.bind("<Configure>", _sync_scroll)

        # Scroll con rueda del ratÃ³n â sÃ³lo sobre el canvas y sus hijos directos
        def _mw(e):
            _cv.yview_scroll(int(-1 * (e.delta / 120)), "units")

        def _bind_mw(widget):
            widget.bind("<MouseWheel>", _mw)
            for child in widget.winfo_children():
                _bind_mw(child)

        # Se llama despuÃ©s de construir todo el contenido
        def _bind_all_children():
            _cv.bind("<MouseWheel>", _mw)
            _bind_mw(inn)
        outer.after(200, _bind_all_children)

        # ââ TÃ­tulo âââââââââââââââââââââââââââââââââââââââââââââââââ
        hdr = tk.Frame(inn, bg=COLOR_FONDO)
        hdr.grid(row=0, column=0, sticky="ew", padx=16, pady=(12, 6))
        tk.Label(hdr, text="Convertir PDF Bancario â Excel",
                 bg=COLOR_FONDO, fg=COLOR_FUCSIA_OSCURO,
                 font=("Segoe UI", 13, "bold")).pack(anchor="w")
        tk.Label(hdr,
                 text="Extrae todas las tablas de estados de cuenta e inversiones con inteligencia artificial",
                 bg=COLOR_FONDO, fg="#888888", font=("Segoe UI", 8)).pack(anchor="w")

        # ââ SecciÃ³n 1: banco ââââââââââââââââââââââââââââââââââââââââ
        s1 = tk.Frame(inn, bg=COLOR_BLANCO,
                      highlightbackground="#D0D0D0", highlightthickness=1)
        s1.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 6))
        s1.columnconfigure(0, weight=1)
        tk.Label(s1, text="1. Selecciona el banco",
                 bg=COLOR_BLANCO, fg=COLOR_TEXTO,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=14, pady=(10, 6))

        banco_grid = tk.Frame(s1, bg=COLOR_BLANCO)
        banco_grid.pack(fill="x", padx=10, pady=(0, 10))
        self._ec_banco_btns = {}
        COLS = 5

        _BANCOS_INFO = [
            ("Auto-detectar",     "#1E6FBF"),
            ("Banorte DÃ©bito",    "#E53935"),
            ("Banorte CrÃ©dito",   "#E53935"),
            ("BBVA DÃ©bito",       "#004481"),
            ("BBVA CrÃ©dito",      "#004481"),
            ("BBVA Pyme",         "#1A6EB5"),
            ("BBVA TDC",          "#004481"),
            ("BBVA LibretÃ³n",     "#004481"),
            ("Banamex DÃ©bito",    "#C62828"),
            ("Banamex CrÃ©dito",   "#C62828"),
            ("Santander DÃ©bito",  "#EC0000"),
            ("Santander CrÃ©dito", "#EC0000"),
            ("HSBC DÃ©bito",       "#DB0011"),
            ("HSBC CrÃ©dito",      "#DB0011"),
            ("Scotiabank DÃ©bito", "#EC111A"),
            ("Scotiabank CrÃ©dito","#EC111A"),
            ("Banregio DÃ©bito",   "#00875A"),
            ("Banregio CrÃ©dito",  "#00875A"),
            ("Inbursa DÃ©bito",    "#1565C0"),
            ("American Express",  "#006FCF"),
            ("Afirme",            "#FF6600"),
        ]

        def _sel_banco(nombre):
            self._ec_banco.set(nombre)
            for n, frm in self._ec_banco_btns.items():
                sel = (n == nombre)
                frm.config(highlightbackground="#1E6FBF" if sel else "#D0D0D0",
                           highlightthickness=2 if sel else 1,
                           bg="#EEF4FF" if sel else COLOR_BLANCO)
                for ch in frm.winfo_children():
                    if isinstance(ch, tk.Label):
                        ch.config(bg="#EEF4FF" if sel else COLOR_BLANCO)

        for idx, (nombre, color) in enumerate(_BANCOS_INFO):
            r, c = divmod(idx, COLS)
            banco_grid.columnconfigure(c, weight=1)
            cell = tk.Frame(banco_grid, bg=COLOR_BLANCO,
                            highlightbackground="#D0D0D0", highlightthickness=1, cursor="hand2")
            cell.grid(row=r, column=c, padx=4, pady=4, sticky="nsew")
            cv2 = tk.Canvas(cell, width=12, height=12, bg=COLOR_BLANCO, highlightthickness=0)
            cv2.pack(anchor="center", pady=(8, 2))
            cv2.create_oval(1, 1, 11, 11, fill=color, outline="")
            tk.Label(cell, text=nombre, bg=COLOR_BLANCO, fg=COLOR_TEXTO,
                     font=("Segoe UI", 8, "bold"), wraplength=110, justify="center").pack(anchor="center")
            tk.Label(cell, text="MXN", bg=COLOR_BLANCO, fg="#AAAAAA",
                     font=("Segoe UI", 7)).pack(anchor="center", pady=(0, 6))
            for w in (cell, cv2): w.bind("<Button-1>", lambda e, n=nombre: _sel_banco(n))
            for ch in cell.winfo_children(): ch.bind("<Button-1>", lambda e, n=nombre: _sel_banco(n))
            self._ec_banco_btns[nombre] = cell
        _sel_banco("Banorte DÃ©bito")
        self._ec_sel_banco = _sel_banco   # accesible desde _ec_elegir_archivo

        # ââ SecciÃ³n 2: archivo ââââââââââââââââââââââââââââââââââââââ
        s2 = tk.Frame(inn, bg=COLOR_BLANCO,
                      highlightbackground="#D0D0D0", highlightthickness=1)
        s2.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 6))
        s2.columnconfigure(0, weight=1)
        tk.Label(s2, text="2. Sube el estado de cuenta PDF",
                 bg=COLOR_BLANCO, fg=COLOR_TEXTO,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=14, pady=(10, 6))

        drop_outer = tk.Frame(s2, bg="#F0F6FF",
                              highlightbackground="#B0C8E8", highlightthickness=2, cursor="hand2")
        drop_outer.pack(fill="x", padx=14, pady=(0, 12))
        self._ec_drop_icon = tk.Label(drop_outer, text="ð",
            font=("Segoe UI", 20), bg="#F0F6FF", fg="#6090C0")
        self._ec_drop_icon.pack(pady=(14, 2))
        self._ec_drop_lbl = tk.Label(drop_outer, text="Clic para seleccionar PDF",
            font=("Segoe UI", 10, "bold"), bg="#F0F6FF", fg="#2B4A70")
        self._ec_drop_lbl.pack()
        self._ec_drop_sub = tk.Label(drop_outer,
            text="Arrastra el PDF aquÃ­ o haz clic para seleccionar",
            font=("Segoe UI", 8), bg="#F0F6FF", fg="#7090B0")
        self._ec_drop_sub.pack(pady=(2, 14))

        def _abrir_pdf(e=None):
            self._ec_elegir_archivo()
            ruta = self._ec_archivo.get()
            if ruta:
                nombre = ruta.split("/")[-1].split("\\")[-1]
                self._ec_drop_icon.config(text="â")
                self._ec_drop_lbl.config(text=nombre, fg=COLOR_FUCSIA_OSCURO,
                                          font=("Segoe UI", 9, "bold"))
                self._ec_drop_sub.config(text="Archivo listo para convertir")

        for w in (drop_outer, self._ec_drop_icon, self._ec_drop_lbl, self._ec_drop_sub):
            w.bind("<Button-1>", _abrir_pdf)
        def _hin(e):
            drop_outer.config(bg="#DFF0FF")
            for lw in (self._ec_drop_icon, self._ec_drop_lbl, self._ec_drop_sub): lw.config(bg="#DFF0FF")
        def _hout(e):
            drop_outer.config(bg="#F0F6FF")
            for lw in (self._ec_drop_icon, self._ec_drop_lbl, self._ec_drop_sub): lw.config(bg="#F0F6FF")
        for w in (drop_outer, self._ec_drop_icon, self._ec_drop_lbl, self._ec_drop_sub):
            w.bind("<Enter>", _hin); w.bind("<Leave>", _hout)

        if _DND_OK:
            def _drop_pdf(event):
                import re as _re
                raw = event.data.strip()
                paths = _re.findall(r'\{([^}]+)\}|(\S+)', raw)
                paths = [a or b for a, b in paths]
                pdf = next((p for p in paths if p.lower().endswith(".pdf")), None)
                if pdf:
                    self._ec_archivo.set(pdf)
                    nombre = pdf.split("/")[-1].split("\\")[-1]
                    self._ec_drop_icon.config(text="â")
                    self._ec_drop_lbl.config(text=nombre, fg=COLOR_FUCSIA_OSCURO,
                                              font=("Segoe UI", 9, "bold"))
                    self._ec_drop_sub.config(text="Archivo listo para convertir")
                else:
                    self._ec_drop_lbl.config(text="â  Solo archivos PDF", fg="red")
            drop_outer.drop_target_register(DND_FILES)
            drop_outer.dnd_bind("<<Drop>>", _drop_pdf)

        # ââ BotÃ³n convertir âââââââââââââââââââââââââââââââââââââââââ
        tk.Button(inn,
            text="â  Extraer tablas y generar Excel",
            bg="#1B2B4B", fg=COLOR_BLANCO,
            font=("Segoe UI", 10, "bold"), relief="flat",
            padx=16, pady=11, cursor="hand2",
            activebackground="#2C4070",
            command=self._ec_convertir,
        ).grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 4))

        # ââ Barra de progreso âââââââââââââââââââââââââââââââââââââââ
        _ec_pb_frame = tk.Frame(inn, bg=COLOR_FONDO)
        _ec_pb_frame.grid(row=4, column=0, sticky="ew", padx=16, pady=(0, 2))
        _ec_pb_frame.grid_remove()
        self._ec_pb_frame = _ec_pb_frame
        self._ec_pb = FunkyProgressBar(_ec_pb_frame, maximum=100, height=70)
        self._ec_pb.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self._ec_pb_lbl = ttk.Label(_ec_pb_frame, text="0 %", width=8,
            foreground=COLOR_FUCSIA_OSCURO, font=("Segoe UI", 8, "bold"))
        self._ec_pb_lbl.pack(side="left")

        # ââ Zona resultado ââââââââââââââââââââââââââââââââââââââââââ
        bot = tk.Frame(inn, bg=COLOR_FONDO)
        bot.grid(row=5, column=0, sticky="ew", padx=8, pady=(0, 8))
        bot.columnconfigure(0, weight=1)

        # Toolbar
        tb = tk.Frame(bot, bg=COLOR_BLANCO,
                      highlightbackground="#D0D0D0", highlightthickness=1)
        tb.grid(row=0, column=0, sticky="ew", pady=(0, 2))

        def _sep():
            tk.Frame(tb, bg="#D0D0D0", width=1).pack(side="left", fill="y", padx=4, pady=3)

        tk.Label(tb, text="ð Resultado:", bg=COLOR_BLANCO, fg=COLOR_TEXTO,
                 font=("Segoe UI", 8)).pack(side="left", padx=(10, 2), pady=4)
        self._ec_lbl_archivo = tk.Label(tb, text="â", bg=COLOR_BLANCO,
            fg=COLOR_FUCSIA_OSCURO, font=("Segoe UI", 8, "bold"))
        self._ec_lbl_archivo.pack(side="left", padx=(0, 6), pady=4)
        _sep()
        ttk.Button(tb, text="ð Copiar",     command=self._ec_copiar).pack(side="left", padx=2, pady=4)
        ttk.Button(tb, text="Î£ Autosuma",   command=self._ec_autosuma).pack(side="left", padx=2, pady=4)
        ttk.Button(tb, text="Quitar dup",   command=self._ec_quitar_dup).pack(side="left", padx=2, pady=4)
        _sep()
        ttk.Button(tb, text="AâZ", command=lambda: self._ec_ordenar_col(self._ec_sort_col or "Fecha", True)).pack(side="left", padx=2, pady=4)
        ttk.Button(tb, text="ZâA", command=lambda: self._ec_ordenar_col(self._ec_sort_col or "Fecha", False)).pack(side="left", padx=2, pady=4)
        ttk.Button(tb, text="ð Buscar",    command=self._ec_toggle_buscar).pack(side="left", padx=2, pady=4)
        ttk.Button(tb, text="â Filtro",     command=self._ec_limpiar_filtro).pack(side="left", padx=2, pady=4)
        _sep()
        ttk.Button(tb, text="ð Abrir Excel", command=self._ec_abrir).pack(side="left", padx=2, pady=4)

        # BÃºsqueda (oculta)
        self._ec_frame_buscar = tk.Frame(bot, bg=COLOR_BLANCO,
                                          highlightbackground="#D0D0D0", highlightthickness=1)
        self._ec_frame_buscar.grid(row=0, column=0, sticky="ew")
        self._ec_frame_buscar.grid_remove()
        tk.Label(self._ec_frame_buscar, text="ð Buscar:", bg=COLOR_BLANCO,
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=(10, 4), pady=5)
        _e = ttk.Entry(self._ec_frame_buscar, textvariable=self._ec_busqueda, width=36)
        _e.pack(side="left", padx=4, pady=5)
        _e.bind("<KeyRelease>", self._ec_filtrar)
        self._ec_lbl_filtro = tk.Label(self._ec_frame_buscar, text="",
            bg=COLOR_BLANCO, fg=COLOR_FUCSIA_OSCURO, font=("Segoe UI", 8, "bold"))
        self._ec_lbl_filtro.pack(side="left", padx=6)

        # Tabla â altura fija grande; el canvas exterior hace el scroll general
        prev = ttk.Frame(bot)
        prev.grid(row=1, column=0, sticky="ew", pady=(2, 0))
        prev.columnconfigure(0, weight=1)
        cols = ("Fecha", "DescripciÃ³n", "DepÃ³sito", "Retiro", "Saldo")
        self._ec_tree = ttk.Treeview(prev, columns=cols, show="headings", height=25)
        for c in cols:
            self._ec_tree.heading(c, text=c, command=lambda _c=c: self._ec_ordenar_col(_c))
            w = 95 if c in ("DepÃ³sito", "Retiro", "Saldo") else (115 if c == "Fecha" else 0)
            self._ec_tree.column(c, width=w,
                anchor="e" if c in ("DepÃ³sito", "Retiro", "Saldo") else "w",
                stretch=(c == "DescripciÃ³n"))
        self._ec_tree.grid(row=0, column=0, sticky="ew")
        # Sin scrollbar propia â el canvas exterior maneja el scroll
        # (pero se puede activar si se necesita scroll rÃ¡pido en la tabla)

        # Resumen
        self._ec_resumen = tk.Frame(bot, bg=COLOR_BLANCO,
                                     highlightbackground="#D0D0D0", highlightthickness=1)
        self._ec_resumen.grid(row=2, column=0, sticky="ew", pady=(4, 0))
        self._ec_lbl_resumen = tk.Label(self._ec_resumen, text="",
            bg=COLOR_BLANCO, fg=COLOR_TEXTO, font=("Segoe UI", 9), justify="left")
        self._ec_lbl_resumen.pack(padx=10, pady=6, anchor="w")

        # ââ MenÃº contextual Treeview: guardar correcciÃ³n ââââââââââââââ
        _ctx_menu = tk.Menu(self._ec_tree, tearoff=0)
        def _ctx_corregir():
            sel = self._ec_tree.selection()
            if not sel: return
            vals = self._ec_tree.item(sel[0])["values"]
            if not vals: return
            desc_orig = str(vals[1]) if len(vals) > 1 else ""
            if not desc_orig: return
            top = tk.Toplevel(self)
            top.title("â Corregir descripciÃ³n")
            top.geometry("520x160")
            top.configure(bg=COLOR_FONDO)
            tk.Label(top, text="DescripciÃ³n original:", bg=COLOR_FONDO,
                     fg="#9CA3AF", font=("Segoe UI", 8)).pack(anchor="w", padx=16, pady=(14, 0))
            tk.Label(top, text=desc_orig, bg=COLOR_FONDO,
                     fg=COLOR_BLANCO, font=("Segoe UI", 9, "bold"), wraplength=480).pack(anchor="w", padx=16)
            tk.Label(top, text="DescripciÃ³n corregida:", bg=COLOR_FONDO,
                     fg="#9CA3AF", font=("Segoe UI", 8)).pack(anchor="w", padx=16, pady=(10, 0))
            _var_corr = tk.StringVar(value=desc_orig)
            _ent = ttk.Entry(top, textvariable=_var_corr, width=64)
            _ent.pack(anchor="w", padx=16, pady=(2, 0))
            _ent.select_range(0, "end"); _ent.focus_set()
            def _guardar():
                nueva = _var_corr.get().strip()
                if nueva and nueva != desc_orig:
                    try:
                        import aprendizaje as _ap
                        _ap.guardar_correccion(desc_orig, nueva)
                        self._ec_tree.set(sel[0], "DescripciÃ³n", nueva)
                        self._log(f"ð§  CorrecciÃ³n guardada: '{desc_orig}' â '{nueva}'", False)
                    except Exception as _e:
                        messagebox.showerror("Error", str(_e), parent=top)
                top.destroy()
            tk.Button(top, text="ð¾ Guardar correcciÃ³n", command=_guardar,
                      bg="#1E3A8A", fg="white", relief="flat", padx=12, pady=4,
                      font=("Segoe UI", 9, "bold")).pack(pady=10)
            top.bind("<Return>", lambda e: _guardar())
            top.transient(self); top.grab_set()
        _ctx_menu.add_command(label="â Corregir descripciÃ³n (guardar)", command=_ctx_corregir)
        def _ctx_popup(event):
            if self._ec_tree.identify_row(event.y):
                self._ec_tree.selection_set(self._ec_tree.identify_row(event.y))
                _ctx_menu.post(event.x_root, event.y_root)
        self._ec_tree.bind("<Button-3>", _ctx_popup)
        self._ec_tree.bind("<Double-1>", lambda e: _ctx_corregir())

        # ââ SecciÃ³n Aprendizaje ââââââââââââââââââââââââââââââââââââââ
        s_ap = tk.Frame(inn, bg=COLOR_BLANCO,
                        highlightbackground="#D0D0D0", highlightthickness=1)
        s_ap.grid(row=6, column=0, sticky="ew", padx=16, pady=(6, 14))
        s_ap.columnconfigure(0, weight=1)
        tk.Label(s_ap, text="ð§  Sistema de Aprendizaje",
                 bg=COLOR_BLANCO, fg=COLOR_TEXTO,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=14, pady=(10, 2))
        tk.Label(s_ap,
                 text="A: corrige descripciones con clic derecho  â¢  B: memoriza bancos automÃ¡ticamente  â¢  C: IA cuando el parser falla",
                 bg=COLOR_BLANCO, fg="#888888", font=("Segoe UI", 8)).pack(anchor="w", padx=14, pady=(0, 8))

        # API Key (OpciÃ³n C)
        api_row = tk.Frame(s_ap, bg=COLOR_BLANCO)
        api_row.pack(fill="x", padx=14, pady=(0, 6))
        tk.Label(api_row, text="API Key Claude (OpciÃ³n C):", bg=COLOR_BLANCO,
                 fg=COLOR_TEXTO, font=("Segoe UI", 8, "bold")).pack(side="left")
        self._ec_api_var = tk.StringVar()
        try:
            import aprendizaje as _ap0
            _k0 = _ap0.obtener_api_key()
            if _k0: self._ec_api_var.set("â¢â¢â¢â¢" + _k0[-4:])
        except Exception: pass
        _api_entry = ttk.Entry(api_row, textvariable=self._ec_api_var, width=44)
        _api_entry.pack(side="left", padx=8)
        def _mostrar_api(e):
            if "â¢" in self._ec_api_var.get():
                self._ec_api_var.set("")
        def _guardar_api():
            val = self._ec_api_var.get().strip()
            if not val or "â¢" in val: return
            try:
                import aprendizaje as _apm
                _apm.guardar_api_key(val)
                self._ec_api_var.set("â¢â¢â¢â¢" + val[-4:])
                messagebox.showinfo("API Key", "â API Key guardada.", parent=self)
            except Exception as _e2:
                messagebox.showerror("Error", str(_e2), parent=self)
        _api_entry.bind("<FocusIn>", _mostrar_api)
        ttk.Button(api_row, text="Guardar", command=_guardar_api).pack(side="left")

        # Botones de gestiÃ³n
        btn_ap = tk.Frame(s_ap, bg=COLOR_BLANCO)
        btn_ap.pack(fill="x", padx=14, pady=(0, 10))
        def _ver_corr():
            try:
                import aprendizaje as _apm
                _apm.abrir_ventana_correcciones(self)
            except Exception as _e3:
                messagebox.showerror("Error", str(_e3), parent=self)
        def _ver_stats():
            try:
                import aprendizaje as _apm
                st = _apm.obtener_estadisticas()
                messagebox.showinfo("EstadÃ­sticas de aprendizaje",
                    f"Correcciones guardadas : {st.get('num_correcciones', 0)}\n"
                    f"Patrones de banco      : {st.get('num_patrones_banco', 0)}\n"
                    f"Archivos procesados    : {st.get('total_archivos_procesados', 0)}\n"
                    f"IA configurada         : {'â SÃ­' if st.get('ia_configurada') else 'â No'}\n"
                    f"Ãltima actualizaciÃ³n   : {st.get('ultima_actualizacion', 'â')}",
                    parent=self)
            except Exception as _e4:
                messagebox.showerror("Error", str(_e4), parent=self)
        ttk.Button(btn_ap, text="â Ver / Editar correcciones",
                   command=_ver_corr).pack(side="left", padx=(0, 8))
        ttk.Button(btn_ap, text="ð EstadÃ­sticas",
                   command=_ver_stats).pack(side="left")


    def _ec_elegir_archivo(self):
        self.lift()
        self.focus_force()
        self.update()
        ruta = filedialog.askopenfilename(
            parent=self,
            initialdir=self.carpeta.get() or os.getcwd(),
            filetypes=[
                ("PDF / Excel", "*.pdf *.xlsx *.xls"),
                ("PDF", "*.pdf"),
                ("Excel", "*.xlsx *.xls"),
                ("Todos", "*.*"),
            ],
        )
        if ruta:
            self._ec_archivo.set(ruta)
            # OpciÃ³n B: sugerir banco por nombre de archivo
            try:
                import aprendizaje as _ap
                sug = _ap.sugerir_banco(ruta)
                if sug and hasattr(self, "_ec_sel_banco"):
                    # Buscar coincidencia exacta o por prefijo en los botones
                    _btns = getattr(self, "_ec_banco_btns", {})
                    _match = next(
                        (n for n in _btns if n == sug or n.startswith(sug.split()[0])),
                        None
                    )
                    if _match:
                        self._ec_sel_banco(_match)
                        self._log(f"ð§  Banco detectado automÃ¡ticamente: {_match}", False)
            except Exception:
                pass

    def _ec_abrir(self):
        if self._ec_resultado and os.path.exists(self._ec_resultado):
            os.startfile(self._ec_resultado)

    def _ec_status(self, msg, color=None):
        """Actualiza la etiqueta de estado dentro del tab (thread-safe vÃ­a after)."""
        color = color or COLOR_FUCSIA_OSCURO
        self._ec_lbl_archivo.config(text=msg, foreground=color)

    def _ec_convertir(self):
        ruta = self._ec_archivo.get().strip()
        if not ruta:
            messagebox.showwarning("Archivo faltante",
                "Selecciona el estado de cuenta antes de continuar.", parent=self)
            return
        if not os.path.exists(ruta):
            messagebox.showwarning("Archivo no encontrado",
                f"No se encontrÃ³ el archivo:\n{ruta}", parent=self)
            return
        saldo_ini = 0.0
        saldo_esp = None

        self.after(0, self._ec_status, "â³ Procesando...", COLOR_AZUL)
        self._ec_pb_frame.grid()
        self._pb_iniciar(self._ec_pb, self._ec_pb_lbl)
        import threading
        threading.Thread(
            target=self._ec_hilo,
            args=(ruta, saldo_ini, saldo_esp),
            daemon=True,
        ).start()

    def _ec_hilo(self, ruta, saldo_ini, saldo_esp):
        import traceback as _tb2
        try:
            # ââ Instalar pdfplumber si falta ââ
            try:
                import pdfplumber
            except ImportError:
                self.after(0, self._ec_status, "â³ Instalando pdfplumber...", COLOR_AZUL)
                import subprocess, sys as _sys
                subprocess.check_call(
                    [_sys.executable, "-m", "pip", "install", "pdfplumber"],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                )
                import pdfplumber

            try:
                import xlsxwriter
            except ImportError:
                self.after(0, self._ec_status, "â³ Instalando xlsxwriter...", COLOR_AZUL)
                import subprocess, sys as _sys
                subprocess.check_call(
                    [_sys.executable, "-m", "pip", "install", "xlsxwriter"],
                    stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                )
                import xlsxwriter

            import openpyxl

            ext = os.path.splitext(ruta)[1].lower()
            movimientos = []

            if ext == ".pdf":
                movimientos = self._ec_leer_pdf(ruta, pdfplumber)
            else:
                movimientos = self._ec_leer_excel(ruta, openpyxl)

            # ââ OpciÃ³n A: aplicar correcciones aprendidas ââââââââââââââ
            if movimientos:
                try:
                    import aprendizaje as _ap
                    movimientos, _n_corr = _ap.aplicar_correcciones(movimientos)
                    if _n_corr:
                        self.after(0, self._log,
                            f"ð§  Aprendizaje: {_n_corr} correcciÃ³n(es) aplicada(s) automÃ¡ticamente.", False)
                except Exception:
                    pass

            # ââ OpciÃ³n C: fallback con IA si parser no encontrÃ³ nada âââ
            if not movimientos:
                try:
                    import aprendizaje as _ap
                    if _ap.obtener_api_key():
                        self.after(0, self._ec_status,
                            "ð¤ Parser fallÃ³ â intentando con IA...", COLOR_AZUL)
                        self.after(0, self._log,
                            "ð¤ Parser convencional sin resultados. Llamando a la API de Claude...", False)
                        _texto_ia = getattr(self, "_ec_dbg_texto", "")
                        _banco_ia = getattr(self, "_ec_dbg_banco", "")
                        movimientos = _ap.parsear_con_ia(_texto_ia, _banco_ia)
                        if movimientos:
                            self.after(0, self._log,
                                f"ð¤ IA extrajo {len(movimientos)} movimientos.", False)
                except Exception as _ia_e:
                    self.after(0, self._log, f"â  IA error: {_ia_e}", True)

            if not movimientos:
                _dbg_banco = getattr(self, "_ec_dbg_banco", "?")
                _dbg_pags  = getattr(self, "_ec_dbg_paginas", 0)
                _dbg_txt   = getattr(self, "_ec_dbg_texto", "")
                _muestra   = _dbg_txt[:800].replace("\n", " | ") if _dbg_txt else "(vacÃ­o)"
                self.after(0, self._pb_error, self._ec_pb, self._ec_pb_lbl)
                self.after(0, self._ec_pb_frame.grid_remove)
                self.after(0, self._log,
                    f"â  Sin movimientos. Banco={_dbg_banco!r}  PÃ¡ginas={_dbg_pags}\n"
                    f"   Texto extraÃ­do (primeros 800 chars):\n   {_muestra}", True)
                self.after(0, self._ec_status, "â  Sin movimientos detectados", COLOR_ERROR)
                self.after(0, messagebox.showwarning,
                    "Sin movimientos",
                    f"No se encontraron movimientos.\n\n"
                    f"Banco: {_dbg_banco}\n"
                    f"PÃ¡ginas con texto: {_dbg_pags}\n\n"
                    "Revisa el Log de actividad para ver el texto extraÃ­do.\n"
                    "(Puede ser un PDF escaneado sin texto seleccionable)")
                return

            # ââ Calcular saldos acumulados ââ
            saldo = saldo_ini
            filas = []
            total_dep = 0.0
            total_ret = 0.0
            for mov in movimientos:
                if len(mov) == 5:
                    # Parser Banorte: saldo real del PDF incluido
                    fecha, desc, dep, ret, saldo = mov
                else:
                    fecha, desc, dep, ret = mov
                    saldo += dep - ret
                total_dep += dep
                total_ret += ret
                filas.append((fecha, desc, dep, ret, saldo))

            # ââ Escribir Excel ââ
            import datetime as _dt
            base = os.path.splitext(os.path.basename(ruta))[0]
            out_dir = self._resultados_dir()
            out_path = os.path.join(out_dir, f"{base}_conciliacion.xlsx")
            # Si el archivo estÃ¡ abierto (Permission denied), usar nombre con timestamp
            if os.path.exists(out_path):
                try:
                    with open(out_path, 'ab'):
                        pass
                except PermissionError:
                    ts = _dt.datetime.now().strftime("%H%M%S")
                    out_path = os.path.join(out_dir, f"{base}_conciliacion_{ts}.xlsx")

            wb = xlsxwriter.Workbook(out_path)
            ws = wb.add_worksheet("Movimientos")
            ws_conc = wb.add_worksheet("ConciliaciÃ³n")

            # Formatos
            fmt_h  = wb.add_format({"bold": True, "bg_color": "#1E6FBF",
                                     "font_color": "#FFFFFF", "border": 1,
                                     "align": "center"})
            fmt_d  = wb.add_format({"num_format": "DD/MM/YYYY", "border": 1})
            fmt_t  = wb.add_format({"border": 1})
            fmt_n  = wb.add_format({"num_format": "#,##0.00", "border": 1})
            fmt_dep = wb.add_format({"num_format": "#,##0.00", "border": 1,
                                      "bg_color": "#E6F2FB"})
            fmt_ret = wb.add_format({"num_format": "#,##0.00", "border": 1,
                                      "bg_color": "#FDE8E4"})
            fmt_sal = wb.add_format({"num_format": "#,##0.00", "border": 1,
                                      "bold": True})
            fmt_neg = wb.add_format({"num_format": "#,##0.00", "border": 1,
                                      "bold": True, "font_color": "#1E6FBF"})
            fmt_ok  = wb.add_format({"num_format": "#,##0.00", "border": 1,
                                      "bold": True, "font_color": "#3E6E96"})
            fmt_lbl = wb.add_format({"bold": True, "border": 1})
            fmt_ttl = wb.add_format({"bold": True, "bg_color": "#CFE7F8",
                                      "num_format": "#,##0.00", "border": 1})

            # ââ Formatos adicionales ââ
            fmt_titulo = wb.add_format({
                "bold": True, "font_size": 14, "font_color": "#1E6FBF",
                "align": "center", "valign": "vcenter",
            })
            fmt_subtit = wb.add_format({
                "bold": True, "font_size": 11, "bg_color": "#AED6F2",
                "border": 1, "align": "center",
            })
            fmt_conc_lbl = wb.add_format({
                "bold": True, "border": 1, "bg_color": "#B7D9EF",
                "align": "left", "indent": 1,
            })
            fmt_conc_val = wb.add_format({
                "num_format": "#,##0.00", "border": 1, "align": "right",
            })
            fmt_conc_tot = wb.add_format({
                "bold": True, "num_format": "#,##0.00", "border": 2,
                "bg_color": "#CFE7F8", "align": "right",
            })
            fmt_conc_ok  = wb.add_format({
                "bold": True, "num_format": "#,##0.00", "border": 2,
                "bg_color": "#E6F2FB", "font_color": "#3E6E96", "align": "right",
            })
            fmt_conc_err = wb.add_format({
                "bold": True, "num_format": "#,##0.00", "border": 2,
                "bg_color": "#FDE8E4", "font_color": "#E14B3D", "align": "right",
            })

            # ââ Hoja Movimientos ââ
            n_filas = len(filas)
            tot_row = n_filas + 1          # fila 0 = encabezado, 1..n = datos

            # TÃ­tulo
            ws.merge_range(0, 0, 0, 4,
                f"Estado de Cuenta â {base}  ({n_filas} movimientos)", fmt_titulo)
            ws.set_row(0, 24)

            # Encabezados (fila 1)
            hdrs = ["Fecha", "DescripciÃ³n", "DepÃ³sito", "Retiro", "Saldo"]
            ws.set_column(0, 0, 14)
            ws.set_column(1, 1, 55)
            ws.set_column(2, 4, 17)
            for c, h in enumerate(hdrs):
                ws.write(1, c, h, fmt_h)

            # Fijar filas de tÃ­tulo + encabezado
            ws.freeze_panes(2, 0)

            # Auto-filtro en fila de encabezados
            ws.autofilter(1, 0, 1 + n_filas, 4)

            # Datos (filas 2 .. n_filas+1)
            for r, (fecha, desc, dep, ret, sal) in enumerate(filas, start=2):
                if isinstance(fecha, str):
                    ws.write(r, 0, fecha, fmt_t)
                else:
                    ws.write_datetime(r, 0, fecha, fmt_d)
                ws.write(r, 1, desc, fmt_t)
                ws.write(r, 2, dep if dep else "", fmt_dep if dep else fmt_t)
                ws.write(r, 3, ret if ret else "", fmt_ret if ret else fmt_t)
                ws.write(r, 4, sal, fmt_sal)

            # Fila totales con fÃ³rmulas SUMA
            tot_row = n_filas + 2           # debajo de los datos
            ws.write(tot_row, 1, "TOTALES", fmt_lbl)
            ws.write_formula(tot_row, 2,
                f"=SUM(C3:C{tot_row})", fmt_ttl)
            ws.write_formula(tot_row, 3,
                f"=SUM(D3:D{tot_row})", fmt_ttl)
            ws.write(tot_row, 4,
                filas[-1][4] if filas else saldo_ini, fmt_ttl)

            # ConfiguraciÃ³n de impresiÃ³n
            ws.set_header(f"&C&B Estado de Cuenta â {base}")
            ws.set_footer("&L&D &T&R PÃ¡gina &P de &N")
            ws.repeat_rows(1)               # repetir encabezado al imprimir
            ws.set_landscape()
            ws.fit_to_pages(1, 0)           # 1 pÃ¡gina de ancho, alto libre

            # ââ Hoja ConciliaciÃ³n ââ
            saldo_fin_real = filas[-1][4] if filas else saldo_ini
            diferencia = (saldo_fin_real - saldo_esp) if saldo_esp is not None else None

            ws_conc.set_column(0, 0, 36)
            ws_conc.set_column(1, 1, 20)

            # TÃ­tulo
            ws_conc.merge_range(0, 0, 0, 1, "ConciliaciÃ³n Bancaria", fmt_titulo)
            ws_conc.set_row(0, 24)
            ws_conc.merge_range(1, 0, 1, 1, base, fmt_subtit)

            conc_data = [
                ("Saldo inicial",               saldo_ini,      fmt_conc_val),
                ("(+) Total depÃ³sitos",          total_dep,      fmt_conc_val),
                ("(-) Total retiros",            total_ret,      fmt_conc_val),
                ("= Saldo final calculado",      saldo_fin_real, fmt_conc_tot),
            ]
            if saldo_esp is not None:
                conc_data.append(("Saldo final esperado (banco)", saldo_esp, fmt_conc_val))
                f_dif = fmt_conc_ok if (diferencia is not None and abs(diferencia) < 0.01) \
                        else fmt_conc_err
                conc_data.append(("Diferencia (calculado â esperado)", diferencia, f_dif))

            for i, (lbl, val, fmt_v) in enumerate(conc_data, start=2):
                ws_conc.write(i, 0, lbl, fmt_conc_lbl)
                ws_conc.write(i, 1, val, fmt_v)

            # Separador visual
            ws_conc.set_row(2, 18)
            ws_conc.set_row(3, 18)
            ws_conc.set_row(4, 18)
            ws_conc.set_row(5, 18)

            wb.close()
            self._ec_resultado = out_path

            # Resumen para UI
            estado = ""
            if saldo_esp is not None:
                if abs(diferencia) < 0.01:
                    estado = "â CONCILIA â diferencia: $0.00"
                else:
                    estado = f"â   NO CONCILIA â diferencia: ${diferencia:,.2f}"
            resumen = (
                f"Movimientos: {len(filas)}   |   "
                f"DepÃ³sitos: ${total_dep:,.2f}   |   "
                f"Retiros: ${total_ret:,.2f}   |   "
                f"Saldo final: ${saldo_fin_real:,.2f}"
                + (f"\n{estado}" if estado else "")
            )

            self.after(0, self._pb_detener, self._ec_pb, self._ec_pb_lbl)
            self.after(0, self._ec_pb_frame.grid_remove)
            self.after(0, self._ec_mostrar_resultado, filas, resumen, out_path)
            self.after(0, self._ec_status,
                f"â {os.path.basename(out_path)}", COLOR_OK)

            # ââ OpciÃ³n B: memorizar banco + registrar uso ââââââââââââââ
            try:
                import aprendizaje as _ap
                _banco_str = self._ec_banco.get().strip().lstrip("ð â").strip()
                _ap.registrar_uso(_banco_str, len(filas), ruta)
                _ap.recordar_banco(ruta, _banco_str)
            except Exception:
                pass

        except Exception as exc:
            det = _tb2.format_exc()
            self.after(0, self._pb_error, self._ec_pb, self._ec_pb_lbl)
            self.after(0, self._ec_pb_frame.grid_remove)
            self.after(0, self._ec_status, f"â Error: {exc}", COLOR_ERROR)
            self.after(0, messagebox.showerror,
                "Error al procesar",
                f"OcurriÃ³ un error:\n\n{exc}\n\n{det[:600]}")

    def _ec_leer_pdf(self, ruta, pdfplumber):
        """Extrae movimientos de un PDF de estado de cuenta.
        Despacha al parser especÃ­fico segÃºn el banco seleccionado,
        o auto-detecta si no se especificÃ³.
        """
        import re

        banco = getattr(self, "_ec_banco", None)
        banco_sel = banco.get().strip() if banco else ""
        # Quitar prefijos de secciÃ³n ("ââ X ââ", "ð ...")
        self._ec_dbg_banco = banco_sel   # para diagnÃ³stico
        banco_key = banco_sel.lstrip("ð â").strip()

        # ââ Recolectar texto completo ââââââââââââââââââââââââââââââââââ
        paginas_texto = []
        paginas_tablas = []
        with pdfplumber.open(ruta) as pdf:
            for page in pdf.pages:
                txt = page.extract_text() or ""
                if txt.strip():
                    paginas_texto.append(txt)
                tbls = page.extract_tables() or []
                paginas_tablas.extend(tbls)

        texto_total = "\n".join(paginas_texto)
        self._ec_dbg_texto   = texto_total        # para diagnÃ³stico
        self._ec_dbg_paginas = len(paginas_texto)  # para diagnÃ³stico

        # ââ Dispatch por banco âââââââââââââââââââââââââââââââââââââââââ
        if banco_key.startswith("Banorte"):
            return self._ec_parsear_banorte(texto_total, ruta=ruta, pdfplumber=pdfplumber)

        if banco_key == "BBVA TDC":
            return self._ec_parsear_bbva_tdc(texto_total, ruta=ruta)

        if banco_key == "BBVA Pyme" or banco_key.startswith("BBVA"):
            # Detectar TDC antes que otras variantes BBVA
            if any(k in texto_total.upper() for k in ("T NEGOC", "LCDIGITAL", "FECHA AUTORIZACION")):
                movs = self._ec_parsear_bbva_tdc(texto_total, ruta=ruta)
                if movs:
                    return movs
            # Cash Management: columnas posicionales (tiene prioridad sobre V45-only)
            if any(k in texto_total.upper() for k in ("CASH MANAGEMENT", "MAESTRA PYME", "OPER LIQ COD")):
                movs = self._ec_parsear_bbva_cashmanagement(ruta, pdfplumber)
                if movs:
                    return movs
            # BBVA Pyme: formato DD/ABR con COD V45 como abono
            if banco_key == "BBVA Pyme":
                movs = self._ec_parsear_bbva_pyme(texto_total)
                if movs:
                    return movs
            # BBVA genÃ©rico: tablas o texto DD/MM/YYYY
            if banco_key.startswith("BBVA"):
                movs = self._ec_parsear_bbva(texto_total, paginas_tablas, banco_key)
                if movs:
                    return movs

        if banco_key.startswith("Banamex"):
            movs = self._ec_parsear_banamex(texto_total, banco_key)
            if movs:
                return movs

        if banco_key.startswith("Santander"):
            movs = self._ec_parsear_santander(texto_total)
            if movs:
                return movs

        if banco_key.startswith("HSBC"):
            movs = self._ec_parsear_hsbc(texto_total)
            if movs:
                return movs

        if banco_key.startswith("Scotiabank"):
            movs = self._ec_parsear_scotiabank(texto_total, paginas_tablas)
            if movs:
                return movs

        if banco_key.startswith("Banregio"):
            movs = self._ec_parsear_banregio(texto_total)
            if movs:
                return movs

        if banco_key.startswith("Inbursa"):
            movs = self._ec_parsear_inbursa(texto_total)
            if movs:
                return movs

        if banco_key.startswith("American Express"):
            movs = self._ec_parsear_amex(texto_total)
            if movs:
                return movs

        if banco_key.startswith("Afirme"):
            movs = self._ec_parsear_afirme(texto_total, ruta=ruta, pdfplumber=pdfplumber)
            if movs:
                return movs

        if banco_key.startswith(("BajÃ­o", "Azteca", "Bancoppel", "Bx+", "CiBanco")):
            pass  # caen al auto-detect genÃ©rico

        # ââ Auto-detecciÃ³n (o banco desconocido) ââââââââââââââââââââââ
        # 1) Banorte por patrÃ³n DD-MMM-YY
        BANORTE_PAT = re.compile(
            r"\d{2}-(?:ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)-\d{2}")
        if BANORTE_PAT.search(texto_total):
            movs = self._ec_parsear_banorte(texto_total)
            if movs:
                return movs

        # 2) BBVA Cash Management (auto-detect por texto)
        if any(k in texto_total.upper() for k in ("CASH MANAGEMENT", "MAESTRA PYME", "OPER LIQ COD")):
            movs = self._ec_parsear_bbva_cashmanagement(ruta, pdfplumber)
            if movs:
                return movs

        # 2b) BBVA TDC (marcadores en texto)
        if any(k in texto_total.upper() for k in ("T NEGOC", "LCDIGITAL")):
            movs = self._ec_parsear_bbva_tdc(texto_total, ruta=ruta)
            if movs:
                return movs

        # 2c) PDF sin texto extraÃ­ble â intentar TDC vÃ­a OCR
        if not texto_total.strip():
            movs = self._ec_parsear_bbva_tdc("", ruta=ruta)
            if movs:
                return movs

        # 3) BBVA por tablas con columnas conocidas
        if any("CARGO" in str(t) or "ABONO" in str(t) for t in paginas_tablas):
            movs = self._ec_parsear_bbva(texto_total, paginas_tablas, "BBVA DÃ©bito")
            if movs:
                return movs

        # 3) GenÃ©rico: lÃ­neas con fecha numÃ©rica
        patron_fecha = re.compile(r"\b(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\b")
        patron_monto = re.compile(r"[\d,]+\.\d{2}")
        movimientos = []
        for linea in texto_total.splitlines():
            if not patron_fecha.search(linea):
                continue
            mov = self._ec_parsear_linea(linea, patron_fecha, patron_monto)
            if mov:
                movimientos.append(mov)
        return movimientos

    # ââ Parsers por banco ââââââââââââââââââââââââââââââââââââââââââââââ

    def _ec_parsear_bbva_tdc(self, texto, ruta=None):
        """
        Parser BBVA T Negoc / LCDigital (Tarjeta de CrÃ©dito Digital).
        Soporta PDFs digitales y PDFs imagen (OCR via pdf2image + pytesseract).

        dep = abonos (pagos) â reducen la deuda
        ret = cargos (compras) â aumentan la deuda
        """
        import re

        _TDC_REAL_STOP = (
            "RESUMEN INFORMATIVO", "SI ESTAS ADHERIDO", "PLAN DE APOYO",
            "RESUMEN DE SUS", "FECHA NOMBRE DE", "FECHA TRANSACCION",
            "SUBTOTAL DE", "TOTAL DE PARCIALIDADES",
        )
        _TDC_SKIP = (
            "IVA :", "IVA:", "INTERES:", "BBVA MEXICO", "AV. PASEO", "LINEA BBVA",
            "T NEGOC", "LCDIGITAL", "ESTADO DE CUENTA", "MOVIMIENTOS EFECTUADOS",
            "AUTORIZACION APLICACION", "TASA ANUAL", "SUCURSAL", "CREDITO DISPONIBLE",
            "LIMITE", "INTERESES", "ESTIMADO TARJETA", "R.F.C.", "NO. DE",
            "PAGO MINIMO", "RENDIMIENTO", "OTROS ABONOS", "OTROS CARGOS",
            "COMISIONES COBRADAS", "MONTO BASE", "DIRECCION", "CENTENO",
            "GRANJAS", "08400", "TOTAL DE", "INCLUIDO EN", "SUBTOTAL",
            "WWW.", "FECHA DE CORTE", "FECHA LIMITE", "CREMACION",
            "CAT ACTUAL", "TASA DE INTERES", "RFC BBA", "REGIMEN",
            "COMPRAS +", "IVA. +", "SALDO INICIAL", "SALDO AL CORTE",
            "FECHA AUTORIZACION", "IMPORTE CARGOS", "IMPORTE ABONOS",
            "LINEA BBVA:", "CIUDAD DE MEXICO", "PAGINA ",
        )

        def _limpiar(raw):
            desc = raw
            desc = re.sub(r'\s*\$\s*[ââ]?\s*-?\s*$', '', desc)
            desc = re.sub(r'\s*[SÂ§s]\$?\s*$', '', desc)
            desc = re.sub(r'\s*\$\s*$', '', desc)
            desc = re.sub(r'\s+[#H*]{2,}\w*', '', desc)
            desc = re.sub(r'\s+\w{8,}\s*$', '', desc)
            desc = re.sub(r'\s+[A-Z]{3}\s+\d{6,}[A-Z0-9]*\s*$', '', desc)
            desc = re.sub(r'\s+[a-z]{2,6}\s*$', '', desc)
            desc = re.sub(r'\s+[a-zA-Z/\\|]{1,2}\s*$', '', desc)
            desc = re.sub(r'\s+', ' ', desc).strip()
            return desc or "â"

        # search (no match) â tolera basura OCR al inicio de lÃ­nea
        pat_tx  = re.compile(r"(\d{2}/\d{2}/\d{2})\s+(\d{2}/\d{2}/\d{2})\s+(.+)$")
        # Abono si: hay signo menos precedido de $/$S/Â§, O si la lÃ­nea termina en "s" sola
        # (s minÃºscula = $ subrayado en la columna ABONOS sin guiÃ³n explÃ­cito)
        pat_neg = re.compile(r'[$SÂ§]\s*[ââ]?\s*-|\b[sSÂ§]\s*$')
        pat_amt = re.compile(r'([\d,]+\.\d{2})')

        def _normalizar_y_unir(lineas):
            """Normaliza lÃ­neas OCR y une lÃ­neas partidas (fecha sin monto + monto sin fecha)."""
            raw = []
            for l in lineas:
                ls = l.strip()
                ls = re.sub(r'(\d{1,3}(?:,\d{3})+)\s(\d{2})(?=\s|$)', r'\1.\2', ls)
                ls = re.sub(r',(?=\.\d)', '', ls)   # "3,339,.54" â "3,339.54"
                if ls:
                    raw.append(ls)
            joined = []
            skip_next = False
            for idx, ls in enumerate(raw):
                if skip_next:
                    skip_next = False
                    continue
                has_date = bool(pat_tx.search(ls))
                has_amt  = bool(pat_amt.search(ls))
                if has_date and not has_amt and idx + 1 < len(raw):
                    nxt = raw[idx + 1]
                    if pat_amt.search(nxt) and not pat_tx.search(nxt):
                        joined.append(ls + ' ' + nxt)
                        skip_next = True
                        continue
                joined.append(ls)
            return joined

        def _procesar_lineas(lineas):
            movs = []
            for ls in _normalizar_y_unir(lineas):
                lu = ls.upper()
                if any(sw in lu for sw in _TDC_REAL_STOP):
                    return movs, True
                if "TOTAL IMPORTES" in lu:
                    if len(pat_amt.findall(ls)) >= 2:
                        return movs, True
                    continue
                # Saltar solo si NO tiene patrÃ³n de fecha â ruido OCR de columnas adyacentes
                if any(sk in lu for sk in _TDC_SKIP) and not pat_tx.search(ls):
                    continue
                m = pat_tx.search(ls)
                if not m:
                    continue
                try:
                    fecha = self._ec_parse_fecha(m.group(2))
                    if isinstance(fecha, str):
                        continue
                except Exception:
                    continue
                desc_raw = m.group(3).strip()
                amounts  = list(re.finditer(r'([\d,]+\.\d{2})', desc_raw))
                if not amounts:
                    continue
                last_m   = amounts[-1]
                try:
                    amount = float(last_m.group(1).replace(",", ""))
                except ValueError:
                    continue   # monto corrupto por OCR
                prefix   = desc_raw[:last_m.start()]
                is_abono = bool(pat_neg.search(prefix))
                dep = amount if is_abono else 0.0
                ret = 0.0   if is_abono else amount
                movs.append((fecha, _limpiar(prefix), dep, ret))
            return movs, False

        # Primero intentar texto extraÃ­ble (PDF digital)
        if texto.strip():
            movs, _ = _procesar_lineas(texto.splitlines())
            if movs:
                return movs

        # OCR fallback para PDFs imagen
        if ruta is None:
            return []
        try:
            from pdf2image import convert_from_path
            import pytesseract
        except ImportError:
            return []

        movimientos = []
        try:
            imgs = convert_from_path(ruta, dpi=200)
        except Exception:
            return []
        for img in imgs:
            text = pytesseract.image_to_string(img, config='--psm 6')
            movs, stop = _procesar_lineas(text.splitlines())
            movimientos.extend(movs)
            if stop:
                break
        return movimientos

    def _ec_parsear_bbva(self, texto, tablas, variante="BBVA DÃ©bito"):
        """Parser BBVA: usa tablas pdfplumber cuando existen,
        cae a texto si no hay tablas Ãºtiles.
        Columnas tÃ­picas: FECHA | DESCRIPCIÃN | CARGO(-) | ABONO(+) | SALDO
        """
        import re
        from datetime import date as _date

        movimientos = []

        # ââ Intentar extraer de tablas ââââââââââââââââââââââââââââââââ
        for tbl in tablas:
            if not tbl or len(tbl) < 2:
                continue
            encabezado = [str(c or "").upper() for c in tbl[0]]
            # Detectar columnas relevantes
            def col_idx(keywords):
                for kw in keywords:
                    for i, h in enumerate(encabezado):
                        if kw in h:
                            return i
                return -1
            i_fecha = col_idx(["FECHA"])
            i_desc  = col_idx(["DESCRIPCI", "CONCEPTO", "MOVIMIENTO"])
            i_cargo = col_idx(["CARGO", "RETIRO", "DÃBITO", "DEBITO"])
            i_abono = col_idx(["ABONO", "DEPÃSITO", "DEPOSITO", "CRÃDITO", "CREDITO"])
            i_saldo = col_idx(["SALDO"])
            if i_fecha < 0 or i_desc < 0:
                continue
            for fila in tbl[1:]:
                if not fila or len(fila) <= max(i_fecha, i_desc):
                    continue
                f_str = str(fila[i_fecha] or "").strip()
                desc  = str(fila[i_desc]  or "").strip()
                if not f_str or not desc:
                    continue
                fecha = self._ec_parse_fecha(f_str)
                if not fecha or isinstance(fecha, str) and len(fecha) > 12:
                    continue
                def _num(idx):
                    if idx < 0 or idx >= len(fila):
                        return 0.0
                    try:
                        return float(str(fila[idx] or "0").replace(",","").replace("$","") or "0")
                    except:
                        return 0.0
                cargo = _num(i_cargo)
                abono = _num(i_abono)
                saldo = _num(i_saldo) if i_saldo >= 0 else None
                dep = abono; ret = cargo
                if saldo is not None:
                    movimientos.append((fecha, desc, dep, ret, saldo))
                else:
                    movimientos.append((fecha, desc, dep, ret))
            if movimientos:
                return movimientos

        # ââ Fallback texto: DD/MM/YYYY o DD-MM-YYYY ââââââââââââââââââ
        pat_f = re.compile(r"(\d{2}[/\-]\d{2}[/\-]\d{4})")
        pat_n = re.compile(r"[\d,]+\.\d{2}")
        for linea in texto.splitlines():
            m = pat_f.search(linea)
            if not m:
                continue
            fecha = self._ec_parse_fecha(m.group(1))
            montos = [float(x.replace(",","")) for x in pat_n.findall(linea)]
            if not montos:
                continue
            desc = linea[:m.start()].strip() + " " + linea[m.end():].strip()
            desc = re.sub(r"[\d,]+\.\d{2}", "", desc).strip() or "â"
            if len(montos) >= 3:
                cargo, abono, saldo = montos[-3], montos[-2], montos[-1]
                dep = abono; ret = cargo
                movimientos.append((fecha, desc, dep, ret, saldo))
            elif len(montos) == 2:
                monto, saldo = montos
                dep = monto; ret = 0.0
                movimientos.append((fecha, desc, dep, ret, saldo))
            elif len(montos) == 1:
                movimientos.append((fecha, desc, montos[0], 0.0))
        return movimientos

    def _ec_parsear_bbva_cashmanagement(self, ruta, pdfplumber):
        """Parser BBVA Cash Management (M.N. S INT y similares).
        Formato: DD/ABR DD/ABR COD DESCRIPCIÃN  [monto_cargo | monto_abono]  saldo_op  saldo_liq
        Usa posiciÃ³n X de cada monto para determinar si es CARGO o ABONO,
        comparando contra las coordenadas X de los encabezados CARGOS/ABONOS.
        """
        import re
        from datetime import date as _date
        from collections import defaultdict

        MESES = {"ENE":1,"FEB":2,"MAR":3,"ABR":4,"MAY":5,"JUN":6,
                 "JUL":7,"AGO":8,"SEP":9,"OCT":10,"NOV":11,"DIC":12}

        pat_fecha  = re.compile(
            r"^(\d{2})/(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)$",
            re.IGNORECASE)
        pat_monto  = re.compile(r"^\d{1,3}(?:,\d{3})*\.\d{2}$")

        movimientos = []

        with pdfplumber.open(ruta) as pdf:
            for page in pdf.pages:
                words = page.extract_words(
                    keep_blank_chars=False, x_tolerance=2, y_tolerance=3)
                if not words:
                    continue

                # ââ Detectar X de columnas CARGOS / ABONOS / SALDO ââââââ
                # Buscamos la FILA que contenga AMBOS "CARGOS" y "ABONOS"
                # para no confundir con el resumen financiero que tambiÃ©n los menciona.
                rows_tmp = defaultdict(list)
                for w in words:
                    rows_tmp[round(w["top"])].append(w)

                x_cargo_hdr = None
                x_abono_hdr = None
                x_saldo_hdr = None
                for y_h in sorted(rows_tmp.keys()):
                    row_w = rows_tmp[y_h]
                    row_texts = [w["text"].upper() for w in row_w]
                    if "CARGOS" in row_texts and "ABONOS" in row_texts:
                        for w in sorted(row_w, key=lambda w: w["x0"]):
                            t = w["text"].upper()
                            if t == "CARGOS" and x_cargo_hdr is None:
                                x_cargo_hdr = w["x0"]
                            elif t == "ABONOS" and x_abono_hdr is None:
                                x_abono_hdr = w["x0"]
                            elif t in ("OPERACIÃN", "OPERACION",
                                       "LIQUIDACIÃN", "LIQUIDACION") \
                                 and x_saldo_hdr is None:
                                x_saldo_hdr = w["x0"]  # primer col. de saldo (menor x)
                        break   # solo la primera fila encabezado

                if x_cargo_hdr is None or x_abono_hdr is None:
                    continue  # pÃ¡gina sin tabla de movimientos

                # Umbral: mitad entre encabezados
                x_sep   = (x_cargo_hdr + x_abono_hdr) / 2
                x_saldo = x_saldo_hdr if x_saldo_hdr else x_abono_hdr + 50

                # ââ Reusar agrupaciÃ³n por fila ya construida ââââââââââ
                rows = rows_tmp

                # ââ Detectar aÃ±o del estado (buscar "AL DD/MM/YYYY") ââ
                anio = _date.today().year
                texto_pag = page.extract_text() or ""
                m_anio = re.search(r"/(\d{4})", texto_pag)
                if m_anio:
                    anio = int(m_anio.group(1))

                # ââ Iterar filas en orden, acumulando continuaciones â
                pat_solo_num = re.compile(r"^[\d\s]+$")
                pat_hex      = re.compile(r"^[0-9A-Fa-f]{8,}$")
                pat_clabe    = re.compile(r"^\d{10,}$")

                _SKIP_CONT = (
                    "BBVA MEXICO", "AV. PASEO DE", "ESTIMADO CLIENTE",
                    "SU ESTADO DE CUENTA", "TOTAL DE MOVIMIENTOS",
                    "TOTAL IMPORTE", "INSTITUCION DE BANCA",
                )

                def _util_cont(tok):
                    """True si el token de continuaciÃ³n aporta info legible."""
                    t = tok.strip()
                    if not t or len(t) < 3:                      return False
                    if pat_clabe.match(t):                        return False  # CLABE pura (solo dÃ­gitos â¥10)
                    if pat_hex.match(t):                          return False  # ID hexadecimal puro
                    if pat_solo_num.match(t):                     return False  # solo dÃ­gitos/espacios
                    if len(t) > 100:                              return False  # texto legal largo
                    # Si empieza con dÃ­gitos pero contiene letras Ãºtiles â incluir
                    # (ej. "0130622KIMBERLY-CLARK..." tiene nombre del emisor)
                    letters = sum(1 for c in t if c.isalpha())
                    if letters < 4:                               return False  # casi sin texto
                    tu = t.upper()
                    if any(s in tu for s in _SKIP_CONT):         return False  # pie de pÃ¡gina
                    return True

                cur_fecha  = None
                cur_desc   = None
                cur_dep    = 0.0
                cur_ret    = 0.0
                cur_saldo  = None
                cur_conts  = []   # lÃ­neas de continuaciÃ³n acumuladas

                def _flush():
                    if cur_fecha is None or cur_desc is None:
                        return
                    # Filtrar y anexar continuaciones Ãºtiles
                    utiles = [c for c in cur_conts if _util_cont(c)]
                    desc_final = cur_desc
                    if utiles:
                        desc_final = cur_desc + " | " + " | ".join(utiles[:3])
                    if cur_saldo is not None:
                        movimientos.append((cur_fecha, desc_final, cur_dep, cur_ret, cur_saldo))
                    else:
                        movimientos.append((cur_fecha, desc_final, cur_dep, cur_ret))

                for y in sorted(rows.keys()):
                    ws = sorted(rows[y], key=lambda w: w["x0"])
                    tokens = [w["text"] for w in ws]
                    xs     = [w["x0"]  for w in ws]

                    if len(tokens) < 2:
                        continue

                    # Â¿Es fila de transacciÃ³n? (primeros 2 tokens = DD/MMM)
                    m1 = pat_fecha.match(tokens[0])
                    m2 = pat_fecha.match(tokens[1]) if len(tokens) > 1 else None

                    if m1 and m2:
                        # Guardar la transacciÃ³n anterior
                        _flush()
                        cur_conts = []

                        try:
                            dia = int(m1.group(1))
                            mes = MESES[m1.group(2).upper()]
                            cur_fecha = _date(anio, mes, dia)
                        except Exception:
                            cur_fecha = None
                            continue

                        cur_dep = cur_ret = 0.0
                        cur_saldo = None
                        for tok, x in zip(tokens, xs):
                            if not pat_monto.match(tok):
                                continue
                            val = float(tok.replace(",", ""))
                            if x >= x_saldo - 5:
                                cur_saldo = val
                            elif x >= x_sep:
                                cur_dep = val
                            else:
                                cur_ret = val

                        if cur_dep == 0.0 and cur_ret == 0.0:
                            cur_fecha = None
                            continue

                        # DescripciÃ³n principal: tokens[3..] hasta el primer monto
                        desc_tokens = []
                        for tok in tokens[3:]:
                            if pat_monto.match(tok):
                                break
                            desc_tokens.append(tok)
                        cur_desc = " ".join(desc_tokens).strip() or (tokens[2] if len(tokens) > 2 else "")

                    elif cur_fecha is not None:
                        # LÃ­nea de continuaciÃ³n â concatenar todos los tokens
                        line = " ".join(tokens).strip()
                        if line:
                            cur_conts.append(line)

                _flush()  # Ãºltima transacciÃ³n de la pÃ¡gina

        return movimientos

    def _ec_parsear_bbva_pyme(self, texto):
        """Parser BBVA Maestra PYME.
        Formato: DD/MMM DD/MMM COD DESCRIPCIÃN [CARGO|ABONO] [saldo_op saldo_liq]
        COD V45 = VENTAS CREDITO â abono; todos los demÃ¡s â cargo.
        Saldo aparece al final de cada grupo de fecha (no en cada lÃ­nea).
        """
        import re
        from datetime import date as _date

        MESES = {"ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
                 "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12}

        y = re.search(r"\b(20\d{2})\b", texto)
        anio = int(y.group(1)) if y else 2026

        pat = re.compile(
            r"^(\d{2})/(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)"
            r"\s+\d{2}/[A-Z]{3}\s+(\w+)\s+(.*?)$", re.IGNORECASE)
        pat_monto = re.compile(r"([\d,]+\.\d{2})")

        # LÃ­neas de cabecera/pie que aparecen en saltos de pÃ¡gina
        _BBVA_SKIP = (
            "BBVA MEXICO", "MAESTRA PYME BBVA", "ESTADO DE CUENTA",
            "PAGINA ", "PÃGINA ", "NO. CUENTA", "NO. CLIENTE",
            "FECHA SALDO", "OPER LIQ COD", "AV. PASEO DE LA REFORMA",
            "R.F.C. BBA", "ALCALDIA", "CIUDAD DE MEXICO",
        )

        movimientos = []
        cur_fecha  = None
        cur_cod    = None
        cur_desc   = None
        cur_dep    = 0.0
        cur_ret    = 0.0
        cur_saldo  = None
        cur_conts  = []   # lÃ­neas de continuaciÃ³n

        def _flush_bbva():
            if cur_fecha is None or cur_desc is None:
                return
            desc = cur_desc
            # Para SPEI: agregar beneficiario de lÃ­neas de continuaciÃ³n
            if cur_conts:
                # Filtrar refs numÃ©ricas, CVE RASTREO y texto legal/promo (>120 chars)
                utiles = []
                for cl in cur_conts[:6]:   # mÃ¡x 6 lÃ­neas de continuaciÃ³n
                    if len(cl) > 120: continue              # texto legal/promo
                    if re.match(r"^\d{6,}", cl): continue   # CLABE / ref numÃ©rica
                    if re.match(r"^BNET", cl, re.I): continue  # CVE RASTREO
                    if re.match(r"^\d+TRANSFERENCIA", cl, re.I): continue
                    utiles.append(cl.strip())
                if utiles:
                    desc = desc + " | " + " ".join(utiles)
            desc = re.sub(r"\s+", " ", desc).strip() or "â"
            if cur_saldo is not None:
                movimientos.append((cur_fecha, desc, cur_dep, cur_ret, cur_saldo))
            else:
                movimientos.append((cur_fecha, desc, cur_dep, cur_ret))

        for line in texto.splitlines():
            lu = line.strip().upper()
            # Ignorar cabeceras de pÃ¡gina
            if any(s in lu for s in _BBVA_SKIP):
                continue

            m = pat.match(line.strip())
            if m:
                _flush_bbva()
                dia = int(m.group(1))
                mes_str = m.group(2).upper()
                cur_cod = m.group(3)
                resto   = m.group(4).strip()
                try:
                    cur_fecha = _date(anio, MESES[mes_str], dia)
                except Exception:
                    cur_fecha = None; continue

                numeros = [float(x.replace(",", "")) for x in pat_monto.findall(resto)]
                if not numeros:
                    cur_fecha = None; continue

                monto = numeros[0]
                is_abono = (cur_cod.upper() == "V45")
                cur_dep  = monto if is_abono else 0.0
                cur_ret  = 0.0 if is_abono else monto
                cur_desc = re.sub(r"\s+", " ", pat_monto.sub("", resto)).strip() or cur_cod
                cur_saldo = numeros[-1] if len(numeros) >= 2 else None
                cur_conts = []
            else:
                if cur_fecha is not None and line.strip():
                    cur_conts.append(line.strip())

        _flush_bbva()

        return movimientos

    def _ec_parsear_banamex(self, texto, variante="Banamex DÃ©bito"):
        """Parser Citibanamex MiCuenta.
        Formato: DD MMM (sin aÃ±o) con descripciÃ³n multi-lÃ­nea.
        Los montos aparecen al final de la Ãºltima lÃ­nea del bloque: monto saldo
        o solo saldo. El aÃ±o se extrae del encabezado del PDF.
        Dep vs ret se determina por cambio de saldo.
        """
        import re
        from datetime import date as _date

        MESES = {"ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
                 "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12}

        anio_m = re.search(r"AL\s+\d+\s+DE\s+\w+\s+DE\s+(20\d{2})", texto, re.I)
        anio = int(anio_m.group(1)) if anio_m else 2026

        si_m = re.search(r"Saldo Anterior\s+\$?([\d,]+\.\d{2})", texto)
        saldo_ant = float(si_m.group(1).replace(",", "")) if si_m else 0.0

        pat_fecha = re.compile(
            r"^(\d{1,2})\s+(ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)\s+(.*?)$",
            re.IGNORECASE)
        # Dos nÃºmeros al final de lÃ­nea: monto saldo
        pat_fin = re.compile(r"([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$")
        pat_num = re.compile(r"([\d,]+\.\d{2})")
        # Encabezados de pÃ¡gina a omitir
        pat_skip = re.compile(
            r"^ESTADO DE CUENTA|^CLIENTE:|^P[aÃ¡]gina\s*\d|^IRVING|"
            r"Centro de Atenci|^Ciudad de M[eÃ©]xico|^Resto del pa[iÃ­]s|"
            r"^DETALLE DE OPERACIONES|^FECHA\s+CONCEPTO|^\d{6}\.",
            re.IGNORECASE)
        # Fin del detalle de operaciones
        pat_fin_det = re.compile(
            r"GRAFICO TRANSACCIONAL|Otros cargos|TOTAL DE MOVIMIENTOS|"
            r"CONCEPTOS\s*$|ClavProd Serv|SUBTOTALES",
            re.IGNORECASE)

        lineas = texto.splitlines()
        i = 0
        movimientos = []
        saldo_act = saldo_ant

        while i < len(lineas):
            linea = lineas[i].strip()
            if pat_fin_det.search(linea):
                break

            m = pat_fecha.match(linea)
            if not m:
                i += 1
                continue

            dia, mes_str, desc_ini = (int(m.group(1)), m.group(2).upper(),
                                      m.group(3).strip())

            if re.search(r"SALDO ANTERIOR", desc_ini, re.I):
                nums = pat_num.findall(desc_ini)
                if nums:
                    saldo_act = float(nums[-1].replace(",", ""))
                i += 1
                continue
            if re.search(r"^(FECHA|CONCEPTO|RETIROS|DEPOSITOS)", desc_ini, re.I):
                i += 1
                continue

            try:
                fecha = _date(anio, MESES[mes_str], dia)
            except Exception:
                i += 1
                continue

            # Recopilar bloque multi-lÃ­nea
            j = i + 1
            bloque_lines = [linea]
            while j < len(lineas):
                nl = lineas[j].strip()
                if pat_fin_det.search(nl):
                    break
                if pat_fecha.match(nl):
                    break
                if pat_skip.search(nl):
                    j += 1
                    continue
                if nl:
                    bloque_lines.append(nl)
                j += 1

            # Ãltima lÃ­nea con dos nÃºmeros al final = monto + saldo
            monto = None
            saldo_nuevo = None
            for bl in reversed(bloque_lines):
                pm = pat_fin.search(bl)
                if pm:
                    monto = float(pm.group(1).replace(",", ""))
                    saldo_nuevo = float(pm.group(2).replace(",", ""))
                    break

            if saldo_nuevo is not None:
                diff = round(saldo_nuevo - saldo_act, 2)
                if monto is None:
                    monto = abs(diff)
                dep = monto if diff >= 0 else 0.0
                ret = 0.0 if diff >= 0 else monto
                saldo_act = saldo_nuevo
                desc = re.sub(r"\s*([\d,]+\.\d{2})\s*", " ", desc_ini).strip()
                desc = re.sub(r"\s+", " ", desc).strip() or "â"
                movimientos.append((fecha, desc, dep, ret, saldo_nuevo))
            i = j

        return movimientos

    def _ec_parsear_santander(self, texto):
        """Parser Santander: DD-MMM-YYYY FOLIO DESCRIPCION dep_o_ret saldo.
        Maneja entradas de una lÃ­nea y entradas SPEI multi-lÃ­nea donde los
        montos aparecen en el Ãºltimo renglÃ³n del bloque.
        """
        import re
        from datetime import date as _date

        MESES = {"ENE":1,"FEB":2,"MAR":3,"ABR":4,"MAY":5,"JUN":6,
                 "JUL":7,"AGO":8,"SEP":9,"OCT":10,"NOV":11,"DIC":12}

        pat_fecha = re.compile(
            r"^(\d{2})-([A-Z]{3})-(\d{4})\s+(\d+)\s+(.*?)$")
        pat_monto = re.compile(r"([\d,]+\.\d{2})")

        movimientos = []
        # Extraer saldo del periodo anterior para que el primer movimiento sea correcto
        saldo_ant = None
        si_m = re.search(r"(?:SALDO\s+FINAL\s+DEL\s+PERIODO\s+ANTERIOR|SALDO\s+ANTERIOR|Saldo\s+inicial)"
                         r"[:\s\$]*([\d,]+\.\d{2})", texto, re.I)
        if si_m:
            try:
                saldo_ant = float(si_m.group(1).replace(",", ""))
            except Exception:
                pass
        lineas = texto.splitlines()
        i = 0

        while i < len(lineas):
            m = pat_fecha.match(lineas[i].strip())
            if not m:
                i += 1
                continue

            dia = int(m.group(1))
            mes_str = m.group(2)
            year = int(m.group(3))
            desc_ini = m.group(5).strip()

            if mes_str not in MESES:
                i += 1
                continue
            try:
                fecha = _date(year, MESES[mes_str], dia)
            except Exception:
                i += 1
                continue

            # Recoger lÃ­neas de continuaciÃ³n hasta la siguiente fecha
            j = i + 1
            bloque = desc_ini
            while j < len(lineas) and not pat_fecha.match(lineas[j].strip()):
                nl = lineas[j].strip()
                if nl:
                    bloque += " " + nl
                j += 1

            montos = [float(x.replace(",","")) for x in pat_monto.findall(bloque)]
            if montos:
                saldo = montos[-1]
                monto = montos[-2] if len(montos) >= 2 else 0.0
                if saldo_ant is not None:
                    diff = round(saldo - saldo_ant, 2)
                    dep, ret = (monto, 0.0) if diff >= 0 else (0.0, monto)
                else:
                    dep, ret = monto, 0.0
                desc = pat_monto.sub("", desc_ini).strip()
                desc = re.sub(r"\s+", " ", desc).strip() or "â"
                saldo_ant = saldo
                movimientos.append((fecha, desc, dep, ret, saldo))
            i = j

        # Fallback: fechas DD/MM/YYYY (algunos PDFs Santander mÃ¡s antiguos)
        if not movimientos:
            pat_f2 = re.compile(r"(\d{2}/\d{2}/\d{4})")
            for linea in texto.splitlines():
                m2 = pat_f2.search(linea)
                if not m2:
                    continue
                fecha = self._ec_parse_fecha(m2.group(1))
                montos = [float(x.replace(",","")) for x in pat_monto.findall(linea)]
                if not montos:
                    continue
                saldo = montos[-1]
                monto = montos[-2] if len(montos) >= 2 else 0.0
                if saldo_ant is not None:
                    diff = round(saldo - saldo_ant, 2)
                    dep, ret = (monto, 0.0) if diff >= 0 else (0.0, monto)
                else:
                    dep, ret = monto, 0.0
                desc = pat_f2.sub("", linea)
                desc = pat_monto.sub("", desc).strip() or "â"
                saldo_ant = saldo
                movimientos.append((fecha, desc, dep, ret, saldo))

        return movimientos

    def _ec_parsear_hsbc(self, texto):
        """Parser HSBC: formato similar a Santander."""
        return self._ec_parsear_santander(texto)

    def _ec_parsear_scotiabank(self, texto, tablas=None):
        """Parser Scotiabank: DD MMM CONCEPTO multi-lÃ­nea, montos con $.
        Formato real: fecha 'DD MMM', descripciÃ³n en mÃºltiples renglones,
        Ãºltimo renglÃ³n tiene '$monto' y '$saldo'.
        """
        import re
        from datetime import date as _date

        MESES = {"ENE":1,"FEB":2,"MAR":3,"ABR":4,"MAY":5,"JUN":6,
                 "JUL":7,"AGO":8,"SEP":9,"OCT":10,"NOV":11,"DIC":12,
                 "JAN":1,"APR":4,"AUG":8,"OCT":10,"NOV":11,"DEC":12}

        # AÃ±o desde encabezado: "01-OCT-18" â 2018, o "Fecha de corte 31-OCT-18"
        anio = 2024
        y4 = re.search(r"\b(20\d{2})\b", texto)
        if y4:
            anio = int(y4.group(1))
        else:
            y2 = re.search(r"\b(\d{2})[/\-](\d{2})[/\-](\d{2})\b", texto)
            if y2:
                anio = 2000 + int(y2.group(3))

        pat_fecha = re.compile(r"^(\d{2})\s+([A-Z]{3})\s+(.*?)$")
        pat_monto_d = re.compile(r"\$([\d,]+\.\d{2})")
        pat_monto = re.compile(r"([\d,]+\.\d{2})")

        # Extraer saldo inicial del encabezado (evita error en primer movimiento)
        movimientos = []
        saldo_ant = None
        si_m = re.search(r"(?:Saldo\s+inicial|SALDO\s+ANTERIOR|Saldo\s+anterior)[^\d]+([\d,]+\.\d{2})", texto, re.I)
        if si_m:
            try:
                saldo_ant = float(si_m.group(1).replace(",", ""))
            except Exception:
                pass
        lineas = texto.splitlines()
        i = 0

        while i < len(lineas):
            m = pat_fecha.match(lineas[i].strip())
            if not m:
                i += 1
                continue

            dia = int(m.group(1))
            mes_str = m.group(2)
            desc_ini = m.group(3).strip()

            if mes_str not in MESES or dia < 1 or dia > 31:
                i += 1
                continue
            try:
                fecha = _date(anio, MESES[mes_str], dia)
            except Exception:
                i += 1
                continue

            # Recoger continuaciÃ³n hasta siguiente lÃ­nea de fecha
            j = i + 1
            bloque_lines = [desc_ini]
            while j < len(lineas) and not pat_fecha.match(lineas[j].strip()):
                nl = lineas[j].strip()
                if nl:
                    bloque_lines.append(nl)
                j += 1
            bloque = " ".join(bloque_lines)

            # Preferir montos con $, si no hay usar sin $
            montos = [float(x.replace(",","")) for x in pat_monto_d.findall(bloque)]
            if not montos:
                montos = [float(x.replace(",","")) for x in pat_monto.findall(bloque)]

            if montos:
                saldo = montos[-1]
                monto = montos[-2] if len(montos) >= 2 else 0.0
                if saldo_ant is not None:
                    diff = round(saldo - saldo_ant, 2)
                    dep, ret = (monto, 0.0) if diff >= 0 else (0.0, monto)
                else:
                    dep, ret = monto, 0.0
                desc = pat_monto_d.sub("", desc_ini)
                desc = pat_monto.sub("", desc).strip() or desc_ini or "â"
                desc = re.sub(r"\s+", " ", desc).strip() or "â"
                saldo_ant = saldo
                movimientos.append((fecha, desc, dep, ret, saldo))
            i = j

        return movimientos

    def _ec_parsear_banregio(self, texto):
        """Parser Banregio: similar a Santander."""
        return self._ec_parsear_santander(texto)

    def _ec_parsear_inbursa(self, texto):
        """Parser para estados de cuenta Inbursa.
        Formato: MMM. DD <referencia> <CONCEPTO> <monto> <saldo>
        Ejemplo:  MAY. 04 ansaccion_ DEPOSITO INBURED 718.48 2,596,167.41
        El saldo viene explÃ­cito en el PDF; se detecta dep vs ret por cambio de saldo.
        """
        import re
        from datetime import date as _date

        MESES = {"ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
                 "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12}

        # Detectar aÃ±o del encabezado (ej. "Del 01 May. 2026")
        year_m = re.search(r"\b(20\d{2})\b", texto)
        anio = int(year_m.group(1)) if year_m else 2026

        pat_linea = re.compile(r"^([A-Z]{3})\.\s{1,3}(\d{1,2})\s+(.*)", re.MULTILINE)
        pat_monto = re.compile(r"([\d,]+\.\d{2})")

        movimientos = []
        saldo_ant = None

        for m in pat_linea.finditer(texto):
            mes_str = m.group(1)
            dia_str = m.group(2)
            resto = m.group(3).strip()

            if mes_str not in MESES:
                continue
            if re.search(r"\b(REFERENCIA|CONCEPTO|FECHA)\b", resto):
                continue

            try:
                fecha = _date(anio, MESES[mes_str], int(dia_str))
            except Exception:
                continue

            montos_raw = pat_monto.findall(resto)
            montos = []
            for s in montos_raw:
                try:
                    montos.append(float(s.replace(",", "")))
                except Exception:
                    pass

            if not montos:
                continue

            saldo = montos[-1]

            # LÃ­nea de saldo inicial
            if "BALANCE INICIAL" in resto.upper():
                saldo_ant = saldo
                continue

            monto = montos[-2] if len(montos) >= 2 else abs(saldo - (saldo_ant or saldo))

            # Determinar dep vs ret por cambio de saldo
            if saldo_ant is not None:
                diff = round(saldo - saldo_ant, 2)
                dep, ret = (monto, 0.0) if diff >= 0 else (0.0, monto)
            else:
                dep, ret = monto, 0.0

            # DescripciÃ³n: quitar montos y referencia inicial
            desc = pat_monto.sub("", resto).strip()
            desc = re.sub(r"^\S+\s+", "", desc).strip()   # quitar referencia
            desc = re.sub(r"\s+", " ", desc).strip() or "â"

            saldo_ant = saldo
            movimientos.append((fecha, desc, dep, ret, saldo))

        return movimientos

    def _ec_parsear_amex(self, texto):
        """Parser American Express Business Gold Card.
        Formato: DD deMES DESCRIPCION importe [CR en lÃ­nea siguiente o misma]
        CR = pago/abono (depÃ³sito); sin CR = cargo (retiro).
        Sin columna de saldo â retorna 4-tuplas (fecha, desc, dep, ret).
        """
        import re
        from datetime import date as _date

        MESES = {
            "enero": 1, "ene": 1, "febrero": 2, "feb": 2, "marzo": 3, "mar": 3,
            "abril": 4, "abr": 4, "mayo": 5, "may": 5, "junio": 6, "jun": 6,
            "julio": 7, "jul": 7, "agosto": 8, "ago": 8,
            "septiembre": 9, "sep": 9, "sept": 9, "octubre": 10, "oct": 10,
            "noviembre": 11, "nov": 11, "diciembre": 12, "dic": 12,
        }
        year_m = re.search(r"\b(20\d{2})\b", texto)
        anio = int(year_m.group(1)) if year_m else 2026

        pat_fecha = re.compile(
            r"^(\d{1,2})\s+de\s*([A-Za-zÃ¡Ã©Ã­Ã³ÃºÃÃÃÃÃ]+)\s+(.*?)$", re.IGNORECASE)
        pat_monto = re.compile(r"([\d,]+\.\d{2})")
        pat_corte = re.compile(
            r"Total de las|Estado de Cuenta P[aÃ¡]g|Este no es un|"
            r"Resumen de|Abreviaci[oÃ³]n|N[uÃº]mero de Cuenta", re.I)
        MAX_CONT = 4  # RFC + REF + 1 extra max

        movimientos = []
        lineas = texto.splitlines()
        i = 0
        while i < len(lineas):
            m = pat_fecha.match(lineas[i].strip())
            if not m:
                i += 1
                continue
            dia = int(m.group(1))
            mes_str = m.group(2).lower().strip()
            desc_ini = m.group(3).strip()
            if mes_str not in MESES:
                i += 1
                continue
            try:
                fecha = _date(anio, MESES[mes_str], dia)
            except Exception:
                i += 1
                continue

            # Recopilar lÃ­neas de continuaciÃ³n (RFC, REF, CARGO X DE Y)
            j = i + 1
            cont = 0
            bloque_lines = [desc_ini]
            while j < len(lineas) and cont < MAX_CONT:
                nl = lineas[j].strip()
                if pat_fecha.match(nl) or pat_corte.search(nl):
                    break
                if nl:
                    bloque_lines.append(nl)
                    cont += 1
                j += 1

            bloque = " ".join(bloque_lines)
            montos = [float(x.replace(",", "")) for x in pat_monto.findall(bloque)]
            if not montos:
                i += 1
                continue

            monto = montos[0]
            # CR solo como palabra standalone para no coincidir con "RFCR..."
            is_cr = (bool(re.search(r"\bCR\b", bloque))
                     or "PAGO RECIBIDO" in bloque.upper())
            dep = monto if is_cr else 0.0
            ret = 0.0 if is_cr else monto

            desc = pat_monto.sub("", desc_ini).strip()
            desc = re.sub(r"\bCR\b", "", desc).strip()
            desc = re.sub(r"\s+", " ", desc).strip() or "â"

            movimientos.append((fecha, desc, dep, ret))
            i += 1

        return movimientos


    def _ec_parsear_afirme(self, texto, ruta=None, pdfplumber=None):
        """Parser para estados de cuenta BANCA AFIRME.
        Extrae DESTINATARIO y CONCEPTO de transacciones SPEI multi-linea,
        filtrando encabezados de pagina y contenido CFDI.
        """
        import re
        from datetime import date as _date
        from collections import defaultdict

        MESES = {"ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
                 "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12}

        per_m = re.search(r'Per[iÃ­]odo\s+de\s+(\d{2})([A-Z]{3})(\d{4})AL', texto, re.IGNORECASE)
        if not per_m:
            per_m = re.search(r'(\d{2})([A-Z]{3})(\d{4})AL', texto)
        if per_m:
            mes  = MESES.get(per_m.group(2).upper(), 1)
            anio = int(per_m.group(3))
        else:
            from datetime import datetime as _dt
            anio, mes = _dt.now().year, _dt.now().month

        ini_m = re.search(r'Saldo\s+inicial\s+\$\s*([\d,]+\.\d{2})', texto)
        saldo_ant = float(ini_m.group(1).replace(",", "")) if ini_m else None

        pat_monto = re.compile(r'\$\s*([\d,]+\.\d{2})')

        # Patrones de lineas que NO son datos de transaccion
        _SKIP_TX = ("descripci", "detalle", "informaci", "resumen", "comision",
                    "deposito", "retiro", "saldo", "ganancia", "rendimient",
                    "numero de", "tasa de", "pagina", "sucursal",
                    "direcci", "cliente", "pyme", "periodo", "fecha de")

        # Patrones de lineas de continuacion a ignorar (encabezados, pie, CFDI)
        _SKIP_CONT = (
            "pagina", "sello digital", "cadena original", "este documento",
            "representaci", "detalle de operaciones", "regimen fiscal",
            "av. ju", "banca afirme", "instituci", "r.f.c. ba",
            "ipab", "condusef", "sus ahorros", "saldo inicial",
            "numero de cuenta", "clave bancaria", "lider pyme",
            "gas 122 sa de cv", "cll poniente", "estado de cuenta al",
            "numero de cliente", "metodo de pago", "tipo o factor",
            "uso cfdi", "claveprodserv", "forma de pago",
            "||1.0|", "dqgfup", "ty5ilfr",  # fragmentos de firma digital
            "20-104",                         # numero de producto en encabezado de pagina
            "dia descripcion", "referencia depositos",  # encabezado de columnas
        )

        import unicodedata as _ud

        def _norm(s):
            """Quita acentos para comparar sin importar acentuacion."""
            return _ud.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii').lower()

        def _es_cont_valida(line):
            ln = _norm(line)
            if any(s in ln for s in _SKIP_CONT):
                return False
            # Rechazar lineas muy largas o con muchos caracteres de base64
            if len(line) > 200:
                return False
            if line.count('+') + line.count('=') + line.count('/') > 10:
                return False
            # Rechazar lineas con doble barra (separador CFDI "||")
            if '||' in line:
                return False
            # Rechazar lineas con secuencias largas base64/hex (firmas digitales)
            import re as _re2
            if _re2.search(r'[A-Za-z0-9+/]{40,}', line):
                return False
            return True

        movimientos = []

        # ââ Extraccion posicional con pdfplumber âââââââââââââââââââââââââ
        if ruta is not None and pdfplumber is not None:
            X_DIA_MAX = 50

            all_rows = []
            with pdfplumber.open(ruta) as pdf:
                for page in pdf.pages:
                    words = page.extract_words(
                        keep_blank_chars=False, x_tolerance=2, y_tolerance=3)
                    if not words:
                        continue
                    rows_tmp = defaultdict(list)
                    for w in words:
                        rows_tmp[round(w["top"])].append(w)
                    for y in sorted(rows_tmp.keys()):
                        ws = sorted(rows_tmp[y], key=lambda w: w["x0"])
                        line = " ".join(w["text"] for w in ws)
                        all_rows.append((y, ws[0]["x0"], line))

            pat_dia = re.compile(r"^(\d{1,2})\s+")
            blocks = []
            cur_dia = None
            cur_rows = []

            for y, x0, line in all_rows:
                m = pat_dia.match(line)
                if m and x0 < X_DIA_MAX:
                    dia = int(m.group(1))
                    if 1 <= dia <= 31:
                        if not any(kw in line.lower() for kw in _SKIP_TX):
                            if cur_dia is not None:
                                blocks.append((cur_dia, cur_rows))
                            cur_dia = dia
                            cur_rows = [line]
                            continue
                if cur_dia is not None and _es_cont_valida(line):
                    cur_rows.append(line)
            if cur_dia is not None:
                blocks.append((cur_dia, cur_rows))

            for dia, rows in blocks:
                # Unir lineas y corregir splits de palabras al borde de columna
                bt = " ".join(rows)
                bt = re.sub(r"\bDESTI\s+NATARIO\b", "DESTINATARIO", bt, flags=re.I)
                bt = re.sub(r"\bCON\s+CEPTO\s*:", "CONCEPTO:", bt)
                bt = re.sub(r"\bINST\s+ITUCION\b", "INSTITUCION", bt, flags=re.I)
                bt = re.sub(r"\(\s*DA\s+TO\s+NO\b", "(DATO NO", bt, flags=re.I)

                # Montos
                montos = pat_monto.findall(bt)
                if len(montos) < 2:
                    continue
                try:
                    saldo = float(montos[-1].replace(",", ""))
                    monto = float(montos[-2].replace(",", ""))
                except Exception:
                    continue

                try:
                    fecha = _date(anio, mes, dia)
                except Exception:
                    continue

                # Descripcion de primera linea (sin numero de dia, sin montos)
                fl = rows[0]
                desc = pat_monto.sub("", fl)
                desc = re.sub(r"\$\s*", "", desc)
                desc = re.sub(r"^\d{1,2}\s+", "", desc)
                desc = re.sub(r"\s+\d{1,6}\s*$", "", desc)
                desc = re.sub(r"\s+", " ", desc).strip() or u"â"

                # DESTINATARIO
                dest_m = re.search(
                    r"DESTINATARIO\s*[:\|]\s*(.+?)(?=\s*(?:\(?DATO\s+NO|RFC\s+DEST|ND\s+CVE|CVE\s*RASTREO|CONCEPTO|$))",
                    bt, re.IGNORECASE)
                dest = ""
                if dest_m:
                    dest = dest_m.group(1).strip()
                    dest = re.sub(r"\s*\(?DA\s+TO\s+NO.+$", "", dest, flags=re.I).strip()
                    dest = re.sub(r"\s*\(?DATO\s+NO.+$",    "", dest, flags=re.I).strip()
                    dest = dest.rstrip("(").strip()

                # CONCEPTO (maximo 80 chars para evitar arrastrar CFDI/firmas)
                conc_m = re.search(
                    r"CONCEPTO\s*[:\|]\s*(.+?)(?=\s*(?:HORA|CVE\s+RASTREO|M[eÃ©]todo|RFC|$))",
                    bt, re.IGNORECASE)
                concepto = conc_m.group(1).strip()[:80] if conc_m else ""

                # Ensamblar descripcion
                if dest:
                    desc += " | " + dest
                if concepto and concepto.upper() not in desc.upper():
                    desc += " | " + concepto

                # Deposito vs retiro por diferencia de saldo
                if saldo_ant is not None:
                    diff = round(saldo - saldo_ant, 2)
                    dep = round(diff,  2) if diff > 0 else 0.0
                    ret = round(-diff, 2) if diff < 0 else 0.0
                else:
                    dep, ret = 0.0, monto

                saldo_ant = saldo
                movimientos.append((fecha, desc, dep, ret, saldo))

            if movimientos:
                return movimientos

        # ââ Fallback: parser de texto (sin pdfplumber) âââââââââââââââââââ
        pat_tx = re.compile(r"^(\d{1,2})\s+([^\n]+)", re.MULTILINE)
        saldo_ant2 = None
        ini_m2 = re.search(r"Saldo\s+inicial\s+\$\s*([\d,]+\.\d{2})", texto)
        if ini_m2:
            saldo_ant2 = float(ini_m2.group(1).replace(",", ""))

        for m in pat_tx.finditer(texto):
            dia   = int(m.group(1))
            resto = m.group(2).strip()
            if dia < 1 or dia > 31:
                continue
            if any(kw in resto.lower() for kw in _SKIP_TX):
                continue
            montos_raw = pat_monto.findall(resto)
            if len(montos_raw) < 2:
                continue
            try:
                saldo = float(montos_raw[-1].replace(",", ""))
                monto = float(montos_raw[-2].replace(",", ""))
                fecha = _date(anio, mes, dia)
            except Exception:
                continue
            desc = pat_monto.sub("", resto)
            desc = re.sub(r"\$\s*", "", desc)
            desc = re.sub(r"\s+\d{4,}\s*$", "", desc)
            desc = re.sub(r"\s+", " ", desc).strip() or u"â"
            if saldo_ant2 is not None:
                diff = round(saldo - saldo_ant2, 2)
                dep = round(diff,  2) if diff > 0 else 0.0
                ret = round(-diff, 2) if diff < 0 else 0.0
            else:
                dep, ret = 0.0, monto
            saldo_ant2 = saldo
            movimientos.append((fecha, desc, dep, ret, saldo))

        return movimientos


    def _ec_parsear_banorte(self, texto, ruta=None, pdfplumber=None):
        """Parser especÃ­fico para estados de cuenta Banorte.
        Usa extracciÃ³n por palabras agrupadas (Y) para capturar descripciones
        multi-lÃ­nea (SPEI con BENEF: y RFC:).
        """
        import re
        from datetime import date as _date
        from collections import defaultdict

        MESES = {"ENE":1,"FEB":2,"MAR":3,"ABR":4,"MAY":5,"JUN":6,
                 "JUL":7,"AGO":8,"SEP":9,"OCT":10,"NOV":11,"DIC":12}

        pat_fecha = re.compile(
            r"^(\d{2}-(?:ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|OCT|NOV|DIC)-\d{2})")
        pat_monto = re.compile(r"(?<!\S)([\d,]+\.\d{2})(?!\S|\d)")

        # ââ ExtracciÃ³n de lÃ­neas ââââââââââââââââââââââââââââââââââââââââââ
        # Preferimos word-based (Y-agrupado) si tenemos acceso al PDF
        if ruta is not None and pdfplumber is not None:
            lineas = []
            with pdfplumber.open(ruta) as _pdf:
                for _page in _pdf.pages:
                    _words = _page.extract_words(
                        keep_blank_chars=False, x_tolerance=1, y_tolerance=3)
                    _rows = defaultdict(list)
                    for _w in _words:
                        _rows[round(_w["top"])].append(_w)
                    for _y in sorted(_rows.keys()):
                        _ws = sorted(_rows[_y], key=lambda w: w["x0"])
                        lineas.append(" ".join(w["text"] for w in _ws))
        else:
            lineas = texto.splitlines()

        # ââ Tokens a ignorar (pies de pÃ¡gina, encabezados) âââââââââââââââ
        # LÃ­neas a ignorar (cabeceras/pies de pÃ¡gina, NO quitar RFC de transacciones)
        _SKIP_HDR = (
            "FECHA DESCRIPCI", "MONTO DEL", "DETALLE DE MOVIMIENTOS",
            "ESTADO DE CUENTA", "ENLACE GLOBAL", "LINEA DIRECTA", "LÃNEA DIRECTA",
            "BANCO MERCANTIL", "CIUDAD DE MEXICO", "NUEVO LEON",
            "PAGINA ", "PÃGINA ", "TIPO DE ENVIO", "INFORMACION DEL",
            "NO. DE CLIENTE", "DATOS DE SUCURSAL",
            "PLAZA:", "TELEFONO:", "RESUMEN INTEGRAL", "RESUMEN DEL PERIODO",
            "BANCO MERCANTIL DEL NORTE", "/63", "5140 5640", "3669 9040",
        )
        # Cabeceras que sÃ­ usan RFC: pero deben ignorarse (solo en contexto de inicio de pÃ¡gina)
        _SKIP_HDR_RFC = ("SUCURSAL:", "NO. DE CUENTA", "CLABE INTERBANCARIA")

        # ââ Construir transacciones con lÃ­neas de continuaciÃ³n âââââââââââ
        movimientos = []
        saldo_anterior = None
        cur_fecha_str  = None
        cur_lineas     = []   # lÃ­neas de descripciÃ³n (sin fecha)

        def _flush(fecha_str, lineas_desc):
            """Procesa la transacciÃ³n acumulada y la agrega a movimientos."""
            nonlocal saldo_anterior
            if not fecha_str or not lineas_desc:
                return
            try:
                d, mes_str, a = fecha_str.split("-")
                mes  = MESES[mes_str]
                anio = 2000 + int(a)
                fecha = _date(anio, mes, int(d))
            except Exception:
                return

            # LÃ­nea 1 contiene los montos; lÃ­neas 2+ son continuaciÃ³n de descripciÃ³n
            linea1    = lineas_desc[0] if lineas_desc else ""
            continuas = lineas_desc[1:] if len(lineas_desc) > 1 else []

            tu = linea1.upper()

            # Saldo anterior
            if "SALDO ANTERIOR" in tu:
                m2 = pat_monto.findall(" " + linea1)
                if m2:
                    try:
                        saldo_anterior = float(m2[-1].replace(",", ""))
                    except Exception:
                        pass
                return

            # Saltamos encabezados
            if any(s in tu for s in ("DETALLE DE MOVIMIENTOS", "MONTO DEL",
                                      "DESCRIPCIÃN", "FECHA DESCRIPCI",
                                      "SALDO FINAL", "TOTAL DE")):
                return

            # Montos y saldo vienen siempre en la primera lÃ­nea
            montos_raw = pat_monto.findall(" " + linea1)
            montos = []
            for s in montos_raw:
                try:
                    montos.append(float(s.replace(",", "")))
                except ValueError:
                    pass
            if not montos:
                return

            saldo = montos[-1]
            if len(montos) >= 2:
                monto = montos[-2]
            else:
                monto = abs(saldo - saldo_anterior) if saldo_anterior is not None else 0.0

            if saldo_anterior is not None:
                diff = round(saldo - saldo_anterior, 2)
                dep, ret = (monto, 0.0) if diff >= 0 else (0.0, monto)
            else:
                if any(k in tu for k in ("DEP", "DEPOSITO", "ABONO",
                                          "SPEI RECIBIDO", "SUPER SERV")):
                    dep, ret = monto, 0.0
                else:
                    dep, ret = 0.0, monto

            # ââ DescripciÃ³n: limpiar montos de lÃ­nea1, agregar continuaciones ââ
            desc1 = linea1.strip()
            for s in montos_raw[-2:]:
                idx = desc1.rfind(s)
                if idx >= 0:
                    desc1 = desc1[:idx].rstrip()
            desc1 = re.sub(r"\s+", " ", desc1).strip()

            # Unir con continuaciones (excluir lÃ­neas que son solo refs numÃ©ricas)
            extras = []
            for cl in continuas:
                cl = cl.strip()
                # Saltar lÃ­neas de solo nÃºmeros/referencias cortas
                if re.match(r"^[\d\s]{1,20}$", cl):
                    continue
                # Saltar CVE RASTREO largo y hora de liquidaciÃ³n
                if "CVE RAST" in cl.upper() or "HORA LIQ" in cl.upper():
                    continue
                extras.append(cl)

            desc_extra = " ".join(extras).strip()

            # Para SPEI saliente: preferir BENEF + RFC del texto completo
            texto_cont = " ".join(continuas)
            if "BENEF:" in texto_cont.upper():
                _bm = re.search(r"BENEF:([^,\(\[]+)", texto_cont, re.IGNORECASE)
                _rm = re.search(r"RFC:\s*([A-Z&\xc0-\xff0-9]{10,13})", texto_cont, re.IGNORECASE)
                _benef = _bm.group(1).strip() if _bm else ""
                _rfc   = _rm.group(1).strip() if _rm else ""
                if _benef:
                    desc = desc1 + " | " + _benef + (" RFC:" + _rfc if _rfc else "")
                else:
                    desc = (desc1 + " " + desc_extra).strip() or "â"
            else:
                desc = (desc1 + (" " + desc_extra if desc_extra else "")).strip() or "â"

            desc = re.sub(r"\s+", " ", desc).strip() or "â"

            saldo_anterior = saldo
            movimientos.append((fecha, desc, dep, ret, saldo))

        for linea in lineas:
            lu = linea.upper()
            # Ignorar pies de pÃ¡gina y encabezados de secciÃ³n
            if any(s in lu for s in _SKIP_HDR):
                continue

            m = pat_fecha.match(linea)
            if m:
                _flush(cur_fecha_str, cur_lineas)
                cur_fecha_str = m.group(1)
                # La primera letra de la descripciÃ³n puede quedar pegada a la fecha
                # (ej: "01-MAY-26C OMPRA" â "COMPRA"). La reunimos.
                _resto = linea[m.end():]
                _resto = re.sub(r"^([A-ZÃÃÃÃÃÃ]) ([A-ZÃÃÃÃÃÃ])", r"\1\2", _resto)
                cur_lineas = [_resto.strip()]
            else:
                if cur_fecha_str is not None:
                    cur_lineas.append(linea.strip())

        # Ãltima transacciÃ³n pendiente
        _flush(cur_fecha_str, cur_lineas)

        return movimientos

    def _ec_parsear_fila(self, fila):
        """Parsea una fila de tabla (lista de celdas). Retorna (fecha,desc,dep,ret) o None."""
        import re
        patron_monto = re.compile(r"^-?[\d,]+\.\d{2}$")
        patron_fecha = re.compile(
            r"\b(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\b"
        )

        celdas = [str(c).strip() if c else "" for c in fila]
        # Buscar fecha
        fecha_str = ""
        desc_parts = []
        montos = []
        for celda in celdas:
            m = patron_fecha.search(celda)
            if m and not fecha_str:
                fecha_str = m.group(1)
            elif patron_monto.match(celda.replace(" ", "").replace("(", "-").replace(")", "")):
                val = celda.replace(",", "").replace(" ", "").replace("(", "-").replace(")", "")
                try:
                    montos.append(float(val))
                except ValueError:
                    pass
            elif celda and not patron_fecha.search(celda):
                desc_parts.append(celda)

        if not fecha_str:
            return None

        fecha = self._ec_parse_fecha(fecha_str)
        desc  = " ".join(desc_parts).strip() or "â"

        dep = ret = 0.0
        if len(montos) == 1:
            v = montos[0]
            dep, ret = (v, 0.0) if v > 0 else (0.0, abs(v))
        elif len(montos) >= 2:
            dep = montos[0] if montos[0] > 0 else 0.0
            ret = montos[1] if montos[1] > 0 else 0.0
            # si dep es negativo interpretarlo como retiro
            if montos[0] < 0 and montos[1] >= 0:
                ret, dep = abs(montos[0]), montos[1]
        else:
            return None

        return (fecha, desc, dep, ret)

    def _ec_parsear_linea(self, linea, patron_fecha, patron_monto):
        """Parsea una lÃ­nea de texto libre. Retorna (fecha,desc,dep,ret) o None."""
        m_fecha = patron_fecha.search(linea)
        if not m_fecha:
            return None
        fecha_str = m_fecha.group(1)
        fecha = self._ec_parse_fecha(fecha_str)

        montos_raw = patron_monto.findall(linea)
        montos = []
        for s in montos_raw:
            try:
                montos.append(float(s.replace(",", "")))
            except ValueError:
                pass

        # DescripciÃ³n: texto entre fecha y primer monto
        pos_fecha_fin = m_fecha.end()
        primer_monto_pos = linea.find(montos_raw[0]) if montos_raw else len(linea)
        desc = linea[pos_fecha_fin:primer_monto_pos].strip(" |-\t") or "â"

        dep = ret = 0.0
        if len(montos) == 1:
            dep = montos[0]; ret = 0.0
        elif len(montos) == 2:
            dep = montos[0]; ret = montos[1]
        elif len(montos) >= 3:
            dep = montos[0]; ret = montos[1]
        else:
            return None

        return (fecha, desc, dep, ret)

    def _ec_leer_excel(self, ruta, openpyxl):
        """Extrae movimientos de un Excel de estado de cuenta."""
        import re
        patron_fecha = re.compile(r"\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}")

        wb = openpyxl.load_workbook(ruta, data_only=True)
        ws = wb.active
        movimientos = []

        # Detectar fila de encabezados
        col_fecha = col_desc = col_dep = col_ret = None
        for row in ws.iter_rows():
            vals = [str(c.value or "").lower() for c in row]
            for i, v in enumerate(vals):
                if "fecha" in v and col_fecha is None:
                    col_fecha = i
                if any(x in v for x in ("concepto", "descripci", "referencia")) \
                        and col_desc is None:
                    col_desc = i
                if any(x in v for x in ("dep", "abono", "cargo_cr")) and col_dep is None:
                    col_dep = i
                if any(x in v for x in ("retiro", "cargo", "egreso")) and col_ret is None:
                    col_ret = i
            if col_fecha is not None:
                break

        if col_fecha is None:
            # fallback: columna 0=fecha, 1=desc, 2=dep, 3=ret
            col_fecha, col_desc, col_dep, col_ret = 0, 1, 2, 3

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                fecha_raw = row[col_fecha] if col_fecha < len(row) else None
                desc_raw  = row[col_desc]  if col_desc  < len(row) else None
                dep_raw   = row[col_dep]   if col_dep   < len(row) else None
                ret_raw   = row[col_ret]   if col_ret   < len(row) else None

                if fecha_raw is None:
                    continue
                fecha = self._ec_parse_fecha(str(fecha_raw)) \
                    if not hasattr(fecha_raw, "year") else fecha_raw

                dep = float(str(dep_raw).replace(",", "")) \
                    if dep_raw not in (None, "", "None") else 0.0
                ret = float(str(ret_raw).replace(",", "")) \
                    if ret_raw not in (None, "", "None") else 0.0
                desc = str(desc_raw or "â").strip()

                if dep == 0.0 and ret == 0.0:
                    continue
                movimientos.append((fecha, desc, abs(dep), abs(ret)))
            except Exception:
                continue

        wb.close()
        return movimientos

    @staticmethod
    def _ec_parse_fecha(s):
        """Convierte una cadena de fecha a datetime.date, o devuelve la cadena original."""
        from datetime import datetime as dt
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d", "%d-%m-%Y",
                    "%d-%m-%y", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return dt.strptime(str(s).strip(), fmt).date()
            except ValueError:
                pass
        return str(s).strip()

    def _ec_mostrar_resultado(self, filas, resumen, out_path):
        # Guardar todas las filas para filtrar/ordenar
        self._ec_filas_todas = list(filas)
        self._ec_sort_col    = None
        self._ec_sort_asc    = True
        self._ec_busqueda.set("")
        if hasattr(self, "_ec_frame_buscar"):
            self._ec_frame_buscar.grid_remove()
        self._ec_refrescar_tree(filas)
        self._ec_lbl_resumen.config(text=resumen)
        self._ec_lbl_archivo.config(text=os.path.basename(out_path))

    def _ec_refrescar_tree(self, filas):
        """Repopula el Treeview con la lista de filas dada."""
        for item in self._ec_tree.get_children():
            self._ec_tree.delete(item)
        for fecha, desc, dep, ret, sal in filas:
            f_str = fecha.strftime("%d/%m/%Y") if hasattr(fecha, "strftime") else str(fecha)
            self._ec_tree.insert("", "end", values=(
                f_str, desc,
                f"${dep:,.2f}" if dep else "",
                f"${ret:,.2f}" if ret else "",
                f"${sal:,.2f}",
            ))

    # ââ Toolbar: Copiar ââââââââââââââââââââââââââââââââââââââââââââââ
    def _ec_copiar(self):
        sel = self._ec_tree.selection()
        items = sel if sel else self._ec_tree.get_children()
        if not items:
            messagebox.showinfo("Sin datos", "No hay movimientos para copiar.", parent=self)
            return
        hdrs = ["Fecha", "DescripciÃ³n", "DepÃ³sito", "Retiro", "Saldo"]
        lineas = ["\t".join(hdrs)]
        for iid in items:
            lineas.append("\t".join(str(v) for v in self._ec_tree.item(iid, "values")))
        self.clipboard_clear()
        self.clipboard_append("\n".join(lineas))
        messagebox.showinfo("Copiado",
            f"{len(lineas)-1} fila(s) copiadas al portapapeles.", parent=self)

    # ââ Toolbar: Autosuma ââââââââââââââââââââââââââââââââââââââââââââ
    def _ec_autosuma(self):
        sel = self._ec_tree.selection()
        items = sel if sel else self._ec_tree.get_children()
        if not items:
            messagebox.showinfo("Sin datos", "No hay movimientos.", parent=self)
            return
        total_dep = total_ret = 0.0
        for iid in items:
            v = self._ec_tree.item(iid, "values")
            for i, col in enumerate(("DepÃ³sito", "Retiro")):
                try:
                    val = float(str(v[2 + i]).replace("$", "").replace(",", "") or "0")
                    if i == 0: total_dep += val
                    else:      total_ret += val
                except ValueError:
                    pass
        n = len(items)
        scope = "seleccionadas" if sel else "todas"
        messagebox.showinfo("Autosuma",
            f"Filas {scope}: {n}\n\n"
            f"Total depÃ³sitos:  ${total_dep:>14,.2f}\n"
            f"Total retiros:    ${total_ret:>14,.2f}\n"
            f"âââââââââââââââââââââââââââââ\n"
            f"Diferencia neta:  ${total_dep - total_ret:>14,.2f}",
            parent=self)

    # ââ Toolbar: Ordenar por columna âââââââââââââââââââââââââââââââââ
    def _ec_ordenar_col(self, col, ascending=None):
        if not self._ec_filas_todas:
            return
        if ascending is None:
            # Toggle si se hace clic en el encabezado
            ascending = not self._ec_sort_asc if col == self._ec_sort_col else True
        self._ec_sort_col = col
        self._ec_sort_asc = ascending
        idx_map = {"Fecha": 0, "DescripciÃ³n": 1, "DepÃ³sito": 2, "Retiro": 3, "Saldo": 4}
        idx = idx_map.get(col, 0)

        def key(fila):
            v = fila[idx]
            if isinstance(v, (int, float)): return v
            try: return float(str(v).replace(",","") or "0")
            except: return str(v)

        # Ordenar la lista completa
        sorted_filas = sorted(self._ec_filas_todas, key=key, reverse=not ascending)
        self._ec_filas_todas = sorted_filas
        # Aplicar filtro activo si existe
        q = self._ec_busqueda.get().strip().lower()
        vista = [f for f in sorted_filas if self._ec_fila_match(f, q)] if q else sorted_filas
        self._ec_refrescar_tree(vista)
        # Actualizar indicador en encabezado
        for c in ("Fecha", "DescripciÃ³n", "DepÃ³sito", "Retiro", "Saldo"):
            arrow = (" â²" if ascending else " â¼") if c == col else ""
            self._ec_tree.heading(c, text=c + arrow)

    # ââ Toolbar: Buscar ââââââââââââââââââââââââââââââââââââââââââââââ
    def _ec_toggle_buscar(self):
        if self._ec_frame_buscar.winfo_ismapped():
            self._ec_frame_buscar.grid_remove()
        else:
            self._ec_frame_buscar.grid()

    @staticmethod
    def _ec_fila_match(fila, q):
        return any(q in str(v).lower() for v in fila)

    def _ec_filtrar(self, event=None):
        q = self._ec_busqueda.get().strip().lower()
        if not q:
            vista = self._ec_filas_todas
        else:
            vista = [f for f in self._ec_filas_todas if self._ec_fila_match(f, q)]
        self._ec_refrescar_tree(vista)
        self._ec_lbl_filtro.config(
            text=f"{len(vista)} de {len(self._ec_filas_todas)} movimientos")

    def _ec_limpiar_filtro(self):
        self._ec_busqueda.set("")
        self._ec_refrescar_tree(self._ec_filas_todas)
        if hasattr(self, "_ec_lbl_filtro"):
            self._ec_lbl_filtro.config(text="")

    # ââ Toolbar: Quitar duplicados âââââââââââââââââââââââââââââââââââ
    def _ec_quitar_dup(self):
        if not self._ec_filas_todas:
            messagebox.showinfo("Sin datos", "No hay movimientos cargados.", parent=self)
            return
        vistos = set()
        unicos = []
        for fila in self._ec_filas_todas:
            key = (str(fila[0]), fila[1].strip())   # fecha + descripciÃ³n
            if key not in vistos:
                vistos.add(key)
                unicos.append(fila)
        removed = len(self._ec_filas_todas) - len(unicos)
        if removed == 0:
            messagebox.showinfo("Quitar duplicados",
                "No se encontraron filas duplicadas.", parent=self)
            return
        res = messagebox.askyesno("Quitar duplicados",
            f"Se encontraron {removed} filas duplicadas (misma fecha + descripciÃ³n).\n"
            "Â¿Eliminarlas de la vista?", parent=self)
        if res:
            self._ec_filas_todas = unicos
            self._ec_refrescar_tree(unicos)


    # ---------------------------------------------------------------- #
    # PESTAÃA RECONCILIACIÃN â CSV â Excel plantilla
    # ---------------------------------------------------------------- #
    def _tab_reconciliacion(self, nb):
        outer = ttk.Frame(nb)
        nb.add(outer, text="  ð ReconciliaciÃ³n  ")
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(0, weight=0)   # controles: altura fija
        outer.rowconfigure(1, weight=1)   # tabla: ocupa el resto

        self._rec_plantilla  = tk.StringVar(value="")
        self._rec_csvs       = []
        self._rec_resultado  = None
        self._rec_wb         = None
        self._rec_wb_name    = ""
        self._rec_wb_dir     = ""

        # ââ Zona superior: controles compactos ââââââââââââââââââââââââââââ
        ctrl = tk.Frame(outer, bg=COLOR_FONDO)
        ctrl.grid(row=0, column=0, sticky="ew")
        ctrl.columnconfigure(0, weight=1)

        # Titulo
        hdr = tk.Frame(ctrl, bg=COLOR_FONDO)
        hdr.grid(row=0, column=0, sticky="ew", padx=16, pady=(8, 4))
        tk.Label(hdr, text="ReconciliaciÃ³n CSV â Excel",
                 bg=COLOR_FONDO, fg="#2E7D32",
                 font=("Segoe UI", 12, "bold")).pack(side="left")

        # Fila de controles: plantilla + CSVs + botÃ³n en una sola lÃ­nea
        row_ctrl = tk.Frame(ctrl, bg=COLOR_FONDO)
        row_ctrl.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 4))
        row_ctrl.columnconfigure(1, weight=1)
        row_ctrl.columnconfigure(3, weight=1)

        # Plantilla
        tk.Label(row_ctrl, text="Plantilla:", bg=COLOR_FONDO,
                 fg=COLOR_TEXTO, font=("Segoe UI", 9, "bold")).grid(
                 row=0, column=0, sticky="w", padx=(0, 6))
        self._rec_lbl_plantilla = tk.Label(row_ctrl, text="Sin seleccionar",
            bg=COLOR_BLANCO, fg="#999999", font=("Segoe UI", 9),
            relief="sunken", padx=6, pady=2)
        self._rec_lbl_plantilla.grid(row=0, column=1, sticky="ew", padx=(0, 6))
        ttk.Button(row_ctrl, text="ð Plantilla",
                   command=self._rec_elegir_plantilla).grid(
                   row=0, column=2, padx=(0, 14))

        # CSVs
        tk.Label(row_ctrl, text="CSV(s):", bg=COLOR_FONDO,
                 fg=COLOR_TEXTO, font=("Segoe UI", 9, "bold")).grid(
                 row=0, column=3, sticky="w", padx=(0, 6))
        self._rec_csv_lbl = tk.Label(row_ctrl,
            text="Sin seleccionar", bg=COLOR_BLANCO,
            fg="#999999", font=("Segoe UI", 9), relief="sunken",
            padx=6, pady=2, cursor="hand2")
        self._rec_csv_lbl.grid(row=0, column=4, sticky="ew", padx=(0, 6))
        row_ctrl.columnconfigure(4, weight=1)

        def _abrir_csv(e=None):
            rutas = filedialog.askopenfilenames(
                title="Seleccionar CSV(s)",
                filetypes=[("CSV","*.csv"),("Todos","*.*")])
            if rutas:
                self._rec_csvs = list(rutas)
                n = len(rutas)
                nombres = ", ".join(os.path.basename(r) for r in rutas[:3])
                if n > 3: nombres += f" (+{n-3} mas)"
                self._rec_csv_lbl.config(
                    text=f"â {n} archivo(s): {nombres}",
                    fg="#2E7D32", font=("Segoe UI", 8, "bold"))
                self._rec_csv_sub.config(text=f"{n} archivo(s) seleccionado(s)")

        # referencia compartida para compatibilidad
        self._rec_csv_icon = self._rec_csv_lbl
        self._rec_csv_sub  = tk.Label(ctrl, text="", bg=COLOR_FONDO,
                                       fg="#66BB6A", font=("Segoe UI", 8))
        self._rec_csv_sub.grid(row=2, column=0, sticky="w", padx=16)

        self._rec_csv_lbl.bind("<Button-1>", _abrir_csv)
        ttk.Button(row_ctrl, text="ð CSV(s)",
                   command=_abrir_csv).grid(row=0, column=5, padx=(0, 0))

        # BotÃ³n generar + barra progreso
        btn_row = tk.Frame(ctrl, bg=COLOR_FONDO)
        btn_row.grid(row=3, column=0, sticky="ew", padx=16, pady=(2, 4))
        btn_row.columnconfigure(0, weight=1)
        tk.Button(btn_row, text="â  Generar Excel Reconciliado",
            bg="#1B5E20", fg=COLOR_BLANCO,
            font=("Segoe UI", 10, "bold"), relief="flat",
            padx=16, pady=8, cursor="hand2",
            activebackground="#2E7D32",
            command=self._rec_convertir,
        ).grid(row=0, column=0, sticky="ew")

        _rec_pb_frame = tk.Frame(ctrl, bg=COLOR_FONDO)
        _rec_pb_frame.grid(row=4, column=0, sticky="ew", padx=16, pady=(0, 2))
        _rec_pb_frame.grid_remove()
        self._rec_pb_frame = _rec_pb_frame
        self._rec_pb = FunkyProgressBar(_rec_pb_frame, maximum=100, height=50)
        self._rec_pb.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self._rec_pb_lbl = ttk.Label(_rec_pb_frame, text="0 %", width=8,
            foreground="#2E7D32", font=("Segoe UI", 8, "bold"))
        self._rec_pb_lbl.pack(side="left")

        # Barra resultado
        tb = tk.Frame(ctrl, bg=COLOR_BLANCO,
                      highlightbackground="#C8E6C9", highlightthickness=1)
        tb.grid(row=5, column=0, sticky="ew", padx=16, pady=(0, 4))
        tk.Label(tb, text="Resultado:", bg=COLOR_BLANCO, fg=COLOR_TEXTO,
                 font=("Segoe UI", 8)).pack(side="left", padx=(10, 2), pady=3)
        self._rec_lbl_arch = tk.Label(tb, text="---", bg=COLOR_BLANCO,
            fg="#2E7D32", font=("Segoe UI", 8, "bold"))
        self._rec_lbl_arch.pack(side="left", padx=(0, 10), pady=3)
        self._rec_btn_guardar = ttk.Button(tb, text="Guardar Excel",
                   command=self._rec_guardar, state="disabled")
        self._rec_btn_guardar.pack(side="right", padx=(6, 2), pady=3)
        ttk.Button(tb, text="Abrir Excel",
                   command=self._rec_abrir).pack(side="right", padx=6, pady=3)

        # Separador
        ttk.Separator(outer, orient="horizontal").grid(
            row=0, column=0, sticky="ew", pady=0)

        # ââ Zona inferior: tabla de resultados (Ã¡rea de trabajo) ââââââââââ
        work = tk.Frame(outer, bg=COLOR_FONDO)
        work.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)
        work.columnconfigure(0, weight=1)
        work.rowconfigure(1, weight=1)

        lbl_area = tk.Frame(work, bg="#1B5E20")
        lbl_area.grid(row=0, column=0, columnspan=2, sticky="ew")
        tk.Label(lbl_area, text="  Ãrea de Trabajo â Reporte de ReconciliaciÃ³n",
                 bg="#1B5E20", fg=COLOR_BLANCO,
                 font=("Segoe UI", 9, "bold")).pack(side="left", pady=4)
        self._rec_lbl_resumen = tk.Label(lbl_area, text="",
            bg="#1B5E20", fg="#A5D6A7", font=("Segoe UI", 8))
        self._rec_lbl_resumen.pack(side="right", padx=10)

        cols = ("Fecha", "TOTAL IG", "TOTAL VTA", "CONCILIACION", "Estado")
        self._rec_tree = ttk.Treeview(work, columns=cols, show="headings")
        _col_w = {"Fecha": 110, "TOTAL IG": 200, "TOTAL VTA": 200, "CONCILIACION": 180, "Estado": 80}
        for c in cols:
            self._rec_tree.heading(c, text=c)
            self._rec_tree.column(c, width=_col_w.get(c, 150),
                minwidth=80,
                anchor="e" if c not in ("Fecha", "Estado") else "center",
                stretch=True)
        self._rec_tree.grid(row=1, column=0, sticky="nsew")

        vsb2 = ttk.Scrollbar(work, orient="vertical", command=self._rec_tree.yview)
        self._rec_tree.configure(yscrollcommand=vsb2.set)
        vsb2.grid(row=1, column=1, sticky="ns")

        hsb2 = ttk.Scrollbar(work, orient="horizontal", command=self._rec_tree.xview)
        self._rec_tree.configure(xscrollcommand=hsb2.set)
        hsb2.grid(row=2, column=0, sticky="ew")

        self._rec_resumen = lbl_area   # compatibilidad

    def _rec_elegir_plantilla(self):
        self.focus_force()
        self.update()
        ruta = filedialog.askopenfilename(
            parent=self,
            title="Seleccionar plantilla Excel",
            filetypes=[("Excel","*.xlsx *.xlsm"),("Todos","*.*")])
        if ruta:
            self._rec_plantilla.set(ruta)
            self._rec_lbl_plantilla.config(
                text=os.path.basename(ruta), fg="#2E7D32",
                font=("Segoe UI", 9, "bold"))

    def _rec_abrir(self):
        if self._rec_resultado and os.path.exists(self._rec_resultado):
            import subprocess
            try: subprocess.Popen(["start", "", self._rec_resultado], shell=True)
            except Exception: pass

    def _rec_guardar(self):
        wb = getattr(self, "_rec_wb", None)
        if wb is None:
            messagebox.showwarning("Sin datos", "Primero genera el reporte.")
            return
        self.focus_force()
        self.update()
        ruta = filedialog.asksaveasfilename(
            parent=self,
            title="Guardar reporte de reconciliacion",
            initialdir=getattr(self, "_rec_wb_dir", os.getcwd()),
            initialfile=getattr(self, "_rec_wb_name", "reconciliacion.xlsx"),
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
        )
        if not ruta:
            return
        try:
            wb.save(ruta)
            self._rec_resultado = ruta
            self._rec_lbl_arch.config(text=os.path.basename(ruta))
            cur = self._rec_lbl_resumen.cget("text")
            self._rec_lbl_resumen.config(
                text=cur.replace("pendiente de guardar", os.path.basename(ruta)))
            self._rec_btn_guardar.config(state="disabled")
            if hasattr(self, "_visor_carpeta"):
                self._visor_carpeta.set(os.path.dirname(ruta))
            if hasattr(self, "_visor_refrescar"):
                self._visor_refrescar()
            messagebox.showinfo("Guardado", "Reporte guardado en:\n" + ruta)
        except Exception as e:
            messagebox.showerror("Error", "No se pudo guardar:\n" + str(e))

    def _rec_convertir(self):
        plantilla = self._rec_plantilla.get().strip()
        # La plantilla es opcional; si se da, debe ser Excel
        if plantilla and os.path.exists(plantilla):
            ext = os.path.splitext(plantilla)[1].lower()
            if ext not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
                messagebox.showerror(
                    "Formato no soportado",
                    f"La plantilla debe ser un archivo Excel (.xlsx).\n"
                    f"El archivo seleccionado es '{os.path.basename(plantilla)}'.\n\n"
                    f"Puedes dejar la plantilla vacÃ­a para generar con estructura automÃ¡tica."
                )
                return
        else:
            plantilla = None   # sin plantilla â estructura automÃ¡tica
        if not self._rec_csvs:
            messagebox.showwarning("CSV", "Selecciona al menos un archivo CSV.")
            return
        self._rec_pb_frame.grid()
        self._pb_iniciar(self._rec_pb, self._rec_pb_lbl)
        for item in self._rec_tree.get_children(): self._rec_tree.delete(item)
        self._rec_lbl_resumen.config(text="")
        import threading
        threading.Thread(target=self._rec_hilo,
                         args=(plantilla, list(self._rec_csvs)), daemon=True).start()

    def _rec_hilo(self, plantilla_path, csv_paths):
        import csv as _csv, re as _re
        from openpyxl.utils import get_column_letter as _gcl
        try:
            import openpyxl
            from openpyxl.styles import PatternFill as _PF, Alignment as _AL
        except ImportError:
            self.after(0, self._log, "Falta openpyxl", True)
            self.after(0, self._pb_error, self._rec_pb, self._rec_pb_lbl)
            self.after(0, self._rec_pb_frame.grid_remove)
            return

        if plantilla_path:
            try:
                wb = openpyxl.load_workbook(plantilla_path)
                if "POLIZA" in wb.sheetnames:
                    ws = wb["POLIZA"]
                else:
                    ws = wb.active
                    self.after(0, self._log,
                        f"Hoja 'POLIZA' no encontrada, usando '{ws.title}'")
            except Exception as e:
                self.after(0, self._log, f"Error abriendo plantilla: {e}", True)
                self.after(0, self._pb_error, self._rec_pb, self._rec_pb_lbl)
                self.after(0, self._rec_pb_frame.grid_remove)
                return

            # ââ Leer hoja CUENTAS âââââââââââââââââââââââââââââââââââââââââ
            cuentas_ing_list    = []
            cuentas_vta_list    = []
            cuentas_ing_by_acct = {}
            cuentas_vta_by_acct = {}
            cuentas_ing_names   = {}
            cuentas_vta_names   = {}
            cuentas_vta_acct_set = set()
            # Buscar hoja CUENTAS â abrir copia data_only para leer valores reales
            # (maneja formulas ="...", texto forzado y cualquier otro formato)
            _cuentas_sheet = next((s for s in wb.sheetnames if s.strip().upper() == "CUENTAS"), None)
            if _cuentas_sheet:
                try:
                    _wb_do = openpyxl.load_workbook(plantilla_path, data_only=True)
                    _cuentas_sheet_do = next(
                        (s for s in _wb_do.sheetnames if s.strip().upper() == "CUENTAS"), None)
                    wc = _wb_do[_cuentas_sheet_do] if _cuentas_sheet_do else wb[_cuentas_sheet]
                except Exception:
                    wc = wb[_cuentas_sheet]

                def _limpiar_val(v):
                    """Devuelve string limpio del valor de celda"""
                    if v is None: return ""
                    s = str(v).strip()
                    # Formato fÃ³rmula texto sin resolver: ="101-01-0001" â 101-01-0001
                    import re as _re_acct
                    _m = _re_acct.match(r'^=?"?([^"=]+)"?$', s)
                    if _m: s = _m.group(1).strip()
                    # Quitar apÃ³strofo de prefijo Excel ('101-01-0001)
                    if s.startswith("'"): s = s[1:].strip()
                    return s

                def _es_cuenta_real(s):
                    """True si parece nÃºmero de cuenta (empieza con dÃ­gito)"""
                    return bool(s) and s[0].isdigit()

                # Filtrar por NOMBRE para permitir que dos productos compartan
                # el mismo nÃºmero de cuenta (ej. ACEITE y COMBUSTIBLES)
                _seen_ing = set(); _seen_vta = set()
                for r in range(1, (wc.max_row or 50) + 1):
                    a = wc.cell(r, 1).value; n = wc.cell(r, 2).value
                    if a is not None and n is not None:
                        an = _limpiar_val(a); nn = _limpiar_val(n)
                        if _es_cuenta_real(an) and nn and nn.upper() not in _seen_ing:
                            _seen_ing.add(nn.upper())
                            cuentas_ing_list.append((an, nn))
                            cuentas_ing_by_acct[an] = nn
                            cuentas_ing_names[nn.upper()] = an
                    va = wc.cell(r, 5).value; vn = wc.cell(r, 6).value
                    if va is not None and vn is not None:
                        van = _limpiar_val(va); vnn = _limpiar_val(vn)
                        if _es_cuenta_real(van) and vnn and vnn.upper() not in _seen_vta:
                            _seen_vta.add(vnn.upper())
                            cuentas_vta_list.append((van, vnn))
                            cuentas_vta_by_acct[van] = vnn
                            cuentas_vta_names[vnn.upper()] = van
                        if _es_cuenta_real(van):
                            cuentas_vta_acct_set.add(van)
                self.after(0, self._log,
                    f"CUENTAS: {len(cuentas_ing_list)} ING, {len(cuentas_vta_list)} VTA")

            # ââ Agregar productos CSV que no estÃ©n en CUENTAS (con cuenta vacÃ­a) ââ
            # Aplica cuando hay CUENTAS cargado pero el CSV tiene productos extra
            if cuentas_ing_by_acct or cuentas_vta_by_acct:
                _pre_skip_x = {"descripcion","islas","total islas","total estacion",
                               "total impuestos","total ingresos","impuestos",""}
                _csv_ing_extra = []; _csv_vta_extra = []
                _ing_names_up = {nn.upper() for _, nn in cuentas_ing_list}
                _vta_names_up = {nn.upper() for _, nn in cuentas_vta_list}
                for _cpx in sorted(csv_paths):
                    _sx = None
                    try:
                        with open(_cpx, "r", encoding="latin-1") as _fx:
                            for _rx in _csv.reader(_fx):
                                if not _rx: continue
                                _nx = _rx[0].strip(); _nlx = _nx.lower()
                                if _nlx == "ingresos": _sx = "ING"; continue
                                if _nlx in ("islas","impuestos"): _sx = "VTA"; continue
                                if not _nlx or _nlx in _pre_skip_x or _nlx.startswith("total"): continue
                                if len(_rx) > 3 and _rx[3].strip():
                                    try:
                                        float(_rx[3].strip().replace(",",""))
                                        if _sx == "ING" and _nx.upper() not in _ing_names_up:
                                            _csv_ing_extra.append(_nx)
                                            _ing_names_up.add(_nx.upper())
                                        elif _sx == "VTA" and _nx.upper() not in _vta_names_up:
                                            _csv_vta_extra.append(_nx)
                                            _vta_names_up.add(_nx.upper())
                                    except ValueError: pass
                    except: pass
                for _enm in _csv_ing_extra:
                    cuentas_ing_list.append(("", _enm))
                    cuentas_ing_by_acct[_enm] = _enm
                for _enm in _csv_vta_extra:
                    cuentas_vta_list.append(("", _enm))
                    cuentas_vta_by_acct[_enm] = _enm
                if _csv_ing_extra or _csv_vta_extra:
                    self.after(0, self._log,
                        f"Productos CSV extra: +{len(_csv_ing_extra)} ING, +{len(_csv_vta_extra)} VTA")

            # ââ DetecciÃ³n dinÃ¡mica de columnas desde fila 3 âââââââââââââââ
            col_to_name  = {}
            col_to_acct  = {}
            proc_col = 8
            scan_up_to = max((ws.max_column or 30) + 10, 40)
            for col in range(1, scan_up_to):
                a2 = ws.cell(row=2, column=col).value
                if a2:
                    col_to_acct[col] = str(a2).strip()
            for col in range(1, scan_up_to):
                h = ws.cell(row=3, column=col).value
                if h:
                    hn = str(h).strip()
                    col_to_name[col] = hn
                    if hn == "PROCESADO":
                        proc_col = col

            hdrs_str = "  ".join(f"{_gcl(c)}={n}" for c, n in sorted(col_to_name.items()) if c > proc_col - 2)
            self.after(0, self._log, f"Enc.fila3: {hdrs_str}")

            import re as _re_tot
            _total_pat = _re_tot.compile(r'^TOTAL\b', _re_tot.IGNORECASE)
            total_cols = sorted([col for col, hn in col_to_name.items()
                                 if _total_pat.match(hn) and col > proc_col])
            _t1 = total_cols[0] if total_cols else proc_col + 11
            _t2 = total_cols[1] if len(total_cols) >= 2 else _t1 + 10
            _simplificada = (_t1 - proc_col) < 5
            tot_ig_col  = _t1
            tot_vta_col = _t2
            conc_col = next((c for c, n in col_to_name.items()
                             if "CONCILIACION" in n and c > tot_vta_col),
                            tot_vta_col + 1)
            ing_start = proc_col + 1; ing_end = tot_ig_col - 1
            vta_start = tot_ig_col + 1; vta_end = tot_vta_col - 1
            ing_ltr_s = _gcl(ing_start); ing_ltr_e = _gcl(ing_end)
            vta_ltr_s = _gcl(vta_start); vta_ltr_e = _gcl(vta_end)
            tig_ltr = _gcl(tot_ig_col); tva_ltr = _gcl(tot_vta_col)

            # ââ Si simplificada sin CUENTAS: pre-escanear CSV para descubrir cuentas ââ
            if _simplificada and not (cuentas_ing_by_acct or cuentas_vta_by_acct):
                _pre_skip2 = {"descripcion","islas","total islas","total estacion",
                              "total impuestos","total ingresos","impuestos",""}
                _csv_ing = []; _csv_vta = []
                _csv_ing_s = set(); _csv_vta_s = set()
                for _cp2 in sorted(csv_paths):
                    _s2 = None
                    try:
                        with open(_cp2,"r",encoding="latin-1") as _f2:
                            for _r2 in _csv.reader(_f2):
                                if not _r2: continue
                                _n2 = _r2[0].strip(); _nl2 = _n2.lower()
                                if _nl2 == "ingresos": _s2 = "ING"; continue
                                if _nl2 in ("islas","impuestos"): _s2 = "VTA"; continue
                                if not _nl2 or _nl2 in _pre_skip2 or _nl2.startswith("total"): continue
                                if len(_r2) > 3 and _r2[3].strip():
                                    try:
                                        float(_r2[3].strip().replace(",",""))
                                        if _s2 == "ING" and _n2 not in _csv_ing_s:
                                            _csv_ing.append(_n2); _csv_ing_s.add(_n2)
                                        elif _s2 == "VTA" and _n2 not in _csv_vta_s:
                                            _csv_vta.append(_n2); _csv_vta_s.add(_n2)
                                    except ValueError: pass
                    except: pass
                # Usar los nombres del CSV como cuentas
                cuentas_ing_list = [("", nm) for nm in _csv_ing]
                cuentas_vta_list = [("", nm) for nm in _csv_vta]
                cuentas_ing_by_acct = {nm: nm for nm in _csv_ing}
                cuentas_vta_by_acct = {nm: nm for nm in _csv_vta}
                self.after(0, self._log,
                    f"CUENTAS desde CSV: {len(_csv_ing)} ING, {len(_csv_vta)} VTA")

            # ââ Expandir si simplificada pero hay CUENTAS âââââââââââââââââ
            if _simplificada and (cuentas_ing_by_acct or cuentas_vta_by_acct):
                n_ing = len(cuentas_ing_list); n_vta = len(cuentas_vta_list)
                if n_ing > 0:
                    ws.insert_cols(tot_ig_col, n_ing)
                    for i, (acct, nombre) in enumerate(cuentas_ing_list):
                        col = proc_col + 1 + i
                        ws.cell(2, col).value = acct; ws.cell(3, col).value = nombre
                        col_to_name[col] = nombre; col_to_acct[col] = acct
                    tot_ig_col += n_ing; tot_vta_col += n_ing; conc_col += n_ing
                if n_vta > 0:
                    ws.insert_cols(tot_vta_col, n_vta)
                    for i, (acct, nombre) in enumerate(cuentas_vta_list):
                        col = tot_ig_col + 1 + i
                        ws.cell(2, col).value = acct; ws.cell(3, col).value = nombre
                        col_to_name[col] = nombre; col_to_acct[col] = acct
                    tot_vta_col += n_vta; conc_col += n_vta
                ing_start = proc_col + 1; ing_end = tot_ig_col - 1
                vta_start = tot_ig_col + 1; vta_end = tot_vta_col - 1
                ing_ltr_s = _gcl(ing_start); ing_ltr_e = _gcl(ing_end)
                vta_ltr_s = _gcl(vta_start); vta_ltr_e = _gcl(vta_end)
                tig_ltr = _gcl(tot_ig_col); tva_ltr = _gcl(tot_vta_col)
                _simplificada = False
                self.after(0, self._log,
                    f"Expandida: {n_ing} ING + {n_vta} VTA  TOTAL_IG={tig_ltr}  TOTAL_VTA={tva_ltr}")

            modo = "SIMPLIFICADA" if _simplificada else "DETALLE"
            self.after(0, self._log,
                f"Modo={modo}  INGRESOS {ing_ltr_s}-{ing_ltr_e}  TOTAL_IG={tig_ltr}  "
                f"VENTAS {vta_ltr_s}-{vta_ltr_e}  TOTAL_VTA={tva_ltr}")

            # ââ Insertar columna BAL_ACCT ââââââââââââââââââââââââââââââââââ
            _BAL_ACCT = '101-01-0003-0001'
            _bal_col = next((c for c in range(ing_start, tot_ig_col)
                             if col_to_acct.get(c) == _BAL_ACCT), None)
            if _bal_col is None and not _simplificada:
                ws.insert_cols(tot_ig_col, 1)
                _bal_col = tot_ig_col
                tot_ig_col += 1; tot_vta_col += 1; conc_col += 1
                vta_start += 1; vta_end += 1
                for _c in sorted([c for c in col_to_name if c >= _bal_col], reverse=True):
                    col_to_name[_c + 1] = col_to_name.pop(_c)
                for _c in sorted([c for c in col_to_acct if c >= _bal_col], reverse=True):
                    col_to_acct[_c + 1] = col_to_acct.pop(_c)
                ws.cell(2, _bal_col).value = _BAL_ACCT
                ws.cell(3, _bal_col).value = "efectivo cuenta dif"
                col_to_name[_bal_col] = "efectivo cuenta dif"
                col_to_acct[_bal_col] = _BAL_ACCT
                ing_ltr_e = _gcl(_bal_col); tig_ltr = _gcl(tot_ig_col)
                vta_ltr_s = _gcl(vta_start); vta_ltr_e = _gcl(vta_end)
                tva_ltr = _gcl(tot_vta_col)
                self.after(0, self._log,
                    f"Col BAL insertada en {_gcl(_bal_col)}")

        else:
            # ââ AUTO MODE: estructura generada desde los CSV âââââââââââââââ
            # Pre-escanear todos los CSV para descubrir cuentas
            _pre_skip = {"descripcion", "islas", "total islas", "total estacion",
                         "total impuestos", "total ingresos", "impuestos", ""}
            _auto_ing = []; _auto_vta = []
            _auto_ing_set = set(); _auto_vta_set = set()
            for _cp in sorted(csv_paths):
                _sec = None
                try:
                    with open(_cp, "r", encoding="latin-1") as _f:
                        for _row in _csv.reader(_f):
                            if not _row: continue
                            _n = _row[0].strip(); _nl = _n.lower()
                            if _nl == "ingresos": _sec = "ING"; continue
                            if _nl in ("islas","impuestos"): _sec = "VTA"; continue
                            if not _nl or _nl in _pre_skip or _nl.startswith("total"): continue
                            if len(_row) > 3 and _row[3].strip():
                                try:
                                    float(_row[3].strip().replace(",",""))
                                    if _sec == "ING" and _n not in _auto_ing_set:
                                        _auto_ing.append(_n); _auto_ing_set.add(_n)
                                    elif _sec == "VTA" and _n not in _auto_vta_set:
                                        _auto_vta.append(_n); _auto_vta_set.add(_n)
                                except ValueError: pass
                except: pass

            n_ing = len(_auto_ing); n_vta = len(_auto_vta)
            self.after(0, self._log, f"Auto: {n_ing} cuentas ING, {n_vta} cuentas VTA")

            # Crear libro con columnas por cuenta
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "POLIZA"
            proc_col    = 8
            tot_ig_col  = proc_col + 1 + n_ing
            tot_vta_col = tot_ig_col + 1 + n_vta
            conc_col    = tot_vta_col + 1
            col_to_name = {}; col_to_acct = {}

            for _c, _h in enumerate(["TIPO DE POLIZA","Fecha","REFERENCIA","CONCEPTO",
                                      "ERROR","UIDD","NUM POLIZA","PROCESADO"], start=1):
                ws.cell(3, _c).value = _h; col_to_name[_c] = _h

            ing_start = proc_col + 1
            for _i, _nm in enumerate(_auto_ing):
                _col = ing_start + _i
                ws.cell(3, _col).value = _nm; col_to_name[_col] = _nm
            ing_end = tot_ig_col - 1

            ws.cell(3, tot_ig_col).value = "TOTAL INGRESOS"
            col_to_name[tot_ig_col] = "TOTAL INGRESOS"

            vta_start = tot_ig_col + 1
            for _i, _nm in enumerate(_auto_vta):
                _col = vta_start + _i
                ws.cell(3, _col).value = _nm; col_to_name[_col] = _nm
            vta_end = tot_vta_col - 1

            ws.cell(3, tot_vta_col).value = "TOTAL VENTAS"
            col_to_name[tot_vta_col] = "TOTAL VENTAS"
            ws.cell(3, conc_col).value = "CONCILIACION"
            col_to_name[conc_col] = "CONCILIACION"

            ing_ltr_s = _gcl(ing_start); ing_ltr_e = _gcl(ing_end)
            vta_ltr_s = _gcl(vta_start); vta_ltr_e = _gcl(vta_end)
            tig_ltr = _gcl(tot_ig_col); tva_ltr = _gcl(tot_vta_col)

            cuentas_ing_list = []; cuentas_vta_list = []
            cuentas_ing_by_acct = {}; cuentas_vta_by_acct = {}
            cuentas_ing_names = {}; cuentas_vta_names = {}
            cuentas_vta_acct_set = set()
            _simplificada = False; _bal_col = None

        # ââ Renumerar fila 1 secuencialmente tras todas las inserciones ââââ
        # Asegura que fila 1 tenga 0,1,2,3... sin huecos en columnas nuevas
        _last_col = max(conc_col, tot_vta_col + 1)
        for _c in range(1, _last_col + 1):
            ws.cell(1, _c).value = _c - 1   # 0-indexed

        # ââ Asegurar header CONCILIACION y colorear fila 3 ââââââââââââââââ
        from openpyxl.styles import Font as _Fnt
        ws.cell(3, conc_col).value = "CONCILIACION"
        def _hdr(col, bg):
            c = ws.cell(3, col)
            c.fill = _PF("solid", fgColor=bg)
            c.font = _Fnt(bold=True, color="FFFFFF", size=9)
            c.alignment = _AL(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, proc_col + 1):         _hdr(col, "595959")
        for col in range(ing_start, tot_ig_col):   _hdr(col, "2E75B6")
        _hdr(tot_ig_col,  "1F4E79")
        for col in range(vta_start, tot_vta_col):  _hdr(col, "C55A11")
        _hdr(tot_vta_col, "833C00")
        _hdr(conc_col,    "7030A0")

        # ââ Mapa fecha â fila existente (evitar duplicados) ââââââââââââââ
        fecha_fila = {}
        for r in range(4, (ws.max_row or 4) + 1):
            v = ws.cell(r, 2).value
            if v:
                from datetime import datetime as _dtx
                if isinstance(v, _dtx):
                    fecha_fila[v.strftime("%d/%m/%Y")] = r
                elif isinstance(v, str) and "/" in v:
                    fecha_fila[v.strip()] = r

        # Primera fila vacÃ­a despuÃ©s de las existentes
        next_empty = (max(fecha_fila.values()) + 1) if fecha_fila else 4

        SKIP = {"descripcion", "estacion gas122 s.a. de c.v.", "islas",
                "total islas", "total estacion", "total impuestos",
                "total ingresos", "impuestos", ""}

        def _fill_colors(r):
            def _s(c, bg, h="right"):
                c.fill = _PF("solid", fgColor=bg)
                c.alignment = _AL(horizontal=h, vertical="center")
            for col in range(1, proc_col + 1):
                _s(ws.cell(r, col), "F2F2F2", "left")
            for col in range(ing_start, tot_ig_col):
                _s(ws.cell(r, col), "DEEAF1")
            _s(ws.cell(r, tot_ig_col), "BDD7EE")
            for col in range(vta_start, tot_vta_col):
                _s(ws.cell(r, col), "FDF2E9")
            _s(ws.cell(r, tot_vta_col), "FAD7AC")
            _s(ws.cell(r, conc_col), "EAD1DC")

        # Encabezados del Ã¡rbol: cuentas ING + TOTAL IG + cuentas VTA + TOTAL VTA + CONCILIACION
        _tree_ing_hdrs = [col_to_name.get(c, f"ING{c}") for c in range(ing_start, tot_ig_col)
                          if col_to_acct.get(c) != _BAL_ACCT]
        _tree_vta_hdrs = [col_to_name.get(c, f"VTA{c}") for c in range(vta_start, tot_vta_col)
                          if col_to_acct.get(c) != _BAL_ACCT and col_to_name.get(c)]
        _tree_vta_src  = [c for c in range(vta_start, tot_vta_col)
                          if col_to_acct.get(c) != _BAL_ACCT and col_to_name.get(c)]
        _tree_all_cols = (["Fecha"] + _tree_ing_hdrs + ["TOTAL IG"] +
                          _tree_vta_hdrs + ["TOTAL VTA", "CONCILIACION", "Estado"])

        # ââ Sistema de aprendizaje: alias CSV â CUENTAS âââââââââââââââââ
        import json as _json, unicodedata as _ud
        _aliases_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "reconciliacion_aliases.json")
        try:
            with open(_aliases_path, "r", encoding="utf-8") as _af:
                _alias_data = _json.load(_af)
        except Exception:
            _alias_data = {}
        # _alias_data formato: {"ing": {"CSV_NAME_UP": "CUENTAS_NAME"}, "vta": {...}}
        _aliases_ing = _alias_data.get("ing", {})
        _aliases_vta = _alias_data.get("vta", {})
        _new_aliases = False  # marcar si se aprendiÃ³ algo nuevo

        def _norm(s):
            """Normaliza: sin acentos, sin espacios extra, mayÃºsculas"""
            return _ud.normalize("NFD", s.upper().strip()).encode("ascii", "ignore").decode()

        def _buscar_alias(vals_dict, header, aliases_dict, tipo):
            """Busca valor en vals_dict por header de CUENTAS.
            Orden: exacto â case-insensitive (auto-aprende) â normalizado (auto-aprende) â alias guardado"""
            nonlocal _new_aliases
            hU = header.upper()
            hN = _norm(header)
            # 1. Exacto
            if header in vals_dict: return vals_dict[header]
            # 2. Case-insensitive â auto-aprende
            for k, kv in vals_dict.items():
                if k.upper() == hU:
                    if k != header:  # diferente capitalizaciÃ³n â guardar alias
                        if k.upper() not in aliases_dict:
                            aliases_dict[k.upper()] = header
                            _new_aliases = True
                    return kv
            # 3. Normalizado (sin acentos) â auto-aprende
            for k, kv in vals_dict.items():
                if _norm(k) == hN:
                    if k.upper() not in aliases_dict:
                        aliases_dict[k.upper()] = header
                        _new_aliases = True
                    return kv
            # 4. Alias guardados de sesiones anteriores
            if hU in aliases_dict.values():
                target = [csv_k for csv_k, ctas_v in aliases_dict.items() if ctas_v.upper() == hU]
                for tk in target:
                    for k, kv in vals_dict.items():
                        if k.upper() == tk:
                            return kv
            return None

        resultados = []
        total_csvs = len(csv_paths)
        data_row   = next_empty

        for idx, csv_path in enumerate(sorted(csv_paths)):
            pct = int((idx / total_csvs) * 90)
            self._rec_pb["value"] = pct
            self.after(0, lambda p=pct: self._rec_pb_lbl.config(text=f"{p} %"))

            csv_vals = {}; ing_vals = {}; vta_vals = {}
            fecha_str = None; _seccion = None
            try:
                with open(csv_path, "r", encoding="latin-1") as f:
                    reader = _csv.reader(f)
                    for row in reader:
                        if not row: continue
                        name = row[0].strip()
                        nl = name.lower()
                        if not fecha_str:
                            m = _re.search(r"(\d{2}/\d{2}/\d{4})", row[0])
                            if m: fecha_str = m.group(1); continue
                        # Detectar secciÃ³n ANTES del SKIP
                        if nl == "ingresos":
                            _seccion = "ING"; continue
                        if nl in ("islas", "impuestos"):
                            _seccion = "VTA"; continue
                        # Saltar vacÃ­os, encabezados SKIP y cualquier fila "total â¦"
                        if not nl or nl in SKIP or nl.startswith("total"):
                            continue
                        if len(row) > 3 and row[3].strip():
                            try:
                                val = float(row[3].strip().replace(",", ""))
                                csv_vals[name] = val
                                if _seccion == "ING": ing_vals[name] = val
                                elif _seccion == "VTA": vta_vals[name] = val
                            except ValueError: pass
            except Exception as e:
                resultados.append((os.path.basename(csv_path), "---", "---", "---", f"Error: {e}"))
                continue

            if not fecha_str:
                resultados.append((os.path.basename(csv_path), "---", "---", "---", "Sin fecha"))
                continue

            from datetime import datetime as _dt
            d_p, mo_p, y_p = fecha_str.split("/")
            fecha_dt  = _dt(int(y_p), int(mo_p), int(d_p))
            ref_text  = f"VENTA DEL {fecha_str}"

            # Usar fila existente si la fecha ya estÃ¡, si no la siguiente vacÃ­a
            wr = fecha_fila.get(fecha_str, data_row)
            if wr == data_row:
                data_row += 1   # avanzar solo si es fila nueva

            # Limpiar fila destino
            for col in range(1, conc_col + 2):
                ws.cell(wr, col).value = None

            # Columnas fijas
            ws.cell(wr, 1).value = "CLI"
            ws.cell(wr, 2).value = fecha_dt
            ws.cell(wr, 2).number_format = "DD/MM/YYYY"
            ws.cell(wr, 3).value = ref_text
            ws.cell(wr, 4).value = f"=C{wr}"

            if _simplificada:
                # Plantilla simplificada: secciones detectadas del CSV
                tot_ig  = round(sum(ing_vals.values()), 2) if ing_vals else round(sum(csv_vals.values()), 2)
                tot_vta = round(sum(vta_vals.values()), 2) if vta_vals else tot_ig
                ws.cell(wr, tot_ig_col).value = round(tot_ig, 2)
                ws.cell(wr, tot_ig_col).number_format = "General"
                ws.cell(wr, tot_vta_col).value = round(tot_vta, 2)
                ws.cell(wr, tot_vta_col).number_format = "General"
                ws.cell(wr, conc_col).value = round(tot_ig - tot_vta, 2)
                ws.cell(wr, conc_col).number_format = "General"
            else:
                # Plantilla con detalle: mapear individualmente + fÃ³rmulas
                # Usa sistema de aprendizaje: exacto â case-insensitive â normalizado â alias
                def _get_ing(col):
                    hdr = col_to_name.get(col, "")
                    if not hdr: return None
                    return _buscar_alias(ing_vals, hdr, _aliases_ing, "ing")

                def _get_vta(col):
                    hdr = col_to_name.get(col, "")
                    acct = col_to_acct.get(col, "")
                    if not hdr: return None
                    # Solo escribir si la cuenta estÃ¡ en CUENTAS VTA
                    if cuentas_vta_acct_set and acct and acct not in cuentas_vta_acct_set:
                        return None
                    return _buscar_alias(vta_vals, hdr, _aliases_vta, "vta")

                for col in range(ing_start, tot_ig_col):
                    if col == _bal_col:
                        continue  # se escribe con signo contrario en bloque de balanceo
                    val = _get_ing(col)
                    ws.cell(wr, col).value = val if val is not None else 0
                    ws.cell(wr, col).number_format = "General"
                ws.cell(wr, tot_ig_col).value = (
                    f"=SUM({ing_ltr_s}{wr}:{ing_ltr_e}{wr})")
                ws.cell(wr, tot_ig_col).number_format = "General"

                for col in range(vta_start, tot_vta_col):
                    val = _get_vta(col)
                    ws.cell(wr, col).value = val if val is not None else 0
                    ws.cell(wr, col).number_format = "General"
                ws.cell(wr, tot_vta_col).value = (
                    f"=SUM({vta_ltr_s}{wr}:{vta_ltr_e}{wr})")
                ws.cell(wr, tot_vta_col).number_format = "General"
                ws.cell(wr, conc_col).value = f"={tig_ltr}{wr}-{tva_ltr}{wr}"
                ws.cell(wr, conc_col).number_format = "General"

                # Totales reales para reporte (sumas de secciones CSV)
                tot_ig  = round(sum(ing_vals.values()), 2)
                tot_vta = round(sum(vta_vals.values()), 2)

                # ââ Escribir PROCESADO en secciÃ³n AZUL con signo contrario ââ
                if _bal_col is not None:
                    _dif = round(tot_ig - tot_vta, 2)
                    _adj = round(-_dif, 2)   # signo contrario â cuadra CONCILIACION a 0
                    ws.cell(wr, _bal_col).value = _adj if abs(_adj) >= 0.01 else 0
                    ws.cell(wr, _bal_col).number_format = "General"
                    tot_ig = round(tot_vta, 2)  # tot_ig ajustado para el Ã¡rbol



            _fill_colors(wr)
            conc_v  = tot_ig - tot_vta
            estado  = "OK" if abs(conc_v) < 1 else f"Dif {conc_v:+,.2f}"

            # Valores por cuenta para el Ã¡rbol
            _ing_row = []
            for _th in _tree_ing_hdrs:
                _v = ing_vals.get(_th) or next((kv for k,kv in ing_vals.items() if k.upper()==_th.upper()), None)
                _ing_row.append(f"{(_v or 0):,.2f}")
            _vta_row = []
            for _ti, _th in zip(_tree_vta_src, _tree_vta_hdrs):
                _a = col_to_acct.get(_ti,"")
                if cuentas_vta_acct_set and _a and _a not in cuentas_vta_acct_set:
                    _vta_row.append("0.00")
                else:
                    _v = vta_vals.get(_th) or next((kv for k,kv in vta_vals.items() if k.upper()==_th.upper()), None)
                    _vta_row.append(f"{(_v or 0):,.2f}")
            resultados.append(
                (fecha_str, *_ing_row, f"{tot_ig:,.2f}", *_vta_row, f"{tot_vta:,.2f}", f"{conc_v:,.2f}", estado))

        from datetime import datetime as _dt2
        ts = _dt2.now().strftime("%Y%m%d_%H%M%S")
        if plantilla_path:
            _base = os.path.splitext(os.path.basename(plantilla_path))[0]
            _dir  = os.path.dirname(plantilla_path)
        else:
            _base = "reconciliacion"
            _dir  = os.path.dirname(csv_paths[0]) if csv_paths else os.getcwd()
        self._rec_wb      = wb
        self._rec_wb_name = _base + "_reconciliado_" + ts + ".xlsx"
        self._rec_wb_dir  = _dir
        # Guardar en temp para que "Abrir Excel" funcione sin necesidad de guardar primero
        import tempfile as _tmp
        _tmp_path = os.path.join(_tmp.gettempdir(), self._rec_wb_name)
        try:
            wb.save(_tmp_path)
            self._rec_resultado = _tmp_path
        except Exception:
            self._rec_resultado = None

        # Guardar aliases aprendidos en esta sesiÃ³n
        if _new_aliases:
            try:
                _alias_data["ing"] = _aliases_ing
                _alias_data["vta"] = _aliases_vta
                with open(_aliases_path, "w", encoding="utf-8") as _af:
                    _json.dump(_alias_data, _af, ensure_ascii=False, indent=2)
                _n_new = sum(1 for k in _aliases_ing) + sum(1 for k in _aliases_vta)
                self.after(0, self._log, f"â Sistema aprendiÃ³ {_n_new} equivalencia(s) de nombres")
            except Exception as _ae:
                self.after(0, self._log, f"Alias: no se pudo guardar ({_ae})")

        self._rec_pb["value"] = 100
        self.after(0, lambda: self._rec_pb_lbl.config(text="100 %"))

        def _ui_done():
            self._pb_detener(self._rec_pb, self._rec_pb_lbl)
            self.after(1200, self._rec_pb_frame.grid_remove)
            self._rec_lbl_arch.config(text="Listo  â  clic en Guardar Excel")
            self._rec_btn_guardar.config(state="normal")
            # Reconfigurar columnas del arbol con todas las cuentas
            self._rec_tree.config(columns=_tree_all_cols)
            for _tc in _tree_all_cols:
                self._rec_tree.heading(_tc, text=_tc)
                if _tc == "Fecha":
                    _tw, _ta, _ts = 95, "center", True
                elif _tc in ("TOTAL IG","TOTAL VTA","CONCILIACION"):
                    _tw, _ta, _ts = 110, "e", True
                elif _tc == "Estado":
                    _tw, _ta, _ts = 70, "center", True
                else:
                    _tw, _ta, _ts = 85, "e", True
                self._rec_tree.column(_tc, width=_tw, minwidth=55, anchor=_ta, stretch=_ts)
            for item in self._rec_tree.get_children(): self._rec_tree.delete(item)
            for r in resultados:
                tag = "ok" if r[-1] == "OK" else "warn"
                self._rec_tree.insert("", "end", values=r, tags=(tag,))
            self._rec_tree.tag_configure("ok",   background="#E8F5E9")
            self._rec_tree.tag_configure("warn",  background="#FFF9C4")
            dias = len(resultados)
            self._rec_lbl_resumen.config(
                text="Dias procesados: " + str(dias) + "  |  pendiente de guardar")
            # Enviar al Visor de Resultados automÃ¡ticamente
            _tmp_r = getattr(self, "_rec_resultado", None)
            if _tmp_r and os.path.exists(_tmp_r):
                self._visor_abrir_directo(_tmp_r)
        self.after(0, _ui_done)


    def _visor_abrir_directo(self, ruta):
        """Carga un archivo directamente en el Visor sin necesidad de que estÃ© en la carpeta."""
        if not hasattr(self, "_visor_lb"):
            return
        # Abrir la pestaÃ±a Visor si estÃ¡ disponible
        if "visor" in self._tab_frames:
            try:
                self.nb.select(self._tab_frames["visor"])
            except Exception:
                pass
        # Insertar al inicio de la lista si no estÃ¡ ya
        if ruta not in self._visor_archivos:
            self._visor_archivos.insert(0, ruta)
            self._visor_lb.insert(0, os.path.basename(ruta))
        idx = self._visor_archivos.index(ruta)
        self._visor_lb.selection_clear(0, "end")
        self._visor_lb.selection_set(idx)
        self._visor_lb.see(idx)
        self._visor_cargar_archivo()

    # ---------------------------------------------------------------- #
    # PESTAÃA VISOR DE RESULTADOS
    # ---------------------------------------------------------------- #
    def _tab_visor(self, nb):
        tab = ttk.Frame(nb)
        nb.add(tab, text="  Visor de Resultados  ")
        self._PALETA = [
            ("#FFF9C4","#7B5800"), ("#B7D9EF","#13283B"),
            ("#E3F2FD","#0D47A1"), ("#E8F5E9","#1B5E20"),
            ("#FFEB3B","#5F4400"), ("#FF9800","#FFFFFF"),
            ("#F44336","#FFFFFF"), ("#9C27B0","#FFFFFF"),
        ]
        # Slots de colores personalizados (None = vacÃ­o)
        self._PALETA_CUSTOM = [None, None, None, None, None, None]
        self._PALETA_CUSTOM_IDX = 0   # prÃ³ximo slot a rellenar
        self._visor_cell_colors  = {}   # {(iid, col_id): (bg, fg)}
        self._visor_color_widgets = {}  # {(iid, col_id): tk.Label}
        self._visor_cell_fonts   = {}   # {(iid, col_id): {'bold','italic','underline','textcolor'}}
        self._visor_cell_edits   = {}
        self._visor_ruta_actual  = None
        self._visor_archivos     = []
        self._visor_filas_ocultas = []  # para filtro

        top = tk.Frame(tab, bg=COLOR_FONDO)
        top.pack(fill="x", padx=10, pady=(8, 2))
        ttk.Label(top, text="Carpeta:").pack(side="left")
        self._visor_carpeta = tk.StringVar(value="")
        ttk.Entry(top, textvariable=self._visor_carpeta, width=42).pack(side="left", padx=4)
        ttk.Button(top, text="Elegir...", command=self._visor_elegir_carpeta).pack(side="left")
        ttk.Button(top, text="Actualizar", command=self._visor_refrescar).pack(side="left", padx=(8, 0))
        ttk.Button(top, text="Guardar cambios", command=self._visor_guardar_cambios).pack(side="right")
        tk.Button(top, text="ð Borrar reporte", command=self._visor_borrar_reporte,
                  bg=COLOR_FUCSIA_OSCURO, fg=COLOR_BLANCO, relief="flat", cursor="hand2",
                  font=("Segoe UI", 9), padx=10, pady=3,
                  activebackground="#5a0a28", activeforeground=COLOR_BLANCO).pack(side="right", padx=(0, 6))

        # ââ Toolbar Ãºnica âââââââââââââââââââââââââââââââââââââââââââââââââ
        tb = tk.Frame(tab, bg=COLOR_TARJETA, pady=3)
        tb.pack(fill="x", padx=10, pady=(0, 4))

        def _sep(p=tb):
            tk.Label(p, text="|", bg=COLOR_TARJETA, fg=COLOR_FUCSIA_SUAVE).pack(side="left", padx=4)

        def tbtn(text, cmd, bold=False, padx=4):
            f = ("Segoe UI", 9, "bold") if bold else ("Segoe UI", 9)
            b = tk.Button(tb, text=text, command=cmd, relief="flat", cursor="hand2",
                          bg=COLOR_TARJETA, fg=COLOR_TEXTO, font=f,
                          activebackground=COLOR_FUCSIA_SUAVE, padx=padx, pady=1)
            b.pack(side="left", padx=1)
            return b

        # Colores fijos
        for bg, fg in self._PALETA:
            b = tk.Button(tb, bg=bg, relief="raised", width=2, height=1,
                          cursor="hand2", bd=1, activebackground=bg,
                          command=lambda b=bg, f=fg: self._visor_aplicar_color(b, f))
            b.pack(side="left", padx=1, pady=2)

        # Slots personalizados
        self._visor_custom_btns = []
        for i in range(len(self._PALETA_CUSTOM)):
            slot_color = self._PALETA_CUSTOM[i]
            bg_c = slot_color[0] if slot_color else COLOR_FONDO
            b = tk.Button(tb, bg=bg_c, relief="groove", width=2, height=1,
                          cursor="hand2", bd=1, activebackground=bg_c,
                          command=lambda idx=i: self._visor_aplicar_custom(idx))
            b.pack(side="left", padx=1, pady=2)
            self._visor_custom_btns.append(b)

        _sep()
        tk.Button(tb, text="ð¨", bg=COLOR_TARJETA, fg=COLOR_FUCSIA_OSCURO,
                  relief="flat", cursor="hand2", font=("Segoe UI", 10),
                  command=self._visor_color_picker, padx=4, pady=1).pack(side="left", padx=1)
        tbtn("â Color", self._visor_quitar_color)
        _sep()
        # Fuente
        tbtn("N", self._visor_toggle_bold,      bold=True, padx=5)
        tbtn("K", self._visor_toggle_italic,     padx=5)
        tbtn("S", self._visor_toggle_underline,  padx=5)
        tk.Button(tb, text="Aâ¾", bg=COLOR_TARJETA, fg=COLOR_TEXTO,
                  relief="flat", cursor="hand2", font=("Segoe UI", 9, "bold"),
                  command=self._visor_color_texto, padx=4, pady=1).pack(side="left", padx=1)
        _sep()
        # Ordenar
        tbtn("â AZ", lambda: self._visor_ordenar(asc=True))
        tbtn("â ZA", lambda: self._visor_ordenar(asc=False))
        _sep()
        # Filtro
        tbtn("â² Filtrar", self._visor_filtrar)
        tbtn("â Filtro",  self._visor_quitar_filtro)
        _sep()
        # EdiciÃ³n
        tbtn("â Editar", self._visor_editar_celda_manual)
        tbtn("+ Fila",   self._visor_agregar_fila)
        tbtn("â Fila",   self._visor_borrar_fila)

        paned = tk.PanedWindow(tab, orient="horizontal", sashwidth=5,
                               bg=COLOR_FUCSIA_SUAVE, relief="flat")
        paned.pack(fill="both", expand=True, padx=10, pady=(0, 6))

        izq = tk.Frame(paned, bg=COLOR_FONDO)
        paned.add(izq, minsize=180)
        ttk.Label(izq, text="Archivos generados:", style="Seccion.TLabel").pack(anchor="w", pady=(4, 2))
        lbf, self._visor_lb = _make_listbox(izq, COLOR_FONDO)
        lbf.pack(fill="both", expand=True)
        self._visor_lb.bind("<<ListboxSelect>>",
            lambda e: self._visor_cargar_archivo() if len(self._visor_lb.curselection()) == 1 else None)
        hoja_row = tk.Frame(izq, bg=COLOR_FONDO)
        hoja_row.pack(fill="x", pady=(4, 0))
        ttk.Label(hoja_row, text="Hoja:").pack(side="left")
        self._visor_hoja = tk.StringVar()
        self._visor_combo_hoja = ttk.Combobox(hoja_row, textvariable=self._visor_hoja,
                                               state="readonly", width=16)
        self._visor_combo_hoja.pack(side="left", padx=4)
        self._visor_combo_hoja.bind("<<ComboboxSelected>>", lambda e: self._visor_mostrar_hoja())

        der = tk.Frame(paned, bg=COLOR_FONDO)
        paned.add(der, minsize=400)
        tree_wrap = tk.Frame(der, bg=COLOR_FONDO)
        tree_wrap.pack(fill="both", expand=True)
        self._visor_tree = ttk.Treeview(tree_wrap, show="headings", selectmode="extended")
        vsb = ttk.Scrollbar(tree_wrap, orient="vertical",   command=self._visor_tree.yview)
        hsb = ttk.Scrollbar(tree_wrap, orient="horizontal", command=self._visor_tree.xview)
        self._visor_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._visor_tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_wrap.rowconfigure(0, weight=1)
        tree_wrap.columnconfigure(0, weight=1)
        self._visor_iid_sel  = None
        self._visor_col_sel  = None
        self._visor_sel_border = None
        self._visor_undo_stack = []   # lista de (iid, col_id, valor_anterior)
        self._visor_tree.bind("<Button-1>",       self._visor_click_celda)
        self._visor_tree.bind("<Double-1>",       self._visor_editar_celda)
        self._visor_tree.bind("<Button-3>",       self._visor_menu_contextual)
        self._visor_tree.bind("<Control-c>",      lambda e: self._visor_copiar())
        self._visor_tree.bind("<Control-C>",      lambda e: self._visor_copiar())
        self._visor_tree.bind("<Control-x>",      lambda e: self._visor_cortar())
        self._visor_tree.bind("<Control-X>",      lambda e: self._visor_cortar())
        self._visor_tree.bind("<Control-v>",      lambda e: self._visor_pegar())
        self._visor_tree.bind("<Control-V>",      lambda e: self._visor_pegar())
        self._visor_tree.bind("<Control-z>",      lambda e: self._visor_deshacer())
        self._visor_tree.bind("<Control-Z>",      lambda e: self._visor_deshacer())
        self._visor_tree.bind("<Delete>",         lambda e: self._visor_borrar_celda())
        self._visor_tree.bind("<BackSpace>",      lambda e: self._visor_borrar_celda())
        self._visor_tree.bind("<Up>",             lambda e: self._visor_navegar("up"))
        self._visor_tree.bind("<Down>",           lambda e: self._visor_navegar("down"))
        self._visor_tree.bind("<Left>",           lambda e: self._visor_navegar("left"))
        self._visor_tree.bind("<Right>",          lambda e: self._visor_navegar("right"))
        self._visor_tree.bind("<Return>",         self._visor_editar_celda)
        self._visor_tree.bind("<Tab>",            lambda e: self._visor_navegar("right"))
        self._visor_tree.focus_set
        # Refrescar labels de color al hacer scroll
        self._visor_tree.bind("<<TreeviewSelect>>", lambda e: None)
        vsb.configure(command=lambda *a: (
            self._visor_tree.yview(*a), self.after(10, self._visor_refresh_cell_labels)))
        hsb.configure(command=lambda *a: (
            self._visor_tree.xview(*a), self.after(10, self._visor_refresh_cell_labels)))
        self._visor_tree.bind("<MouseWheel>",
            lambda e: self.after(30, self._visor_refresh_cell_labels))
        self._visor_tree.bind("<Configure>",
            lambda e: self.after(10, self._visor_refresh_cell_labels))

        info_row = tk.Frame(der, bg=COLOR_FONDO)
        info_row.pack(fill="x", pady=(2, 0))
        self._visor_label_info = ttk.Label(info_row, text="", style="Seccion.TLabel")
        self._visor_label_info.pack(side="left")
        self._visor_label_celda = tk.Label(info_row, text="",
            bg=COLOR_TARJETA, fg=COLOR_FUCSIA_OSCURO,
            font=("Segoe UI", 8), relief="flat", padx=8, pady=2)
        self._visor_label_celda.pack(side="right", padx=4)
        self.after(200, self._visor_init)

    def _visor_init(self):
        try:
            carpeta = self._resultados_dir()
            if os.path.isdir(carpeta):
                self._visor_carpeta.set(carpeta)
                self._visor_refrescar()
        except Exception:
            pass

    def _visor_elegir_carpeta(self):
        from tkinter import filedialog
        self.lift()
        self.focus_force()
        self.update()
        d = filedialog.askdirectory(title="Seleccionar carpeta de resultados", parent=self)
        if d:
            self._visor_carpeta.set(d)
            self._visor_refrescar()

    def _visor_refrescar(self):
        if not hasattr(self, "_visor_lb"):
            return
        import glob as _glob
        self._visor_lb.delete(0, "end")
        carpeta = self._visor_carpeta.get().strip()
        if not carpeta or not os.path.isdir(carpeta):
            return
        archivos = sorted(_glob.glob(os.path.join(carpeta, "*.xlsx")),
                          key=os.path.getmtime, reverse=True)
        self._visor_archivos = archivos
        for a in archivos:
            self._visor_lb.insert("end", os.path.basename(a))

    def _visor_cargar_archivo(self):
        sel = self._visor_lb.curselection()
        if not sel:
            return
        ruta = self._visor_archivos[sel[0]]
        try:
            import openpyxl
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            hojas = wb.sheetnames
            wb.close()
            self._visor_ruta_actual = ruta
            for lbl in self._visor_color_widgets.values():
                try: lbl.destroy()
                except Exception: pass
            self._visor_cell_colors   = {}
            self._visor_color_widgets = {}
            self._visor_cell_fonts    = {}
            self._visor_filas_ocultas = []
            self._visor_cell_edits    = {}
            self._visor_undo_stack    = []
            self._visor_combo_hoja["values"] = hojas
            if hojas:
                self._visor_hoja.set(hojas[0])
                self._visor_mostrar_hoja()
        except Exception as e:
            self._visor_label_info.config(text=f"Error al abrir: {e}")

    def _visor_mostrar_hoja(self):
        hoja = self._visor_hoja.get()
        ruta = self._visor_ruta_actual
        if not ruta or not hoja:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            ws = wb[hoja]
            filas = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as e:
            self._visor_label_info.config(text=f"Error: {e}")
            return
        if not filas:
            self._visor_label_info.config(text="Hoja vacia.")
            return
        # Destruir labels visuales antes de limpiar los dicts
        for lbl in self._visor_color_widgets.values():
            try: lbl.destroy()
            except Exception: pass
        self._visor_cell_colors   = {}
        self._visor_color_widgets = {}
        self._visor_cell_fonts    = {}
        self._visor_cell_edits    = {}
        self._visor_filas_ocultas = []
        self._visor_undo_stack    = []
        self._visor_iid_sel       = None
        self._visor_col_sel       = None
        self._visor_label_celda.config(text="")
        if self._visor_sel_border:
            try: self._visor_sel_border.destroy()
            except Exception: pass
            self._visor_sel_border = None
        encabezados = [str(c) if c is not None else "" for c in filas[0]]
        self._visor_cols = encabezados
        self._visor_tree.delete(*self._visor_tree.get_children())
        self._visor_tree["columns"] = encabezados
        for col in encabezados:
            self._visor_tree.heading(col, text=col, anchor="w")
            self._visor_tree.column(col, width=130, minwidth=50, anchor="w")
        for fila in filas[1:]:
            vals = [str(v) if v is not None else "" for v in fila]
            self._visor_tree.insert("", "end", values=vals)
        self._visor_label_info.config(
            text=f"{len(filas)-1} filas  |  {len(encabezados)} columnas  |  Doble clic para editar")


    # ââ Color picker âââââââââââââââââââââââââââââââââââââââââââââââââ
    def _visor_color_picker(self):
        from tkinter import colorchooser
        resultado = colorchooser.askcolor(title="Elige color de celda", parent=self)
        if not resultado or not resultado[1]:
            return
        bg = resultado[1]
        r, g, b = int(bg[1:3], 16), int(bg[3:5], 16), int(bg[5:7], 16)
        lum = 0.299*r + 0.587*g + 0.114*b
        fg = "#FFFFFF" if lum < 128 else "#000000"
        # Guardar en prÃ³ximo slot personalizado
        idx = self._PALETA_CUSTOM_IDX % len(self._PALETA_CUSTOM)
        self._PALETA_CUSTOM[idx] = (bg, fg)
        self._PALETA_CUSTOM_IDX = (idx + 1) % len(self._PALETA_CUSTOM)
        # Actualizar botÃ³n del slot
        try:
            btn = self._visor_custom_btns[idx]
            btn.config(bg=bg, activebackground=bg,
                       command=lambda b=bg, f=fg: self._visor_aplicar_color(b, f))
        except Exception:
            pass
        self._visor_aplicar_color(bg, fg)

    def _visor_aplicar_custom(self, idx):
        """Aplica el color del slot personalizado si estÃ¡ definido."""
        slot = self._PALETA_CUSTOM[idx]
        if slot:
            self._visor_aplicar_color(slot[0], slot[1])
        else:
            # Slot vacÃ­o: abrir picker y guardar en este slot
            from tkinter import colorchooser
            resultado = colorchooser.askcolor(title="Elige color para este slot", parent=self)
            if not resultado or not resultado[1]:
                return
            bg = resultado[1]
            r, g, b = int(bg[1:3], 16), int(bg[3:5], 16), int(bg[5:7], 16)
            lum = 0.299*r + 0.587*g + 0.114*b
            fg = "#FFFFFF" if lum < 128 else "#000000"
            self._PALETA_CUSTOM[idx] = (bg, fg)
            try:
                btn = self._visor_custom_btns[idx]
                btn.config(bg=bg, activebackground=bg,
                           command=lambda b=bg, f=fg: self._visor_aplicar_color(b, f))
            except Exception:
                pass
            self._visor_aplicar_color(bg, fg)

    def _visor_color_texto(self):
        from tkinter import colorchooser
        resultado = colorchooser.askcolor(title="Elige color de texto", parent=self)
        if not resultado or not resultado[1]:
            return
        color = resultado[1]
        iid, col_id = self._visor_celda_actual()
        if not iid:
            return
        est = self._visor_cell_fonts.setdefault((iid, col_id), {})
        est["textcolor"] = color
        self._visor_refresh_cell_labels()

    # ââ Fuente âââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    def _visor_toggle_bold(self):
        iid, col_id = self._visor_celda_actual()
        if not iid:
            return
        est = self._visor_cell_fonts.setdefault((iid, col_id), {})
        est["bold"] = not est.get("bold", False)
        self._visor_refresh_cell_labels()

    def _visor_toggle_italic(self):
        iid, col_id = self._visor_celda_actual()
        if not iid:
            return
        est = self._visor_cell_fonts.setdefault((iid, col_id), {})
        est["italic"] = not est.get("italic", False)
        self._visor_refresh_cell_labels()

    def _visor_toggle_underline(self):
        iid, col_id = self._visor_celda_actual()
        if not iid:
            return
        est = self._visor_cell_fonts.setdefault((iid, col_id), {})
        est["underline"] = not est.get("underline", False)
        self._visor_refresh_cell_labels()

    # ââ Ordenar ââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    def _visor_ordenar(self, asc=True):
        _, col_id = self._visor_celda_actual()
        if not col_id:
            messagebox.showinfo("Ordenar", "Selecciona una celda de la columna a ordenar.", parent=self)
            return
        items = [(self._visor_tree.set(i, col_id), i) for i in self._visor_tree.get_children()]
        def _key(v):
            try: return (0, float(v[0].replace(",", ".")))
            except Exception: return (1, v[0].lower())
        items.sort(key=_key, reverse=not asc)
        for idx, (_, iid) in enumerate(items):
            self._visor_tree.move(iid, "", idx)
        self.after(20, self._visor_refresh_cell_labels)

    # ââ Filtrar ââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    def _visor_filtrar(self):
        _, col_id = self._visor_celda_actual()
        if not col_id:
            messagebox.showinfo("Filtrar", "Selecciona una celda de la columna a filtrar.", parent=self)
            return
        cols = self._visor_tree["columns"]
        try:
            col_num  = int(col_id.replace("#", "")) - 1
            col_name = cols[col_num] if col_num < len(cols) else col_id
        except Exception:
            col_name = col_id

        # Recolectar todos los valores Ãºnicos (incluyendo filas ocultas)
        todos_iids = list(self._visor_tree.get_children()) + list(self._visor_filas_ocultas)
        valores_unicos = sorted(
            {self._visor_tree.set(i, col_id) for i in todos_iids},
            key=lambda v: (v == "", v.lower())
        )

        dlg = tk.Toplevel(self)
        dlg.title(f"Filtrar: {col_name}")
        dlg.geometry("300x420")
        dlg.resizable(True, True)
        dlg.configure(bg=COLOR_FONDO)
        dlg.grab_set()

        # Encabezado
        tk.Label(dlg, text=f"Columna: {col_name}",
                 bg=COLOR_FUCSIA, fg=COLOR_BLANCO,
                 font=("Segoe UI", 10, "bold"), padx=14, pady=8,
                 anchor="w").pack(fill="x")

        # Buscador
        buscar_var = tk.StringVar()
        buscar_frame = tk.Frame(dlg, bg=COLOR_FONDO)
        buscar_frame.pack(fill="x", padx=12, pady=(10, 4))
        tk.Label(buscar_frame, text="ð", bg=COLOR_FONDO, font=("Segoe UI", 10)).pack(side="left")
        buscar_entry = tk.Entry(buscar_frame, textvariable=buscar_var,
                                font=("Segoe UI", 10), bg=COLOR_BLANCO, fg=COLOR_TEXTO,
                                relief="solid", bd=1)
        buscar_entry.pack(side="left", fill="x", expand=True, padx=(4, 0))

        # Botones Seleccionar todo / Ninguno
        ctrl_frame = tk.Frame(dlg, bg=COLOR_FONDO)
        ctrl_frame.pack(fill="x", padx=12, pady=(2, 4))

        # Lista de valores con checkboxes
        list_frame = tk.Frame(dlg, bg=COLOR_BLANCO, relief="solid", bd=1)
        list_frame.pack(fill="both", expand=True, padx=12)
        canvas_chk = tk.Canvas(list_frame, bg=COLOR_BLANCO, highlightthickness=0)
        scroll_chk = ttk.Scrollbar(list_frame, orient="vertical", command=canvas_chk.yview)
        canvas_chk.configure(yscrollcommand=scroll_chk.set)
        scroll_chk.pack(side="right", fill="y")
        canvas_chk.pack(side="left", fill="both", expand=True)
        inner_frame = tk.Frame(canvas_chk, bg=COLOR_BLANCO)
        canvas_win = canvas_chk.create_window((0, 0), window=inner_frame, anchor="nw")

        checks = {}   # valor â BooleanVar
        chk_widgets = {}  # valor â (Frame, Checkbutton)

        def _rebuild_lista(filtro=""):
            for w in inner_frame.winfo_children():
                w.destroy()
            checks.clear()
            chk_widgets.clear()
            for val in valores_unicos:
                if filtro and filtro.lower() not in val.lower():
                    continue
                var = tk.BooleanVar(value=True)
                checks[val] = var
                row = tk.Frame(inner_frame, bg=COLOR_BLANCO)
                row.pack(fill="x", pady=1)
                cb = tk.Checkbutton(row, text=val if val != "" else "(vacÃ­o)",
                                    variable=var, bg=COLOR_BLANCO, fg=COLOR_TEXTO,
                                    selectcolor=COLOR_FUCSIA_SUAVE,
                                    activebackground=COLOR_FONDO,
                                    font=("Segoe UI", 9), anchor="w", padx=6)
                cb.pack(fill="x")
                chk_widgets[val] = cb
            inner_frame.update_idletasks()
            canvas_chk.configure(scrollregion=canvas_chk.bbox("all"))
            canvas_chk.itemconfig(canvas_win, width=canvas_chk.winfo_width())

        def _todo():
            for v in checks.values(): v.set(True)
        def _ninguno():
            for v in checks.values(): v.set(False)

        tk.Button(ctrl_frame, text="â Todo", command=_todo,
                  bg=COLOR_TARJETA, fg=COLOR_FUCSIA_OSCURO, relief="flat",
                  font=("Segoe UI", 8), cursor="hand2", padx=8, pady=2).pack(side="left")
        tk.Button(ctrl_frame, text="â Ninguno", command=_ninguno,
                  bg=COLOR_TARJETA, fg=COLOR_FUCSIA_OSCURO, relief="flat",
                  font=("Segoe UI", 8), cursor="hand2", padx=8, pady=2).pack(side="left", padx=(4, 0))

        buscar_var.trace_add("write", lambda *a: _rebuild_lista(buscar_var.get()))
        canvas_chk.bind("<Configure>",
            lambda e: canvas_chk.itemconfig(canvas_win, width=e.width))
        _rebuild_lista()
        buscar_entry.focus_set()

        def aplicar():
            seleccionados = {v for v, var in checks.items() if var.get()}
            dlg.destroy()
            self._visor_quitar_filtro()
            if not seleccionados or seleccionados == set(valores_unicos):
                return
            ocultar = [i for i in self._visor_tree.get_children()
                       if self._visor_tree.set(i, col_id) not in seleccionados]
            for iid in ocultar:
                self._visor_tree.detach(iid)
            self._visor_filas_ocultas = ocultar
            self._visor_label_info.config(
                text=f"Filtro activo en '{col_name}' â "
                     f"{len(self._visor_tree.get_children())} filas visibles")

        # BotÃ³n Aplicar
        tk.Button(dlg, text="Aplicar filtro", command=aplicar,
                  bg=COLOR_FUCSIA, fg=COLOR_BLANCO, font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=16, pady=8, cursor="hand2",
                  activebackground=COLOR_FUCSIA_OSCURO).pack(pady=10)

    def _visor_quitar_filtro(self):
        for iid in self._visor_filas_ocultas:
            try:
                self._visor_tree.reattach(iid, "", "end")
            except Exception:
                pass
        self._visor_filas_ocultas = []
        self._visor_label_info.config(text="")
        self.after(20, self._visor_refresh_cell_labels)


    # ââ Teclado visor âââââââââââââââââââââââââââââââââââââââââââââââââââ

    def _visor_celda_actual(self):
        """Devuelve (iid, col_id) de la celda seleccionada o None."""
        if self._visor_iid_sel and self._visor_col_sel:
            return self._visor_iid_sel, self._visor_col_sel
        sel = self._visor_tree.selection()
        if sel:
            return sel[0], "#1"
        return None, None

    def _visor_copiar(self):
        iid, col_id = self._visor_celda_actual()
        if not iid:
            return
        valor = self._visor_tree.set(iid, col_id)
        self.clipboard_clear()
        self.clipboard_append(valor)

    def _visor_cortar(self):
        iid, col_id = self._visor_celda_actual()
        if not iid:
            return
        valor = self._visor_tree.set(iid, col_id)
        self.clipboard_clear()
        self.clipboard_append(valor)
        self._visor_undo_stack.append((iid, col_id, valor))
        self._visor_tree.set(iid, col_id, "")

    def _visor_pegar(self):
        iid, col_id = self._visor_celda_actual()
        if not iid:
            return
        try:
            texto = self.clipboard_get()
        except Exception:
            return
        viejo = self._visor_tree.set(iid, col_id)
        self._visor_undo_stack.append((iid, col_id, viejo))
        self._visor_tree.set(iid, col_id, texto)

    def _visor_deshacer(self):
        if not self._visor_undo_stack:
            return
        iid, col_id, valor_anterior = self._visor_undo_stack.pop()
        try:
            self._visor_tree.set(iid, col_id, valor_anterior)
            self._visor_iid_sel = iid
            self._visor_col_sel = col_id
            self._visor_tree.selection_set(iid)
            self._visor_tree.see(iid)
        except Exception:
            pass

    def _visor_borrar_celda(self):
        iid, col_id = self._visor_celda_actual()
        if not iid:
            return
        viejo = self._visor_tree.set(iid, col_id)
        self._visor_undo_stack.append((iid, col_id, viejo))
        self._visor_tree.set(iid, col_id, "")

    def _visor_navegar(self, direccion):
        iid, col_id = self._visor_celda_actual()
        if not iid:
            return
        cols   = list(self._visor_tree["columns"])
        todos  = list(self._visor_tree.get_children())
        try:
            col_num = int(col_id.replace("#", "")) - 1
            row_num = todos.index(iid)
        except (ValueError, IndexError):
            return
        if direccion == "right":
            col_num = min(col_num + 1, len(cols) - 1)
        elif direccion == "left":
            col_num = max(col_num - 1, 0)
        elif direccion == "down":
            row_num = min(row_num + 1, len(todos) - 1)
        elif direccion == "up":
            row_num = max(row_num - 1, 0)
        nuevo_iid    = todos[row_num]
        nuevo_col_id = f"#{col_num + 1}"
        self._visor_iid_sel = nuevo_iid
        self._visor_col_sel = nuevo_col_id
        self._visor_tree.selection_set(nuevo_iid)
        self._visor_tree.see(nuevo_iid)
        # Actualizar label de celda
        try:
            col_letra = ""
            n = col_num + 1
            while n:
                col_letra = chr(64 + ((n - 1) % 26 + 1)) + col_letra
                n = (n - 1) // 26
            self._visor_label_celda.config(text=f"Celda: {col_letra}{row_num + 2}")
        except Exception:
            pass

    def _visor_draw_sel_border(self):
        """Dibuja un borde fucsia alrededor de la celda seleccionada."""
        # Destruir borde anterior
        if self._visor_sel_border:
            try:
                self._visor_sel_border.destroy()
            except Exception:
                pass
            self._visor_sel_border = None
        iid    = self._visor_iid_sel
        col_id = self._visor_col_sel
        if not iid or not col_id:
            return
        try:
            bbox = self._visor_tree.bbox(iid, col_id)
            if not bbox:
                return
            x, y, w, h = bbox
            bw = 2  # grosor del borde
            # Marco exterior fucsia
            outer = tk.Frame(self._visor_tree, bg=COLOR_FUCSIA, bd=0, highlightthickness=0)
            outer.place(x=x - bw, y=y - bw, width=w + bw*2, height=h + bw*2)
            # Marco interior blanco (muestra el contenido de la celda)
            color_info = self._visor_cell_colors.get((iid, col_id))
            font_info  = self._visor_cell_fonts.get((iid, col_id), {})
            bg = color_info[0] if color_info else COLOR_BLANCO
            fg = font_info.get("textcolor") or (color_info[1] if color_info else COLOR_TEXTO)
            bold    = font_info.get("bold", False)
            italic  = font_info.get("italic", False)
            underline = font_info.get("underline", False)
            style = []
            if bold:   style.append("bold")
            if italic: style.append("italic")
            font_t = ("Segoe UI", 9) + ((" ".join(style),) if style else ())
            text = self._visor_tree.set(iid, col_id)
            inner = tk.Label(outer, text=text, bg=bg, fg=fg,
                             font=font_t, anchor="w", padx=4,
                             relief="flat", bd=0,
                             underline=0 if underline else -1)
            inner.place(x=bw, y=bw, width=w, height=h)
            # Reenviar eventos desde el borde al Ã¡rbol
            for widget in (outer, inner):
                widget.bind("<Button-1>",
                    lambda e, i=iid, c=col_id: self._visor_click_desde_label(e, i, c))
                widget.bind("<Double-1>",
                    lambda e, i=iid, c=col_id: self._visor_editar_desde_label(i, c))
                widget.bind("<Button-3>",
                    lambda e, i=iid, c=col_id: self._visor_menu_contextual_label(e, i, c))
                widget.bind("<MouseWheel>",
                    lambda e: (self._visor_tree.yview_scroll(int(-1*(e.delta/120)), "units"),
                               self.after(30, self._visor_refresh_cell_labels)))
            outer.lift()
            self._visor_sel_border = outer
        except Exception:
            pass

    def _visor_click_celda(self, event):
        region = self._visor_tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        iid    = self._visor_tree.identify_row(event.y)
        col_id = self._visor_tree.identify_column(event.x)
        self._visor_iid_sel = iid
        self._visor_col_sel = col_id
        # Quitar selecciÃ³n visual de tkinter (ya desactivada por estilo)
        self._visor_tree.selection_set(iid)
        # Mostrar referencia de celda (ej: B3)
        try:
            col_num = int(col_id.replace("#", ""))
            col_letra = ""
            n = col_num
            while n:
                col_letra = chr(64 + ((n - 1) % 26 + 1)) + col_letra
                n = (n - 1) // 26
            row_num = self._visor_tree.index(iid) + 2  # +2 por encabezado
            self._visor_label_celda.config(text=f"Celda: {col_letra}{row_num}")
        except Exception:
            pass
        self._visor_draw_sel_border()

    def _visor_aplicar_color(self, bg, fg):
        iid, col_id = self._visor_celda_actual()
        if not iid:
            return
        self._visor_cell_colors[(iid, col_id)] = (bg, fg)
        self._visor_refresh_cell_labels()

    def _visor_quitar_color(self):
        iid, col_id = self._visor_celda_actual()
        if not iid:
            return
        self._visor_cell_colors.pop((iid, col_id), None)
        self._visor_refresh_cell_labels()

    def _visor_refresh_cell_labels(self):
        """Coloca Labels sobre cada celda con color y/o formato de fuente."""
        # Destruir labels antiguos
        for lbl in self._visor_color_widgets.values():
            try:
                lbl.destroy()
            except Exception:
                pass
        self._visor_color_widgets.clear()
        # UniÃ³n de celdas con color O con formato de fuente
        todas = set(self._visor_cell_colors) | set(self._visor_cell_fonts)
        for (iid, col_id) in todas:
            try:
                bbox = self._visor_tree.bbox(iid, col_id)
                if not bbox:
                    continue
                x, y, w, h = bbox
                color_info = self._visor_cell_colors.get((iid, col_id))
                font_info  = self._visor_cell_fonts.get((iid, col_id), {})
                bg  = color_info[0] if color_info else COLOR_BLANCO
                fg  = font_info.get("textcolor") or (color_info[1] if color_info else COLOR_TEXTO)
                bold      = font_info.get("bold", False)
                italic    = font_info.get("italic", False)
                underline = font_info.get("underline", False)
                style = []
                if bold:      style.append("bold")
                if italic:    style.append("italic")
                font_tuple = ("Segoe UI", 9) + ((" ".join(style),) if style else ())
                text = self._visor_tree.set(iid, col_id)
                lbl = tk.Label(
                    self._visor_tree, text=text,
                    bg=bg, fg=fg,
                    font=font_tuple, anchor="w", padx=4,
                    relief="flat", bd=0,
                    underline=0 if underline else -1,
                )
                lbl.place(x=x, y=y, width=w, height=h)
                lbl.bind("<Button-1>",
                    lambda e, i=iid, c=col_id: self._visor_click_desde_label(e, i, c))
                lbl.bind("<Double-1>",
                    lambda e, i=iid, c=col_id: self._visor_editar_desde_label(i, c))
                lbl.bind("<Button-3>",
                    lambda e, i=iid, c=col_id: self._visor_menu_contextual_label(e, i, c))
                lbl.bind("<MouseWheel>",
                    lambda e: (self._visor_tree.yview_scroll(int(-1*(e.delta/120)), "units"),
                               self.after(30, self._visor_refresh_cell_labels)))
                self._visor_color_widgets[(iid, col_id)] = lbl
            except Exception:
                pass
        self._visor_draw_sel_border()

    def _visor_click_desde_label(self, event, iid, col_id):
        self._visor_iid_sel = iid
        self._visor_col_sel = col_id
        self._visor_tree.selection_set(iid)
        self._visor_tree.see(iid)
        self._visor_tree.focus_set()
        try:
            col_num = int(col_id.replace("#", ""))
            col_letra = ""
            n = col_num
            while n:
                col_letra = chr(64 + ((n - 1) % 26 + 1)) + col_letra
                n = (n - 1) // 26
            row_num = self._visor_tree.index(iid) + 2
            self._visor_label_celda.config(text=f"Celda: {col_letra}{row_num}")
        except Exception:
            pass

    def _visor_editar_desde_label(self, iid, col_id):
        self._visor_iid_sel = iid
        self._visor_col_sel = col_id
        self._visor_editar_celda()

    def _visor_menu_contextual_label(self, event, iid, col_id):
        self._visor_iid_sel = iid
        self._visor_col_sel = col_id
        self._visor_tree.selection_set(iid)
        self._visor_menu_contextual(event)

    def _visor_editar_celda(self, event=None):
        if event:
            if self._visor_tree.identify_region(event.x, event.y) != "cell":
                return
            col_id = self._visor_tree.identify_column(event.x)
            iid    = self._visor_tree.identify_row(event.y)
        elif self._visor_iid_sel and self._visor_col_sel:
            iid    = self._visor_iid_sel
            col_id = self._visor_col_sel
        else:
            sel = self._visor_tree.selection()
            if not sel:
                return
            iid    = sel[0]
            col_id = "#1"
        try:
            bbox = self._visor_tree.bbox(iid, col_id)
        except Exception:
            return
        if not bbox:
            return
        x, y, w, h = bbox
        valor_actual = self._visor_tree.set(iid, col_id)
        entry = tk.Entry(self._visor_tree, font=("Segoe UI", 9),
                         bg=COLOR_BLANCO, fg=COLOR_TEXTO,
                         highlightbackground=COLOR_FUCSIA, highlightthickness=1,
                         relief="flat")
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, valor_actual)
        entry.select_range(0, "end")
        entry.focus_set()
        def _guardar(e=None):
            self._visor_tree.set(iid, col_id, entry.get())
            entry.destroy()
        entry.bind("<Return>",   _guardar)
        entry.bind("<Tab>",      _guardar)
        entry.bind("<Escape>",   lambda e: entry.destroy())
        entry.bind("<FocusOut>", _guardar)

    def _visor_editar_celda_manual(self):
        self._visor_editar_celda(event=None)

    def _visor_menu_contextual(self, event):
        iid = self._visor_tree.identify_row(event.y)
        if iid and iid not in self._visor_tree.selection():
            self._visor_tree.selection_set(iid)
        menu = tk.Menu(self, tearoff=0)
        sub = tk.Menu(menu, tearoff=0)
        for bg, fg, nombre in self._PALETA:
            sub.add_command(label=f"  {nombre}",
                            command=lambda b=bg, f=fg: self._visor_aplicar_color(b, f))
        menu.add_cascade(label="Color de fila", menu=sub)
        menu.add_command(label="Quitar color", command=self._visor_quitar_color)
        menu.add_separator()
        menu.add_command(label="Editar celda", command=self._visor_editar_celda_manual)
        menu.add_command(label="Agregar fila abajo", command=self._visor_agregar_fila)
        menu.add_command(label="Eliminar fila", command=self._visor_borrar_fila)
        menu.add_separator()
        menu.add_command(label="Guardar cambios en Excel", command=self._visor_guardar_cambios)
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _visor_agregar_fila(self):
        cols = self._visor_tree["columns"]
        if not cols:
            return
        sel = self._visor_tree.selection()
        pos = self._visor_tree.index(sel[-1]) + 1 if sel else "end"
        iid = self._visor_tree.insert("", pos, values=[""] * len(cols))
        self._visor_tree.selection_set(iid)
        self._visor_tree.see(iid)

    def _visor_borrar_fila(self):
        for iid in self._visor_tree.selection():
            # Limpiar colores de todas las celdas de la fila eliminada
            keys_to_del = [k for k in self._visor_cell_colors if k[0] == iid]
            for k in keys_to_del:
                self._visor_cell_colors.pop(k, None)
                lbl = self._visor_color_widgets.pop(k, None)
                if lbl:
                    try: lbl.destroy()
                    except Exception: pass
            self._visor_tree.delete(iid)

    def _visor_borrar_reporte(self):
        sel = self._visor_lb.curselection()
        if sel:
            rutas = [self._visor_archivos[i] for i in sel]
        elif self._visor_ruta_actual:
            rutas = [self._visor_ruta_actual]
        else:
            messagebox.showinfo("Borrar reporte",
                "Selecciona uno o mÃ¡s reportes de la lista primero.", parent=self)
            return

        n = len(rutas)
        msg = (f"Â¿Eliminar permanentemente {n} reporte(s)?\n"
               + "\n".join(os.path.basename(r) for r in rutas[:5])
               + ("\n..." if n > 5 else "")
               + "\n\nEsta acciÃ³n no se puede deshacer.")
        if not messagebox.askyesno("Borrar reporte", msg, parent=self):
            return

        borrados = []
        errores  = []
        for ruta in rutas:
            try:
                if os.path.exists(ruta):
                    os.remove(ruta)
                # Si era el archivo cargado, limpiar visor
                if ruta == self._visor_ruta_actual:
                    for lbl in self._visor_color_widgets.values():
                        try: lbl.destroy()
                        except Exception: pass
                    self._visor_cell_colors   = {}
                    self._visor_color_widgets = {}
                    self._visor_cell_fonts    = {}
                    self._visor_filas_ocultas = []
                    self._visor_ruta_actual   = None
                    self._visor_tree.delete(*self._visor_tree.get_children())
                    self._visor_tree["columns"] = []
                    self._visor_combo_hoja.set("")
                    self._visor_combo_hoja["values"] = []
                    self._visor_label_celda.config(text="")
                borrados.append(ruta)
            except Exception as e:
                errores.append(f"{os.path.basename(ruta)}: {e}")

        # Eliminar de la lista de mayor a menor Ã­ndice para no desplazar
        indices_a_borrar = sorted(
            [self._visor_archivos.index(r) for r in borrados if r in self._visor_archivos],
            reverse=True
        )
        for idx in indices_a_borrar:
            self._visor_lb.delete(idx)
            self._visor_archivos.pop(idx)

        txt = f"{len(borrados)} reporte(s) eliminado(s)."
        if errores:
            txt += f"  Errores: {'; '.join(errores)}"
        self._visor_label_info.config(text=txt)

    def _visor_guardar_cambios(self):
        ruta = self._visor_ruta_actual
        hoja = self._visor_hoja.get() if hasattr(self, "_visor_hoja") else ""
        if not ruta or not hoja:
            messagebox.showinfo("Visor", "Abre un archivo primero.", parent=self)
            return
        try:
            import openpyxl
            from openpyxl.styles import PatternFill
            wb = openpyxl.load_workbook(ruta)
            ws = wb[hoja]
            all_iids = list(self._visor_tree.get_children())
            cols = self._visor_tree["columns"]
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
                    cell.fill  = PatternFill()
            for r_idx, iid in enumerate(all_iids, start=2):
                row_vals = [self._visor_tree.set(iid, col) for col in cols]
                for c_idx, val in enumerate(row_vals, start=1):
                    col_id = f"#{c_idx}"
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    color_info = self._visor_cell_colors.get((iid, col_id))
                    font_info  = self._visor_cell_fonts.get((iid, col_id), {})
                    from openpyxl.styles import Font as OXFont
                    fill_color = color_info[0].lstrip("#") if color_info else None
                    text_color = font_info.get("textcolor", color_info[1] if color_info else None)
                    if fill_color:
                        cell.fill = PatternFill("solid", fgColor=fill_color)
                    cell.font = OXFont(
                        color=text_color.lstrip("#") if text_color else "000000",
                        bold=font_info.get("bold", False),
                        italic=font_info.get("italic", False),
                        underline="single" if font_info.get("underline") else None,
                    )
            wb.save(ruta)
            self._visor_label_info.config(text=f"Guardado: {os.path.basename(ruta)}")
            self._log(f"Visor: cambios guardados en {os.path.basename(ruta)}", ok=True)
        except Exception as e:
            messagebox.showerror("Error al guardar", str(e), parent=self)

    # ---------------------------------------------------------------- #
    # PESTAÃA: VENTAS DEL DÃA
    # ---------------------------------------------------------------- #
    def _tab_ventas_dia(self, nb):
        tab = ttk.Frame(nb)
        nb.add(tab, text="  â½ Ventas del DÃ­a  ")
        tab.columnconfigure(0, weight=1)

        # ââ Instrucciones ââ
        info = ttk.Frame(tab, style="Tarjeta.TFrame")
        info.grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 4))
        ttk.Label(info,
            text="Genera la pÃ³liza de Ventas del DÃ­a en formato Excel. "
                 "Selecciona el archivo de control de despachos y la plantilla de cuentas.",
            style="Tarjeta.TLabel", wraplength=900,
        ).pack(padx=12, pady=8)

        # ââ SelecciÃ³n de archivos ââ
        card = ttk.Frame(tab, style="Tarjeta.TFrame")
        card.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 4))
        card.columnconfigure(1, weight=1)

        self._vd_despachos   = tk.StringVar(value="")
        self._vd_plantilla   = tk.StringVar(value="")
        self._vd_resultado_path = None

        ttk.Label(card, text="ð Control de despachos (.xlsx):", width=30,
                  style="Tarjeta.TLabel").grid(row=0, column=0, sticky="w", padx=10, pady=(10, 4))
        ttk.Entry(card, textvariable=self._vd_despachos).grid(
            row=0, column=1, sticky="ew", padx=6, pady=(10, 4))
        ttk.Button(card, text="Elegir...",
            command=lambda: self._vd_elegir_archivo(self._vd_despachos)).grid(
            row=0, column=2, padx=6, pady=(10, 4))

        ttk.Label(card, text="ð Plantilla de cuentas (.xlsx):", width=30,
                  style="Tarjeta.TLabel").grid(row=1, column=0, sticky="w", padx=10, pady=(0, 10))
        ttk.Entry(card, textvariable=self._vd_plantilla).grid(
            row=1, column=1, sticky="ew", padx=6, pady=(0, 10))
        ttk.Button(card, text="Elegir...",
            command=lambda: self._vd_elegir_archivo(self._vd_plantilla)).grid(
            row=1, column=2, padx=6, pady=(0, 10))

        # Cuentas IEPS (valores fijos, sin UI)
        self._vd_ieps_gs = tk.StringVar(value="401-01-0001-0006-0001")
        self._vd_ieps_gp = tk.StringVar(value="401-01-0001-0006-0002")
        self._vd_ieps_gd = tk.StringVar(value="401-01-0001-0006-0003")

        # ââ BotÃ³n generar ââ
        ttk.Button(tab,
            text="â½  Generar PÃ³liza Ventas del DÃ­a",
            style="Grande.TButton",
            command=self._vd_generar,
        ).grid(row=3, column=0, sticky="ew", padx=8, pady=(6, 2))

        # ââ Barra de progreso VD ââ
        _vd_pb_frame = tk.Frame(tab, bg=COLOR_FONDO)
        _vd_pb_frame.grid(row=4, column=0, sticky="ew", padx=8, pady=(2, 2))
        _vd_pb_frame.grid_remove()
        self._vd_pb_frame = _vd_pb_frame
        self._vd_pb = FunkyProgressBar(_vd_pb_frame, maximum=100, height=70)
        self._vd_pb.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self._vd_pb_lbl = ttk.Label(_vd_pb_frame, text="0 %", width=8,
            foreground=COLOR_FUCSIA_OSCURO, font=("Segoe UI", 8, "bold"))
        self._vd_pb_lbl.pack(side="left")

        # ââ Barra de descarga ââ
        self._vd_barra = ttk.Frame(tab, style="Tarjeta.TFrame")
        self._vd_barra.grid(row=5, column=0, sticky="ew", padx=8, pady=(0, 6))

        ttk.Label(self._vd_barra, text="ð Resultado:", style="Tarjeta.TLabel").pack(
            side="left", padx=(10, 4), pady=8)
        self._vd_lbl_archivo = ttk.Label(self._vd_barra, text="â",
            foreground=COLOR_FUCSIA_OSCURO, font=("Segoe UI", 9, "bold"),
            background=COLOR_TARJETA)
        self._vd_lbl_archivo.pack(side="left", padx=4, pady=8)

        ttk.Button(self._vd_barra, text="ð  Abrir",
            command=self._vd_abrir).pack(side="right", padx=4, pady=6)
        ttk.Button(self._vd_barra, text="ð¾  Guardar como...",
            command=self._vd_guardar_como).pack(side="right", padx=4, pady=6)

        # ââ Vista previa inline ââ
        tab.rowconfigure(6, weight=1)
        vd_visor_wrap = ttk.Frame(tab, style="Tarjeta.TFrame")
        vd_visor_wrap.grid(row=6, column=0, sticky="nsew", padx=8, pady=(0, 8))
        vd_visor_wrap.columnconfigure(0, weight=1)
        vd_visor_wrap.rowconfigure(1, weight=1)
        ttk.Label(vd_visor_wrap, text="ð  Vista previa del resultado",
                  style="Tarjeta.TLabel",
                  font=("Segoe UI", 9, "bold")).grid(
            row=0, column=0, sticky="w", padx=10, pady=(6, 2))
        _vd_tv_frame = tk.Frame(vd_visor_wrap, bg=COLOR_FONDO)
        _vd_tv_frame.grid(row=1, column=0, sticky="nsew", padx=6, pady=(0, 6))
        _vd_tv_frame.columnconfigure(0, weight=1)
        _vd_tv_frame.rowconfigure(0, weight=1)
        self._vd_tv = ttk.Treeview(_vd_tv_frame, show="headings", height=6)
        _vd_vsb = ttk.Scrollbar(_vd_tv_frame, orient="vertical",   command=self._vd_tv.yview)
        _vd_hsb = ttk.Scrollbar(_vd_tv_frame, orient="horizontal", command=self._vd_tv.xview)
        self._vd_tv.configure(yscrollcommand=_vd_vsb.set, xscrollcommand=_vd_hsb.set)
        self._vd_tv.grid(row=0, column=0, sticky="nsew")
        _vd_vsb.grid(row=0, column=1, sticky="ns")
        _vd_hsb.grid(row=1, column=0, sticky="ew")

    def _vd_cargar_visor_inline(self, path):
        tv = getattr(self, "_vd_tv", None)
        if tv is None:
            return
        for item in tv.get_children():
            tv.delete(item)
        if not path or not os.path.isfile(path):
            return
        try:
            import openpyxl as _opxl
            wb = _opxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not all_rows:
                return
            headers = [str(c) if c is not None else "" for c in all_rows[0]]
            tv["columns"] = headers
            tv["show"] = "headings"
            for h in headers:
                tv.heading(h, text=h)
                tv.column(h, width=140, anchor="w", stretch=True)
            for row in all_rows[1:51]:
                vals = [str(c) if c is not None else "" for c in row]
                tv.insert("", "end", values=vals)
        except Exception as _e:
            self._log(f"  â  Vista previa: {_e}", True)

    def _vd_elegir_archivo(self, var):
        init = self.carpeta.get() or os.getcwd()
        if not os.path.isdir(init):
            init = os.getcwd()
        self.lift()
        self.focus_force()
        self.update()
        ruta = filedialog.askopenfilename(
            parent=self,
            initialdir=init,
            title="Seleccionar archivo Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos los archivos", "*.*")],
        )
        if ruta:
            var.set(ruta)

    def _vd_abrir(self):
        if self._vd_resultado_path and os.path.exists(self._vd_resultado_path):
            os.startfile(self._vd_resultado_path)
        else:
            messagebox.showwarning("Sin archivo", "Primero genera la pÃ³liza.")

    def _vd_guardar_como(self):
        if not self._vd_resultado_path or not os.path.exists(self._vd_resultado_path):
            messagebox.showwarning("Sin archivo", "Primero genera la pÃ³liza.")
            return
        self.lift()
        self.focus_force()
        self.update()
        destino = filedialog.asksaveasfilename(
            parent=self,
            title="Guardar pÃ³liza como...",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=os.path.basename(self._vd_resultado_path),
        )
        if destino:
            import shutil
            shutil.copy2(self._vd_resultado_path, destino)
            messagebox.showinfo("Guardado", f"Guardado en:\n{destino}")

    def _vd_abrir_en_visor(self, ruta_xlsx, carpeta):
        """Tras generar, carga el archivo automÃ¡ticamente en el Visor de Resultados."""
        try:
            # 1. Cambiar al tab Visor de Resultados
            for i in range(self.nb.index("end")):
                if "Visor" in self.nb.tab(i, "text"):
                    self.nb.select(i)
                    break
            # 2. Apuntar la carpeta del visor a donde se guardÃ³
            self._visor_carpeta.set(carpeta)
            self._visor_refrescar()
            # 3. Seleccionar el archivo reciÃ©n generado en la lista
            nombre = os.path.basename(ruta_xlsx)
            for idx, arch in enumerate(self._visor_archivos):
                if os.path.basename(arch) == nombre:
                    self._visor_lb.selection_clear(0, "end")
                    self._visor_lb.selection_set(idx)
                    self._visor_lb.see(idx)
                    self._visor_cargar_archivo()
                    break
        except Exception:
            pass  # No crÃ­tico â el archivo ya fue generado

    def _vd_generar(self):
        try:
            despachos = self._vd_despachos.get().strip()
            plantilla = self._vd_plantilla.get().strip()
            if not despachos or not os.path.isfile(despachos):
                messagebox.showwarning("Archivo faltante", "Elige el archivo de control de despachos.")
                return
            if not plantilla or not os.path.isfile(plantilla):
                messagebox.showwarning("Archivo faltante", "Elige la plantilla de cuentas.")
                return
            self._vd_lbl_archivo.config(text="Generando...")
            self._vd_pb_frame.grid()
            self._pb_iniciar(self._vd_pb, self._vd_pb_lbl)
            threading.Thread(target=self._vd_generar_hilo, args=(despachos, plantilla), daemon=True).start()
        except Exception as _e:
            import traceback as _tb
            messagebox.showerror("Error al iniciar",
                f"No se pudo iniciar la generaciÃ³n:\n{_e}\n\n{_tb.format_exc()[-600:]}")

    def _vd_generar_hilo(self, despachos_path, plantilla_path):
        try:
            # ââ Verificar xlsxwriter ââââââââââââââââââââââââââââââââââââââââââ
            try:
                import xlsxwriter
            except ImportError:
                self.after(0, self._pb_error, self._vd_pb, self._vd_pb_lbl)
                self.after(0, self._vd_pb_frame.grid_remove)
                self.after(0, messagebox.showerror, "LibrerÃ­a faltante",
                    "Falta la librerÃ­a 'xlsxwriter'.\n\n"
                    "InstÃ¡lala ejecutando en la terminal:\n"
                    "    pip install xlsxwriter\n\n"
                    "Luego reinicia la aplicaciÃ³n.")
                self.after(0, self._vd_lbl_archivo.config, {"text": "Error"})
                return

            import openpyxl
            from collections import defaultdict

            import re as _re

            # âââ Helpers match/detecciÃ³n (equivalentes al mÃ³dulo Streamlit) ââââââ
            def _vd_norm(s):
                return str(s or "").replace("\n", " ").strip().upper()
            def _vd_clean(s):
                s = str(s or "").replace("\n", " ")
                s = _re.sub(r"\(.*?\)", "", s)
                s = s.replace(".", " ")
                return _re.sub(r"\s+", " ", s).strip().upper()
            def _vd_sig_words(s):
                return {w for w in _vd_clean(s).split() if len(w) > 3}
            def _vd_match_cliente(raw, clientes_list):
                raw_s = str(raw or "").strip()
                raw_n = _vd_norm(raw_s)
                if not raw_n:
                    for nm in clientes_list:
                        if _vd_norm(nm) == "CONTADO":
                            return nm
                    return clientes_list[0] if clientes_list else raw_s
                for nm in clientes_list:
                    if _vd_norm(nm) == raw_n:
                        return nm
                for nm in clientes_list:
                    nm_n = _vd_norm(nm)
                    if raw_n in nm_n or nm_n in raw_n:
                        return nm
                raw_c = _vd_clean(raw_s)
                for nm in clientes_list:
                    nm_c = _vd_clean(nm)
                    if raw_c in nm_c or nm_c in raw_c:
                        return nm
                raw_words = _vd_sig_words(raw_s)
                best, best_score = None, 0
                for nm in clientes_list:
                    nm_words = _vd_sig_words(nm)
                    score = len(raw_words & nm_words)
                    if score > best_score:
                        best_score, best = score, nm
                if best_score > 0:
                    return best
                return raw_s
            def _vd_detectar_columnas(header_row):
                col_map = {}
                for i, h in enumerate(header_row):
                    if h is not None:
                        col_map[str(h).strip()] = i
                defaults = {
                    "FechaHora": 0, "Producto": 3, "Subtotal": 6,
                    "Iva": 7, "Ieps": 8, "Importe": 9, "Cliente": 16,
                }
                return {k: col_map.get(k, v) for k, v in defaults.items()}

            # ââ Leer despachos (soporta .xls y .xlsx) ââââââââââââââââââââââââ
            self.after(0, self._log, "â½ Leyendo control de despachos...")
            ext = os.path.splitext(despachos_path)[1].lower()
            if ext == '.xls':
                # Formato antiguo: intentar con xlrd
                try:
                    import xlrd
                    _wb_xls = xlrd.open_workbook(despachos_path)
                    _ws_xls = _wb_xls.sheet_by_index(0)
                    _rows_raw = [
                        tuple(_ws_xls.row_values(r))
                        for r in range(_ws_xls.nrows)
                    ]
                    _col_map = _vd_detectar_columnas(_rows_raw[0])
                    data = _rows_raw[1:]  # saltar encabezado
                    self.after(0, self._log, f"  {len(data):,} registros leÃ­dos (.xls vÃ­a xlrd).")
                except ImportError:
                    self.after(0, self._pb_error, self._vd_pb, self._vd_pb_lbl)
                    self.after(0, self._vd_pb_frame.grid_remove)
                    self.after(0, messagebox.showerror, "Formato .xls no soportado",
                        "El archivo de despachos estÃ¡ en formato antiguo .xls\n"
                        "y la librerÃ­a 'xlrd' no estÃ¡ instalada.\n\n"
                        "Soluciones:\n"
                        "  1. Abre el archivo en Excel y guÃ¡rdalo como .xlsx\n"
                        "  2. O instala xlrd:  pip install xlrd==1.2.0")
                    self.after(0, self._vd_lbl_archivo.config, {"text": "Error"})
                    return
            else:
                wb_desp = openpyxl.load_workbook(despachos_path, data_only=True, read_only=True)
                _all_rows = list(wb_desp.active.iter_rows(values_only=True))
                wb_desp.close()
                _col_map = _vd_detectar_columnas(_all_rows[0])
                data = _all_rows[1:]
                self.after(0, self._log, f"  {len(data):,} registros leÃ­dos (.xlsx).")

            self.after(0, self._log, "â½ Leyendo plantilla de cuentas...")
            wb_tpl = openpyxl.load_workbook(plantilla_path, data_only=True)

            cuentas_map = {
                self._vd_ieps_gs.get().strip(): "IEPS De Gasolina Magna",
                self._vd_ieps_gp.get().strip(): "IEPS de Premium",
                self._vd_ieps_gd.get().strip(): "IEPS de Diesel",
            }
            # Buscar hoja de cuentas (nombre con espacio posible)
            hoja_cuentas = None
            for sn in wb_tpl.sheetnames:
                if sn.strip().upper() == "CUENTAS":
                    hoja_cuentas = wb_tpl[sn]
                    break
            if hoja_cuentas is None:
                self.after(0, self._log, "  â  No se encontrÃ³ hoja 'cuentas' en la plantilla.", True)

            # ── Extraer listas dinámicas de la hoja CUENTAS ──────────────────
            _dyn_clientes = []   # [(num_cuenta, nombre)] — columnas H/I (7,8)
            _dyn_prods    = []   # [(num_cuenta, nombre)] — columnas L/M (11,12)
            if hoja_cuentas is not None:
                for r in hoja_cuentas.iter_rows(values_only=True):
                    _nc  = r[7]  if len(r) > 7  else None
                    _nm  = r[8]  if len(r) > 8  else None
                    if _nc and _nm:
                        _ncs = str(_nc).strip(); _nms = str(_nm).strip().replace('\n', '').strip()
                        if _ncs and _ncs[0].isdigit() and _nms:
                            cuentas_map[_ncs] = _nms
                            _dyn_clientes.append((_ncs, _nms))
                    _pc  = r[11] if len(r) > 11 else None
                    _pm  = r[12] if len(r) > 12 else None
                    if _pc and _pm:
                        _pcs = str(_pc).strip(); _pms = str(_pm).strip().replace('\n', '').strip()
                        if _pcs and _pcs[0].isdigit() and _pms:
                            cuentas_map[_pcs] = _pms
                            _dyn_prods.append((_pcs, _pms))

            # Listas de respaldo hardcoded
            _CLIENTES_FB = [
                "T BANORTE","Contado","T EDENRED (ACCOR)","T EFECTIVALE",
                "M Y T INTEGRALES PARA LA SALUD","T AMERICAN EXPRESS",
                "CENTRO DE DISTRIBUCION ORIENTE","T ULTRAGAS","T PLUXEE MEXICO",
                "JUAN ANTONIO CRUZ MONDRAGON","T INBURSA","V EFECTIVALE",
                "ETIQUETAS CCL","MARIA DEL CARMEN PEÃALOZA PEIMBERT",
                "ROTULACIONES E IMPRESIONES MEXICANAS SA DE CV",
                "ALMACENAJE Y DISTRIBUCION TRANSGALLA","PETRO ASFALTOS DEL SURESTE",
                "GRAFIARTE DELA","MARICELA GONZALEZ RODRIGUEZ","ADEPT SERVICES MEXICO",
            ]
            _CUENTAS_FB = [
                "105-01-0001-0004","105-01-0001-0001","105-01-0002-0002","105-01-0002-0001",
                "105-01-0003-3600","105-01-0001-0007","105-01-0003-0601","105-01-0002-0003",
                "105-01-0002-0007","105-01-0003-2700","105-01-0002-0006","105-01-0002-0005",
                "105-01-0003-1200","105-01-0003-3601","105-01-0003-5400","105-01-0003-0002",
                "105-01-0003-4701","105-01-0003-1800","105-01-0003-3602","105-01-0003-0001",
            ]
            if _dyn_clientes:
                CUENTAS_TPL  = [nc for nc, _ in _dyn_clientes]
                CLIENTES_TPL = [nm for _, nm in _dyn_clientes]
                self.after(0, self._log, f"  {len(CLIENTES_TPL)} clientes leídos de hoja CUENTAS.")
            else:
                CUENTAS_TPL  = _CUENTAS_FB
                CLIENTES_TPL = _CLIENTES_FB
                self.after(0, self._log, "  Usando lista predeterminada.")
            NOMBRES_TPL = [cuentas_map.get(c, cli) for c, cli in zip(CUENTAS_TPL, CLIENTES_TPL)]

            META_HDRS = ["TIPO DE POLIZA","Fecha","REFERENCIA","CONCEPTO","ERROR","UIDD","NUM POLIZA","PROCESADO"]
            N_META = len(META_HDRS)
            _PRODS_FB = ["GS","GP","GD","IVA","IEPS GS","IEPS GP","IEPS GD"]
            _CTAS_PROD_FB = [
                "401-01-0001-0001","401-01-0001-0002","401-01-0001-0003","209-01",
                self._vd_ieps_gs.get().strip(),
                self._vd_ieps_gp.get().strip(),
                self._vd_ieps_gd.get().strip(),
            ]
            if len(_dyn_prods) == 7:
                CTAS_PROD = [pc for pc, _ in _dyn_prods]
                PRODS     = [pm for _, pm in _dyn_prods]
            else:
                PRODS     = _PRODS_FB
                CTAS_PROD = _CTAS_PROD_FB
            NOMS_PROD = [cuentas_map.get(c, p) for c, p in zip(CTAS_PROD, PRODS)]

            # ââ Columnas detectadas ââââââââââââââââââââââââââââ
            C_FECHA    = _col_map["FechaHora"]
            C_PROD     = _col_map["Producto"]
            C_SUBTOTAL = _col_map["Subtotal"]
            C_IVA      = _col_map["Iva"]
            C_IEPS     = _col_map["Ieps"]
            C_IMPORTE  = _col_map["Importe"]
            C_CLIENTE  = _col_map["Cliente"]
            self.after(0, self._log,
                f"  Columnas: Producto={C_PROD} Importe={C_IMPORTE} Cliente={C_CLIENTE}")

            cli_day   = defaultdict(float)
            prod_day  = defaultdict(float)
            iva_day   = defaultdict(float)
            ieps_prod = defaultdict(float)

            for r in data:
                try:
                    fecha   = str(r[C_FECHA])[:10] if r[C_FECHA] is not None else ""
                    cliente = _vd_match_cliente(str(r[C_CLIENTE] or "").strip(), CLIENTES_TPL)
                    prod    = str(r[C_PROD] or "")
                    cli_day[(fecha, cliente)]  += float(r[C_IMPORTE]  or 0)
                    prod_day[(fecha, prod)]    += float(r[C_SUBTOTAL] or 0)
                    iva_day[fecha]             += float(r[C_IVA]      or 0)
                    ieps_prod[(fecha, prod)]   += float(r[C_IEPS]     or 0)
                except Exception:
                    continue

            fechas = sorted(set(k[0] for k in cli_day))
            N_CLI  = len(CLIENTES_TPL)
            N_PROD = len(PRODS)
            OFF    = N_META
            COL_TOT1  = OFF + N_CLI
            COL_PROD0 = OFF + N_CLI + 1
            COL_TOT2  = OFF + N_CLI + 1 + N_PROD
            COL_CONC  = OFF + N_CLI + 1 + N_PROD + 1
            TOTAL_COLS = COL_CONC + 1

            # Nombre del archivo de salida basado en el archivo de entrada
            base = os.path.splitext(os.path.basename(despachos_path))[0]
            carpeta_sal = (self.salida_dir.get().strip() or
                           os.path.join(os.path.dirname(despachos_path), "resultados"))
            os.makedirs(carpeta_sal, exist_ok=True)
            OUT = os.path.join(carpeta_sal, f"poliza_ventas_dia_{base}.xlsx")

            wb  = xlsxwriter.Workbook(OUT, {'constant_memory': True})
            ws  = wb.add_worksheet("poliza IA")

            def fmt(d): return wb.add_format(d)
            CURR = 'General'
            BASE = {'font_name':'Arial','font_size':9,'border':1,'border_color':'#CCCCCC','valign':'vcenter'}

            f_title   = fmt({'bold':True,'font_color':'#FFFFFF','bg_color':'#C2185B','font_size':12,'align':'center','valign':'vcenter','font_name':'Arial'})
            f_acct    = fmt({**BASE,'bold':True,'bg_color':'#3B0764','font_color':'#FFFFFF','align':'center','font_size':8})
            f_hdr_m   = fmt({**BASE,'bold':True,'bg_color':'#6A1B9A','font_color':'#FFFFFF','align':'center','text_wrap':True,'font_size':8})
            f_hdr_c   = fmt({**BASE,'bold':True,'bg_color':'#1565C0','font_color':'#FFFFFF','align':'center','text_wrap':True,'font_size':8})
            f_hdr_p   = fmt({**BASE,'bold':True,'bg_color':'#0D47A1','font_color':'#FFFFFF','align':'center','text_wrap':True,'font_size':8})
            f_hdr_tot = fmt({**BASE,'bold':True,'bg_color':'#1B5E20','font_color':'#FFFFFF','align':'center','font_size':8})
            f_hdr_con = fmt({**BASE,'bold':True,'bg_color':'#E65100','font_color':'#FFFFFF','align':'center','font_size':8})
            f_fecha   = fmt({**BASE,'bold':True,'bg_color':'#F3E5F5','align':'center','num_format':'dd/mm/yyyy'})
            f_meta    = fmt({**BASE,'bg_color':'#EDE7F6','align':'left','font_size':8})
            f_num0    = fmt({**BASE,'bg_color':'#FFFFFF','align':'right','num_format':CURR})
            f_num1    = fmt({**BASE,'bg_color':'#B7D9EF','align':'right','num_format':CURR})
            f_tot0    = fmt({**BASE,'bold':True,'bg_color':'#E6F2FB','align':'right','num_format':CURR})
            f_tot1    = fmt({**BASE,'bold':True,'bg_color':'#DCEDC8','align':'right','num_format':CURR})
            f_conc0   = fmt({**BASE,'bold':True,'bg_color':'#FFF9C4','font_color':'#E65100','align':'right','num_format':CURR})
            f_conc1   = fmt({**BASE,'bold':True,'bg_color':'#FFF176','font_color':'#E65100','align':'right','num_format':CURR})
            f_grand   = fmt({**BASE,'bold':True,'bg_color':'#1B5E20','font_color':'#FFFFFF','align':'right','num_format':CURR,'border':2,'border_color':'#000000'})
            f_grand_l = fmt({**BASE,'bold':True,'bg_color':'#1B5E20','font_color':'#FFFFFF','align':'center','border':2,'border_color':'#000000'})
            f_grand_c = fmt({**BASE,'bold':True,'bg_color':'#E65100','font_color':'#FFFFFF','align':'right','num_format':CURR,'border':2,'border_color':'#000000'})

            ws.set_row(0, 14); ws.set_row(1, 24); ws.set_row(2, 18); ws.set_row(3, 50)

            # Fila 0: numeración 0-based
            for c in range(TOTAL_COLS): ws.write(0, c, c, f_acct)

            # Fila 1: título
            titulo = "VENTAS DEL DIA — SUPER SERVICIO PERIFERICO"
            ws.merge_range(1, 0, 1, TOTAL_COLS-1, titulo, f_title)

            # Fila 2: número de cuenta
            for c in range(N_META): ws.write(2, c, c, f_acct)
            for i, acct in enumerate(CUENTAS_TPL): ws.write(2, OFF + i, acct, f_acct)
            ws.write(2, COL_TOT1, COL_TOT1, f_acct)
            for i, acct in enumerate(CTAS_PROD): ws.write(2, COL_PROD0 + i, acct, f_acct)
            ws.write(2, COL_TOT2, COL_TOT2, f_acct)
            ws.write(2, COL_CONC, COL_CONC, f_acct)

            # Fila 3: encabezados
            for i, h in enumerate(META_HDRS): ws.write(3, i, h, f_hdr_m)
            for i, nom in enumerate(NOMBRES_TPL): ws.write(3, OFF+i, nom, f_hdr_c)
            ws.write(3, COL_TOT1, "TOTAL B2", f_hdr_tot)
            for i, nom in enumerate(NOMS_PROD): ws.write(3, COL_PROD0+i, nom, f_hdr_p)
            ws.write(3, COL_TOT2, "TOTAL B2", f_hdr_tot)
            ws.write(3, COL_CONC, "CONCILIACION", f_hdr_con)

            ws.set_column(0, 0, 14); ws.set_column(1, 1, 12)
            for c in range(2, N_META): ws.set_column(c, c, 20)
            for i in range(N_CLI): ws.set_column(OFF+i, OFF+i, 14)
            ws.set_column(COL_TOT1, COL_TOT1, 13)
            for i in range(N_PROD): ws.set_column(COL_PROD0+i, COL_PROD0+i, 13)
            ws.set_column(COL_TOT2, COL_TOT2, 13)
            ws.set_column(COL_CONC, COL_CONC, 14)
            ws.freeze_panes(4, 2)

            gran_cli  = [0.0] * N_CLI
            gran_tot1 = 0.0
            gran_prod = [0.0] * N_PROD
            gran_tot2 = 0.0
            gran_conc = 0.0

            for ri, fecha in enumerate(fechas):
                row = ri + 4
                fn  = f_num0 if ri%2==0 else f_num1
                ft  = f_tot0 if ri%2==0 else f_tot1
                fc  = f_conc0 if ri%2==0 else f_conc1

                # Convertir fecha a objeto date para Excel
                try:
                    from datetime import datetime as _dt
                    _fd = _dt.strptime(fecha[:10], '%Y-%m-%d')
                    fecha_display = _fd.strftime('%d/%m/%Y')
                    _fecha_val = _fd.date()
                except Exception:
                    fecha_display = fecha
                    _fecha_val = fecha

                ws.write(row, 0, "D", f_meta)
                ws.write_datetime(row, 1, _fecha_val, f_fecha)
                ws.write(row, 2, "VENTA DEL DIA " + fecha_display, f_meta)
                ws.write(row, 3, "VENTA DEL DIA " + fecha_display, f_meta)
                for c in range(4, N_META): ws.write(row, c, "", f_meta)

                total_b2 = 0.0
                for i, cli in enumerate(CLIENTES_TPL):
                    v = round(cli_day.get((fecha, cli), 0.0), 2)
                    ws.write(row, OFF+i, v if v else None, fn)
                    total_b2 += v
                    gran_cli[i] += v
                ws.write(row, COL_TOT1, round(total_b2, 2), ft)
                gran_tot1 += total_b2

                prod_vals = [
                    prod_day.get((fecha,"GS"), 0.0),
                    prod_day.get((fecha,"GP"), 0.0),
                    prod_day.get((fecha,"GD"), 0.0),
                    iva_day.get(fecha, 0.0),
                    ieps_prod.get((fecha,"GS"), 0.0),
                    ieps_prod.get((fecha,"GP"), 0.0),
                    ieps_prod.get((fecha,"GD"), 0.0),
                ]
                total_prod = sum(prod_vals)
                for i, v in enumerate(prod_vals):
                    ws.write(row, COL_PROD0+i, round(v,2) if v else None, fn)
                    gran_prod[i] += v
                ws.write(row, COL_TOT2, round(total_prod, 2), ft)
                gran_tot2 += total_prod

                diferencia = round(total_b2 - total_prod, 2)
                ws.write(row, COL_CONC, diferencia, fc)
                gran_conc += diferencia

            tr = len(fechas) + 4
            ws.merge_range(tr, 0, tr, N_META-1, "TOTAL GENERAL", f_grand_l)
            for i in range(N_CLI): ws.write(tr, OFF+i, round(gran_cli[i], 2), f_grand)
            ws.write(tr, COL_TOT1, round(gran_tot1, 2), f_grand)
            for i in range(N_PROD): ws.write(tr, COL_PROD0+i, round(gran_prod[i], 2), f_grand)
            ws.write(tr, COL_TOT2, round(gran_tot2, 2), f_grand)
            ws.write(tr, COL_CONC, round(gran_conc, 2), f_grand_c)

            wb.close()
            self._vd_resultado_path = OUT
            nombre = os.path.basename(OUT)
            self.after(0, self._pb_detener, self._vd_pb, self._vd_pb_lbl)
            self.after(0, self._vd_pb_frame.grid_remove)
            self.after(0, self._vd_lbl_archivo.config, {"text": nombre})
            self.after(0, self._log, f"â PÃ³liza generada: {nombre}", True)
            self.after(0, self._vd_abrir_en_visor, OUT, carpeta_sal)
            self.after(0, self._vd_cargar_visor_inline, OUT)

        except Exception as e:
            import traceback as _tb2
            self.after(0, self._pb_error, self._vd_pb, self._vd_pb_lbl)
            self.after(0, self._vd_pb_frame.grid_remove)
            self.after(0, self._log, f"â Error al generar pÃ³liza: {e}", True)
            self.after(0, self._vd_lbl_archivo.config, {"text": "Error"})
            self.after(0, messagebox.showerror, "Error al generar pÃ³liza",
                f"{e}\n\n{_tb2.format_exc()[-500:]}")


    # âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    # MÃDULO: VENTAS DEL DÃA (grupo: PÃ³liza + ConciliaciÃ³n)
    # âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    def _tab_ventas_grupo(self, nb):
        """PestaÃ±a contenedora: une Ventas del DÃ­a y Control de Despacho vs Ventas."""
        outer = ttk.Frame(nb)
        nb.add(outer, text="  â½ Ventas del DÃ­a  ")
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(0, weight=1)

        inner_nb = ttk.Notebook(outer)
        inner_nb.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)

        self._tab_ventas_dia(inner_nb)   # sub-pestaÃ±a: Generar PÃ³liza (sin modificar)
        self._tab_vd_conc(inner_nb)      # sub-pestaÃ±a: Control de Despacho vs Ventas

    # âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    # MÃDULO: CONTROL DE DESPACHO vs VENTAS DEL DÃA
    # âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    def _tab_vd_conc(self, nb):
        import os, threading
        from tkinter import filedialog, messagebox as mb

        tab = ttk.Frame(nb)
        nb.add(tab, text="  ð Despachos vs Ventas  ")
        tab.columnconfigure(0, weight=1)

        # ââ Encabezado ââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
        hdr = tk.Frame(tab, bg="#B45309", height=40)
        hdr.grid(row=0, column=0, sticky="ew")
        tk.Label(hdr, text="ð  CONTROL DE DESPACHO vs VENTAS DEL DÃA",
                 bg="#B45309", fg="white",
                 font=("Segoe UI", 11, "bold")).pack(side="left", padx=14, pady=8)

        # ââ SelecciÃ³n de archivos ââââââââââââââââââââââââââââââââââââââââââââââ
        card = ttk.Frame(tab, style="Tarjeta.TFrame")
        card.grid(row=1, column=0, sticky="ew", padx=8, pady=(8, 4))
        card.columnconfigure(1, weight=1)

        self._vdc_despachos = tk.StringVar(value="")
        self._vdc_ventas    = tk.StringVar(value="")
        self._vdc_resultado_path = None

        ttk.Label(card, text="ð Control de Despachos (.xlsx):", width=32,
                  style="Tarjeta.TLabel").grid(row=0, column=0, sticky="w", padx=10, pady=(10, 4))
        ttk.Entry(card, textvariable=self._vdc_despachos).grid(
            row=0, column=1, sticky="ew", padx=6, pady=(10, 4))
        ttk.Button(card, text="Elegir...",
                   command=lambda: self._vdc_elegir_archivo(self._vdc_despachos)).grid(
            row=0, column=2, padx=(0, 10), pady=(10, 4))

        ttk.Label(card, text="ð Ventas del DÃ­a (.xlsx):", width=32,
                  style="Tarjeta.TLabel").grid(row=1, column=0, sticky="w", padx=10, pady=(4, 10))
        ttk.Entry(card, textvariable=self._vdc_ventas).grid(
            row=1, column=1, sticky="ew", padx=6, pady=(4, 10))
        ttk.Button(card, text="Elegir...",
                   command=lambda: self._vdc_elegir_archivo(self._vdc_ventas)).grid(
            row=1, column=2, padx=(0, 10), pady=(4, 10))

        # ââ BotÃ³n generar âââââââââââââââââââââââââââââââââââââââââââââââââââââ
        ttk.Button(tab, text="â¡  Generar ConciliaciÃ³n",
                   command=self._vdc_generar,
                   style="Accent.TButton").grid(row=2, column=0, pady=6)

        # ââ Barra de progreso âââââââââââââââââââââââââââââââââââââââââââââââââ
        _vdc_pb_frame = tk.Frame(tab, bg=COLOR_FONDO)
        _vdc_pb_frame.grid(row=3, column=0, sticky="ew", padx=8, pady=(2, 2))
        _vdc_pb_frame.grid_remove()
        self._vdc_pb_frame = _vdc_pb_frame
        self._vdc_pb = FunkyProgressBar(_vdc_pb_frame, maximum=100, height=70)
        self._vdc_pb.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self._vdc_pb_lbl = ttk.Label(_vdc_pb_frame, text="0 %", width=8,
                                     style="Tarjeta.TLabel")
        self._vdc_pb_lbl.pack(side="left")

        # ââ Barra inferior resultado âââââââââââââââââââââââââââââââââââââââââââ
        self._vdc_barra = ttk.Frame(tab, style="Tarjeta.TFrame")
        self._vdc_barra.grid(row=4, column=0, sticky="ew", padx=8, pady=(0, 6))

        ttk.Label(self._vdc_barra, text="ð Resultado:", style="Tarjeta.TLabel").pack(
            side="left", padx=(10, 4), pady=8)
        self._vdc_lbl_archivo = ttk.Label(self._vdc_barra, text="â",
                                          style="Tarjeta.TLabel", foreground="#6B7280")
        self._vdc_lbl_archivo.pack(side="left", padx=4, pady=8)
        ttk.Button(self._vdc_barra, text="ð  Abrir",
                   command=self._vdc_abrir).pack(side="right", padx=4, pady=6)
        ttk.Button(self._vdc_barra, text="ð¾  Guardar como...",
                   command=self._vdc_guardar_como).pack(side="right", padx=4, pady=6)

        # ââ Vista previa (Treeview) ââââââââââââââââââââââââââââââââââââââââââââ
        tab.rowconfigure(5, weight=1)
        visor_wrap = ttk.Frame(tab, style="Tarjeta.TFrame")
        visor_wrap.grid(row=5, column=0, sticky="nsew", padx=8, pady=(0, 8))
        visor_wrap.columnconfigure(0, weight=1)
        visor_wrap.rowconfigure(1, weight=1)

        ttk.Label(visor_wrap, text="ð  Vista previa del resultado",
                  style="Tarjeta.TLabel",
                  font=("Segoe UI", 9, "bold")).grid(
            row=0, column=0, sticky="w", padx=10, pady=(6, 2))

        tv_frame = tk.Frame(visor_wrap, bg=COLOR_FONDO)
        tv_frame.grid(row=1, column=0, sticky="nsew", padx=6, pady=(0, 6))
        tv_frame.columnconfigure(0, weight=1)
        tv_frame.rowconfigure(0, weight=1)

        self._vdc_tv = ttk.Treeview(tv_frame, show="headings", height=6)
        vsb = ttk.Scrollbar(tv_frame, orient="vertical",   command=self._vdc_tv.yview)
        hsb = ttk.Scrollbar(tv_frame, orient="horizontal", command=self._vdc_tv.xview)
        self._vdc_tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._vdc_tv.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

    def _vdc_cargar_visor(self, path):
        tv = getattr(self, "_vdc_tv", None)
        if tv is None:
            return
        for item in tv.get_children():
            tv.delete(item)
        if not path or not __import__("os").path.isfile(path):
            return
        try:
            import openpyxl as _opxl
            wb = _opxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not all_rows:
                return
            headers = [str(c) if c is not None else "" for c in all_rows[0]]
            tv["columns"] = headers
            tv["show"] = "headings"
            for h in headers:
                tv.heading(h, text=h)
                tv.column(h, width=140, anchor="w", stretch=True)
            for row in all_rows[1:51]:
                vals = [str(c) if c is not None else "" for c in row]
                tv.insert("", "end", values=vals)
            self._log(f"  Vista previa: {min(len(all_rows)-1, 50)} fila(s).")
        except Exception as _e:
            self._log(f"  â  Vista previa no disponible: {_e}", True)

    def _vdc_elegir_archivo(self, var):
        from tkinter import filedialog
        ruta = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")]
        )
        if ruta:
            var.set(ruta)

    def _vdc_abrir(self):
        import os
        if self._vdc_resultado_path and os.path.exists(self._vdc_resultado_path):
            os.startfile(self._vdc_resultado_path)
        else:
            from tkinter import messagebox
            messagebox.showwarning("Sin resultado", "Primero genera la conciliaciÃ³n.")

    def _vdc_guardar_como(self):
        import os, shutil
        from tkinter import filedialog, messagebox
        if not self._vdc_resultado_path or not os.path.exists(self._vdc_resultado_path):
            messagebox.showwarning("Sin resultado", "Primero genera la conciliaciÃ³n.")
            return
        destino = filedialog.asksaveasfilename(
            title="Guardar conciliaciÃ³n como...",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=os.path.basename(self._vdc_resultado_path),
        )
        if destino:
            shutil.copy2(self._vdc_resultado_path, destino)
            self._log(f"ð¾ Guardado en: {destino}")

    def _vdc_generar(self):
        from tkinter import messagebox
        desp = self._vdc_despachos.get().strip()
        vd   = self._vdc_ventas.get().strip()
        if not desp or not vd:
            messagebox.showwarning("Faltan archivos",
                "Selecciona el Control de Despachos y el archivo de Ventas del DÃ­a.")
            return
        self._vdc_lbl_archivo.config(text="Generando...")
        self._vdc_pb_frame.grid()
        self._pb_iniciar(self._vdc_pb, self._vdc_pb_lbl)
        import threading
        threading.Thread(target=self._vdc_generar_hilo, args=(desp, vd), daemon=True).start()

    def _vdc_generar_hilo(self, desp_path, vd_path):
        import os, tempfile
        from datetime import datetime
        from collections import defaultdict
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            self.after(0, self._pb_error, self._vdc_pb, self._vdc_pb_lbl)
            self.after(0, self._vdc_pb_frame.grid_remove)
            self.after(0, self._log, f"â Falta librerÃ­a: {e}", True)
            return

        try:
            self.after(0, self._log, "ð Leyendo Control de Despachos...")
            wb1 = openpyxl.load_workbook(desp_path, read_only=True, data_only=True)
            hoja1 = wb1.sheetnames[0]
            rows1 = list(wb1[hoja1].iter_rows(values_only=True))
            wb1.close()

            self.after(0, self._log, "ð Leyendo Ventas del DÃ­a...")
            wb2 = openpyxl.load_workbook(vd_path, read_only=True, data_only=True)
            hoja2 = wb2.sheetnames[0]
            rows2 = list(wb2[hoja2].iter_rows(values_only=True))
            wb2.close()
        except Exception as e:
            self.after(0, self._pb_error, self._vdc_pb, self._vdc_pb_lbl)
            self.after(0, self._vdc_pb_frame.grid_remove)
            self.after(0, self._log, f"â Error al leer archivos: {e}", True)
            return

        try:
            # ââ Procesar Despachos âââââââââââââââââââââââââââââââââââââââââââââ
            dia = defaultdict(lambda: {"GS_sub":0.0,"GP_sub":0.0,"GD_sub":0.0,
                                       "iva":0.0,"ieps":0.0,"importe":0.0,"uuids":set()})
            data = defaultdict(lambda: {"nombre":"","fechas":set(),
                "GS":{"sub":0.0,"iva":0.0,"ieps":0.0,"tot":0.0},
                "GP":{"sub":0.0,"iva":0.0,"ieps":0.0,"tot":0.0},
                "GD":{"sub":0.0,"iva":0.0,"ieps":0.0,"tot":0.0},
            })
            sin_uuid = defaultdict(lambda: {
                "GS":{"sub":0.0,"iva":0.0,"ieps":0.0,"tot":0.0,"n":0},
                "GP":{"sub":0.0,"iva":0.0,"ieps":0.0,"tot":0.0,"n":0},
                "GD":{"sub":0.0,"iva":0.0,"ieps":0.0,"tot":0.0,"n":0},
            })

            for r in rows1[1:]:
                if not r or r[0] is None: continue
                try: dt = datetime.strptime(str(r[0])[:10], "%Y-%m-%d").date()
                except: continue
                prod = str(r[3]).strip().upper() if r[3] else ""
                if prod not in ("GS","GP","GD"): continue
                sub=float(r[6] or 0); iva=float(r[7] or 0)
                ieps=float(r[8] or 0); imp=float(r[9] or 0)
                cliente=str(r[16]).strip() if r[16] else "SIN CLIENTE"
                uuid=str(r[22]).strip() if r[22] else "-----"
                rfc=str(r[23]).strip() if r[23] else "SIN-RFC"
                nombre=str(r[24]).strip() if r[24] else ""
                d=dia[dt]
                d[f"{prod}_sub"]+=sub; d["iva"]+=iva
                d["ieps"]+=ieps; d["importe"]+=imp; d["uuids"].add(uuid)
                if uuid == "-----":
                    s=sin_uuid[(cliente,dt)][prod]
                    s["sub"]+=sub; s["iva"]+=iva; s["ieps"]+=ieps; s["tot"]+=imp; s["n"]+=1
                else:
                    k=(rfc,uuid); e=data[k]
                    e["nombre"]=nombre; e["fechas"].add(dt)
                    e[prod]["sub"]+=sub; e[prod]["iva"]+=iva
                    e[prod]["ieps"]+=ieps; e[prod]["tot"]+=imp

            # ââ Procesar Ventas del DÃ­a ââââââââââââââââââââââââââââââââââââââââ
            venta = {}
            for r in rows2[3:]:
                if not r or r[1] is None: continue
                fecha = r[1]
                if isinstance(fecha, datetime): fecha = fecha.date()
                else: continue
                venta[fecha] = {
                    "GS":float(r[29] or 0),"GP":float(r[30] or 0),"GD":float(r[31] or 0),
                    "IVA":float(r[32] or 0),"IEPS":float(r[33] or 0),"TOTAL":float(r[35] or 0)
                }

            fechas = sorted(dia.keys())
            self.after(0, self._log, f"â {len(fechas)} dÃ­as encontrados en Despachos.")

            # ââ Estilos ââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
            MF = '#,##0.00'
            F_WHITE = PatternFill("solid", fgColor="FFFFFF")
            def fill(c): return PatternFill("solid", fgColor=c)
            thin = Side(style="thin", color="CCCCCC")
            bord = Border(left=thin, right=thin, top=thin, bottom=thin)
            al_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
            al_r = Alignment(horizontal="right",  vertical="center")
            al_l = Alignment(horizontal="left",   vertical="center")

            F_HDR=fill("1E3A8A"); F_GS=fill("DBEAFE"); F_GP=fill("D1FAE5"); F_GD=fill("FEF9C3")
            F_TOT=fill("FFC000"); F_OK=fill("C6EFCE"); F_WRN=fill("FFEB9C"); F_ERR=fill("FFC7CE")
            F_IVA=fill("FCE7F3"); F_UUID=fill("F1F5F9"); F_RFC=fill("E0F2FE")
            F_DARK=fill("111827"); F_GRAY=fill("374151")
            F_SIN=fill("FFF7ED"); F_SIN_H=fill("EA580C")
            F_SIN_TOT=fill("F97316"); F_SIN_ROW=fill("FFEDD5")
            F_TITULO=fill("0F172A")

            def cw(ws, row, col, val=None, fi=None, al=None, fmt=None, bold=False, sz=9, fc="000000"):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = fi if fi is not None else F_WHITE
                c.font = Font(color=fc, bold=bold, size=sz)
                if al:  c.alignment = al
                if fmt: c.number_format = fmt
                c.border = bord
                return c

            def mw(ws, r, c1, c2, val, fi, bold=False, sz=9, fc="000000", al=None):
                ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
                cell = ws.cell(r, c1, val)
                cell.fill = fi; cell.font = Font(color=fc, bold=bold, size=sz)
                cell.border = bord; cell.alignment = al or al_l
                for ci in range(c1+1, c2+1):
                    ws.cell(r, ci).fill = fi; ws.cell(r, ci).border = bord

            wb = openpyxl.Workbook()

            # ââ HOJA 1 â CONCILIACIÃN DIARIA ââââââââââââââââââââââââââââââââââ
            ws = wb.active; ws.title = "ConciliaciÃ³n"
            mw(ws,1,1,18,"CONTROL DE DESPACHO vs VENTAS DEL DÃA",F_HDR,bold=True,sz=12,fc="FFFFFF",al=al_c)
            ws.row_dimensions[1].height = 26
            for s,e,txt,fi in [(1,1,"FECHA",F_HDR),(2,8,"CONTROL DE DESPACHOS",F_DARK),
                                (9,13,"VENTAS DEL DÃA",F_GRAY),(14,18,"DIFERENCIAS",F_HDR)]:
                if s<e: mw(ws,2,s,e,txt,fi,bold=True,sz=9,fc="FFFFFF",al=al_c)
                else:   cw(ws,2,s,val=txt,fi=fi,al=al_c,bold=True,sz=9,fc="FFFFFF")
            ws.row_dimensions[2].height = 18
            cols3 = [
                ("Fecha",12,F_HDR,"FFFFFF"),
                ("G-Super\nSubtotal",14,F_GS,"1E3A8A"),("G-Premium\nSubtotal",14,F_GP,"166534"),
                ("G-Diesel\nSubtotal",14,F_GD,"92400E"),
                ("IVA\nUnificado",14,F_IVA,"9D174D"),("IEPS\nUnificado",14,F_IVA,"9D174D"),
                ("TOTAL\nImporte",15,F_DARK,"FFFFFF"),("UUIDs\n#",9,F_UUID,"374151"),
                ("GS\n(VtaDÃ­a)",13,F_GS,"1E3A8A"),("GP\n(VtaDÃ­a)",13,F_GP,"166534"),
                ("GD\n(VtaDÃ­a)",13,F_GD,"92400E"),
                ("IVA\n(VtaDÃ­a)",13,F_IVA,"9D174D"),("IEPS\n(VtaDÃ­a)",13,F_IVA,"9D174D"),
                ("Dif\nGS",12,F_HDR,"FFFFFF"),("Dif\nGP",12,F_HDR,"FFFFFF"),
                ("Dif\nGD",12,F_HDR,"FFFFFF"),("Dif\nIVA",12,F_HDR,"FFFFFF"),
                ("DIF\nTOTAL",13,F_HDR,"FFFFFF"),
            ]
            for i,(h,w,fi,fc) in enumerate(cols3,1):
                c=ws.cell(3,i,h); c.fill=fi; c.font=Font(color=fc,bold=True,size=8)
                c.alignment=al_c; c.border=bord
                ws.column_dimensions[get_column_letter(i)].width=w
            ws.row_dimensions[3].height = 34

            ROW = 4; tots = defaultdict(float)
            total_dias = len(fechas)
            for idx, fecha in enumerate(fechas):
                self.after(0, self._pb_step, self._vdc_pb, self._vdc_pb_lbl)
                d=dia[fecha]; v=venta.get(fecha,{})
                gs_s=d["GS_sub"]; gp_s=d["GP_sub"]; gd_s=d["GD_sub"]
                iva_u=d["iva"]; ieps_u=d["ieps"]; tot=d["importe"]
                v_gs=v.get("GS",0); v_gp=v.get("GP",0); v_gd=v.get("GD",0)
                v_iva=v.get("IVA",0); v_ieps=v.get("IEPS",0); v_tot=v.get("TOTAL",0)
                dif_gs=round(gs_s-v_gs,2); dif_gp=round(gp_s-v_gp,2)
                dif_gd=round(gd_s-v_gd,2); dif_iva=round(iva_u-v_iva,2)
                dif_tot=round(tot-v_tot,2)
                fi_d=F_OK if abs(dif_tot)<1 else F_WRN if abs(dif_tot)<1000 else F_ERR
                vals_row = [
                    (fecha.strftime("%d/%m/%Y"),None),(gs_s,F_GS),(gp_s,F_GP),(gd_s,F_GD),
                    (iva_u,F_IVA),(ieps_u,F_IVA),(tot,F_DARK),(len(d["uuids"]),F_UUID),
                    (v_gs,F_GS),(v_gp,F_GP),(v_gd,F_GD),(v_iva,F_IVA),(v_ieps,F_IVA),
                    (dif_gs,fi_d),(dif_gp,fi_d),(dif_gd,fi_d),(dif_iva,fi_d),(dif_tot,fi_d),
                ]
                fcs_r = ["000000","1E3A8A","166534","92400E","9D174D","9D174D","FFFFFF","374151",
                         "1E3A8A","166534","92400E","9D174D","9D174D",
                         "000000","000000","000000","000000","000000"]
                fmts_r = [None,MF,MF,MF,MF,MF,MF,"#,##0",MF,MF,MF,MF,MF,MF,MF,MF,MF,MF]
                for i,((val,fi),fc,fmt) in enumerate(zip(vals_row,fcs_r,fmts_r),1):
                    cw(ws,ROW,i,val=val,fi=fi,al=al_c if i in(1,8) else al_r,fmt=fmt,sz=9,fc=fc)
                for k,kv in [("gs",gs_s),("gp",gp_s),("gd",gd_s),("iva",iva_u),
                              ("ieps",ieps_u),("tot",tot),("vgs",v_gs),("vgp",v_gp),
                              ("vgd",v_gd),("viva",v_iva),("vieps",v_ieps),("vtot",v_tot)]:
                    tots[k]+=kv
                ws.row_dimensions[ROW].height=15; ROW+=1

            tv = ["TOTAL",tots["gs"],tots["gp"],tots["gd"],tots["iva"],tots["ieps"],tots["tot"],
                  sum(len(dia[f]["uuids"]) for f in fechas),
                  tots["vgs"],tots["vgp"],tots["vgd"],tots["viva"],tots["vieps"],
                  round(tots["gs"]-tots["vgs"],2),round(tots["gp"]-tots["vgp"],2),
                  round(tots["gd"]-tots["vgd"],2),round(tots["iva"]-tots["viva"],2),
                  round(tots["tot"]-tots["vtot"],2)]
            for i,v in enumerate(tv,1):
                c=ws.cell(ROW,i,v); c.fill=F_TOT; c.font=Font(bold=True,size=9)
                c.border=bord; c.alignment=al_l if i==1 else al_r
                if 2<=i<=18: c.number_format=MF if i!=8 else "#,##0"
            ws.row_dimensions[ROW].height=22; ws.freeze_panes="B4"

            # ââ HOJA 2 â DETALLE UUID ââââââââââââââââââââââââââââââââââââââââââ
            ws2 = wb.create_sheet("Detalle UUID")
            for i,w in enumerate([20,30,42,14,15,13,13,15],1):
                ws2.column_dimensions[get_column_letter(i)].width=w

            colores_alt = [fill("EFF6FF"),fill("F0FFFE"),fill("FFFBEB"),fill("FDF4FF"),fill("FFF1F2")]
            PRODS = [
                ("GS","G-SUPER",   fill("1D4ED8"),F_GS,"1D4ED8"),
                ("GP","G-PREMIUM", fill("15803D"),F_GP,"14532D"),
                ("GD","G-DIESEL",  fill("92400E"),F_GD,"451A03"),
            ]
            CUR = 1
            grand_total = {"GS":defaultdict(float),"GP":defaultdict(float),"GD":defaultdict(float)}

            mw(ws2,CUR,1,8,"TABLA DE UUIDs",F_TITULO,bold=True,sz=14,fc="FFFFFF",al=al_c)
            ws2.row_dimensions[CUR].height=34; CUR+=1

            self.after(0, self._pb_step, self._vdc_pb, self._vdc_pb_lbl)

            for prod,label,fi_hdr,fi_prod,fc_txt in PRODS:
                mw(ws2,CUR,1,8,f"  â TABLA {label}",fi_hdr,bold=True,sz=12,fc="FFFFFF",al=al_l)
                ws2.row_dimensions[CUR].height=28; CUR+=1

                claves_sin = sorted(
                    [(cli,dt) for (cli,dt) in sin_uuid if sin_uuid[(cli,dt)][prod]["n"]>0],
                    key=lambda x:(x[0],x[1])
                )
                if claves_sin:
                    mw(ws2,CUR,1,8,"  â   SIN COMPROBANTE FISCAL (UUID = -----)",
                       F_SIN_H,bold=True,sz=9,fc="FFFFFF",al=al_l)
                    ws2.row_dimensions[CUR].height=20; CUR+=1
                    for i,h in enumerate(["Cliente (col Q)","â","â SIN UUID â",
                                          "Fecha","Subtotal","IVA","IEPS","TOTAL"],1):
                        c=ws2.cell(CUR,i,h); c.fill=F_SIN
                        c.font=Font(color="92400E",bold=True,size=8)
                        c.alignment=al_c; c.border=bord
                    ws2.row_dimensions[CUR].height=16; CUR+=1
                    prev_cli=None; cli_tots=defaultdict(float); sin_grand=defaultdict(float)
                    for (cli,dt) in claves_sin:
                        if prev_cli is not None and cli!=prev_cli:
                            mw(ws2,CUR,1,3,f"  Subtotal  {prev_cli}",F_RFC,bold=True,sz=8,fc="7C2D12",al=al_l)
                            cw(ws2,CUR,4,fi=F_RFC)
                            for ci,kk in enumerate(["sub","iva","ieps","tot"],5):
                                cw(ws2,CUR,ci,val=cli_tots[kk],fi=F_RFC,al=al_r,fmt=MF,bold=True,sz=8,fc="7C2D12")
                            for k in ["sub","iva","ieps","tot"]: cli_tots[k]=0.0
                            ws2.row_dimensions[CUR].height=14; CUR+=1
                        prev_cli=cli; s=sin_uuid[(cli,dt)][prod]
                        cw(ws2,CUR,1,val=cli,fi=F_SIN_ROW,al=al_l,sz=8,fc="7C2D12")
                        cw(ws2,CUR,2,val="â SIN NOMBRE â",fi=F_SIN_ROW,al=al_l,sz=8,fc="9A3412")
                        cw(ws2,CUR,3,val="â SIN COMPROBANTE FISCAL â",fi=F_SIN_ROW,al=al_c,sz=8,fc="9A3412")
                        cw(ws2,CUR,4,val=dt.strftime("%d/%m/%Y"),fi=F_SIN_ROW,al=al_c,sz=8)
                        cw(ws2,CUR,5,val=s["sub"],fi=F_SIN_ROW,al=al_r,fmt=MF,sz=8)
                        cw(ws2,CUR,6,val=s["iva"],fi=F_SIN_ROW,al=al_r,fmt=MF,sz=8)
                        cw(ws2,CUR,7,val=s["ieps"],fi=F_SIN_ROW,al=al_r,fmt=MF,sz=8)
                        cw(ws2,CUR,8,val=s["tot"],fi=F_SIN_ROW,al=al_r,fmt=MF,sz=8)
                        for k,kv in [("sub",s["sub"]),("iva",s["iva"]),("ieps",s["ieps"]),("tot",s["tot"])]:
                            cli_tots[k]+=kv; sin_grand[k]+=kv
                        ws2.row_dimensions[CUR].height=13; CUR+=1
                    if prev_cli:
                        mw(ws2,CUR,1,3,f"  Subtotal  {prev_cli}",F_RFC,bold=True,sz=8,fc="7C2D12",al=al_l)
                        cw(ws2,CUR,4,fi=F_RFC)
                        for ci,kk in enumerate(["sub","iva","ieps","tot"],5):
                            cw(ws2,CUR,ci,val=cli_tots[kk],fi=F_RFC,al=al_r,fmt=MF,bold=True,sz=8,fc="7C2D12")
                        ws2.row_dimensions[CUR].height=14; CUR+=1
                    mw(ws2,CUR,1,4,f"  TOTAL SIN COMPROBANTE â {label}",F_SIN_TOT,bold=True,sz=9,fc="FFFFFF",al=al_l)
                    for ci,kk in enumerate(["sub","iva","ieps","tot"],5):
                        cw(ws2,CUR,ci,val=sin_grand[kk],fi=F_SIN_TOT,al=al_r,fmt=MF,bold=True,sz=9,fc="FFFFFF")
                    ws2.row_dimensions[CUR].height=18; CUR+=1
                    gt=grand_total[prod]
                    for k in ["sub","iva","ieps","tot"]: gt[f"sin_{k}"]+=sin_grand[k]

                mw(ws2,CUR,1,8,"  CON COMPROBANTE FISCAL (CFDI / UUID)",F_HDR,bold=True,sz=9,fc="FFFFFF",al=al_l)
                ws2.row_dimensions[CUR].height=16; CUR+=1
                for i,h in enumerate(["RFC","Nombre / RazÃ³n Social","UUID / Folio Fiscal",
                                       "Fechas","Subtotal","IVA","IEPS","TOTAL"],1):
                    c=ws2.cell(CUR,i,h); c.fill=fi_prod
                    c.font=Font(color=fc_txt,bold=True,size=9); c.alignment=al_c; c.border=bord
                ws2.row_dimensions[CUR].height=16; CUR+=1

                claves = sorted([(rfc,uuid) for (rfc,uuid),e in data.items() if e[prod]["tot"]>0],
                                 key=lambda x:(x[0],x[1]))
                rfc_col={}; cidx=0; prev_rfc=None; rfc_tots=defaultdict(float); gt=grand_total[prod]
                for (rfc,uuid) in claves:
                    e=data[(rfc,uuid)]
                    if rfc not in rfc_col: rfc_col[rfc]=colores_alt[cidx % len(colores_alt)]; cidx+=1
                    if prev_rfc is not None and rfc!=prev_rfc:
                        mw(ws2,CUR,1,4,f"  Subtotal  {prev_rfc}",F_RFC,bold=True,sz=8,fc="1E40AF",al=al_l)
                        for ci,kk in enumerate(["sub","iva","ieps","tot"],5):
                            cw(ws2,CUR,ci,val=rfc_tots[kk],fi=F_RFC,al=al_r,fmt=MF,bold=True,sz=8,fc="1E40AF")
                        for k in ["sub","iva","ieps","tot"]: rfc_tots[k]=0.0
                        ws2.row_dimensions[CUR].height=14; CUR+=1
                    prev_rfc=rfc; fi_r=rfc_col[rfc]
                    sub=e[prod]["sub"]; iva=e[prod]["iva"]; ieps=e[prod]["ieps"]; tot_v=e[prod]["tot"]
                    fechas_str=", ".join(sorted(f.strftime("%d/%m") for f in e["fechas"]))
                    for i,val in enumerate([rfc,e["nombre"],uuid,fechas_str,sub,iva,ieps,tot_v],1):
                        c=ws2.cell(CUR,i,val); c.fill=fi_r; c.font=Font(size=8)
                        c.border=bord; c.alignment=al_l if i<=4 else al_r
                        if i>4: c.number_format=MF
                    for k,kv in [("sub",sub),("iva",iva),("ieps",ieps),("tot",tot_v)]:
                        rfc_tots[k]+=kv; gt[k]+=kv
                    ws2.row_dimensions[CUR].height=13; CUR+=1
                if prev_rfc:
                    mw(ws2,CUR,1,4,f"  Subtotal  {prev_rfc}",F_RFC,bold=True,sz=8,fc="1E40AF",al=al_l)
                    for ci,kk in enumerate(["sub","iva","ieps","tot"],5):
                        cw(ws2,CUR,ci,val=rfc_tots[kk],fi=F_RFC,al=al_r,fmt=MF,bold=True,sz=8,fc="1E40AF")
                    ws2.row_dimensions[CUR].height=14; CUR+=1
                mw(ws2,CUR,1,4,f"  TOTAL CON COMPROBANTE â {label}",fill("166534"),bold=True,sz=9,fc="FFFFFF",al=al_l)
                for ci,kk in enumerate(["sub","iva","ieps","tot"],5):
                    cw(ws2,CUR,ci,val=gt.get(kk,0),fi=fill("166534"),al=al_r,fmt=MF,bold=True,sz=9,fc="FFFFFF")
                ws2.row_dimensions[CUR].height=18; CUR+=1
                gran_sub=gt.get("sub",0)+gt.get("sin_sub",0)
                gran_iva=gt.get("iva",0)+gt.get("sin_iva",0)
                gran_ieps=gt.get("ieps",0)+gt.get("sin_ieps",0)
                gran_tot=gt.get("tot",0)+gt.get("sin_tot",0)
                mw(ws2,CUR,1,4,f"  â GRAN TOTAL {label}",F_TOT,bold=True,sz=11,al=al_l)
                for ci,v in enumerate([gran_sub,gran_iva,gran_ieps,gran_tot],5):
                    cw(ws2,CUR,ci,val=v,fi=F_TOT,al=al_r,fmt=MF,bold=True,sz=11)
                ws2.row_dimensions[CUR].height=26; CUR+=2

            ws2.freeze_panes="A2"

            # ââ HOJA 3 â RESUMEN ââââââââââââââââââââââââââââââââââââââââââââââ
            ws3 = wb.create_sheet("Resumen")
            for i,w in enumerate([30,18,18,18,14,16],1):
                ws3.column_dimensions[get_column_letter(i)].width=w
            mw(ws3,1,1,6,"RESUMEN â CONTROL DE DESPACHO vs VENTAS DEL DÃA",F_HDR,bold=True,sz=13,fc="FFFFFF",al=al_c)
            ws3.row_dimensions[1].height=32

            tot_gs=tots["gs"]; tot_gp=tots["gp"]; tot_gd=tots["gd"]
            tot_iva=tots["iva"]; tot_ieps=tots["ieps"]; tot_imp=tots["tot"]
            v_tot_gs=tots["vgs"]; v_tot_gp=tots["vgp"]; v_tot_gd=tots["vgd"]
            v_tot_iva=tots["viva"]; v_tot_ieps=tots["vieps"]; v_tot_total=tots["vtot"]
            uuid_gs=len({(r,u) for (r,u),e in data.items() if e["GS"]["tot"]>0})
            uuid_gp=len({(r,u) for (r,u),e in data.items() if e["GP"]["tot"]>0})
            uuid_gd=len({(r,u) for (r,u),e in data.items() if e["GD"]["tot"]>0})
            uuid_total=len(data); rfc_total=len({r for (r,u) in data.keys()})
            TTOT=max(tot_gs+tot_gp+tot_gd,1)

            sin_resumen = defaultdict(lambda:{"GS":0.0,"GP":0.0,"GD":0.0,"n":0})
            for (cli,dt),pd in sin_uuid.items():
                for pr in ("GS","GP","GD"):
                    sin_resumen[cli][pr]+=pd[pr]["tot"]; sin_resumen[cli]["n"]+=pd[pr]["n"]

            R=3
            mw(ws3,R,1,6,"VENTAS POR PRODUCTO",F_DARK,bold=True,sz=10,fc="FFFFFF",al=al_l)
            ws3.row_dimensions[R].height=22; R+=1
            for i,h in enumerate(["Producto","Subtotal Despachos","Ventas del DÃ­a","Diferencia","% Total","CFDIs"],1):
                cw(ws3,R,i,val=h,fi=F_HDR,al=al_c,bold=True,sz=8,fc="FFFFFF")
            ws3.row_dimensions[R].height=16; R+=1
            for lbl,desp,pol,nuuid,fi_r in [
                    ("G-Super (GS)",tot_gs,v_tot_gs,uuid_gs,F_GS),
                    ("G-Premium (GP)",tot_gp,v_tot_gp,uuid_gp,F_GP),
                    ("G-Diesel (GD)",tot_gd,v_tot_gd,uuid_gd,F_GD)]:
                dif=round(desp-pol,2); fi_dif=F_OK if abs(dif)<1 else F_WRN if abs(dif)<500 else F_ERR
                cw(ws3,R,1,val=lbl,fi=fi_r,al=al_l,sz=9)
                cw(ws3,R,2,val=desp,fi=fi_r,al=al_r,fmt=MF,sz=9)
                cw(ws3,R,3,val=pol,fi=fi_r,al=al_r,fmt=MF,sz=9)
                cw(ws3,R,4,val=dif,fi=fi_dif,al=al_r,fmt=MF,bold=True,sz=9)
                cw(ws3,R,5,val=desp/TTOT,fi=fi_r,al=al_r,fmt="0.0%",sz=9)
                cw(ws3,R,6,val=nuuid,fi=fi_r,al=al_r,fmt="#,##0",sz=9)
                ws3.row_dimensions[R].height=16; R+=1
            TTOT_r=tot_gs+tot_gp+tot_gd; VTOT_r=v_tot_gs+v_tot_gp+v_tot_gd
            dif_sub=round(TTOT_r-VTOT_r,2)
            fi_dt=F_OK if abs(dif_sub)<1 else F_WRN if abs(dif_sub)<500 else F_ERR
            cw(ws3,R,1,val="TOTAL",fi=F_TOT,al=al_l,bold=True,sz=10)
            cw(ws3,R,2,val=TTOT_r,fi=F_TOT,al=al_r,fmt=MF,bold=True,sz=10)
            cw(ws3,R,3,val=VTOT_r,fi=F_TOT,al=al_r,fmt=MF,bold=True,sz=10)
            cw(ws3,R,4,val=dif_sub,fi=fi_dt,al=al_r,fmt=MF,bold=True,sz=10)
            cw(ws3,R,5,val=1.0,fi=F_TOT,al=al_r,fmt="0.0%",bold=True,sz=10)
            cw(ws3,R,6,val=uuid_total,fi=F_TOT,al=al_r,fmt="#,##0",bold=True,sz=10)
            ws3.row_dimensions[R].height=22; R+=2

            sin_tot_gs=sin_tot_gp=sin_tot_gd=sin_tot_n=0
            mw(ws3,R,1,6,"â   SIN COMPROBANTE FISCAL â POR CLIENTE (col Q)",F_SIN_H,bold=True,sz=10,fc="FFFFFF",al=al_l)
            ws3.row_dimensions[R].height=22; R+=1
            for i,h in enumerate(["Cliente (col Q)","G-Super","G-Premium","G-Diesel","TOTAL","# Tx"],1):
                cw(ws3,R,i,val=h,fi=F_SIN,al=al_c,bold=True,sz=8,fc="92400E")
            ws3.row_dimensions[R].height=16; R+=1
            for cli in sorted(sin_resumen.keys()):
                d=sin_resumen[cli]; tc=d["GS"]+d["GP"]+d["GD"]
                cw(ws3,R,1,val=cli,fi=F_SIN_ROW,al=al_l,sz=9,fc="7C2D12")
                cw(ws3,R,2,val=d["GS"],fi=F_SIN_ROW,al=al_r,fmt=MF,sz=9)
                cw(ws3,R,3,val=d["GP"],fi=F_SIN_ROW,al=al_r,fmt=MF,sz=9)
                cw(ws3,R,4,val=d["GD"],fi=F_SIN_ROW,al=al_r,fmt=MF,sz=9)
                cw(ws3,R,5,val=tc,fi=F_SIN_ROW,al=al_r,fmt=MF,bold=True,sz=9)
                cw(ws3,R,6,val=d["n"],fi=F_SIN_ROW,al=al_r,fmt="#,##0",sz=9)
                sin_tot_gs+=d["GS"]; sin_tot_gp+=d["GP"]; sin_tot_gd+=d["GD"]; sin_tot_n+=d["n"]
                ws3.row_dimensions[R].height=16; R+=1
            cw(ws3,R,1,val="TOTAL SIN COMPROBANTE",fi=F_SIN_TOT,al=al_l,bold=True,sz=10,fc="FFFFFF")
            cw(ws3,R,2,val=sin_tot_gs,fi=F_SIN_TOT,al=al_r,fmt=MF,bold=True,sz=10,fc="FFFFFF")
            cw(ws3,R,3,val=sin_tot_gp,fi=F_SIN_TOT,al=al_r,fmt=MF,bold=True,sz=10,fc="FFFFFF")
            cw(ws3,R,4,val=sin_tot_gd,fi=F_SIN_TOT,al=al_r,fmt=MF,bold=True,sz=10,fc="FFFFFF")
            cw(ws3,R,5,val=sin_tot_gs+sin_tot_gp+sin_tot_gd,fi=F_SIN_TOT,al=al_r,fmt=MF,bold=True,sz=10,fc="FFFFFF")
            cw(ws3,R,6,val=sin_tot_n,fi=F_SIN_TOT,al=al_r,fmt="#,##0",bold=True,sz=10,fc="FFFFFF")
            ws3.row_dimensions[R].height=22; R+=2

            mw(ws3,R,1,6,"IMPUESTOS",F_DARK,bold=True,sz=10,fc="FFFFFF",al=al_l)
            ws3.row_dimensions[R].height=22; R+=1
            for i,h in enumerate(["Concepto","Despachos","Ventas del DÃ­a","Diferencia","",""],1):
                cw(ws3,R,i,val=h,fi=F_HDR,al=al_c,bold=True,sz=8,fc="FFFFFF")
            ws3.row_dimensions[R].height=16; R+=1
            for lbl,desp,pol in [("IVA 16%",tot_iva,v_tot_iva),("IEPS",tot_ieps,v_tot_ieps),
                                  ("TOTAL con impuestos",tot_imp,v_tot_total)]:
                dif=round(desp-pol,2); fi_dif=F_OK if abs(dif)<1 else F_WRN if abs(dif)<500 else F_ERR
                is_tot="TOTAL" in lbl; fi_r=F_TOT if is_tot else fill("FCE7F3")
                cw(ws3,R,1,val=lbl,fi=fi_r,al=al_l,bold=is_tot,sz=9)
                cw(ws3,R,2,val=desp,fi=fi_r,al=al_r,fmt=MF,bold=is_tot,sz=9)
                cw(ws3,R,3,val=pol,fi=fi_r,al=al_r,fmt=MF,bold=is_tot,sz=9)
                cw(ws3,R,4,val=dif,fi=fi_dif,al=al_r,fmt=MF,bold=True,sz=9)
                cw(ws3,R,5,fi=fi_r); cw(ws3,R,6,fi=fi_r)
                ws3.row_dimensions[R].height=16; R+=1
            R+=1
            mw(ws3,R,1,6,"MÃTRICAS",F_DARK,bold=True,sz=10,fc="FFFFFF",al=al_l)
            ws3.row_dimensions[R].height=22; R+=1
            F_LIGHT=fill("F8FAFC")
            for lbl,val,fi_v in [
                    ("Total CFDIs Ãºnicos",uuid_total,F_GS),
                    ("   â³ G-Super",uuid_gs,F_GS),("   â³ G-Premium",uuid_gp,F_GP),
                    ("   â³ G-Diesel",uuid_gd,F_GD),("Clientes Ãºnicos (RFC)",rfc_total,F_UUID),
                    ("Tx sin comprobante",sin_tot_n,F_SIN),("DÃ­as conciliados",len(fechas),F_UUID)]:
                cw(ws3,R,1,val=lbl,fi=F_LIGHT,al=al_l,sz=9)
                cw(ws3,R,2,val=val,fi=fi_v,al=al_r,fmt="#,##0",bold=True,sz=11)
                for ci in range(3,7): cw(ws3,R,ci,fi=F_LIGHT)
                ws3.row_dimensions[R].height=18; R+=1

            wb.move_sheet("Resumen", offset=-2)

            # ââ Guardar ââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
            carpeta_sal = os.path.dirname(desp_path)
            base = os.path.splitext(os.path.basename(desp_path))[0]
            OUT = os.path.join(carpeta_sal, f"Conciliacion_DespachoVentas_{base}.xlsx")
            wb.save(OUT)
            nombre = os.path.basename(OUT)

            self._vdc_resultado_path = OUT
            self.after(0, self._pb_detener, self._vdc_pb, self._vdc_pb_lbl)
            self.after(0, self._vdc_pb_frame.grid_remove)
            self.after(0, self._vdc_lbl_archivo.config, {"text": nombre})
            self.after(0, self._log, f"â ConciliaciÃ³n generada: {nombre}")
            self.after(0, self._vdc_cargar_visor, OUT)

        except Exception as e:
            import traceback as _tb
            self.after(0, self._pb_error, self._vdc_pb, self._vdc_pb_lbl)
            self.after(0, self._vdc_pb_frame.grid_remove)
            self.after(0, self._log, f"â Error: {e}", True)
            self.after(0, self._vdc_lbl_archivo.config, {"text": "Error"})
            from tkinter import messagebox
            self.after(0, messagebox.showerror, "Error",
                       f"{e}\n\n{_tb.format_exc()[-500:]}")

    # âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    # MÃDULO: CONCILIACIÃN CONTROL DE DESPACHO VS SAT
    # âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    def _tab_concilia_sat(self, nb):
        import os, threading
        from tkinter import filedialog, messagebox as mb

        tab = ttk.Frame(nb)
        nb.add(tab, text="  ð ConciliaciÃ³n SAT  ")

        # ââ Encabezado ââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
        hdr = tk.Frame(tab, bg="#0A7A4A", height=40)
        hdr.pack(fill="x")
        tk.Label(hdr, text="ð  CONCILIACIÃN  CONTROL DE DESPACHO  vs  SAT (CFDI/XML)",
                 bg="#0A7A4A", fg="white",
                 font=("Segoe UI", 11, "bold")).pack(side="left", padx=14, pady=8)

        # ââ Estado de archivos ââââââââââââââââââââââââââââââââââââââââââââââââ
        self._csat_excel_path = tk.StringVar(value="")
        self._csat_xml_paths  = []

        drop_frame = tk.Frame(tab, bg=COLOR_FONDO)
        drop_frame.pack(fill="x", padx=12, pady=(10,4))
        drop_frame.columnconfigure(0, weight=1)
        drop_frame.columnconfigure(1, weight=1)

        def _zona(parent, titulo, subtitulo, col):
            z = tk.Frame(parent, bg="#E8F5E9", bd=2, relief="groove")
            z.grid(row=0, column=col, padx=6, pady=4, sticky="nsew")
            tk.Label(z, text=titulo, bg="#E8F5E9", fg="#1A6B3A",
                     font=("Segoe UI", 12)).pack(pady=(14,2))
            tk.Label(z, text=subtitulo, bg="#E8F5E9", fg="#555",
                     font=("Segoe UI", 8), wraplength=280, justify="center").pack()
            info = tk.Label(z, text="Sin archivo", bg="#E8F5E9", fg="#0A7A4A",
                            font=("Segoe UI", 8, "bold"), wraplength=280)
            info.pack(pady=(4,2))
            return z, info

        zone_xl,  self._csat_lbl_xl  = _zona(drop_frame,
            "ð  Arrastra el Excel aquÃ­",
            "Control de Despachos (.xls / .xlsx)\no clic â seleccionar", 0)
        zone_xml, self._csat_lbl_xml = _zona(drop_frame,
            "ð  Arrastra los XMLs aquÃ­",
            "Archivos CFDI (.xml) â uno o varios\no clic â seleccionar", 1)

        ttk.Button(zone_xl,  text="Seleccionar Excel...",
                   command=self._csat_cargar_excel).pack(pady=(2,12))
        ttk.Button(zone_xml, text="Seleccionar XMLs...",
                   command=self._csat_cargar_xmls).pack(pady=(2,12))

        def _hon(z):
            z.config(bg="#C8E6C9")
            for w in z.winfo_children():
                try: w.config(bg="#C8E6C9")
                except: pass
        def _hoff(z):
            z.config(bg="#E8F5E9")
            for w in z.winfo_children():
                try: w.config(bg="#E8F5E9")
                except: pass

        for z, cmd in ((zone_xl, self._csat_cargar_excel),
                       (zone_xml, self._csat_cargar_xmls)):
            z.bind("<Enter>",    lambda e, zz=z: _hon(zz))
            z.bind("<Leave>",    lambda e, zz=z: _hoff(zz))
            z.bind("<Button-1>", lambda e, c=cmd: c())
            for w in z.winfo_children():
                w.bind("<Enter>",    lambda e, zz=z: _hon(zz))
                w.bind("<Leave>",    lambda e, zz=z: _hoff(zz))

        if _DND_OK:
            def _drop_xl(event):
                p = event.data.strip().strip("{}")
                if p.lower().endswith((".xls",".xlsx")):
                    self._csat_excel_path.set(p)
                    self._csat_lbl_xl.config(text="â " + os.path.basename(p))
                    self._csat_log_write(f"ð Excel: {p}", "ok")
                _hoff(zone_xl)

            def _drop_xml(event):
                import shlex, re
                raw = event.data.strip()
                paths = re.findall(r"\{([^}]+)\}|([^\s]+)", raw)
                paths = [a or b for a,b in paths]
                xmls  = [p for p in paths if p.lower().endswith(".xml")]
                if xmls:
                    self._csat_xml_paths = xmls
                    n = len(xmls)
                    self._csat_lbl_xml.config(
                        text=f"â {n} XML{'s' if n>1 else ''} cargado{'s' if n>1 else ''}")
                    self._csat_log_write(f"ð {n} XML(s):", "ok")
                    for p in xmls:
                        self._csat_log_write(f"   â¢ {os.path.basename(p)}", "neutral")
                _hoff(zone_xml)

            zone_xl.drop_target_register(DND_FILES)
            zone_xl.dnd_bind("<<Drop>>",      _drop_xl)
            zone_xl.dnd_bind("<<DragEnter>>", lambda e: _hon(zone_xl))
            zone_xl.dnd_bind("<<DragLeave>>", lambda e: _hoff(zone_xl))
            zone_xml.drop_target_register(DND_FILES)
            zone_xml.dnd_bind("<<Drop>>",      _drop_xml)
            zone_xml.dnd_bind("<<DragEnter>>", lambda e: _hon(zone_xml))
            zone_xml.dnd_bind("<<DragLeave>>", lambda e: _hoff(zone_xml))

        self._csat_xml_lbl = tk.StringVar(value="Sin XMLs cargados")
        ttk.Separator(tab, orient="horizontal").pack(fill="x", padx=12, pady=4)


        # ââ Botones âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
        btn_frame = tk.Frame(tab, bg=COLOR_FONDO)
        btn_frame.pack(fill="x", padx=12, pady=6)
        ttk.Button(btn_frame, text="â¡  Generar ConciliaciÃ³n",
                   style="Accion.TButton",
                   command=self._csat_generar).pack(side="left", padx=4)
        ttk.Button(btn_frame, text="ð  Limpiar",
                   command=self._csat_limpiar).pack(side="left", padx=4)
        self._csat_btn_abrir_conc = ttk.Button(
            btn_frame, text="ð  Abrir ConciliaciÃ³n",
            command=lambda: self._csat_abrir_archivo("conc"), state="disabled")
        self._csat_btn_abrir_conc.pack(side="left", padx=4)
        self._csat_btn_abrir_xl = ttk.Button(
            btn_frame, text="ð  Abrir Excel con UUID",
            command=lambda: self._csat_abrir_archivo("xl"), state="disabled")
        self._csat_btn_abrir_xl.pack(side="left", padx=4)
        self._csat_out_paths = {}

        # ââ Log / resultado âââââââââââââââââââââââââââââââââââââââââââââââââââ
        log_frame = ttk.Frame(tab, style="Tarjeta.TFrame")
        log_frame.pack(fill="both", expand=True, padx=12, pady=(4, 12))

        tk.Label(log_frame, text="Resultado de conciliaciÃ³n:",
                 bg=COLOR_TARJETA, fg=COLOR_TEXTO,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=8, pady=(8, 2))

        self._csat_log = tk.Text(log_frame, height=18, state="disabled",
                                  font=("Consolas", 9), bg="#F5F5F5",
                                  relief="flat", wrap="none")
        sb_y = ttk.Scrollbar(log_frame, command=self._csat_log.yview)
        sb_x = ttk.Scrollbar(log_frame, orient="horizontal",
                              command=self._csat_log.xview)
        self._csat_log.configure(yscrollcommand=sb_y.set,
                                  xscrollcommand=sb_x.set)
        sb_y.pack(side="right", fill="y")
        sb_x.pack(side="bottom", fill="x")
        self._csat_log.pack(fill="both", expand=True, padx=(8, 0), pady=(0, 4))

        # Tags colores log
        self._csat_log.tag_config("ok",      foreground="#1A6B3A", font=("Consolas", 9, "bold"))
        self._csat_log.tag_config("err",     foreground="#C0392B", font=("Consolas", 9, "bold"))
        self._csat_log.tag_config("hdr",     foreground="#1E3A5F", font=("Consolas", 9, "bold"))
        self._csat_log.tag_config("neutral", foreground="#555555")

    def _csat_log_write(self, txt, tag="neutral"):
        self._csat_log.config(state="normal")
        self._csat_log.insert("end", txt + "\n", tag)
        self._csat_log.see("end")
        self._csat_log.config(state="disabled")

    def _csat_abrir_archivo(self, tipo):
        import os, subprocess, sys
        p = self._csat_out_paths.get(tipo, "")
        if p and os.path.isfile(p):
            if sys.platform == "win32":
                os.startfile(p)
            else:
                subprocess.Popen(["xdg-open", p])
        else:
            from tkinter import messagebox as mb
            mb.showinfo("No disponible", "Primero genera la conciliaciÃ³n.")

    def _csat_limpiar(self):
        self._csat_out_paths = {}
        try:
            self._csat_btn_abrir_conc.config(state="disabled")
            self._csat_btn_abrir_xl.config(state="disabled")
        except Exception:
            pass
        self._csat_excel_path.set("")
        self._csat_xml_paths = []
        self._csat_xml_lbl.set("Sin XMLs cargados")
        self._csat_log.config(state="normal")
        self._csat_log.delete("1.0", "end")
        self._csat_log.config(state="disabled")

    def _csat_cargar_excel(self):
        from tkinter import filedialog
        self.lift()
        self.focus_force()
        self.update()
        p = filedialog.askopenfilename(
            parent=self,
            title="Selecciona el reporte Excel de Control de Despachos",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")])
        if p:
            self._csat_excel_path.set(p)
            self._csat_log_write(f"ð Excel cargado: {p}", "ok")

    def _csat_cargar_xmls(self):
        from tkinter import filedialog
        self.lift()
        self.focus_force()
        self.update()
        paths = filedialog.askopenfilenames(
            parent=self,
            title="Selecciona uno o varios archivos XML (CFDI)",
            filetypes=[("XML CFDI", "*.xml"), ("Todos", "*.*")])
        if paths:
            self._csat_xml_paths = list(paths)
            n = len(paths)
            self._csat_xml_lbl.set(f"{n} XML{'s' if n>1 else ''} cargado{'s' if n>1 else ''}")
            self._csat_log_write(f"ð {n} XML(s) cargado(s):", "ok")
            for p in paths:
                self._csat_log_write(f"   â¢ {os.path.basename(p)}", "neutral")

    def _csat_generar(self):
        import threading
        xl = self._csat_excel_path.get()
        xmls = self._csat_xml_paths
        if not xl:
            from tkinter import messagebox as mb
            mb.showwarning("Falta Excel", "Carga primero el reporte Excel de Control de Despachos.")
            return
        if not xmls:
            from tkinter import messagebox as mb
            mb.showwarning("Falta XML", "Carga al menos un archivo XML (CFDI).")
            return
        t = threading.Thread(target=self._csat_procesar, args=(xl, xmls), daemon=True)
        t.start()

    def _csat_procesar(self, xl_path, xml_paths):
        try:
            self._csat_procesar_inner(xl_path, xml_paths)
        except Exception as _ex:
            import traceback
            self._csat_log_write("\nâ ERROR INESPERADO:\n" + traceback.format_exc(), "err")

    def _csat_procesar_inner(self, xl_path, xml_paths):
        import xml.etree.ElementTree as ET
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import os

        self._csat_log.config(state="normal")
        self._csat_log.delete("1.0", "end")
        self._csat_log.config(state="disabled")

        self._csat_log_write("=" * 72, "hdr")
        self._csat_log_write("  CONCILIACIÃN CONTROL DE DESPACHO vs SAT (CFDI)", "hdr")
        self._csat_log_write("=" * 72, "hdr")

        # ââ 1. Parsear XMLs âââââââââââââââââââââââââââââââââââââââââââââââââââ
        NS = {"cfdi": "http://www.sat.gob.mx/cfd/4",
              "tfd":  "http://www.sat.gob.mx/TimbreFiscalDigital"}
        xmls_data = []
        self._csat_log_write("\nð Leyendo XMLs...", "hdr")
        for xp in xml_paths:
            try:
                root = ET.parse(xp).getroot()
                tfd  = root.find(".//tfd:TimbreFiscalDigital", NS)
                uuid = tfd.get("UUID", "").strip().upper() if tfd is not None else ""
                rec  = root.find("cfdi:Receptor", NS)
                conc = root.find(".//cfdi:Concepto", NS)
                imp  = root.find(".//cfdi:Impuestos", NS)
                subtotal = float(root.get("SubTotal", 0))
                total    = float(root.get("Total", 0))
                iva      = float(imp.get("TotalImpuestosTrasladados", 0)) if imp is not None else 0
                if iva == 0 and total > subtotal:
                    iva = round(total - subtotal, 2)
                xmls_data.append({
                    "uuid":      uuid,
                    "fecha":     root.get("Fecha","")[:10],
                    "serie":     root.get("Serie",""),
                    "folio":     root.get("Folio",""),
                    "receptor":  rec.get("Nombre","") if rec is not None else "",
                    "rfc_rec":   rec.get("Rfc","")    if rec is not None else "",
                    "concepto":  conc.get("Descripcion","") if conc is not None else "",
                    "subtotal":  subtotal,
                    "iva":       iva,
                    "total":     total,
                    "archivo":   os.path.basename(xp),
                })
            except Exception as e:
                self._csat_log_write(f"   â {os.path.basename(xp)}: {e}", "err")

        self._csat_log_write(f"   â {len(xmls_data)} XMLs leÃ­dos correctamente", "ok")
        if not xmls_data:
            self._csat_log_write("\nâ Sin XMLs vÃ¡lidos.", "err"); return

        # ââ 2. Leer Excel â detectar columna FolioFiscal ââââââââââââââââââââââ
        self._csat_log_write("\nð Leyendo Excel...", "hdr")
        try:
            is_xls = xl_path.lower().endswith(".xls")
            folio_col_idx = None   # Ã­ndice 0-based de la col FolioFiscal
            excel_rows = []        # lista de dicts por fila

            if is_xls:
                try:
                    import xlrd as _xl
                except ImportError:
                    import subprocess, sys
                    subprocess.check_call([sys.executable,"-m","pip","install","xlrd"],
                                          stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    import xlrd as _xl
                bk  = _xl.open_workbook(xl_path)
                sht = bk.sheet_by_index(0)
                # Buscar fila de encabezado y columna FolioFiscal
                hdr_row = 0
                for ri in range(min(10, sht.nrows)):
                    row_vals = [str(sht.cell_value(ri,ci)).strip() for ci in range(sht.ncols)]
                    matches = [ci for ci,v in enumerate(row_vals)
                               if "foliofiscal" in v.lower().replace(" ","")
                               or v.lower() in ("uuid","folio fiscal","folio_fiscal","uidd")]
                    if matches:
                        hdr_row = ri
                        folio_col_idx = matches[0]
                        self._csat_log_write(
                            f"   â Columna FolioFiscal encontrada: col {folio_col_idx+1} "
                            f"('{sht.cell_value(ri, folio_col_idx)}')", "ok")
                        break
                if folio_col_idx is None:
                    self._csat_log_write("   â No se encontrÃ³ columna FolioFiscal en el Excel", "err"); return
                # Leer filas de datos
                for ri in range(hdr_row+1, sht.nrows):
                    uuid_xl = str(sht.cell_value(ri, folio_col_idx)).strip().upper()
                    if not uuid_xl or uuid_xl in ("NONE",""):
                        continue
                    # Leer importe (buscar col "Importe" o "Total" o "Subtotal")
                    importe = 0
                    for ci in range(sht.ncols):
                        hv = str(sht.cell_value(hdr_row, ci)).strip().lower()
                        if hv in ("importe","total","subtotal"):
                            try: importe = float(sht.cell_value(ri, ci) or 0)
                            except: pass
                            break
                    excel_rows.append({"uuid": uuid_xl, "importe": importe})
            else:
                from openpyxl import load_workbook as _lw
                wb_r = _lw(xl_path, data_only=True)
                ws_r = wb_r.active
                hdr_row_xl = None
                for ri in range(1, min(12, ws_r.max_row+1)):
                    for ci in range(1, ws_r.max_column+1):
                        v = str(ws_r.cell(ri,ci).value or "").strip()
                        if "foliofiscal" in v.lower().replace(" ","")                                 or v.lower() in ("uuid","folio fiscal","folio_fiscal","uidd"):
                            folio_col_idx = ci - 1   # 0-based
                            hdr_row_xl    = ri
                            self._csat_log_write(
                                f"   â Columna FolioFiscal: col {ci} ('{v}')", "ok")
                            break
                    if folio_col_idx is not None: break
                if folio_col_idx is None:
                    self._csat_log_write("   â No se encontrÃ³ columna FolioFiscal", "err"); return
                # Pre-calcular col de importe UNA SOLA VEZ
                imp_col_idx = None
                for ci, cell in enumerate(ws_r[hdr_row_xl]):
                    if str(cell.value or "").strip().lower() in ("importe","total","subtotal"):
                        imp_col_idx = ci
                        break
                for row in ws_r.iter_rows(min_row=hdr_row_xl+1, values_only=True):
                    uuid_xl = str(row[folio_col_idx] or "").strip().upper()
                    if not uuid_xl or uuid_xl == "NONE": continue
                    importe = 0
                    if imp_col_idx is not None:
                        try: importe = float(row[imp_col_idx] or 0)
                        except: pass
                    excel_rows.append({"uuid": uuid_xl, "importe": importe})

            # Ãndice UUID â lista de importes en Excel
            excel_idx = {}
            for r in excel_rows:
                excel_idx.setdefault(r["uuid"], []).append(r["importe"])
            self._csat_log_write(f"   â {len(excel_rows)} filas leÃ­das  |  {len(excel_idx)} UUIDs Ãºnicos en Excel", "ok")

        except Exception as e:
            import traceback
            self._csat_log_write(f"\nâ Error leyendo Excel: {e}", "err")
            self._csat_log_write(traceback.format_exc(), "err"); return

        # ââ 3. Generar reporte de conciliaciÃ³n ââââââââââââââââââââââââââââââââ
        self._csat_log_write("\nð Conciliando...", "hdr")

        wb3  = Workbook()
        ws3  = wb3.active
        ws3.title = "ConciliaciÃ³n UUID"
        ws3.sheet_view.showGridLines = False

        AZUL="1E3A5F"; AZUL2="2E75B6"; AZUL3="D6E4F0"
        VERDE="C8E6C9"; ROJO="FFCDD2"; AMBAR="FFF9C4"; GRIS="F5F5F5"
        thin=Side(style="thin",color="BBBBBB")
        brd=Border(left=thin,right=thin,top=thin,bottom=thin)

        def cx(cell, v=None, bg=None, fg="000000", bold=False, sz=9,
               align="left", nf=None, wrap=False, border=True):
            if v is not None: cell.value = v
            cell.font=Font(name="Arial",bold=bold,color=fg,size=sz)
            if bg: cell.fill=PatternFill("solid",start_color=bg)
            cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
            if border: cell.border=brd
            if nf: cell.number_format=nf

        ws3.merge_cells("A1:L1")
        cx(ws3["A1"],"CONCILIACIÃN  CONTROL DE DESPACHO  vs  SAT (CFDI/XML)",
           bg=AZUL,fg="FFFFFF",bold=True,sz=13,align="center",border=False)
        ws3.row_dimensions[1].height=30
        ws3.merge_cells("A2:L2")
        cx(ws3["A2"],
           "Match: UUID (XML) = FolioFiscal (Excel)   |   SUPER SERVICIO PERIFERICO",
           bg="0A7A4A",fg="FFFFFF",sz=9,align="center",border=False)
        ws3.row_dimensions[2].height=18

        HDRS=[
            ("No.",4),("Folio Fiscal\n(UUID â XML)",42),
            ("Folio Fiscal\n(UIDD â Excel)",42),
            ("Â¿Coincide?",12),("Fecha",12),
            ("Receptor XML",28),("Concepto",26),
            ("SubTotal\nXML",14),("IVA\nXML",12),("Total\nXML",14),
            ("Importe\nExcel",14),("Diferencia",14),("Estado",14),
        ]
        for j,(h,w) in enumerate(HDRS,1):
            cx(ws3.cell(4,j),h,bg=AZUL2,fg="FFFFFF",bold=True,sz=9,align="center",wrap=True)
            ws3.column_dimensions[ws3.cell(4,j).column_letter].width=w
        ws3.row_dimensions[4].height=36

        self._csat_log_write(
            f"\n{'No.':<5}{'Fecha':<12}{'Match':<7}{'Total XML':>12}{'Imp.Excel':>13}{'Dif':>12}", "hdr")
        self._csat_log_write("-"*62,"hdr")

        ok_count = 0
        for i,x in enumerate(xmls_data, start=5):
            bg    = GRIS if i%2==0 else "FFFFFF"
            uuid  = x["uuid"]
            found = uuid in excel_idx
            imp_xl= sum(excel_idx[uuid]) if found else 0
            diff  = imp_xl - x["total"]
            m_bg  = VERDE if found else ROJO
            m_txt = "â  SÃ" if found else "â  NO"
            d_bg  = VERDE if abs(diff)<0.02 else (AMBAR if abs(diff)<100 else ROJO)
            est   = "â Conciliado" if found and abs(diff)<0.02 else                     ("â  Dif. menor" if found and abs(diff)<100 else                     ("â  Dif. mayor" if found else "â No encontrado"))
            est_bg= VERDE if "Conciliado" in est else (AMBAR if "â " in est else ROJO)
            if found: ok_count += 1

            row_data=[
                (i-4,bg,"center",None),(uuid,bg,"left",None),
                (m_txt,m_bg,"center",None),(x["fecha"],bg,"center",None),
                (x.get("receptor", x.get("nombre_rec","")),bg,"left",None),(x["concepto"],bg,"left",None),
                (x["subtotal"],bg,"right","$#,##0.00"),(x["iva"],bg,"right","$#,##0.00"),
                (x["total"],bg,"right","$#,##0.00"),
                (imp_xl if found else None,bg,"right","$#,##0.00"),
                (diff if found else None,d_bg,"right","$#,##0.00;[Red]($#,##0.00)"),
                (est,est_bg,"center",None),
            ]
            for j,(v,cbg,al,nf) in enumerate(row_data,1):
                cx(ws3.cell(i,j),v,bg=cbg,align=al,nf=nf,sz=9)
            ws3.row_dimensions[i].height=20

            tag = "ok" if found else "err"
            self._csat_log_write(
                f"{i-4:<5}{x['fecha']:<12}{'â' if found else 'â':<7}"
                f"${x['total']:>11,.2f}  ${imp_xl:>11,.2f}  ${diff:>10,.2f}", tag)

        # Fila totales
        tr = 5+len(xmls_data)
        ws3.merge_cells(f"A{tr}:F{tr}")
        cx(ws3.cell(tr,1),"TOTALES",bg=AZUL,fg="FFFFFF",bold=True,sz=10,align="right")
        for j,v,nf in [
            (7,sum(x["subtotal"] for x in xmls_data),"$#,##0.00"),
            (8,sum(x["iva"]      for x in xmls_data),"$#,##0.00"),
            (9,sum(x["total"]    for x in xmls_data),"$#,##0.00"),
            (10,sum(sum(excel_idx.get(x["uuid"],[])) for x in xmls_data),"$#,##0.00"),
        ]:
            cx(ws3.cell(tr,j),v,bg=AZUL,fg="FFFFFF",bold=True,sz=10,align="right",nf=nf)
        ws3.row_dimensions[tr].height=22

        nr=tr+2
        ws3.merge_cells(f"A{nr}:L{nr}")
        cx(ws3.cell(nr,1),
           f"Procesados: {len(xmls_data)} XMLs  |  Encontrados en Excel: {ok_count}  |  "
           f"No encontrados: {len(xmls_data)-ok_count}",
           bg=AZUL3,sz=9,wrap=True,border=False)
        ws3.row_dimensions[nr].height=24

        # Guardar en carpeta del script si no hay permisos en la del Excel
        import tempfile
        _cand = os.path.join(os.path.dirname(xl_path), "Conciliacion_DespachoVsSAT.xlsx")
        try:
            wb3.save(_cand)
            out_conc = _cand
        except PermissionError:
            # Intentar carpeta del script
            _app_dir = os.path.dirname(os.path.abspath(__file__))
            out_conc = os.path.join(_app_dir, "Conciliacion_DespachoVsSAT.xlsx")
            try:
                wb3.save(out_conc)
            except PermissionError:
                # Ãltimo recurso: carpeta temporal del sistema
                out_conc = os.path.join(tempfile.gettempdir(), "Conciliacion_DespachoVsSAT.xlsx")
                wb3.save(out_conc)
            self._csat_log_write(
                f"   â¹ Sin permisos en carpeta origen â guardado en: {out_conc}", "neutral")
        xl_out = xl_path

        total_xml = sum(x["total"] for x in xmls_data)
        self._csat_log_write("\n"+"="*72,"hdr")
        self._csat_log_write(f"  XMLs procesados   : {len(xmls_data)}","ok")
        self._csat_log_write(f"  Encontrados       : {ok_count} â","ok")
        self._csat_log_write(f"  No encontrados    : {len(xmls_data)-ok_count} â","err")
        self._csat_log_write(f"  Total XMLs        : ${total_xml:,.2f}","ok")
        self._csat_log_write(f"\nð¾ {out_conc}","ok")
        self._csat_log_write("="*72,"hdr")
        self._csat_out_paths = {"conc": out_conc, "xl": xl_out}
        try:
            self._csat_btn_abrir_conc.config(state="normal")
            self._csat_btn_abrir_xl.config(state="normal")
        except Exception:
            pass


        # ââ Parsear XMLs ââââââââââââââââââââââââââââââââââââââââââââââââââââââ
        NS = {
            "cfdi": "http://www.sat.gob.mx/cfd/4",
            "tfd":  "http://www.sat.gob.mx/TimbreFiscalDigital",
        }
        xmls_data = []
        self._csat_log_write("\nð Leyendo XMLs...", "hdr")
        for xp in xml_paths:
            try:
                tree = ET.parse(xp)
                root = tree.getroot()
                uuid = ""
                fecha_timb = ""
                tfd = root.find(".//tfd:TimbreFiscalDigital", NS)
                if tfd is not None:
                    uuid = tfd.get("UUID", "")
                    fecha_timb = tfd.get("FechaTimbrado", "")[:10]

                fecha_emision = root.get("Fecha", "")[:10]
                subtotal  = float(root.get("SubTotal", 0))
                total     = float(root.get("Total", 0))
                serie     = root.get("Serie", "")
                folio     = root.get("Folio", "")
                rec = root.find("cfdi:Receptor", NS)
                nombre_rec = rec.get("Nombre", "") if rec is not None else ""
                rfc_rec    = rec.get("Rfc", "")    if rec is not None else ""

                # IVA
                imp = root.find(".//cfdi:Impuestos", NS)
                iva = float(imp.get("TotalImpuestosTrasladados", 0)) if imp is not None else round(total - subtotal, 2)

                # Concepto
                conc = root.find(".//cfdi:Concepto", NS)
                desc = conc.get("Descripcion", "") if conc is not None else ""

                xmls_data.append({
                    "uuid": uuid, "fecha": fecha_emision,
                    "serie": serie, "folio": folio,
                    "nombre_rec": nombre_rec, "rfc_rec": rfc_rec,
                    "subtotal": subtotal, "iva": iva, "total": total,
                    "concepto": desc, "archivo": os.path.basename(xp),
                })
                self._csat_log_write(
                    f"   â {os.path.basename(xp):40s} UUID: {uuid[:18]}...  ${total:,.2f}", "ok")
            except Exception as e:
                self._csat_log_write(f"   â Error en {os.path.basename(xp)}: {e}", "err")

        if not xmls_data:
            self._csat_log_write("\nâ No se pudo leer ningÃºn XML.", "err")
            return

        # ââ Leer Excel (.xls o .xlsx) y llenar UIDD âââââââââââââââââââââââââ
        self._csat_log_write("\nð Leyendo reporte Excel...", "hdr")
        try:
            is_xls = xl_path.lower().endswith(".xls")
            if is_xls:
                try:
                    import xlrd as _xlrd
                except ImportError:
                    import subprocess, sys
                    self._csat_log_write("   â¹ Instalando xlrd...", "neutral")
                    subprocess.check_call([sys.executable, "-m", "pip", "install", "xlrd"],
                                          stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                    import xlrd as _xlrd
                _xls_book = _xlrd.open_workbook(xl_path)
                _xls_sht  = _xls_book.sheet_by_index(0)
                # Leer datos directamente desde xlrd
                fecha_data = {}  # fecha â {total_b2, row_idx}
                for ri in range(3, _xls_sht.nrows):  # fila 4+ (0-indexed=3)
                    tipo = _xls_sht.cell_value(ri, 0)
                    if str(tipo).strip().upper() != 'D':
                        continue
                    raw_fecha = _xls_sht.cell(ri, 1)
                    if raw_fecha.ctype == _xlrd.XL_CELL_DATE:
                        from datetime import datetime as _dt
                        fecha = _xlrd.xldate_as_datetime(
                            raw_fecha.value, _xls_book.datemode).strftime("%Y-%m-%d")
                    else:
                        fecha = str(raw_fecha.value)[:10]
                    total_b2 = _xls_sht.cell_value(ri, 28) or 0  # col AC (idx 28)
                    fecha_data[fecha] = {"total_b2": total_b2, "row_idx": ri}

                # Convertir a openpyxl para guardar con UUIDs
                from openpyxl import Workbook as _WB2
                _wb_out = _WB2()
                _ws_out = _wb_out.active
                for ri in range(_xls_sht.nrows):
                    for ci in range(_xls_sht.ncols):
                        cell = _xls_sht.cell(ri, ci)
                        if cell.ctype == _xlrd.XL_CELL_DATE:
                            from datetime import datetime as _dt2
                            val = _xlrd.xldate_as_datetime(cell.value, _xls_book.datemode)
                            _ws_out.cell(ri+1, ci+1).value = val
                        else:
                            _ws_out.cell(ri+1, ci+1).value = cell.value

                # Llenar UIDD (col F = idx 6 en openpyxl)
                reporte = {}
                for x in xmls_data:
                    fecha = x["fecha"]
                    if fecha in fecha_data:
                        ri = fecha_data[fecha]["row_idx"]
                        _ws_out.cell(ri+1, 6).value = x["uuid"]
                        _ws_out.cell(ri+1, 6).font = Font(name="Arial", size=9, color="1A6B3A")
                        total_b2 = fecha_data[fecha]["total_b2"]
                        reporte[fecha] = {"uuid_excel": x["uuid"], "total_b2": total_b2}
                        self._csat_log_write(
                            f"   â Fecha {fecha}: UIDD actualizado  |  Total B2: ${total_b2:,.2f}", "ok")
                    else:
                        reporte[fecha] = {"uuid_excel": "", "total_b2": 0}
                        self._csat_log_write(f"   â  Fecha {fecha}: no encontrada en Excel", "err")

                xl_out = os.path.join(os.path.dirname(xl_path),
                                      "control_despachos_con_uuid.xlsx")
                _wb_out.save(xl_out)

            self._csat_log_write(f"\nð¾ Excel actualizado: {os.path.basename(xl_out)}", "ok")

        except Exception as e:
            self._csat_log_write(f"\nâ Error leyendo Excel: {e}", "err")
            import traceback
            self._csat_log_write(traceback.format_exc(), "err")
            return

        # ââ Generar reporte de conciliaciÃ³n âââââââââââââââââââââââââââââââââââ
        self._csat_log_write("\nð Generando reporte de conciliaciÃ³n...", "hdr")

        wb3  = Workbook()
        ws3  = wb3.active
        ws3.title = "ConciliaciÃ³n UUID"
        ws3.sheet_view.showGridLines = False

        AZUL="1E3A5F"; AZUL2="2E75B6"; AZUL3="D6E4F0"
        VERDE="C8E6C9"; ROJO="FFCDD2"; AMBAR="FFF9C4"; GRIS="F5F5F5"
        thin=Side(style="thin",color="BBBBBB")
        brd=Border(left=thin,right=thin,top=thin,bottom=thin)

        def cx(cell, v=None, bg=None, fg="000000", bold=False, sz=9,
               align="left", nf=None, wrap=False, border=True):
            if v is not None: cell.value = v
            cell.font=Font(name="Arial",bold=bold,color=fg,size=sz)
            if bg: cell.fill=PatternFill("solid",start_color=bg)
            cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
            if border: cell.border=brd
            if nf: cell.number_format=nf

        ws3.merge_cells("A1:K1")
        cx(ws3["A1"],"CONCILIACIÃN  CONTROL DE DESPACHO  vs  SAT (CFDI/XML)",
           bg=AZUL,fg="FFFFFF",bold=True,sz=13,align="center",border=False)
        ws3.row_dimensions[1].height=30

        ws3.merge_cells("A2:K2")
        cx(ws3["A2"],
           "Match: UUID (XML) = Folio Fiscal / UIDD (Excel)   |   SUPER SERVICIO PERIFERICO  |  RFC: SSP950710PL9",
           bg="0A7A4A",fg="FFFFFF",sz=9,align="center",border=False)
        ws3.row_dimensions[2].height=18

        HDRS=[
            ("No.",4),("Folio Fiscal\n(UUID â XML)",42),
            ("Folio Fiscal\n(UIDD â Excel)",42),
            ("Â¿Coincide?",13),("Fecha",12),
            ("Receptor XML",26),("Concepto",28),
            ("SubTotal XML",14),("IVA XML",12),
            ("Total XML",14),("Total B2 Excel",16),("Diferencia",14),
        ]
        for j,(h,w) in enumerate(HDRS,1):
            cx(ws3.cell(4,j),h,bg=AZUL2,fg="FFFFFF",bold=True,sz=9,align="center",wrap=True)
            ws3.column_dimensions[ws3.cell(4,j).column_letter].width=w
        ws3.row_dimensions[4].height=38

        self._csat_log_write(f"\n{'No.':<5}{'Fecha':<12}{'Match':<7}{'Total XML':>12}{'Imp.Excel':>13}{'Dif':>12}", "hdr")
        self._csat_log_write("-"*62, "hdr")

        ok_count = 0
        for i,x in enumerate(xmls_data, start=5):
            bg    = GRIS if i%2==0 else "FFFFFF"
            uuid  = x["uuid"]
            found = uuid in excel_idx
            imp_xl= sum(excel_idx[uuid]) if found else 0
            diff  = imp_xl - x["total"]
            m_bg  = VERDE if found else ROJO
            m_txt = "â  SÃ" if found else "â  NO"
            d_bg  = VERDE if abs(diff)<0.02 else (AMBAR if abs(diff)<100 else ROJO)
            est   = "â Conciliado" if found and abs(diff)<0.02 else \
                    ("â  Dif.menor"  if found and abs(diff)<100  else \
                    ("â  Dif.mayor"  if found else "â No encontrado"))
            est_bg= VERDE if "Conciliado" in est else (AMBAR if "â " in est else ROJO)
            if found: ok_count += 1

            uuid_xl_val = uuid if found else "â"
            row_data=[
                (i-4,bg,"center",None),(uuid,bg,"left",None),
                (uuid_xl_val,m_bg,"left",None),
                (m_txt,m_bg,"center",None),(x["fecha"],bg,"center",None),
                (x.get("receptor", x.get("nombre_rec","")),bg,"left",None),(x["concepto"],bg,"left",None),
                (x["subtotal"],bg,"right","$#,##0.00"),(x["iva"],bg,"right","$#,##0.00"),
                (x["total"],bg,"right","$#,##0.00"),
                (imp_xl if found else None,bg,"right","$#,##0.00"),
                (diff   if found else None,d_bg,"right","$#,##0.00;[Red]($#,##0.00)"),
                (est,est_bg,"center",None),
            ]
            for j,(v,cbg,al,nf) in enumerate(row_data,1):
                cx(ws3.cell(i,j),v,bg=cbg,align=al,nf=nf,sz=9)
            ws3.row_dimensions[i].height=20
            tag = "ok" if found else "err"
            self._csat_log_write(
                f"{i-4:<5}{x['fecha']:<12}{'â' if found else 'â':<7}"
                f"${x['total']:>11,.2f}  ${imp_xl:>11,.2f}  ${diff:>10,.2f}", tag)

        # Fila totales
        tr = 5+len(xmls_data)
        ws3.merge_cells(f"A{tr}:F{tr}")
        cx(ws3.cell(tr,1),"TOTALES",bg=AZUL,fg="FFFFFF",bold=True,sz=10,align="right")
        for jj,vv,nff in [
            (7, sum(x["subtotal"] for x in xmls_data),"$#,##0.00"),
            (8, sum(x["iva"]      for x in xmls_data),"$#,##0.00"),
            (9, sum(x["total"]    for x in xmls_data),"$#,##0.00"),
            (10,sum(sum(excel_idx.get(x["uuid"],[])) for x in xmls_data),"$#,##0.00"),
        ]:
            cx(ws3.cell(tr,jj),vv,bg=AZUL,fg="FFFFFF",bold=True,sz=10,align="right",nf=nff)
        ws3.row_dimensions[tr].height=22

        nr=tr+2
        ws3.merge_cells(f"A{nr}:L{nr}")
        cx(ws3.cell(nr,1),
           f"Procesados: {len(xmls_data)} XMLs  |  Encontrados: {ok_count}  |  No encontrados: {len(xmls_data)-ok_count}",
           bg=AZUL3,sz=9,wrap=True,border=False)
        ws3.row_dimensions[nr].height=24

        # Guardar con manejo de permisos
        import tempfile
        _cand = os.path.join(os.path.dirname(xl_path),"Conciliacion_DespachoVsSAT.xlsx")
        try:
            wb3.save(_cand); out_conc = _cand
        except PermissionError:
            _app_dir = os.path.dirname(os.path.abspath(__file__))
            out_conc = os.path.join(_app_dir,"Conciliacion_DespachoVsSAT.xlsx")
            try: wb3.save(out_conc)
            except PermissionError:
                out_conc = os.path.join(tempfile.gettempdir(),"Conciliacion_DespachoVsSAT.xlsx")
                wb3.save(out_conc)
            self._csat_log_write(f"   â¹ Guardado en: {out_conc}", "neutral")


        total_xml = sum(x["total"] for x in xmls_data)
        self._csat_log_write("\n" + "=" * 70, "hdr")
        self._csat_log_write(f"  XMLs procesados : {len(xmls_data)}", "ok")
        self._csat_log_write(f"  Total XMLs      : ${total_xml:,.2f}", "ok")
        self._csat_log_write(f"\nð¾ ConciliaciÃ³n guardada en:\n   {out_conc}", "ok")
        self._csat_log_write("=" * 70, "hdr")

    def _tab_configuracion(self, nb):
        tab = ttk.Frame(nb)
        nb.add(tab, text="  â ConfiguraciÃ³n  ")

        card = ttk.Frame(tab, style="Tarjeta.TFrame")
        card.pack(fill="x", padx=12, pady=12)
        card.columnconfigure(1, weight=1)

        ttk.Label(card, text="â  Opciones generales", style="Encabezado.TLabel").grid(
            row=0, column=0, columnspan=3, sticky="w", padx=12, pady=(10, 8))

        # Carpeta de salida personalizada
        ttk.Label(card, text="Carpeta de salida:", style="Tarjeta.TLabel", width=24).grid(
            row=1, column=0, sticky="w", padx=12, pady=4)
        ttk.Entry(card, textvariable=self.salida_dir).grid(row=1, column=1, sticky="ew", padx=6)
        ttk.Button(card, text="Elegir...",
                   command=self._elegir_salida).grid(row=1, column=2, padx=6)
        ttk.Label(card,
            text="  (si vacÃ­o, se usa 'resultados/' dentro de la carpeta de trabajo)",
            style="Tarjeta.TLabel", font=("Segoe UI", 8),
        ).grid(row=2, column=1, columnspan=2, sticky="w", padx=6)

        ttk.Separator(card, orient="horizontal").grid(row=3, column=0, columnspan=3, sticky="ew", padx=12, pady=10)

        # Prefijo de archivos
        ttk.Label(card, text="Prefijo de archivos generados:", style="Tarjeta.TLabel", width=24).grid(
            row=4, column=0, sticky="w", padx=12, pady=4)
        ttk.Entry(card, textvariable=self.prefijo_salida, width=30).grid(row=4, column=1, sticky="w", padx=6)
        ttk.Label(card,
            text="  Ej: 'MAYO2026_' â archivos como MAYO2026_Nomina_...",
            style="Tarjeta.TLabel", font=("Segoe UI", 8),
        ).grid(row=5, column=1, columnspan=2, sticky="w", padx=6)

        ttk.Separator(card, orient="horizontal").grid(row=6, column=0, columnspan=3, sticky="ew", padx=12, pady=10)

        # Info del sistema
        ttk.Label(card, text="InformaciÃ³n del sistema:", style="Tarjeta.TLabel", width=24).grid(
            row=7, column=0, sticky="nw", padx=12, pady=4)

        info_lines = [
            f"Python: {sys.version.split()[0]}",
            f"Motor contable: {'cargado â' if cn else 'NO encontrado â'}",
            f"Carpeta actual: {os.getcwd()}",
        ]
        info_txt = tk.Text(card, height=4, width=60,
            bg=COLOR_AZUL_SUAVE, fg=COLOR_AZUL,
            relief="flat", font=("Consolas", 9), state="normal",
            highlightthickness=0)
        info_txt.insert("end", "\n".join(info_lines))
        info_txt.config(state="disabled")
        info_txt.grid(row=7, column=1, columnspan=2, sticky="w", padx=6, pady=4)

        ttk.Button(card, text="ð¾  Aplicar configuraciÃ³n",
                   command=lambda: self._log("ConfiguraciÃ³n guardada.", ok=True)).grid(
            row=8, column=0, columnspan=3, padx=12, pady=12)

    # ---------------------------------------------------------------- #
    # PANEL DE LOG (compartido)
    # ---------------------------------------------------------------- #
    def _panel_log(self):
        # Contenedor raÃ­z empacado al fondo â se reserva ANTES del notebook
        log_outer = tk.Frame(self._content_frame, bg=COLOR_FONDO)
        log_outer.pack(side="bottom", fill="x")

        ttk.Label(log_outer, text="ð Registro de actividad:", style="Seccion.TLabel").pack(
            anchor="w", padx=14, pady=(4, 2))
        log_frame = tk.Frame(log_outer, bg="#0D1117")
        log_frame.pack(fill="both", expand=False, padx=10, pady=(0, 4))
        log_scroll = ttk.Scrollbar(log_frame, orient="vertical")
        self.log_text = tk.Text(
            log_frame, height=6, yscrollcommand=log_scroll.set,
            state="disabled", wrap="word",
            bg="#0D1117", fg="#C9D1D9",
            relief="flat", highlightthickness=1,
            highlightbackground="#1F2937", highlightcolor="#1E3A8A",
            font=("Consolas", 9), padx=8, pady=6,
        )
        log_scroll.config(command=self.log_text.yview)
        self.log_text.pack(side="left", fill="both", expand=True)
        log_scroll.pack(side="right", fill="y")
        self.log_text.tag_config("error", foreground=COLOR_ERROR,  font=("Consolas", 9, "bold"))
        self.log_text.tag_config("ok",    foreground="#3FB950",     font=("Consolas", 9, "bold"))

        bottom = ttk.Frame(log_outer)
        bottom.pack(fill="x", padx=10, pady=(0, 6))
        ttk.Button(bottom, text="ð  Limpiar log",
                   command=self._limpiar_log).pack(side="left")

    # ================================================================ #
    # ACCIONES DE UI
    # ================================================================ #

    def _elegir_carpeta(self):
        self.lift()
        self.focus_force()
        self.update()
        r = filedialog.askdirectory(initialdir=self.carpeta.get() or os.getcwd())
        if r:
            self.carpeta.set(r)

    def _elegir_excel(self):
        self.lift()
        self.focus_force()
        self.update()
        r = filedialog.askopenfilename(
            parent=self,
            initialdir=self.carpeta.get() or os.getcwd(),
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos los archivos", "*.*")],
        )
        if r:
            self.excel_path.set(r)

    def _elegir_salida(self):
        self.lift()
        self.focus_force()
        self.update()
        r = filedialog.askdirectory(initialdir=self.carpeta.get() or os.getcwd())
        if r:
            self.salida_dir.set(r)

    # ââ Elegir carpeta / catÃ¡logo independiente por mÃ³dulo âââââââââââ #
    def _elegir_carpeta_pagos(self):
        self.lift(); self.focus_force(); self.update()
        rutas = filedialog.askopenfilenames(
            parent=self,
            title="Selecciona los PDFs de nÃ³mina",
            initialdir=self.carpeta_pagos.get() or os.getcwd(),
            filetypes=[("Archivos PDF", "*.pdf")],
        )
        if not rutas:
            return

        # Carpeta de trabajo = carpeta del primer PDF seleccionado
        carpeta = os.path.dirname(rutas[0])
        self.carpeta_pagos.set(carpeta)

        # Limpiar listas actuales
        for lb in self.listboxes.values():
            lb.delete(0, "end")
        for key in ("nomina", "complementos", "vacaciones"):
            self.archivos_por_categoria[key] = {}

        self._log(f"[Pagos] {len(rutas)} PDF(s) seleccionado(s).")
        no_reconocidos = []
        for path in sorted(rutas):
            categoria, descripcion = clasificar_pdf(path)
            rel = os.path.relpath(path, carpeta)
            etiqueta = f"{rel}   â   {descripcion}"
            if categoria in ("nomina", "complementos", "vacaciones"):
                self.archivos_por_categoria[categoria][path] = descripcion
                self.listboxes[categoria].insert("end", etiqueta)
                self.listboxes[categoria].select_set("end")
            elif categoria != "prestamos":
                no_reconocidos.append(f"{os.path.basename(path)} â {categoria}")
                self._log(f"  Sin clasificar: {os.path.basename(path)} ({categoria})")

        for key in ("nomina", "complementos", "vacaciones"):
            n = len(self.archivos_por_categoria[key])
            self._log(f"  {CATEGORIA_LABELS.get(key, key)}: {n} archivo(s)")

        if no_reconocidos and self.otros_label:
            self.otros_label.config(text="Sin clasificar: " + "; ".join(no_reconocidos))
        elif self.otros_label:
            self.otros_label.config(text="")

        self._log("[Pagos] PDFs cargados.", ok=True)

    def _elegir_excel_pagos(self):
        self.lift(); self.focus_force(); self.update()
        r = filedialog.askopenfilename(
            parent=self,
            initialdir=self.carpeta_pagos.get() or os.getcwd(),
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos los archivos", "*.*")],
        )
        if r:
            self.excel_path_pagos.set(r)

    def _elegir_carpeta_prestamos(self):
        self.lift(); self.focus_force(); self.update()
        r = filedialog.askdirectory(initialdir=self.carpeta_prestamos.get() or os.getcwd())
        if r:
            self.carpeta_prestamos.set(r)

    # ââ Escaneo independiente: Pagos Bancarios âââââââââââââââââââââââ #
    def _escanear_pagos(self):
        if cn is None:
            self._log("No puedo escanear: falta conciliacion_nomina.py.", error=True)
            return
        carpeta = self.carpeta_pagos.get()
        if not os.path.isdir(carpeta):
            messagebox.showerror("Carpeta invÃ¡lida", "Elige una carpeta vÃ¡lida primero.")
            return

        # Limpiar listas de pagos (NO toca prestamos)
        for lb in self.listboxes.values():
            lb.delete(0, "end")
        for key in ("nomina", "complementos", "vacaciones"):
            self.archivos_por_categoria[key] = {}

        pdfs = []
        for raiz, dirs, archivos in os.walk(carpeta):
            for f in sorted(archivos):
                if f.lower().endswith(".pdf"):
                    pdfs.append(os.path.join(raiz, f))
        pdfs.sort()

        todos = []
        for raiz, dirs, archivos in os.walk(carpeta):
            for f in archivos:
                todos.append(f)
        self._log(f"[Pagos] Carpeta: {carpeta}")
        self._log(f"[Pagos] Archivos encontrados: {len(todos)}  |  PDFs: {len(pdfs)}")

        if todos and not pdfs:
            ext_enc = list({os.path.splitext(f)[1].lower() for f in todos})
            self._log(f"Tipos encontrados: {', '.join(ext_enc)}", error=True)
            return
        if not pdfs:
            self._log("[Pagos] No encontrÃ© PDFs en esa carpeta.", error=True)
            return

        self._log(f"[Pagos] Escaneando {len(pdfs)} PDF(s)...")
        no_reconocidos = []
        for path in pdfs:
            categoria, descripcion = clasificar_pdf(path)
            rel = os.path.relpath(path, carpeta)
            etiqueta = f"{rel}   â   {descripcion}"
            if categoria in ("nomina", "complementos", "vacaciones"):
                self.archivos_por_categoria[categoria][path] = descripcion
                self.listboxes[categoria].insert("end", etiqueta)
                self.listboxes[categoria].select_set("end")
            elif categoria != "prestamos":
                no_reconocidos.append(f"{os.path.basename(path)} â {categoria}")
                self._log(f"  Sin clasificar: {os.path.basename(path)} ({categoria})")

        for key in ("nomina", "complementos", "vacaciones"):
            n = len(self.archivos_por_categoria[key])
            self._log(f"  {CATEGORIA_LABELS.get(key, key)}: {n} archivo(s)")

        if no_reconocidos and self.otros_label:
            self.otros_label.config(text="Sin clasificar: " + "; ".join(no_reconocidos))
        elif self.otros_label:
            self.otros_label.config(text="")

        self._log("[Pagos] Escaneo terminado.", ok=True)

    # ââ Escaneo independiente: PrÃ©stamos âââââââââââââââââââââââââââââ #
    def _escanear_prestamos(self):
        if cn is None:
            self._log("No puedo escanear: falta conciliacion_nomina.py.", error=True)
            return
        carpeta = self.carpeta_prestamos.get()
        if not os.path.isdir(carpeta):
            messagebox.showerror("Carpeta invÃ¡lida", "Elige una carpeta vÃ¡lida primero.")
            return

        # Limpiar solo la lista de prÃ©stamos (NO toca pagos)
        if self.lb_prestamos:
            self.lb_prestamos.delete(0, "end")
        self.archivos_por_categoria["prestamos"] = {}

        pdfs = []
        for raiz, dirs, archivos in os.walk(carpeta):
            for f in sorted(archivos):
                if f.lower().endswith(".pdf"):
                    pdfs.append(os.path.join(raiz, f))
        pdfs.sort()

        self._log(f"[PrÃ©stamos] Carpeta: {carpeta}")
        self._log(f"[PrÃ©stamos] PDFs encontrados: {len(pdfs)}")

        if not pdfs:
            self._log("[PrÃ©stamos] No encontrÃ© PDFs en esa carpeta.", error=True)
            return

        self._log(f"[PrÃ©stamos] Escaneando {len(pdfs)} PDF(s)...")
        for path in pdfs:
            categoria, descripcion = clasificar_pdf(path)
            rel = os.path.relpath(path, carpeta)
            etiqueta = f"{rel}   â   {descripcion}"
            if categoria == "prestamos":
                self.archivos_por_categoria["prestamos"][path] = descripcion
                self.lb_prestamos.insert("end", etiqueta)
                self.lb_prestamos.select_set("end")
            else:
                self._log(f"  No es prÃ©stamo: {os.path.basename(path)} ({categoria})")

        n = len(self.archivos_por_categoria["prestamos"])
        self._log(f"  PrÃ©stamos: {n} archivo(s)")
        self._log("[PrÃ©stamos] Escaneo terminado.", ok=True)

    def _log(self, texto, error=False, ok=False):
        self.log_text.config(state="normal")
        tag = "error" if error else ("ok" if ok else None)
        marca = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{marca}] {texto}\n", tag)
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def _limpiar_log(self):
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.config(state="disabled")

    def _sel_todos(self, key):
        self.listboxes[key].select_set(0, "end")

    def _sel_ninguno(self, key):
        self.listboxes[key].select_clear(0, "end")

    def _resultados_dir(self, base_hint=None):
        """Devuelve la carpeta de resultados.

        Prioridad:
        1. Carpeta personalizada (campo Salida en ConfiguraciÃ³n) si existe.
        2. base_hint / carpeta seleccionada de PDFs si existe y no es sistema.
        3. Carpeta del script como respaldo final.
        """
        custom = self.salida_dir.get().strip()
        if custom and os.path.isdir(custom):
            out = custom
            os.makedirs(out, exist_ok=True)
            return out

        # Candidatos en orden de preferencia
        candidatos = []
        if base_hint:
            candidatos.append(base_hint)
        carpeta_pdfs = self.carpeta_pagos.get().strip()
        if carpeta_pdfs:
            candidatos.append(carpeta_pdfs)
        # Carpeta del script
        candidatos.append(os.path.dirname(os.path.abspath(__file__)))

        for base in candidatos:
            # Rechazar rutas de sistema de Windows
            base_norm = os.path.normcase(base)
            if any(base_norm.startswith(os.path.normcase(p))
                   for p in [os.environ.get('WINDIR', 'C:\\Windows'),
                              os.environ.get('SystemRoot', 'C:\\Windows')]):
                continue
            if os.path.isdir(base):
                out = os.path.join(base, "resultados")
                try:
                    os.makedirs(out, exist_ok=True)
                    return out
                except PermissionError:
                    continue

        # Ãltimo recurso: carpeta Documentos del usuario
        out = os.path.join(os.path.expanduser("~"), "Documents", "resultados_nomina")
        os.makedirs(out, exist_ok=True)
        return out

    def _nombre_salida(self, base):
        prefijo = self.prefijo_salida.get().strip()
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefijo}{base}_{ts}.xlsx"

    def _abrir_resultados(self):
        out = self._resultados_dir()
        try:
            os.startfile(out)
        except AttributeError:
            messagebox.showinfo("Carpeta de resultados", out)

    # ---------------------------------------------------------------- #
    # ESCANEO
    # ---------------------------------------------------------------- #
    def _escanear(self):
        if cn is None:
            self._log("No puedo escanear: falta conciliacion_nomina.py.", error=True)
            return
        carpeta = self.carpeta.get()
        if not os.path.isdir(carpeta):
            messagebox.showerror("Carpeta invÃ¡lida", "Elige una carpeta vÃ¡lida primero.")
            return

        # Limpiar todo
        for lb in self.listboxes.values():
            lb.delete(0, "end")
        if self.lb_prestamos:
            self.lb_prestamos.delete(0, "end")
        self.archivos_por_categoria = {
            "nomina": {}, "complementos": {}, "vacaciones": {}, "prestamos": {}
        }

        # Buscar PDFs en carpeta raÃ­z Y subcarpetas
        pdfs = []
        for raiz, dirs, archivos in os.walk(carpeta):
            for f in sorted(archivos):
                if f.lower().endswith(".pdf"):
                    pdfs.append(os.path.join(raiz, f))
        pdfs.sort()


        # Reportar quÃ© hay en la carpeta para diagnÃ³stico
        todos = []
        for raiz, dirs, archivos in os.walk(carpeta):
            for f in archivos:
                todos.append(f)
        self._log(f"Carpeta: {carpeta}")
        self._log(f"Total archivos encontrados: {len(todos)}  |  PDFs: {len(pdfs)}")
        if todos and not pdfs:
            ext_encontradas = list({os.path.splitext(f)[1].lower() for f in todos})
            self._log(f"Tipos de archivo encontrados: {', '.join(ext_encontradas)}", error=True)
            self._log("Esta pestaÃ±a procesa PDFs bancarios. Â¿Los archivos son XML? Usa la pestaÃ±a 'ProvisiÃ³n de NÃ³mina'.", error=True)
            return
        if not pdfs:
            self._log("No encontrÃ© ningÃºn PDF en esa carpeta ni subcarpetas.", error=True)
            return

        self._log(f"Escaneando {len(pdfs)} PDF(s)...")
        no_reconocidos = []

        for path in pdfs:
            categoria, descripcion = clasificar_pdf(path)
            nombre = os.path.basename(path)
            rel = os.path.relpath(path, carpeta)
            etiqueta = f"{rel}   â   {descripcion}"

            if categoria in self.archivos_por_categoria:
                self.archivos_por_categoria[categoria][path] = descripcion
                if categoria == "prestamos":
                    self.lb_prestamos.insert("end", etiqueta)
                    self.lb_prestamos.select_set("end")
                else:
                    self.listboxes[categoria].insert("end", etiqueta)
                    self.listboxes[categoria].select_set("end")
            else:
                no_reconocidos.append(f"{nombre} â {categoria}")
                self._log(f"  Sin clasificar: {nombre} ({categoria})")

        for key in ("nomina", "complementos", "vacaciones", "prestamos"):
            n = len(self.archivos_por_categoria[key])
            self._log(f"  {CATEGORIA_LABELS[key]}: {n} archivo(s)")

        if no_reconocidos:
            msg = "Sin clasificar: " + "; ".join(no_reconocidos)
            self.otros_label.config(text=msg)
        else:
            self.otros_label.config(text="")

        self._log("Escaneo terminado.", ok=True)

    def _seleccionados(self, categoria):
        lb = self.listboxes[categoria]
        paths = list(self.archivos_por_categoria[categoria].keys())
        return [paths[i] for i in lb.curselection()]

    def _seleccionados_prestamos(self):
        paths = list(self.archivos_por_categoria["prestamos"].keys())
        return [paths[i] for i in self.lb_prestamos.curselection()]

    def _ejecutar(self, categoria):
        if cn is None:
            self._log("Falta conciliacion_nomina.py.", error=True)
            return
        archivos = self._seleccionados(categoria)
        if not archivos:
            messagebox.showwarning("Sin selecciÃ³n", f"No hay archivos seleccionados en '{CATEGORIA_LABELS.get(categoria, categoria)}'.")
            return
        try:
            self._conc_pb_frame.grid()
            self._pb_iniciar(self._conc_pb, self._conc_pb_lbl)
        except Exception:
            pass
        threading.Thread(target=self._ejecutar_en_hilo,
                         args=(categoria, archivos,), daemon=True).start()

    def _ejecutar_todo(self):
        if cn is None:
            self._log("Falta conciliacion_nomina.py.", error=True)
            return
        nom  = self._seleccionados("nomina")
        comp = self._seleccionados("complementos")
        vac  = self._seleccionados("vacaciones")
        todos = nom + comp + vac
        if not todos:
            messagebox.showwarning("Nada seleccionado", "No hay archivos seleccionados.")
            return
        try:
            self._conc_pb_frame.grid()
            self._pb_iniciar(self._conc_pb, self._conc_pb_lbl)
        except Exception:
            pass
        threading.Thread(target=self._ejecutar_todo_en_hilo,
                         args=(todos,), daemon=True).start()

    def _ejecutar_todo_en_hilo(self, todos_pdfs):
        _pb      = getattr(self, "_conc_pb",       None)
        _lbl     = getattr(self, "_conc_pb_lbl",   None)
        _pb_frm  = getattr(self, "_conc_pb_frame", None)
        try:
            out_dir  = self._resultados_dir()
            catalogo = self._cargar_catalogo_hilo()
            total    = len(todos_pdfs)
            self.after(0, self._log,
                       f"Generando Excel de Pagos Bancarios con {total} PDF(s)...")
            out_path = os.path.join(out_dir,
                                    self._nombre_salida("PagosBancarios_Consolidado"))
            cn.escribir_pagos_bancarios_todo(todos_pdfs, catalogo, out_path)
            self.after(0, self._log,
                       f"  [OK] {os.path.basename(out_path)}", False, True)
            self.after(0, self._log, "Listo.", False, True)
            self.after(0, self._conc_set_resultado, out_path)
            self.after(0, self._conc_cargar_visor, out_path)
            if _pb: self.after(0, self._pb_detener, _pb, _lbl)
            # Ocultar barra de progreso para que los botones Abrir/Guardar queden visibles
            if _pb_frm: self.after(50, _pb_frm.grid_remove)
        except Exception as exc:
            self.after(0, self._log,
                       f"[Error] {exc}\n{traceback.format_exc()}", True)
            if _pb: self.after(0, self._pb_error, _pb, _lbl)

    def _cargar_catalogo_hilo(self):
        """Carga el catÃ¡logo de cuentas.
        Devuelve {"empleados": {nombre_norm: cuenta}, "prestamos": {nombre_norm: cuenta}}.
        Usa load_poliza() de conciliacion_nomina para leer ambas tablas."""
        catalogo = {"empleados": {}, "prestamos": {}}
        cat_path = self.excel_path_pagos.get().strip()
        if not cat_path or not os.path.isfile(cat_path):
            return catalogo
        try:
            mapa = cn.load_poliza(cat_path)
            for key in ("empleados", "prestamos"):
                df = mapa.get(key)
                if df is not None and not df.empty:
                    catalogo[key] = {idx: str(row["Cuenta"]) for idx, row in df.iterrows()}
            n_emp  = len(catalogo["empleados"])
            n_pres = len(catalogo["prestamos"])
            self.after(0, self._log,
                       f"  CatÃ¡logo: {n_emp} empleado(s), {n_pres} prÃ©stamo(s) cargado(s).")
        except Exception as _e:
            self.after(0, self._log, f"  â  No se pudo leer CatÃ¡logo: {_e}", True)
        return catalogo

    def _conc_cargar_visor(self, path):
        """Carga las primeras filas del Excel en el Treeview del visor."""
        tv = getattr(self, "_conc_tv", None)
        if tv is None:
            return
        for item in tv.get_children():
            tv.delete(item)
        if not path or not os.path.isfile(path):
            return
        try:
            import openpyxl as _opxl
            wb = _opxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not all_rows:
                return
            headers = [str(c) if c is not None else "" for c in all_rows[0]]
            tv["columns"] = headers
            tv["show"] = "headings"
            for h in headers:
                tv.heading(h, text=h)
                tv.column(h, width=160, anchor="w", stretch=True)
            for row in all_rows[1:51]:
                vals = [str(c) if c is not None else "" for c in row]
                tv.insert("", "end", values=vals)
            self._log(f"  Vista previa: {min(len(all_rows)-1, 50)} fila(s).")
        except Exception as _e:
            self._log(f"  â  Vista previa: error al leer Excel â {_e}", True)
            tv["columns"] = ["error"]
            tv.heading("error", text="No se pudo cargar la vista previa")
            tv.insert("", "end", values=(str(_e),))

    # ââ Helpers genÃ©ricos de barra de progreso ââââââââââââââââââââââââ #
    def _pb_iniciar(self, bar=None, lbl=None):
        """Inicia la barra global y la barra de mÃ³dulo (si se pasa)."""
        try:
            # Barra global
            self._global_pb["value"] = 0
            self._global_pb_lbl.config(text="Procesando...")
            self._pb_step(self._global_pb, self._global_pb_lbl)
        except Exception:
            pass
        try:
            # Barra de mÃ³dulo (opcional)
            if bar is not None and bar is not getattr(self, "_global_pb", None):
                bar["value"] = 0
                if lbl:
                    lbl.config(text="0 %")
        except Exception:
            pass

    def _pb_step(self, bar, lbl):
        try:
            val = bar["value"]
            if val < 95:
                bar["value"] = min(val + 1.5, 95)
                pct = int(bar["value"])
                if lbl:
                    lbl.config(text=f"{pct} %" if bar is not self._global_pb else f"{pct}%  Procesando...")
                if isinstance(bar, FunkyProgressBar):
                    bar.animate_step()
                bar._after_id = self.after(80, self._pb_step, bar, lbl)
        except Exception:
            pass

    def _pb_detener(self, bar=None, lbl=None, texto="â Listo"):
        """Detiene animaciÃ³n y lleva al 100%."""
        try:
            # Detener barra global
            if hasattr(self._global_pb, "_after_id") and self._global_pb._after_id:
                self.after_cancel(self._global_pb._after_id)
                self._global_pb._after_id = None
            self._global_pb["value"] = 100
            if isinstance(self._global_pb, FunkyProgressBar):
                self._global_pb._draw()
            self._global_pb_lbl.config(text="â Listo")
            # Resetear a 0 despuÃ©s de 2 segundos
            self.after(2000, self._pb_reset_global)
        except Exception:
            pass
        try:
            if bar is not None and bar is not getattr(self, "_global_pb", None):
                if hasattr(bar, "_after_id") and bar._after_id:
                    self.after_cancel(bar._after_id)
                    bar._after_id = None
                bar["value"] = 100
                if isinstance(bar, FunkyProgressBar):
                    bar._draw()
                if lbl:
                    lbl.config(text=texto)
        except Exception:
            pass

    def _pb_reset_global(self):
        """Regresa la barra global a 0 en estado reposo."""
        try:
            if self._global_pb["value"] >= 100:
                self._global_pb["value"] = 0
                self._global_pb_lbl.config(text="Listo")
        except Exception:
            pass

    def _pb_error(self, bar, lbl=None):
        """Detiene la barra de progreso y marca error."""
        try:
            if hasattr(bar, "_after_id") and bar._after_id:
                self.after_cancel(bar._after_id)
                bar._after_id = None
            bar["value"] = 0
            if lbl:
                lbl.config(text="Error")
        except Exception:
            pass

    # ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    # MÃDULO â CONCILIACIÃN BANCO VS AUXILIAR
    # ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    _CBA_ACENTO   = "#6B3FA0"
    _CBA_ACENTO_L = "#EDE7F6"
    _DEP_ACENTO   = "#1565C0"
    _DEP_ACENTO_L = "#E3F2FD"

    def _tab_conc_banco_aux(self, nb):
        outer = ttk.Frame(nb)
        nb.add(outer, text="  ð Banco vs Auxiliar  ")
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(1, weight=1)

        self._cba_ruta_banco = tk.StringVar(value="")
        self._cba_ruta_aux   = tk.StringVar(value="")
        self._cba_resultado  = None
        self._cba_wb         = None
        self._cba_wb_name    = ""
        self._cba_wb_dir     = ""

        # ââ Panel de controles ââââââââââââââââââââââââââââââââââââââââ
        ctrl = tk.Frame(outer, bg=COLOR_FONDO)
        ctrl.grid(row=0, column=0, sticky="ew")
        ctrl.columnconfigure(0, weight=1)

        # Titulo
        hdr = tk.Frame(ctrl, bg=COLOR_FONDO)
        hdr.grid(row=0, column=0, sticky="ew", padx=16, pady=(8, 4))
        tk.Label(hdr, text="Conciliacion Banco vs Auxiliar Contable",
                 bg=COLOR_FONDO, fg=self._CBA_ACENTO,
                 font=("Segoe UI", 12, "bold")).pack(side="left")

        # Selectores de archivo
        fsel = tk.Frame(ctrl, bg=COLOR_FONDO)
        fsel.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 4))
        fsel.columnconfigure(1, weight=1)
        fsel.columnconfigure(4, weight=1)

        tk.Label(fsel, text="Banco (.xlsx):", bg=COLOR_FONDO,
                 fg=COLOR_TEXTO, font=("Segoe UI", 9, "bold")).grid(
                 row=0, column=0, sticky="w", padx=(0, 6))
        self._cba_lbl_banco = tk.Label(fsel, text="Sin seleccionar",
            bg=COLOR_BLANCO, fg="#999999", font=("Segoe UI", 9),
            relief="sunken", padx=6, pady=2, cursor="hand2")
        self._cba_lbl_banco.grid(row=0, column=1, sticky="ew", padx=(0, 6))
        self._cba_lbl_banco.bind("<Button-1>", lambda e: self._cba_elegir_banco())
        ttk.Button(fsel, text="Banco",
                   command=self._cba_elegir_banco).grid(row=0, column=2, padx=(0, 18))

        tk.Label(fsel, text="Auxiliar (.xlsx):", bg=COLOR_FONDO,
                 fg=COLOR_TEXTO, font=("Segoe UI", 9, "bold")).grid(
                 row=0, column=3, sticky="w", padx=(0, 6))
        self._cba_lbl_aux = tk.Label(fsel, text="Sin seleccionar",
            bg=COLOR_BLANCO, fg="#999999", font=("Segoe UI", 9),
            relief="sunken", padx=6, pady=2, cursor="hand2")
        self._cba_lbl_aux.grid(row=0, column=4, sticky="ew", padx=(0, 6))
        self._cba_lbl_aux.bind("<Button-1>", lambda e: self._cba_elegir_aux())
        ttk.Button(fsel, text="Auxiliar",
                   command=self._cba_elegir_aux).grid(row=0, column=5, padx=0)

        # Boton generar
        btn_row = tk.Frame(ctrl, bg=COLOR_FONDO)
        btn_row.grid(row=2, column=0, sticky="ew", padx=16, pady=(4, 4))
        btn_row.columnconfigure(0, weight=1)
        self._cba_btn_gen = tk.Button(btn_row,
            text="  â Generar Conciliacion",
            bg=self._CBA_ACENTO, fg=COLOR_BLANCO,
            font=("Segoe UI", 10, "bold"), relief="flat",
            padx=16, pady=8, cursor="hand2",
            activebackground="#4A2070",
            command=self._cba_generar)
        self._cba_btn_gen.grid(row=0, column=0, sticky="ew")

        # Barra de progreso
        self._cba_pb_frame = tk.Frame(ctrl, bg=COLOR_FONDO)
        self._cba_pb_frame.grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 2))
        self._cba_pb_frame.grid_remove()
        self._cba_pb = FunkyProgressBar(self._cba_pb_frame, maximum=100, height=14)
        self._cba_pb.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self._cba_pb_lbl = ttk.Label(self._cba_pb_frame, text="", width=10,
            foreground=self._CBA_ACENTO, font=("Segoe UI", 8, "bold"))
        self._cba_pb_lbl.pack(side="left")

        # Barra de resultado
        tb = tk.Frame(ctrl, bg=COLOR_BLANCO,
                      highlightbackground="#D1C4E9", highlightthickness=1)
        tb.grid(row=4, column=0, sticky="ew", padx=16, pady=(0, 6))
        tk.Label(tb, text="Resultado:", bg=COLOR_BLANCO, fg=COLOR_TEXTO,
                 font=("Segoe UI", 8)).pack(side="left", padx=(10, 2), pady=3)
        self._cba_lbl_arch = tk.Label(tb, text="---", bg=COLOR_BLANCO,
            fg=self._CBA_ACENTO, font=("Segoe UI", 8, "bold"))
        self._cba_lbl_arch.pack(side="left", padx=(0, 10), pady=3)
        self._cba_btn_guardar = ttk.Button(tb, text="ð¾ Guardar como",
            command=self._cba_guardar_como, state="disabled")
        self._cba_btn_guardar.pack(side="right", padx=(6, 2), pady=3)
        self._cba_btn_abrir = ttk.Button(tb, text="ð Abrir en Excel",
            command=self._cba_abrir, state="disabled")
        self._cba_btn_abrir.pack(side="right", padx=6, pady=3)

        # ââ Area de trabajo âââââââââââââââââââââââââââââââââââââââââââ
        ttk.Separator(outer, orient="horizontal").grid(
            row=0, column=0, sticky="sew")   # al fondo del row=0

        work = tk.Frame(outer, bg=COLOR_FONDO)
        work.grid(row=1, column=0, sticky="nsew")
        work.columnconfigure(0, weight=1)
        work.rowconfigure(2, weight=1)

        # Header area de trabajo
        lbl_area = tk.Frame(work, bg=self._CBA_ACENTO)
        lbl_area.grid(row=0, column=0, columnspan=2, sticky="ew")
        tk.Label(lbl_area, text="  Area de Trabajo  â  Vista previa",
                 bg=self._CBA_ACENTO, fg=COLOR_BLANCO,
                 font=("Segoe UI", 9, "bold")).pack(side="left", pady=4)
        self._cba_lbl_resumen = tk.Label(lbl_area, text="",
            bg=self._CBA_ACENTO, fg="#CE93D8", font=("Segoe UI", 8))
        self._cba_lbl_resumen.pack(side="right", padx=10)

        # Leyenda de colores
        leyenda_bar = tk.Frame(work, bg=COLOR_FONDO)
        leyenda_bar.grid(row=1, column=0, columnspan=2, sticky="ew", padx=8, pady=(3, 1))
        for txt_l, bg_l, fg_l in [
                ("â Deposito conciliado",  "#C8E6C9", "#1B5E20"),
                ("â Retiro conciliado",    "#BBDEFB", "#1F3864"),
                ("â No conciliado",        "#FFCDD2", "#B71C1C"),
                ("â Solo en banco",        "#FFF9C4", "#7D5B00")]:
            tk.Label(leyenda_bar, text=f"  {txt_l}  ",
                     bg=bg_l, fg=fg_l,
                     font=("Segoe UI", 8), relief="flat",
                     padx=4, pady=1).pack(side="left", padx=2)

        # Treeview
        cols_tree = ("Fecha", "Descripcion / Cliente", "Deposito", "Retiro", "Estado", "Monto Aux")
        self._cba_tree = ttk.Treeview(work, columns=cols_tree,
                                       show="headings", selectmode="browse")
        col_w   = {"Fecha": 100, "Descripcion / Cliente": 380,
                   "Deposito": 120, "Retiro": 120, "Estado": 170, "Monto Aux": 120}
        col_anc = {"Fecha": "center", "Descripcion / Cliente": "w",
                   "Deposito": "e", "Retiro": "e", "Estado": "center", "Monto Aux": "e"}
        for c in cols_tree:
            self._cba_tree.heading(c, text=c)
            self._cba_tree.column(c, width=col_w.get(c, 120),
                minwidth=60, anchor=col_anc.get(c, "w"), stretch=True)
        self._cba_tree.grid(row=2, column=0, sticky="nsew")

        # Tags de color en treeview
        self._cba_tree.tag_configure("conc",       foreground="#1B5E20", background="#F1F8E9")
        self._cba_tree.tag_configure("conc_ret",   foreground="#1F3864", background="#E3F2FD")
        self._cba_tree.tag_configure("no_conc",    foreground="#B71C1C", background="#FFEBEE")
        self._cba_tree.tag_configure("solo_banco", foreground="#7D5B00", background="#FFFDE7")
        self._cba_tree.tag_configure("retiro",     foreground="#1F3864", background="#EFF6FF")

        vsb = ttk.Scrollbar(work, orient="vertical",   command=self._cba_tree.yview)
        hsb = ttk.Scrollbar(work, orient="horizontal", command=self._cba_tree.xview)
        self._cba_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.grid(row=2, column=1, sticky="ns")
        hsb.grid(row=3, column=0, sticky="ew")

    def _cba_elegir_banco(self):
        self.focus_force(); self.update()
        ruta = filedialog.askopenfilename(
            parent=self, title="Seleccionar Excel del Estado de Cuenta (Banco)",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if ruta:
            self._cba_ruta_banco.set(ruta)
            self._cba_lbl_banco.config(
                text=os.path.basename(ruta), fg="#1B5E20",
                font=("Segoe UI", 9, "bold"))
            self._log(f"Banco cargado: {os.path.basename(ruta)}")

    def _cba_elegir_aux(self):
        self.focus_force(); self.update()
        ruta = filedialog.askopenfilename(
            parent=self, title="Seleccionar Excel del Auxiliar Contable",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if ruta:
            self._cba_ruta_aux.set(ruta)
            self._cba_lbl_aux.config(
                text=os.path.basename(ruta), fg="#1B5E20",
                font=("Segoe UI", 9, "bold"))
            self._log(f"Auxiliar cargado: {os.path.basename(ruta)}")

    def _cba_generar(self):
        banco = self._cba_ruta_banco.get().strip()
        aux   = self._cba_ruta_aux.get().strip()
        if not banco:
            messagebox.showwarning("Archivo faltante",
                "Selecciona el Excel del estado de cuenta bancario.", parent=self)
            return
        if not aux:
            messagebox.showwarning("Archivo faltante",
                "Selecciona el Excel del auxiliar contable.", parent=self)
            return
        for item in self._cba_tree.get_children():
            self._cba_tree.delete(item)
        self._cba_lbl_resumen.config(text="")
        self._cba_lbl_arch.config(text="---")
        self._cba_btn_abrir.config(state="disabled")
        self._cba_btn_guardar.config(state="disabled")
        self._cba_pb_frame.grid()
        self._pb_iniciar(self._cba_pb, self._cba_pb_lbl)
        self._cba_btn_gen.config(state="disabled")
        import threading
        threading.Thread(target=self._cba_hilo,
                         args=(banco, aux), daemon=True).start()

    def _cba_hilo(self, ruta_banco, ruta_aux):
        import re as _re, io as _io, os as _os
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        except ImportError:
            self.after(0, self._log, "Falta openpyxl", True)
            self.after(0, self._pb_error, self._cba_pb, self._cba_pb_lbl)
            self.after(0, self._cba_pb_frame.grid_remove)
            self.after(0, lambda: self._cba_btn_gen.config(state="normal"))
            return

        def _log(msg, err=False):
            self.after(0, self._log, msg, err)

        def _num(row, c):
            v = row[c] if len(row) > c else None
            return float(v) if isinstance(v, (int, float)) else None

        def _str(row, c):
            v = row[c] if len(row) > c else None
            return str(v) if v is not None else ""

        def _parse_client(dd, dp):
            m = _re.search(r"CLIENTE\((.+?)\)#", dd)
            if m: return m.group(1)
            m2 = _re.search(r"\[P:\d+\](.+)", dd)
            if m2: return m2.group(1).strip()
            return dd.strip()[:60] or dp[:60]

        try:
            _log(f"Banco:    {_os.path.basename(ruta_banco)}")
            _log(f"Auxiliar: {_os.path.basename(ruta_aux)}")

            # ââ Leer Banco ââââââââââââââââââââââââââââââââââââââââââââ
            _log("Leyendo estado de cuenta bancario...")
            wb1 = openpyxl.load_workbook(ruta_banco, read_only=True, data_only=True)
            ws1 = wb1.active
            rows1 = list(ws1.iter_rows(values_only=True))
            wb1.close()

            # Detectar encabezado banco (sin limite de filas)
            hdr_row = 1; col_f=0; col_d=1; col_dep=2; col_ret=3; col_sal=4
            for i, row in enumerate(rows1):
                vals = [str(v).strip().lower() if v else "" for v in row]
                if "fecha" in vals:
                    hdr_row = i + 1
                    col_f   = vals.index("fecha")
                    col_d   = next((j for j,v in enumerate(vals) if "desc" in v), 1)
                    col_dep = next((j for j,v in enumerate(vals) if "dep"  in v), 2)
                    col_ret = next((j for j,v in enumerate(vals) if "ret"  in v), 3)
                    col_sal = next((j for j,v in enumerate(vals) if "saldo" in v), 4)
                    break

            bank = []
            for row in rows1[hdr_row:]:
                if len(row) <= col_f or row[col_f] is None: continue
                bank.append({
                    "fecha":    row[col_f],
                    "desc":     _str(row, col_d),
                    "deposito": _num(row, col_dep),
                    "retiro":   _num(row, col_ret),
                    "saldo":    _num(row, col_sal),
                })
            deposits = [b for b in bank if b["deposito"] and b["deposito"] > 0]
            retiros  = [b for b in bank if b["retiro"]   and b["retiro"]   > 0]
            _log(f"Banco: {len(bank)} movimientos ({len(deposits)} dep, {len(retiros)} ret).")

            # ââ Leer Auxiliar âââââââââââââââââââââââââââââââââââââââââ
            _log("Leyendo auxiliar contable...")
            wb2 = openpyxl.load_workbook(ruta_aux, read_only=True, data_only=True)
            ws2 = wb2.active
            rows2 = list(ws2.iter_rows(values_only=True))
            wb2.close()

            # Detectar encabezado auxiliar (sin limite de filas)
            hdr2 = -1; cf=4; cp=3; cdet=7; cdp=8; ccargo=9; cabono=10
            for i, row in enumerate(rows2):
                vals = [str(v).strip().lower() if v else "" for v in row]
                if "cargo" in vals and "fecha" in vals:
                    hdr2   = i
                    cf     = vals.index("fecha")
                    ccargo = vals.index("cargo")
                    cabono = next((j for j,v in enumerate(vals) if "abono"   in v), cabono)
                    cp     = next((j for j,v in enumerate(vals) if "num" in v or "num. p" in v), cp)
                    cdet   = next((j for j,v in enumerate(vals) if "detalle" in v), cdet)
                    cdp    = next((j for j,v in enumerate(vals) if "desc" in v and j != cdet), cdp)
                    break
            data_start = hdr2 + 1 if hdr2 >= 0 else 0
            _log(f"Auxiliar: encabezado en fila {hdr2+1 if hdr2>=0 else 'N/A'} | Fecha={cf} Cargo={ccargo} Abono={cabono}")

            from datetime import datetime as _dt
            aux_cargo = []; aux_abono = []
            for row in rows2[data_start:]:
                if len(row) <= cf or row[cf] is None: continue
                if not isinstance(row[cf], _dt): continue
                cargo = _num(row, ccargo) or 0
                abono = _num(row, cabono) or 0
                dd = _str(row, cdet); dp = _str(row, cdp)
                pol = row[cp] if len(row) > cp else ""
                client = _parse_client(dd, dp)
                if cargo > 0:
                    aux_cargo.append({"fecha":row[cf],"poliza":pol,"client":client,"monto":cargo,"dd":dd,"dp":dp})
                if abono > 0:
                    aux_abono.append({"fecha":row[cf],"poliza":pol,"client":client,"monto":abono,"dd":dd,"dp":dp})
            _log(f"Auxiliar: {len(aux_cargo)} cargos, {len(aux_abono)} abonos.")

            # ââ Matching depositos <-> cargos âââââââââââââââââââââââââ
            used_dep = set(); used_cargo = set()
            for i, a in enumerate(aux_cargo):
                best, best_d = None, 999
                for j, b in enumerate(deposits):
                    if j in used_dep: continue
                    if abs(b["deposito"] - a["monto"]) <= 0.05:
                        try: delta = abs((b["fecha"] - a["fecha"]).days)
                        except: delta = 0
                        if delta < best_d: best, best_d = j, delta
                if best is not None:
                    used_dep.add(best); used_cargo.add(i)
                    aux_cargo[i]["match"] = deposits[best]; deposits[best]["match_aux"] = a

            # ââ Matching retiros <-> abonos âââââââââââââââââââââââââââ
            used_ret = set(); used_abono = set()
            for i, a in enumerate(aux_abono):
                best, best_d = None, 999
                for j, b in enumerate(retiros):
                    if j in used_ret: continue
                    if abs(b["retiro"] - a["monto"]) <= 0.05:
                        try: delta = abs((b["fecha"] - a["fecha"]).days)
                        except: delta = 0
                        if delta < best_d: best, best_d = j, delta
                if best is not None:
                    used_ret.add(best); used_abono.add(i)
                    aux_abono[i]["match"] = retiros[best]; retiros[best]["match_aux"] = a

            n_conc_dep = len(used_cargo);  n_nconc_dep = len(aux_cargo) - n_conc_dep
            n_conc_ret = len(used_abono);  n_nconc_ret = len(aux_abono) - n_conc_ret
            _log(f"Depositos: {n_conc_dep}/{len(aux_cargo)} conciliados.")
            _log(f"Retiros:   {n_conc_ret}/{len(aux_abono)} conciliados.")

            # ââ Estilos âââââââââââââââââââââââââââââââââââââââââââââââ
            thin = Side(style="thin", color="BFBFBF")
            BORD    = Border(left=thin, right=thin, top=thin, bottom=thin)
            HDR_FL  = PatternFill("solid", fgColor="1F3864")
            HDR2_FL = PatternFill("solid", fgColor="4A1E8A")
            HDR_FN  = Font(color="FFFFFF", bold=True)
            C_GREEN = PatternFill("solid", fgColor="C6EFCE"); F_GREEN = Font(color="375623")
            C_RED   = PatternFill("solid", fgColor="FFC7CE"); F_RED   = Font(color="9C0006")
            C_YELL  = PatternFill("solid", fgColor="FFEB9C"); F_YELL  = Font(color="7D5B00")
            C_BLUE  = PatternFill("solid", fgColor="DDEBF7"); F_BLUE  = Font(color="1F3864")
            GRAY_FL = PatternFill("solid", fgColor="F2F2F2")

            def _hdr(cell, text, fill=None):
                cell.value = text; cell.fill = fill or HDR_FL; cell.font = HDR_FN
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = BORD

            def _srow(ws, r, nc, fill, font):
                for c in range(1, nc + 1):
                    ws.cell(r, c).fill = fill; ws.cell(r, c).font = font
                    ws.cell(r, c).border = BORD

            wb_out = openpyxl.Workbook()

            # ââ Hoja Depositos ââââââââââââââââââââââââââââââââââââââââ
            ws_dep = wb_out.active; ws_dep.title = "Depositos"
            ws_dep.merge_cells("A1:F1")
            t = ws_dep["A1"]
            t.value = "BANCO - " + _os.path.basename(ruta_banco).upper() + "  |  Depositos vs Auxiliar"
            t.fill = HDR_FL; t.font = Font(color="FFFFFF", bold=True, size=12)
            t.alignment = Alignment(horizontal="center", vertical="center")
            ws_dep.row_dimensions[1].height = 22
            for c, h in enumerate(["Fecha","Descripcion","Deposito","Estado","Monto Auxiliar","Cliente / Poliza"], 1):
                _hdr(ws_dep.cell(2, c), h)
            ws_dep.row_dimensions[2].height = 28; ws_dep.freeze_panes = "A3"
            ri = 3
            for i, a in enumerate(aux_cargo):
                matched = i in used_cargo; b = a.get("match", {})
                fill = C_GREEN if matched else C_RED
                font = F_GREEN if matched else F_RED
                estado = "Conciliado" if matched else "No Conciliado"
                fecha_s = (b.get("fecha") or a["fecha"])
                desc_s  = b.get("desc", "") if matched else (a["dp"] or a["dd"] or a["client"])
                monto_b = b.get("deposito", "") if matched else ""
                pol_s   = f"Pol.{a['poliza']} - {a['client']}"
                ws_dep.cell(ri, 1, fecha_s).number_format = "DD/MM/YYYY"
                ws_dep.cell(ri, 2, desc_s)
                if monto_b != "": ws_dep.cell(ri, 3, monto_b).number_format = "#,##0.00"
                ws_dep.cell(ri, 4, estado)
                ws_dep.cell(ri, 5, a["monto"]).number_format = "#,##0.00"
                ws_dep.cell(ri, 6, pol_s)
                _srow(ws_dep, ri, 6, fill, font); ri += 1
            for b in [x for j,x in enumerate(deposits) if j not in used_dep]:
                ws_dep.cell(ri, 1, b["fecha"]).number_format = "DD/MM/YYYY"
                ws_dep.cell(ri, 2, b["desc"])
                ws_dep.cell(ri, 3, b["deposito"]).number_format = "#,##0.00"
                ws_dep.cell(ri, 4, "Solo en Banco")
                _srow(ws_dep, ri, 6, C_YELL, F_YELL); ri += 1
            for col, w in zip("ABCDEF", [14,60,14,18,14,40]):
                ws_dep.column_dimensions[col].width = w

            # ââ Hoja Retiros ââââââââââââââââââââââââââââââââââââââââââ
            ws_ret = wb_out.create_sheet("Retiros")
            ws_ret.merge_cells("A1:F1")
            t2 = ws_ret["A1"]
            t2.value = "BANCO - " + _os.path.basename(ruta_banco).upper() + "  |  Retiros vs Auxiliar"
            t2.fill = HDR2_FL; t2.font = Font(color="FFFFFF", bold=True, size=12)
            t2.alignment = Alignment(horizontal="center", vertical="center")
            ws_ret.row_dimensions[1].height = 22
            for c, h in enumerate(["Fecha","Descripcion","Retiro","Estado","Monto Auxiliar","Descripcion Poliza"], 1):
                _hdr(ws_ret.cell(2, c), h, HDR2_FL)
            ws_ret.row_dimensions[2].height = 28; ws_ret.freeze_panes = "A3"
            ri = 3
            for i, a in enumerate(aux_abono):
                matched = i in used_abono; b = a.get("match", {})
                fill = C_BLUE if matched else C_RED
                font = F_BLUE if matched else F_RED
                estado = "Conciliado" if matched else "No Conciliado"
                fecha_s = (b.get("fecha") or a["fecha"])
                desc_s  = b.get("desc", "") if matched else (a["dp"] or a["dd"] or "")
                monto_b = b.get("retiro", "") if matched else ""
                ws_ret.cell(ri, 1, fecha_s).number_format = "DD/MM/YYYY"
                ws_ret.cell(ri, 2, desc_s)
                if monto_b != "": ws_ret.cell(ri, 3, monto_b).number_format = "#,##0.00"
                ws_ret.cell(ri, 4, estado)
                ws_ret.cell(ri, 5, a["monto"]).number_format = "#,##0.00"
                ws_ret.cell(ri, 6, a["dd"])
                _srow(ws_ret, ri, 6, fill, font); ri += 1
            for b in [x for j,x in enumerate(retiros) if j not in used_ret]:
                ws_ret.cell(ri, 1, b["fecha"]).number_format = "DD/MM/YYYY"
                ws_ret.cell(ri, 2, b["desc"])
                ws_ret.cell(ri, 3, b["retiro"]).number_format = "#,##0.00"
                ws_ret.cell(ri, 4, "Solo en Banco")
                _srow(ws_ret, ri, 6, C_YELL, F_YELL); ri += 1
            for col, w in zip("ABCDEF", [14,60,14,18,14,50]):
                ws_ret.column_dimensions[col].width = w

            # ââ Hoja Resumen ââââââââââââââââââââââââââââââââââââââââââ
            ws_r = wb_out.create_sheet("Resumen")
            ws_r.merge_cells("A1:C1")
            t3 = ws_r["A1"]
            t3.value = "RESUMEN DE CONCILIACION"
            t3.fill = HDR_FL; t3.font = Font(color="FFFFFF", bold=True, size=13)
            t3.alignment = Alignment(horizontal="center", vertical="center")
            ws_r.row_dimensions[1].height = 26
            solo_dep_n = len(deposits) - len(used_dep)
            solo_ret_n = len(retiros)  - len(used_ret)
            datos_res = [
                ("DEPOSITOS","",""),
                ("Total polizas auxiliar", len(aux_cargo), ""),
                ("Conciliadas", n_conc_dep, ""),
                ("No conciliadas", n_nconc_dep, ""),
                ("Solo en banco (sin poliza)", solo_dep_n, ""),
                ("","",""),
                ("RETIROS / PAGOS","",""),
                ("Total polizas auxiliar", len(aux_abono), ""),
                ("Conciliadas", n_conc_ret, ""),
                ("No conciliadas", n_nconc_ret, ""),
                ("Solo en banco (sin poliza)", solo_ret_n, ""),
                ("","",""),
                ("BANCO TOTAL","",""),
                ("Total movimientos", len(bank), ""),
                ("Depositos", len(deposits), ""),
                ("Retiros", len(retiros), ""),
            ]
            for ri2, (label, val, fmt) in enumerate(datos_res, 2):
                ws_r.cell(ri2, 1, label)
                if val != "":
                    c = ws_r.cell(ri2, 2, val)
                    if fmt: c.number_format = fmt
                if label in ("DEPOSITOS","RETIROS / PAGOS","BANCO TOTAL"):
                    ws_r.cell(ri2, 1).font = Font(bold=True, color="1F3864", size=11)
                    ws_r.cell(ri2, 1).fill = GRAY_FL
            r_det = len(datos_res) + 4
            ws_r.merge_cells(f"A{r_det}:D{r_det}")
            h_det = ws_r.cell(r_det, 1, "DEPOSITOS NO CONCILIADOS (DETALLE)")
            h_det.font = Font(bold=True, color="B71C1C"); h_det.fill = GRAY_FL
            r_det += 1
            for hdx, htxt in enumerate(["Fecha","Cliente","Monto","Nota"], 1):
                _hdr(ws_r.cell(r_det, hdx), htxt)
            r_det += 1
            for i, a in enumerate(aux_cargo):
                if i not in used_cargo:
                    ws_r.cell(r_det, 1, a["fecha"]).number_format = "DD/MM/YYYY"
                    ws_r.cell(r_det, 2, a["client"])
                    ws_r.cell(r_det, 3, a["monto"]).number_format = "#,##0.00"
                    close = [b for b in deposits if abs((b["deposito"] or 0) - a["monto"]) < 200]
                    ws_r.cell(r_det, 4, f"Mas cercano: ${close[0]['deposito']:,.2f}" if close else "No encontrado")
                    _srow(ws_r, r_det, 4, C_RED, F_RED); r_det += 1
            for col, w in zip("ABCD", [28,34,16,55]):
                ws_r.column_dimensions[col].width = w

            buf = _io.BytesIO()
            wb_out.save(buf); buf.seek(0)
            self._cba_wb = wb_out
            nombre_base = "CONCILIACION_" + _os.path.splitext(_os.path.basename(ruta_banco))[0].upper() + ".xlsx"
            self._cba_wb_name = nombre_base
            self._cba_wb_dir  = _os.path.dirname(ruta_banco)

            import tempfile
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", prefix="conc_bco_aux_", delete=False)
            tmp.write(buf.getvalue()); tmp.close()
            self._cba_resultado = tmp.name

            resumen_txt = (f"Dep: {n_conc_dep}/{len(aux_cargo)}  |  "
                           f"Ret: {n_conc_ret}/{len(aux_abono)}  |  "
                           f"No conciliados: {n_nconc_dep} dep, {n_nconc_ret} ret")
            _log(f"Listo. {resumen_txt}")
            self.after(0, self._pb_detener, self._cba_pb, self._cba_pb_lbl)
            self.after(100, self._cba_pb_frame.grid_remove)
            self.after(0, lambda: self._cba_btn_gen.config(state="normal"))
            self.after(0, lambda nb=nombre_base: self._cba_lbl_arch.config(text=nb))
            self.after(0, lambda: self._cba_btn_abrir.config(state="normal"))
            self.after(0, lambda: self._cba_btn_guardar.config(state="normal"))
            self.after(0, lambda t=resumen_txt: self._cba_lbl_resumen.config(text=t))

            # ââ Treeview ââââââââââââââââââââââââââââââââââââââââââââââ
            tree_data = []
            for i, b in enumerate(bank):
                fecha_s = b["fecha"].strftime("%d/%m/%Y") if hasattr(b["fecha"], "strftime") else str(b["fecha"])
                if b["deposito"] and b["deposito"] > 0:
                    ma = b.get("match_aux")
                    if ma:
                        tag = "conc"; estado = "Conciliado"
                        monto_aux = f"${ma['monto']:,.2f}"
                        desc_show = ma["client"] + "  -  " + b["desc"][:70]
                    else:
                        tag = "solo_banco"; estado = "Solo Banco"
                        monto_aux = ""; desc_show = b["desc"][:100]
                    tree_data.append((fecha_s, desc_show, f"${b['deposito']:,.2f}", "", estado, monto_aux, tag))
                elif b["retiro"] and b["retiro"] > 0:
                    ma = b.get("match_aux")
                    if ma:
                        tag = "conc_ret"; estado = "Retiro Conc."
                        monto_aux = f"${ma['monto']:,.2f}"
                    else:
                        tag = "retiro"; estado = "Retiro"
                        monto_aux = ""
                    tree_data.append((fecha_s, b["desc"][:100], "",
                        f"${b['retiro']:,.2f}", estado, monto_aux, tag))

            def _fill(data):
                for d in data:
                    self._cba_tree.insert("", "end", values=d[:6], tags=(d[6],))
            self.after(0, lambda d=tree_data: _fill(d))

        except Exception as exc:
            import traceback
            _log(f"Error: {exc}", True)
            _log(traceback.format_exc(), True)
            self.after(0, self._pb_error, self._cba_pb, self._cba_pb_lbl)
            self.after(100, self._cba_pb_frame.grid_remove)
            self.after(0, lambda: self._cba_btn_gen.config(state="normal"))

    def _cba_abrir(self):
        ruta = getattr(self, "_cba_resultado", None)
        if ruta and os.path.exists(ruta):
            import subprocess
            try:
                subprocess.Popen(["start", "", ruta], shell=True)
            except Exception:
                pass

    def _cba_guardar_como(self):
        wb = getattr(self, "_cba_wb", None)
        if wb is None:
            messagebox.showwarning("Sin datos",
                "Primero genera la conciliacion.", parent=self)
            return
        self.focus_force(); self.update()
        ruta = filedialog.asksaveasfilename(
            parent=self,
            title="Guardar conciliacion como...",
            initialdir=getattr(self, "_cba_wb_dir", os.getcwd()),
            initialfile=getattr(self, "_cba_wb_name", "conciliacion.xlsx"),
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
        )
        if not ruta:
            return
        try:
            wb.save(ruta)
            self._cba_resultado = ruta
            self._cba_wb_dir  = os.path.dirname(ruta)
            self._cba_wb_name = os.path.basename(ruta)
            self._cba_lbl_arch.config(text=os.path.basename(ruta))
            self._cba_btn_guardar.config(state="disabled")
            if hasattr(self, "_visor_carpeta"):
                self._visor_carpeta.set(os.path.dirname(ruta))
            if hasattr(self, "_visor_refrescar"):
                self._visor_refrescar()
            messagebox.showinfo("Guardado",
                f"Archivo guardado en:\n{ruta}", parent=self)
        except Exception as e:
            messagebox.showerror("Error al guardar",
                f"No se pudo guardar:\n{e}", parent=self)

    # ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    # MÃDULO: DEPÃSITOS BANCARIOS
    # ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

    def _tab_depositos_bancarios(self, nb):
        outer = ttk.Frame(nb)
        nb.add(outer, text="  ð¦ DepÃ³sitos Bancarios  ")
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(1, weight=1)

        self._dep_ruta_bbva       = tk.StringVar(value="")
        self._dep_ruta_banorte    = tk.StringVar(value="")
        self._dep_ruta_inbursa    = tk.StringVar(value="")
        self._dep_ruta_plantilla  = tk.StringVar(value="")
        self._dep_wb           = None
        self._dep_resultado    = None
        self._dep_wb_name      = ""
        self._dep_wb_dir       = ""

        # ââ Panel de controles ââââââââââââââââââââââââââââââââââââââââââââââââââ
        ctrl = tk.Frame(outer, bg=COLOR_FONDO)
        ctrl.grid(row=0, column=0, sticky="ew")
        ctrl.columnconfigure(0, weight=1)

        hdr = tk.Frame(ctrl, bg=COLOR_FONDO)
        hdr.grid(row=0, column=0, sticky="ew", padx=16, pady=(8, 4))
        tk.Label(hdr, text="DepÃ³sitos Bancarios  â  BBVA Â· Banorte Â· Inbursa",
                 bg=COLOR_FONDO, fg=self._DEP_ACENTO,
                 font=("Segoe UI", 12, "bold")).pack(side="left")

        # Selectores de archivo (3 bancos)
        fsel = tk.Frame(ctrl, bg=COLOR_FONDO)
        fsel.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 4))
        fsel.columnconfigure(1, weight=1)
        fsel.columnconfigure(4, weight=1)
        fsel.columnconfigure(7, weight=1)

        bancos_cfg = [
            ("ð¦ BBVA (.xlsx):",    "_dep_lbl_bbva",    self._dep_elegir_bbva,    "#0D47A1", 0),
            ("ð¥ Banorte (.xlsx):", "_dep_lbl_banorte", self._dep_elegir_banorte, "#B71C1C", 3),
            ("ð§ Inbursa (.xlsx):", "_dep_lbl_inbursa", self._dep_elegir_inbursa, "#BF360C", 6),
        ]
        for lbl_txt, lbl_attr, cmd, fg_sel, c0 in bancos_cfg:
            tk.Label(fsel, text=lbl_txt, bg=COLOR_FONDO, fg=COLOR_TEXTO,
                     font=("Segoe UI", 9, "bold")).grid(
                     row=0, column=c0, sticky="w", padx=(12 if c0 > 0 else 0, 4))
            lbl = tk.Label(fsel, text="Sin seleccionar",
                bg=COLOR_BLANCO, fg="#999999", font=("Segoe UI", 9),
                relief="sunken", padx=6, pady=2, cursor="hand2")
            lbl.grid(row=0, column=c0+1, sticky="ew", padx=(0, 4))
            lbl.bind("<Button-1>", lambda e, c=cmd: c())
            setattr(self, lbl_attr, lbl)
            ttk.Button(fsel, text="ð", width=3,
                       command=cmd).grid(row=0, column=c0+2, padx=(0, 4))

        # Selector plantilla
        fsel.columnconfigure(1, weight=1)  # ya estaba, reforzar expansiÃ³n col 1
        tk.Label(fsel, text="ð PLANTILLA DE DEPOSITOS:",
                 bg=COLOR_FONDO, fg=COLOR_TEXTO,
                 font=("Segoe UI", 9, "bold")).grid(
                 row=1, column=0, sticky="w", padx=(0, 4), pady=(6, 0))
        self._dep_lbl_plantilla = tk.Label(fsel, text="Sin seleccionar",
            bg=COLOR_BLANCO, fg="#999999", font=("Segoe UI", 9),
            relief="sunken", padx=6, pady=2, cursor="hand2")
        self._dep_lbl_plantilla.grid(row=1, column=1, columnspan=7,
            sticky="ew", padx=(0, 4), pady=(6, 0))
        self._dep_lbl_plantilla.bind("<Button-1>", lambda e: self._dep_elegir_plantilla())
        ttk.Button(fsel, text="ð", width=3,
                   command=self._dep_elegir_plantilla).grid(
                   row=1, column=8, padx=(0, 4), pady=(6, 0))

        # BotÃ³n generar
        btn_row = tk.Frame(ctrl, bg=COLOR_FONDO)
        btn_row.grid(row=2, column=0, sticky="ew", padx=16, pady=(4, 4))
        btn_row.columnconfigure(0, weight=1)
        self._dep_btn_gen = tk.Button(btn_row,
            text="  â Generar PÃ³liza de DepÃ³sitos",
            bg=self._DEP_ACENTO, fg=COLOR_BLANCO,
            font=("Segoe UI", 10, "bold"), relief="flat",
            padx=16, pady=8, cursor="hand2",
            activebackground="#0D47A1",
            command=self._dep_generar)
        self._dep_btn_gen.grid(row=0, column=0, sticky="ew")

        # Barra de progreso
        self._dep_pb_frame = tk.Frame(ctrl, bg=COLOR_FONDO)
        self._dep_pb_frame.grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 2))
        self._dep_pb_frame.grid_remove()
        self._dep_pb = FunkyProgressBar(self._dep_pb_frame, maximum=100, height=14)
        self._dep_pb.pack(side="left", fill="x", expand=True, padx=(0, 6))
        self._dep_pb_lbl = ttk.Label(self._dep_pb_frame, text="", width=10,
            foreground=self._DEP_ACENTO, font=("Segoe UI", 8, "bold"))
        self._dep_pb_lbl.pack(side="left")

        # Barra de resultado
        tb = tk.Frame(ctrl, bg=COLOR_BLANCO,
                      highlightbackground="#BBDEFB", highlightthickness=1)
        tb.grid(row=4, column=0, sticky="ew", padx=16, pady=(0, 6))
        tk.Label(tb, text="Resultado:", bg=COLOR_BLANCO, fg=COLOR_TEXTO,
                 font=("Segoe UI", 8)).pack(side="left", padx=(10, 2), pady=3)
        self._dep_lbl_arch = tk.Label(tb, text="---", bg=COLOR_BLANCO,
            fg=self._DEP_ACENTO, font=("Segoe UI", 8, "bold"))
        self._dep_lbl_arch.pack(side="left", padx=(0, 10), pady=3)
        self._dep_btn_guardar = ttk.Button(tb, text="ð¾ Guardar como",
            command=self._dep_guardar_como, state="disabled")
        self._dep_btn_guardar.pack(side="right", padx=(6, 2), pady=3)
        self._dep_btn_abrir = ttk.Button(tb, text="ð Abrir en Excel",
            command=self._dep_abrir, state="disabled")
        self._dep_btn_abrir.pack(side="right", padx=6, pady=3)

        # ââ Ãrea de trabajo âââââââââââââââââââââââââââââââââââââââââââââââââââââ
        ttk.Separator(outer, orient="horizontal").grid(row=0, column=0, sticky="sew")

        work = tk.Frame(outer, bg=COLOR_FONDO)
        work.grid(row=1, column=0, sticky="nsew")
        work.columnconfigure(0, weight=1)
        work.rowconfigure(2, weight=1)

        lbl_area = tk.Frame(work, bg=self._DEP_ACENTO)
        lbl_area.grid(row=0, column=0, columnspan=2, sticky="ew")
        tk.Label(lbl_area, text="  Ãrea de Trabajo  â  Vista previa de pÃ³liza",
                 bg=self._DEP_ACENTO, fg=COLOR_BLANCO,
                 font=("Segoe UI", 9, "bold")).pack(side="left", pady=4)
        self._dep_lbl_resumen = tk.Label(lbl_area, text="",
            bg=self._DEP_ACENTO, fg="#90CAF9", font=("Segoe UI", 8))
        self._dep_lbl_resumen.pack(side="right", padx=10)

        # Leyenda de colores por banco
        leyenda_bar = tk.Frame(work, bg=COLOR_FONDO)
        leyenda_bar.grid(row=1, column=0, columnspan=2, sticky="ew", padx=8, pady=(3, 1))
        for txt_l, bg_l, fg_l in [
                ("â BBVA",    "#C5DCF5", "#0D47A1"),
                ("â Inbursa", "#FFE0B5", "#BF360C"),
                ("â Banorte", "#FFCCCC", "#B71C1C")]:
            tk.Label(leyenda_bar, text=f"  {txt_l}  ",
                     bg=bg_l, fg=fg_l, font=("Segoe UI", 8),
                     relief="flat", padx=4, pady=1).pack(side="left", padx=2)

        # Treeview
        cols_tree = ("Fecha", "Banco", "DescripciÃ³n", "Cta. Cargo", "Cta. Abono", "Monto")
        self._dep_tree = ttk.Treeview(work, columns=cols_tree,
                                       show="headings", selectmode="browse")
        col_w   = {"Fecha": 90, "Banco": 80, "DescripciÃ³n": 310,
                   "Cta. Cargo": 140, "Cta. Abono": 180, "Monto": 110}
        col_anc = {"Fecha": "center", "Banco": "center", "DescripciÃ³n": "w",
                   "Cta. Cargo": "center", "Cta. Abono": "w", "Monto": "e"}
        for c in cols_tree:
            self._dep_tree.heading(c, text=c)
            self._dep_tree.column(c, width=col_w.get(c, 100),
                minwidth=60, anchor=col_anc.get(c, "w"), stretch=True)
        self._dep_tree.grid(row=2, column=0, sticky="nsew")

        self._dep_tree.tag_configure("bbva",     foreground="#0D47A1", background="#DDEEFF")
        self._dep_tree.tag_configure("bbva_alt", foreground="#0D47A1", background="#C5DCF5")
        self._dep_tree.tag_configure("inbu",     foreground="#BF360C", background="#FFF0DC")
        self._dep_tree.tag_configure("inbu_alt", foreground="#BF360C", background="#FFE0B5")
        self._dep_tree.tag_configure("bnrt",     foreground="#B71C1C", background="#FFE4E4")
        self._dep_tree.tag_configure("bnrt_alt", foreground="#B71C1C", background="#FFCCCC")

        vsb = ttk.Scrollbar(work, orient="vertical",   command=self._dep_tree.yview)
        hsb = ttk.Scrollbar(work, orient="horizontal", command=self._dep_tree.xview)
        self._dep_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.grid(row=2, column=1, sticky="ns")
        hsb.grid(row=3, column=0, sticky="ew")

    def _dep_elegir_bbva(self):
        self.focus_force(); self.update()
        ruta = filedialog.askopenfilename(parent=self,
            title="Estado de cuenta BBVA (.xlsx)",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if ruta:
            self._dep_ruta_bbva.set(ruta)
            self._dep_lbl_bbva.config(text=os.path.basename(ruta), fg="#0D47A1")

    def _dep_elegir_banorte(self):
        self.focus_force(); self.update()
        ruta = filedialog.askopenfilename(parent=self,
            title="Estado de cuenta Banorte (.xlsx)",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if ruta:
            self._dep_ruta_banorte.set(ruta)
            self._dep_lbl_banorte.config(text=os.path.basename(ruta), fg="#B71C1C")

    def _dep_elegir_inbursa(self):
        self.focus_force(); self.update()
        ruta = filedialog.askopenfilename(parent=self,
            title="Estado de cuenta Inbursa (.xlsx)",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if ruta:
            self._dep_ruta_inbursa.set(ruta)
            self._dep_lbl_inbursa.config(text=os.path.basename(ruta), fg="#BF360C")

    def _dep_elegir_plantilla(self):
        self.focus_force(); self.update()
        ruta = filedialog.askopenfilename(parent=self,
            title="Plantilla de DepÃ³sitos (.xlsx)",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if ruta:
            self._dep_ruta_plantilla.set(ruta)
            self._dep_lbl_plantilla.config(text=os.path.basename(ruta), fg="#1565C0")

    def _dep_generar(self):
        archivos = {}
        if self._dep_ruta_bbva.get():    archivos["BBVA"]    = self._dep_ruta_bbva.get()
        if self._dep_ruta_banorte.get(): archivos["BANORTE"] = self._dep_ruta_banorte.get()
        if self._dep_ruta_inbursa.get(): archivos["INBURSA"] = self._dep_ruta_inbursa.get()
        if not archivos:
            messagebox.showwarning("Sin archivos",
                "Selecciona al menos un estado de cuenta.", parent=self)
            return
        self._dep_btn_gen.config(state="disabled")
        self._dep_btn_abrir.config(state="disabled")
        self._dep_btn_guardar.config(state="disabled")
        self._dep_pb_frame.grid()
        self._dep_pb.reset()
        self._dep_pb_lbl.config(text="0%")
        self._dep_lbl_resumen.config(text="")
        for item in self._dep_tree.get_children():
            self._dep_tree.delete(item)
        import threading
        threading.Thread(target=self._dep_hilo, args=(archivos,), daemon=True).start()

    def _dep_hilo(self, archivos):
        import re as _re, tempfile as _tmp, os as _os
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.after(0, self._log, "Falta openpyxl â pip install openpyxl", True)
            self.after(0, lambda: self._dep_btn_gen.config(state="normal"))
            self.after(0, self._dep_pb_frame.grid_remove)
            return

        def _log(msg, err=False): self.after(0, self._log, msg, err)
        def _pb(v, lbl=""):
            self.after(0, self._dep_pb.set_value, v)
            self.after(0, self._dep_pb_lbl.config, {"text": lbl or f"{v}%"})

        CARGOS = {
            "BANORTE": ("102-01-0001-0001", "Banorte"),
            "BBVA":    ("102-01-0001-0003", "BBVA"),
            "INBURSA": ("102-01-0001-0002", "INBURSA"),
        }
        ABONOS = [
            {"col":12,"cuenta":"106-01-0001-0010","nombre":"DEPOSITO EN TRANSITO T AMERICAN EXPRESS"},
            {"col":13,"cuenta":"106-01-0001-0005","nombre":"DEPOSITO EN TRANSITO  EFECTIVALE"},
            {"col":14,"cuenta":"106-01-0001-0009","nombre":"DEPOSITO EN TRANSITO  TICKET CARD EDENRED"},
            {"col":15,"cuenta":"101-01-0001",      "nombre":"Fondo Fijo de Caja"},
            {"col":16,"cuenta":"106-01-0001-0011", "nombre":"DEPOSITO EN TRANSITO T BANORTE"},
            {"col":17,"cuenta":"106-01-0001-0003", "nombre":"DEPOSITO EN TRANSITO  SMARTBT - SHELL FLEET"},
            {"col":18,"cuenta":"106-01-0001-0013", "nombre":"BBVA"},
            {"col":19,"cuenta":"106-01-0001-0012", "nombre":"DEPOSITO EN TRANSITO T INBURSA"},
        ]
        COL_ABONO_IDX = {a["col"]: a for a in ABONOS}
        CARGO_COL = {"BBVA": 9, "BANORTE": 8, "INBURSA": 10}
        CARGO_CTA = {"BBVA": "102-01-0001-0003", "BANORTE": "102-01-0001-0001", "INBURSA": "102-01-0001-0002"}

        def _clf_bbva(desc, monto):
            d = desc.upper()
            if "AMEX" in d: return 12
            if ("VENTAS PUNTOS TDC" in d or "VENTAS CREDITO" in d
                    or "TERMINALES PUNTO DE VENTA" in d
                    or "VENTAS TDC INTER" in d or "TDC INTER" in d): return 18
            if "DEPOSITO EN EFECTIVO" in d or "DEP.EFECTIVO" in d or "DEP EN EFECTIVO" in d: return 15
            return None
        def _clf_inbursa(desc, monto):
            return 19 if "INBURED" in desc.upper() else None
        def _clf_banorte(desc, monto):
            d = desc.upper()
            if "DEP. CH." in d or "CHEQUE SBC" in d: return None
            if "COMPENSACION DESFASE" in d: return None
            if "SPEI RECIBIDO" in d and "SERVICIOS FELUSA" in d: return None
            if "SHELL" in d or "SMARTBT" in d: return 17
            if "AMERICAN EXPRESS" in d or "BCO:0124" in d or "CITI MEXICO" in d:
                m = _re.search(r"HR LIQ:\s*(\d{2}):", desc)
                return 17 if (int(m.group(1)) if m else 0) >= 12 else 12
            if "EFECTIVALE" in d or "EFE8908015L3" in d or "BCO:0014" in d or (
                    "SANTANDER" in d and "SPEI RECIBIDO" in d): return 13
            if "EDENRED" in d or "HSBCPGMD" in d or "BCO:0021" in d: return 14
            if "DEP.EFECTIVO" in d or "DEPOSITO EN EFECTIVO" in d: return 15
            if "07277262C" in d or "07277262D" in d or (
                    "SERV" in d and _re.search(r'\d{5,}[CD]', d)):
                if "AMERICAN" in d: return 12
                if "EFECTIVALE" in d: return 13
                if "EDENRED" in d or "TICKET" in d: return 14
                if "SHELL" in d or "SMARTBT" in d: return 17
                if "INBURSA" in d: return 19
                return 16
            return None

        CLSF = {"BBVA": _clf_bbva, "BANORTE": _clf_banorte, "INBURSA": _clf_inbursa}

        def leer_banco(ruta, banco):
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            ini = 2
            for i, r in enumerate(rows):
                if r and str(r[0]).strip().lower() == "fecha":
                    ini = i + 1; break
            ok, nc = [], []
            fn = CLSF[banco]
            for r in rows[ini:]:
                if not r or r[0] is None: continue
                fecha = r[0]
                if not hasattr(fecha, "day"): continue
                desc  = str(r[1] or "").strip()
                monto = r[2]
                if not monto or monto <= 0: continue
                col_ab = fn(desc, monto)
                if col_ab is None:
                    nc.append({"fecha":fecha,"banco":banco,"desc":desc[:80],"monto":monto})
                else:
                    ok.append({"fecha":fecha,"banco":banco,"ref":f"DEPOSITOS {banco}",
                               "desc":desc,"monto":monto,
                               "col_cargo":CARGO_COL[banco],"col_abono":col_ab})
            return ok, nc

        try:
            todos_ok, todos_nc = [], []
            n_bancos = len(archivos)
            for i, (banco, ruta) in enumerate(archivos.items()):
                _log(f"Leyendo {banco}: {_os.path.basename(ruta)}")
                _pb(int(i/n_bancos*55), f"Leyendo {banco}...")
                ok, nc = leer_banco(ruta, banco)
                todos_ok.extend(ok); todos_nc.extend(nc)
                total_m = sum(r["monto"] for r in ok)
                _log(f"  {banco}: {len(ok)} clasificados (${total_m:,.2f})")
                if nc: _log(f"  {banco}: {len(nc)} sin clasificar â omitidos")

            if not todos_ok:
                _log("Sin depÃ³sitos clasificables.", True)
                self.after(0, lambda: self._dep_btn_gen.config(state="normal"))
                self.after(0, self._dep_pb_frame.grid_remove)
                return

            _log("Generando Excel..."); _pb(60, "Generando...")

            F_FECHA=PatternFill("solid",fgColor="FFC000"); F_REF=PatternFill("solid",fgColor="7030A0")
            F_CONC=PatternFill("solid",fgColor="00B050");  F_ERROR=PatternFill("solid",fgColor="843C0C")
            F_ADMIN=PatternFill("solid",fgColor="D0D6DC"); F_BANCO=PatternFill("solid",fgColor="BDD7EE")
            F_ABONO=PatternFill("solid",fgColor="FFE699"); F_GRAY2=PatternFill("solid",fgColor="BFBFBF")
            F_TIPO=PatternFill("solid",fgColor="E2EFDA");  F_NONE=PatternFill(fill_type=None)
            FILLS_BANCO={
                "BBVA":   (PatternFill("solid",fgColor="DDEEFF"),PatternFill("solid",fgColor="C5DCF5")),
                "INBURSA":(PatternFill("solid",fgColor="FFF0DC"),PatternFill("solid",fgColor="FFE0B5")),
                "BANORTE":(PatternFill("solid",fgColor="FFE4E4"),PatternFill("solid",fgColor="FFCCCC")),
            }
            FMT_N="#,##0.00"; FMT_D="DD/MM/YYYY"
            AC=Alignment(horizontal="center",vertical="center")
            AL=Alignment(horizontal="left",  vertical="center")
            AR=Alignment(horizontal="right", vertical="center")
            ACW=Alignment(horizontal="center",vertical="center",wrap_text=True)

            def fnt(bold=False,color="000000"):
                return Font(name="Aptos Narrow",size=11,bold=bold,color=color)
            def sc(ws,row,col,value=None,font=None,fill=None,align=None,nf=None):
                c=ws.cell(row=row,column=col,value=value)
                if font: c.font=font
                if fill: c.fill=fill
                if align: c.alignment=align
                if nf: c.number_format=nf
                return c

            wb_o=openpyxl.Workbook(); ws_p=wb_o.active; ws_p.title="POLIZA"
            for idx in range(21): sc(ws_p,1,idx+1,idx,fnt(),F_NONE,AC)
            for pi,banco in [(8,"BANORTE"),(9,"BBVA"),(10,"INBURSA")]:
                sc(ws_p,2,pi+1,CARGOS[banco][0],fnt(),F_GRAY2,AC)
            for a in ABONOS: sc(ws_p,2,a["col"]+1,a["cuenta"],fnt(),F_GRAY2,AC)
            for col,lbl,fill,font in [
                (1,"TIPO DE POLIZA",F_TIPO,fnt()),(2,"Fecha",F_FECHA,fnt(bold=True,color="FFFFFF")),
                (3,"REFERENCIA",F_REF,fnt(bold=True,color="FFFFFF")),(4,"CONCEPTO",F_CONC,fnt(bold=True,color="FFFFFF")),
                (5,"ERROR",F_ERROR,fnt(bold=True,color="FFFFFF")),(6,"UIDD",F_ADMIN,fnt(bold=True,color="FFFFFF")),
                (7,"NUM POLIZA",F_ADMIN,fnt(bold=True,color="FFFFFF")),(8,"PROCESADO",F_ADMIN,fnt(bold=True,color="FFFFFF"))]:
                sc(ws_p,3,col,lbl,font,fill,AC)
            for pi,banco in [(8,"BANORTE"),(9,"BBVA"),(10,"INBURSA")]:
                sc(ws_p,3,pi+1,CARGOS[banco][1],fnt(bold=True,color="FFFFFF"),F_BANCO,AC)
            sc(ws_p,3,12,"TOTAL CARGOS",fnt(color="FF0000"),F_NONE,AC)
            for a in ABONOS: sc(ws_p,3,a["col"]+1,a["nombre"],fnt(bold=True),F_ABONO,ACW)
            sc(ws_p,3,21,"TOTAL ABONOS",fnt(color="FF0000"),F_NONE,AC)
            sc(ws_p,3,22,"DIFERENCIA",  fnt(color="FF0000"),F_NONE,AC)

            orden_b={"BBVA":0,"INBURSA":1,"BANORTE":2}
            reg_s=sorted(todos_ok,key=lambda x:(orden_b[x["banco"]],x["fecha"]))
            for fn_num,r in enumerate(reg_s,start=4):
                f1,f2=FILLS_BANCO[r["banco"]]; fr=f1 if fn_num%2==0 else f2
                fd=fnt(); m=r["monto"]
                def dat(col,val,nfmt=None,al=AC,_row=fn_num,_fr=fr,_fd=fd):
                    c=sc(ws_p,_row,col,val,_fd,_fr,al,nfmt); return c
                dat(1,"I"); dat(2,r["fecha"],FMT_D)
                dat(3,r["ref"],al=AL); dat(4,r["ref"],al=AL)
                for col in [5,6,7,8]: dat(col,None)
                dat(r["col_cargo"]+1,m,FMT_N,AR)
                dat(12,m,FMT_N,AR)
                dat(r["col_abono"]+1,m,FMT_N,AR)
                dat(21,m,FMT_N,AR)
                cd=ws_p.cell(row=fn_num,column=22,
                    value=f"={get_column_letter(12)}{fn_num}-{get_column_letter(21)}{fn_num}")
                cd.font=fd; cd.fill=fr; cd.alignment=AR; cd.number_format=FMT_N

            for cn,w in {1:14.4,2:13.0,3:36.9,4:26.3,5:6.7,6:5.3,7:11.3,8:11.6,
                         9:16.0,10:16.0,11:16.0,12:16.3,13:32.0,14:26.0,15:30.0,
                         16:18.0,17:28.0,18:32.0,19:12.0,20:28.0,21:14.1,22:12.6}.items():
                ws_p.column_dimensions[get_column_letter(cn)].width=w
            ws_p.freeze_panes="B4"

            wc=wb_o.create_sheet("CUENTAS"); fp=fnt()
            sc(wc,1,1,"CARGOS",fp,F_NONE,Alignment(horizontal="center"))
            sc(wc,1,4,"ABONOS",fp,F_NONE,Alignment(horizontal="center"))
            sc(wc,2,1,"NÂ° Cuenta",fp,F_NONE); sc(wc,2,2,"Banco",fp,F_NONE)
            sc(wc,2,4,"NÂ° Cuenta",fp,F_NONE); sc(wc,2,5,"Nombre del Deposito",fp,F_NONE)
            for i,(banco,(cta,nom)) in enumerate(CARGOS.items(),start=3):
                sc(wc,i,1,cta,fp,F_NONE); sc(wc,i,2,nom,fp,F_NONE)
            for i,a in enumerate(ABONOS,start=3):
                sc(wc,i,4,a["cuenta"],fp,F_NONE); sc(wc,i,5,a["nombre"],fp,F_NONE)
            for cn,w in {1:16.3,2:10.0,4:16.3,5:42.9}.items():
                wc.column_dimensions[get_column_letter(cn)].width=w

            _pb(90,"Guardando...")
            tmp=_tmp.NamedTemporaryFile(suffix=".xlsx",delete=False,prefix="DEPOSITOS_BANCARIOS_")
            wb_o.save(tmp.name); tmp.close()
            self._dep_wb=wb_o; self._dep_resultado=tmp.name
            from datetime import datetime as _dt
            mes=_dt.now().strftime("%Y-%m")
            self._dep_wb_name=f"DEPOSITOS BANCARIOS {mes}.xlsx"
            self._dep_wb_dir=_os.path.expanduser("~\\Desktop")
            _pb(100,"Â¡Listo!")

            total_g=sum(r["monto"] for r in todos_ok)
            bbva_n=sum(1 for r in todos_ok if r["banco"]=="BBVA")
            bnrt_n=sum(1 for r in todos_ok if r["banco"]=="BANORTE")
            inbu_n=sum(1 for r in todos_ok if r["banco"]=="INBURSA")
            res=(f"  BBVA:{bbva_n}  Inbursa:{inbu_n}  Banorte:{bnrt_n}  "
                 f"Total:{len(todos_ok)} movs  ${total_g:,.2f}")
            self.after(0,self._dep_lbl_resumen.config,{"text":res})
            self.after(0,self._dep_cargar_visor,reg_s,COL_ABONO_IDX,CARGO_CTA)
            self.after(0,self._dep_lbl_arch.config,{"text":self._dep_wb_name})
            self.after(0,self._dep_btn_abrir.config,{"state":"normal"})
            self.after(0,self._dep_btn_guardar.config,{"state":"normal"})
            _log(f"â Listo â {len(todos_ok)} movimientos | Total: ${total_g:,.2f}")
            if todos_nc: _log(f"â   {len(todos_nc)} movimientos no clasificados (omitidos)")

        except Exception as exc:
            import traceback as _tb
            _log(f"Error: {exc}", True)
            _log(_tb.format_exc(), True)
        finally:
            self.after(0, lambda: self._dep_btn_gen.config(state="normal"))
            self.after(500, self._dep_pb_frame.grid_remove)

    def _dep_cargar_visor(self, registros, col_abono_idx, cargo_cta):
        for item in self._dep_tree.get_children():
            self._dep_tree.delete(item)
        TAG_MAP = {
            "BBVA":    ("bbva",    "bbva_alt"),
            "INBURSA": ("inbu",    "inbu_alt"),
            "BANORTE": ("bnrt",    "bnrt_alt"),
        }
        cnt = {"BBVA": 0, "INBURSA": 0, "BANORTE": 0}
        for r in registros:
            banco = r["banco"]
            cnt[banco] += 1
            t1, t2 = TAG_MAP[banco]
            tag = t1 if cnt[banco] % 2 == 1 else t2
            fecha_s = r["fecha"].strftime("%d/%m/%Y") if hasattr(r["fecha"],"strftime") else str(r["fecha"])
            ab = col_abono_idx.get(r["col_abono"], {})
            ab_nom = ab.get("nombre","")[:30]
            self._dep_tree.insert("","end", values=(
                fecha_s, banco, r["desc"][:55],
                cargo_cta.get(banco,""), ab_nom,
                f"${r['monto']:,.2f}"
            ), tags=(tag,))

    def _dep_abrir(self):
        ruta = getattr(self, "_dep_resultado", None)
        if ruta and os.path.exists(ruta):
            import subprocess
            try: subprocess.Popen(["start","",ruta], shell=True)
            except Exception: pass

    def _dep_guardar_como(self):
        wb = getattr(self, "_dep_wb", None)
        if wb is None:
            messagebox.showwarning("Sin datos","Primero genera la pÃ³liza.",parent=self)
            return
        self.focus_force(); self.update()
        ruta = filedialog.asksaveasfilename(
            parent=self,
            title="Guardar pÃ³liza de depÃ³sitos como...",
            initialdir=getattr(self,"_dep_wb_dir",os.getcwd()),
            initialfile=getattr(self,"_dep_wb_name","DEPOSITOS BANCARIOS.xlsx"),
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx"),("Todos","*.*")],
        )
        if not ruta: return
        try:
            wb.save(ruta)
            self._dep_resultado=ruta
            self._dep_wb_dir=os.path.dirname(ruta)
            self._dep_wb_name=os.path.basename(ruta)
            self._dep_lbl_arch.config(text=os.path.basename(ruta))
            self._dep_btn_guardar.config(state="disabled")
            messagebox.showinfo("Guardado",f"Archivo guardado en:\n{ruta}",parent=self)
        except Exception as e:
            messagebox.showerror("Error al guardar",f"No se pudo guardar:\n{e}",parent=self)


class App(TkinterDnD.Tk if _DND_OK else tk.Tk):
    _CONFIG = os.path.join(os.path.dirname(os.path.abspath(__file__)), "auxiliar_config.json")

    def __init__(self):
        super().__init__()
        # Root siempre oculto â evita que withdraw afecte a los workspaces hijos
        self.withdraw()
        self.title("AUXILIAR DE REGISTROS")
        self.workspaces = {}

        self._aplicar_estilo()

        # Launcher como Toplevel independiente del root
        self._launcher_top = tk.Toplevel(self)
        self._launcher_top.title("AUXILIAR DE REGISTROS")
        self._launcher_top.geometry("500x430")
        self._launcher_top.resizable(False, False)
        self._launcher_top.configure(bg=COLOR_FONDO)
        self._launcher_top.protocol("WM_DELETE_WINDOW", self._on_cerrar_launcher)

        self._construir_launcher()
        self._cargar_empresas()
        self.protocol("WM_DELETE_WINDOW", self._on_cerrar_launcher)

    def _on_cerrar_launcher(self):
        """Cerrar el selector no cierra la app si hay workspaces abiertos."""
        import os as _os
        activos = {k: v for k, v in self.workspaces.items() if v.winfo_exists()}
        # Log de diagnÃ³stico para detectar estado inconsistente
        try:
            _log_path = _os.path.join(_os.path.dirname(__file__), "debug_launcher.txt")
            import time as _t
            with open(_log_path, "a", encoding="utf-8") as _lf:
                _lf.write(
                    f"{_t.strftime('%H:%M:%S')} on_cerrar_launcher: "
                    f"workspaces={list(self.workspaces.keys())} "
                    f"activos={len(activos)}\n"
                )
        except Exception:
            pass
        if activos or self.workspaces:
            # alpha=0: invisible sin withdraw â no afecta otras ventanas en Windows
            self._launcher_top.wm_attributes("-alpha", 0.0)
        else:
            self.destroy()

    def _aplicar_estilo(self):
        s = ttk.Style(self)
        try:
            s.theme_use("clam")
        except tk.TclError:
            pass
        s.configure(".", background=COLOR_FONDO, foreground=COLOR_TEXTO, font=("Segoe UI", 9))
        s.configure("TFrame",     background=COLOR_FONDO)
        s.configure("TLabel",     background=COLOR_FONDO, foreground=COLOR_TEXTO)
        s.configure("TSeparator", background=COLOR_FUCSIA_SUAVE)
        s.configure("TNotebook",  background=COLOR_FONDO, tabmargins=[2, 5, 0, 0])
        s.configure("TNotebook.Tab",
            background="#93C5FD", foreground="#1E3A8A",
            font=("Georgia", 10, "bold"), padding=[14, 8])
        s.map("TNotebook.Tab",
            background=[("selected", "#FFF0F5"), ("active", "#BFDBFE")],
            foreground=[("selected", "#0F172A"), ("active", "#0F172A")])
        s.configure("Tarjeta.TFrame",  background=COLOR_TARJETA, relief="flat")
        s.configure("Tarjeta.TLabel",  background=COLOR_TARJETA, foreground=COLOR_FUCSIA_OSCURO)
        s.configure("Titulo.TLabel",
            background="#1E3A8A", foreground="#FBCFE8",
            font=("Georgia", 14, "bold"), padding=(14, 12))
        s.configure("Seccion.TLabel",
            background=COLOR_FONDO, foreground=COLOR_FUCSIA_OSCURO,
            font=("Segoe UI", 10, "bold"))
        s.configure("Grande.TButton",
            background="#F43F8A", foreground="white",
            font=("Segoe UI", 11, "bold"), padding=(12, 8))
        s.configure("Encabezado.TLabel",
            background=COLOR_TARJETA, foreground=COLOR_FUCSIA_OSCURO,
            font=("Segoe UI", 10, "bold"))
        s.configure("NoTabs.TNotebook", tabmargins=[0, 0, 0, 0])
        s.layout("NoTabs.TNotebook", [("Notebook.client", {"sticky": "nswe"})])
        s.configure("NoTabs.TNotebook.Tab",
                    font=("Segoe UI", 1), padding=[0, 0], focuscolor="")
        s.map("NoTabs.TNotebook.Tab",
              foreground=[("selected", COLOR_FONDO), ("active", COLOR_FONDO), ("!selected", COLOR_FONDO)],
              background=[("selected", COLOR_FONDO), ("active", COLOR_FONDO), ("!selected", COLOR_FONDO)])


    def _construir_launcher(self):
        dlg = self._launcher_top   # construir sobre el Toplevel, no el root oculto
        hdr = tk.Frame(dlg, bg="#1E3A8A", height=60)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        _hdr_bg = hdr.cget("background")
        try:
            from PIL import Image, ImageTk
            import pathlib as _pl
            _img_path = _pl.Path(__file__).parent / "logo_ready.png"
            _img = Image.open(_img_path).resize((40, 40), Image.LANCZOS)
            self._logo_img = ImageTk.PhotoImage(_img)
            tk.Label(hdr, image=self._logo_img, bg=_hdr_bg,
                     bd=0).pack(side="left", padx=(12, 0))
        except Exception:
            pass
        tk.Label(hdr, text="AUXILIAR DE REGISTROS", bg="#1E3A8A", fg="white",
                 font=("Georgia", 14, "bold")).pack(expand=True)

        fr = ttk.Frame(dlg, padding=16)
        fr.pack(fill="both", expand=True)

        ttk.Label(fr, text="Espacios de trabajo:", style="Seccion.TLabel").pack(anchor="w", pady=(0, 4))

        lb_fr = tk.Frame(fr, bg=COLOR_FONDO)
        lb_fr.pack(fill="both", expand=True)

        sb = ttk.Scrollbar(lb_fr)
        sb.pack(side="right", fill="y")

        self.lb_ws = tk.Listbox(lb_fr, yscrollcommand=sb.set,
                                 font=("Segoe UI", 11), selectmode="single",
                                 bg=COLOR_TARJETA, fg=COLOR_TEXTO,
                                 selectbackground=COLOR_FUCSIA_OSCURO,
                                 selectforeground="white",
                                 relief="flat", bd=0, highlightthickness=1,
                                 highlightcolor=COLOR_FUCSIA_SUAVE,
                                 activestyle="none")
        self.lb_ws.pack(side="left", fill="both", expand=True)
        sb.config(command=self.lb_ws.yview)
        self.lb_ws.bind("<Double-1>", lambda e: self._abrir_seleccionada())

        btn_fr = ttk.Frame(fr)
        btn_fr.pack(fill="x", pady=(10, 0))

        ttk.Button(btn_fr, text=" Abrir", style="Grande.TButton",
                   command=self._abrir_seleccionada).pack(side="left", padx=(0, 6))
        ttk.Button(btn_fr, text=" Nueva empresa",
                   command=self._nueva_empresa).pack(side="left", padx=(0, 6))
        ttk.Button(btn_fr, text=" Eliminar",
                   command=self._eliminar_empresa).pack(side="right")

        self._panel_log()

    def _cargar_empresas(self):
        import json as _j
        self._cfg = {}
        if os.path.isfile(self._CONFIG):
            try:
                data = _j.loads(open(self._CONFIG, "rb").read().decode("utf-8"))
                if isinstance(data, dict):
                    self._cfg = data
                elif isinstance(data, list):
                    for nombre in data:
                        if isinstance(nombre, str):
                            self._cfg[nombre] = {"modulos": []}
            except Exception:
                self._cfg = {}
        orden = self._cfg.get("_orden", sorted(k for k in self._cfg if not k.startswith("_")))
        self.lb_ws.delete(0, "end")
        for n in orden:
            if not n.startswith("_"):
                self.lb_ws.insert("end", n)

    def _guardar_modulos(self, nombre, mods):
        import json as _j
        self._cfg.setdefault(nombre, {})["modulos"] = mods
        if "_orden" not in self._cfg:
            self._cfg["_orden"] = [k for k in self._cfg if not k.startswith("_")]
        try:
            with open(self._CONFIG, "w", encoding="utf-8") as f:
                _j.dump(self._cfg, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _abrir_seleccionada(self):
        sel = self.lb_ws.curselection()
        if not sel:
            return
        nombre = self.lb_ws.get(sel[0])
        if nombre in self.workspaces and self.workspaces[nombre].winfo_exists():
            self.workspaces[nombre].lift()
            return
        cfg_e = self._cfg.get(nombre, {})
        ws = None
        try:
            ws = WorkspaceWindow(self, nombre=nombre,
                                 modulos=cfg_e.get("modulos", []),
                                 modulos_disponibles=cfg_e.get("modulos_disponibles", None),
                                 on_modulos_change=lambda _ws_nombre, mods: self._guardar_modulos(nombre, mods))
        except Exception as _e:
            import traceback as _tb
            import os as _os
            _log_path = _os.path.join(_os.path.dirname(__file__), "debug_launcher.txt")
            try:
                with open(_log_path, "a", encoding="utf-8") as _lf:
                    _lf.write(f"ERROR abriendo workspace '{nombre}':\n{_tb.format_exc()}\n")
            except Exception:
                pass
            # Destruir ventana parcial si quedÃ³ creada
            if ws is not None:
                try:
                    ws.destroy()
                except Exception:
                    pass
            from tkinter import messagebox as _mb
            _mb.showerror("Error", f"No se pudo abrir '{nombre}':\n{_e}", parent=self._launcher_top)
            return
        self.workspaces[nombre] = ws
        ws.protocol("WM_DELETE_WINDOW", lambda n=nombre: self._cerrar_workspace(n))
        # Log de apertura exitosa
        import os as _os, time as _t
        try:
            _log_path = _os.path.join(_os.path.dirname(__file__), "debug_launcher.txt")
            with open(_log_path, "a", encoding="utf-8") as _lf:
                _lf.write(f"{_t.strftime('%H:%M:%S')} workspace abierto OK: '{nombre}'\n")
        except Exception:
            pass
        self._launcher_top.wm_attributes("-alpha", 0.0)   # ocultar launcher

    def _cerrar_workspace(self, nombre):
        if nombre in self.workspaces:
            try:
                self.workspaces[nombre].destroy()
            except Exception:
                pass
            del self.workspaces[nombre]
        abiertos = [w for w in self.workspaces.values() if w.winfo_exists()]
        if not abiertos:
            self._launcher_top.wm_attributes("-alpha", 1.0)
            self._launcher_top.lift()
            self._launcher_top.focus_force()

    def _nueva_empresa(self):
        dlg = tk.Toplevel(self._launcher_top)
        dlg.title("Nueva empresa")
        dlg.geometry("340x130")
        dlg.resizable(False, False)
        dlg.configure(bg=COLOR_FONDO)
        dlg.transient(self._launcher_top)
        dlg.grab_set()
        ttk.Label(dlg, text="Nombre de la empresa:").pack(pady=(18, 4))
        var = tk.StringVar()
        ent = ttk.Entry(dlg, textvariable=var, width=32)
        ent.pack(pady=4)
        ent.focus_set()

        def _ok():
            n = var.get().strip()
            if not n:
                return
            if n not in self._cfg:
                self._cfg[n] = {"modulos": []}
                if "_orden" not in self._cfg:
                    self._cfg["_orden"] = []
                if n not in self._cfg["_orden"]:
                    self._cfg["_orden"].append(n)
                self._guardar_modulos(n, [])
                self.lb_ws.insert("end", n)
            dlg.destroy()

        ttk.Button(dlg, text="Crear", command=_ok).pack(pady=8)
        dlg.bind("<Return>", lambda e: _ok())

    def _eliminar_empresa(self):
        sel = self.lb_ws.curselection()
        if not sel:
            return
        nombre = self.lb_ws.get(sel[0])
        if not messagebox.askyesno("Eliminar", f"Eliminar '{nombre}'", parent=self):
            return
        if nombre in self._cfg:
            del self._cfg[nombre]
        if "_orden" in self._cfg and nombre in self._cfg["_orden"]:
            self._cfg["_orden"].remove(nombre)
        self.lb_ws.delete(sel[0])
        import json as _j
        try:
            with open(self._CONFIG, "w", encoding="utf-8") as f:
                _j.dump(self._cfg, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _panel_log(self):
        dlg = self._launcher_top
        log_frame = tk.Frame(dlg, bg=COLOR_FUCSIA_OSCURO, height=28)
        log_frame.pack(side="bottom", fill="x")
        log_frame.pack_propagate(False)
        tk.Label(log_frame, text=" Registro de actividad:", bg=COLOR_FUCSIA_OSCURO,
                 fg="#A8CDE8", font=("Segoe UI", 8, "bold")).pack(
                 side="left", padx=(10, 4), pady=4)
        self._lbl_log = tk.Label(log_frame, text="", bg=COLOR_FUCSIA_OSCURO,
                                  fg=COLOR_FUCSIA_SUAVE, font=("Segoe UI", 8), anchor="w")
        self._lbl_log.pack(side="left", fill="x", expand=True, padx=4)

    def _log(self, msg, ok=False, error=False):
        color = "#7FFFD4" if ok else ("#FF6B6B" if error else COLOR_FUCSIA_SUAVE)
        try:
            self._lbl_log.config(text=msg, fg=color)
        except Exception:
            pass

    def _set_progreso(self, pct, texto=""):
        pass


if __name__ == "__main__":
    import traceback as _tb
    _err_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "error_log.txt")
    try:
        app = App()
        app.mainloop()
    except Exception as _e:
        with open(_err_path, "w", encoding="utf-8") as _f:
            _f.write(_tb.format_exc())
        raise
