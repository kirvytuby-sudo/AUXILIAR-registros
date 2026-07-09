"""AUXILIAR DE REGISTROS — Control de Despacho vs Ventas del Día"""
import io
import streamlit as st
from datetime import datetime
from collections import defaultdict

st.set_page_config(
    page_title="Control Despacho vs Ventas · Auxiliar",
    page_icon="📊",
    layout="wide",
)

st.markdown("""<style>
[data-testid="stAppViewContainer"] { background: #dbeafe; }
.header-bar{background:#B45309;padding:18px 28px;border-radius:10px;margin-bottom:20px;}
.header-bar h1{color:#FFFFFF;font-size:1.5rem;margin:0;}
.header-bar p{color:#FEF3C7;margin:4px 0 0;font-size:.9rem;}
#MainMenu{visibility:hidden;}footer{visibility:hidden;}
</style>
<div class="header-bar">
  <h1>📊 Control de Despacho vs Ventas del Día</h1>
  <p>Concilia el control de despachos contra la póliza de ventas del día · 3 hojas: Conciliación, Detalle UUID, Resumen</p>
</div>
""", unsafe_allow_html=True)

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as _e:
    st.error(f"❌ Librería faltante: {_e}. Verifica requirements.txt.")
    st.stop()


# ══════════════════════════════════════════════════════════════════════════════
# LÓGICA DE PROCESAMIENTO
# ══════════════════════════════════════════════════════════════════════════════

def procesar(desp_bytes: bytes, vd_bytes: bytes):
    """
    Procesa despachos + ventas del día y devuelve (excel_bytes, logs, resumen_dias).
    """
    logs = []

    # ── Leer archivos ─────────────────────────────────────────────────────────
    logs.append("📂 Leyendo Control de Despachos...")
    wb1 = openpyxl.load_workbook(io.BytesIO(desp_bytes), read_only=True, data_only=True)
    rows1 = list(wb1[wb1.sheetnames[0]].iter_rows(values_only=True))
    wb1.close()

    logs.append("📂 Leyendo Ventas del Día...")
    wb2 = openpyxl.load_workbook(io.BytesIO(vd_bytes), read_only=True, data_only=True)
    rows2 = list(wb2[wb2.sheetnames[0]].iter_rows(values_only=True))
    wb2.close()

    # ── Procesar Despachos ────────────────────────────────────────────────────
    dia = defaultdict(lambda: {"GS_sub": 0.0, "GP_sub": 0.0, "GD_sub": 0.0,
                               "iva": 0.0, "ieps": 0.0, "importe": 0.0, "uuids": set()})
    data = defaultdict(lambda: {
        "nombre": "", "fechas": set(),
        "GS": {"sub": 0.0, "iva": 0.0, "ieps": 0.0, "tot": 0.0},
        "GP": {"sub": 0.0, "iva": 0.0, "ieps": 0.0, "tot": 0.0},
        "GD": {"sub": 0.0, "iva": 0.0, "ieps": 0.0, "tot": 0.0},
    })
    sin_uuid = defaultdict(lambda: {
        "GS": {"sub": 0.0, "iva": 0.0, "ieps": 0.0, "tot": 0.0, "n": 0},
        "GP": {"sub": 0.0, "iva": 0.0, "ieps": 0.0, "tot": 0.0, "n": 0},
        "GD": {"sub": 0.0, "iva": 0.0, "ieps": 0.0, "tot": 0.0, "n": 0},
    })

    for r in rows1[1:]:
        if not r or r[0] is None:
            continue
        try:
            dt = datetime.strptime(str(r[0])[:10], "%Y-%m-%d").date()
        except Exception:
            continue
        prod = str(r[3]).strip().upper() if r[3] else ""
        if prod not in ("GS", "GP", "GD"):
            continue
        sub = float(r[6] or 0); iva = float(r[7] or 0)
        ieps = float(r[8] or 0); imp = float(r[9] or 0)
        cliente = str(r[16]).strip() if r[16] else "SIN CLIENTE"
        uuid = str(r[22]).strip() if r[22] else "-----"
        rfc = str(r[23]).strip() if r[23] else "SIN-RFC"
        nombre = str(r[24]).strip() if r[24] else ""
        d = dia[dt]
        d[f"{prod}_sub"] += sub; d["iva"] += iva
        d["ieps"] += ieps; d["importe"] += imp; d["uuids"].add(uuid)
        if uuid == "-----":
            s = sin_uuid[(cliente, dt)][prod]
            s["sub"] += sub; s["iva"] += iva; s["ieps"] += ieps; s["tot"] += imp; s["n"] += 1
        else:
            k = (rfc, uuid); e = data[k]
            e["nombre"] = nombre; e["fechas"].add(dt)
            e[prod]["sub"] += sub; e[prod]["iva"] += iva
            e[prod]["ieps"] += ieps; e[prod]["tot"] += imp

    # ── Procesar Ventas del Día ───────────────────────────────────────────────
    venta = {}
    for r in rows2[3:]:
        if not r or r[1] is None:
            continue
        fecha = r[1]
        if isinstance(fecha, datetime):
            fecha = fecha.date()
        else:
            continue
        venta[fecha] = {
            "GS": float(r[29] or 0), "GP": float(r[30] or 0), "GD": float(r[31] or 0),
            "IVA": float(r[32] or 0), "IEPS": float(r[33] or 0), "TOTAL": float(r[35] or 0),
        }

    fechas = sorted(dia.keys())
    logs.append(f"✅ {len(fechas)} días encontrados en Despachos.")

    # ── Estilos ───────────────────────────────────────────────────────────────
    MF = '#,##0.00'
    F_WHITE = PatternFill("solid", fgColor="FFFFFF")
    def fill(c): return PatternFill("solid", fgColor=c)
    thin = Side(style="thin", color="CCCCCC")
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    al_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
    al_r = Alignment(horizontal="right",  vertical="center")
    al_l = Alignment(horizontal="left",   vertical="center")

    F_HDR = fill("1E3A8A"); F_GS = fill("DBEAFE"); F_GP = fill("D1FAE5"); F_GD = fill("FEF9C3")
    F_TOT = fill("FFC000"); F_OK = fill("C6EFCE"); F_WRN = fill("FFEB9C"); F_ERR = fill("FFC7CE")
    F_IVA = fill("FCE7F3"); F_UUID = fill("F1F5F9"); F_RFC = fill("E0F2FE")
    F_DARK = fill("111827"); F_GRAY = fill("374151")
    F_SIN = fill("FFF7ED"); F_SIN_H = fill("EA580C")
    F_SIN_TOT = fill("F97316"); F_SIN_ROW = fill("FFEDD5")
    F_TITULO = fill("0F172A")

    def cw(ws, row, col, val=None, fi=None, al=None, fmt=None, bold=False, sz=9, fc="000000"):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fi if fi is not None else F_WHITE
        c.font = Font(color=fc, bold=bold, size=sz)
        if al:  c.alignment = al
        if fmt: c.number_format = fmt
        c.border = bord
        return c

    def mw(ws, r, c1, c2, val, fi, bold=False, sz=9, fc="000000", al=None):
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(r, c1, val)
        cell.fill = fi; cell.font = Font(color=fc, bold=bold, size=sz)
        cell.border = bord; cell.alignment = al or al_l
        for ci in range(c1 + 1, c2 + 1):
            ws.cell(r, ci).fill = fi; ws.cell(r, ci).border = bord

    wb = openpyxl.Workbook()

    # ══ HOJA 1 — CONCILIACIÓN DIARIA ══════════════════════════════════════════
    ws = wb.active; ws.title = "Conciliación"
    mw(ws, 1, 1, 18, "CONTROL DE DESPACHO vs VENTAS DEL DÍA", F_HDR, bold=True, sz=12, fc="FFFFFF", al=al_c)
    ws.row_dimensions[1].height = 26
    for s, e, txt, fi in [(1, 1, "FECHA", F_HDR), (2, 8, "CONTROL DE DESPACHOS", F_DARK),
                          (9, 13, "VENTAS DEL DÍA", F_GRAY), (14, 18, "DIFERENCIAS", F_HDR)]:
        if s < e: mw(ws, 2, s, e, txt, fi, bold=True, sz=9, fc="FFFFFF", al=al_c)
        else:     cw(ws, 2, s, val=txt, fi=fi, al=al_c, bold=True, sz=9, fc="FFFFFF")
    ws.row_dimensions[2].height = 18
    cols3 = [
        ("Fecha", 12, F_HDR, "FFFFFF"),
        ("G-Super\nSubtotal", 14, F_GS, "1E3A8A"), ("G-Premium\nSubtotal", 14, F_GP, "166534"),
        ("G-Diesel\nSubtotal", 14, F_GD, "92400E"),
        ("IVA\nUnificado", 14, F_IVA, "9D174D"), ("IEPS\nUnificado", 14, F_IVA, "9D174D"),
        ("TOTAL\nImporte", 15, F_DARK, "FFFFFF"), ("UUIDs\n#", 9, F_UUID, "374151"),
        ("GS\n(VtaDía)", 13, F_GS, "1E3A8A"), ("GP\n(VtaDía)", 13, F_GP, "166534"),
        ("GD\n(VtaDía)", 13, F_GD, "92400E"),
        ("IVA\n(VtaDía)", 13, F_IVA, "9D174D"), ("IEPS\n(VtaDía)", 13, F_IVA, "9D174D"),
        ("Dif\nGS", 12, F_HDR, "FFFFFF"), ("Dif\nGP", 12, F_HDR, "FFFFFF"),
        ("Dif\nGD", 12, F_HDR, "FFFFFF"), ("Dif\nIVA", 12, F_HDR, "FFFFFF"),
        ("DIF\nTOTAL", 13, F_HDR, "FFFFFF"),
    ]
    for i, (h, w, fi, fc) in enumerate(cols3, 1):
        c = ws.cell(3, i, h); c.fill = fi; c.font = Font(color=fc, bold=True, size=8)
        c.alignment = al_c; c.border = bord
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 34

    ROW = 4; tots = defaultdict(float)
    resumen_dias = []
    for idx, fecha in enumerate(fechas):
        d = dia[fecha]; v = venta.get(fecha, {})
        gs_s = d["GS_sub"]; gp_s = d["GP_sub"]; gd_s = d["GD_sub"]
        iva_u = d["iva"]; ieps_u = d["ieps"]; tot = d["importe"]
        v_gs = v.get("GS", 0); v_gp = v.get("GP", 0); v_gd = v.get("GD", 0)
        v_iva = v.get("IVA", 0); v_ieps = v.get("IEPS", 0); v_tot = v.get("TOTAL", 0)
        dif_gs = round(gs_s - v_gs, 2); dif_gp = round(gp_s - v_gp, 2)
        dif_gd = round(gd_s - v_gd, 2); dif_iva = round(iva_u - v_iva, 2)
        dif_tot = round(tot - v_tot, 2)
        fi_d = F_OK if abs(dif_tot) < 1 else F_WRN if abs(dif_tot) < 1000 else F_ERR
        vals_row = [
            (fecha.strftime("%d/%m/%Y"), None), (gs_s, F_GS), (gp_s, F_GP), (gd_s, F_GD),
            (iva_u, F_IVA), (ieps_u, F_IVA), (tot, F_DARK), (len(d["uuids"]), F_UUID),
            (v_gs, F_GS), (v_gp, F_GP), (v_gd, F_GD), (v_iva, F_IVA), (v_ieps, F_IVA),
            (dif_gs, fi_d), (dif_gp, fi_d), (dif_gd, fi_d), (dif_iva, fi_d), (dif_tot, fi_d),
        ]
        fcs_r = ["000000", "1E3A8A", "166534", "92400E", "9D174D", "9D174D", "FFFFFF", "374151",
                 "1E3A8A", "166534", "92400E", "9D174D", "9D174D",
                 "000000", "000000", "000000", "000000", "000000"]
        fmts_r = [None, MF, MF, MF, MF, MF, MF, "#,##0", MF, MF, MF, MF, MF, MF, MF, MF, MF, MF]
        for i, ((val, fi), fc, fmt) in enumerate(zip(vals_row, fcs_r, fmts_r), 1):
            cw(ws, ROW, i, val=val, fi=fi, al=al_c if i in (1, 8) else al_r, fmt=fmt, sz=9, fc=fc)
        for k, kv in [("gs", gs_s), ("gp", gp_s), ("gd", gd_s), ("iva", iva_u),
                      ("ieps", ieps_u), ("tot", tot), ("vgs", v_gs), ("vgp", v_gp),
                      ("vgd", v_gd), ("viva", v_iva), ("vieps", v_ieps), ("vtot", v_tot)]:
            tots[k] += kv
        resumen_dias.append({
            "Fecha": fecha.strftime("%d/%m/%Y"),
            "Tot Despachos": round(tot, 2),
            "Tot VtaDía": round(v_tot, 2),
            "Diferencia": round(dif_tot, 2),
        })
        ws.row_dimensions[ROW].height = 15; ROW += 1

    tv = ["TOTAL", tots["gs"], tots["gp"], tots["gd"], tots["iva"], tots["ieps"], tots["tot"],
          sum(len(dia[f]["uuids"]) for f in fechas),
          tots["vgs"], tots["vgp"], tots["vgd"], tots["viva"], tots["vieps"],
          round(tots["gs"] - tots["vgs"], 2), round(tots["gp"] - tots["vgp"], 2),
          round(tots["gd"] - tots["vgd"], 2), round(tots["iva"] - tots["viva"], 2),
          round(tots["tot"] - tots["vtot"], 2)]
    for i, v in enumerate(tv, 1):
        c = ws.cell(ROW, i, v); c.fill = F_TOT; c.font = Font(bold=True, size=9)
        c.border = bord; c.alignment = al_l if i == 1 else al_r
        if 2 <= i <= 18: c.number_format = MF if i != 8 else "#,##0"
    ws.row_dimensions[ROW].height = 22; ws.freeze_panes = "B4"

    # ══ HOJA 2 — DETALLE UUID ═════════════════════════════════════════════════
    ws2 = wb.create_sheet("Detalle UUID")
    for i, w in enumerate([20, 30, 42, 14, 15, 13, 13, 15], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    colores_alt = [fill("EFF6FF"), fill("F0FFFE"), fill("FFFBEB"), fill("FDF4FF"), fill("FFF1F2")]
    PRODS = [
        ("GS", "G-SUPER",   fill("1D4ED8"), F_GS, "1D4ED8"),
        ("GP", "G-PREMIUM", fill("15803D"), F_GP, "14532D"),
        ("GD", "G-DIESEL",  fill("92400E"), F_GD, "451A03"),
    ]
    CUR = 1
    grand_total = {"GS": defaultdict(float), "GP": defaultdict(float), "GD": defaultdict(float)}

    mw(ws2, CUR, 1, 8, "TABLA DE UUIDs", F_TITULO, bold=True, sz=14, fc="FFFFFF", al=al_c)
    ws2.row_dimensions[CUR].height = 34; CUR += 1

    for prod, label, fi_hdr, fi_prod, fc_txt in PRODS:
        mw(ws2, CUR, 1, 8, f"  ▌ TABLA {label}", fi_hdr, bold=True, sz=12, fc="FFFFFF", al=al_l)
        ws2.row_dimensions[CUR].height = 28; CUR += 1

        claves_sin = sorted(
            [(cli, dt) for (cli, dt) in sin_uuid if sin_uuid[(cli, dt)][prod]["n"] > 0],
            key=lambda x: (x[0], x[1])
        )
        if claves_sin:
            mw(ws2, CUR, 1, 8, "  ⚠  SIN COMPROBANTE FISCAL (UUID = -----)",
               F_SIN_H, bold=True, sz=9, fc="FFFFFF", al=al_l)
            ws2.row_dimensions[CUR].height = 20; CUR += 1
            for i, h in enumerate(["Cliente (col Q)", "—", "— SIN UUID —",
                                   "Fecha", "Subtotal", "IVA", "IEPS", "TOTAL"], 1):
                c = ws2.cell(CUR, i, h); c.fill = F_SIN
                c.font = Font(color="92400E", bold=True, size=8)
                c.alignment = al_c; c.border = bord
            ws2.row_dimensions[CUR].height = 16; CUR += 1
            prev_cli = None; cli_tots = defaultdict(float); sin_grand = defaultdict(float)
            for (cli, dt) in claves_sin:
                if prev_cli is not None and cli != prev_cli:
                    mw(ws2, CUR, 1, 3, f"  Subtotal  {prev_cli}", F_RFC, bold=True, sz=8, fc="7C2D12", al=al_l)
                    cw(ws2, CUR, 4, fi=F_RFC)
                    for ci, kk in enumerate(["sub", "iva", "ieps", "tot"], 5):
                        cw(ws2, CUR, ci, val=cli_tots[kk], fi=F_RFC, al=al_r, fmt=MF, bold=True, sz=8, fc="7C2D12")
                    for k in ["sub", "iva", "ieps", "tot"]: cli_tots[k] = 0.0
                    ws2.row_dimensions[CUR].height = 14; CUR += 1
                prev_cli = cli; s = sin_uuid[(cli, dt)][prod]
                cw(ws2, CUR, 1, val=cli,                      fi=F_SIN_ROW, al=al_l, sz=8, fc="7C2D12")
                cw(ws2, CUR, 2, val="— SIN NOMBRE —",         fi=F_SIN_ROW, al=al_l, sz=8, fc="9A3412")
                cw(ws2, CUR, 3, val="— SIN COMPROBANTE FISCAL —", fi=F_SIN_ROW, al=al_c, sz=8, fc="9A3412")
                cw(ws2, CUR, 4, val=dt.strftime("%d/%m/%Y"),  fi=F_SIN_ROW, al=al_c, sz=8)
                cw(ws2, CUR, 5, val=s["sub"],  fi=F_SIN_ROW, al=al_r, fmt=MF, sz=8)
                cw(ws2, CUR, 6, val=s["iva"],  fi=F_SIN_ROW, al=al_r, fmt=MF, sz=8)
                cw(ws2, CUR, 7, val=s["ieps"], fi=F_SIN_ROW, al=al_r, fmt=MF, sz=8)
                cw(ws2, CUR, 8, val=s["tot"],  fi=F_SIN_ROW, al=al_r, fmt=MF, sz=8)
                for k, kv in [("sub", s["sub"]), ("iva", s["iva"]), ("ieps", s["ieps"]), ("tot", s["tot"])]:
                    cli_tots[k] += kv; sin_grand[k] += kv
                ws2.row_dimensions[CUR].height = 13; CUR += 1
            if prev_cli:
                mw(ws2, CUR, 1, 3, f"  Subtotal  {prev_cli}", F_RFC, bold=True, sz=8, fc="7C2D12", al=al_l)
                cw(ws2, CUR, 4, fi=F_RFC)
                for ci, kk in enumerate(["sub", "iva", "ieps", "tot"], 5):
                    cw(ws2, CUR, ci, val=cli_tots[kk], fi=F_RFC, al=al_r, fmt=MF, bold=True, sz=8, fc="7C2D12")
                ws2.row_dimensions[CUR].height = 14; CUR += 1
            mw(ws2, CUR, 1, 4, f"  TOTAL SIN COMPROBANTE — {label}", F_SIN_TOT, bold=True, sz=9, fc="FFFFFF", al=al_l)
            for ci, kk in enumerate(["sub", "iva", "ieps", "tot"], 5):
                cw(ws2, CUR, ci, val=sin_grand[kk], fi=F_SIN_TOT, al=al_r, fmt=MF, bold=True, sz=9, fc="FFFFFF")
            ws2.row_dimensions[CUR].height = 18; CUR += 1
            gt = grand_total[prod]
            for k in ["sub", "iva", "ieps", "tot"]: gt[f"sin_{k}"] += sin_grand[k]

        mw(ws2, CUR, 1, 8, "  CON COMPROBANTE FISCAL (CFDI / UUID)", F_HDR, bold=True, sz=9, fc="FFFFFF", al=al_l)
        ws2.row_dimensions[CUR].height = 16; CUR += 1
        for i, h in enumerate(["RFC", "Nombre / Razón Social", "UUID / Folio Fiscal",
                               "Fechas", "Subtotal", "IVA", "IEPS", "TOTAL"], 1):
            c = ws2.cell(CUR, i, h); c.fill = fi_prod
            c.font = Font(color=fc_txt, bold=True, size=9); c.alignment = al_c; c.border = bord
        ws2.row_dimensions[CUR].height = 16; CUR += 1

        claves = sorted([(rfc, uuid) for (rfc, uuid), e in data.items() if e[prod]["tot"] > 0],
                        key=lambda x: (x[0], x[1]))
        rfc_col = {}; cidx = 0; prev_rfc = None; rfc_tots = defaultdict(float); gt = grand_total[prod]
        for (rfc, uuid) in claves:
            e = data[(rfc, uuid)]
            if rfc not in rfc_col: rfc_col[rfc] = colores_alt[cidx % len(colores_alt)]; cidx += 1
            if prev_rfc is not None and rfc != prev_rfc:
                mw(ws2, CUR, 1, 4, f"  Subtotal  {prev_rfc}", F_RFC, bold=True, sz=8, fc="1E40AF", al=al_l)
                for ci, kk in enumerate(["sub", "iva", "ieps", "tot"], 5):
                    cw(ws2, CUR, ci, val=rfc_tots[kk], fi=F_RFC, al=al_r, fmt=MF, bold=True, sz=8, fc="1E40AF")
                for k in ["sub", "iva", "ieps", "tot"]: rfc_tots[k] = 0.0
                ws2.row_dimensions[CUR].height = 14; CUR += 1
            prev_rfc = rfc; fi_r = rfc_col[rfc]
            sub = e[prod]["sub"]; iva = e[prod]["iva"]; ieps = e[prod]["ieps"]; tot_v = e[prod]["tot"]
            fechas_str = ", ".join(sorted(f.strftime("%d/%m") for f in e["fechas"]))
            for i, val in enumerate([rfc, e["nombre"], uuid, fechas_str, sub, iva, ieps, tot_v], 1):
                c = ws2.cell(CUR, i, val); c.fill = fi_r; c.font = Font(size=8)
                c.border = bord; c.alignment = al_l if i <= 4 else al_r
                if i > 4: c.number_format = MF
            for k, kv in [("sub", sub), ("iva", iva), ("ieps", ieps), ("tot", tot_v)]:
                rfc_tots[k] += kv; gt[k] += kv
            ws2.row_dimensions[CUR].height = 13; CUR += 1
        if prev_rfc:
            mw(ws2, CUR, 1, 4, f"  Subtotal  {prev_rfc}", F_RFC, bold=True, sz=8, fc="1E40AF", al=al_l)
            for ci, kk in enumerate(["sub", "iva", "ieps", "tot"], 5):
                cw(ws2, CUR, ci, val=rfc_tots[kk], fi=F_RFC, al=al_r, fmt=MF, bold=True, sz=8, fc="1E40AF")
            ws2.row_dimensions[CUR].height = 14; CUR += 1
        mw(ws2, CUR, 1, 4, f"  TOTAL CON COMPROBANTE — {label}", fill("166534"), bold=True, sz=9, fc="FFFFFF", al=al_l)
        for ci, kk in enumerate(["sub", "iva", "ieps", "tot"], 5):
            cw(ws2, CUR, ci, val=gt.get(kk, 0), fi=fill("166534"), al=al_r, fmt=MF, bold=True, sz=9, fc="FFFFFF")
        ws2.row_dimensions[CUR].height = 18; CUR += 1
        gran_sub  = gt.get("sub", 0) + gt.get("sin_sub", 0)
        gran_iva  = gt.get("iva", 0) + gt.get("sin_iva", 0)
        gran_ieps = gt.get("ieps", 0) + gt.get("sin_ieps", 0)
        gran_tot  = gt.get("tot", 0) + gt.get("sin_tot", 0)
        mw(ws2, CUR, 1, 4, f"  ★ GRAN TOTAL {label}", F_TOT, bold=True, sz=11, al=al_l)
        for ci, v in enumerate([gran_sub, gran_iva, gran_ieps, gran_tot], 5):
            cw(ws2, CUR, ci, val=v, fi=F_TOT, al=al_r, fmt=MF, bold=True, sz=11)
        ws2.row_dimensions[CUR].height = 26; CUR += 2

    ws2.freeze_panes = "A2"

    # ══ HOJA 3 — RESUMEN ══════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Resumen")
    for i, w in enumerate([30, 18, 18, 18, 14, 16], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    mw(ws3, 1, 1, 6, "RESUMEN — CONTROL DE DESPACHO vs VENTAS DEL DÍA", F_HDR, bold=True, sz=13, fc="FFFFFF", al=al_c)
    ws3.row_dimensions[1].height = 32

    tot_gs = tots["gs"]; tot_gp = tots["gp"]; tot_gd = tots["gd"]
    tot_iva = tots["iva"]; tot_ieps = tots["ieps"]; tot_imp = tots["tot"]
    v_tot_gs = tots["vgs"]; v_tot_gp = tots["vgp"]; v_tot_gd = tots["vgd"]
    v_tot_iva = tots["viva"]; v_tot_ieps = tots["vieps"]; v_tot_total = tots["vtot"]
    uuid_gs   = len({(r, u) for (r, u), e in data.items() if e["GS"]["tot"] > 0})
    uuid_gp   = len({(r, u) for (r, u), e in data.items() if e["GP"]["tot"] > 0})
    uuid_gd   = len({(r, u) for (r, u), e in data.items() if e["GD"]["tot"] > 0})
    uuid_total = len(data); rfc_total = len({r for (r, u) in data.keys()})
    TTOT = max(tot_gs + tot_gp + tot_gd, 1)

    sin_resumen = defaultdict(lambda: {"GS": 0.0, "GP": 0.0, "GD": 0.0, "n": 0})
    for (cli, dt), pd in sin_uuid.items():
        for pr in ("GS", "GP", "GD"):
            sin_resumen[cli][pr] += pd[pr]["tot"]; sin_resumen[cli]["n"] += pd[pr]["n"]

    R = 3
    mw(ws3, R, 1, 6, "VENTAS POR PRODUCTO", F_DARK, bold=True, sz=10, fc="FFFFFF", al=al_l)
    ws3.row_dimensions[R].height = 22; R += 1
    for i, h in enumerate(["Producto", "Subtotal Despachos", "Ventas del Día", "Diferencia", "% Total", "CFDIs"], 1):
        cw(ws3, R, i, val=h, fi=F_HDR, al=al_c, bold=True, sz=8, fc="FFFFFF")
    ws3.row_dimensions[R].height = 16; R += 1
    for lbl, desp, pol, nuuid, fi_r in [
            ("G-Super (GS)", tot_gs, v_tot_gs, uuid_gs, F_GS),
            ("G-Premium (GP)", tot_gp, v_tot_gp, uuid_gp, F_GP),
            ("G-Diesel (GD)", tot_gd, v_tot_gd, uuid_gd, F_GD)]:
        dif = round(desp - pol, 2); fi_dif = F_OK if abs(dif) < 1 else F_WRN if abs(dif) < 500 else F_ERR
        cw(ws3, R, 1, val=lbl,  fi=fi_r, al=al_l, sz=9)
        cw(ws3, R, 2, val=desp, fi=fi_r, al=al_r, fmt=MF, sz=9)
        cw(ws3, R, 3, val=pol,  fi=fi_r, al=al_r, fmt=MF, sz=9)
        cw(ws3, R, 4, val=dif,  fi=fi_dif, al=al_r, fmt=MF, bold=True, sz=9)
        cw(ws3, R, 5, val=desp / TTOT, fi=fi_r, al=al_r, fmt="0.0%", sz=9)
        cw(ws3, R, 6, val=nuuid, fi=fi_r, al=al_r, fmt="#,##0", sz=9)
        ws3.row_dimensions[R].height = 16; R += 1
    TTOT_r = tot_gs + tot_gp + tot_gd; VTOT_r = v_tot_gs + v_tot_gp + v_tot_gd
    dif_sub = round(TTOT_r - VTOT_r, 2)
    fi_dt = F_OK if abs(dif_sub) < 1 else F_WRN if abs(dif_sub) < 500 else F_ERR
    cw(ws3, R, 1, val="TOTAL",  fi=F_TOT, al=al_l, bold=True, sz=10)
    cw(ws3, R, 2, val=TTOT_r,  fi=F_TOT, al=al_r, fmt=MF, bold=True, sz=10)
    cw(ws3, R, 3, val=VTOT_r,  fi=F_TOT, al=al_r, fmt=MF, bold=True, sz=10)
    cw(ws3, R, 4, val=dif_sub, fi=fi_dt, al=al_r, fmt=MF, bold=True, sz=10)
    cw(ws3, R, 5, val=1.0,     fi=F_TOT, al=al_r, fmt="0.0%", bold=True, sz=10)
    cw(ws3, R, 6, val=uuid_total, fi=F_TOT, al=al_r, fmt="#,##0", bold=True, sz=10)
    ws3.row_dimensions[R].height = 22; R += 2

    sin_tot_gs = sin_tot_gp = sin_tot_gd = sin_tot_n = 0
    mw(ws3, R, 1, 6, "⚠  SIN COMPROBANTE FISCAL — POR CLIENTE (col Q)", F_SIN_H, bold=True, sz=10, fc="FFFFFF", al=al_l)
    ws3.row_dimensions[R].height = 22; R += 1
    for i, h in enumerate(["Cliente (col Q)", "G-Super", "G-Premium", "G-Diesel", "TOTAL", "# Tx"], 1):
        cw(ws3, R, i, val=h, fi=F_SIN, al=al_c, bold=True, sz=8, fc="92400E")
    ws3.row_dimensions[R].height = 16; R += 1
    for cli in sorted(sin_resumen.keys()):
        d = sin_resumen[cli]; tc = d["GS"] + d["GP"] + d["GD"]
        cw(ws3, R, 1, val=cli,   fi=F_SIN_ROW, al=al_l, sz=9, fc="7C2D12")
        cw(ws3, R, 2, val=d["GS"], fi=F_SIN_ROW, al=al_r, fmt=MF, sz=9)
        cw(ws3, R, 3, val=d["GP"], fi=F_SIN_ROW, al=al_r, fmt=MF, sz=9)
        cw(ws3, R, 4, val=d["GD"], fi=F_SIN_ROW, al=al_r, fmt=MF, sz=9)
        cw(ws3, R, 5, val=tc,    fi=F_SIN_ROW, al=al_r, fmt=MF, bold=True, sz=9)
        cw(ws3, R, 6, val=d["n"], fi=F_SIN_ROW, al=al_r, fmt="#,##0", sz=9)
        sin_tot_gs += d["GS"]; sin_tot_gp += d["GP"]; sin_tot_gd += d["GD"]; sin_tot_n += d["n"]
        ws3.row_dimensions[R].height = 16; R += 1
    cw(ws3, R, 1, val="TOTAL SIN COMPROBANTE", fi=F_SIN_TOT, al=al_l, bold=True, sz=10, fc="FFFFFF")
    cw(ws3, R, 2, val=sin_tot_gs, fi=F_SIN_TOT, al=al_r, fmt=MF, bold=True, sz=10, fc="FFFFFF")
    cw(ws3, R, 3, val=sin_tot_gp, fi=F_SIN_TOT, al=al_r, fmt=MF, bold=True, sz=10, fc="FFFFFF")
    cw(ws3, R, 4, val=sin_tot_gd, fi=F_SIN_TOT, al=al_r, fmt=MF, bold=True, sz=10, fc="FFFFFF")
    cw(ws3, R, 5, val=sin_tot_gs + sin_tot_gp + sin_tot_gd, fi=F_SIN_TOT, al=al_r, fmt=MF, bold=True, sz=10, fc="FFFFFF")
    cw(ws3, R, 6, val=sin_tot_n, fi=F_SIN_TOT, al=al_r, fmt="#,##0", bold=True, sz=10, fc="FFFFFF")
    ws3.row_dimensions[R].height = 22; R += 2

    mw(ws3, R, 1, 6, "IMPUESTOS", F_DARK, bold=True, sz=10, fc="FFFFFF", al=al_l)
    ws3.row_dimensions[R].height = 22; R += 1
    for i, h in enumerate(["Concepto", "Despachos", "Ventas del Día", "Diferencia", "", ""], 1):
        cw(ws3, R, i, val=h, fi=F_HDR, al=al_c, bold=True, sz=8, fc="FFFFFF")
    ws3.row_dimensions[R].height = 16; R += 1
    for lbl, desp, pol in [("IVA 16%", tot_iva, v_tot_iva), ("IEPS", tot_ieps, v_tot_ieps),
                            ("TOTAL con impuestos", tot_imp, v_tot_total)]:
        dif = round(desp - pol, 2); fi_dif = F_OK if abs(dif) < 1 else F_WRN if abs(dif) < 500 else F_ERR
        is_tot = "TOTAL" in lbl; fi_r = F_TOT if is_tot else fill("FCE7F3")
        cw(ws3, R, 1, val=lbl,  fi=fi_r, al=al_l, bold=is_tot, sz=9)
        cw(ws3, R, 2, val=desp, fi=fi_r, al=al_r, fmt=MF, bold=is_tot, sz=9)
        cw(ws3, R, 3, val=pol,  fi=fi_r, al=al_r, fmt=MF, bold=is_tot, sz=9)
        cw(ws3, R, 4, val=dif,  fi=fi_dif, al=al_r, fmt=MF, bold=True, sz=9)
        cw(ws3, R, 5, fi=fi_r); cw(ws3, R, 6, fi=fi_r)
        ws3.row_dimensions[R].height = 16; R += 1
    R += 1
    mw(ws3, R, 1, 6, "MÉTRICAS", F_DARK, bold=True, sz=10, fc="FFFFFF", al=al_l)
    ws3.row_dimensions[R].height = 22; R += 1
    F_LIGHT = fill("F8FAFC")
    for lbl, val, fi_v in [
            ("Total CFDIs únicos", uuid_total, F_GS),
            ("   ↳ G-Super",  uuid_gs, F_GS), ("   ↳ G-Premium", uuid_gp, F_GP),
            ("   ↳ G-Diesel", uuid_gd, F_GD), ("Clientes únicos (RFC)", rfc_total, F_UUID),
            ("Tx sin comprobante", sin_tot_n, F_SIN), ("Días conciliados", len(fechas), F_UUID)]:
        cw(ws3, R, 1, val=lbl, fi=F_LIGHT, al=al_l, sz=9)
        cw(ws3, R, 2, val=val, fi=fi_v, al=al_r, fmt="#,##0", bold=True, sz=11)
        for ci in range(3, 7): cw(ws3, R, ci, fi=F_LIGHT)
        ws3.row_dimensions[R].height = 18; R += 1

    wb.move_sheet("Resumen", offset=-2)

    # ── Serializar ────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    dif_final = round(tots["tot"] - tots["vtot"], 2)
    if abs(dif_final) < 1:
        logs.append(f"✅ Conciliación cuadra: diferencia total = {dif_final:,.2f}")
    else:
        logs.append(f"⚠️  Diferencia total: {dif_final:,.2f}")

    return buf.read(), logs, resumen_dias


# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("### 📂 Archivos de entrada")
col1, col2 = st.columns(2)
with col1:
    desp_file = st.file_uploader(
        "📊 Control de Despachos (.xlsx)",
        type=["xlsx"],
        help="Archivo de control de despachos. Col D=producto, Col Q=cliente, Col W=UUID.",
    )
with col2:
    vd_file = st.file_uploader(
        "⛽ Póliza Ventas del Día (.xlsx)",
        type=["xlsx"],
        help="Excel de póliza generado por el módulo Ventas del Día.",
    )

st.markdown("")
generar = st.button(
    "📊  Generar Conciliación",
    type="primary",
    disabled=(desp_file is None or vd_file is None),
    use_container_width=True,
)

if desp_file is None or vd_file is None:
    st.info("👆 Selecciona ambos archivos para comenzar.")

if generar and desp_file is not None and vd_file is not None:
    with st.spinner("Procesando..."):
        try:
            excel_bytes, logs, resumen_dias = procesar(
                desp_file.read(),
                vd_file.read(),
            )

            base = desp_file.name.rsplit(".", 1)[0]
            nombre_salida = f"Conciliacion_DespachoVentas_{base}.xlsx"

            st.success(f"✅ Conciliación generada — {len(resumen_dias)} día(s)")

            st.download_button(
                label="💾  Descargar Excel",
                data=excel_bytes,
                file_name=nombre_salida,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            # ── Tabla resumen por día ──────────────────────────────────────
            if resumen_dias:
                st.markdown("### 📋 Resumen por día")
                import pandas as pd
                df = pd.DataFrame(resumen_dias)

                def _color_dif(v):
                    if abs(v) < 1:
                        return "background-color:#C6EFCE; color:#276221"
                    elif abs(v) < 1000:
                        return "background-color:#FFEB9C; color:#9C5700"
                    return "background-color:#FFC7CE; color:#9C0006"

                styled = (
                    df.style
                    .format({
                        "Tot Despachos": "{:,.2f}",
                        "Tot VtaDía":    "{:,.2f}",
                        "Diferencia":    "{:,.2f}",
                    })
                    .map(_color_dif, subset=["Diferencia"])
                )
                st.dataframe(styled, use_container_width=True, hide_index=True)

            # ── Log ────────────────────────────────────────────────────────
            with st.expander("📋 Log de procesamiento"):
                for line in logs:
                    st.text(line)

        except Exception as exc:
            import traceback
            st.error(f"❌ Error: {exc}")
            with st.expander("Detalle del error"):
                st.code(traceback.format_exc())

st.markdown("---")
st.caption("Módulo Control de Despacho vs Ventas · AUXILIAR DE REGISTROS")
