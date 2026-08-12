"""
Módulo: Depósitos Bancarios
Genera automáticamente la póliza de depósitos bancarios a partir de los
estados de cuenta de BBVA, Banorte e Inbursa.
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import re
from datetime import datetime

st.set_page_config(page_title="Depósitos Bancarios", page_icon="🏦", layout="wide")

# ─── Estilos ──────────────────────────────────────────────────────────────────
import _theme
_theme.aplicar_header("🏦 Depósitos Bancarios", "BBVA, Banorte e Inbursa → póliza de depósitos con cuentas de tránsito")
st.caption("Genera la póliza contable desde los estados de cuenta de BBVA, Banorte e Inbursa.")

# ─── Constantes contables ──────────────────────────────────────────────────────
CARGOS = {
    "BANORTE": ("102-01-0001-0001", "Banorte"),
    "BBVA":    ("102-01-0001-0003", "BBVA"),
    "INBURSA": ("102-01-0001-0002", "INBURSA"),
}

# Orden de columnas de abonos (cols 12-19 en la póliza)
ABONOS = [
    {"col": 12, "cuenta": "106-01-0001-0010", "nombre": "DEPOSITO EN TRANSITO T AMERICAN EXPRESS"},
    {"col": 13, "cuenta": "106-01-0001-0005", "nombre": "DEPOSITO EN TRANSITO  EFECTIVALE"},
    {"col": 14, "cuenta": "106-01-0001-0009", "nombre": "DEPOSITO EN TRANSITO  TICKET CARD EDENRED"},
    {"col": 15, "cuenta": "101-01-0001",       "nombre": "Fondo Fijo de Caja"},
    {"col": 16, "cuenta": "106-01-0001-0011",  "nombre": "DEPOSITO EN TRANSITO T BANORTE"},
    {"col": 17, "cuenta": "106-01-0001-0003",  "nombre": "DEPOSITO EN TRANSITO  SMARTBT - SHELL FLEET"},
    {"col": 18, "cuenta": "106-01-0001-0013",  "nombre": "BBVA"},
    {"col": 19, "cuenta": "106-01-0001-0012",  "nombre": "DEPOSITO EN TRANSITO T INBURSA"},
]

COL_ABONO_IDX = {a["col"]: a for a in ABONOS}

# ─── Clasificación de depósitos ───────────────────────────────────────────────
def _hora_citi(desc: str) -> int:
    """Extrae la hora del SPEI de CITI para distinguir AMEX vs Shell."""
    m = re.search(r"HR LIQ:\s*(\d{2}):", desc)
    return int(m.group(1)) if m else 0


def clasificar_bbva(desc: str, monto: float):
    """
    BBVA:
      *AMEX* en cualquier parte           → col 12 (AMEX transit)
      VENTAS PUNTOS TDC / VENTAS CREDITO
        / TERMINALES PV / TDC INTER       → col 18 (Dep. Tránsito Tarjetas Bancomer)
      DEPOSITO EN EFECTIVO                → col 15 (Caja 101-01-0001)
      Cualquier otra descripción          → None  (no clasificado)
    """
    d = desc.upper()
    if "AMEX" in d:
        return 12
    if ("VENTAS PUNTOS TDC" in d or "VENTAS CREDITO" in d
            or "TERMINALES PUNTO DE VENTA" in d
            or "VENTAS TDC INTER" in d or "TDC INTER" in d):
        return 18
    if "DEPOSITO EN EFECTIVO" in d or "DEP.EFECTIVO" in d or "DEP EN EFECTIVO" in d:
        return 15
    return None


def clasificar_inbursa(desc: str, monto: float):
    """INBURSA: solo 'INBURED' → INBURSA transit (col 19). Resto → None."""
    d = desc.upper()
    if "INBURED" in d:
        return 19
    return None


def clasificar_banorte(desc: str, monto: float):
    """
    BANORTE: busca la palabra clave en la descripción y devuelve la columna
    de abono. Devuelve None si no hay coincidencia (no se registra).
    """
    d = desc.upper()

    # Devoluciones SPEI → nunca son un depósito real
    if "DEV.SPEI" in d or "DEVOLUCION SPEI" in d or d.startswith("DEV."):
        return None

    # Cheques → no clasificar
    if "DEP. CH." in d or "CHEQUE SBC" in d:
        return None

    # Compensación desfase → ignorar
    if "COMPENSACION DESFASE" in d:
        return None

    # FELUSA SPEI (transferencia interna desde cualquier banco) → no clasificar
    if "SPEI RECIBIDO" in d and "SERVICIOS FELUSA" in d:
        return None

    # SHELL FLEET / SMARTBT (excluye BNET genérico que no es Shell)
    if "SHELL" in d or "SMARTBT" in d:
        return 17

    # AMERICAN EXPRESS / CITI MEXICO
    if "AMERICAN EXPRESS" in d or "BCO:0124" in d or "CITI MEXICO" in d:
        m = re.search(r"HR LIQ:\s*(\d{2}):", desc)
        hora = int(m.group(1)) if m else 0
        return 17 if hora >= 12 else 12

    # EFECTIVALE / SANTANDER
    if "EFECTIVALE" in d or "EFE8908015L3" in d or "BCO:0014" in d or (
            "SANTANDER" in d and "SPEI RECIBIDO" in d):
        return 13

    # EDENRED / HSBC
    if "EDENRED" in d or "HSBCPGMD" in d or "BCO:0021" in d:
        return 14

    # Efectivo en caja
    if "DEP.EFECTIVO" in d or "DEPOSITO EN EFECTIVO" in d:
        return 15

    # INBURSA via SPEI (BCO:0036 = código Inbursa en SPEI) → tránsito Inbursa
    if "BCO:0036" in d or ("INBURSA" in d and "SPEI RECIBIDO" in d):
        return 19

    # GASnGO → tránsito GASnGO-BANORTE (match exacto del nombre de empresa)
    if "GASNGO" in d or "GASN GO" in d:
        return 16

    # ICIGAS
    if "ICIGAS" in d:
        return 17

    # SERV <NOMBRE> <DÍGITOS>C/D
    if "07277262C" in d or "07277262D" in d or (
            "SERV" in d and re.search(r'\d{5,}[CD]', d)):
        if "AMERICAN" in d: return 12
        if "EFECTIVALE" in d: return 13
        if "EDENRED" in d or "TICKET" in d: return 14
        if "SHELL" in d or "SMARTBT" in d: return 17
        if "INBURSA" in d: return 19
        return 16

    return None


CLASIFICADORES = {
    "BBVA":    clasificar_bbva,
    "BANORTE": clasificar_banorte,
    "INBURSA": clasificar_inbursa,
}

CARGO_COL = {"BBVA": 9, "BANORTE": 8, "INBURSA": 10}
# ─── Lectura de estados de cuenta ─────────────────────────────────────────────
def leer_banco(file_obj, banco: str) -> tuple[list, list]:
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    inicio = 2
    for i, r in enumerate(rows):
        if r and str(r[0]).strip().lower() == "fecha":
            inicio = i + 1
            break

    ok, sin_clasif = [], []
    fn = CLASIFICADORES[banco]
    col_cargo = CARGO_COL[banco]

    for idx, r in enumerate(rows[inicio:], start=inicio):
        if not r or r[0] is None:
            continue
        fecha = r[0]
        if not hasattr(fecha, "day"):
            continue
        desc  = str(r[1] or "").strip()
        monto = r[2]
        if not monto or monto <= 0:
            continue

        col_abono = fn(desc, monto)
        if col_abono is None:
            sin_clasif.append({"fecha": fecha, "banco": banco,
                               "descripcion": desc[:80], "monto": monto})
        else:
            ok.append({
                "fecha":      fecha,
                "banco":      banco,
                "ref":        f"DEPOSITOS {banco}",
                "desc":       desc,
                "monto":      monto,
                "col_cargo":  col_cargo,
                "col_abono":  col_abono,
                "fila_excel": idx + 1,
            })

    return ok, sin_clasif


def marcar_estado_cuenta(file_bytes: bytes, filas_usadas: set, banco: str = "") -> bytes:
    BANK_STYLE = {
        "BBVA":    ("A9C9EF", "004481"),
        "BANORTE": ("FBBDBD", "C8102E"),
        "INBURSA": ("C8E6C9", "1B5E20"),
    }
    fondo, color_txt = BANK_STYLE.get(banco.upper(), ("A8D5B5", "1A5276"))
    row_fill = PatternFill("solid", fgColor=fondo)
    _s = Side(style="thin", color=color_txt)
    row_border = Border(left=_s, right=_s, top=_s, bottom=_s)

    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    ws = wb.active
    for fila in sorted(filas_usadas):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=fila, column=col)
            cell.fill = row_fill
            cell.border = row_border
            old = cell.font
            cell.font = Font(
                name=old.name or "Calibri",
                size=old.size or 11,
                bold=True,
                color=color_txt,
                italic=old.italic,
            )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
# ─── Generación del Excel ────────────────────────────────────────────────────
def generar_excel(registros: list, plantilla=None) -> bytes:
    FONT_NAME = "Aptos Narrow"
    FONT_SIZE = 11

    F_ADMIN   = PatternFill("solid", fgColor="1E293B")
    F_GRAY2   = PatternFill("solid", fgColor="1E293B")
    # Fondos de fila — familia azul por banco
    F_BBVA_1  = PatternFill("solid", fgColor="EFF6FF")  # blue-50
    F_BBVA_2  = PatternFill("solid", fgColor="DBEAFE")  # blue-100
    F_BNT_1   = PatternFill("solid", fgColor="EEF2FF")  # indigo-50
    F_BNT_2   = PatternFill("solid", fgColor="E0E7FF")  # indigo-100
    F_INB_1   = PatternFill("solid", fgColor="F0F9FF")  # sky-50
    F_INB_2   = PatternFill("solid", fgColor="E0F2FE")  # sky-100
    F_NONE    = PatternFill(fill_type=None)

    _S = Side(style="thin", color="999999")
    BORDER = Border(left=_S, right=_S, top=_S, bottom=_S)
    _SH = Side(style="medium", color="FFFFFF")
    BORDER_H = Border(left=_SH, right=_SH, top=_SH, bottom=_SH)

    def fnt(bold=False, color="000000", size=None, italic=False):
        return Font(name=FONT_NAME, size=size or FONT_SIZE,
                    bold=bold, color=color, italic=italic)

    A_CTR   = Alignment(horizontal="center", vertical="center", wrap_text=False)
    A_CTR_W = Alignment(horizontal="center", vertical="center", wrap_text=True)
    A_LEFT  = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    A_RIGHT = Alignment(horizontal="right",  vertical="center", wrap_text=False)

    FMT_NUM  = '#,##0.00'
    FMT_DATE = 'DD/MM/YYYY'

    def set_cell(ws, row, col, value=None, font=None, fill=F_NONE,
                 align=A_CTR, num_format=None, border=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:       c.font = font
        if fill:       c.fill = fill
        if align:      c.alignment = align
        if num_format: c.number_format = num_format
        if border:     c.border = border
        return c
    # ── Keyword → col semántico (para color de encabezado) ───────────────────────
    _ABONO_KW = [
        (["AMEX", "AMERICAN"],                    12),
        (["EFECTIVALE"],                           13),
        (["EDENRED", "TICKET", "TICKETCARD"],      14),
        (["FONDO", "CAJA"],                        15),
        (["GASNGO", "GASN"],                       16),
        (["SHELL", "SMARTBT", "ICIGAS"],           17),
        (["BANCOMER", "BBVA", "TDC", "VISA"],      18),
        (["INBURSA"],                              19),
    ]
    # Col semántico → color header de abono
    _SEM_FILL = {
        12: PatternFill("solid", fgColor="312E81"),  # AMEX — índigo-900
        13: PatternFill("solid", fgColor="3730A3"),  # EFECTIVALE — índigo-800
        14: PatternFill("solid", fgColor="1D4ED8"),  # EDENRED — blue-700
        15: PatternFill("solid", fgColor="1E40AF"),  # CAJA — blue-800
        16: PatternFill("solid", fgColor="0C4A6E"),  # GASNGO — sky-900
        17: PatternFill("solid", fgColor="164E63"),  # SHELL/ICIGAS — cyan-900
        18: PatternFill("solid", fgColor="1E3A8A"),  # BBVA — blue-800
        19: PatternFill("solid", fgColor="0F4C81"),  # INBURSA — blue oscuro
    }
    _DEFAULT_AB_FILL = PatternFill("solid", fgColor="1E3A8A")

    # Cargo banco → (fill header, color texto datos) — familia azul/índigo/cielo
    _CARGO_STYLE = {
        "BANORTE": (PatternFill("solid", fgColor="312E81"), "4338CA"),  # índigo
        "BBVA":    (PatternFill("solid", fgColor="1E3A8A"), "1D4ED8"),  # azul
        "INBURSA": (PatternFill("solid", fgColor="0C4A6E"), "0369A1"),  # cielo
    }

    # ── Construir listas 100 % desde CUENTAS ─────────────────────────────────────
    # cargos_list  : [{banco, cuenta, col}]  — col = Excel col real (1-indexed)
    # abonos_raw   : [{cuenta, nombre, sem_col}]  — col asignada tras cargos
    cargos_list = []
    abonos_raw  = []

    if plantilla is not None:
        wb = openpyxl.load_workbook(plantilla)
        ws = wb["POLIZA"] if "POLIZA" in wb.sheetnames else wb.active
        for _ri in range(1, ws.max_row + 1):
            for _ci in range(1, ws.max_column + 1):
                ws.cell(row=_ri, column=_ci).value = None

        _hoja_cuentas = next(
            (s for s in wb.sheetnames if s.strip().upper() == "CUENTAS"), None)
        if _hoja_cuentas:
            _wc = wb[_hoja_cuentas]
            for _row in _wc.iter_rows(min_row=3, values_only=True):
                # CARGOS (cols 1-2)
                _cta   = str(_row[0] or "").strip()
                _banco = str(_row[1] or "").strip().upper()
                if _cta and _banco:
                    for _key in ("BANORTE", "BBVA", "INBURSA"):
                        if _key in _banco and not any(c["banco"] == _key for c in cargos_list):
                            cargos_list.append({"banco": _key, "cuenta": _cta, "col": 0})
                            break
                # ABONOS (cols 4-5)
                _cta_ab  = str(_row[3] or "").strip() if len(_row) > 3 else ""
                _nom_raw = (str(_row[4] or "").strip().replace("\n", " ").strip()
                            if len(_row) > 4 else "")
                if _cta_ab and _nom_raw:
                    _nom_up  = _nom_raw.upper().replace(" ", "")
                    _sem_col = None
                    for _kws, _sc in _ABONO_KW:
                        if any(kw.replace(" ", "") in _nom_up for kw in _kws):
                            _sem_col = _sc
                            break
                    abonos_raw.append({"cuenta": _cta_ab, "nombre": _nom_raw,
                                       "sem_col": _sem_col})
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "POLIZA"

    # Defaults si no hay plantilla o CUENTAS vacía
    if not cargos_list:
        cargos_list = [{"banco": b, "cuenta": CARGOS[b][0], "col": 0}
                       for b in ("BANORTE", "BBVA", "INBURSA")]
    if not abonos_raw:
        abonos_raw = [{"cuenta": a["cuenta"], "nombre": a["nombre"],
                       "sem_col": a["col"]} for a in ABONOS]

    # ── Asignar columnas Excel reales (todo 1-indexed, sin +1 posterior) ────────
    CARGO_START = 9
    for _i, _c in enumerate(cargos_list):
        _c["col"] = CARGO_START + _i

    N_CARGOS       = len(cargos_list)
    COL_TOT_CARGOS = CARGO_START + N_CARGOS      # Excel col TOTAL CARGOS
    ABONO_START    = COL_TOT_CARGOS + 1           # Excel col primer abono

    abonos_efectivos  = []
    semantic_to_actual = {}
    for _i, _a in enumerate(abonos_raw):
        _acol = ABONO_START + _i
        abonos_efectivos.append({**_a, "col": _acol})
        if _a["sem_col"] is not None:
            semantic_to_actual[_a["sem_col"]] = _acol

    N_ABONOS       = len(abonos_efectivos)
    COL_TOT_ABONOS = ABONO_START + N_ABONOS
    COL_DIFERENCIA = COL_TOT_ABONOS + 1
    N_COLS         = COL_DIFERENCIA

    # Mapa banco → col cargo real
    cargos_col_map = {_c["banco"]: _c["col"] for _c in cargos_list}

    # ── Fills de color ────────────────────────────────────────────────────────────
    F_META_HDR = PatternFill("solid", fgColor="0F172A")   # slate-950
    F_TOT_CARG = PatternFill("solid", fgColor="6D28D9")   # violet-700
    F_TOT_ABON = PatternFill("solid", fgColor="0369A1")   # sky-700
    F_DIFF_HDR = PatternFill("solid", fgColor="4C1D95")   # violet-900
    F_CTA_BG   = PatternFill("solid", fgColor="1E293B")   # slate-800

    # ── Fila 1: numeración ────────────────────────────────────────────────────────
    for _idx in range(N_COLS):
        set_cell(ws, 1, _idx + 1, value=_idx,
                 font=fnt(color="94A3B8", size=8), fill=F_NONE, align=A_CTR)

    # ── Fila 2: N° de cuentas ─────────────────────────────────────────────────────
    fnt_cta = fnt(bold=False, color="CBD5E1", size=8, italic=True)
    for _c in cargos_list:
        set_cell(ws, 2, _c["col"], value=_c["cuenta"],
                 font=fnt_cta, fill=F_CTA_BG, align=A_CTR, border=BORDER)
    for _ab in abonos_efectivos:
        set_cell(ws, 2, _ab["col"], value=_ab["cuenta"],
                 font=fnt_cta, fill=F_CTA_BG, align=A_CTR, border=BORDER)

    # ── Fila 3: encabezados ───────────────────────────────────────────────────────
    fnt_h = lambda: fnt(bold=True, color="FFFFFF")
    for _col, _lbl in [(1,"TIPO"),(2,"FECHA"),(3,"REFERENCIA"),(4,"CONCEPTO"),
                       (5,"ERROR"),(6,"UIDD"),(7,"NÚM PÓLIZA"),(8,"PROCESADO")]:
        set_cell(ws, 3, _col, value=_lbl, font=fnt_h(), fill=F_META_HDR,
                 align=A_CTR, border=BORDER_H)
    for _c in cargos_list:
        _fh, _ = _CARGO_STYLE.get(_c["banco"], (_DEFAULT_AB_FILL, "FFFFFF"))
        set_cell(ws, 3, _c["col"],
                 value=CARGOS.get(_c["banco"], ("", _c["banco"]))[1].strip() or _c["banco"],
                 font=fnt_h(), fill=_fh, align=A_CTR, border=BORDER_H)
    set_cell(ws, 3, COL_TOT_CARGOS, value="TOTAL CARGOS",
             font=fnt_h(), fill=F_TOT_CARG, align=A_CTR, border=BORDER_H)
    for _ab in abonos_efectivos:
        _fh = _SEM_FILL.get(_ab.get("sem_col"), _DEFAULT_AB_FILL)
        set_cell(ws, 3, _ab["col"], value=_ab["nombre"],
                 font=fnt_h(), fill=_fh, align=A_CTR_W, border=BORDER_H)
    set_cell(ws, 3, COL_TOT_ABONOS, value="TOTAL ABONOS",
             font=fnt_h(), fill=F_TOT_ABON, align=A_CTR, border=BORDER_H)
    set_cell(ws, 3, COL_DIFERENCIA, value="DIFERENCIA",
             font=fnt_h(), fill=F_DIFF_HDR, align=A_CTR, border=BORDER_H)

    # ── Filas de datos ────────────────────────────────────────────────────────────
    MESES_ES = {1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",5:"MAYO",6:"JUNIO",
                7:"JULIO",8:"AGOSTO",9:"SEPTIEMBRE",10:"OCTUBRE",11:"NOVIEMBRE",12:"DICIEMBRE"}
    orden_banco = {"BBVA": 0, "INBURSA": 1, "BANORTE": 2}
    registros_sorted = sorted(registros,
                               key=lambda x: (
                                   getattr(x["fecha"],"year",0),
                                   getattr(x["fecha"],"month",0),
                                   orden_banco.get(x["banco"], 9),
                                   x["fecha"]))

    # Inyectar separadores de mes
    _rows_excel = []
    _prev_mes   = None
    for _rx in registros_sorted:
        _mk = (getattr(_rx["fecha"],"year",0), getattr(_rx["fecha"],"month",0))
        if _mk != _prev_mes:
            _rows_excel.append({"_type": "mes_hdr",
                                 "label": f"{MESES_ES.get(_mk[1],'')} {_mk[0]}"})
            _prev_mes = _mk
        _rows_excel.append(_rx)

    FILLS_BANCO = {
        "BBVA":    (F_BBVA_1, F_BBVA_2),
        "INBURSA": (F_INB_1,  F_INB_2),
        "BANORTE": (F_BNT_1,  F_BNT_2),
    }
    BANCO_COLOR = {"BBVA": "1D4ED8", "BANORTE": "4338CA", "INBURSA": "0369A1"}

    fila_num    = 4
    _row_parity = 0
    for _item in _rows_excel:
        if _item.get("_type") == "mes_hdr":
            # Fila separadora de mes (fondo oscuro, texto centrado)
            ws.merge_cells(start_row=fila_num, start_column=1,
                           end_row=fila_num, end_column=N_COLS)
            _mhc = ws.cell(row=fila_num, column=1, value=_item["label"])
            _mhc.font      = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
            _mhc.fill      = F_META_HDR
            _mhc.alignment = A_CTR
            _mhc.border    = BORDER_H
            ws.row_dimensions[fila_num].height = 22
            fila_num    += 1
            _row_parity  = 0
            continue
        r            = _item
        _row_parity += 1
        _f1, _f2    = FILLS_BANCO.get(r["banco"], (F_BBVA_1, F_BBVA_2))
        fill_row    = _f1 if _row_parity % 2 == 1 else _f2
        fn_dat      = fnt(color="1E293B")
        _, _col_color = _CARGO_STYLE.get(r["banco"], (_DEFAULT_AB_FILL, "374151"))

        def dat(col, val, num_fmt=None, align=A_CTR, _fn=fila_num, _fr=fill_row):
            c = set_cell(ws, _fn, col, value=val,
                         font=fn_dat, fill=_fr, align=align, border=BORDER)
            if num_fmt: c.number_format = num_fmt
            return c

        monto = r["monto"]
        set_cell(ws, fila_num, 1, value="I",
                 font=fnt(bold=True, color=_col_color),
                 fill=fill_row, align=A_CTR, border=BORDER)
        # Fecha → formato real de fecha (no texto)
        _cf = ws.cell(row=fila_num, column=2, value=r["fecha"])
        _cf.font = fn_dat; _cf.fill = fill_row
        _cf.alignment = A_CTR; _cf.border = BORDER
        _cf.number_format = FMT_DATE
        dat(3, r["ref"],  align=A_LEFT)
        dat(4, r["ref"],  align=A_LEFT)
        for _c_ in [5, 6, 7, 8]:
            dat(_c_, None)

        # Cargo dinámico por banco
        _cargo_col = cargos_col_map.get(r["banco"])
        if _cargo_col:
            dat(_cargo_col, monto, num_fmt=FMT_NUM, align=A_RIGHT)
        dat(COL_TOT_CARGOS, monto, num_fmt=FMT_NUM, align=A_RIGHT)

        # Abono: col semántico → col real
        _actual_ab = semantic_to_actual.get(r["col_abono"])
        if _actual_ab is not None:
            dat(_actual_ab, monto, num_fmt=FMT_NUM, align=A_RIGHT)
        dat(COL_TOT_ABONOS, monto, num_fmt=FMT_NUM, align=A_RIGHT)

        _cL  = get_column_letter(COL_TOT_CARGOS)
        _cU  = get_column_letter(COL_TOT_ABONOS)
        _cd  = ws.cell(row=fila_num, column=COL_DIFERENCIA,
                       value=f"={_cL}{fila_num}-{_cU}{fila_num}")
        _cd.font = fnt(bold=True, color="4C1D95")
        _cd.fill = fill_row; _cd.border = BORDER
        _cd.alignment = A_RIGHT; _cd.number_format = FMT_NUM
        ws.row_dimensions[fila_num].height = 18
        fila_num += 1

    # ── Anchos de columna ─────────────────────────────────────────────────────────
    for _cn, _w in {1:5.0, 2:13.0, 3:38.0, 4:28.0,
                    5:6.7, 6:5.3, 7:11.3, 8:11.6}.items():
        ws.column_dimensions[get_column_letter(_cn)].width = _w
    for _c in cargos_list:
        ws.column_dimensions[get_column_letter(_c["col"])].width = 16.0
    ws.column_dimensions[get_column_letter(COL_TOT_CARGOS)].width = 15.0
    for _ab in abonos_efectivos:
        ws.column_dimensions[get_column_letter(_ab["col"])].width = 20.0
    ws.column_dimensions[get_column_letter(COL_TOT_ABONOS)].width = 15.0
    ws.column_dimensions[get_column_letter(COL_DIFERENCIA)].width  = 13.0
    ws.row_dimensions[1].height = 12
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 36
    ws.freeze_panes = "B4"

    # ── Hoja RESUMEN ──────────────────────────────────────────────────────────────
    from collections import defaultdict
    from datetime import date as _date_cls

    ws_res = wb.create_sheet("RESUMEN")

    # Fills
    F_R_TITLE  = PatternFill("solid", fgColor="0F172A")
    F_R_HDR    = PatternFill("solid", fgColor="1E293B")
    F_R_TOT_M  = PatternFill("solid", fgColor="1D4ED8")   # total mes
    F_R_TOT_G  = PatternFill("solid", fgColor="0F172A")   # total global
    F_R_DIFF_G = PatternFill("solid", fgColor="065F46")
    F_R_DIFF_R = PatternFill("solid", fgColor="7F1D1D")
    F_R_ODD    = PatternFill("solid", fgColor="EEF2FF")
    F_R_EVN    = PatternFill("solid", fgColor="E0F2FE")
    F_R_BNK = {
        "BANORTE": PatternFill("solid", fgColor="312E81"),
        "BBVA":    PatternFill("solid", fgColor="1E3A8A"),
        "INBURSA": PatternFill("solid", fgColor="0C4A6E"),
    }
    # 12 tonos azul/índigo/cielo para meses
    _MES_PAL = [
        "0F172A","1E3A8A","312E81","0C4A6E","164E63","0F4C81",
        "1E40AF","0369A1","3730A3","0E7490","1D4ED8","083344",
    ]
    _BS  = Side(style="thin",   color="CBD5E1")
    _BSH = Side(style="medium", color="64748B")
    _BSW = Side(style="medium", color="FFFFFF")
    B_R  = Border(left=_BS,  right=_BS,  top=_BS,  bottom=_BS)
    B_RH = Border(left=_BSH, right=_BSH, top=_BSH, bottom=_BSH)
    B_RW = Border(left=_BSW, right=_BSW, top=_BSW, bottom=_BSW)

    def _rsc(row, col, value=None, bold=False, color="FFFFFF", fill=None,
             align=None, nf=None, border=None, italic=False, size=None):
        c = ws_res.cell(row=row, column=col, value=value)
        c.font = Font(name=FONT_NAME, size=size or FONT_SIZE,
                      bold=bold, color=color, italic=italic)
        if fill is not None: c.fill = fill
        if align:  c.alignment = align
        if nf:     c.number_format = nf
        if border: c.border = border
        return c

    def _fill6(row, fill):
        for _c in range(1, 7):
            ws_res.cell(row=row, column=_c).fill = fill

    # ── Acumular ──────────────────────────────────────────────────────────────────
    _bank_d       = defaultdict(lambda: {"count": 0, "total": 0.0})
    _bank_abon_d  = defaultdict(lambda: defaultdict(lambda: {"count": 0, "total": 0.0}))
    _no_cls_d     = defaultdict(lambda: {"count": 0, "total": 0.0})
    _mes_bank_d   = defaultdict(lambda: defaultdict(lambda: {"count": 0, "total": 0.0}))
    _mes_abon_d   = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: {"count": 0, "total": 0.0})))
    _mes_no_cls_d = defaultdict(lambda: defaultdict(lambda: {"count": 0, "total": 0.0}))

    for _r in registros:
        _mk2 = (getattr(_r["fecha"],"year",0), getattr(_r["fecha"],"month",0))
        _bn  = _r["banco"]; _mn = _r["monto"]
        _bank_d[_bn]["count"] += 1;        _bank_d[_bn]["total"] += _mn
        _mes_bank_d[_mk2][_bn]["count"] += 1; _mes_bank_d[_mk2][_bn]["total"] += _mn
        _ac = semantic_to_actual.get(_r["col_abono"])
        if _ac is not None:
            _bank_abon_d[_bn][_ac]["count"] += 1; _bank_abon_d[_bn][_ac]["total"] += _mn
            _mes_abon_d[_mk2][_bn][_ac]["count"] += 1; _mes_abon_d[_mk2][_bn][_ac]["total"] += _mn
        else:
            _no_cls_d[_bn]["count"] += 1; _no_cls_d[_bn]["total"] += _mn
            _mes_no_cls_d[_mk2][_bn]["count"] += 1; _mes_no_cls_d[_mk2][_bn]["total"] += _mn

    _meses_sorted = sorted(_mes_bank_d.keys())
    _tot_carg     = sum(v["total"] for v in _bank_d.values())
    _tot_abon_all = sum(sum(ad["total"] for ad in ba.values()) for ba in _bank_abon_d.values())
    _dif_conc     = _tot_carg - _tot_abon_all

    MESES_R = {1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",5:"MAYO",6:"JUNIO",
               7:"JULIO",8:"AGOSTO",9:"SEPTIEMBRE",10:"OCTUBRE",11:"NOVIEMBRE",12:"DICIEMBRE"}

    # ── Título (sin fecha) ────────────────────────────────────────────────────────
    ws_res.merge_cells("A1:F1")
    _tc = ws_res.cell(row=1, column=1, value="RESUMEN DE DEPÓSITOS BANCARIOS")
    _tc.font = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
    _tc.fill = F_R_TITLE; _tc.alignment = A_CTR
    ws_res.row_dimensions[1].height = 32

    _row = 3

    # ══ Por mes → por banco → desglose de abonos ══════════════════════════════════
    for _mi, _mk2 in enumerate(_meses_sorted):
        _mfill   = _MES_PAL[_mi % len(_MES_PAL)]
        _mfill_p = PatternFill("solid", fgColor=_mfill)
        _mes_lbl = f"{MESES_R.get(_mk2[1], str(_mk2[1]))} {_mk2[0]}"
        _mes_tot = sum(_mes_bank_d[_mk2][_ci["banco"]]["total"] for _ci in cargos_list)
        _mes_cnt = sum(_mes_bank_d[_mk2][_ci["banco"]]["count"] for _ci in cargos_list)

        # ── Cabecera de mes ────────────────────────────────────────────────────────
        ws_res.merge_cells(f"A{_row}:F{_row}")
        _rsc(_row, 1, f"  📅  {_mes_lbl}", bold=True, color="FFFFFF",
             fill=_mfill_p, align=A_LEFT, border=B_RW, size=12)
        ws_res.row_dimensions[_row].height = 26; _row += 1

        # ── Encabezado de columnas ─────────────────────────────────────────────────
        for _cn2, _lb2 in [(1,"BANCO / TIPO DE ABONO"),(2,"N° CUENTA"),
                            (3,"# MOV"),(4,"TOTAL"),(5,"% BANCO"),(6,"% MES")]:
            _rsc(_row, _cn2, _lb2, bold=True, color="CBD5E1",
                 fill=F_R_HDR, align=A_CTR, border=B_RH, size=8)
        ws_res.row_dimensions[_row].height = 16; _row += 1

        # ── Bancos ────────────────────────────────────────────────────────────────
        for _ci in cargos_list:
            _bnc   = _ci["banco"]
            _bnk_f = F_R_BNK.get(_bnc, F_R_HDR)
            _bd2   = _mes_bank_d[_mk2][_bnc]
            if _bd2["total"] == 0:
                continue
            _pct_m = (_bd2["total"] / _mes_tot * 100) if _mes_tot else 0
            _pct_t = (_bd2["total"] / _tot_carg * 100) if _tot_carg else 0

            # Fila banco
            _rsc(_row,1,f"  {_bnc}",bold=True,color="FFFFFF",fill=_bnk_f,align=A_LEFT,border=B_RW)
            _rsc(_row,2,_ci["cuenta"],bold=False,color="CBD5E1",fill=_bnk_f,align=A_CTR,border=B_RW,size=9,italic=True)
            _rsc(_row,3,_bd2["count"],bold=True,color="FFFFFF",fill=_bnk_f,align=A_CTR,border=B_RW,nf="#,##0")
            _rsc(_row,4,_bd2["total"],bold=True,color="FFFFFF",fill=_bnk_f,align=A_RIGHT,border=B_RW,nf=FMT_NUM)
            _rsc(_row,5,_pct_m/100,bold=True,color="FFFFFF",fill=_bnk_f,align=A_CTR,border=B_RW,nf="0.0%")
            _rsc(_row,6,_pct_t/100,bold=False,color="CBD5E1",fill=_bnk_f,align=A_CTR,border=B_RW,nf="0.0%")
            ws_res.row_dimensions[_row].height = 18; _row += 1

            # Filas de abono detalle
            _abs2  = _mes_abon_d[_mk2].get(_bnc, {})
            _ab_i  = 0
            for _ab2 in abonos_efectivos:
                _ad = _abs2.get(_ab2["col"])
                if not _ad or _ad["total"] == 0:
                    continue
                _fl  = F_R_ODD if _ab_i % 2 == 0 else F_R_EVN
                _pb  = (_ad["total"] / _bd2["total"] * 100) if _bd2["total"] else 0
                _pt  = (_ad["total"] / _mes_tot      * 100) if _mes_tot      else 0
                _rsc(_row,1,f"    {_ab2['nombre']}",color="1E293B",fill=_fl,align=A_LEFT,border=B_R,size=10)
                _rsc(_row,2,_ab2["cuenta"],color="64748B",fill=_fl,align=A_CTR,border=B_R,size=9,italic=True)
                _rsc(_row,3,_ad["count"],color="1E293B",fill=_fl,align=A_CTR,border=B_R,nf="#,##0")
                _rsc(_row,4,_ad["total"],color="1E293B",fill=_fl,align=A_RIGHT,border=B_R,nf=FMT_NUM)
                _rsc(_row,5,_pb/100,color="1E293B",fill=_fl,align=A_CTR,border=B_R,nf="0.0%")
                _rsc(_row,6,_pt/100,color="475569",fill=_fl,align=A_CTR,border=B_R,nf="0.0%")
                ws_res.row_dimensions[_row].height = 16; _row += 1
                _ab_i += 1

            # Sin clasificar
            _nc2 = _mes_no_cls_d[_mk2].get(_bnc)
            if _nc2 and _nc2["total"] > 0:
                _fl  = F_R_ODD if _ab_i % 2 == 0 else F_R_EVN
                _pb  = (_nc2["total"] / _bd2["total"] * 100) if _bd2["total"] else 0
                _pt  = (_nc2["total"] / _mes_tot      * 100) if _mes_tot      else 0
                _rsc(_row,1,"    Sin clasificar",color="DC2626",fill=_fl,align=A_LEFT,border=B_R,italic=True)
                _rsc(_row,2,"—",color="DC2626",fill=_fl,align=A_CTR,border=B_R,size=9)
                _rsc(_row,3,_nc2["count"],color="DC2626",fill=_fl,align=A_CTR,border=B_R,nf="#,##0")
                _rsc(_row,4,_nc2["total"],color="DC2626",fill=_fl,align=A_RIGHT,border=B_R,nf=FMT_NUM)
                _rsc(_row,5,_pb/100,color="DC2626",fill=_fl,align=A_CTR,border=B_R,nf="0.0%")
                _rsc(_row,6,_pt/100,color="DC2626",fill=_fl,align=A_CTR,border=B_R,nf="0.0%")
                ws_res.row_dimensions[_row].height = 16; _row += 1

        # ── Total del mes ──────────────────────────────────────────────────────────
        _pgt = (_mes_tot / _tot_carg * 100) if _tot_carg else 0
        _rsc(_row,1,f"TOTAL  {_mes_lbl}",bold=True,color="FFFFFF",fill=_mfill_p,align=A_LEFT,border=B_RW,size=11)
        _rsc(_row,2,"",fill=_mfill_p,border=B_RW)
        _rsc(_row,3,_mes_cnt,bold=True,color="FFFFFF",fill=_mfill_p,align=A_CTR,border=B_RW,nf="#,##0")
        _rsc(_row,4,_mes_tot,bold=True,color="FFFFFF",fill=_mfill_p,align=A_RIGHT,border=B_RW,nf=FMT_NUM)
        _rsc(_row,5,1.0,bold=True,color="FFFFFF",fill=_mfill_p,align=A_CTR,border=B_RW,nf="0.0%")
        _rsc(_row,6,_pgt/100,bold=True,color="FFFFFF",fill=_mfill_p,align=A_CTR,border=B_RW,nf="0.0%")
        ws_res.row_dimensions[_row].height = 22; _row += 2

    # ── Total global ───────────────────────────────────────────────────────────────
    _grand_cnt = sum(v["count"] for v in _bank_d.values())
    _rsc(_row,1,"TOTAL GENERAL",bold=True,color="FFFFFF",fill=F_R_TOT_G,align=A_LEFT,border=B_RW,size=12)
    _rsc(_row,2,"",fill=F_R_TOT_G,border=B_RW)
    _rsc(_row,3,_grand_cnt,bold=True,color="FFFFFF",fill=F_R_TOT_G,align=A_CTR,border=B_RW,nf="#,##0")
    _rsc(_row,4,_tot_carg,bold=True,color="FFFFFF",fill=F_R_TOT_G,align=A_RIGHT,border=B_RW,nf=FMT_NUM)
    _rsc(_row,5,1.0,bold=True,color="FFFFFF",fill=F_R_TOT_G,align=A_CTR,border=B_RW,nf="0.0%")
    _rsc(_row,6,"",fill=F_R_TOT_G,border=B_RW)
    ws_res.row_dimensions[_row].height = 24; _row += 2

    # ══ Conciliación ══════════════════════════════════════════════════════════════
    ws_res.merge_cells(f"A{_row}:F{_row}")
    _rsc(_row, 1, "CONCILIACIÓN", bold=True, color="FFFFFF",
         fill=PatternFill("solid", fgColor="1E3A8A"),
         align=A_LEFT, border=B_RW, size=11)
    ws_res.row_dimensions[_row].height = 24; _row += 1

    for _lbl_c, _val_c in [
        ("Total Cargos (suma de todos los meses)", _tot_carg),
        ("Total Abonos clasificados",              _tot_abon_all),
    ]:
        _rsc(_row,1,_lbl_c,color="1E293B",fill=F_R_ODD,align=A_LEFT,border=B_R)
        for _c in range(2,6): _rsc(_row,_c,"",fill=F_R_ODD,border=B_R)
        _rsc(_row,4,_val_c,bold=True,color="1E293B",fill=F_R_ODD,align=A_RIGHT,border=B_R,nf=FMT_NUM)
        _rsc(_row,6,"",fill=F_R_ODD,border=B_R); _row += 1

    _diff_fill  = F_R_DIFF_G if abs(_dif_conc) < 0.01 else F_R_DIFF_R
    _diff_label = "CONCILIADO" if abs(_dif_conc) < 0.01 else "DIFERENCIA"
    _rsc(_row,1,"DIFERENCIA",bold=True,color="FFFFFF",fill=_diff_fill,align=A_LEFT,border=B_RW,size=12)
    _rsc(_row,2,"",fill=_diff_fill,border=B_RW)
    _rsc(_row,3,"",fill=_diff_fill,border=B_RW)
    _rsc(_row,4,_dif_conc,bold=True,color="FFFFFF",fill=_diff_fill,align=A_RIGHT,border=B_RW,nf=FMT_NUM,size=12)
    _rsc(_row,5,_diff_label,bold=True,color="FFFFFF",fill=_diff_fill,align=A_CTR,border=B_RW,size=10)
    _rsc(_row,6,"",fill=_diff_fill,border=B_RW)
    ws_res.row_dimensions[_row].height = 26

    for _cn2,_w2 in {1:40.0,2:22.0,3:9.0,4:18.0,5:10.0,6:10.0}.items():
        ws_res.column_dimensions[get_column_letter(_cn2)].width = _w2
    ws_res.row_dimensions[1].height = 32
    ws_res.freeze_panes = "A3"

    # ── Hoja CUENTAS (solo cuando no hay plantilla) ───────────────────────────────
    if plantilla is None:
        wc = wb.create_sheet("CUENTAS")
        fn_th  = fnt(bold=True, color="FFFFFF")
        fn_row = fnt(color="000000")
        F_TH_C = PatternFill("solid", fgColor="4B5563")
        F_TH_A = PatternFill("solid", fgColor="15803D")
        F_ROW1 = PatternFill("solid", fgColor="EFF6FF")
        F_ROW2 = PatternFill("solid", fgColor="F0FDF4")
        A_L  = Alignment(horizontal="left",   vertical="center")
        A_C2 = Alignment(horizontal="center", vertical="center")
        set_cell(wc,1,1,"CARGOS",font=fn_th,fill=F_TH_C,align=A_C2,border=BORDER_H)
        set_cell(wc,1,2,"",      font=fn_th,fill=F_TH_C,align=A_C2,border=BORDER_H)
        set_cell(wc,1,4,"ABONOS",font=fn_th,fill=F_TH_A,align=A_C2,border=BORDER_H)
        set_cell(wc,1,5,"",      font=fn_th,fill=F_TH_A,align=A_C2,border=BORDER_H)
        for _cn2,_lb2,_fl2 in [(1,"N° Cuenta",F_TH_C),(2,"Banco",F_TH_C),
                                 (4,"N° Cuenta",F_TH_A),(5,"Nombre",F_TH_A)]:
            set_cell(wc,2,_cn2,_lb2,font=fn_th,fill=_fl2,align=A_C2,border=BORDER_H)
        for _i2,_ci2 in enumerate(cargos_list,start=3):
            _fl2 = F_ROW1 if _i2%2==0 else F_ROW2
            set_cell(wc,_i2,1,_ci2["cuenta"],font=fn_row,fill=_fl2,align=A_L,border=BORDER)
            set_cell(wc,_i2,2,_ci2["banco"], font=fn_row,fill=_fl2,align=A_L,border=BORDER)
        for _i2,_ab2 in enumerate(abonos_efectivos,start=3):
            _fl2 = F_ROW1 if _i2%2==0 else F_ROW2
            set_cell(wc,_i2,4,_ab2["cuenta"],font=fn_row,fill=_fl2,align=A_L,border=BORDER)
            set_cell(wc,_i2,5,_ab2["nombre"],font=fn_row,fill=_fl2,align=A_L,border=BORDER)
        for _cn2,_w2 in {1:20.0,2:12.0,4:20.0,5:48.0}.items():
            wc.column_dimensions[get_column_letter(_cn2)].width = _w2
        wc.row_dimensions[1].height = 22; wc.row_dimensions[2].height = 22

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
# ─── UI ─────────────────────────────────────────────────────────────────────
col1, col2, col3 = st.columns(3)

with col1:
    st.markdown('<div class="upload-box">', unsafe_allow_html=True)
    st.markdown("**🟦 BBVA Bancomer** *(uno o varios meses)*")
    files_bbva = st.file_uploader("Estado de cuenta BBVA (.xlsx)",
                                   type=["xlsx"], key="bbva",
                                   accept_multiple_files=True,
                                   label_visibility="collapsed")
    if files_bbva:
        st.caption(f"📎 {len(files_bbva)} archivo(s) cargado(s)")
    st.markdown('</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="upload-box">', unsafe_allow_html=True)
    st.markdown("**🟥 Banorte** *(uno o varios meses)*")
    files_banorte = st.file_uploader("Estado de cuenta Banorte (.xlsx)",
                                      type=["xlsx"], key="banorte",
                                      accept_multiple_files=True,
                                      label_visibility="collapsed")
    if files_banorte:
        st.caption(f"📎 {len(files_banorte)} archivo(s) cargado(s)")
    st.markdown('</div>', unsafe_allow_html=True)

with col3:
    st.markdown('<div class="upload-box">', unsafe_allow_html=True)
    st.markdown("**🟧 Inbursa** *(uno o varios meses)*")
    files_inbursa = st.file_uploader("Estado de cuenta Inbursa (.xlsx)",
                                      type=["xlsx"], key="inbursa",
                                      accept_multiple_files=True,
                                      label_visibility="collapsed")
    if files_inbursa:
        st.caption(f"📎 {len(files_inbursa)} archivo(s) cargado(s)")
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div class="upload-box">', unsafe_allow_html=True)
st.markdown("**📋 PLANTILLA DE DEPOSITOS**  *(opcional — si se sube, los datos se escriben sobre ella)*")
file_plantilla = st.file_uploader(
    "Plantilla de Depósitos (.xlsx)",
    type=["xlsx", "xlsm"], key="plantilla",
    label_visibility="collapsed",
)
st.markdown('</div>', unsafe_allow_html=True)

st.divider()

if "dep_result" not in st.session_state:
    st.session_state.dep_result = None

_any_files = any([files_bbva, files_banorte, files_inbursa])

if st.button("⚙️ Generar Póliza", type="primary", disabled=not _any_files):

    todos_ok, todos_nc = [], []
    archivos_bytes = {}   # banco -> [raw_bytes, ...]

    banco_files = [
        (files_bbva,    "BBVA",    "🟦"),
        (files_banorte, "BANORTE", "🟥"),
        (files_inbursa, "INBURSA", "🟧"),
    ]

    total_files = sum(len(fl or []) for fl, _, _ in banco_files)
    progress = st.progress(0)
    done = 0

    for fl, banco, ico in banco_files:
        if not fl:
            continue
        archivos_bytes[banco] = []
        for fi, f in enumerate(fl):
            raw = f.read()
            archivos_bytes[banco].append(raw)
            with st.spinner(f"{ico} Leyendo {banco} — archivo {fi+1}/{len(fl)}..."):
                try:
                    ok, nc = leer_banco(BytesIO(raw), banco)
                    for r in ok:
                        r["archivo_idx"] = fi  # para marcar el archivo correcto
                    todos_ok.extend(ok)
                    todos_nc.extend(nc)
                    st.success(f"{ico} {banco} [{fi+1}]: **{len(ok)}** depósitos"
                               + (f", {len(nc)} sin clasificar" if nc else ""))
                except Exception as e:
                    st.error(f"Error leyendo {banco} [{fi+1}]: {e}")
            done += 1
            progress.progress(done / max(total_files, 1))

    if todos_ok:
        excel_bytes = generar_excel(todos_ok, plantilla=file_plantilla)

        # Nombre con rango de fechas
        _fechas = [r["fecha"] for r in todos_ok if hasattr(r["fecha"], "month")]
        if _fechas:
            _fmin, _fmax = min(_fechas), max(_fechas)
            if _fmin.year == _fmax.year and _fmin.month == _fmax.month:
                _rango = _fmin.strftime("%Y-%m")
            else:
                _rango = f"{_fmin.strftime('%Y-%m')} al {_fmax.strftime('%Y-%m')}"
        else:
            _rango = datetime.now().strftime("%Y-%m")
        nombre_archivo = f"DEPOSITOS BANCARIOS {_rango}.xlsx"

        # Marcar cada archivo fuente por separado
        marked = {}   # banco -> [{ico, bytes, n, nombre}]
        for fl, banco, ico in banco_files:
            if banco not in archivos_bytes:
                continue
            marked[banco] = []
            for fi, raw in enumerate(archivos_bytes[banco]):
                filas = {r["fila_excel"] for r in todos_ok
                         if r["banco"] == banco and r.get("archivo_idx") == fi}
                if filas:
                    _fn = fl[fi].name if fl and fi < len(fl) else f"{banco}_{fi+1}.xlsx"
                    marked[banco].append({
                        "ico":    ico,
                        "bytes":  marcar_estado_cuenta(raw, filas, banco),
                        "n":      len(filas),
                        "nombre": _fn,
                    })

        st.session_state.dep_result = {
            "todos_ok":       todos_ok,
            "todos_nc":       todos_nc,
            "excel_bytes":    excel_bytes,
            "nombre_archivo": nombre_archivo,
            "marked":         marked,
        }
    else:
        st.session_state.dep_result = None

if st.session_state.dep_result:
    res      = st.session_state.dep_result
    todos_ok = res["todos_ok"]
    todos_nc = res["todos_nc"]

    st.subheader("📊 Resumen")
    df_ok = pd.DataFrame(todos_ok)
    total_general = df_ok["monto"].sum()

    # Métricas por banco
    sc = st.columns(4)
    for ci, (banco, ico) in enumerate([("BBVA","🟦"),("BANORTE","🟥"),("INBURSA","🟧")]):
        sub = df_ok[df_ok["banco"] == banco]
        sc[ci].metric(f"{ico} {banco}",
                      f"${sub['monto'].sum():,.2f}",
                      f"{len(sub)} mov.")
    sc[3].metric("💰 TOTAL", f"${total_general:,.2f}", f"{len(df_ok)} mov.")

    # Métricas por mes (si hay más de uno)
    _meses_ok = sorted(df_ok["fecha"].apply(
        lambda x: x.strftime("%Y-%m") if hasattr(x,"strftime") else "").unique())
    if len(_meses_ok) > 1:
        st.caption(f"Meses procesados: {', '.join(_meses_ok)}")
        _mes_cols = st.columns(min(len(_meses_ok), 6))
        for _mi, _mes in enumerate(_meses_ok):
            _sub_m = df_ok[df_ok["fecha"].apply(
                lambda x: x.strftime("%Y-%m") if hasattr(x,"strftime") else "") == _mes]
            _mes_cols[_mi % 6].metric(_mes, f"${_sub_m['monto'].sum():,.2f}",
                                       f"{len(_sub_m)} mov.")

    st.subheader("📋 Vista previa")
    abono_nombre = {a["col"]: a["nombre"][:30] for a in ABONOS}
    cargo_nombre = {9:"BANORTE",10:"BBVA",11:"INBURSA"}
    _ord_banco   = {"BBVA":0,"BANORTE":1,"INBURSA":2}

    preview_rows = []
    for r in sorted(todos_ok, key=lambda x:(
                        getattr(x["fecha"],"year",0),
                        getattr(x["fecha"],"month",0),
                        _ord_banco.get(x["banco"],9),
                        x["fecha"])):
        preview_rows.append({
            "Mes":         r["fecha"].strftime("%Y-%m") if hasattr(r["fecha"],"strftime") else "",
            "Fecha":       r["fecha"].strftime("%d/%m/%Y") if hasattr(r["fecha"],"strftime") else str(r["fecha"]),
            "Banco":       r["banco"],
            "Abono":       abono_nombre.get(r["col_abono"], str(r["col_abono"])),
            "Monto":       r["monto"],
            "Descripción": r["desc"][:55],
        })

    st.dataframe(pd.DataFrame(preview_rows).style.format({"Monto":"${:,.2f}"}),
                 use_container_width=True, height=400)

    if todos_nc:
        st.subheader(f"⚠️ Sin clasificar ({len(todos_nc)} movimientos — no incluidos)")
        df_nc = pd.DataFrame(todos_nc)
        df_nc["fecha"] = df_nc["fecha"].apply(
            lambda x: x.strftime("%d/%m/%Y") if hasattr(x,"strftime") else str(x))
        st.dataframe(df_nc.style.format({"monto":"${:,.2f}"}), use_container_width=True)

    st.subheader("💾 Descargar")
    st.download_button(
        label="📥 Descargar Póliza Excel",
        data=res["excel_bytes"],
        file_name=res["nombre_archivo"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )

    if res["marked"]:
        st.subheader("🗙️ Estados de cuenta marcados")
        st.caption("Las filas incluidas en la póliza aparecen resaltadas en el color del banco.")
        for banco, info_list in res["marked"].items():
            for _idx2, info in enumerate(info_list):
                _lbl = f"{info['ico']} {info['nombre']} — {info['n']} mov. marcados"
                st.download_button(
                    label=_lbl,
                    data=info["bytes"],
                    file_name=f"MARCADO_{banco}_{_idx2+1}_{datetime.now().strftime('%Y-%m')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_marcado_{banco}_{_idx2}",
                )

else:
    st.info("⬆️ Sube al menos un archivo de depósito para continuar.")
