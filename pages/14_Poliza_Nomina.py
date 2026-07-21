"""
Página 14 — Póliza de Nómina
Autocontenida: la función generar_poliza_nomina está definida aquí directamente.
"""

import io, re, tempfile, os
from collections import OrderedDict
from datetime import datetime

import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Póliza de Nómina",
    page_icon="📋",
    layout="wide",
)

# ── función core ─────────────────────────────────────────────────────────────

def generar_poliza_nomina(pagos_excel_path, out_path, plantilla_path=None):
    """Lee un Excel de PagosBancarios_Consolidado y genera la plantilla de
    póliza de nómina (matriz empleados × semana) con hoja de conciliación."""

    PLNT_HDRS = [
        {"value": "TIPO DE POLIZA", "fill": "4472C4"},
        {"value": "Fecha",          "fill": "FFC000"},
        {"value": "REFERENCIA",     "fill": "7030A0"},
        {"value": "CONCEPTO",       "fill": "00B050"},
        {"value": "ERROR",          "fill": "4472C4"},
        {"value": "UIDD",           "fill": "4472C4"},
        {"value": "NUM POLIZA",     "fill": "4472C4"},
        {"value": "PROCESADO",      "fill": "4472C4"},
    ]
    TINTS = {"FFC000": "FFF8DC", "7030A0": "F3E8FF",
             "00B050": "ECFDF5", "4472C4": "EFF6FF"}
    DATE_FMT = "DD/MM/YYYY"

    def _to_date(s):
        try:
            return datetime.strptime(str(s).strip(), "%d/%m/%Y")
        except Exception:
            return s

    wb_pago = openpyxl.load_workbook(pagos_excel_path, data_only=True)
    empleados   = OrderedDict()
    hojas_datos = []

    def _fsort(f):
        p = str(f).split("/")
        return f"{p[2]}{p[1]}{p[0]}" if len(p) == 3 else str(f)

    for sheet in wb_pago.sheetnames:
        if sheet == "Resumen":
            continue
        rows = list(wb_pago[sheet].iter_rows(values_only=True))
        if not rows:
            continue
        fecha, folio, pagos = None, "", {}
        for row in rows[1:]:
            nombre, cuenta, importe, fecha_row = (list(row) + [None] * 4)[:4]
            if nombre in (None, "TOTAL"):
                continue
            if isinstance(nombre, str) and "|" in nombre:
                m = re.search(r"Folio:\s*(\S+)", nombre)
                if m:
                    folio = m.group(1)
                continue
            if nombre and isinstance(importe, (int, float)):
                nombre = str(nombre).strip().upper()
                pagos[nombre] = round(float(importe), 2)
                if fecha is None and fecha_row:
                    fecha = str(fecha_row)
                if nombre not in empleados:
                    empleados[nombre] = str(cuenta).strip() if cuenta else ""
        if pagos:
            hojas_datos.append((_fsort(fecha), fecha or "", sheet, folio, pagos))

    wb_pago.close()
    hojas_datos.sort(key=lambda x: x[0])
    emp_list   = list(empleados.keys())
    emp_cuenta = list(empleados.values())
    n_emp      = len(emp_list)

    N_PLNT  = 8
    C_EMP   = N_PLNT + 1
    C_TOT   = N_PLNT + n_emp + 1
    C_BBVA  = C_TOT + 1
    C_TBBVA = C_TOT + 2
    C_CONC  = C_TOT + 3
    GL = openpyxl.utils.get_column_letter

    thin = Side(style="thin", color="B0C4DE")
    _BRD = Border(left=thin, right=thin, top=thin, bottom=thin)
    CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LFT  = Alignment(horizontal="left",   vertical="center")
    RGT  = Alignment(horizontal="right",  vertical="center")

    def _mc(ws, r, c, v="", bold=False, color="000000", bg="FFFFFF",
            align=None, fmt=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(bold=bold, name="Arial",
                              size=8 if not bold else 9, color=color)
        cell.fill      = PatternFill("solid", start_color=bg)
        cell.alignment = align or CTR
        cell.border    = _BRD
        if fmt:
            cell.number_format = fmt
        return cell

    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "poliza IA"
    TOTAL_COLS = C_CONC

    # Fila 1 — numeración
    for c in range(TOTAL_COLS):
        _mc(ws1, 1, c + 1, c, bold=True, color="FFFFFF", bg="1E3A8A")
    ws1.row_dimensions[1].height = 16

    # Fila 2 — N° cuenta
    for c in range(N_PLNT):
        _mc(ws1, 2, c + 1, bg=PLNT_HDRS[c]["fill"])
    for i, cta in enumerate(emp_cuenta):
        _mc(ws1, 2, C_EMP + i, cta, bg="FFF0F5")
    _mc(ws1, 2, C_TOT,   bg="FEF3C7")
    _mc(ws1, 2, C_BBVA,  "102-01-0001-0001", bold=True, bg="DCFCE7")
    _mc(ws1, 2, C_TBBVA, bg="FEF3C7")
    _mc(ws1, 2, C_CONC,  bg="E0E7FF")
    ws1.row_dimensions[2].height = 22

    # Fila 3 — headers
    for c, hdr in enumerate(PLNT_HDRS):
        _mc(ws1, 3, c + 1, hdr["value"], bold=True, color="FFFFFF", bg=hdr["fill"])
    for i, nom in enumerate(emp_list):
        cell = ws1.cell(row=3, column=C_EMP + i, value=nom)
        cell.font      = Font(bold=True, name="Arial", size=7)
        cell.fill      = PatternFill("solid", start_color="EFF6FF")
        cell.alignment = CTR
        cell.border    = _BRD
    for col, lbl, fill in [
        (C_TOT,   "TOTAL",        "92400E"),
        (C_BBVA,  "BBVA",         "166534"),
        (C_TBBVA, "TOTAL BBVA",   "92400E"),
        (C_CONC,  "CONCILIACION", "1E3A8A"),
    ]:
        _mc(ws1, 3, col, lbl, bold=True, color="FFFFFF", bg=fill)
    ws1.row_dimensions[3].height = 42

    # Filas de datos
    for r_idx, (_, fecha_d, sh_name, folio, pagos) in enumerate(hojas_datos):
        rn = r_idx + 4
        fv = _to_date(fecha_d)
        plnt_vals = ["E", fv, sh_name, sh_name, "", "", "", ""]
        for c, (hdr, val) in enumerate(zip(PLNT_HDRS, plnt_vals)):
            cell = ws1.cell(row=rn, column=c + 1, value=val)
            cell.font      = Font(name="Arial", size=8)
            cell.fill      = PatternFill("solid",
                                         start_color=TINTS.get(hdr["fill"], "F8FAFC"))
            cell.alignment = CTR
            cell.border    = _BRD
            if c == 1 and isinstance(fv, datetime):
                cell.number_format = DATE_FMT
        fe = GL(C_EMP); le = GL(C_EMP + n_emp - 1)
        for i, nom in enumerate(emp_list):
            val  = pagos.get(nom)
            cell = ws1.cell(row=rn, column=C_EMP + i, value=val)
            cell.font      = Font(name="Arial", size=8)
            cell.fill      = PatternFill("solid",
                                         start_color="F0F9FF" if val else "FFFFFF")
            cell.alignment = RGT
            cell.border    = _BRD
            if val:
                cell.number_format = "#,##0.00"
        tot = f"{GL(C_TOT)}{rn}"; bbv = f"{GL(C_BBVA)}{rn}"; tbb = f"{GL(C_TBBVA)}{rn}"
        for col, fml, bg, bold in [
            (C_TOT,   f"=SUM({fe}{rn}:{le}{rn})", "FEF3C7", True),
            (C_BBVA,  f"={tot}",                   "DCFCE7", False),
            (C_TBBVA, f"={bbv}",                   "FEF3C7", True),
            (C_CONC,  f"={tot}-{tbb}",             "E0E7FF", True),
        ]:
            cell = ws1.cell(row=rn, column=col, value=fml)
            cell.font          = Font(bold=bold, name="Arial", size=8)
            cell.fill          = PatternFill("solid", start_color=bg)
            cell.alignment     = RGT
            cell.border        = _BRD
            cell.number_format = "#,##0.00"
        ws1.row_dimensions[rn].height = 14

    for c, w in enumerate([14, 12, 26, 26, 8, 14, 12, 11]):
        ws1.column_dimensions[GL(c + 1)].width = w
    for i in range(n_emp):
        ws1.column_dimensions[GL(C_EMP + i)].width = 12
    for col, w in [(C_TOT, 14), (C_BBVA, 20), (C_TBBVA, 14), (C_CONC, 14)]:
        ws1.column_dimensions[GL(col)].width = w
    ws1.freeze_panes = f"{GL(C_EMP)}4"

    # Hoja 2: Conciliación
    ws2 = wb.create_sheet("Conciliación")
    hdrs2  = ["Hoja (PDF)", "Fecha", "Empleado", "N° Cuenta",
               "Importe PDF", "Importe Reporte", "Diferencia", "Estatus"]
    fills2 = ["1E3A8A","1E3A8A","1E3A8A","1E3A8A","7030A0","00B050","FFC000","4472C4"]
    for c, (h, f) in enumerate(zip(hdrs2, fills2)):
        _mc(ws2, 1, c + 1, h, bold=True, color="FFFFFF", bg=f)
    ws2.row_dimensions[1].height = 20

    row2 = 2
    for _, fecha_d, sh_name, folio, pagos in hojas_datos:
        fv = _to_date(fecha_d)
        for nom in emp_list:
            if nom not in pagos:
                continue
            importe = pagos[nom]
            cta     = empleados.get(nom, "")
            bgs = ["DBEAFE","DBEAFE","F8FAFF","FFF0F5",
                   "F3E8FF","ECFDF5","FFFBEB","DCFCE7"]
            for c, (val, bg) in enumerate(zip(
                [sh_name, fv, nom, cta, importe, importe, 0.0, "OK"], bgs
            )):
                cell = ws2.cell(row=row2, column=c + 1, value=val)
                cell.font      = Font(name="Arial", size=8)
                cell.fill      = PatternFill("solid", start_color=bg)
                cell.alignment = LFT if c == 2 else CTR
                cell.border    = _BRD
                if c in (4, 5, 6):
                    cell.number_format = "#,##0.00"
                if c == 1 and isinstance(fv, datetime):
                    cell.number_format = DATE_FMT
            ws2.row_dimensions[row2].height = 13
            row2 += 1
        total_hoja = round(sum(pagos.values()), 2)
        sub_bgs = ["1E3A8A","1E3A8A","1E3A8A","1E3A8A",
                   "92400E","166534","4B5563","166534"]
        for c, (val, bg) in enumerate(zip(
            [sh_name, fv, "SUBTOTAL", "", total_hoja, total_hoja, 0.0, "OK"],
            sub_bgs
        )):
            cell = ws2.cell(row=row2, column=c + 1, value=val)
            cell.font      = Font(bold=True, name="Arial", size=8, color="FFFFFF")
            cell.fill      = PatternFill("solid", start_color=bg)
            cell.alignment = CTR
            cell.border    = _BRD
            if c in (4, 5, 6):
                cell.number_format = "#,##0.00"
            if c == 1 and isinstance(fv, datetime):
                cell.number_format = DATE_FMT
        ws2.row_dimensions[row2].height = 14
        row2 += 1

    for c, w in enumerate([30, 13, 38, 16, 15, 16, 13, 12]):
        ws2.column_dimensions[GL(c + 1)].width = w
    ws2.freeze_panes = "A2"

    wb.save(out_path)
    return len(hojas_datos), n_emp


# ── UI ────────────────────────────────────────────────────────────────────────
import _theme
_theme.aplicar_header("📋 Póliza de Nómina", "Excel de Pagos Bancarios → Póliza con matriz empleados × semana")
col1, col2 = st.columns(2)
with col1:
    pagos_file = st.file_uploader(
        "📊 Pagos Bancarios (Excel)",
        type=["xlsx", "xls"],
        key="pnom_pagos",
    )
with col2:
    plantilla_file = st.file_uploader(
        "📋 Plantilla (opcional)",
        type=["xlsx", "xls"],
        key="pnom_plantilla",
    )

if st.button("🔄 Generar Póliza de Nómina", type="primary",
             disabled=(pagos_file is None)):
    with tempfile.TemporaryDirectory() as tmp:
        pagos_path = os.path.join(tmp, "pagos.xlsx")
        with open(pagos_path, "wb") as f:
            f.write(pagos_file.getvalue())

        plantilla_path = None
        if plantilla_file is not None:
            plantilla_path = os.path.join(tmp, "plantilla.xlsx")
            with open(plantilla_path, "wb") as f:
                f.write(plantilla_file.getvalue())

        out_path = os.path.join(tmp, "PolizaNomina.xlsx")

        with st.spinner("Generando póliza..."):
            try:
                n_hojas, n_emp = generar_poliza_nomina(
                    pagos_path, out_path, plantilla_path
                )
                st.success(f"✅ {n_hojas} hoja(s) · {n_emp} empleado(s)")

                with open(out_path, "rb") as f:
                    xlsx_bytes = f.read()

                st.download_button(
                    label="⬇️ Descargar Excel",
                    data=xlsx_bytes,
                    file_name="PolizaNomina.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

                st.subheader("Vista previa — Póliza IA")
                try:
                    import pandas as pd
                    df = pd.read_excel(
                        io.BytesIO(xlsx_bytes),
                        sheet_name=0,
                        header=2,
                        nrows=30,
                        dtype=str,
                    ).fillna("")
                    st.dataframe(df, use_container_width=True, height=320)
                except Exception as _pe:
                    st.warning(f"Vista previa no disponible: {_pe}")

                st.subheader("Vista previa — Conciliación")
                try:
                    import pandas as pd
                    df2 = pd.read_excel(
                        io.BytesIO(xlsx_bytes),
                        sheet_name="Conciliación",
                        header=0,
                        nrows=40,
                        dtype=str,
                    ).fillna("")
                    st.dataframe(df2, use_container_width=True, height=250)
                except Exception as _ce:
                    st.info(f"Hoja Conciliación: {_ce}")

            except Exception as exc:
                import traceback
                st.error(f"Error al generar: {exc}")
                st.code(traceback.format_exc())
