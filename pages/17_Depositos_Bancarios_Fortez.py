"""
Módulo: Depósitos Bancarios FORTEZ
Lee plantilla (hoja CUENTAS) + estado de cuenta → genera póliza Excel.
Las cuentas se leen dinámicamente: si se agregan, eliminan o cambian
en la plantilla, el reporte se adapta automáticamente.
"""

import streamlit as st
from io import BytesIO
import tempfile, os

st.set_page_config(
    page_title="Depósitos Bancarios FORTEZ",
    page_icon="🏦",
    layout="wide",
)

import _theme
_theme.aplicar_header(
    "🏦 Depósitos Bancarios FORTEZ",
    "Plantilla con hoja CUENTAS + Estado de cuenta → póliza con cargos y abonos clasificados",
)

# ── Importar motor ────────────────────────────────────────────────────────────
try:
    from fortez_depositos import generar_poliza_fortez, _leer_cuentas
    _motor_ok = True
except Exception as _e:
    _motor_ok = False
    st.error(f"No se pudo cargar el módulo de generación: {_e}")
    st.stop()

# ── Carga de archivos ─────────────────────────────────────────────────────────
col_a, col_b = st.columns(2)

with col_a:
    st.markdown("#### 📋 Plantilla Excel")
    st.caption("Debe contener la hoja **CUENTAS** con cuentas de cargo (cols A-B) y abono (cols D-E).")
    f_plantilla = st.file_uploader(
        "Selecciona la plantilla", type=["xlsx"],
        key="fortez_plantilla", label_visibility="collapsed",
    )

with col_b:
    st.markdown("#### 📈 Estado de Cuenta")
    st.caption("Excel generado por el módulo **Estado de Cuenta** (hoja *Movimientos*).")
    f_estado = st.file_uploader(
        "Selecciona el estado de cuenta", type=["xlsx"],
        key="fortez_ec", label_visibility="collapsed",
    )

# ── Leer bytes UNA sola vez al inicio (evita problemas con seek en Cloud) ─────
bytes_plantilla = f_plantilla.read() if f_plantilla else None
bytes_estado    = f_estado.read()    if f_estado    else None

# ── Selector de banco ─────────────────────────────────────────────────────────
banco_sel = st.selectbox(
    "Banco",
    ["BANORTE", "BBVA", "INBURSA", "BANAMEX", "HSBC", "SANTANDER"],
    index=0,
    key="fortez_banco",
)

# ── Vista previa de cuentas ───────────────────────────────────────────────────
if bytes_plantilla:
    try:
        import openpyxl
        wb_prev = openpyxl.load_workbook(BytesIO(bytes_plantilla))
        if "CUENTAS" not in wb_prev.sheetnames:
            st.warning(
                f"⚠️ La plantilla no tiene hoja **CUENTAS** "
                f"(hojas encontradas: {', '.join(wb_prev.sheetnames)})."
            )
        else:
            cargos_prev, abonos_prev = _leer_cuentas(wb_prev["CUENTAS"])
            with st.expander("👁 Cuentas detectadas en la plantilla", expanded=True):
                col_c, col_ab = st.columns(2)
                with col_c:
                    st.markdown("**CARGOS**")
                    for cg in cargos_prev:
                        st.markdown(f"- `{cg['num']}` — {cg['banco']}")
                with col_ab:
                    st.markdown("**ABONOS**")
                    for ab in abonos_prev:
                        st.markdown(f"- `{ab['num']}` — {ab['nombre']}")
    except Exception as ex:
        st.caption(f"No se pudo previsualizar CUENTAS: {ex}")

# ── Botón generar ─────────────────────────────────────────────────────────────
st.divider()
generar = st.button(
    "⚙️ Generar Póliza",
    disabled=(bytes_plantilla is None or bytes_estado is None),
    type="primary",
    use_container_width=True,
)

if generar:
    # Validar hoja CUENTAS antes de proceder
    try:
        import openpyxl as _ox
        _wb_check = _ox.load_workbook(BytesIO(bytes_plantilla))
        if "CUENTAS" not in _wb_check.sheetnames:
            st.error(
                f"La plantilla no tiene hoja **CUENTAS**. "
                f"Hojas disponibles: {', '.join(_wb_check.sheetnames)}"
            )
            st.stop()
    except Exception as ex:
        st.error(f"No se pudo leer la plantilla: {ex}")
        st.stop()

    with st.spinner("Generando póliza…"):
        try:
            # Escribir bytes a archivos temporales (fortez_depositos usa rutas de archivo)
            tmp_p   = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp_p.write(bytes_plantilla); tmp_p.close()

            tmp_e   = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp_e.write(bytes_estado); tmp_e.close()

            tmp_out = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp_out.close()

            n_total, n_sin = generar_poliza_fortez(
                ruta_plantilla     = tmp_p.name,
                ruta_estado_cuenta = tmp_e.name,
                ruta_salida        = tmp_out.name,
                banco_override     = banco_sel,
            )

            with open(tmp_out.name, "rb") as fh:
                datos_excel = fh.read()

            for f in [tmp_p.name, tmp_e.name, tmp_out.name]:
                try: os.unlink(f)
                except: pass

            # ── Resultado ─────────────────────────────────────────────────────
            n_cls = n_total - n_sin
            st.success(
                f"✅ **{n_cls:,}** movimientos clasificados  |  "
                f"**{n_sin:,}** sin clasificar  |  "
                f"**{n_total:,}** total depósitos"
            )

            # Vista previa RESUMEN
            import pandas as pd
            wb_res = _ox.load_workbook(BytesIO(datos_excel))
            if "RESUMEN" in wb_res.sheetnames:
                ws_r = wb_res["RESUMEN"]
                data_res = []
                for r in range(2, ws_r.max_row + 1):
                    num = ws_r.cell(r, 1).value
                    nom = ws_r.cell(r, 2).value
                    cnt = ws_r.cell(r, 3).value
                    imp = ws_r.cell(r, 4).value
                    if nom:
                        data_res.append({
                            "N° Cuenta":   num or "",
                            "Nombre":      nom,
                            "Movimientos": cnt or "",
                            "Importe":     f"${imp:,.2f}" if isinstance(imp, (int, float)) else "",
                        })
                if data_res:
                    st.markdown("##### Resumen por cuenta")
                    st.dataframe(
                        pd.DataFrame(data_res),
                        use_container_width=True,
                        hide_index=True,
                    )

            # ── Descarga ──────────────────────────────────────────────────────
            nombre_ec  = (f_estado.name if f_estado else "ec").replace(".xlsx", "")
            nombre_sal = f"DEPOSITOS_FORTEZ_{banco_sel}_{nombre_ec}.xlsx"
            st.download_button(
                label="💾 Descargar póliza Excel (original)",
                data=datos_excel,
                file_name=nombre_sal,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
            import _viewer as _vwr
            _vwr.show(datos_excel, filename=nombre_sal, key="fortez_vwr")

        except Exception as ex:
            st.error(f"Error al generar: {ex}")
            import traceback
            st.code(traceback.format_exc(), language="python")
