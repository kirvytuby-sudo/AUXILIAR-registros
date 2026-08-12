"""
Página 18 — Póliza para Pago de Nómina
Inputs : PagosBancarios_Consolidado.xlsx  +  PLANTILLA PAGO DE NOMINA.xlsx
Output : Excel con dos hojas:
           • poliza IA    — matriz lotes × empleados (cargos) + columna banco (abono)
           • Conciliacion — comparativo fuente vs generado por lote
"""

import io
from collections import OrderedDict

import streamlit as st
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Póliza Pago de Nómina",
    page_icon="💸",
    layout="wide",
)

# ── constantes ────────────────────────────────────────────────────────────────
BANCO_NUM = "1020100010001"
BANCO_NOM = "BANCO BBVA"
TIPO_POL  = "EA"

AZUL      = "1E3A8A"
AZUL2     = "2563EB"
AMBAR     = "FEF08A"
VERDE     = "DCFCE7"
GRIS      = "F9FAFB"
MORADO    = "EDE9FE"
VERDE_OK  = "D1FAE5"
ROJO      = "FEE2E2"

# ── helpers de estilo ─────────────────────────────────────────────────────────
def _fill(hex_):
    return PatternFill("solid", fgColor=hex_)

_TH  = Side(style="thin", color="BBBBBB")
_BOR = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)
_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LFT = Alignment(horizontal="left",   vertical="center")
_RGT = Alignment(horizontal="right",  vertical="center")


def _w(ws, r, c, v, bold=False, color="000000", sz=8,
       bg=None, al=None, nf=None):
    cell = ws.cell(r, c)
    cell.value  = v
    cell.font   = Font(bold=bold, color=color, size=sz)
    cell.border = _BOR
    if bg: cell.fill           = _fill(bg)
    if al: cell.alignment      = al
    if nf: cell.number_format  = nf


# ── función generadora ────────────────────────────────────────────────────────
def generar_poliza_pago(bytes_consolidado: bytes,
                        bytes_plantilla:   bytes) -> tuple[bytes, int, int, float]:
    """
    Devuelve (xlsx_bytes, n_lotes, n_empleados, gran_total).
    """
    # -- leer consolidado
    wb_c   = openpyxl.load_workbook(io.BytesIO(bytes_consolidado), data_only=True)
    ws_res = wb_c["Resumen"]

    # Mapa nombre_stripped → nombre_real para tolerar espacios finales en sheet names
    _sheet_map = {s.strip(): s for s in wb_c.sheetnames}

    def _get_sheet(hoja_str):
        """Devuelve la hoja por nombre exacto o por nombre strippeado."""
        if hoja_str in wb_c.sheetnames:
            return wb_c[hoja_str]
        real = _sheet_map.get(hoja_str.strip())
        return wb_c[real] if real else None

    resumen = []
    for r in range(2, ws_res.max_row + 1):
        desc    = ws_res.cell(r, 1).value
        hoja    = ws_res.cell(r, 2).value
        fecha   = ws_res.cell(r, 3).value
        regs    = ws_res.cell(r, 4).value
        imp_src = ws_res.cell(r, 5).value
        if desc and hoja:
            resumen.append((
                str(desc).strip(), str(hoja).strip(), fecha,
                regs, float(imp_src) if imp_src else 0.0
            ))

    # empleados únicos en orden de aparición
    empleados_ord = OrderedDict()
    for desc, hoja, *_ in resumen:
        if _get_sheet(hoja) is None:
            continue
        ws = _get_sheet(hoja)
        for r in range(2, ws.max_row + 1):
            n  = ws.cell(r, 1).value
            nc = ws.cell(r, 2).value
            im = ws.cell(r, 3).value
            if not n or str(n).strip().upper() in ("TOTAL", "") or not im:
                continue
            try:
                float(im)
            except (TypeError, ValueError):
                continue
            key = (str(nc).strip() if nc else "", str(n).strip())
            empleados_ord[key] = True

    empleados = list(empleados_ord.keys())   # [(num_cta, nombre), ...]
    N_EMP     = len(empleados)

    # índices de columna
    C_EMP_INI = 9
    C_T1      = C_EMP_INI + N_EMP
    C_BANCO   = C_T1 + 1
    C_T2      = C_T1 + 2
    C_CONC    = C_T1 + 3

    # -- abrir plantilla e insertar columnas de empleados
    wb_out = openpyxl.load_workbook(io.BytesIO(bytes_plantilla))
    ws_pol = wb_out["poliza IA"]
    # La plantilla tiene 10 cols (A-J). I=TOTAL 1, J=TOTAL 2.
    # Insertar N_EMP + 2 cols antes de col 9 (desplaza TOTAL 1 y TOTAL 2).
    ws_pol.insert_cols(9, N_EMP + 2)

    # fila 1 — numeración
    for c in range(1, C_CONC + 1):
        _w(ws_pol, 1, c, c - 1 if c <= C_T1 else None,
           bold=True, color="FFFFFF", bg=AZUL, al=_CTR)

    # fila 2 — N° de cuenta
    for c in range(1, 9):
        _w(ws_pol, 2, c, None, bg="D6E4FF", al=_CTR)
    for i, (nc, _) in enumerate(empleados):
        _w(ws_pol, 2, C_EMP_INI + i, nc, bg="DBEAFE", al=_CTR)
    _w(ws_pol, 2, C_T1,    None,      bg="FEF9C3", al=_CTR)
    _w(ws_pol, 2, C_BANCO, BANCO_NUM, bold=True, color="166534", bg=VERDE, al=_CTR)
    _w(ws_pol, 2, C_T2,    None,      bg="FEF9C3", al=_CTR)
    _w(ws_pol, 2, C_CONC,  None,      bg=MORADO,   al=_CTR)

    # fila 3 — headers
    for c, h in enumerate(
        ["TIPO DE POLIZA", "Fecha", "REFERENCIA", "CONCEPTO",
         "ERROR", "UIDD", "NUM POLIZA", "PROCESADO"], 1
    ):
        _w(ws_pol, 3, c, h, bold=True, color="FFFFFF", bg=AZUL, al=_CTR)
    for i, (_, nombre) in enumerate(empleados):
        _w(ws_pol, 3, C_EMP_INI + i, nombre, bold=True, color="FFFFFF", bg=AZUL2, al=_CTR)
    _w(ws_pol, 3, C_T1,   "TOTAL 1",      bold=True, color="FFFFFF", bg=AZUL,     al=_CTR)
    _w(ws_pol, 3, C_BANCO, BANCO_NOM,     bold=True, color="FFFFFF", bg="166534", al=_CTR)
    _w(ws_pol, 3, C_T2,   "TOTAL 2",      bold=True, color="FFFFFF", bg=AZUL,     al=_CTR)
    _w(ws_pol, 3, C_CONC, "CONCILIACION", bold=True, color="7C3AED", bg=MORADO,   al=_CTR)

    # filas de datos — una por lote del Resumen
    poliza_rows = []
    ROW = 4
    for desc, hoja, fecha, regs_src, imp_src in resumen:
        ws = _get_sheet(hoja)
        if ws is None:
            continue
        lote = {}
        for r in range(2, ws.max_row + 1):
            n  = ws.cell(r, 1).value
            nc = ws.cell(r, 2).value
            im = ws.cell(r, 3).value
            if not n or str(n).strip().upper() in ("TOTAL", "") or not im:
                continue
            try:
                im = float(im)
            except (TypeError, ValueError):
                continue
            key = (str(nc).strip() if nc else "", str(n).strip())
            lote[key] = im
        if not lote:
            continue

        total_cargo = sum(lote.values())
        poliza_rows.append(dict(
            desc=desc, hoja=hoja, fecha=fecha,
            regs_src=regs_src, imp_src=imp_src,
            regs_gen=len(lote), imp_gen=total_cargo
        ))

        bg_r = "FFFFFF" if ROW % 2 == 0 else GRIS
        _w(ws_pol, ROW, 1, TIPO_POL, bg=bg_r, al=_CTR)
        _w(ws_pol, ROW, 2, fecha,    bg=bg_r, al=_CTR)
        _w(ws_pol, ROW, 3, desc,     bg=bg_r, al=_LFT)
        _w(ws_pol, ROW, 4, desc,     bg=bg_r, al=_LFT)
        for c in range(5, 9):
            _w(ws_pol, ROW, c, None, bg=bg_r)
        for i in range(N_EMP):
            _w(ws_pol, ROW, C_EMP_INI + i,
               lote.get(empleados[i]), bg=bg_r, al=_RGT, nf='#,##0.00')
        _w(ws_pol, ROW, C_T1,    total_cargo, bold=True, bg=AMBAR, al=_RGT, nf='#,##0.00')
        _w(ws_pol, ROW, C_BANCO, total_cargo, bold=True, bg=VERDE,  al=_RGT, nf='#,##0.00')
        _w(ws_pol, ROW, C_T2,    total_cargo, bold=True, bg=AMBAR, al=_RGT, nf='#,##0.00')
        conc_f = (f"={get_column_letter(C_T1)}{ROW}"
                  f"-{get_column_letter(C_T2)}{ROW}")
        _w(ws_pol, ROW, C_CONC, conc_f, bold=True, bg=MORADO, al=_RGT, nf='#,##0.00')
        ROW += 1

    # anchos
    ws_pol.column_dimensions["A"].width = 9
    ws_pol.column_dimensions["B"].width = 11
    ws_pol.column_dimensions["C"].width = 30
    ws_pol.column_dimensions["D"].width = 30
    for c in range(5, 9):
        ws_pol.column_dimensions[get_column_letter(c)].width = 7
    for c in range(C_EMP_INI, C_CONC + 1):
        ws_pol.column_dimensions[get_column_letter(c)].width = 12
    ws_pol.column_dimensions[get_column_letter(C_BANCO)].width = 15
    ws_pol.row_dimensions[2].height = 20
    ws_pol.row_dimensions[3].height = 32
    ws_pol.freeze_panes = f"C4"

    # ── hoja Conciliacion ─────────────────────────────────────────────────────
    ws_conc = wb_out.create_sheet("Conciliacion")

    ws_conc.merge_cells("A1:I1")
    cel = ws_conc["A1"]
    cel.value     = "CONCILIACIÓN — PAGOS CONSOLIDADOS vs PÓLIZA GENERADA"
    cel.font      = Font(bold=True, color="FFFFFF", size=11)
    cel.fill      = _fill(AZUL)
    cel.alignment = _CTR
    ws_conc.row_dimensions[1].height = 28

    cols_hdr = [
        ("Descripción",       28), ("Hoja",          22), ("Fecha",             12),
        ("Registros\nFuente", 10), ("Importe\nFuente",14), ("Registros\nPóliza", 10),
        ("Importe\nPóliza",   14), ("Diferencia",     13), ("Estado",            13),
    ]
    for ci, (h, wd) in enumerate(cols_hdr, 1):
        _w(ws_conc, 2, ci, h, bold=True, color="FFFFFF", bg=AZUL, al=_CTR)
        ws_conc.column_dimensions[get_column_letter(ci)].width = wd
    ws_conc.row_dimensions[2].height = 30

    grand_src = grand_gen = 0.0
    for ri, pr in enumerate(poliza_rows, 3):
        dif  = round(pr["imp_gen"] - pr["imp_src"], 2)
        ok   = abs(dif) < 0.01
        bg_r = "FFFFFF" if ri % 2 == 0 else "F0F4FF"
        bg_d = VERDE_OK if ok else ROJO
        clr  = "166534" if ok else "991B1B"
        _w(ws_conc, ri, 1, pr["desc"],     bg=bg_r, al=_LFT)
        _w(ws_conc, ri, 2, pr["hoja"],     bg=bg_r, al=_LFT, sz=7)
        _w(ws_conc, ri, 3, pr["fecha"],    bg=bg_r, al=_CTR)
        _w(ws_conc, ri, 4, pr["regs_src"], bg=bg_r, al=_CTR)
        _w(ws_conc, ri, 5, pr["imp_src"],  bg=bg_r, al=_RGT, nf='#,##0.00')
        _w(ws_conc, ri, 6, pr["regs_gen"], bg=bg_r, al=_CTR)
        _w(ws_conc, ri, 7, pr["imp_gen"],  bg=bg_r, al=_RGT, nf='#,##0.00')
        _w(ws_conc, ri, 8, dif,            bg=bg_d, al=_RGT, nf='#,##0.00',
           bold=not ok, color=clr)
        _w(ws_conc, ri, 9, "✅ OK" if ok else "⚠️ DIFERENCIA",
           bg=bg_d, al=_CTR, bold=True, color=clr)
        grand_src += pr["imp_src"]
        grand_gen += pr["imp_gen"]

    # fila gran total
    RT = len(poliza_rows) + 3
    ws_conc.merge_cells(f"A{RT}:D{RT}")
    dif_g = round(grand_gen - grand_src, 2)
    ok_g  = abs(dif_g) < 0.01
    _w(ws_conc, RT, 1, "GRAN TOTAL",
       bold=True, color="FFFFFF", bg=AZUL, al=_CTR, sz=9)
    _w(ws_conc, RT, 5, grand_src,
       bold=True, bg=AMBAR, al=_RGT, nf='#,##0.00', sz=9)
    _w(ws_conc, RT, 6, None, bg=AMBAR)
    _w(ws_conc, RT, 7, grand_gen,
       bold=True, bg=AMBAR, al=_RGT, nf='#,##0.00', sz=9)
    _w(ws_conc, RT, 8, dif_g,
       bold=True, bg=VERDE_OK if ok_g else ROJO, al=_RGT, nf='#,##0.00',
       color="166534" if ok_g else "991B1B", sz=9)
    _w(ws_conc, RT, 9,
       "✅ CUADRADO" if ok_g else "⚠️ NO CUADRA",
       bold=True, bg=VERDE_OK if ok_g else ROJO,
       color="166534" if ok_g else "991B1B", al=_CTR, sz=9)
    ws_conc.row_dimensions[RT].height = 22

    # bloque resumen
    RS = RT + 2
    ws_conc.merge_cells(f"A{RS}:I{RS}")
    cel2 = ws_conc.cell(RS, 1)
    cel2.value     = "RESUMEN GENERAL"
    cel2.font      = Font(bold=True, color="FFFFFF", size=9)
    cel2.fill      = _fill("374151")
    cel2.alignment = _CTR
    cel2.border    = _BOR

    datos_res = [
        ("Lotes en fuente (Resumen)",  len(resumen)),
        ("Lotes procesados en póliza", len(poliza_rows)),
        ("Empleados únicos",           N_EMP),
        ("Total importe fuente",       f"${grand_src:,.2f}"),
        ("Total importe póliza",       f"${grand_gen:,.2f}"),
        ("Diferencia neta",            f"${dif_g:,.2f}"),
    ]
    for li, (lbl, val) in enumerate(datos_res, RS + 1):
        ws_conc.merge_cells(f"A{li}:E{li}")
        _w(ws_conc, li, 1, lbl, bold=True, bg="F3F4F6", al=_LFT)
        ws_conc.merge_cells(f"F{li}:I{li}")
        es_dif = "Diferencia" in lbl
        bg_v   = (VERDE_OK if ok_g else ROJO) if es_dif else "FFFFFF"
        clr_v  = ("166534" if ok_g else "991B1B") if es_dif else "000000"
        _w(ws_conc, li, 6, val, bold=True, bg=bg_v, color=clr_v, al=_LFT)

    ws_conc.freeze_panes = "A3"

    # serializar
    buf = io.BytesIO()
    wb_out.save(buf)
    return buf.getvalue(), len(poliza_rows), N_EMP, grand_gen


# ── UI ────────────────────────────────────────────────────────────────────────
import _theme
_theme.aplicar_header(
    "💸 Póliza para Pago de Nómina",
    "PagosBancarios_Consolidado.xlsx + Plantilla → Póliza contable con conciliación",
)

col1, col2 = st.columns(2)
with col1:
    f_consolidado = st.file_uploader(
        "📊 PagosBancarios_Consolidado.xlsx",
        type=["xlsx"],
        key="ppn_consolidado",
    )
with col2:
    f_plantilla = st.file_uploader(
        "📋 PLANTILLA PAGO DE NOMINA.xlsx",
        type=["xlsx"],
        key="ppn_plantilla",
    )

listo = f_consolidado is not None and f_plantilla is not None

if st.button("🔄 Generar Póliza", type="primary", disabled=not listo):
    bytes_cons  = f_consolidado.read()
    bytes_tmpl  = f_plantilla.read()

    with st.spinner("Generando póliza…"):
        try:
            xlsx_bytes, n_lotes, n_emp, gran_total = generar_poliza_pago(
                bytes_cons, bytes_tmpl
            )
            st.success(
                f"✅ {n_lotes} lote(s) · {n_emp} empleado(s) · "
                f"Total: **${gran_total:,.2f}**"
            )

            st.download_button(
                label="⬇️ Descargar Excel",
                data=xlsx_bytes,
                file_name="Poliza_Pago_Nomina.xlsx",
                mime=(
                    "application/vnd.openxmlformats-"
                    "officedocument.spreadsheetml.sheet"
                ),
            )

            # vista previa — póliza
            st.subheader("Vista previa — póliza IA")
            try:
                import pandas as pd
                df = pd.read_excel(
                    io.BytesIO(xlsx_bytes),
                    sheet_name="poliza IA",
                    header=2,
                    nrows=30,
                    dtype=str,
                ).fillna("")
                st.dataframe(df, use_container_width=True, height=300)
            except Exception as _e:
                st.warning(f"Vista previa no disponible: {_e}")

            # vista previa — conciliación
            st.subheader("Vista previa — Conciliacion")
            try:
                import pandas as pd
                df2 = pd.read_excel(
                    io.BytesIO(xlsx_bytes),
                    sheet_name="Conciliacion",
                    header=1,
                    nrows=40,
                    dtype=str,
                ).fillna("")
                st.dataframe(df2, use_container_width=True, height=250)
            except Exception as _e2:
                st.info(f"Hoja Conciliacion: {_e2}")

        except Exception as exc:
            import traceback
            st.error(f"Error al generar: {exc}")
            st.code(traceback.format_exc())
