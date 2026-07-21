"""
AUXILIAR DE REGISTROS — Módulo: Provisión de Nómina
XMLs CFDI de nómina → plantilla Excel SINUBE con columnas dinámicas.
"""
import os, re, tempfile, shutil
import xml.etree.ElementTree as ET
from datetime import datetime
import streamlit as st

st.set_page_config(
    page_title="Provisión de Nómina · Auxiliar",
    page_icon="📋",
    layout="wide",
)

import _theme
_theme.aplicar_header("📋 Provisión de Nómina", "Archivos XML (CFDI de Nómina) → Plantilla Excel SINUBE con columnas dinámicas")
# ── Estado de sesión ───────────────────────────────────────────────────────────
if "pn_resultado_bytes" not in st.session_state:
    st.session_state.pn_resultado_bytes = None
if "pn_tmp" not in st.session_state:
    st.session_state.pn_tmp = tempfile.mkdtemp(prefix="pn_")

TMP = st.session_state.pn_tmp

def norm(s):
    return re.sub(r'\s+', ' ', str(s or '').strip().upper())

# ── Sección 1: Carga de archivos ───────────────────────────────────────────────
st.subheader("1️⃣  Cargar archivos")
col_plantilla, col_xmls = st.columns([1, 1])

with col_plantilla:
    plantilla_file = st.file_uploader(
        "📊 Plantilla Excel SINUBE (.xlsx)",
        type=["xlsx"],
        help="Archivo con hojas POLIZA y CUENTAS.",
    )

with col_xmls:
    xml_files = st.file_uploader(
        "📁 Archivos XML (CFDI de Nómina) — puedes seleccionar varios",
        type=["xml"],
        accept_multiple_files=True,
    )

if not plantilla_file:
    st.info("👆 Carga la plantilla Excel SINUBE y los XMLs de nómina para comenzar.")
    st.stop()

if not xml_files:
    st.warning("⚠️ Carga al menos un archivo XML para procesar.")
    st.stop()

# ── Sección 2: Procesar ────────────────────────────────────────────────────────
st.subheader("2️⃣  Generar provisión")
st.caption(f"Plantilla: **{plantilla_file.name}** · XMLs cargados: **{len(xml_files)}**")

if st.button("⚙️  Procesar Provisión de Nómina", type="primary", use_container_width=False):
    logs = []
    errores = []

    with st.spinner("Procesando…"):
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

            # ── Guardar plantilla en tmp ──────────────────────────────────────
            plantilla_path = os.path.join(TMP, plantilla_file.name)
            with open(plantilla_path, "wb") as f:
                f.write(plantilla_file.read())

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = os.path.join(TMP, f"Provision_Nomina_{ts}.xlsx")
            shutil.copy2(plantilla_path, out_path)

            wb = openpyxl.load_workbook(out_path)

            # ── Auto-detectar hoja POLIZA ─────────────────────────────────────
            ws = None
            for sn in wb.sheetnames:
                if sn.strip().upper() in ('POLIZA', 'PÓLIZA', 'POLIZA IA'):
                    ws = wb[sn]
                    break
            if ws is None:
                ws = wb.active

            # Limpiar filas de datos (mantener encabezados filas 1-3)
            for r in range(ws.max_row, 3, -1):
                ws.delete_rows(r)

            # ── Leer hoja CUENTAS ─────────────────────────────────────────────
            ws_c = None
            for sn in wb.sheetnames:
                if sn.strip().upper() == 'CUENTAS':
                    ws_c = wb[sn]
                    break

            perc_cat     = []   # [(cuenta, nombre_upper)]  cols A/B
            ded_cat      = []   # [(cuenta, nombre_upper)]  cols M/N
            emp_reg      = {}   # nombre_upper -> cuenta     cols G/H
            emp_asim     = {}   # nombre_upper -> cuenta     cols J/K
            emp_prestamo = {}   # nombre_upper -> cuenta     cols D/E

            if ws_c is not None:
                for _row in ws_c.iter_rows(min_row=2, values_only=True):
                    def _v(idx): return _row[idx] if len(_row) > idx else None
                    _a, _b = _v(0), _v(1)
                    _d, _e = _v(3), _v(4)
                    _g, _h = _v(6), _v(7)
                    _j, _k = _v(9), _v(10)
                    _m, _n = _v(12), _v(13)
                    if _a and _b: perc_cat.append((str(_a).strip(), norm(_b)))
                    if _d and _e: emp_prestamo[norm(_e)] = str(_d).strip()
                    if _g and _h: emp_reg[norm(_h)]  = str(_g).strip()
                    if _j and _k: emp_asim[norm(_k)] = str(_j).strip()
                    if _m and _n: ded_cat.append((str(_m).strip(), norm(_n)))

            logs.append(
                f"CUENTAS: {len(perc_cat)} percepciones, "
                f"{len(emp_reg)+len(emp_asim)} empleados, "
                f"{len(emp_prestamo)} préstamos, {len(ded_cat)} deducciones."
            )

            # ── Localizar TOTAL 1 y TOTAL 2 en fila 3 ────────────────────────
            col_tot1 = col_tot2 = None
            for c in range(1, ws.max_column + 1):
                v = norm(ws.cell(3, c).value or '')
                if 'TOTAL 1' in v:   col_tot1 = c
                elif 'TOTAL 2' in v: col_tot2 = c

            if col_tot1 is None or col_tot2 is None:
                st.error("❌ No se encontraron columnas TOTAL 1 / TOTAL 2 en fila 3 de la plantilla.")
                st.stop()

            col_perc_start = col_tot1

            # ── Insertar percepciones ANTES de TOTAL 1 ───────────────────────
            perc_cols = {}
            for i, (cta, nombre) in enumerate(perc_cat):
                col_ins = col_tot1 + i
                ws.insert_cols(col_ins)
                ws.cell(2, col_ins, value=cta)
                ws.cell(3, col_ins, value=nombre)
                perc_cols[nombre] = col_ins
            col_tot1 += len(perc_cat)
            col_tot2 += len(perc_cat)

            # ── Insertar deducciones DESPUÉS de TOTAL 1 ──────────────────────
            ded_cols = {}
            for i, (cta, nombre) in enumerate(ded_cat):
                col_ins = col_tot1 + 1 + i
                ws.insert_cols(col_ins)
                ws.cell(2, col_ins, value=cta)
                ws.cell(3, col_ins, value=nombre)
                ded_cols[nombre] = col_ins
            col_tot2 += len(ded_cat)

            logs.append(f"Estructura: TOTAL 1=col{col_tot1}  |  TOTAL 2=col{col_tot2}")

            # ── Parsear XMLs ──────────────────────────────────────────────────
            filas = []
            for uf in xml_files:
                xml_bytes = uf.read()
                xml_name  = uf.name
                try:
                    root = ET.fromstring(xml_bytes)
                    receptor = nomina = None
                    uuid = ''
                    for el in root.iter():
                        local = el.tag.split('}')[-1] if '}' in el.tag else el.tag
                        if local == 'Receptor' and receptor is None:
                            receptor = el
                        elif local == 'Nomina' and 'FechaPago' in el.attrib and nomina is None:
                            nomina = el
                        elif local == 'TimbreFiscalDigital' and not uuid:
                            uuid = el.get('UUID', '')

                    if nomina is None:
                        errores.append(f"{xml_name} — sin nodo Nomina (se omite)")
                        continue

                    nombre = re.sub(r'\s+', ' ', (receptor.get('Nombre', '') if receptor is not None else '').strip())
                    rfc    = (receptor.get('Rfc', '').strip() if receptor is not None else '')
                    total  = float(root.get('Total', 0))
                    fecha  = datetime.strptime(nomina.get('FechaPago'), '%Y-%m-%d')

                    perc = {}
                    for el in root.iter():
                        if el.tag.split('}')[-1] == 'Percepcion':
                            c = norm(el.get('Concepto', ''))
                            v = (float(el.get('ImporteGravado', 0)) +
                                 float(el.get('ImporteExento',  0)))
                            if c and v:
                                perc[c] = perc.get(c, 0) + v

                    ded = {}
                    for el in root.iter():
                        if el.tag.split('}')[-1] == 'Deduccion':
                            c = norm(el.get('Concepto', ''))
                            v = float(el.get('Importe', 0))
                            if c and v:
                                ded[c] = ded.get(c, 0) + v

                    otro_prest = 0.0
                    for el in root.iter():
                        if el.tag.split('}')[-1] == 'OtroPago':
                            c = norm(el.get('Concepto', ''))
                            v = float(el.get('Importe', 0))
                            if 'PRESTAMO' in c and v:
                                otro_prest += v

                    es_asimilado = 'ASIMILABLES A SALARIOS' in perc

                    filas.append({
                        'fecha':      fecha,
                        'rfc':        rfc,
                        'nombre':     nombre,
                        'nombre_up':  norm(nombre),
                        'uuid':       uuid,
                        'total':      total,
                        'perc':       perc,
                        'ded':        ded,
                        'otro_prest': otro_prest,
                        'asimilado':  es_asimilado,
                        'ref':        f"NOMINA DEL {fecha.strftime('%d/%m/%Y')}",
                    })
                    logs.append(f"✓ {xml_name}  |  {nombre}  |  ${total:,.2f}")
                except Exception as e:
                    errores.append(f"{xml_name}: {e}")

            if not filas:
                st.error("❌ No se procesó ningún XML.")
                st.stop()

            filas.sort(key=lambda f: f['fecha'])

            # ── Columnas OtroPago PRESTAMO (antes TOTAL 1) ───────────────────
            otro_prest_cols = {}
            vistos_otro_prest = []
            for f in filas:
                if f['otro_prest'] and f['nombre_up'] not in vistos_otro_prest:
                    vistos_otro_prest.append(f['nombre_up'])
            for nu in vistos_otro_prest:
                cta = emp_prestamo.get(nu, '')
                col_ins = col_tot1
                ws.insert_cols(col_ins)
                ws.cell(2, col_ins, value=cta)
                ws.cell(3, col_ins, value=f"PREST.OTROP. {nu}")
                otro_prest_cols[nu] = col_ins
                col_tot1 += 1
                col_tot2 += 1
                if cta:
                    logs.append(f"  + OtroPago Préstamo {nu}  →  {cta}")
                else:
                    errores.append(f"Sin cuenta préstamo D/E para: '{nu}'")

            if otro_prest_cols:
                n_shift = len(otro_prest_cols)
                ded_cols = {k: v + n_shift for k, v in ded_cols.items()}

            # ── Columnas PRESTAMO deducción (entre ded y neto) ───────────────
            prest_cols = {}
            vistos_prest = []
            for f in filas:
                for c_name in f['ded']:
                    if 'PRESTAMO' in c_name and f['nombre_up'] not in vistos_prest:
                        vistos_prest.append(f['nombre_up'])
                        break
            for nu in vistos_prest:
                cta = emp_prestamo.get(nu, '')
                col_ins = col_tot2
                ws.insert_cols(col_ins)
                ws.cell(2, col_ins, value=cta)
                ws.cell(3, col_ins, value=f"PREST. {nu}")
                prest_cols[nu] = col_ins
                col_tot2 += 1
                if cta:
                    logs.append(f"  + Préstamo deducción {nu}  →  {cta}")
                else:
                    errores.append(f"Sin cuenta préstamo D/E para: '{nu}'")

            # ── Columnas neto por empleado (antes TOTAL 2) ────────────────────
            vistos = []
            for f in filas:
                if f['nombre_up'] and f['nombre_up'] not in vistos:
                    vistos.append(f['nombre_up'])

            emp_cols = {}
            for nu in vistos:
                cta = emp_reg.get(nu) or emp_asim.get(nu) or ''
                col_ins = col_tot2
                ws.insert_cols(col_ins)
                ws.cell(2, col_ins, value=cta)
                ws.cell(3, col_ins, value=nu)
                emp_cols[nu] = col_ins
                col_tot2 += 1
                if cta:
                    logs.append(f"  + {nu}  →  {cta}")
                else:
                    errores.append(f"Sin cuenta en CUENTAS G/H o J/K: '{nu}'")

            col_conc = col_tot2 + 1
            ws.cell(3, col_conc, value='CONCILIACION')

            # ── Fila 1: numeración 0-based ────────────────────────────────────
            _fill_num = PatternFill("solid", fgColor="1F3864")
            _font_num = Font(name="Segoe UI", bold=True, color="FFFFFF", size=8)
            _align_ctr = Alignment(horizontal="center", vertical="center")
            _side_thin = Side(style="thin", color="3A6BAF")
            _border_num = Border(
                left=_side_thin, right=_side_thin,
                top=_side_thin, bottom=_side_thin,
            )
            ws.row_dimensions[1].height = 16
            for _c in range(1, col_conc + 1):
                _cell = ws.cell(1, _c, value=_c - 1)
                _cell.fill = _fill_num
                _cell.font = _font_num
                _cell.alignment = _align_ctr
                _cell.border = _border_num

            # ── Mapa de encabezados fila 3 ────────────────────────────────────
            col_meta = {}
            for c in range(1, ws.max_column + 1):
                v = norm(ws.cell(3, c).value or '')
                if v:
                    col_meta[v] = c

            perc_no_mapeadas = set()
            ded_no_mapeadas  = set()

            # ── Escribir filas de datos ───────────────────────────────────────
            nf = 4
            for fila in filas:
                if 'TIPO DE POLIZA' in col_meta:
                    ws.cell(nf, col_meta['TIPO DE POLIZA'], value='D')
                if 'FECHA' in col_meta:
                    c = ws.cell(nf, col_meta['FECHA'], value=fila['fecha'])
                    c.number_format = 'DD/MM/YYYY'
                if 'REFERENCIA' in col_meta:
                    ws.cell(nf, col_meta['REFERENCIA'], value=fila['ref'])
                if 'CONCEPTO' in col_meta:
                    ws.cell(nf, col_meta['CONCEPTO'], value=fila['ref'])
                if 'UIDD' in col_meta:
                    ws.cell(nf, col_meta['UIDD'],
                            value=f"{fila['uuid']};{fila['rfc']};{fila['total']}")

                # Percepciones
                for c_name, c_val in fila['perc'].items():
                    _c = perc_cols.get(c_name)
                    if _c and c_val:
                        ws.cell(nf, _c, value=c_val).number_format = '#,##0.00'
                    elif c_val and not _c:
                        perc_no_mapeadas.add(c_name)

                # Deducciones
                _prest_acum = 0.0
                for c_name, c_val in fila['ded'].items():
                    if 'PRESTAMO' in c_name:
                        _prest_acum += c_val
                        continue
                    col_key = c_name
                    if fila.get('asimilado') and c_name == 'ISR':
                        col_key = 'ISR ASIMILADOS'
                    _c = ded_cols.get(col_key)
                    if _c and c_val:
                        ws.cell(nf, _c, value=c_val).number_format = '#,##0.00'
                    elif c_val and not _c:
                        ded_no_mapeadas.add(col_key)
                if _prest_acum:
                    _cp = prest_cols.get(fila['nombre_up'])
                    if _cp:
                        ws.cell(nf, _cp, value=_prest_acum).number_format = '#,##0.00'
                    else:
                        ded_no_mapeadas.add('PRESTAMO (sin col empleado)')

                # OtroPago PRESTAMO
                if fila['otro_prest']:
                    _cp_otro = otro_prest_cols.get(fila['nombre_up'])
                    if _cp_otro:
                        ws.cell(nf, _cp_otro, value=fila['otro_prest']).number_format = '#,##0.00'
                    else:
                        perc_no_mapeadas.add('PRESTAMO OtroPago (sin col empleado)')

                # Neto del empleado
                _ce = emp_cols.get(fila['nombre_up'])
                if _ce:
                    ws.cell(nf, _ce, value=fila['total']).number_format = '#,##0.00'

                # TOTAL 1
                p_s = get_column_letter(col_perc_start)
                p_e = get_column_letter(col_tot1 - 1)
                ws.cell(nf, col_tot1,
                    value=f"=SUM({p_s}{nf}:{p_e}{nf})").number_format = '#,##0.00'

                # TOTAL 2
                d_s = get_column_letter(col_tot1 + 1)
                d_e = get_column_letter(col_tot2 - 1)
                ws.cell(nf, col_tot2,
                    value=f"=SUM({d_s}{nf}:{d_e}{nf})").number_format = '#,##0.00'

                # CONCILIACION
                t1l = get_column_letter(col_tot1)
                t2l = get_column_letter(col_tot2)
                ws.cell(nf, col_conc,
                    value=f"={t1l}{nf}-{t2l}{nf}").number_format = '#,##0.00'

                nf += 1

            # ── Estilos ───────────────────────────────────────────────────────
            def _fill(hex6): return PatternFill("solid", fgColor=hex6)
            def _font(bold=False, color="000000", size=9, italic=False):
                return Font(name="Segoe UI", bold=bold, italic=italic,
                            color=color, size=size)
            def _border(color="C9D9EE"):
                s = Side(style="thin", color=color)
                return Border(left=s, right=s, top=s, bottom=s)

            C_MID   = "2F5496"; C_BLUE  = "4472C4"
            C_AMBER = "FFC000"; C_AMBER2 = "FFF2CC"
            C_GREEN = "375623"; C_GREEN2 = "E2EFDA"
            C_ROW1  = "FFFFFF"; C_ROW2  = "D6E4F7"
            C_WHITE = "FFFFFF"; C_NAVY  = "1F3864"
            brd = _border()

            # Fila 2 (cuentas)
            ws.row_dimensions[2].height = 14
            for c in range(1, col_conc + 1):
                cell = ws.cell(2, c)
                if c in (col_tot1, col_tot2):
                    cell.fill = _fill(C_AMBER); cell.font = _font(italic=True, color=C_NAVY, size=8)
                elif c == col_conc:
                    cell.fill = _fill(C_GREEN); cell.font = _font(italic=True, color=C_WHITE, size=8)
                else:
                    cell.fill = _fill(C_MID); cell.font = _font(color=C_WHITE, size=8, italic=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = brd

            # Fila 3 (encabezados)
            ws.row_dimensions[3].height = 42
            for c in range(1, col_conc + 1):
                cell = ws.cell(3, c)
                if c in (col_tot1, col_tot2):
                    cell.fill = _fill(C_AMBER); cell.font = _font(bold=True, color=C_NAVY, size=9)
                elif c == col_conc:
                    cell.fill = _fill(C_GREEN); cell.font = _font(bold=True, color=C_WHITE, size=9)
                else:
                    cell.fill = _fill(C_BLUE); cell.font = _font(bold=True, color=C_WHITE, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = brd

            # Filas de datos
            for r in range(4, nf):
                ws.row_dimensions[r].height = 14
                fondo = _fill(C_ROW1 if r % 2 != 0 else C_ROW2)
                for c in range(1, col_conc + 1):
                    cell = ws.cell(r, c)
                    if c >= col_perc_start:
                        cell.number_format = '#,##0.00'
                    if c in (col_tot1, col_tot2):
                        cell.fill = _fill(C_AMBER2); cell.font = _font(bold=True, color=C_NAVY, size=9)
                    elif c == col_conc:
                        cell.fill = _fill(C_GREEN2); cell.font = _font(bold=True, color=C_NAVY, size=9)
                    else:
                        cell.fill = fondo; cell.font = _font(size=9)
                    cell.alignment = Alignment(
                        horizontal="right" if c >= col_perc_start else "left",
                        vertical="center")
                    cell.border = brd

            wb.save(out_path)
            wb.close()

            with open(out_path, "rb") as f:
                st.session_state.pn_resultado_bytes = f.read()
            st.session_state.pn_nombre = f"Provision_Nomina_{ts}.xlsx"

            # ── Diagnóstico ───────────────────────────────────────────────────
            if perc_no_mapeadas:
                errores.append(
                    "Percepciones sin columna en CUENTAS A/B (CONCILIACION puede ≠ 0): "
                    + ", ".join(sorted(perc_no_mapeadas)))
            if ded_no_mapeadas:
                errores.append(
                    "Deducciones sin columna en CUENTAS M/N (CONCILIACION puede ≠ 0): "
                    + ", ".join(sorted(ded_no_mapeadas)))

            emp_sin_cuenta = [nu for nu in vistos if not (emp_reg.get(nu) or emp_asim.get(nu))]
            if emp_sin_cuenta:
                errores.append(
                    "Empleados sin cuenta en G/H o J/K: " + ", ".join(emp_sin_cuenta))

            st.success(f"✅ {len(filas)} XML(s) procesados correctamente.")

        except Exception as e:
            import traceback
            st.error(f"❌ Error inesperado: {e}")
            with st.expander("Ver detalle"):
                st.code(traceback.format_exc())
            st.stop()

    # ── Logs y advertencias ────────────────────────────────────────────────────
    if logs:
        with st.expander("📋 Log de procesamiento", expanded=False):
            st.code("\n".join(logs))
    if errores:
        with st.expander("⚠️ Advertencias", expanded=True):
            for e in errores:
                st.warning(e)

# ── Descarga ───────────────────────────────────────────────────────────────────
if st.session_state.pn_resultado_bytes:
    nombre_dl = st.session_state.get("pn_nombre", "Provision_Nomina.xlsx")
    st.download_button(
        label="⬇️  Descargar Excel de Provisión",
        data=st.session_state.pn_resultado_bytes,
        file_name=nombre_dl,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="secondary",
    )

st.markdown("---")
st.caption("Módulo Provisión de Nómina · v2.0")
