"""AUXILIAR DE REGISTROS — Ventas del Día (Control de Despachos → Póliza Excel)"""
import streamlit as st
import io
from collections import defaultdict
from datetime import datetime

st.set_page_config(page_title="Ventas del Día · Auxiliar", page_icon="⛽", layout="wide")

st.markdown("""<style>
[data-testid="stAppViewContainer"] { background: #dbeafe; }
.header-bar{background:#1E3A8A;padding:18px 28px;border-radius:10px;margin-bottom:20px;}
.header-bar h1{color:#FBCFE8;font-size:1.5rem;margin:0;}
.header-bar p{color:#93C5FD;margin:4px 0 0;font-size:.9rem;}
#MainMenu{visibility:hidden;}footer{visibility:hidden;}
</style>
<div class="header-bar">
  <h1>⛽ Ventas del Día</h1>
  <p>Genera la póliza contable de ventas diarias desde el control de despachos</p>
</div>
""", unsafe_allow_html=True)

try:
    import xlsxwriter
    import openpyxl
except ImportError as _e:
    st.error(f"❌ Librería faltante: {_e}. Verifica requirements.txt.")
    st.stop()

# ── Cuentas IEPS (fijas) ────────────────────────────────────────────────────
IEPS_GS = "401-01-0001-0006-0001"
IEPS_GP = "401-01-0001-0006-0002"
IEPS_GD = "401-01-0001-0006-0003"

CLIENTES_TPL = [
    "T. BANORTE","Contado","T. EDENRED (TICKET CAR)","T. EFECTICARD",
    "T. SODEXO (GASOPASS)","T. AMERICAN EXPRESS","V. EFECTIVALE",
    "ACEROS Y CEMENTOS TEPEPAN SA DE CV","CONCESIONARIA KIOTO SA DE CV",
    "EXCELENCIA COREANA S.A. DE C.V.","GOTESA SA DE CV",
    "INGENIERIA Y SERVICIO EN TRATAMIENTO DE AGUA SA DE CV",
    "MF CONSTRUCCIONES Y ASOCIADOS SA DE CV","PHIEMES SA DE CV",
    "SUOMI PUBLICITY AND LEAFLETING",
]
CUENTAS_TPL = [
    "105-01-0001-0004","105-01-0001-0001","105-01-0002-0002","105-01-0002-0009",
    "105-01-0002-0008","105-01-0001-0007","105-01-0002-0005",
    "105-01-0003-0005","105-01-0003-0602",
    "105-01-0003-1200","105-01-0003-1801",
    "105-01-0003-2400",
    "105-01-0003-3601","105-01-0003-4700",
    "105-01-0003-5701",
]

META_HDRS = ["TIPO DE POLIZA", "Fecha", "REFERENCIA", "CONCEPTO", "ERROR", "UIDD", "NUM POLIZA", "PROCESADO"]
PRODS     = ["GS", "GP", "GD", "IVA", "IEPS GS", "IEPS GP", "IEPS GD"]
CTAS_PROD = [
    "401-01-0001-0001", "401-01-0001-0002", "401-01-0001-0003", "209-01",
    IEPS_GS, IEPS_GP, IEPS_GD,
]


def _leer_cuentas_plantilla(plantilla_bytes):
    """Lee la hoja CUENTAS de la plantilla; retorna (cuentas_map, dyn_clientes, dyn_prods)."""
    cuentas_map = {
        IEPS_GS: "IEPS De Gasolina Magna",
        IEPS_GP: "IEPS de Premium",
        IEPS_GD: "IEPS de Diesel",
    }
    dyn_clientes = []  # [(num_cuenta, nombre)] — col H/I (7,8)
    dyn_prods    = []  # [(num_cuenta, nombre)] — col L/M (11,12)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(plantilla_bytes), data_only=True)
        hoja = None
        for sn in wb.sheetnames:
            if sn.strip().upper() == "CUENTAS":
                hoja = wb[sn]
                break
        if hoja:
            for r in hoja.iter_rows(values_only=True):
                # Clientes — col H(7) cuenta, I(8) nombre
                _nc = r[7] if len(r) > 7 else None
                _nm = r[8] if len(r) > 8 else None
                if _nc is not None and _nm is not None:
                    _ncs = str(_nc).strip()
                    _nms = str(_nm).strip().replace('\n', '').strip()
                    if _ncs and _nms:          # sin filtro isdigit() — acepta cualquier clave
                        cuentas_map[_ncs] = _nms
                        dyn_clientes.append((_ncs, _nms))
                # Productos — col L(11) cuenta, M(12) nombre
                _pc = r[11] if len(r) > 11 else None
                _pm = r[12] if len(r) > 12 else None
                if _pc is not None and _pm is not None:
                    _pcs = str(_pc).strip()
                    _pms = str(_pm).strip().replace('\n', '').strip()
                    if _pcs and _pms:
                        cuentas_map[_pcs] = _pms
                        dyn_prods.append((_pcs, _pms))
        wb.close()
    except Exception:
        pass
    return cuentas_map, dyn_clientes, dyn_prods


import re as _re

def _norm(s):
    """Normaliza nombre para comparación: mayúsculas, sin espacios/saltos de línea extra."""
    return str(s or "").replace('\n', ' ').strip().upper()

def _clean(s):
    """Limpieza profunda: quita paréntesis, reemplaza puntos por espacio, comprime espacios."""
    s = str(s or "").replace('\n', ' ')
    s = _re.sub(r'\(.*?\)', '', s)   # quitar (ACCOR), (SA DE CV), etc.
    s = s.replace('.', ' ')           # T.EDENRED → T EDENRED
    return _re.sub(r'\s+', ' ', s).strip().upper()

def _sig_words(s):
    """Palabras significativas (>3 chars) del texto limpio."""
    return {w for w in _clean(s).split() if len(w) > 3}

def _match_cliente(raw, clientes_list):
    """Mapea nombre de cliente del despacho al nombre exacto de la plantilla.
    1) Vacío → busca 'CONTADO' en la lista (o primer elemento).
    2) Exacto normalizado.
    3) Subcadena normalizada (ej. 'T.EDENRED' ↔ 'Clientes T.EDENRED').
    4) Subcadena limpia: sin puntos ni paréntesis
       (ej. 'T EFECTIVALE' ↔ 'Clientes T.EFECTIVALE',
            'T EDENRED (ACCOR)' ↔ 'Clientes T.EDENRED').
    5) Palabra significativa en común
       (ej. 'T BANORTE' ↔ 'Clientes Tarjeta banorte' → BANORTE).
    6) Sin coincidencia → retorna el nombre original (aparecerá en log ⚠️).
    """
    raw_s = str(raw or "").strip()
    raw_n = _norm(raw_s)
    if not raw_n:
        for nm in clientes_list:
            if _norm(nm) == "CONTADO":
                return nm
        return clientes_list[0] if clientes_list else raw_s

    # Paso 1: exacto
    for nm in clientes_list:
        if _norm(nm) == raw_n:
            return nm

    # Paso 2: subcadena normal
    for nm in clientes_list:
        nm_n = _norm(nm)
        if raw_n in nm_n or nm_n in raw_n:
            return nm

    # Paso 3: subcadena limpia (sin puntos/paréntesis)
    raw_c = _clean(raw_s)
    for nm in clientes_list:
        nm_c = _clean(nm)
        if raw_c in nm_c or nm_c in raw_c:
            return nm

    # Paso 4: palabra significativa en común (última opción)
    raw_words = _sig_words(raw_s)
    best, best_score = None, 0
    for nm in clientes_list:
        nm_words = _sig_words(nm)
        score = len(raw_words & nm_words)
        if score > best_score:
            best_score, best = score, nm
    if best_score > 0:
        return best

    # Sin coincidencia
    return raw_s


def _detectar_columnas(header_row, logs):
    """
    Lee la fila de encabezado y retorna un dict {nombre_col: índice}.
    Soporta MENA/PERIFERICO (27 cols, Cliente en r[16]) y
    VALLEJO (28 cols con None extra en r[14], Cliente en r[17]).
    """
    col_map = {}
    for i, h in enumerate(header_row):
        if h is not None:
            col_map[str(h).strip()] = i
    # Columnas esperadas con fallbacks seguros
    defaults = {
        'FechaHora': 0, 'Producto': 3, 'Subtotal': 6,
        'Iva': 7, 'Ieps': 8, 'Importe': 9, 'Cliente': 16,
    }
    result = {k: col_map.get(k, v) for k, v in defaults.items()}
    logs.append(
        f"  Columnas detectadas → Producto:{result['Producto']} "
        f"Importe:{result['Importe']} Cliente:{result['Cliente']}"
    )
    return result


def _leer_despachos(file_bytes, filename, logs):
    """Lee el archivo de despachos (.xlsx o .xls).
    Retorna (data, col_map) — data sin encabezado, col_map con índices detectados."""
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else 'xlsx'
    if ext == 'xls':
        try:
            import xlrd
            wb_xls = xlrd.open_workbook(file_contents=file_bytes)
            ws_xls = wb_xls.sheet_by_index(0)
            rows = [tuple(ws_xls.row_values(r)) for r in range(ws_xls.nrows)]
            col_map = _detectar_columnas(rows[0], logs)
            data = rows[1:]
            logs.append(f"  {len(data):,} registros leídos (.xls vía xlrd).")
            return data, col_map
        except ImportError:
            raise RuntimeError(
                "El archivo está en formato antiguo .xls y la librería 'xlrd' no está instalada.\n"
                "Solución: Abre el archivo en Excel y guárdalo como .xlsx."
            )
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        all_rows = list(wb.active.iter_rows(values_only=True))
        wb.close()
        col_map = _detectar_columnas(all_rows[0], logs)
        data = all_rows[1:]
        logs.append(f"  {len(data):,} registros leídos (.xlsx).")
        return data, col_map


def procesar_ventas(despachos_bytes, despachos_nombre, plantilla_bytes=None):
    """
    Procesa el control de despachos y genera la póliza Excel.
    Retorna (excel_bytes: bytes, logs: list[str]).
    """
    logs = []

    logs.append("⛽ Leyendo control de despachos...")
    data, col_map = _leer_despachos(despachos_bytes, despachos_nombre, logs)
    C_FECHA    = col_map['FechaHora']
    C_PROD     = col_map['Producto']
    C_SUBTOTAL = col_map['Subtotal']
    C_IVA      = col_map['Iva']
    C_IEPS     = col_map['Ieps']
    C_IMPORTE  = col_map['Importe']
    C_CLIENTE  = col_map['Cliente']

    # ── Leer plantilla de cuentas ──────────────────────────────────────────
    if plantilla_bytes:
        logs.append("⛽ Leyendo plantilla de cuentas...")
        cuentas_map, _dyn_cli, _dyn_prods = _leer_cuentas_plantilla(plantilla_bytes)
        logs.append(f"  {len(cuentas_map)} cuentas cargadas.")
    else:
        cuentas_map = {
            IEPS_GS: "IEPS De Gasolina Magna",
            IEPS_GP: "IEPS de Premium",
            IEPS_GD: "IEPS de Diesel",
        }
        _dyn_cli, _dyn_prods = [], []
        logs.append("  ℹ Sin plantilla — usando nombres predeterminados.")

    # Listas dinámicas (desde hoja CUENTAS) o fallback hardcoded
    if _dyn_cli:
        _cuentas_tpl  = [nc for nc, _ in _dyn_cli]
        _clientes_tpl = [nm for _, nm in _dyn_cli]
        logs.append(f"  {len(_clientes_tpl)} clientes leídos de hoja CUENTAS.")
    else:
        _cuentas_tpl  = CUENTAS_TPL
        _clientes_tpl = CLIENTES_TPL
    if len(_dyn_prods) == 7:
        _ctas_prod = [pc for pc, _ in _dyn_prods]
        _prods     = [pm for _, pm in _dyn_prods]
    else:
        _ctas_prod = CTAS_PROD
        _prods     = PRODS

    NOMS_PROD   = [cuentas_map.get(c, p) for c, p in zip(_ctas_prod, _prods)]

    # ── Acumular por fecha / cliente / producto ────────────────────────────
    cli_day   = defaultdict(float)
    prod_day  = defaultdict(float)
    iva_day   = defaultdict(float)
    ieps_prod = defaultdict(float)

    for r in data:
        try:
            fecha       = str(r[C_FECHA])[:10] if r[C_FECHA] is not None else ""
            cliente_raw = str(r[C_CLIENTE] or "").strip()
            cliente     = _match_cliente(cliente_raw, _clientes_tpl)
            prod        = str(r[C_PROD] or "")
            cli_day[(fecha, cliente)]  += float(r[C_IMPORTE]  or 0)
            prod_day[(fecha, prod)]    += float(r[C_SUBTOTAL] or 0)
            iva_day[fecha]             += float(r[C_IVA]      or 0)
            ieps_prod[(fecha, prod)]   += float(r[C_IEPS]     or 0)
        except Exception:
            continue

    fechas = sorted(set(k[0] for k in cli_day if k[0]))
    if not fechas:
        raise RuntimeError("No se encontraron datos de ventas en el archivo.")
    logs.append(f"  {len(fechas)} fecha(s) detectada(s): {fechas[0]} … {fechas[-1]}")

    # ── Diagnóstico clientes ───────────────────────────────────────────────
    all_cli_despachos = sorted(set(k[1] for k in cli_day.keys() if k[1]))
    logs.append(f"  Clientes únicos en despachos ({len(all_cli_despachos)}): {', '.join(all_cli_despachos)}")
    logs.append(f"  Cuentas cargadas de plantilla ({len(_clientes_tpl)}): {', '.join(_clientes_tpl)}")

    # Clientes en despachos sin cuenta en plantilla
    _sin_cuenta = [cli for cli in all_cli_despachos if cli not in _clientes_tpl]
    if _sin_cuenta:
        logs.append(f"  ⚠️ En despachos pero SIN cuenta en plantilla: {', '.join(_sin_cuenta)}")

    # Filtrar: solo clientes con importe > 0 en alguna fecha
    _activos = [(nc, nm) for nc, nm in zip(_cuentas_tpl, _clientes_tpl)
                if any(cli_day.get((f, nm), 0.0) for f in fechas)]
    _sin_datos = [nm for nm in _clientes_tpl
                  if not any(cli_day.get((f, nm), 0.0) for f in fechas)]
    if _sin_datos:
        logs.append(f"  ℹ️ Sin importe (se omiten): {', '.join(_sin_datos)}")
    if _activos:
        _cuentas_tpl = [nc for nc, _ in _activos]
        _clientes_tpl = [nm for _, nm in _activos]
    logs.append(f"  ✅ {len(_clientes_tpl)} cliente(s) con importe en la póliza.")

    # NOMBRES_TPL con lista ya filtrada
    NOMBRES_TPL = [cuentas_map.get(c, cli) for c, cli in zip(_cuentas_tpl, _clientes_tpl)]

    # ── Índices de columnas ────────────────────────────────────────────────
    N_META    = len(META_HDRS)      # 8
    N_CLI     = len(_clientes_tpl)  # dinámico desde hoja CUENTAS
    N_PROD    = len(_prods)         # 7 si dinámico, si no fallback
    OFF       = N_META            # 8  → inicio clientes
    COL_TOT1  = OFF + N_CLI       # 28 → TOTAL B2 clientes
    COL_PROD0 = OFF + N_CLI + 1   # 29 → inicio productos
    COL_TOT2  = OFF + N_CLI + 1 + N_PROD  # 36 → TOTAL B2 productos
    COL_CONC  = OFF + N_CLI + 1 + N_PROD + 1  # 37 → CONCILIACION
    TOTAL_COLS = COL_CONC + 1     # 38

    # ── Generar Excel con xlsxwriter ───────────────────────────────────────
    logs.append("⛽ Generando póliza Excel...")
    buf = io.BytesIO()
    wb  = xlsxwriter.Workbook(buf, {'in_memory': True})
    ws  = wb.add_worksheet("poliza IA")

    def fmt(d):
        return wb.add_format(d)

    CURR = 'General'
    BASE = {'font_name': 'Arial', 'font_size': 9, 'border': 1,
            'border_color': '#CCCCCC', 'valign': 'vcenter'}

    f_acct    = fmt({**BASE, 'bold': True, 'bg_color': '#3B0764', 'font_color': '#FFFFFF',
                     'align': 'center', 'font_size': 8})
    f_hdr_m   = fmt({**BASE, 'bold': True, 'bg_color': '#6A1B9A', 'font_color': '#FFFFFF',
                     'align': 'center', 'text_wrap': True, 'font_size': 8})
    f_hdr_c   = fmt({**BASE, 'bold': True, 'bg_color': '#1565C0', 'font_color': '#FFFFFF',
                     'align': 'center', 'text_wrap': True, 'font_size': 8})
    f_hdr_p   = fmt({**BASE, 'bold': True, 'bg_color': '#0D47A1', 'font_color': '#FFFFFF',
                     'align': 'center', 'text_wrap': True, 'font_size': 8})
    f_hdr_tot = fmt({**BASE, 'bold': True, 'bg_color': '#1B5E20', 'font_color': '#FFFFFF',
                     'align': 'center', 'font_size': 8})
    f_hdr_con = fmt({**BASE, 'bold': True, 'bg_color': '#E65100', 'font_color': '#FFFFFF',
                     'align': 'center', 'font_size': 8})
    f_fecha   = fmt({**BASE, 'bold': True, 'bg_color': '#F3E5F5', 'align': 'center', 'num_format': 'dd/mm/yyyy'})
    f_meta    = fmt({**BASE, 'bg_color': '#EDE7F6', 'align': 'left', 'font_size': 8})
    f_num0    = fmt({**BASE, 'bg_color': '#FFFFFF', 'align': 'right', 'num_format': CURR})
    f_num1    = fmt({**BASE, 'bg_color': '#B7D9EF', 'align': 'right', 'num_format': CURR})
    f_tot0    = fmt({**BASE, 'bold': True, 'bg_color': '#E6F2FB', 'align': 'right', 'num_format': CURR})
    f_tot1    = fmt({**BASE, 'bold': True, 'bg_color': '#DCEDC8', 'align': 'right', 'num_format': CURR})
    f_conc0   = fmt({**BASE, 'bold': True, 'bg_color': '#FFF9C4', 'font_color': '#E65100',
                     'align': 'right', 'num_format': CURR})
    f_conc1   = fmt({**BASE, 'bold': True, 'bg_color': '#FFF176', 'font_color': '#E65100',
                     'align': 'right', 'num_format': CURR})
    f_grand   = fmt({**BASE, 'bold': True, 'bg_color': '#1B5E20', 'font_color': '#FFFFFF',
                     'align': 'right', 'num_format': CURR, 'border': 2, 'border_color': '#000000'})
    f_grand_l = fmt({**BASE, 'bold': True, 'bg_color': '#1B5E20', 'font_color': '#FFFFFF',
                     'align': 'center', 'border': 2, 'border_color': '#000000'})
    f_grand_c = fmt({**BASE, 'bold': True, 'bg_color': '#E65100', 'font_color': '#FFFFFF',
                     'align': 'right', 'num_format': CURR, 'border': 2, 'border_color': '#000000'})

    # ── Fila 0: número de columna / número de cuenta ──────────────────────
    ws.set_row(0, 18)
    ws.set_row(1, 50)
    for c in range(N_META): ws.write(0, c, c + 1, f_acct)
    for i, acct in enumerate(_cuentas_tpl): ws.write(0, OFF + i, acct, f_acct)
    ws.write(0, COL_TOT1, COL_TOT1 + 1, f_acct)
    for i, acct in enumerate(_ctas_prod): ws.write(0, COL_PROD0 + i, acct, f_acct)
    ws.write(0, COL_TOT2, COL_TOT2 + 1, f_acct)
    ws.write(0, COL_CONC, COL_CONC + 1, f_acct)

    # ── Fila 1: encabezados ────────────────────────────────────────────────
    for i, h in enumerate(META_HDRS):
        ws.write(1, i, h, f_hdr_m)
    for i, nom in enumerate(NOMBRES_TPL):
        ws.write(1, OFF + i, nom, f_hdr_c)
    ws.write(1, COL_TOT1, "TOTAL B2", f_hdr_tot)
    for i, nom in enumerate(NOMS_PROD):
        ws.write(1, COL_PROD0 + i, nom, f_hdr_p)
    ws.write(1, COL_TOT2, "TOTAL B2", f_hdr_tot)
    ws.write(1, COL_CONC, "CONCILIACION", f_hdr_con)

    # ── Anchos de columna ──────────────────────────────────────────────────
    ws.set_column(0, 0, 14)
    ws.set_column(1, 1, 12)
    for c in range(2, N_META):
        ws.set_column(c, c, 20)
    for i in range(N_CLI):
        ws.set_column(OFF + i, OFF + i, 14)
    ws.set_column(COL_TOT1, COL_TOT1, 13)
    for i in range(N_PROD):
        ws.set_column(COL_PROD0 + i, COL_PROD0 + i, 13)
    ws.set_column(COL_TOT2, COL_TOT2, 13)
    ws.set_column(COL_CONC, COL_CONC, 14)
    ws.freeze_panes(2, 2)

    # ── Filas de datos ─────────────────────────────────────────────────────
    gran_cli  = [0.0] * N_CLI
    gran_tot1 = 0.0
    gran_prod = [0.0] * N_PROD
    gran_tot2 = 0.0
    gran_conc = 0.0

    for ri, fecha in enumerate(fechas):
        row = ri + 2
        fn  = f_num0 if ri % 2 == 0 else f_num1
        ft  = f_tot0 if ri % 2 == 0 else f_tot1
        fc  = f_conc0 if ri % 2 == 0 else f_conc1

        # Convertir fecha a objeto date para Excel
        try:
            _fd = datetime.strptime(fecha[:10], '%Y-%m-%d')
            fecha_display = _fd.strftime('%d/%m/%Y')
            _fecha_val = _fd.date()
        except Exception:
            fecha_display = fecha
            _fecha_val = fecha

        ws.write(row, 0, "D", f_meta)
        ws.write_datetime(row, 1, _fecha_val, f_fecha)
        ws.write(row, 2, "VENTA DEL DIA " + fecha_display, f_meta)
        ws.write(row, 3, "VENTA DEL DIA " + fecha_display, f_meta)
        for c in range(4, N_META):
            ws.write(row, c, "", f_meta)

        # Clientes
        total_b2 = 0.0
        for i, cli in enumerate(_clientes_tpl):
            v = round(cli_day.get((fecha, cli), 0.0), 2)
            ws.write(row, OFF + i, v if v else None, fn)
            total_b2 += v
            gran_cli[i] += v
        ws.write(row, COL_TOT1, round(total_b2, 2), ft)
        gran_tot1 += total_b2

        # Productos
        prod_vals = [
            prod_day.get((fecha, "GS"), 0.0),
            prod_day.get((fecha, "GP"), 0.0),
            prod_day.get((fecha, "GD"), 0.0),
            iva_day.get(fecha, 0.0),
            ieps_prod.get((fecha, "GS"), 0.0),
            ieps_prod.get((fecha, "GP"), 0.0),
            ieps_prod.get((fecha, "GD"), 0.0),
        ]
        total_prod = sum(prod_vals)
        for i, v in enumerate(prod_vals):
            ws.write(row, COL_PROD0 + i, round(v, 2) if v else None, fn)
            gran_prod[i] += v
        ws.write(row, COL_TOT2, round(total_prod, 2), ft)
        gran_tot2 += total_prod

        diferencia = round(total_b2 - total_prod, 2)
        ws.write(row, COL_CONC, diferencia, fc)
        gran_conc += diferencia

    # ── Fila totales generales ─────────────────────────────────────────────
    tr = len(fechas) + 2
    ws.merge_range(tr, 0, tr, N_META - 1, "TOTAL GENERAL", f_grand_l)
    for i in range(N_CLI):
        ws.write(tr, OFF + i, round(gran_cli[i], 2), f_grand)
    ws.write(tr, COL_TOT1, round(gran_tot1, 2), f_grand)
    for i in range(N_PROD):
        ws.write(tr, COL_PROD0 + i, round(gran_prod[i], 2), f_grand)
    ws.write(tr, COL_TOT2, round(gran_tot2, 2), f_grand)
    ws.write(tr, COL_CONC, round(gran_conc, 2), f_grand_c)

    wb.close()
    buf.seek(0)

    logs.append(f"✅ Póliza generada — {len(fechas)} fecha(s), {TOTAL_COLS} columnas.")
    if abs(gran_conc) < 0.02:
        logs.append(f"✅ CONCILIACION cuadra: {gran_conc:,.2f}")
    else:
        logs.append(f"⚠️  CONCILIACION con diferencia: {gran_conc:,.2f}")

    # ── Resumen por fecha (para mostrar en UI) ─────────────────────────────
    resumen = []
    for fecha in fechas:
        try:
            fd = datetime.strptime(fecha[:10], '%Y-%m-%d')
            fd_str = fd.strftime('%d/%m/%Y')
        except Exception:
            fd_str = fecha
        tot_cli  = sum(cli_day.get((fecha, cli), 0.0) for cli in _clientes_tpl)
        tot_prod = (
            prod_day.get((fecha, "GS"), 0.0) +
            prod_day.get((fecha, "GP"), 0.0) +
            prod_day.get((fecha, "GD"), 0.0) +
            iva_day.get(fecha, 0.0) +
            ieps_prod.get((fecha, "GS"), 0.0) +
            ieps_prod.get((fecha, "GP"), 0.0) +
            ieps_prod.get((fecha, "GD"), 0.0)
        )
        resumen.append({
            "Fecha": fd_str,
            "Total Clientes": round(tot_cli, 2),
            "Total Productos": round(tot_prod, 2),
            "Diferencia": round(tot_cli - tot_prod, 2),
        })

    return buf.read(), logs, resumen


# ══════════════════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════════════════

st.markdown("### 📂 Archivos de entrada")
col1, col2 = st.columns([1, 1])
with col1:
    despachos_file = st.file_uploader(
        "📊 Control de despachos (.xlsx / .xls)",
        type=["xlsx", "xls"],
        help="Archivo de control de despachos generado por el sistema de ventas.",
    )
with col2:
    plantilla_file = st.file_uploader(
        "📋 Plantilla de cuentas (.xlsx) — opcional",
        type=["xlsx"],
        help="Plantilla SINUBE con hoja 'CUENTAS'. Si no se proporciona, se usan nombres predeterminados.",
    )

st.markdown("")
generar = st.button(
    "⛽  Generar Póliza Ventas del Día",
    type="primary",
    disabled=despachos_file is None,
    use_container_width=True,
)

if despachos_file is None:
    st.info("👆 Selecciona al menos el archivo de control de despachos para comenzar.")

if generar and despachos_file is not None:
    with st.spinner("Procesando..."):
        try:
            plantilla_bytes = plantilla_file.read() if plantilla_file else None
            excel_bytes, logs, resumen = procesar_ventas(
                despachos_file.read(),
                despachos_file.name,
                plantilla_bytes,
            )

            base = despachos_file.name.rsplit('.', 1)[0]
            nombre_salida = f"poliza_ventas_dia_{base}.xlsx"

            st.success(f"✅ Póliza generada — {len(resumen)} fecha(s)")

            st.download_button(
                label="💾  Descargar póliza Excel",
                data=excel_bytes,
                file_name=nombre_salida,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            # ── Tabla resumen ──────────────────────────────────────────────
            if resumen:
                st.markdown("### 📊 Resumen por fecha")
                import pandas as pd
                df = pd.DataFrame(resumen)
                # Color diferencia
                def _color_diff(v):
                    if abs(v) < 0.02:
                        return "background-color:#D1FAE5; color:#065F46"
                    return "background-color:#FEF3C7; color:#92400E"

                styled = (
                    df.style
                    .format({
                        "Total Clientes":  "{:,.2f}",
                        "Total Productos": "{:,.2f}",
                        "Diferencia":      "{:,.2f}",
                    })
                    .map(_color_diff, subset=["Diferencia"])
                )
                st.dataframe(styled, use_container_width=True, hide_index=True)

            # ── Log ────────────────────────────────────────────────────────
            with st.expander("📋 Log de procesamiento"):
                for line in logs:
                    st.text(line)

        except Exception as exc:
            import traceback
            st.error(f"❌ Error al generar póliza: {exc}")
            with st.expander("Detalle del error"):
                st.code(traceback.format_exc())

st.markdown("---")
st.caption("Módulo Ventas del Día · AUXILIAR DE REGISTROS")
