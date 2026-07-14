"""AUXILIAR DE REGISTROS — Conciliación Control de Despacho vs SAT (CFDI/XML)"""
import io
import xml.etree.ElementTree as ET
from collections import defaultdict

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(
    page_title="Conciliación SAT",
    page_icon="🔗",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
[data-testid="stAppViewContainer"] { background: #dbeafe; }
    .hdr { background:#0A7A4A; padding:18px 24px; border-radius:10px; margin-bottom:20px; }
    .hdr h2 { color:#fff; margin:0; font-size:1.3rem; }
    .hdr p  { color:#A7F3D0; margin:4px 0 0; font-size:.9rem; }
    #MainMenu { visibility:hidden; }
    footer    { visibility:hidden; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hdr">
  <h2>🔗 Conciliación Control de Despacho vs SAT (CFDI/XML)</h2>
  <p>Compara los UUIDs de los XMLs del SAT contra la columna FolioFiscal del Excel de despachos.</p>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# LÓGICA
# ─────────────────────────────────────────────────────────────────────────────
NS = {
    "cfdi": "http://www.sat.gob.mx/cfd/4",
    "tfd":  "http://www.sat.gob.mx/TimbreFiscalDigital",
}


def _parsear_xmls(xml_files):
    """Lee los UploadedFile de XML y retorna lista de dicts."""
    data, logs = [], []
    for f in xml_files:
        try:
            root = ET.parse(io.BytesIO(f.read())).getroot()
            tfd  = root.find(".//tfd:TimbreFiscalDigital", NS)
            uuid = tfd.get("UUID", "").strip().upper() if tfd is not None else ""
            rec  = root.find("cfdi:Receptor", NS)
            conc = root.find(".//cfdi:Concepto", NS)
            imp  = root.find(".//cfdi:Impuestos", NS)
            subtotal = float(root.get("SubTotal", 0))
            total    = float(root.get("Total", 0))
            iva      = float(imp.get("TotalImpuestosTrasladados", 0)) if imp is not None else 0
            if iva == 0 and total > subtotal:
                iva = round(total - subtotal, 2)
            data.append({
                "uuid":     uuid,
                "fecha":    root.get("Fecha", "")[:10],
                "serie":    root.get("Serie", ""),
                "folio":    root.get("Folio", ""),
                "receptor": rec.get("Nombre", "") if rec is not None else "",
                "rfc_rec":  rec.get("Rfc", "")   if rec is not None else "",
                "concepto": conc.get("Descripcion", "") if conc is not None else "",
                "subtotal": subtotal,
                "iva":      iva,
                "total":    total,
                "archivo":  f.name,
            })
            logs.append(f"  ✔ {f.name}  UUID: {uuid[:8]}...")
        except Exception as e:
            logs.append(f"  ✘ {f.name}: {e}")
    return data, logs


def _leer_excel(excel_bytes, excel_name):
    """Lee el Excel de Control de Despachos y retorna (excel_rows, logs)."""
    logs = []
    excel_rows = []
    folio_col_idx = None
    imp_col_idx   = None

    is_xls = excel_name.lower().endswith(".xls")
    if is_xls:
        try:
            import xlrd
        except ImportError:
            import subprocess, sys
            subprocess.check_call([sys.executable, "-m", "pip", "install", "xlrd"],
                                  stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            import xlrd
        bk  = xlrd.open_workbook(file_contents=excel_bytes)
        sht = bk.sheet_by_index(0)
        hdr_row = 0
        for ri in range(min(10, sht.nrows)):
            row_vals = [str(sht.cell_value(ri, ci)).strip() for ci in range(sht.ncols)]
            matches = [ci for ci, v in enumerate(row_vals)
                       if "foliofiscal" in v.lower().replace(" ", "")
                       or v.lower() in ("uuid", "folio fiscal", "folio_fiscal", "uidd")]
            if matches:
                hdr_row = ri
                folio_col_idx = matches[0]
                logs.append(f"  ✔ Columna FolioFiscal: col {folio_col_idx+1} "
                             f"('{sht.cell_value(ri, folio_col_idx)}')")
                break
        if folio_col_idx is None:
            return None, ["  ✘ No se encontró columna FolioFiscal en el Excel"]
        # Col de importe
        for ci in range(sht.ncols):
            hv = str(sht.cell_value(hdr_row, ci)).strip().lower()
            if hv in ("importe", "total", "subtotal"):
                imp_col_idx = ci
                break
        for ri in range(hdr_row + 1, sht.nrows):
            uuid_xl = str(sht.cell_value(ri, folio_col_idx)).strip().upper()
            if not uuid_xl or uuid_xl in ("NONE", ""):
                continue
            importe = 0
            if imp_col_idx is not None:
                try: importe = float(sht.cell_value(ri, imp_col_idx) or 0)
                except: pass
            excel_rows.append({"uuid": uuid_xl, "importe": importe})
    else:
        from openpyxl import load_workbook
        wb_r = load_workbook(io.BytesIO(excel_bytes), data_only=True)
        ws_r = wb_r.active
        hdr_row_xl = None
        for ri in range(1, min(12, ws_r.max_row + 1)):
            for ci in range(1, ws_r.max_column + 1):
                v = str(ws_r.cell(ri, ci).value or "").strip()
                if ("foliofiscal" in v.lower().replace(" ", "")
                        or v.lower() in ("uuid", "folio fiscal", "folio_fiscal", "uidd")):
                    folio_col_idx = ci - 1
                    hdr_row_xl    = ri
                    logs.append(f"  ✔ Columna FolioFiscal: col {ci} ('{v}')")
                    break
            if folio_col_idx is not None:
                break
        if folio_col_idx is None:
            return None, ["  ✘ No se encontró columna FolioFiscal"]
        for ci, cell in enumerate(ws_r[hdr_row_xl]):
            if str(cell.value or "").strip().lower() in ("importe", "total", "subtotal"):
                imp_col_idx = ci
                break
        for row in ws_r.iter_rows(min_row=hdr_row_xl + 1, values_only=True):
            uuid_xl = str(row[folio_col_idx] or "").strip().upper()
            if not uuid_xl or uuid_xl == "NONE":
                continue
            importe = 0
            if imp_col_idx is not None:
                try: importe = float(row[imp_col_idx] or 0)
                except: pass
            excel_rows.append({"uuid": uuid_xl, "importe": importe})

    logs.append(f"  ✔ {len(excel_rows)} filas leídas")
    return excel_rows, logs


def _generar_excel(xmls_data, excel_rows):
    """Genera el Excel de conciliación. Retorna bytes."""
    AZUL  = "1E3A5F"; AZUL2 = "2E75B6"; AZUL3 = "D6E4F0"
    VERDE = "C8E6C9"; ROJO  = "FFCDD2"; AMBAR = "FFF9C4"; GRIS = "F5F5F5"
    thin  = Side(style="thin", color="BBBBBB")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cx(cell, v=None, bg=None, fg="000000", bold=False, sz=9,
           align="left", nf=None, wrap=False, border=True):
        if v is not None:
            cell.value = v
        cell.font      = Font(name="Arial", bold=bold, color=fg, size=sz)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        if border:
            cell.border = brd
        if nf:
            cell.number_format = nf

    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliación UUID"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:M1")
    cx(ws["A1"], "CONCILIACIÓN  CONTROL DE DESPACHO  vs  SAT (CFDI/XML)",
       bg=AZUL, fg="FFFFFF", bold=True, sz=13, align="center", border=False)
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:M2")
    cx(ws["A2"], "Match: UUID (XML) = FolioFiscal (Excel)   |   SUPER SERVICIO PERIFERICO",
       bg="0A7A4A", fg="FFFFFF", sz=9, align="center", border=False)
    ws.row_dimensions[2].height = 18

    HDRS = [
        ("No.", 5), ("Folio Fiscal\n(UUID — XML)", 42),
        ("Folio Fiscal\n(UIDD — Excel)", 42),
        ("¿Coincide?", 12), ("Fecha", 12),
        ("Receptor XML", 28), ("Concepto", 26),
        ("SubTotal\nXML", 14), ("IVA\nXML", 12), ("Total\nXML", 14),
        ("Importe\nExcel", 14), ("Diferencia", 14), ("Estado", 16),
    ]
    for j, (h, w) in enumerate(HDRS, 1):
        cx(ws.cell(4, j), h, bg=AZUL2, fg="FFFFFF", bold=True, sz=9,
           align="center", wrap=True)
        ws.column_dimensions[ws.cell(4, j).column_letter].width = w
    ws.row_dimensions[4].height = 36

    # Índice UUID → lista importes Excel
    excel_idx = defaultdict(list)
    for r in excel_rows:
        excel_idx[r["uuid"]].append(r["importe"])

    ok_count  = 0
    resumen   = []

    for i, x in enumerate(xmls_data, start=5):
        bg   = GRIS if i % 2 == 0 else "FFFFFF"
        uuid = x["uuid"]
        found  = uuid in excel_idx
        imp_xl = sum(excel_idx[uuid]) if found else 0
        diff   = imp_xl - x["total"]
        m_bg   = VERDE if found else ROJO
        m_txt  = "✔  SÍ" if found else "✘  NO"
        d_bg   = VERDE if abs(diff) < 0.02 else (AMBAR if abs(diff) < 100 else ROJO)
        if found and abs(diff) < 0.02:
            est, est_bg = "✔ Conciliado", VERDE
        elif found and abs(diff) < 100:
            est, est_bg = "⚠ Dif. menor", AMBAR
        elif found:
            est, est_bg = "⚠ Dif. mayor", ROJO
        else:
            est, est_bg = "✘ No encontrado", ROJO
        if found:
            ok_count += 1

        row_data = [
            (i - 4, bg, "center", None),
            (uuid,  bg, "left",   None),
            (uuid if found else "", m_bg, "left", None),
            (m_txt, m_bg, "center", None),
            (x["fecha"],    bg,  "center", None),
            (x["receptor"], bg,  "left",   None),
            (x["concepto"], bg,  "left",   None),
            (x["subtotal"], bg,  "right",  "$#,##0.00"),
            (x["iva"],      bg,  "right",  "$#,##0.00"),
            (x["total"],    bg,  "right",  "$#,##0.00"),
            (imp_xl if found else None, bg,  "right", "$#,##0.00"),
            (diff   if found else None, d_bg,"right", "$#,##0.00;[Red]($#,##0.00)"),
            (est, est_bg, "center", None),
        ]
        for j, (v, cbg, al, nf) in enumerate(row_data, 1):
            cx(ws.cell(i, j), v, bg=cbg, align=al, nf=nf, sz=9)
        ws.row_dimensions[i].height = 20

        resumen.append({
            "No.":          i - 4,
            "UUID (XML)":   uuid[:20] + "..." if len(uuid) > 20 else uuid,
            "Fecha":        x["fecha"],
            "Receptor":     x["receptor"][:30],
            "Total XML":    x["total"],
            "Importe Excel": imp_xl if found else None,
            "Diferencia":   round(diff, 2) if found else None,
            "Estado":       est,
        })

    # Fila totales
    tr = 5 + len(xmls_data)
    ws.merge_cells(f"A{tr}:G{tr}")
    cx(ws.cell(tr, 1), "TOTALES", bg=AZUL, fg="FFFFFF", bold=True, sz=10, align="right")
    for j, v, nf in [
        (8,  sum(x["subtotal"] for x in xmls_data), "$#,##0.00"),
        (9,  sum(x["iva"]      for x in xmls_data), "$#,##0.00"),
        (10, sum(x["total"]    for x in xmls_data), "$#,##0.00"),
        (11, sum(sum(excel_idx.get(x["uuid"], [])) for x in xmls_data), "$#,##0.00"),
    ]:
        cx(ws.cell(tr, j), v, bg=AZUL, fg="FFFFFF", bold=True, sz=10, align="right", nf=nf)
    ws.row_dimensions[tr].height = 22

    nr = tr + 2
    ws.merge_cells(f"A{nr}:M{nr}")
    cx(ws.cell(nr, 1),
       f"Procesados: {len(xmls_data)} XMLs  |  "
       f"Encontrados: {ok_count}  |  No encontrados: {len(xmls_data)-ok_count}",
       bg=AZUL3, sz=9, wrap=True, border=False)
    ws.row_dimensions[nr].height = 24

    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read(), resumen, ok_count


# ─────────────────────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("### 📁 Archivos de entrada")
col1, col2 = st.columns([1, 1])
with col1:
    excel_file = st.file_uploader(
        "📄 Control de Despachos (.xlsx / .xls)",
        type=["xlsx", "xls"],
        help="Debe contener una columna FolioFiscal, UUID o UIDD.",
    )
with col2:
    xml_files = st.file_uploader(
        "📑 Archivos XML (CFDI del SAT)",
        type=["xml"],
        accept_multiple_files=True,
        help="Selecciona uno o varios CFDIs (.xml).",
    )

st.markdown("")
generar = st.button(
    "🔗  Generar Conciliación",
    type="primary",
    disabled=(excel_file is None or not xml_files),
    use_container_width=True,
)

if not excel_file or not xml_files:
    st.info("📋 Carga el Excel de Control de Despachos y al menos un XML para comenzar.")

if generar and excel_file and xml_files:
    with st.spinner("Procesando..."):
        logs_all = []
        try:
            # 1. Parsear XMLs
            logs_all.append("📑 Leyendo XMLs...")
            xmls_data, xml_logs = _parsear_xmls(xml_files)
            logs_all.extend(xml_logs)
            logs_all.append(f"  Total: {len(xmls_data)} XMLs leídos")

            if not xmls_data:
                st.error("✘ No se pudo leer ningún XML válido.")
            else:
                # 2. Leer Excel
                logs_all.append("\n📄 Leyendo Excel...")
                excel_bytes  = excel_file.read()
                excel_rows, xl_logs = _leer_excel(excel_bytes, excel_file.name)
                logs_all.extend(xl_logs)

                if excel_rows is None:
                    st.error("\n".join(xl_logs))
                else:
                    logs_all.append(f"  Total: {len(excel_rows)} filas en Excel")

                    # 3. Generar Excel
                    logs_all.append("\n📊 Conciliando y generando reporte...")
                    excel_out, resumen, ok_count = _generar_excel(xmls_data, excel_rows)

                    total_xml = sum(x["total"] for x in xmls_data)
                    no_enc    = len(xmls_data) - ok_count

                    logs_all.append(f"\n{'='*60}")
                    logs_all.append(f"  XMLs procesados   : {len(xmls_data)}")
                    logs_all.append(f"  Encontrados       : {ok_count} ✔")
                    logs_all.append(f"  No encontrados    : {no_enc} ✘")
                    logs_all.append(f"  Total XMLs        : ${total_xml:,.2f}")
                    logs_all.append(f"{'='*60}")

                    # Métricas
                    m1, m2, m3, m4 = st.columns(4)
                    m1.metric("XMLs procesados", len(xmls_data))
                    m2.metric("✔ Encontrados",   ok_count)
                    m3.metric("✘ No encontrados", no_enc)
                    m4.metric("Total XML",        f"${total_xml:,.2f}")

                    st.success(f"✅ Conciliación generada — {ok_count}/{len(xmls_data)} encontrados")

                    nombre_salida = f"Conciliacion_DespachoVsSAT_{excel_file.name.rsplit('.',1)[0]}.xlsx"
                    st.download_button(
                        label="📥  Descargar Excel de Conciliación",
                        data=excel_out,
                        file_name=nombre_salida,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                    # Tabla resumen
                    if resumen:
                        st.markdown("### 📊 Resumen")
                        import pandas as pd
                        df = pd.DataFrame(resumen)

                        def _color_estado(v):
                            if "Conciliado" in str(v):
                                return "background-color:#D1FAE5; color:#065F46"
                            if "⚠" in str(v):
                                return "background-color:#FEF3C7; color:#92400E"
                            return "background-color:#FEE2E2; color:#991B1B"

                        styled = (
                            df.style
                            .format({
                                "Total XML":      "{:,.2f}",
                                "Importe Excel":  "{:,.2f}",
                                "Diferencia":     "{:,.2f}",
                            }, na_rep="—")
                            .map(_color_estado, subset=["Estado"])
                        )
                        st.dataframe(styled, use_container_width=True, hide_index=True)

        except Exception as exc:
            import traceback
            st.error(f"✘ Error: {exc}")
            with st.expander("Detalle del error"):
                st.code(traceback.format_exc())
            logs_all.append(f"\n✘ ERROR: {exc}")

        with st.expander("📋 Log de procesamiento"):
            st.text("\n".join(logs_all))

st.markdown("---")
st.caption("Módulo Conciliación SAT · AUXILIAR DE REGISTROS")
