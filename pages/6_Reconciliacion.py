"""AUXILIAR DE REGISTROS — Reconciliación CSV → Excel (SINUBE)"""
import streamlit as st
import io, csv as _csv_mod, re as _re, unicodedata as _ud
from datetime import datetime

st.set_page_config(page_title="Reconciliación · Auxiliar", page_icon="📑", layout="wide")

import _theme
_theme.aplicar_header("📑 Reconciliación", "Reconciliación contable con plantilla SINUBE")
try:
    import openpyxl
    from openpyxl.styles import PatternFill as _PF, Alignment as _AL, Font as _Fnt
    from openpyxl.utils import get_column_letter as _gcl
except ImportError:
    st.error("❌ `openpyxl` no está instalado.")
    st.stop()


# ─── Función principal ─────────────────────────────────────────────────────────
def procesar_reconciliacion(plantilla_bytes, csv_dict):
    """
    plantilla_bytes : bytes del Excel plantilla, o None
    csv_dict        : {nombre: bytes} de los CSV
    Retorna (wb, resultados, logs)
    """
    logs = []
    _BAL_ACCT = "101-01-0003-0001"

    def _limpiar_val(v):
        if v is None: return ""
        s = str(v).strip()
        m = _re.match(r'^=?"?([^"=]+)"?$', s)
        if m: s = m.group(1).strip()
        if s.startswith("'"): s = s[1:].strip()
        return s

    def _es_cuenta_real(s):
        return bool(s) and s[0].isdigit()

    def _norm(s):
        return _ud.normalize("NFD", s.upper().strip()).encode("ascii","ignore").decode()

    _aliases_ing = {}
    _aliases_vta = {}

    def _buscar_alias(vals_dict, header, aliases_dict):
        hU = header.upper()
        hN = _norm(header)
        if header in vals_dict: return vals_dict[header]
        for k, kv in vals_dict.items():
            if k.upper() == hU: return kv
        for k, kv in vals_dict.items():
            if _norm(k) == hN: return kv
        return None

    cuentas_ing_list = []; cuentas_vta_list = []
    cuentas_ing_by_acct = {}; cuentas_vta_by_acct = {}
    cuentas_vta_acct_set = set()
    col_to_name = {}; col_to_acct = {}
    _bal_col = None; _simplificada = False

    # ── MODO PLANTILLA ─────────────────────────────────────────────────────────
    if plantilla_bytes:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(plantilla_bytes))
            ws = wb["POLIZA"] if "POLIZA" in wb.sheetnames else wb.active
            if "POLIZA" not in wb.sheetnames:
                logs.append(f"Hoja 'POLIZA' no encontrada, usando '{ws.title}'")
        except Exception as e:
            logs.append(f"Error abriendo plantilla: {e}")
            return None, [], logs

        # Leer hoja CUENTAS
        _cuentas_sheet = next((s for s in wb.sheetnames if s.strip().upper() == "CUENTAS"), None)
        if _cuentas_sheet:
            try:
                _wb_do = openpyxl.load_workbook(io.BytesIO(plantilla_bytes), data_only=True)
                _cs_do = next((s for s in _wb_do.sheetnames if s.strip().upper() == "CUENTAS"), None)
                wc = _wb_do[_cs_do] if _cs_do else wb[_cuentas_sheet]
            except Exception:
                wc = wb[_cuentas_sheet]

            _seen_ing = set(); _seen_vta = set()
            for r in range(1, (wc.max_row or 50) + 1):
                a = wc.cell(r,1).value; n = wc.cell(r,2).value
                if a is not None and n is not None:
                    an = _limpiar_val(a); nn = _limpiar_val(n)
                    if _es_cuenta_real(an) and nn and nn.upper() not in _seen_ing:
                        _seen_ing.add(nn.upper())
                        cuentas_ing_list.append((an, nn))
                        cuentas_ing_by_acct[an] = nn
                va = wc.cell(r,5).value; vn = wc.cell(r,6).value
                if va is not None and vn is not None:
                    van = _limpiar_val(va); vnn = _limpiar_val(vn)
                    if _es_cuenta_real(van) and vnn and vnn.upper() not in _seen_vta:
                        _seen_vta.add(vnn.upper())
                        cuentas_vta_list.append((van, vnn))
                        cuentas_vta_by_acct[van] = vnn
                    if _es_cuenta_real(van):
                        cuentas_vta_acct_set.add(van)
            logs.append(f"CUENTAS: {len(cuentas_ing_list)} ING, {len(cuentas_vta_list)} VTA")

        # Agregar productos CSV extra no en CUENTAS
        if cuentas_ing_by_acct or cuentas_vta_by_acct:
            _pre_skip = {"descripcion","islas","total islas","total estacion",
                         "total impuestos","total ingresos","impuestos",""}
            _csv_ing_extra = []; _csv_vta_extra = []
            _ing_up = {nn.upper() for _,nn in cuentas_ing_list}
            _vta_up = {nn.upper() for _,nn in cuentas_vta_list}
            for _name, _bytes in sorted(csv_dict.items()):
                _sx = None
                try:
                    for _rx in _csv_mod.reader(io.StringIO(_bytes.decode("latin-1"))):
                        if not _rx: continue
                        _nx = _rx[0].strip(); _nlx = _nx.lower()
                        if _nlx == "ingresos": _sx = "ING"; continue
                        if _nlx in ("islas","impuestos"): _sx = "VTA"; continue
                        if not _nlx or _nlx in _pre_skip or _nlx.startswith("total"): continue
                        if len(_rx) > 3 and _rx[3].strip():
                            try:
                                float(_rx[3].strip().replace(",",""))
                                if _sx == "ING" and _nx.upper() not in _ing_up:
                                    _csv_ing_extra.append(_nx); _ing_up.add(_nx.upper())
                                elif _sx == "VTA" and _nx.upper() not in _vta_up:
                                    _csv_vta_extra.append(_nx); _vta_up.add(_nx.upper())
                            except ValueError: pass
                except: pass
            for nm in _csv_ing_extra: cuentas_ing_list.append(("", nm)); cuentas_ing_by_acct[nm] = nm
            for nm in _csv_vta_extra: cuentas_vta_list.append(("", nm)); cuentas_vta_by_acct[nm] = nm
            if _csv_ing_extra or _csv_vta_extra:
                logs.append(f"Productos CSV extra: +{len(_csv_ing_extra)} ING, +{len(_csv_vta_extra)} VTA")

        # Detección dinámica columnas fila 3
        proc_col = 8
        scan_up_to = max((ws.max_column or 30) + 10, 40)
        for col in range(1, scan_up_to):
            a2 = ws.cell(row=2, column=col).value
            if a2: col_to_acct[col] = str(a2).strip()
        for col in range(1, scan_up_to):
            h = ws.cell(row=3, column=col).value
            if h:
                hn = str(h).strip()
                col_to_name[col] = hn
                if hn == "PROCESADO": proc_col = col

        _re_tot = _re.compile(r'^TOTAL\b', _re.IGNORECASE)
        total_cols = sorted([c for c,n in col_to_name.items()
                             if _re_tot.match(n) and c > proc_col])
        _t1 = total_cols[0] if total_cols else proc_col + 11
        _t2 = total_cols[1] if len(total_cols) >= 2 else _t1 + 10
        _simplificada = (_t1 - proc_col) < 5
        tot_ig_col = _t1; tot_vta_col = _t2
        conc_col = next((c for c,n in col_to_name.items()
                         if "CONCILIACION" in n and c > tot_vta_col), tot_vta_col + 1)
        ing_start = proc_col + 1
        vta_start = tot_ig_col + 1

        # Si simplificada sin CUENTAS: pre-escanear CSV
        if _simplificada and not (cuentas_ing_by_acct or cuentas_vta_by_acct):
            _pre2 = {"descripcion","islas","total islas","total estacion",
                     "total impuestos","total ingresos","impuestos",""}
            _ci=[]; _cv=[]; _ci_s=set(); _cv_s=set()
            for _n, _b in sorted(csv_dict.items()):
                _s2=None
                try:
                    for _r2 in _csv_mod.reader(io.StringIO(_b.decode("latin-1"))):
                        if not _r2: continue
                        _n2=_r2[0].strip(); _nl2=_n2.lower()
                        if _nl2=="ingresos": _s2="ING"; continue
                        if _nl2 in ("islas","impuestos"): _s2="VTA"; continue
                        if not _nl2 or _nl2 in _pre2 or _nl2.startswith("total"): continue
                        if len(_r2)>3 and _r2[3].strip():
                            try:
                                float(_r2[3].strip().replace(",",""))
                                if _s2=="ING" and _n2 not in _ci_s: _ci.append(_n2);_ci_s.add(_n2)
                                elif _s2=="VTA" and _n2 not in _cv_s: _cv.append(_n2);_cv_s.add(_n2)
                            except ValueError: pass
                except: pass
            cuentas_ing_list=[("",nm) for nm in _ci]; cuentas_vta_list=[("",nm) for nm in _cv]
            cuentas_ing_by_acct={nm:nm for nm in _ci}; cuentas_vta_by_acct={nm:nm for nm in _cv}
            logs.append(f"CUENTAS desde CSV: {len(_ci)} ING, {len(_cv)} VTA")

        # Expandir si simplificada con CUENTAS
        if _simplificada and (cuentas_ing_by_acct or cuentas_vta_by_acct):
            n_ing=len(cuentas_ing_list); n_vta=len(cuentas_vta_list)
            if n_ing > 0:
                ws.insert_cols(tot_ig_col, n_ing)
                for i,(acct,nombre) in enumerate(cuentas_ing_list):
                    col=proc_col+1+i
                    ws.cell(2,col).value=acct; ws.cell(3,col).value=nombre
                    col_to_name[col]=nombre; col_to_acct[col]=acct
                tot_ig_col+=n_ing; tot_vta_col+=n_ing; conc_col+=n_ing
            if n_vta > 0:
                ws.insert_cols(tot_vta_col, n_vta)
                for i,(acct,nombre) in enumerate(cuentas_vta_list):
                    col=tot_ig_col+1+i
                    ws.cell(2,col).value=acct; ws.cell(3,col).value=nombre
                    col_to_name[col]=nombre; col_to_acct[col]=acct
                tot_vta_col+=n_vta; conc_col+=n_vta
            vta_start=tot_ig_col+1
            _simplificada=False
            logs.append(f"Expandida: {n_ing} ING + {n_vta} VTA")

        # Insertar columna BAL (efectivo cuenta dif)
        _bal_col = next((c for c in range(ing_start, tot_ig_col)
                         if col_to_acct.get(c) == _BAL_ACCT), None)
        if _bal_col is None and not _simplificada:
            ws.insert_cols(tot_ig_col, 1)
            _bal_col = tot_ig_col
            tot_ig_col += 1; tot_vta_col += 1; conc_col += 1
            vta_start += 1
            for _c in sorted([c for c in col_to_name if c >= _bal_col], reverse=True):
                col_to_name[_c+1] = col_to_name.pop(_c)
            for _c in sorted([c for c in col_to_acct if c >= _bal_col], reverse=True):
                col_to_acct[_c+1] = col_to_acct.pop(_c)
            ws.cell(2,_bal_col).value = _BAL_ACCT
            ws.cell(3,_bal_col).value = "efectivo cuenta dif"
            col_to_name[_bal_col] = "efectivo cuenta dif"
            col_to_acct[_bal_col] = _BAL_ACCT
            logs.append(f"Col BAL insertada en {_gcl(_bal_col)}")

    # ── MODO AUTO (sin plantilla) ──────────────────────────────────────────────
    else:
        _pre = {"descripcion","islas","total islas","total estacion",
                "total impuestos","total ingresos","impuestos",""}
        _ai=[]; _av=[]; _ai_s=set(); _av_s=set()
        for _n, _b in sorted(csv_dict.items()):
            _sec=None
            try:
                for _row in _csv_mod.reader(io.StringIO(_b.decode("latin-1"))):
                    if not _row: continue
                    _nm=_row[0].strip(); _nl=_nm.lower()
                    if _nl=="ingresos": _sec="ING"; continue
                    if _nl in ("islas","impuestos"): _sec="VTA"; continue
                    if not _nl or _nl in _pre or _nl.startswith("total"): continue
                    if len(_row)>3 and _row[3].strip():
                        try:
                            float(_row[3].strip().replace(",",""))
                            if _sec=="ING" and _nm not in _ai_s: _ai.append(_nm);_ai_s.add(_nm)
                            elif _sec=="VTA" and _nm not in _av_s: _av.append(_nm);_av_s.add(_nm)
                        except ValueError: pass
            except: pass

        n_ing=len(_ai); n_vta=len(_av)
        logs.append(f"Auto: {n_ing} cuentas ING, {n_vta} cuentas VTA")
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="POLIZA"
        proc_col=8
        tot_ig_col=proc_col+1+n_ing; tot_vta_col=tot_ig_col+1+n_vta; conc_col=tot_vta_col+1
        for _c,_h in enumerate(["TIPO DE POLIZA","Fecha","REFERENCIA","CONCEPTO",
                                 "ERROR","UIDD","NUM POLIZA","PROCESADO"],start=1):
            ws.cell(3,_c).value=_h; col_to_name[_c]=_h
        ing_start=proc_col+1
        for _i,_nm in enumerate(_ai):
            col=ing_start+_i; ws.cell(3,col).value=_nm; col_to_name[col]=_nm
        ws.cell(3,tot_ig_col).value="TOTAL INGRESOS"; col_to_name[tot_ig_col]="TOTAL INGRESOS"
        vta_start=tot_ig_col+1
        for _i,_nm in enumerate(_av):
            col=vta_start+_i; ws.cell(3,col).value=_nm; col_to_name[col]=_nm
        ws.cell(3,tot_vta_col).value="TOTAL VENTAS"; col_to_name[tot_vta_col]="TOTAL VENTAS"
        ws.cell(3,conc_col).value="CONCILIACION"; col_to_name[conc_col]="CONCILIACION"
        cuentas_ing_list=[]; cuentas_vta_list=[]; cuentas_ing_by_acct={}; cuentas_vta_by_acct={}
        _simplificada=False; _bal_col=None

    # ── Variables de letras (con ing_ltr_e = tot_ig_col-1 para incluir BAL) ───
    ing_ltr_s = _gcl(ing_start)
    ing_ltr_e = _gcl(tot_ig_col - 1)   # correcto incluso con BAL insertado
    vta_ltr_s = _gcl(vta_start)
    vta_ltr_e = _gcl(tot_vta_col - 1)
    tig_ltr   = _gcl(tot_ig_col)
    tva_ltr   = _gcl(tot_vta_col)

    # Renumerar fila 1
    _last_col = max(conc_col, tot_vta_col + 1)
    for _c in range(1, _last_col + 1):
        ws.cell(1, _c).value = _c - 1

    # Colorear encabezados fila 3
    ws.cell(3, conc_col).value = "CONCILIACION"
    def _hdr(col, bg):
        c = ws.cell(3, col)
        c.fill = _PF("solid", fgColor=bg)
        c.font = _Fnt(bold=True, color="FFFFFF", size=9)
        c.alignment = _AL(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, proc_col+1):        _hdr(col, "595959")
    for col in range(ing_start, tot_ig_col): _hdr(col, "2E75B6")
    _hdr(tot_ig_col, "1F4E79")
    for col in range(vta_start, tot_vta_col): _hdr(col, "C55A11")
    _hdr(tot_vta_col, "833C00")
    _hdr(conc_col, "7030A0")

    # Mapa fecha → fila existente
    fecha_fila = {}
    for r in range(4, (ws.max_row or 4) + 1):
        v = ws.cell(r, 2).value
        if v:
            if isinstance(v, datetime):
                fecha_fila[v.strftime("%d/%m/%Y")] = r
            elif isinstance(v, str) and "/" in v:
                fecha_fila[v.strip()] = r
    next_empty = (max(fecha_fila.values()) + 1) if fecha_fila else 4

    SKIP = {"descripcion","estacion gas122 s.a. de c.v.","islas",
            "total islas","total estacion","total impuestos",
            "total ingresos","impuestos",""}

    def _fill_colors(r):
        def _s(c, bg, h="right"):
            c.fill = _PF("solid", fgColor=bg)
            c.alignment = _AL(horizontal=h, vertical="center")
        for col in range(1, proc_col+1): _s(ws.cell(r,col), "F2F2F2", "left")
        for col in range(ing_start, tot_ig_col): _s(ws.cell(r,col), "DEEAF1")
        _s(ws.cell(r, tot_ig_col), "BDD7EE")
        for col in range(vta_start, tot_vta_col): _s(ws.cell(r,col), "FDF2E9")
        _s(ws.cell(r, tot_vta_col), "FAD7AC")
        _s(ws.cell(r, conc_col), "EAD1DC")

    _tree_ing_hdrs = [col_to_name.get(c, f"ING{c}") for c in range(ing_start, tot_ig_col)
                      if col_to_acct.get(c) != _BAL_ACCT]
    _tree_vta_hdrs = [col_to_name.get(c, f"VTA{c}") for c in range(vta_start, tot_vta_col)
                      if col_to_acct.get(c) != _BAL_ACCT and col_to_name.get(c)]
    _tree_vta_src  = [c for c in range(vta_start, tot_vta_col)
                      if col_to_acct.get(c) != _BAL_ACCT and col_to_name.get(c)]

    resultados = []
    data_row = next_empty

    for csv_name, csv_bytes in sorted(csv_dict.items()):
        csv_vals = {}; ing_vals = {}; vta_vals = {}
        fecha_str = None; _seccion = None
        try:
            for row in _csv_mod.reader(io.StringIO(csv_bytes.decode("latin-1"))):
                if not row: continue
                name = row[0].strip(); nl = name.lower()
                if not fecha_str:
                    m = _re.search(r"(\d{2}/\d{2}/\d{4})", row[0])
                    if m: fecha_str = m.group(1); continue
                if nl == "ingresos": _seccion = "ING"; continue
                if nl in ("islas","impuestos"): _seccion = "VTA"; continue
                if not nl or nl in SKIP or nl.startswith("total"): continue
                if len(row) > 3 and row[3].strip():
                    try:
                        val = float(row[3].strip().replace(",",""))
                        csv_vals[name] = val
                        if _seccion == "ING": ing_vals[name] = val
                        elif _seccion == "VTA": vta_vals[name] = val
                    except ValueError: pass
        except Exception as e:
            resultados.append((csv_name, "---", "---", "---", f"Error: {e}"))
            continue

        if not fecha_str:
            resultados.append((csv_name, "---", "---", "---", "Sin fecha"))
            continue

        d_p, mo_p, y_p = fecha_str.split("/")
        fecha_dt = datetime(int(y_p), int(mo_p), int(d_p))
        ref_text = f"VENTA DEL {fecha_str}"

        wr = fecha_fila.get(fecha_str, data_row)
        if wr == data_row: data_row += 1

        for col in range(1, conc_col + 2):
            ws.cell(wr, col).value = None

        ws.cell(wr, 1).value = "CLI"
        ws.cell(wr, 2).value = fecha_dt
        ws.cell(wr, 2).number_format = "DD/MM/YYYY"
        ws.cell(wr, 3).value = ref_text
        ws.cell(wr, 4).value = f"=C{wr}"

        if _simplificada:
            tot_ig  = round(sum(ing_vals.values()), 2) if ing_vals else round(sum(csv_vals.values()), 2)
            tot_vta = round(sum(vta_vals.values()), 2) if vta_vals else tot_ig
            ws.cell(wr, tot_ig_col).value = round(tot_ig, 2)
            ws.cell(wr, tot_vta_col).value = round(tot_vta, 2)
            ws.cell(wr, conc_col).value = round(tot_ig - tot_vta, 2)
        else:
            def _get_ing(col):
                hdr = col_to_name.get(col, "")
                return _buscar_alias(ing_vals, hdr, _aliases_ing) if hdr else None

            def _get_vta(col):
                hdr = col_to_name.get(col, "")
                acct = col_to_acct.get(col, "")
                if not hdr: return None
                if cuentas_vta_acct_set and acct and acct not in cuentas_vta_acct_set:
                    return None
                return _buscar_alias(vta_vals, hdr, _aliases_vta)

            for col in range(ing_start, tot_ig_col):
                if col == _bal_col: continue
                val = _get_ing(col)
                ws.cell(wr, col).value = val if val is not None else 0
                ws.cell(wr, col).number_format = "General"
            ws.cell(wr, tot_ig_col).value = f"=SUM({ing_ltr_s}{wr}:{ing_ltr_e}{wr})"

            for col in range(vta_start, tot_vta_col):
                val = _get_vta(col)
                ws.cell(wr, col).value = val if val is not None else 0
                ws.cell(wr, col).number_format = "General"
            ws.cell(wr, tot_vta_col).value = f"=SUM({vta_ltr_s}{wr}:{vta_ltr_e}{wr})"
            ws.cell(wr, conc_col).value = f"={tig_ltr}{wr}-{tva_ltr}{wr}"

            tot_ig  = round(sum(ing_vals.values()), 2)
            tot_vta = round(sum(vta_vals.values()), 2)

            if _bal_col is not None:
                _dif = round(tot_ig - tot_vta, 2)
                _adj = round(-_dif, 2)
                ws.cell(wr, _bal_col).value = _adj if abs(_adj) >= 0.01 else 0
                ws.cell(wr, _bal_col).number_format = "General"
                tot_ig = round(tot_vta, 2)

        _fill_colors(wr)
        conc_v = tot_ig - tot_vta
        estado = "OK" if abs(conc_v) < 1 else f"Dif {conc_v:+,.2f}"

        _ing_row = []
        for _th in _tree_ing_hdrs:
            _v = ing_vals.get(_th) or next((kv for k,kv in ing_vals.items() if k.upper()==_th.upper()), None)
            _ing_row.append(f"{(_v or 0):,.2f}")

        _vta_row = []
        for _ti, _th in zip(_tree_vta_src, _tree_vta_hdrs):
            _a = col_to_acct.get(_ti, "")
            if cuentas_vta_acct_set and _a and _a not in cuentas_vta_acct_set:
                _vta_row.append("0.00")
            else:
                _v = vta_vals.get(_th) or next((kv for k,kv in vta_vals.items() if k.upper()==_th.upper()), None)
                _vta_row.append(f"{(_v or 0):,.2f}")

        resultados.append(
            (fecha_str, *_ing_row, f"{tot_ig:,.2f}", *_vta_row, f"{tot_vta:,.2f}", f"{conc_v:,.2f}", estado))

    logs.append(f"✅ {len(resultados)} día(s) procesado(s)")
    return wb, resultados, logs


# ─── UI ────────────────────────────────────────────────────────────────────────
col1, col2 = st.columns([1, 2])
with col1:
    plantilla_file = st.file_uploader(
        "📂 Plantilla Excel (opcional)",
        type=["xlsx", "xlsm"],
        help="Plantilla SINUBE con hojas POLIZA y CUENTAS. Sin plantilla se genera estructura automática."
    )
with col2:
    csv_files = st.file_uploader(
        "📂 CSV(s) de ventas",
        type=["csv"],
        accept_multiple_files=True,
        help="Uno o más archivos CSV exportados del sistema de islas."
    )

st.markdown("---")

if csv_files:
    n_csv = len(csv_files)
    plantilla_nombre = plantilla_file.name if plantilla_file else "Sin plantilla (modo automático)"
    col_a, col_b = st.columns([3, 1])
    with col_a:
        st.info(f"📄 Plantilla: **{plantilla_nombre}**  |  📋 CSV(s): **{n_csv}** archivo(s)")
    with col_b:
        generar = st.button("⚙ Generar Excel Reconciliado", type="primary", use_container_width=True)

    if generar:
        plantilla_bytes = plantilla_file.read() if plantilla_file else None
        csv_dict = {f.name: f.read() for f in csv_files}

        with st.spinner("Procesando reconciliación..."):
            try:
                wb, resultados, logs = procesar_reconciliacion(plantilla_bytes, csv_dict)
            except Exception as e:
                import traceback
                st.error(f"❌ Error inesperado: {e}")
                with st.expander("Detalle del error"):
                    st.code(traceback.format_exc())
                st.stop()

        if wb is None:
            st.error("❌ No se pudo generar el Excel. Revisa el log.")
        else:
            _buf = io.BytesIO()
            wb.save(_buf)
            _buf.seek(0)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            _nombre = f"reconciliacion_{ts}.xlsx"

            ok_count = sum(1 for r in resultados if r[-1] == "OK")
            warn_count = len(resultados) - ok_count
            st.success(f"✅ {len(resultados)} día(s) procesado(s) — {ok_count} OK, {warn_count} con diferencia")

            st.download_button(
                label="⬇ Descargar Excel Reconciliado",
                data=_buf,
                file_name=_nombre,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            if resultados:
                st.markdown("### 📊 Reporte de Reconciliación")
                import pandas as pd

                _display = []
                for r in resultados:
                    _display.append({
                        "Fecha":        r[0],
                        "TOTAL IG":     r[-4],
                        "TOTAL VTA":    r[-3],
                        "CONCILIACION": r[-2],
                        "Estado":       r[-1],
                    })
                df = pd.DataFrame(_display)

                def _color_row(row):
                    if row["Estado"] == "OK":
                        return ["background-color:#E8F5E9"] * len(row)
                    elif "Error" in str(row["Estado"]) or "Sin fecha" in str(row["Estado"]):
                        return ["background-color:#FFEBEE"] * len(row)
                    else:
                        return ["background-color:#FFF9C4"] * len(row)

                styled = df.style.apply(_color_row, axis=1)
                st.dataframe(styled, use_container_width=True, hide_index=True)

        if logs:
            with st.expander("📋 Log de proceso"):
                for msg in logs:
                    st.text(msg)
else:
    st.markdown("""
    <div style="text-align:center;padding:40px;color:#6B7280;">
      <div style="font-size:3rem;">📑</div>
      <p style="font-size:1.1rem;margin-top:10px;">Selecciona al menos un archivo CSV para comenzar.</p>
      <p style="font-size:.9rem;color:#9CA3AF;">
        La plantilla Excel es opcional — sin ella se genera estructura automática desde los CSV.
      </p>
    </div>
    """, unsafe_allow_html=True)

st.markdown("---")
st.caption("Módulo Reconciliación · AUXILIAR DE REGISTROS")
