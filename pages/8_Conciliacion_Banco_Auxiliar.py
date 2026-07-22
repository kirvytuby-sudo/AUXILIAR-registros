"""
pages/8_Conciliacion_Banco_Auxiliar.py — v2.2
Paso 1: exacto    ±$TOL1,    fecha ±DIAS días
Paso 2: combinado ±$TOL_N,   fecha ±DIAS días, máx MAX_COMBO entradas
Paso 3: texto     ±$TOL_TEXT, fecha ±DIAS días, similitud ≥ SIM_MIN
Auto-sugerencia de parámetros basada en near-misses de no conciliados.
"""
import difflib
import io
import itertools
import os
import traceback
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime

import streamlit as st

st.set_page_config(
    page_title="Conciliación Banco vs Auxiliar | Auxiliar de Registros",
    page_icon="🔀",
    layout="wide",
)

import _theme
_theme.aplicar_header("🔀 Conciliación Banco vs Auxiliar", "Compara movimientos bancarios contra el auxiliar contable")

MESES = {1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",5:"MAYO",6:"JUNIO",
         7:"JULIO",8:"AGOSTO",9:"SEPTIEMBRE",10:"OCTUBRE",11:"NOVIEMBRE",12:"DICIEMBRE"}

# ── Utilidades ────────────────────────────────────────────────────────────────

def _to_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            return from_excel(v).date()
        except Exception: return None
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try: return datetime.strptime(s, fmt).date()
        except ValueError: continue
    return None


def _to_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).replace(",", "").replace("$", "").strip())
    except Exception: return 0.0


def _detect_header(rows, keywords):
    for i, row in enumerate(rows):
        vals = [str(v).strip().lower() if v is not None else "" for v in row]
        if all(kw in vals for kw in keywords):
            return i, {v: j for j, v in enumerate(vals) if v}
    return None, {}


def _col(mapping, *candidates):
    for cand in candidates:
        for k, v in mapping.items():
            if cand in k: return v
    return None


def text_sim(a, b):
    """Similitud de texto [0-1] entre descripción de banco y concepto auxiliar."""
    if not a or not b: return 0.0
    a = a.lower().strip(); b = b.lower().strip()
    return difflib.SequenceMatcher(None, a, b).ratio()


def _read_banco(wb):
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx, mapping = _detect_header(rows, ["fecha"])
    if hdr_idx is None:
        raise ValueError("No se encontró encabezado con columna 'Fecha' en el archivo del banco.")
    col_f   = _col(mapping, "fecha")
    col_d   = _col(mapping, "desc", "concepto", "referencia", "movimiento")
    col_dep = _col(mapping, "dep", "abono", "crédito", "credito", "entrada", "depósito", "deposito")
    col_ret = _col(mapping, "ret", "cargo", "débito", "debito", "salida", "retiro")
    col_sal = _col(mapping, "saldo", "sal")
    if col_f is None:
        raise ValueError("No se detectó la columna 'Fecha' en el banco.")
    movs = []
    for row in rows[hdr_idx + 1:]:
        if not row or all(v is None for v in row): continue
        fecha = _to_date(row[col_f] if col_f < len(row) else None)
        if fecha is None: continue
        desc = str(row[col_d] if col_d is not None and col_d < len(row) else "").strip()
        dep  = _to_float(row[col_dep] if col_dep is not None and col_dep < len(row) else 0)
        ret  = _to_float(row[col_ret] if col_ret is not None and col_ret < len(row) else 0)
        sal  = _to_float(row[col_sal] if col_sal is not None and col_sal < len(row) else 0)
        movs.append({"fecha": fecha, "desc": desc, "dep": dep, "ret": ret, "sal": sal})
    return movs


def _read_auxiliar(wb):
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx, mapping = _detect_header(rows, ["cargo", "fecha"])
    if hdr_idx is None:
        raise ValueError("No se encontró encabezado con columnas 'Fecha' y 'Cargo' en el auxiliar.")
    col_f = _col(mapping, "fecha")
    col_pol = None
    for c in ("núm. póliza", "num. póliza", "número póliza", "num poliza", "num. póliza"):
        if c in mapping: col_pol = mapping[c]; break
    if col_pol is None:
        for c in ("póliza", "poliza", "pol"):
            for k, v in mapping.items():
                if c in k and not k.startswith("tipo"): col_pol = v; break
            if col_pol is not None: break
    col_dp = None
    for c in ("desc. póliza", "desc poliza", "descripción póliza", "desc. detalle", "descripción", "descripcion", "desc"):
        if c in mapping: col_dp = mapping[c]; break
    if col_dp is None: col_dp = _col(mapping, "desc")
    col_cargo = _col(mapping, "cargo")
    col_abono = _col(mapping, "abono")

    def _s(row, col):
        if col is None or col >= len(row): return ""
        v = row[col]; s = "" if v is None else str(v).strip()
        return "" if s.lower() in ("none", "nan", "#n/a", "") else s

    aux_cargo = []; aux_abono = []
    for row in rows[hdr_idx + 1:]:
        if not row or all(v is None for v in row): continue
        fecha = _to_date(row[col_f] if col_f is not None and col_f < len(row) else None)
        if fecha is None: continue
        pol     = _s(row, col_pol)
        concepto = _s(row, col_dp)
        cargo = _to_float(row[col_cargo] if col_cargo is not None and col_cargo < len(row) else 0)
        abono = _to_float(row[col_abono] if col_abono is not None and col_abono < len(row) else 0)
        base = {"fecha": fecha, "poliza": pol, "concepto": concepto, "matched": False}
        if cargo > 0: aux_cargo.append({**base, "monto": cargo})
        if abono > 0: aux_abono.append({**base, "monto": abono})
    return aux_cargo, aux_abono


# ── Algoritmo de conciliación ─────────────────────────────────────────────────

def fecha_ok(f1, f2, dias):
    return abs((f1 - f2).days) <= dias


def conciliar(banco_movs, aux_pool, monto_key, tol1, tol_n, dias, tol_text, sim_min, max_combo,
              combo_min_pct=0.01, combo_sim_min=0.10):
    """
    Tres pasos:
      1. Exacto:    1-a-1, fecha ±dias, monto ±tol1
      2. Combinado: N-a-1, fecha ±dias, suma ±tol_n, máx max_combo entradas
                    · cada entrada ≥ max($1, combo_min_pct × total_banco)
                    · al menos una entrada con similitud texto ≥ combo_sim_min
      3. Texto:     1-a-1, fecha ±dias, monto ±tol_text, similitud texto ≥ sim_min
    """
    libre = [dict(a) for a in aux_pool]
    results = []; matched_idx = set()

    # Paso 1 — exacto
    for bi, banco in enumerate(banco_movs):
        target = banco[monto_key]
        for aux in libre:
            if aux["matched"] or not fecha_ok(banco["fecha"], aux["fecha"], dias): continue
            if abs(aux["monto"] - target) <= tol1:
                results.append({"tipo": "✅ EXACTO", "banco": banco,
                                 "aux_entries": [aux], "diferencia": aux["monto"] - target})
                aux["matched"] = True; matched_idx.add(bi); break

    # Paso 2 — combinado
    libres_p2 = [a for a in libre if not a["matched"]]
    for bi, banco in enumerate(banco_movs):
        if bi in matched_idx: continue
        target = banco[monto_key]
        if target < 10: continue
        min_monto = max(1.0, target * combo_min_pct)   # ej. 1% de $2600 = $26
        desc_b = banco["desc"]
        candidatos = [a for a in libres_p2
                      if not a["matched"] and fecha_ok(banco["fecha"], a["fecha"], dias)
                      and a["monto"] >= min_monto                   # filtro monto mínimo
                      and a["monto"] <= target + tol_n]
        found = False
        for n in range(2, min(max_combo + 1, len(candidatos) + 1)):
            for combo in itertools.combinations(candidatos, n):
                if abs(sum(c["monto"] for c in combo) - target) <= tol_n:
                    # filtro descripción: al menos una entrada debe parecerse
                    max_sim = max(text_sim(desc_b, c["concepto"]) for c in combo)
                    if max_sim < combo_sim_min:
                        continue
                    results.append({"tipo": f"🔀 COMBINADO ({n})", "banco": banco,
                                    "aux_entries": list(combo),
                                    "diferencia": sum(c["monto"] for c in combo) - target})
                    for c in combo: c["matched"] = True
                    matched_idx.add(bi); found = True; break
            if found: break

    # Paso 3 — similitud de texto
    libres_p3 = [a for a in libre if not a["matched"]]
    for bi, banco in enumerate(banco_movs):
        if bi in matched_idx: continue
        target = banco[monto_key]
        desc_b = banco["desc"]
        candidatos = [
            (a, text_sim(desc_b, a["concepto"]))
            for a in libres_p3
            if not a["matched"] and fecha_ok(banco["fecha"], a["fecha"], dias)
            and abs(a["monto"] - target) <= tol_text
        ]
        candidatos.sort(key=lambda x: -x[1])
        for aux, sim in candidatos:
            if sim >= sim_min:
                results.append({"tipo": f"🔤 TEXTO ({sim:.0%})", "banco": banco,
                                 "aux_entries": [aux], "diferencia": aux["monto"] - target})
                aux["matched"] = True; matched_idx.add(bi); break

    sin_banco = [banco_movs[i] for i in range(len(banco_movs)) if i not in matched_idx]
    sin_aux   = [a for a in libre if not a["matched"]]
    return results, sin_banco, sin_aux


def analizar_near_misses(sin_banco, sin_aux, monto_key, dias_act, tol_text_act, sim_act):
    """
    Para cada item bancario sin conciliar busca el mejor candidato en sin_aux.
    Devuelve lista de near-misses y parámetros sugeridos.
    """
    near = []
    for b in sin_banco:
        mejor = None; mejor_sim = -1
        for a in sin_aux:
            dd = abs((b["fecha"] - a["fecha"]).days)
            da = abs(a["monto"] - b[monto_key])
            if da > 500 or dd > 60: continue
            sim = text_sim(b["desc"], a["concepto"])
            score = sim - da / 5000 - dd / 200
            if score > mejor_sim:
                mejor_sim = score
                mejor = {"damt": da, "ddias": dd, "sim": sim,
                         "desc_b": b["desc"], "conc_a": a["concepto"],
                         "monto_b": b[monto_key], "monto_a": a["monto"],
                         "fecha_b": b["fecha"], "fecha_a": a["fecha"]}
        if mejor:
            near.append(mejor)

    sugs = {}
    if near:
        solo_monto = [m for m in near if m["ddias"] <= dias_act and m["sim"] >= sim_act]
        if solo_monto:
            vals = sorted(m["damt"] for m in solo_monto)
            sugs["tol_text"] = round(vals[int(len(vals) * 0.8)], 2)

        solo_fecha = [m for m in near if m["damt"] <= tol_text_act and m["sim"] >= sim_act]
        if solo_fecha:
            vals = sorted(m["ddias"] for m in solo_fecha)
            sugs["dias"] = int(vals[int(len(vals) * 0.8)])

        solo_texto = [m for m in near if m["damt"] <= tol_text_act and m["ddias"] <= dias_act]
        if solo_texto:
            vals = sorted(m["sim"] for m in solo_texto)
            raw = vals[int(len(vals) * 0.2)]  # 20th percentile (lower = más permisivo)
            sugs["sim_min"] = max(0.20, round(raw - 0.05, 2))

    return near, sugs


# ── Generación de Excel ───────────────────────────────────────────────────────

def _generar_excel(res_dep, sin_dep_banco, sin_dep_aux,
                   res_ret, sin_ret_banco, sin_ret_aux, meses_ord):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)

    EXACTO  = PatternFill("solid", fgColor="D1FAE5")   # verde
    COMBO   = PatternFill("solid", fgColor="FCE4D6")   # naranja claro
    TEXTO   = PatternFill("solid", fgColor="EDE9FE")   # lila
    AUXF    = PatternFill("solid", fgColor="FFF3CD")
    DEP_M   = PatternFill("solid", fgColor="1E40AF")
    RET_M   = PatternFill("solid", fgColor="831843")
    DEP_S   = PatternFill("solid", fgColor="1D4ED8")
    RET_S   = PatternFill("solid", fgColor="9D174D")
    DEP_R   = PatternFill("solid", fgColor="DBEAFE")
    RET_R   = PatternFill("solid", fgColor="FCE7F3")
    DEP_A   = PatternFill("solid", fgColor="BFDBFE")
    RET_A   = PatternFill("solid", fgColor="FBCFE8")
    TOT_F   = PatternFill("solid", fgColor="FEF3C7")
    GRAND_F = PatternFill("solid", fgColor="FCD34D")
    BLANK   = PatternFill("solid", fgColor="F9FAFB")

    def hdr(ws, r, c, v, f=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.fill = f or PatternFill("solid", fgColor="1E3A8A")
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = brd

    def dc(ws, r, c, v, fill=None, fmt=None, bold=False, align="left"):
        cell = ws.cell(row=r, column=c, value=v)
        if fill: cell.fill = fill
        if fmt:  cell.number_format = fmt
        cell.font = Font(bold=bold, size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = brd

    CONC_H = ["ESTATUS","FECHA BANCO","MONTO BANCO","DESCRIPCIÓN BANCO",
               "FECHA AUX","PÓLIZA","CONCEPTO AUX","MONTO AUX","DIFERENCIA"]
    CONC_W = [22, 13, 15, 50, 13, 10, 50, 15, 12]

    def _row_fill(tipo):
        if tipo.startswith("✅"): return EXACTO
        if tipo.startswith("🔀"): return COMBO
        return TEXTO

    def write_conc_por_mes(ws, results, mkey, mfill):
        for ci, w in enumerate(CONC_W, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"
        por_mes = defaultdict(list)
        for r in results: por_mes[r["banco"]["fecha"].month].append(r)
        row = 1; grand_total = 0
        for mes in meses_ord:
            its = por_mes.get(mes, [])
            if not its: continue
            ws.row_dimensions[row].height = 24
            for ci in range(1, 10):
                cell = ws.cell(row=row, column=ci, value=f"  {MESES[mes]}" if ci == 1 else "")
                cell.fill = mfill; cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(vertical="center"); cell.border = brd
            row += 1
            for ci, h in enumerate(CONC_H, 1): hdr(ws, row, ci, h)
            row += 1; subtotal = 0
            for r2 in its:
                fill = _row_fill(r2["tipo"])
                a0 = r2["aux_entries"][0]; b = r2["banco"]
                dc(ws, row, 1, r2["tipo"], fill)
                dc(ws, row, 2, b["fecha"], fill, fmt="DD/MM/YYYY", align="center")
                dc(ws, row, 3, b[mkey], fill, fmt='#,##0.00', align="right")
                dc(ws, row, 4, b["desc"], fill)
                dc(ws, row, 5, a0["fecha"], fill, fmt="DD/MM/YYYY", align="center")
                dc(ws, row, 6, a0["poliza"], fill, align="center")
                dc(ws, row, 7, a0["concepto"], fill)
                dc(ws, row, 8, a0["monto"], fill, fmt='#,##0.00', align="right")
                dc(ws, row, 9, r2["diferencia"], fill, fmt='#,##0.00', align="right")
                subtotal += b[mkey]; row += 1
                for a in r2["aux_entries"][1:]:
                    for ci in range(1, 5): dc(ws, row, ci, "", AUXF)
                    dc(ws, row, 5, a["fecha"], AUXF, fmt="DD/MM/YYYY", align="center")
                    dc(ws, row, 6, a["poliza"], AUXF, align="center")
                    dc(ws, row, 7, a["concepto"], AUXF)
                    dc(ws, row, 8, a["monto"], AUXF, fmt='#,##0.00', align="right")
                    dc(ws, row, 9, "", AUXF); row += 1
            for ci in range(1, 10): dc(ws, row, ci, "", TOT_F)
            dc(ws, row, 1, f"Subtotal {MESES[mes]}", TOT_F, bold=True, align="right")
            dc(ws, row, 3, subtotal, TOT_F, fmt='#,##0.00', align="right", bold=True)
            grand_total += subtotal; row += 1
        for ci in range(1, 10): dc(ws, row, ci, "", GRAND_F)
        dc(ws, row, 1, "TOTAL", GRAND_F, bold=True, align="right")
        dc(ws, row, 3, grand_total, GRAND_F, fmt='#,##0.00', align="right", bold=True)

    def write_pair_by_month(ws, start_row,
                             items_l, monto_l, label_l, fill_l, mfill_l, sfill_l, sec_l,
                             items_r, monto_r, label_r, pol_r,  fill_r, mfill_r, sfill_r, sec_r):
        row = start_row; L = (1, 2, 3); R = (5, 6, 7, 8); SEP = 4
        ws.row_dimensions[row].height = 26
        for ci in L:
            cell = ws.cell(row=row, column=ci, value=sec_l if ci == L[0] else "")
            cell.fill = sfill_l; cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(vertical="center"); cell.border = brd
        ws.cell(row=row, column=SEP, value="").border = brd
        for ci in R:
            cell = ws.cell(row=row, column=ci, value=sec_r if ci == R[0] else "")
            cell.fill = sfill_r; cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(vertical="center"); cell.border = brd
        row += 1
        por_mes_l = defaultdict(list); por_mes_r = defaultdict(list)
        for it in items_l: por_mes_l[it["fecha"].month].append(it)
        for it in items_r: por_mes_r[it["fecha"].month].append(it)
        grand_l = 0; grand_r = 0
        for mes in meses_ord:
            its_l = sorted(por_mes_l.get(mes, []), key=lambda x: x["fecha"])
            its_r = sorted(por_mes_r.get(mes, []), key=lambda x: x["fecha"])
            if not its_l and not its_r: continue
            ws.row_dimensions[row].height = 20
            for ci in L:
                cell = ws.cell(row=row, column=ci, value=f"  {MESES[mes]}" if ci == L[0] else "")
                cell.fill = mfill_l; cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(vertical="center"); cell.border = brd
            ws.cell(row=row, column=SEP, value="").border = brd
            for ci in R:
                cell = ws.cell(row=row, column=ci, value=f"  {MESES[mes]}" if ci == R[0] else "")
                cell.fill = mfill_r; cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.alignment = Alignment(vertical="center"); cell.border = brd
            row += 1
            hdr(ws, row, L[0], "FECHA",             mfill_l)
            hdr(ws, row, L[1], "MONTO",             mfill_l)
            hdr(ws, row, L[2], "DESCRIPCION BANCO", mfill_l)
            ws.cell(row=row, column=SEP, value="").border = brd
            hdr(ws, row, R[0], "FECHA",       mfill_r)
            hdr(ws, row, R[1], "MONTO",       mfill_r)
            hdr(ws, row, R[2], "CONCEPTO AUX",mfill_r)
            hdr(ws, row, R[3], "POLIZA",      mfill_r)
            row += 1; sub_l = 0; sub_r = 0
            for i in range(max(len(its_l), len(its_r))):
                ws.cell(row=row, column=SEP, value="").border = brd
                if i < len(its_l):
                    it = its_l[i]; m = monto_l(it); sub_l += m
                    dc(ws, row, L[0], it["fecha"], fill_l, fmt="DD/MM/YYYY", align="center")
                    dc(ws, row, L[1], m,           fill_l, fmt='#,##0.00',   align="right")
                    dc(ws, row, L[2], label_l(it), fill_l)
                else:
                    for ci in L: dc(ws, row, ci, "", BLANK)
                if i < len(its_r):
                    it = its_r[i]; m = monto_r(it); sub_r += m
                    dc(ws, row, R[0], it["fecha"], fill_r, fmt="DD/MM/YYYY", align="center")
                    dc(ws, row, R[1], m,           fill_r, fmt='#,##0.00',   align="right")
                    dc(ws, row, R[2], label_r(it), fill_r)
                    dc(ws, row, R[3], pol_r(it),   fill_r, align="center")
                else:
                    for ci in R: dc(ws, row, ci, "", BLANK)
                row += 1
            ws.cell(row=row, column=SEP, value="").border = brd
            dc(ws, row, L[0], f"Subtotal {MESES[mes]}", TOT_F, bold=True, align="right")
            dc(ws, row, L[1], sub_l, TOT_F, fmt='#,##0.00', align="right", bold=True)
            dc(ws, row, L[2], "", TOT_F)
            dc(ws, row, R[0], f"Subtotal {MESES[mes]}", TOT_F, bold=True, align="right")
            dc(ws, row, R[1], sub_r, TOT_F, fmt='#,##0.00', align="right", bold=True)
            dc(ws, row, R[2], "", TOT_F); dc(ws, row, R[3], "", TOT_F)
            grand_l += sub_l; grand_r += sub_r; row += 1
        ws.cell(row=row, column=SEP, value="").border = brd
        dc(ws, row, L[0], "TOTAL", GRAND_F, bold=True, align="right")
        dc(ws, row, L[1], grand_l, GRAND_F, fmt='#,##0.00', align="right", bold=True)
        dc(ws, row, L[2], "", GRAND_F)
        dc(ws, row, R[0], "TOTAL", GRAND_F, bold=True, align="right")
        dc(ws, row, R[1], grand_r, GRAND_F, fmt='#,##0.00', align="right", bold=True)
        dc(ws, row, R[2], "", GRAND_F); dc(ws, row, R[3], "", GRAND_F)
        return row + 2

    wb = Workbook()
    ws1 = wb.active; ws1.title = "\U0001f4b0 Depósitos"; ws1.sheet_properties.tabColor = "1E3A8A"
    write_conc_por_mes(ws1, res_dep, "dep", DEP_M)

    ws2 = wb.create_sheet("\U0001f4b3 Cargos"); ws2.sheet_properties.tabColor = "BE185D"
    write_conc_por_mes(ws2, res_ret, "ret", RET_M)

    ws3 = wb.create_sheet("⚠ Sin conciliar"); ws3.sheet_properties.tabColor = "D97706"
    for ci, w in {1: 13, 2: 15, 3: 52, 4: 2, 5: 13, 6: 15, 7: 50, 8: 10}.items():
        ws3.column_dimensions[get_column_letter(ci)].width = w
    row = 1
    row = write_pair_by_month(
        ws3, row,
        sin_dep_banco, lambda x: x["dep"],   lambda x: x["desc"],    DEP_R, DEP_M, DEP_S,
        "BANCO — DEPÓSITOS SIN CONCILIAR",
        sin_dep_aux,   lambda x: x["monto"], lambda x: x["concepto"], lambda x: x.get("poliza", ""),
        DEP_A, DEP_M, PatternFill("solid", fgColor="1E40AF"),
        "AUXILIAR — CARGOS SIN CONCILIAR")
    row = write_pair_by_month(
        ws3, row,
        sin_ret_banco, lambda x: x["ret"],   lambda x: x["desc"],    RET_R, RET_M, RET_S,
        "BANCO — CARGOS SIN CONCILIAR",
        sin_ret_aux,   lambda x: x["monto"], lambda x: x["concepto"], lambda x: x.get("poliza", ""),
        RET_A, RET_M, PatternFill("solid", fgColor="831843"),
        "AUXILIAR — ABONOS SIN CONCILIAR")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ── Estado de sesión ──────────────────────────────────────────────────────────
for key in ("cba_resultado_bytes", "cba_resumen", "cba_sugerencias"):
    if key not in st.session_state:
        st.session_state[key] = None

# ── UI — Parámetros ───────────────────────────────────────────────────────────
with st.expander("⚙️ Parámetros de conciliación", expanded=False):
    st.caption("Ajusta las tolerancias manualmente o usa las sugerencias automáticas abajo.")
    c1, c2, c3 = st.columns(3)
    with c1:
        p_tol1          = st.slider("Exacto — tolerancia monto ($)",        0.01, 5.0,  0.05, 0.01, key="p_tol1")
        p_tol_n         = st.slider("Combo  — tolerancia suma ($)",          0.5,  20.0, 2.0,  0.5,  key="p_toln")
        p_combo_min_pct = st.slider("Combo  — monto mínimo por entrada (%)", 0,    10,   1,    1,    key="p_cminpct",
                                    help="Cada entrada del combo debe ser al menos este % del total del banco. Evita combinar centavos con pagos grandes.")
    with c2:
        p_dias          = st.slider("Días de tolerancia fecha",              0,    15,   3,    1,    key="p_dias")
        p_tol_text      = st.slider("Texto  — tolerancia monto ($)",         1.0,  50.0, 5.0,  1.0,  key="p_tolt")
        p_combo_sim_min = st.slider("Combo  — similitud mínima descripción", 0,    50,   10,   5,    key="p_csim",
                                    format="%d%%",
                                    help="Al menos una entrada del combo debe parecerse en descripción al banco. 10% es muy permisivo; sube si hay falsos positivos.")
    with c3:
        p_sim_min       = st.slider("Texto  — similitud mínima",             0.20, 0.90, 0.45, 0.05, key="p_sim",
                                    format="%.2f",
                                    help="0.45 = 45% de similitud entre descripción del banco y concepto del auxiliar")
        p_max_combo     = st.slider("Combo  — máx. entradas",                2,    8,    6,    1,    key="p_maxc")

    if st.session_state.cba_sugerencias:
        sugs = st.session_state.cba_sugerencias
        st.markdown("**💡 Sugerencias basadas en la última corrida:**")
        cols = st.columns(4)
        labels = {"tol_text": "Tolerancia texto ($)", "dias": "Días fecha", "sim_min": "Similitud mín."}
        keys   = {"tol_text": "p_tolt",              "dias": "p_dias",     "sim_min": "p_sim"}
        for i, (k, v) in enumerate(sugs.items()):
            with cols[i % 4]:
                st.metric(labels.get(k, k), v)
        st.caption("👆 Mueve los sliders a estos valores y presiona **Generar** de nuevo para mejorar el resultado.")

# ── UI — Carga de archivos ────────────────────────────────────────────────────
st.subheader("1️⃣  Cargar archivos")
col_banco, col_aux = st.columns(2)

with col_banco:
    banco_files = st.file_uploader(
        "🏦 Archivo(s) del banco (.xlsx)",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        help="Uno o más estados de cuenta en Excel. Se combinan automáticamente.",
        key="cba_banco",
    )

with col_aux:
    aux_file = st.file_uploader(
        "📋 Auxiliar contable (.xlsx)",
        type=["xlsx", "xls"],
        help="Auxiliar con columnas Fecha, Póliza, Concepto, Cargo, Abono.",
        key="cba_aux",
    )

# ── Leyenda ───────────────────────────────────────────────────────────────────
st.markdown("""
<div style="margin:0.5rem 0 1rem; font-size:0.85rem; display:flex; gap:1rem; flex-wrap:wrap;">
  <span style="background:#D1FAE5; padding:2px 8px; border-radius:4px;">✅ Exacto</span>
  <span style="background:#FCE4D6; padding:2px 8px; border-radius:4px;">🔀 Combinado</span>
  <span style="background:#EDE9FE; padding:2px 8px; border-radius:4px;">🔤 Texto (similitud)</span>
  <span style="background:#DBEAFE; padding:2px 8px; border-radius:4px;">⚠ Sin conciliar (banco)</span>
  <span style="background:#BFDBFE; padding:2px 8px; border-radius:4px;">⚠ Sin conciliar (aux)</span>
</div>
""", unsafe_allow_html=True)


# ── Botón generar ─────────────────────────────────────────────────────────────
st.subheader("2️⃣  Generar conciliación")
col_btn, col_dl = st.columns([1, 2])
with col_btn:
    can_gen = bool(banco_files) and aux_file is not None
    generar = st.button(
        "🔀 Generar conciliación",
        disabled=not can_gen,
        type="primary",
        use_container_width=True,
    )
    if not can_gen:
        st.caption("Carga al menos un archivo del banco y el auxiliar para continuar.")

if generar:
    st.session_state.cba_resultado_bytes = None
    st.session_state.cba_resumen = None
    st.session_state.cba_sugerencias = None

    with st.spinner("Leyendo archivos..."):
        try:
            from openpyxl import load_workbook
            movs_banco = []; archivos_leidos = []
            for bf in banco_files:
                wb_b = load_workbook(filename=io.BytesIO(bf.read()), read_only=True, data_only=True)
                movs_bf = _read_banco(wb_b); wb_b.close()
                movs_banco.extend(movs_bf)
                archivos_leidos.append(f"{bf.name} ({len(movs_bf)} movimientos)")
            movs_banco.sort(key=lambda m: m["fecha"])

            wb_aux = load_workbook(filename=io.BytesIO(aux_file.read()), read_only=True, data_only=True)
            aux_cargo_raw, aux_abono_raw = _read_auxiliar(wb_aux); wb_aux.close()
        except Exception as e:
            st.error(f"❌ Error al leer archivos: {e}")
            with st.expander("Ver detalle"): st.code(traceback.format_exc())
            st.stop()

    with st.spinner("Conciliando (3 pasos: exacto · combinaciones · texto)..."):
        try:
            if movs_banco:
                f_min = min(m["fecha"] for m in movs_banco)
                f_max = max(m["fecha"] for m in movs_banco)
                f_min = date(f_min.year, f_min.month, 1)
                f_max = date(f_max.year, f_max.month, monthrange(f_max.year, f_max.month)[1])
                aux_cargo_rng = [a for a in aux_cargo_raw if f_min <= a["fecha"] <= f_max]
                aux_abono_rng = [a for a in aux_abono_raw if f_min <= a["fecha"] <= f_max]
            else:
                aux_cargo_rng = []; aux_abono_rng = []

            deps_banco = [m for m in movs_banco if m["dep"] > 0]
            rets_banco = [m for m in movs_banco if m["ret"] > 0]
            meses_ord  = sorted(set(m["fecha"].month for m in movs_banco))

            params = dict(tol1=p_tol1, tol_n=p_tol_n, dias=p_dias,
                          tol_text=p_tol_text, sim_min=p_sim_min, max_combo=p_max_combo,
                          combo_min_pct=p_combo_min_pct / 100,
                          combo_sim_min=p_combo_sim_min / 100)

            res_dep, sin_dep_banco, sin_dep_aux = conciliar(deps_banco, aux_cargo_rng, "dep", **params)
            res_ret, sin_ret_banco, sin_ret_aux = conciliar(rets_banco, aux_abono_rng, "ret", **params)

            # Near-miss analysis para sugerencias
            near_dep, sugs_dep = analizar_near_misses(sin_dep_banco, sin_dep_aux, "dep",
                                                       p_dias, p_tol_text, p_sim_min)
            near_ret, sugs_ret = analizar_near_misses(sin_ret_banco, sin_ret_aux, "ret",
                                                       p_dias, p_tol_text, p_sim_min)
            # Combinar sugerencias (más permisivo)
            sugs = {}
            for k in set(list(sugs_dep.keys()) + list(sugs_ret.keys())):
                vals = [v for v in [sugs_dep.get(k), sugs_ret.get(k)] if v is not None]
                if vals: sugs[k] = max(vals)
            st.session_state.cba_sugerencias = sugs if sugs else None

        except Exception as e:
            st.error(f"❌ Error en conciliación: {e}")
            with st.expander("Ver detalle"): st.code(traceback.format_exc())
            st.stop()

    with st.spinner("Generando Excel..."):
        try:
            excel_bytes = _generar_excel(
                res_dep, sin_dep_banco, sin_dep_aux,
                res_ret, sin_ret_banco, sin_ret_aux,
                meses_ord)
            st.session_state.cba_resultado_bytes = excel_bytes

            exactos_dep = sum(1 for r in res_dep if r["tipo"].startswith("✅"))
            combo_dep   = sum(1 for r in res_dep if r["tipo"].startswith("🔀"))
            texto_dep   = sum(1 for r in res_dep if r["tipo"].startswith("🔤"))
            exactos_ret = sum(1 for r in res_ret if r["tipo"].startswith("✅"))
            combo_ret   = sum(1 for r in res_ret if r["tipo"].startswith("🔀"))
            texto_ret   = sum(1 for r in res_ret if r["tipo"].startswith("🔤"))

            st.session_state.cba_resumen = {
                "conc_dep": len(res_dep),   "total_dep": len(deps_banco),
                "exactos_dep": exactos_dep, "combo_dep": combo_dep, "texto_dep": texto_dep,
                "sin_dep_b": len(sin_dep_banco), "sin_dep_a": len(sin_dep_aux),
                "conc_ret": len(res_ret),   "total_ret": len(rets_banco),
                "exactos_ret": exactos_ret, "combo_ret": combo_ret, "texto_ret": texto_ret,
                "sin_ret_b": len(sin_ret_banco), "sin_ret_a": len(sin_ret_aux),
                "n_banco": len(movs_banco),
                "archivos_banco": archivos_leidos,
                "near_dep": near_dep, "near_ret": near_ret,
            }
        except Exception as e:
            st.error(f"❌ Error generando Excel: {e}")
            with st.expander("Ver detalle"): st.code(traceback.format_exc())
            st.stop()

# ── Resultado ─────────────────────────────────────────────────────────────────
if st.session_state.cba_resultado_bytes and st.session_state.cba_resumen:
    res = st.session_state.cba_resumen
    st.success("✅ Conciliación completada")

    if res.get("archivos_banco"):
        n = len(res["archivos_banco"])
        with st.expander(f"📂 {n} archivo{'s' if n > 1 else ''} del banco", expanded=n > 1):
            for a in res["archivos_banco"]: st.markdown(f"- {a}")

    c1, c2, c3, c4 = st.columns(4)
    pct_dep = res["conc_dep"] / res["total_dep"] * 100 if res["total_dep"] else 0
    pct_ret = res["conc_ret"] / res["total_ret"] * 100 if res["total_ret"] else 0
    sin_tot = res["sin_dep_b"] + res["sin_ret_b"]

    with c1:
        st.metric("Movimientos banco", res["n_banco"])
    with c2:
        st.metric("Depósitos conciliados",
                  f"{res['conc_dep']}/{res['total_dep']}",
                  delta=f"{pct_dep:.1f}% · {res['exactos_dep']}✅ {res['combo_dep']}🔀 {res['texto_dep']}🔤")
    with c3:
        st.metric("Cargos conciliados",
                  f"{res['conc_ret']}/{res['total_ret']}",
                  delta=f"{pct_ret:.1f}% · {res['exactos_ret']}✅ {res['combo_ret']}🔀 {res['texto_ret']}🔤")
    with c4:
        st.metric("Sin conciliar (banco)", sin_tot,
                  delta=f"Dep: {res['sin_dep_b']} · Cargos: {res['sin_ret_b']}",
                  delta_color="inverse")

    st.download_button(
        label="⬇️  Descargar Excel de conciliación",
        data=st.session_state.cba_resultado_bytes,
        file_name="Conciliacion_Banco_Auxiliar.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )
    st.caption("3 hojas: 💰 Depósitos · 💳 Cargos · ⚠ Sin conciliar (lado a lado por mes)")

    # ── Near-miss analysis ────────────────────────────────────────────────────
    near_all = res.get("near_dep", []) + res.get("near_ret", [])
    if near_all:
        with st.expander(f"📊 Análisis de no conciliados — {len(near_all)} near-misses", expanded=False):
            st.markdown("Los items sin conciliar más cercanos a tener un match:")
            top = sorted(near_all, key=lambda x: x["sim"], reverse=True)[:20]
            rows_tbl = []
            for m in top:
                rows_tbl.append({
                    "Fecha banco": m["fecha_b"].strftime("%d/%m/%Y"),
                    "Monto banco": f"${m['monto_b']:,.2f}",
                    "Descripción banco": m["desc_b"][:40],
                    "Fecha aux": m["fecha_a"].strftime("%d/%m/%Y"),
                    "Monto aux": f"${m['monto_a']:,.2f}",
                    "Concepto aux": m["conc_a"][:40],
                    "Δ monto": f"${m['damt']:.2f}",
                    "Δ días": m["ddias"],
                    "Similitud txt": f"{m['sim']:.0%}",
                })
            st.dataframe(rows_tbl, use_container_width=True)

            if st.session_state.cba_sugerencias:
                st.markdown("**💡 Para mejorar el resultado, ajusta los parámetros arriba a:**")
                for k, v in st.session_state.cba_sugerencias.items():
                    labels = {"tol_text": "Tolerancia texto ($)", "dias": "Días fecha", "sim_min": "Similitud mín."}
                    st.markdown(f"- **{labels.get(k,k)}** → `{v}`")

else:
    st.markdown("""
<div style="text-align:center; padding: 3rem 2rem; background:#F8FAFC; border-radius:12px; margin-top:1rem;">
  <div style="font-size:3rem; margin-bottom:1rem;">🔀</div>
  <h3 style="color:#1E3A8A; margin:0 0 .5rem;">Conciliación Banco vs Auxiliar</h3>
  <p style="color:#64748B; max-width:480px; margin:0 auto;">
    Carga el Excel del banco y el auxiliar contable, luego presiona
    <strong>Generar conciliación</strong>. El resultado se descarga como Excel
    con hojas de Depósitos conciliados, Cargos conciliados y Sin conciliar.
  </p>
</div>
""", unsafe_allow_html=True)

st.markdown("---")
st.caption("Módulo Conciliación Banco vs Auxiliar · v2.3  ·  Exacto ±$0.05 · Combo ±$2.00 (min 1% por entrada, sim ≥10%) · Texto ≥45% · Fecha ±3 días · Auto-sugerencias")
